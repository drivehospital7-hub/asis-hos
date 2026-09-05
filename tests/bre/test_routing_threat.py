"""Routing threat harness (Slice 2 task 2.4, strict TDD).

Threat: POST /procesar dispatch must never 500 on unknown Tipo Factura;
per-tipo orquestador exceptions are contained; AREA_UNIFICADA contract holds.
No engine refactor. No prod connection (engine flag OFF / forbidden session).
"""
from __future__ import annotations

from unittest.mock import patch

from openpyxl import Workbook

from app.constants import AREA_UNIFICADA
from app.services.transversales.column_indices import get_column_indices

REQUIRED_HEADERS: dict[str, str] = {
    "numero_factura": "Número Factura",
    "tipo_factura_descripcion": "Tipo Factura Descripción",
    "codigo": "Código",
    "cantidad": "Cantidad",
    "centro_costo": "Centro Costo",
    "ide_contrato": "IDE Contrato",
    "codigo_entidad_cobrar": "Cód Entidad Cobrar",
}

UNKNOWN_TIPO = "Tipo Inexistente XYZ"


def _build_workbook(rows: list[tuple[str, str]]) -> Workbook:
    """Minimal workbook with exact headers; rows are (factura, tipo)."""
    wb = Workbook()
    ws = wb.active
    headers = list(REQUIRED_HEADERS.values())
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    for r, (fact, tipo) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=fact)
        ws.cell(row=r, column=2, value=tipo)
        ws.cell(row=r, column=3, value="890701")
        ws.cell(row=r, column=4, value=1)
        ws.cell(row=r, column=5, value="URGENCIAS")
        ws.cell(row=r, column=6, value="986")
        ws.cell(row=r, column=7, value="EPSI05")
    return wb


def _indices(wb: Workbook) -> dict[str, int | None]:
    headers = [wb.active.cell(row=1, column=c).value for c in range(1, wb.active.max_column + 1)]
    indices, _ = get_column_indices(headers, REQUIRED_HEADERS)
    return indices


def _forbidden_session(*args, **kwargs):  # pragma: no cover
    raise AssertionError("routing harness must not open a DB connection")


class TestRoutingThreat:
    def test_unknown_tipo_returns_empty_unificada(self, caplog) -> None:
        """Unknown Tipo Factura only: empty unificada set + warning, never 500."""
        from app.services.unified_processor import process_unified

        wb = _build_workbook([("FAC-UNK-001", UNKNOWN_TIPO)])
        indices = _indices(wb)
        with (
            patch("app.services.unified_processor.is_rule_engine_enabled", return_value=False),
            patch("app.database.get_session", _forbidden_session),
        ):
            with caplog.at_level("WARNING"):
                resultado, _ = process_unified(wb.active, indices)
        assert resultado["area"] == AREA_UNIFICADA
        assert resultado["problemas"]["normalizados"] == []
        assert resultado["tipos_procesados"] == []
        assert any("tipo" in rec.message.lower() or "conocido" in rec.message.lower() for rec in caplog.records)

    def test_per_tipo_dispatch_ignores_unknown(self) -> None:
        """Mixed known + unknown: known dispatches, unknown ignored, never 500."""
        from app.services.unified_processor import process_unified

        wb = _build_workbook([("FAC-URG-001", "Urgencias"), ("FAC-UNK-002", UNKNOWN_TIPO)])
        indices = _indices(wb)
        with (
            patch("app.services.unified_processor.is_rule_engine_enabled", return_value=False),
            patch("app.database.get_session", _forbidden_session),
        ):
            resultado, _ = process_unified(wb.active, indices)
        assert resultado["area"] == AREA_UNIFICADA
        assert resultado["tipos_procesados"] == ["Urgencias"]

    def test_per_tipo_exception_contained(self) -> None:
        """Orquestador raising for one tipo: contained, others continue, never 500."""
        from app.services import unified_processor
        from app.services.unified_processor import process_unified

        wb = _build_workbook([("FAC-URG-001", "Urgencias")])
        indices = _indices(wb)
        real_get = unified_processor._get_orquestador

        def _boom(tipo: str):
            if tipo == "Urgencias":
                raise RuntimeError("boom")
            return real_get(tipo)

        with (
            patch.object(unified_processor, "_get_orquestador", side_effect=_boom),
            patch("app.services.unified_processor.is_rule_engine_enabled", return_value=False),
            patch("app.database.get_session", _forbidden_session),
        ):
            resultado, _ = process_unified(wb.active, indices)
        assert resultado["area"] == AREA_UNIFICADA
        assert resultado["problemas"]["normalizados"] == []

    def test_area_unificada_contract(self) -> None:
        """Unified envelope always carries AREA_UNIFICADA + tipos_procesados."""
        from app.services.unified_processor import process_unified

        wb = _build_workbook([("FAC-UNK-001", UNKNOWN_TIPO)])
        indices = _indices(wb)
        with (
            patch("app.services.unified_processor.is_rule_engine_enabled", return_value=False),
            patch("app.database.get_session", _forbidden_session),
        ):
            resultado, responsables = process_unified(wb.active, indices)
        assert resultado["area"] == AREA_UNIFICADA
        assert isinstance(resultado["tipos_procesados"], list)
        assert isinstance(responsables, dict)
