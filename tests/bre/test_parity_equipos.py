"""EQUI parity harness (Slice 2, strict TDD).

Unified /procesar (process_unified) must match the legacy equipos_basicos
orchestrator on the same fixture: order-insensitive problem-set compare,
missing-column tolerance, exact header matching. Routes stay delegators.
No engine refactor. No prod connection (engine flag OFF / mocked session).

Routing note: process_unified dispatches "Odontología" to
detect_all_problems_odontologia_por_responsable, which routes non-Lopez
responsables to detect_all_problems_equipos_basicos. The fixture therefore
uses Tipo Factura Descripción="Odontología" with a non-Lopez responsable.
"""
from __future__ import annotations

from unittest.mock import MagicMock, patch

from openpyxl import Workbook

from app.constants import CONVENIO_PYP
from app.services.transversales.column_indices import get_column_indices

# Non-Lopez responsable -> por_responsable routes rows to equipos_basicos.
EQUI_RESP = "Perez Gomez Ana Maria"

EQUI_REQUIRED_HEADERS: dict[str, str] = {
    "numero_factura": "Número Factura",
    "identificacion": "Nº Identificación",
    "convenio_facturado": "Convenio Facturado",
    "codigo": "Código",
    "tipo_procedimiento": "Tipo Procedimiento",
    "cantidad": "Cantidad",
    "vlr_subsidiado": "Vlr. Subsidiado",
    "vlr_procedimiento": "Vlr. Procedimiento",
    "tipo_factura_descripcion": "Tipo Factura Descripción",
    "responsable_cierra": "Responsable Cierra Facturar",
    "centro_costo": "Centro Costo",
    "tipo_identificacion": "Tipo Identificación",
    "fec_factura": "Fec. Factura",
}

# Unified enriches each normalized row with tipo_factura; legacy does not.
UNIFIED_ONLY_KEYS = frozenset({"tipo_factura"})


def _build_equi_workbook() -> Workbook:
    """Minimal EQUI fixture with exact headers (shared ODO rules)."""
    wb = Workbook()
    ws = wb.active
    headers = list(EQUI_REQUIRED_HEADERS.values())
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    rows = [
        ("FAC-EQUI-001", "PAC-201", CONVENIO_PYP, "890203", "Consultas", 1, 1000, 2000),
        ("FAC-EQUI-002", "PAC-201", CONVENIO_PYP, "997002", "Consultas", 2, 1000, 2000),
        ("FAC-EQUI-003", "PAC-202", "Asistencial", "997106", "Preventivo", 1, 1500.75, 2000),
        ("FAC-EQUI-004", "PAC-203", CONVENIO_PYP, "997301", "Preventivo", 12, 1000, 2000),
    ]
    for r, (fact, pac, conv, cod, tipo, cant, vs, vp) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=fact)
        ws.cell(row=r, column=2, value=pac)
        ws.cell(row=r, column=3, value=conv)
        ws.cell(row=r, column=4, value=cod)
        ws.cell(row=r, column=5, value=tipo)
        ws.cell(row=r, column=6, value=cant)
        ws.cell(row=r, column=7, value=vs)
        ws.cell(row=r, column=8, value=vp)
        ws.cell(row=r, column=9, value="Odontología")
        ws.cell(row=r, column=10, value=EQUI_RESP)
        ws.cell(row=r, column=11, value="EQUIPOS BASICOS")
        ws.cell(row=r, column=12, value="CC")
        ws.cell(row=r, column=13, value="2024-01-15")
    return wb


def _indices(wb: Workbook) -> dict[str, int | None]:
    headers = [wb.active.cell(row=1, column=c).value for c in range(1, wb.active.max_column + 1)]
    indices, _ = get_column_indices(headers, EQUI_REQUIRED_HEADERS)
    return indices


def _norm_key(row: dict) -> tuple[str, str, str, str]:
    """Order-insensitive problem identity (ignores unified-only enrichment)."""
    return (
        str(row.get("tipo_error", "")),
        str(row.get("factura", "")),
        str(row.get("descripcion", "")),
        str(row.get("procedimiento", "")),
    )


def _forbidden_session(*args, **kwargs):  # pragma: no cover
    raise AssertionError("parity harness must not open a DB connection")


class TestEquiParity:
    def test_unified_matches_legacy_order_insensitive(self) -> None:
        """Engine (flag OFF, no DB) unified output == legacy equipos detect_all."""
        from app.services.equipos_basicos.detect_all import (
            detect_all_problems_equipos_basicos,
        )
        from app.services.unified_processor import process_unified

        wb = _build_equi_workbook()
        indices = _indices(wb)
        with (
            patch("app.services.equipos_basicos.detect_all.is_rule_engine_enabled", return_value=False),
            patch("app.services.unified_processor.is_rule_engine_enabled", return_value=False),
            patch("app.database.get_session", _forbidden_session),
        ):
            legacy, _ = detect_all_problems_equipos_basicos(wb.active, indices)
            unified, _ = process_unified(wb.active, indices)
        assert unified["area"] == "unificada"
        assert legacy["area"] == "equipos_basicos"
        assert sorted(map(_norm_key, unified["problemas"]["normalizados"])) == sorted(
            map(_norm_key, legacy["problemas"]["normalizados"])
        )
        assert unified["tipos_procesados"] == ["Odontología"]

    def test_engine_mocked_ruta_exception_parity(self) -> None:
        """Engine ON (mocked RuleBasedDetector): ruta_duplicada sets match legacy."""
        from app.services.equipos_basicos.detect_all import (
            detect_all_problems_equipos_basicos,
        )
        from app.services.unified_processor import process_unified

        wb = _build_equi_workbook()
        indices = _indices(wb)
        payload = [
            {"identificacion": "PAC-201", "factura": "FAC-EQUI-001", "cantidad": 3},
            {"identificacion": "PAC-202", "factura": "FAC-EQUI-003", "cantidad": 3},
        ]

        def _mock_detector(name, session, **kwargs):
            detector = MagicMock()
            detector.detect.return_value = list(payload) if name == "ruta_duplicada" else []
            return detector

        with (
            patch("app.services.equipos_basicos.detect_all.is_rule_engine_enabled", return_value=True),
            patch("app.services.unified_processor.is_rule_engine_enabled", return_value=True),
            patch("app.services.engine.session_manager.SessionManager") as mock_session_mgr,
            patch("app.services.engine.rule_based_detector.RuleBasedDetector") as mock_detector_cls,
            patch("app.database.get_session", return_value=MagicMock()),
        ):
            mock_session_mgr.return_value.__enter__.return_value = MagicMock()
            mock_detector_cls.side_effect = _mock_detector
            legacy, _ = detect_all_problems_equipos_basicos(wb.active, indices)
            unified, _ = process_unified(wb.active, indices)
        legacy_ids = {r["identificacion"] for r in legacy["problemas"]["ruta_duplicada"]}
        unified_ids = {r["identificacion"] for r in unified["problemas"]["ruta_duplicada"]}
        assert unified_ids == legacy_ids

    def test_missing_column_tolerance(self) -> None:
        """Fixture without Cantidad: affected detectors return [], never raise."""
        from app.services.equipos_basicos.detect_all import (
            detect_all_problems_equipos_basicos,
        )
        from app.services.transversales.cantidades_anomalas import detect_cantidades_anomalas
        from app.services.unified_processor import process_unified

        wb = _build_equi_workbook()
        indices = _indices(wb)
        indices["cantidad"] = None
        with (
            patch("app.services.equipos_basicos.detect_all.is_rule_engine_enabled", return_value=False),
            patch("app.services.unified_processor.is_rule_engine_enabled", return_value=False),
            patch("app.database.get_session", _forbidden_session),
        ):
            assert detect_cantidades_anomalas(wb.active, indices) == []
            legacy, _ = detect_all_problems_equipos_basicos(wb.active, indices)
            unified, _ = process_unified(wb.active, indices)
        assert legacy["problemas"]["cantidades_anomalas"] == []
        assert sorted(map(_norm_key, unified["problemas"]["normalizados"])) == sorted(
            map(_norm_key, legacy["problemas"]["normalizados"])
        )

    def test_exact_header_matching(self) -> None:
        """Near-miss 'Codigo' must NOT map; exact 'Código' must map."""
        required = {"codigo": "Código"}
        indices, missing = get_column_indices(["Codigo", "Número Factura"], required)
        assert indices["codigo"] is None
        assert "Código" in missing
        indices, missing = get_column_indices(["Código", "Número Factura"], required)
        assert indices["codigo"] == 0
        assert missing == []
