"""URG parity harness (Slice 2, strict TDD).

Unified /procesar (process_unified) must match the legacy urgencias
orchestrator on the same fixture: order-insensitive problem-set compare,
IDE/centro-costo paths, missing-column tolerance, exact header matching.
No engine refactor. No prod connection (engine flag OFF / mocked session).
"""
from __future__ import annotations

from unittest.mock import MagicMock, patch

from openpyxl import Workbook

from app.services.transversales.column_indices import get_column_indices

RESP = "Perez Gomez Ana Maria"
CENTRO_URG = "URGENCIAS"
CENTRO_HOSP = "HOSPITALIZACIÓN - ESTANCIA GENERAL"

# Unified enriches each normalized row with tipo_factura; legacy does not.
UNIFIED_ONLY_KEYS = frozenset({"tipo_factura"})

URG_REQUIRED_HEADERS: dict[str, str] = {
    "numero_factura": "Número Factura",
    "identificacion": "Nº Identificación",
    "codigo": "Código",
    "procedimiento": "Procedimiento",
    "cantidad": "Cantidad",
    "vlr_subsidiado": "Vlr. Subsidiado",
    "vlr_procedimiento": "Vlr. Procedimiento",
    "tipo_factura_descripcion": "Tipo Factura Descripción",
    "responsable_cierra": "Responsable Cierra Facturar",
    "centro_costo": "Centro Costo",
    "codigo_tipo_procedimiento": "Código Tipo Procedimiento",
    "codigo_entidad_cobrar": "Cód Entidad Cobrar",
    "ide_contrato": "IDE Contrato",
    "fec_factura": "Fec. Factura",
    "tarifario": "Tarifario",
    "laboratorio": "Laboratorio",
    "tipo_identificacion": "Tipo Identificación",
}


def _build_urg_workbook() -> Workbook:
    """Minimal URG fixture with exact headers (shared style with ODO slice)."""
    wb = Workbook()
    ws = wb.active
    headers = list(URG_REQUIRED_HEADERS.values())
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    rows = [
        # cantidad violation (890701 qty 2 > 1)
        ("FAC-URG-001", "PAC-101", "890701", "Procedimiento A", 2, 1000, 2000, CENTRO_URG, "986"),
        # centro-costo violation (Urgencias + HOSPITALIZACIÓN)
        ("FAC-URG-002", "PAC-102", "861101", "Procedimiento B", 1, 1000, 2000, CENTRO_HOSP, "986"),
        # IDE violation (906340 + EPSI05 expects 986)
        ("FAC-URG-003", "PAC-103", "906340", "Procedimiento C", 1, 1000, 2000, CENTRO_URG, "000"),
        # clean IDE row (906340 + EPSI05 -> 986)
        ("FAC-URG-004", "PAC-104", "906340", "Procedimiento D", 1, 1000, 2000, CENTRO_URG, "986"),
        # decimales violation (Vlr. Subsidiado 1500.75)
        ("FAC-URG-005", "PAC-105", "890701", "Procedimiento E", 1, 1500.75, 2000, CENTRO_URG, "986"),
    ]
    for r, (fact, pac, cod, proc, cant, vs, vp, centro, ide) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=fact)
        ws.cell(row=r, column=2, value=pac)
        ws.cell(row=r, column=3, value=cod)
        ws.cell(row=r, column=4, value=proc)
        ws.cell(row=r, column=5, value=cant)
        ws.cell(row=r, column=6, value=vs)
        ws.cell(row=r, column=7, value=vp)
        ws.cell(row=r, column=8, value="Urgencias")
        ws.cell(row=r, column=9, value=RESP)
        ws.cell(row=r, column=10, value=centro)
        ws.cell(row=r, column=11, value="09")
        ws.cell(row=r, column=12, value="EPSI05")
        ws.cell(row=r, column=13, value=ide)
        ws.cell(row=r, column=14, value="2024-01-15")
        ws.cell(row=r, column=15, value="Subsidiado")
        ws.cell(row=r, column=16, value="No")
        ws.cell(row=r, column=17, value="CC")
    return wb


def _indices(wb: Workbook) -> dict[str, int | None]:
    headers = [wb.active.cell(row=1, column=c).value for c in range(1, wb.active.max_column + 1)]
    indices, _ = get_column_indices(headers, URG_REQUIRED_HEADERS)
    return indices


def _norm_key(row: dict) -> tuple[str, str, str, str]:
    """Order-insensitive problem identity (ignores unified-only enrichment)."""
    return (
        str(row.get("tipo_error", "")),
        str(row.get("factura", "")),
        str(row.get("descripcion", "")),
        str(row.get("procedimiento", "")),
    )


def _forbidden_session(*args, **kwargs):  # pragma: no cover
    raise AssertionError("parity harness must not open a DB connection")


class TestUrgParity:
    def test_unified_matches_legacy_order_insensitive(self) -> None:
        """Engine (flag OFF, no DB) unified output == legacy detect_all."""
        from app.services.unified_processor import process_unified
        from app.services.urgencias.detect_all import detect_all_problems_urgencias

        wb = _build_urg_workbook()
        indices = _indices(wb)
        with (
            patch("app.services.urgencias.detect_all.is_rule_engine_enabled", return_value=False),
            patch("app.services.unified_processor.is_rule_engine_enabled", return_value=False),
            patch("app.database.get_session", _forbidden_session),
        ):
            legacy, _ = detect_all_problems_urgencias(wb.active, indices)
            unified, _ = process_unified(wb.active, indices)
        assert unified["area"] == "unificada"
        assert legacy["area"] == "urgencias"
        assert unified["tipos_procesados"] == ["Urgencias"]
        assert sorted(map(_norm_key, unified["problemas"]["normalizados"])) == sorted(
            map(_norm_key, legacy["problemas"]["normalizados"])
        )

    def test_ide_centro_costo_detector_paths(self) -> None:
        """Legacy IDE/centro-costo detectors flag wrong rows, spare clean row."""
        from app.services.urgencias.centro_costo_urgencias import detect_centro_costo_urgencias
        from app.services.urgencias.ide_contrato_urgencias import detect_ide_contrato_urgencias

        wb = _build_urg_workbook()
        indices = _indices(wb)
        ide = detect_ide_contrato_urgencias(wb.active, indices)
        centros = detect_centro_costo_urgencias(wb.active, indices)
        ide_facturas = {item["factura"] for item in ide}
        centros_facturas = {item["factura"] for item in centros}
        assert "FAC-URG-003" in ide_facturas
        assert "FAC-URG-004" not in ide_facturas
        assert "FAC-URG-002" in centros_facturas

    def test_engine_mocked_ide_centro_parity(self) -> None:
        """Engine ON (mocked RuleBasedDetector): IDE/centro sets match legacy."""
        from app.services.unified_processor import process_unified
        from app.services.urgencias.detect_all import detect_all_problems_urgencias

        wb = _build_urg_workbook()
        indices = _indices(wb)
        centros_payload = [
            {"factura": "FAC-URG-002", "codigo": "861101", "procedimiento": "Procedimiento B",
             "centro_actual": CENTRO_HOSP, "centro_deberia": CENTRO_URG, "prioridad": 2},
        ]
        ide_payload = [
            {"factura": "FAC-URG-003", "codigo": "906340", "entidad": "EPSI05",
             "ide_contrato_actual": "000", "ide_contrato_deberia": "986"},
        ]

        def _mock_detector(name, session, **kwargs):
            detector = MagicMock()
            if name == "centro_costo_urgencias_valido":
                detector.detect.return_value = list(centros_payload)
            elif name == "ide_contrato_urgencias_valido":
                # NOTE: legacy extends ide_contrato_simple_urgencias on top;
                # payload on exactly one rule name keeps normalized sets 1:1.
                detector.detect.return_value = list(ide_payload)
            else:
                detector.detect.return_value = []
            return detector

        with (
            patch("app.services.urgencias.detect_all.is_rule_engine_enabled", return_value=True),
            patch("app.services.unified_processor.is_rule_engine_enabled", return_value=True),
            patch("app.services.engine.session_manager.SessionManager") as mock_session_mgr,
            patch("app.services.engine.rule_based_detector.RuleBasedDetector") as mock_detector_cls,
            patch("app.database.get_session", return_value=MagicMock()),
        ):
            mock_session_mgr.return_value.__enter__.return_value = MagicMock()
            mock_detector_cls.side_effect = _mock_detector
            legacy, _ = detect_all_problems_urgencias(wb.active, indices)
            unified, _ = process_unified(wb.active, indices)
        assert {i["factura"] for i in legacy["problemas"]["ide_contrato"]} == {"FAC-URG-003"}
        assert {i["factura"] for i in unified["problemas"]["ide_contrato"]} == {"FAC-URG-003"}
        assert sorted(map(_norm_key, unified["problemas"]["normalizados"])) == sorted(
            map(_norm_key, legacy["problemas"]["normalizados"])
        )

    def test_missing_column_tolerance(self) -> None:
        """Fixture without Cantidad/IDE: affected detectors return [], never raise."""
        from app.services.unified_processor import process_unified
        from app.services.urgencias.cantidades_urgencias import detect_cantidades_urgencias
        from app.services.urgencias.detect_all import detect_all_problems_urgencias
        from app.services.urgencias.ide_contrato_urgencias import detect_ide_contrato_urgencias

        wb = _build_urg_workbook()
        indices = _indices(wb)
        indices["cantidad"] = None
        indices["ide_contrato"] = None
        with (
            patch("app.services.urgencias.detect_all.is_rule_engine_enabled", return_value=False),
            patch("app.services.unified_processor.is_rule_engine_enabled", return_value=False),
            patch("app.database.get_session", _forbidden_session),
        ):
            assert detect_cantidades_urgencias(wb.active, indices) == []
            assert detect_ide_contrato_urgencias(wb.active, indices) == []
            legacy, _ = detect_all_problems_urgencias(wb.active, indices)
            unified, _ = process_unified(wb.active, indices)
        assert legacy["problemas"]["cantidades_urgencias"] == []
        assert sorted(map(_norm_key, unified["problemas"]["normalizados"])) == sorted(
            map(_norm_key, legacy["problemas"]["normalizados"])
        )

    def test_exact_header_matching(self) -> None:
        """Near-miss 'Codigo' must NOT map; exact 'Código' must map."""
        required = {"codigo": "Código"}
        indices, missing = get_column_indices(["Codigo", "Número Factura"], required)
        assert indices["codigo"] is None
        assert "Código" in missing
        indices, missing = get_column_indices(["Código", "Número Factura"], required)
        assert indices["codigo"] == 0
        assert missing == []
