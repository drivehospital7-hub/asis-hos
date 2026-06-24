"""Snapshot tests Phase 7: legacy vs engine output for cross-reference rules.

Verifies that the DB-backed engine, when pointed at the same seed SQL
conditions, detects similar issues as the legacy detectors:
  1. codigo_entidad — malformed entidad_afiliacion text
  2. cups_sin_contrato — CUPS not in procedimiento catalog
"""

from __future__ import annotations

import pytest
from unittest.mock import MagicMock
from openpyxl import Workbook


# ── Helpers ─────────────────────────────────────────────────────────────────

def _mock_session_with_rule(rule_name, dominio, descripcion, condiciones_dicts):
    """Create a mock session that returns a Regla + Condicion tree."""
    from app.models import Regla

    session = MagicMock()
    mock_query = MagicMock()
    mock_query.filter.return_value = mock_query
    mock_query.order_by.return_value = mock_query

    regla = Regla(
        id=1, nombre=rule_name, dominio=dominio,
        estado="active", version=1, prioridad=10, severidad="warning",
        descripcion=descripcion,
    )
    mock_query.first.return_value = regla

    cond_mocks = []
    for cd in condiciones_dicts:
        m = MagicMock()
        m.id = cd["id"]
        m.regla_id = cd.get("regla_id", 1)
        m.padre_id = cd["padre_id"]
        m.tipo = cd["tipo"]
        m.operador = cd.get("operador")
        m.fuente_datos = cd.get("fuente_datos")
        m.valor_esperado = cd.get("valor_esperado")
        m.orden = cd.get("orden", 0)
        cond_mocks.append(m)

    mock_query.all.return_value = cond_mocks
    session.query.return_value = mock_query
    return session


def _build_indices(*col_names):
    """Build indices dict: col_name → 0-based index."""
    return {name: i for i, name in enumerate(col_names)}


def _run_engine_detection(rule_name, dominio, descripcion, condiciones, ws, indices):
    """Run engine detection against a worksheet with mocked session."""
    from app.services.engine.rule_based_detector import RuleBasedDetector
    session = _mock_session_with_rule(rule_name, dominio, descripcion, condiciones)
    detector = RuleBasedDetector(rule_name, session)
    return detector.detect(ws, indices)


def _get_facturas_from_results(results):
    """Extract factura strings from detection results."""
    facturas = set()
    for r in results:
        if isinstance(r, dict):
            facturas.add(r.get("factura", ""))
    return facturas


# ── Test: codigo_entidad ────────────────────────────────────────────────────

class TestCodigoEntidad:
    """Engine must detect entidad_afiliacion rows without code pattern ({...})."""

    def test_detects_missing_brace_pattern(self):
        """Entidad Afiliación without '{' → detected."""
        condiciones = [
            {
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "NOT", "orden": 0,
            },
            {
                "id": 2, "padre_id": 1,
                "tipo": "atomic", "operador": "contains",
                "fuente_datos": "invoice.entidad_afiliacion",
                "valor_esperado": "{",
                "orden": 0,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="ENTIDAD_AFILIACION")
        ws.cell(row=1, column=3, value="CODIGO_ENTIDAD_COBRAR")
        # Row 2: malformed — no '{' in entidad_afiliacion
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="EMSSANAR SIN CODIGO Contributivo")
        ws.cell(row=2, column=3, value="ESSC18")
        # Row 3: valid — has '{...}'
        ws.cell(row=3, column=1, value="F002")
        ws.cell(row=3, column=2, value="EMSSANAR - {ESSC18} «Contributivo»")
        ws.cell(row=3, column=3, value="ESSC18")

        indices = _build_indices("numero_factura", "entidad_afiliacion", "codigo_entidad_cobrar")
        results = _run_engine_detection(
            "codigo_entidad", "transversal",
            "Código entidad",
            condiciones, ws, indices,
        )
        facturas = _get_facturas_from_results(results)

        assert "F001" in facturas, "F001 should be detected (no code pattern)"
        assert "F002" not in facturas, "F002 should NOT be detected (has code pattern)"

    def test_empty_entidad_afiliacion_detected(self):
        """Empty entidad_afiliacion → detected (no '{')."""
        condiciones = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "NOT", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "contains",
             "fuente_datos": "invoice.entidad_afiliacion", "valor_esperado": "{", "orden": 0},
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="ENTIDAD_AFILIACION")
        ws.cell(row=2, column=1, value="F003")
        ws.cell(row=2, column=2, value="")

        indices = _build_indices("numero_factura", "entidad_afiliacion")
        results = _run_engine_detection(
            "codigo_entidad", "transversal",
            "Código entidad",
            condiciones, ws, indices,
        )
        facturas = _get_facturas_from_results(results)
        assert "F003" in facturas, "Empty entidad_afiliacion should be detected"

    def test_multiple_rows_mixed_detection(self):
        """Mixed rows: only those without '{' are detected."""
        condiciones = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "NOT", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "contains",
             "fuente_datos": "invoice.entidad_afiliacion", "valor_esperado": "{", "orden": 0},
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="ENTIDAD_AFILIACION")
        ws.cell(row=2, column=1, value="F010")
        ws.cell(row=2, column=2, value="EPS SIN FORMATO")
        ws.cell(row=3, column=1, value="F011")
        ws.cell(row=3, column=2, value="EPS CORRECTA - {EPSS12} «Subsidiado»")
        ws.cell(row=4, column=1, value="F012")
        ws.cell(row=4, column=2, value="OTRA SIN CODIGO")

        indices = _build_indices("numero_factura", "entidad_afiliacion")
        results = _run_engine_detection(
            "codigo_entidad", "transversal",
            "Código entidad",
            condiciones, ws, indices,
        )
        facturas = _get_facturas_from_results(results)

        assert "F010" in facturas
        assert "F011" not in facturas
        assert "F012" in facturas


# ── Test: cups_sin_contrato ─────────────────────────────────────────────────

class TestCupsSinContrato:
    """Engine must detect CUPS that do NOT exist in the procedimiento table."""

    def _make_exists_in_db_conditions(self):
        """Build the NOT(exists_in_db) condition tree."""
        return [
            {
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "NOT", "orden": 0,
            },
            {
                "id": 2, "padre_id": 1,
                "tipo": "atomic", "operador": "exists_in_db",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": {"table": "procedimiento", "field": "cups"},
                "orden": 0,
            },
        ]

    def test_code_exists_no_detection(self):
        """CUPS found in DB → NOT inverts → no detection."""
        from app.services.engine.evaluators import EVALUATOR_REGISTRY
        from app.services.engine.context import EvaluationContext

        evaluator = EVALUATOR_REGISTRY["exists_in_db"]
        evaluator._cache.clear()

        condiciones = self._make_exists_in_db_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=2, column=1, value="F100")
        ws.cell(row=2, column=2, value="990203")

        indices = _build_indices("numero_factura", "codigo")

        # Mock session that returns a row for 990203
        from app.models import Regla
        session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        regla = Regla(id=1, nombre="cups_sin_contrato", dominio="transversal",
                      estado="active", version=1, prioridad=10, severidad="error",
                      descripcion="CUPS sin contrato")
        mock_query.first.return_value = regla

        cond_mocks = []
        for cd in condiciones:
            m = MagicMock()
            m.id = cd["id"]
            m.padre_id = cd["padre_id"]
            m.tipo = cd["tipo"]
            m.operador = cd.get("operador")
            m.fuente_datos = cd.get("fuente_datos")
            m.valor_esperado = cd.get("valor_esperado")
            m.orden = cd.get("orden", 0)
            cond_mocks.append(m)
        mock_query.all.return_value = cond_mocks
        session.query.return_value = mock_query

        # DB session: 990203 exists
        db_session = MagicMock()
        db_session.execute.return_value.fetchone.return_value = (1,)
        session._db_session = db_session  # Not used, the engine uses the session passed to RuleBasedDetector

        from app.services.engine.rule_based_detector import RuleBasedDetector
        detector = RuleBasedDetector("cups_sin_contrato", session)
        results = detector.detect(ws, indices)

        # With the mocked session, the engine creates EvaluationContext with session=session
        # but session here is the mock with query() for models. The exists_in_db
        # evaluator uses context.session.execute() which is on the session mock.
        # Since we mock session (the same one passed to detector), execute() should work.
        facturas = _get_facturas_from_results(results)
        assert "F100" not in facturas, "990203 exists in DB → no detection"

    def test_code_not_exists_detected(self):
        """CUPS not found in DB → NOT inverts → detected."""
        from app.services.engine.evaluators import EVALUATOR_REGISTRY
        evaluator = EVALUATOR_REGISTRY["exists_in_db"]
        evaluator._cache.clear()

        condiciones = self._make_exists_in_db_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=2, column=1, value="F200")
        ws.cell(row=2, column=2, value="ZZZZZZ")

        indices = _build_indices("numero_factura", "codigo")

        from app.models import Regla
        session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        regla = Regla(id=2, nombre="cups_sin_contrato", dominio="transversal",
                      estado="active", version=1, prioridad=10, severidad="error",
                      descripcion="CUPS sin contrato")
        mock_query.first.return_value = regla

        cond_mocks = []
        for cd in condiciones:
            m = MagicMock()
            m.id = cd["id"]
            m.padre_id = cd["padre_id"]
            m.tipo = cd["tipo"]
            m.operador = cd.get("operador")
            m.fuente_datos = cd.get("fuente_datos")
            m.valor_esperado = cd.get("valor_esperado")
            m.orden = cd.get("orden", 0)
            cond_mocks.append(m)
        mock_query.all.return_value = cond_mocks
        session.query.return_value = mock_query

        # DB session mock: returns None (no match)
        session.execute.return_value.fetchone.return_value = None

        from app.services.engine.rule_based_detector import RuleBasedDetector
        detector = RuleBasedDetector("cups_sin_contrato", session)
        results = detector.detect(ws, indices)

        facturas = _get_facturas_from_results(results)
        assert "F200" in facturas, "ZZZZZZ does not exist in DB → detected"

    def test_empty_code_detected_as_missing(self):
        """Empty/None codigo → exists_in_db returns False → NOT → True → detected.

        Note: unlike the legacy detector which skips empty codes, the simplified
        engine version detects them as missing from the catalog. This is acceptable
        for the placeholder implementation.
        """
        condiciones = self._make_exists_in_db_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=2, column=1, value="F300")
        ws.cell(row=2, column=2, value=None)

        indices = _build_indices("numero_factura", "codigo")

        from app.models import Regla
        session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        regla = Regla(id=3, nombre="cups_sin_contrato", dominio="transversal",
                      estado="active", version=1, prioridad=10, severidad="error",
                      descripcion="CUPS sin contrato")
        mock_query.first.return_value = regla

        cond_mocks = []
        for cd in condiciones:
            m = MagicMock()
            m.id = cd["id"]
            m.padre_id = cd["padre_id"]
            m.tipo = cd["tipo"]
            m.operador = cd.get("operador")
            m.fuente_datos = cd.get("fuente_datos")
            m.valor_esperado = cd.get("valor_esperado")
            m.orden = cd.get("orden", 0)
            cond_mocks.append(m)
        mock_query.all.return_value = cond_mocks
        session.query.return_value = mock_query

        from app.services.engine.rule_based_detector import RuleBasedDetector
        detector = RuleBasedDetector("cups_sin_contrato", session)
        results = detector.detect(ws, indices)

        facturas = _get_facturas_from_results(results)
        # Empty code → exists_in_db returns False → NOT inverts → problem detected
        assert "F300" in facturas, "Empty codigo is treated as missing from catalog"

    def test_no_conditions_returns_empty(self):
        """Rule with no conditions returns empty list."""
        condiciones = []

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=2, column=1, value="F400")
        ws.cell(row=2, column=2, value="990203")

        indices = _build_indices("numero_factura", "codigo")

        from app.models import Regla
        session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        regla = Regla(id=4, nombre="cups_sin_contrato", dominio="transversal",
                      estado="active", version=1, prioridad=10, severidad="error",
                      descripcion="CUPS sin contrato")
        mock_query.first.return_value = regla
        mock_query.all.return_value = []
        session.query.return_value = mock_query

        from app.services.engine.rule_based_detector import RuleBasedDetector
        detector = RuleBasedDetector("cups_sin_contrato", session)
        results = detector.detect(ws, indices)
        assert results == []

    def test_codigo_entidad_with_contains_operator(self):
        """Direct contains test: engine uses the contains evaluator correctly."""
        condiciones = [
            {"id": 1, "padre_id": None, "tipo": "atomic", "operador": "contains",
             "fuente_datos": "invoice.entidad_afiliacion", "valor_esperado": "{", "orden": 0},
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="ENTIDAD_AFILIACION")
        ws.cell(row=2, column=1, value="FX01")
        ws.cell(row=2, column=2, value="EPS - {CODE}")

        indices = _build_indices("numero_factura", "entidad_afiliacion")
        results = _run_engine_detection(
            "test_contains", "transversal", "Test contains",
            condiciones, ws, indices,
        )
        facturas = _get_facturas_from_results(results)
        assert "FX01" in facturas, "Contains '{' should match '{CODE}' text"
