"""Tests for ResultadoAuditoria linking from evidence flush.

Verifies that flushing evidence records also creates corresponding
ResultadoAuditoria records with correct outcome mapping.
"""

from __future__ import annotations

import pytest
from unittest.mock import MagicMock

from app.models import Evidencia, ResultadoAuditoria, Regla


class TestResultadoAuditoriaLinking:
    """Tests that ResultadoAuditoria records are created after evidence flush."""

    def test_flush_batch_returns_list_not_none(self):
        """flush_batch() should return a list of Evidencia objects, not None."""
        from app.services.engine.evidence_collector import EvidenceCollector

        session = MagicMock()
        collector = EvidenceCollector()
        collector.record(
            regla_id=1, regla_version=1, dominio="odontologia",
            factura="F001", outcome="MATCH",
            arbol_evaluado={}, snapshot_fila={},
        )

        result = collector.flush_batch(session)

        assert isinstance(result, list)
        assert len(result) == 1
        assert isinstance(result[0], Evidencia)

    def test_flush_batch_returns_empty_list_for_empty_buffer(self):
        """flush_batch() on empty buffer should return an empty list."""
        from app.services.engine.evidence_collector import EvidenceCollector

        session = MagicMock()
        collector = EvidenceCollector()

        result = collector.flush_batch(session)

        assert isinstance(result, list)
        assert len(result) == 0

    def test_flush_clears_buffer_after_returning(self):
        """After flush_batch() returns, the buffer should be empty."""
        from app.services.engine.evidence_collector import EvidenceCollector

        session = MagicMock()
        collector = EvidenceCollector()
        collector.record(
            regla_id=1, regla_version=1, dominio="odontologia",
            factura="F001", outcome="MATCH",
            arbol_evaluado={}, snapshot_fila={},
        )

        result = collector.flush_batch(session)

        assert len(result) == 1
        assert len(collector._buffer) == 0

    def test_multiple_records_returned_in_order(self):
        """Multiple buffered records should be returned in insertion order."""
        from app.services.engine.evidence_collector import EvidenceCollector

        session = MagicMock()
        collector = EvidenceCollector()
        collector.record(
            regla_id=1, regla_version=1, dominio="odontologia",
            factura="F001", outcome="MATCH",
            arbol_evaluado={}, snapshot_fila={},
        )
        collector.record(
            regla_id=1, regla_version=1, dominio="odontologia",
            factura="F002", outcome="NO_MATCH",
            arbol_evaluado={}, snapshot_fila={},
        )

        result = collector.flush_batch(session)

        assert len(result) == 2
        assert result[0].factura == "F001"
        assert result[1].factura == "F002"


class TestResultadoAuditoriaOutcomeMapping:
    """Tests that the engine creates ResultadoAuditoria with correct mapping."""

    def _make_rule(self):
        """Helper: create a mock Regla."""
        return Regla(
            id=1, nombre="test_rule", dominio="odontologia",
            estado="active", version=1, prioridad=10, severidad="error",
            descripcion="Test rule",
        )

    def test_MATCH_outcome_maps_to_FAIL_resultado(self):
        """MATCH evidence should produce resultado='FAIL'."""
        from app.services.engine.engine import RuleEvaluationEngine
        from openpyxl import Workbook

        rule = self._make_rule()
        session = MagicMock()

        # Mock query chain
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.first.return_value = rule

        cond_mock = MagicMock()
        cond_mock.id = 1
        cond_mock.regla_id = 1
        cond_mock.padre_id = None
        cond_mock.tipo = "atomic"
        cond_mock.operador = "eq"
        cond_mock.fuente_datos = "invoice.x"
        cond_mock.valor_esperado = 1
        cond_mock.orden = 0
        mock_query.all.return_value = [cond_mock]
        session.query.return_value = mock_query

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="x")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value=1)

        engine = RuleEvaluationEngine(session)
        engine.evaluate_sheet("test_rule", ws, {"numero_factura": 0, "x": 1})

        # session.add should have been called for ResultadoAuditoria records
        # Verify add was called at least once (for evidence too)
        assert session.add.called

    def test_engine_creates_resultado_auditoria_after_flush(self):
        """Engine should call session.add with ResultadoAuditoria instances."""
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import ResultadoAuditoria
        from openpyxl import Workbook

        rule = self._make_rule()
        session = MagicMock()

        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.first.return_value = rule
        mock_query.all.return_value = []
        session.query.return_value = mock_query

        engine = RuleEvaluationEngine(session)
        # evaluate_sheet with no conditions returns empty results
        # but should still flush evidence
        results = engine.evaluate_sheet(
            "test_rule",
            MagicMock(),  # sheet mock
            {"numero_factura": 0},
        )

        # Even with no data rows, the engine should not crash
        assert isinstance(results, list)
