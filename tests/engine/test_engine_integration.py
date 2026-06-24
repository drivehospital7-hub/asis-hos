"""Integration test for full engine flow — load rule, evaluate, collect evidence."""

from __future__ import annotations

import pytest
from unittest.mock import MagicMock

from app.models import Regla
from app.services.engine.context import EvaluationContext
from app.services.engine.condition_evaluator import ConditionEvaluator
from app.services.engine.evidence_collector import EvidenceCollector
from app.services.engine.exception_handler import ExceptionHandler


class TestFullEngineIntegration:
    """Integration tests exercising multiple engine components together."""

    def test_full_flow_with_mocks(self):
        """Test engine orchestrator with mock DB session."""
        from app.services.engine.engine import RuleEvaluationEngine
        from openpyxl import Workbook

        # Setup mock session
        session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.first.return_value = Regla(
            id=1, nombre="test_decimales", dominio="odontologia",
            estado="active", version=1, prioridad=10, severidad="error",
            descripcion="Test decimal detection",
        )
        # Conditions for the rule
        from unittest.mock import MagicMock as M

        cond1 = M()
        cond1.id = 1
        cond1.regla_id = 1
        cond1.padre_id = None
        cond1.tipo = "atomic"
        cond1.operador = "eq"
        cond1.fuente_datos = "invoice.convenio_facturado"
        cond1.valor_esperado = "PyP"
        cond1.orden = 0

        mock_query.all.return_value = [cond1]
        session.query.return_value = mock_query

        # Build worksheet
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CONVENIO_FACTURADO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="PyP")
        ws.cell(row=3, column=1, value="F002")
        ws.cell(row=3, column=2, value="Asistencial")

        indices = {"numero_factura": 0, "convenio_facturado": 1}

        engine = RuleEvaluationEngine(session)
        results = engine.evaluate_sheet("test_decimales", ws, indices)

        # F001 matches PyP → should be detected
        assert len(results) == 1
        assert results[0]["factura"] == "F001"
        assert results[0]["regla"] == "test_decimales"

    def test_evidence_immutable_after_flush(self):
        """Evidence is insert-only; flush clears buffer but doesn't delete from DB."""
        session = MagicMock()
        collector = EvidenceCollector()
        collector.record(
            regla_id=1, regla_version=1, dominio="odontologia",
            factura="F001", outcome="MATCH",
            arbol_evaluado={}, snapshot_fila={},
        )
        collector.flush_batch(session)
        # Buffer should be empty (flushed)
        assert len(collector._buffer) == 0
        # session.add_all was called (evidence persisted)
        session.add_all.assert_called_once()
        session.flush.assert_called_once()

    def test_error_tree_produces_error_outcome(self):
        """Unknown operator in condition tree → ERROR outcome, no crash."""
        evaluator = ConditionEvaluator()
        ctx = EvaluationContext(invoice_data={"x": 1})
        tree = {
            "tipo": "atomic",
            "operador": "nonexistent_op",
            "fuente_datos": "invoice.x",
            "valor_esperado": 1,
            "_children": [],
        }
        result = evaluator.evaluate(tree, ctx)
        assert result["outcome"] is False
        assert result.get("error") is not None

    def test_skip_exception_prevents_evaluation(self):
        """When exception returns 'skip', the engine skips evaluation for that row."""
        handler = ExceptionHandler()
        from app.models import Regla, Excepcion

        session = MagicMock()
        exc = Excepcion(
            id=1, regla_id=1, tipo_efecto="skip",
            condicion_json={"convenio": "PyP"},
            activo=True,
        )
        mock_q = MagicMock()
        mock_q.filter.return_value = mock_q
        mock_q.all.return_value = [exc]
        session.query.return_value = mock_q

        rule = Regla(id=1, nombre="test", dominio="odontologia", estado="active", prioridad=10)
        ctx = EvaluationContext(invoice_data={"convenio": "PyP"})

        effect, overrides = handler.apply_exceptions(rule, ctx, session)
        assert effect == "skip"
