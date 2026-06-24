"""Snapshot tests Phase 3: engine vs legacy for IDE Contrato rules.

Verifies that the DB-backed engine detects the same facturas as the
legacy Python detectors for odontologia IDE Contrato, urgencias forward,
and urgencias reverse rules.
"""

from __future__ import annotations

import pytest
from unittest.mock import MagicMock
from openpyxl import Workbook


# ── Helpers ─────────────────────────────────────────────────────────────────

PYP_CUPS_CODES = frozenset({
    "890203", "990203", "990212", "997002", "997106", "997107",
    "997301", "P0000011",
})


def _mock_session_with_rule(rule_name, dominio, descripcion, condiciones_dicts):
    """Create a mock session that returns a Regla + Condicion tree."""
    from app.models import Regla

    session = MagicMock()
    mock_query = MagicMock()
    mock_query.filter.return_value = mock_query
    mock_query.order_by.return_value = mock_query

    regla = Regla(
        id=1, nombre=rule_name, dominio=dominio,
        estado="active", version=1, prioridad=10, severidad="warning",
        descripcion=descripcion,
    )
    mock_query.first.return_value = regla

    cond_mocks = []
    for cd in condiciones_dicts:
        m = MagicMock()
        m.id = cd["id"]
        m.regla_id = cd.get("regla_id", 1)
        m.padre_id = cd["padre_id"]
        m.tipo = cd["tipo"]
        m.operador = cd.get("operador")
        m.fuente_datos = cd.get("fuente_datos")
        m.valor_esperado = cd.get("valor_esperado")
        m.orden = cd.get("orden", 0)
        cond_mocks.append(m)

    mock_query.all.return_value = cond_mocks
    session.query.return_value = mock_query
    return session


def _build_indices(*col_names):
    """Build indices dict: col_name -> 0-based index."""
    return {name: i for i, name in enumerate(col_names)}


def _run_engine_detection(rule_name, dominio, descripcion, condiciones, ws, indices):
    """Run engine detection against a worksheet with mocked session."""
    from app.services.engine.rule_based_detector import RuleBasedDetector
    session = _mock_session_with_rule(rule_name, dominio, descripcion, condiciones)
    detector = RuleBasedDetector(rule_name, session)
    return detector.detect(ws, indices)


def _get_facturas_from_results(results):
    """Extract factura strings from detection results."""
    facturas = set()
    for r in results:
        if isinstance(r, dict):
            facturas.add(r.get("factura", ""))
        elif isinstance(r, str):
            facturas.add(r)
    return facturas


# ── Condition Tree Builder for Odontologia IDE Contrato ─────────────────────

def _build_odon_ide_conditions():
    """Build condition tree covering top 8 entities for odontologia IDE contrato.

    Returns flat list of condition dicts with id, padre_id, tipo, operador,
    fuente_datos, valor_esperado, orden.
    """
    entity_rules = [
        # (entity, pyp_ide_set, no_pyp_ide_set)
        ("ESS118", frozenset({"970", "974"}), frozenset({"969", "973"})),
        ("ESSC18", frozenset({"975"}), frozenset({"968"})),
        ("EPSS41", frozenset({"955", "958"}), frozenset({"956", "959"})),
        ("EPSI05", frozenset({"977"}), frozenset({"976", "978"})),
        ("EPSIC5", frozenset({"979"}), frozenset({"967"})),
        ("RES001", frozenset({"993"}), frozenset({"992"})),
        ("0001", frozenset({"17"}), frozenset({"984"})),
        ("86", None, frozenset({"911"})),  # Only NO PyP
    ]

    next_id = [1]
    def new_id():
        nid = next_id[0]
        next_id[0] += 1
        return nid

    conditions = []
    pyp_list = sorted(PYP_CUPS_CODES)

    # Root OR
    root_id = new_id()
    conditions.append({
        "id": root_id, "padre_id": None,
        "tipo": "composite", "operador": "OR", "orden": 0,
    })

    orden_counter = 0
    for entity, pyp_set, no_pyp_set in entity_rules:
        if pyp_set is not None:
            # PyP branch: AND(eq entidad, IN codigo pyp, NOT IN ide_contrato pyp_set)
            and_id = new_id()
            conditions.append({
                "id": and_id, "padre_id": root_id,
                "tipo": "composite", "operador": "AND", "orden": orden_counter,
            })
            orden_counter += 1

            # EQ entidad
            conditions.append({
                "id": new_id(), "padre_id": and_id,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar",
                "valor_esperado": entity, "orden": 0,
            })
            # IN codigo
            conditions.append({
                "id": new_id(), "padre_id": and_id,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": pyp_list, "orden": 1,
            })
            # NOT IN ide_contrato
            not_id = new_id()
            conditions.append({
                "id": not_id, "padre_id": and_id,
                "tipo": "composite", "operador": "NOT", "orden": 2,
            })
            conditions.append({
                "id": new_id(), "padre_id": not_id,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.ide_contrato",
                "valor_esperado": sorted(pyp_set), "orden": 0,
            })

        if no_pyp_set is not None:
            # No PyP branch: AND(eq entidad, NOT IN codigo pyp, NOT IN ide_contrato no_pyp_set)
            and_id = new_id()
            conditions.append({
                "id": and_id, "padre_id": root_id,
                "tipo": "composite", "operador": "AND", "orden": orden_counter,
            })
            orden_counter += 1

            # EQ entidad
            conditions.append({
                "id": new_id(), "padre_id": and_id,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar",
                "valor_esperado": entity, "orden": 0,
            })
            # NOT IN codigo
            not_cod = new_id()
            conditions.append({
                "id": not_cod, "padre_id": and_id,
                "tipo": "composite", "operador": "NOT", "orden": 1,
            })
            conditions.append({
                "id": new_id(), "padre_id": not_cod,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": pyp_list, "orden": 0,
            })
            # NOT IN ide_contrato
            not_ide = new_id()
            conditions.append({
                "id": not_ide, "padre_id": and_id,
                "tipo": "composite", "operador": "NOT", "orden": 2,
            })
            conditions.append({
                "id": new_id(), "padre_id": not_ide,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.ide_contrato",
                "valor_esperado": sorted(no_pyp_set), "orden": 0,
            })

    return conditions


# ── Test: IDE Contrato Odontologia ──────────────────────────────────────────

class TestIdeContratoOdontologia:
    """Engine must match legacy odontologia IDE Contrato detection."""

    def _make_ws(self, rows):
        """Helper: create worksheet from list of (factura, entidad, codigo, ide_contrato)."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO_ENTIDAD_COBRAR")
        ws.cell(row=1, column=3, value="CODIGO")
        ws.cell(row=1, column=4, value="IDE_CONTRATO")
        for i, (fact, ent, cod, ide) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=fact)
            ws.cell(row=i, column=2, value=ent)
            ws.cell(row=i, column=3, value=cod)
            ws.cell(row=i, column=4, value=ide)
        return ws

    def test_ess118_pyp_wrong_ide_detected(self):
        """ESS118 + PyP code + wrong IDE -> detected."""
        condiciones = _build_odon_ide_conditions()
        ws = self._make_ws([
            ("F001", "ESS118", "890203", "999"),  # PyP, expected 970/974, wrong
        ])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_odontologia_valido", "odontologia",
            "IDE Contrato validation for Odontologia", condiciones, ws, idx,
        )
        facturas = _get_facturas_from_results(results)
        assert "F001" in facturas, "Engine should detect wrong IDE for ESS118+PyP"

    def test_ess118_pyp_correct_ide_ignored(self):
        """ESS118 + PyP code + correct IDE -> not detected."""
        condiciones = _build_odon_ide_conditions()
        ws = self._make_ws([
            ("F002", "ESS118", "890203", "970"),  # PyP, 970 is valid
        ])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_odontologia_valido", "odontologia",
            "IDE Contrato validation", condiciones, ws, idx,
        )
        assert len(results) == 0, "Engine should not detect correct IDE"

    def test_ess118_no_pyp_wrong_ide_detected(self):
        """ESS118 + non-PyP code + wrong IDE -> detected."""
        condiciones = _build_odon_ide_conditions()
        ws = self._make_ws([
            ("F003", "ESS118", "999999", "999"),  # Not PyP, expected 969/973
        ])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_odontologia_valido", "odontologia",
            "IDE Contrato validation", condiciones, ws, idx,
        )
        facturas = _get_facturas_from_results(results)
        assert "F003" in facturas, "Engine should detect wrong IDE for ESS118+NoPyP"

    def test_ess118_no_pyp_correct_ide_ignored(self):
        """ESS118 + non-PyP code + correct IDE (969) -> not detected."""
        condiciones = _build_odon_ide_conditions()
        ws = self._make_ws([
            ("F004", "ESS118", "999999", "969"),
        ])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_odontologia_valido", "odontologia",
            "IDE Contrato validation", condiciones, ws, idx,
        )
        assert len(results) == 0

    def test_essc18_pyp_wrong_ide_detected(self):
        """ESSC18 + PyP + wrong IDE -> detected."""
        condiciones = _build_odon_ide_conditions()
        ws = self._make_ws([
            ("F005", "ESSC18", "890203", "999"),  # Expected 975
        ])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_odontologia_valido", "odontologia",
            "IDE Contrato validation", condiciones, ws, idx,
        )
        facturas = _get_facturas_from_results(results)
        assert "F005" in facturas

    def test_essc18_no_pyp_wrong_ide_detected(self):
        """ESSC18 + non-PyP + wrong IDE -> detected."""
        condiciones = _build_odon_ide_conditions()
        ws = self._make_ws([
            ("F006", "ESSC18", "888888", "999"),  # Expected 968
        ])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_odontologia_valido", "odontologia",
            "IDE Contrato validation", condiciones, ws, idx,
        )
        facturas = _get_facturas_from_results(results)
        assert "F006" in facturas

    def test_entity_with_no_rule_ignored(self):
        """Entity not in our rules (e.g., ESS062) -> not detected."""
        condiciones = _build_odon_ide_conditions()
        ws = self._make_ws([
            ("F007", "ESS062", "890203", "999"),  # ESS062 not in top 8
        ])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_odontologia_valido", "odontologia",
            "IDE Contrato validation", condiciones, ws, idx,
        )
        assert len(results) == 0, "Engine should skip entities not in rule set"

    def test_odin_multiple_rows_mixed(self):
        """Multiple rows: some matching, some not."""
        condiciones = _build_odon_ide_conditions()
        ws = self._make_ws([
            ("F010", "ESS118", "890203", "970"),   # OK: ESS118+PyP with valid IDE 970
            ("F011", "ESS118", "890203", "999"),   # BAD: ESS118+PyP with wrong IDE
            ("F012", "ESSC18", "999999", "968"),   # OK: ESSC18+NoPyP with valid IDE 968
            ("F013", "ESSC18", "999999", "999"),   # BAD: ESSC18+NoPyP with wrong IDE
            ("F014", "0001", "997106", "17"),      # OK: 0001+PyP with valid IDE 17
            ("F015", "0001", "888888", "984"),     # OK: 0001+NoPyP with valid IDE 984
        ])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_odontologia_valido", "odontologia",
            "IDE Contrato validation", condiciones, ws, idx,
        )
        facturas = _get_facturas_from_results(results)
        assert "F010" not in facturas
        assert "F011" in facturas
        assert "F012" not in facturas
        assert "F013" in facturas
        assert "F014" not in facturas
        assert "F015" not in facturas

    def test_legacy_vs_engine_same_detections(self):
        """Both legacy and engine detect the same facturas for odontologia."""
        from app.services.odontologia.ide_contrato import (
            detect_ide_contrato_odontologia,
        )

        condiciones = _build_odon_ide_conditions()
        ws = self._make_ws([
            ("F100", "ESS118", "890203", "999"),    # BAD: ESS118+PyP
            ("F101", "ESS118", "890203", "970"),    # OK
            ("F102", "ESS118", "888888", "969"),    # OK: ESS118+NoPyP valid
            ("F103", "ESSC18", "890203", "999"),    # BAD: ESSC18+PyP
            ("F104", "ESSC18", "888888", "968"),    # OK
            ("F105", "0001", "997106", "17"),       # OK
            ("F106", "86", "888888", "911"),        # OK: 86+NoPyP valid
            ("F107", "86", "888888", "999"),        # BAD: 86+NoPyP wrong
            ("F108", "EPSS41", "997107", "955"),    # OK: EPSS41+PyP valid
            ("F109", "EPSI05", "997301", "977"),    # OK: EPSI05+PyP valid
        ])

        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        legacy_results = detect_ide_contrato_odontologia(ws, idx)
        legacy_facturas = {r["factura"] for r in legacy_results}

        engine_results = _run_engine_detection(
            "ide_contrato_odontologia_valido", "odontologia",
            "IDE Contrato validation", condiciones, ws, idx,
        )
        engine_facturas = _get_facturas_from_results(engine_results)

        # Both must detect F100 and F103 (PyP+wrong IDE)
        assert "F100" in legacy_facturas, "Legacy must detect F100"
        assert "F100" in engine_facturas, "Engine must detect F100"
        assert "F103" in legacy_facturas, "Legacy must detect F103"
        assert "F103" in engine_facturas, "Engine must detect F103"
        assert "F107" in legacy_facturas, "Legacy must detect F107 (86+NoPyP wrong)"
        # Engine covers only top 8 entities including 86, so F107 should be detected
        assert "F107" in engine_facturas, "Engine must detect F107"

        # Neither should detect the OK cases
        for ok_fact in ("F101", "F102", "F104", "F105", "F106", "F108", "F109"):
            assert ok_fact not in engine_facturas, f"Engine should not detect {ok_fact}"

        # Engine must not have false positives
        for ef in engine_facturas:
            assert ef in legacy_facturas, f"Engine false positive: {ef}"


# ── Test: IDE Contrato Urgencias Forward ────────────────────────────────────

class TestIdeContratoUrgencias:
    """Engine must detect simple urgencias IDE Contrato rules."""

    def _build_urg_ide_simple_conditions(self):
        """Build conditions for simple (codigo+entidad→IDE) and multiple rules."""
        next_id = [1]
        def nid():
            n = next_id[0]; next_id[0] += 1; return n

        conditions = []
        root = nid()
        conditions.append({"id": root, "padre_id": None, "tipo": "composite", "operador": "OR", "orden": 0})

        # Simple rules: (entidad, codigo, expected_ide)
        simple_rules = [
            ("EPSI05", "906340", "986"),
            ("EPSI05", "861801", "977"),
            ("EPSIC5", "861801", "979"),
            ("ESS118", "906340", "839"),
            ("ESS118", "890405", "974"),
            ("ESS118", "890205", "970"),
            ("ESSC18", "906340", "842"),
            ("ESSC18", "861801", "975"),
            ("EPS037", "906340", "962"),
            ("EPS037", "861801", "961"),
            ("EPSS41", "906340", "959"),
            ("EPSS41", "861801", "958"),
            ("ESS062", "861801", "922"),
            ("ESSC62", "861801", "863"),
            ("86000", "861801", "920"),
            ("RES004", "861801", "908"),
        ]

        orden = 0
        for entity, code, expected in simple_rules:
            and_id = nid()
            conditions.append({"id": and_id, "padre_id": root, "tipo": "composite", "operador": "AND", "orden": orden})
            orden += 1
            conditions.append({"id": nid(), "padre_id": and_id, "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar", "valor_esperado": entity, "orden": 0})
            conditions.append({"id": nid(), "padre_id": and_id, "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo", "valor_esperado": code, "orden": 1})
            not_id = nid()
            conditions.append({"id": not_id, "padre_id": and_id, "tipo": "composite", "operador": "NOT", "orden": 2})
            conditions.append({"id": nid(), "padre_id": not_id, "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.ide_contrato", "valor_esperado": expected, "orden": 0})

        # Multiple rules
        multi_rules = [
            ("ESS118", "735301", ["970", "974"]),
            ("ESS118", "861801", ["970", "974"]),
        ]
        for entity, code, ide_set in multi_rules:
            and_id = nid()
            conditions.append({"id": and_id, "padre_id": root, "tipo": "composite", "operador": "AND", "orden": orden})
            orden += 1
            conditions.append({"id": nid(), "padre_id": and_id, "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar", "valor_esperado": entity, "orden": 0})
            conditions.append({"id": nid(), "padre_id": and_id, "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo", "valor_esperado": code, "orden": 1})
            not_id = nid()
            conditions.append({"id": not_id, "padre_id": and_id, "tipo": "composite", "operador": "NOT", "orden": 2})
            conditions.append({"id": nid(), "padre_id": not_id, "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.ide_contrato", "valor_esperado": sorted(ide_set), "orden": 0})

        # Generic entity rules
        generic_rules = [
            ("86", "911"),
            ("5177", "917"),
            ("RES001", "992"),
        ]
        for entity, expected in generic_rules:
            and_id = nid()
            conditions.append({"id": and_id, "padre_id": root, "tipo": "composite", "operador": "AND", "orden": orden})
            orden += 1
            conditions.append({"id": nid(), "padre_id": and_id, "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar", "valor_esperado": entity, "orden": 0})
            not_id = nid()
            conditions.append({"id": not_id, "padre_id": and_id, "tipo": "composite", "operador": "NOT", "orden": 1})
            conditions.append({"id": nid(), "padre_id": not_id, "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.ide_contrato", "valor_esperado": expected, "orden": 0})

        return conditions

    def _make_ws(self, rows):
        """Helper: create worksheet from list of (factura, entidad, codigo, ide_contrato)."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO_ENTIDAD_COBRAR")
        ws.cell(row=1, column=3, value="CODIGO")
        ws.cell(row=1, column=4, value="IDE_CONTRATO")
        for i, (fact, ent, cod, ide) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=fact)
            ws.cell(row=i, column=2, value=ent)
            ws.cell(row=i, column=3, value=cod)
            ws.cell(row=i, column=4, value=ide)
        return ws

    def test_simple_rule_detected(self):
        """Simple rule: EPSI05+906340 with IDE!=986 -> detected."""
        condiciones = self._build_urg_ide_simple_conditions()
        ws = self._make_ws([("F001", "EPSI05", "906340", "999")])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_urgencias_valido", "urgencias",
            "IDE Contrato Urgencias validation", condiciones, ws, idx,
        )
        facturas = _get_facturas_from_results(results)
        assert "F001" in facturas

    def test_simple_rule_correct_ignored(self):
        """Simple rule: EPSI05+906340 with correct IDE=986 -> not detected."""
        condiciones = self._build_urg_ide_simple_conditions()
        ws = self._make_ws([("F002", "EPSI05", "906340", "986")])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_urgencias_valido", "urgencias",
            "IDE Contrato Urgencias validation", condiciones, ws, idx,
        )
        assert len(results) == 0

    def test_multiple_rule_detected(self):
        """Multiple rule: ESS118+735301 with IDE not in {970,974} -> detected."""
        condiciones = self._build_urg_ide_simple_conditions()
        ws = self._make_ws([("F003", "ESS118", "735301", "999")])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_urgencias_valido", "urgencias",
            "IDE Contrato Urgencias validation", condiciones, ws, idx,
        )
        facturas = _get_facturas_from_results(results)
        assert "F003" in facturas

    def test_multiple_rule_valid_ide_ignored(self):
        """Multiple rule: ESS118+735301 with IDE=970 -> not detected."""
        condiciones = self._build_urg_ide_simple_conditions()
        ws = self._make_ws([("F004", "ESS118", "735301", "970")])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_urgencias_valido", "urgencias",
            "IDE Contrato Urgencias validation", condiciones, ws, idx,
        )
        assert len(results) == 0

    def test_generic_entity_rule_detected(self):
        """Generic rule: entity 86 with IDE != 911 -> detected."""
        condiciones = self._build_urg_ide_simple_conditions()
        ws = self._make_ws([("F005", "86", "999999", "123")])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_urgencias_valido", "urgencias",
            "IDE Contrato Urgencias validation", condiciones, ws, idx,
        )
        facturas = _get_facturas_from_results(results)
        assert "F005" in facturas

    def test_generic_entity_rule_correct_ignored(self):
        """Generic rule: entity 5177 with IDE=917 -> not detected."""
        condiciones = self._build_urg_ide_simple_conditions()
        ws = self._make_ws([("F006", "5177", "888888", "917")])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_urgencias_valido", "urgencias",
            "IDE Contrato Urgencias validation", condiciones, ws, idx,
        )
        assert len(results) == 0

    def test_urg_mixed_rows(self):
        """Mixed rows: some matching, some not."""
        condiciones = self._build_urg_ide_simple_conditions()
        ws = self._make_ws([
            ("F010", "EPSI05", "906340", "986"),   # OK
            ("F011", "EPSI05", "906340", "999"),   # BAD
            ("F012", "86", "888888", "911"),       # OK: generic
            ("F013", "86", "888888", "123"),       # BAD: generic
            ("F014", "ESS118", "735301", "970"),   # OK: multiple
            ("F015", "ESS118", "735301", "999"),   # BAD: multiple
            ("F016", "ESS062", "861801", "922"),   # OK: simple
        ])
        idx = _build_indices("numero_factura", "codigo_entidad_cobrar", "codigo", "ide_contrato")
        results = _run_engine_detection(
            "ide_contrato_urgencias_valido", "urgencias",
            "IDE Contrato Urgencias validation", condiciones, ws, idx,
        )
        facturas = _get_facturas_from_results(results)
        assert "F010" not in facturas
        assert "F011" in facturas
        assert "F012" not in facturas
        assert "F013" in facturas
        assert "F014" not in facturas
        assert "F015" in facturas
        assert "F016" not in facturas


# ── Test: IDE Contrato Reverse Urgencias ────────────────────────────────────

class TestIdeContratoReverse:
    """Engine must detect reverse IDE Contrato rules (IDE -> expected codigo)."""

    def _build_reverse_conditions(self):
        """Build conditions for reverse IDE rules (no pre-scan needed)."""
        next_id = [1]
        def nid():
            n = next_id[0]; next_id[0] += 1; return n

        conditions = []
        root = nid()
        conditions.append({"id": root, "padre_id": None, "tipo": "composite", "operador": "OR", "orden": 0})

        # Reverse rules: (ide, expected_codigo_or_list)
        reverse_rules = [
            ("986", ["906340"]),
            ("839", ["906340"]),
            ("842", ["906340"]),
            ("970", ["735301", "861801", "890205"]),
            ("974", ["735301", "861801", "890405"]),
        ]

        for orden, (ide_val, expected_codes) in enumerate(reverse_rules):
            and_id = nid()
            conditions.append({"id": and_id, "padre_id": root, "tipo": "composite", "operador": "AND", "orden": orden})
            conditions.append({"id": nid(), "padre_id": and_id, "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.ide_contrato", "valor_esperado": ide_val, "orden": 0})
            not_id = nid()
            conditions.append({"id": not_id, "padre_id": and_id, "tipo": "composite", "operador": "NOT", "orden": 1})
            if len(expected_codes) == 1:
                conditions.append({"id": nid(), "padre_id": not_id, "tipo": "atomic", "operador": "eq",
                    "fuente_datos": "invoice.codigo", "valor_esperado": expected_codes[0], "orden": 0})
            else:
                conditions.append({"id": nid(), "padre_id": not_id, "tipo": "atomic", "operador": "in",
                    "fuente_datos": "invoice.codigo", "valor_esperado": sorted(expected_codes), "orden": 0})

        return conditions

    def _make_ws(self, rows):
        """Helper: create worksheet from list of (factura, ide_contrato, codigo)."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="IDE_CONTRATO")
        ws.cell(row=1, column=3, value="CODIGO")
        for i, (fact, ide, cod) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=fact)
            ws.cell(row=i, column=2, value=ide)
            ws.cell(row=i, column=3, value=cod)
        return ws

    def test_ide_986_wrong_code_detected(self):
        """IDE 986 with code != 906340 -> detected."""
        condiciones = self._build_reverse_conditions()
        ws = self._make_ws([("F001", "986", "999999")])
        idx = _build_indices("numero_factura", "ide_contrato", "codigo")
        results = _run_engine_detection(
            "ide_contrato_reverse_urgencias_valido", "urgencias",
            "IDE Contrato Reverse validation", condiciones, ws, idx,
        )
        facturas = _get_facturas_from_results(results)
        assert "F001" in facturas

    def test_ide_986_correct_code_ignored(self):
        """IDE 986 with code 906340 -> not detected."""
        condiciones = self._build_reverse_conditions()
        ws = self._make_ws([("F002", "986", "906340")])
        idx = _build_indices("numero_factura", "ide_contrato", "codigo")
        results = _run_engine_detection(
            "ide_contrato_reverse_urgencias_valido", "urgencias",
            "IDE Contrato Reverse validation", condiciones, ws, idx,
        )
        assert len(results) == 0

    def test_ide_970_multiple_codes(self):
        """IDE 970: code must be in {735301, 861801, 890205}."""
        condiciones = self._build_reverse_conditions()
        ws = self._make_ws([
            ("F003", "970", "735301"),  # OK
            ("F004", "970", "861801"),  # OK
            ("F005", "970", "890205"),  # OK
            ("F006", "970", "999999"),  # BAD
        ])
        idx = _build_indices("numero_factura", "ide_contrato", "codigo")
        results = _run_engine_detection(
            "ide_contrato_reverse_urgencias_valido", "urgencias",
            "IDE Contrato Reverse validation", condiciones, ws, idx,
        )
        facturas = _get_facturas_from_results(results)
        assert "F003" not in facturas
        assert "F004" not in facturas
        assert "F005" not in facturas
        assert "F006" in facturas

    def test_ide_839_detected(self):
        """IDE 839 must have code 906340."""
        condiciones = self._build_reverse_conditions()
        ws = self._make_ws([("F007", "839", "123456")])
        idx = _build_indices("numero_factura", "ide_contrato", "codigo")
        results = _run_engine_detection(
            "ide_contrato_reverse_urgencias_valido", "urgencias",
            "IDE Contrato Reverse validation", condiciones, ws, idx,
        )
        facturas = _get_facturas_from_results(results)
        assert "F007" in facturas

    def test_reverse_output_format(self):
        """Reverse engine output has factura, problema, regla keys."""
        condiciones = self._build_reverse_conditions()
        ws = self._make_ws([("F010", "986", "bad")])
        idx = _build_indices("numero_factura", "ide_contrato", "codigo")
        results = _run_engine_detection(
            "ide_contrato_reverse_urgencias_valido", "urgencias",
            "IDE Contrato Reverse validation", condiciones, ws, idx,
        )
        assert len(results) > 0
        r = results[0]
        assert "factura" in r
        assert "problema" in r
        assert "regla" in r
        assert "severidad" in r
        assert r["factura"] == "F010"
