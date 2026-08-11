"""Tests for PymRutasDxEvaluator — operator 'pym_rutas_dx_check'.

Replaces the PYM_RUTAS + Dx + pre-scan part of detect_ide_contrato_intramural().
Tests use mock data_sheet for pre_scan and direct evaluate calls.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.engine.context import EvaluationContext


class TestPymRutasDxEvaluator:
    """PymRutasDxEvaluator — PYM_RUTAS + Dx + pre-scan validation."""

    @pytest.fixture
    def evaluator(self):
        from app.services.engine.evaluators import PymRutasDxEvaluator
        return PymRutasDxEvaluator()

    def test_operator_name(self):
        """Operator name should be 'pym_rutas_dx_check'."""
        from app.services.engine.evaluators import PymRutasDxEvaluator
        assert PymRutasDxEvaluator.operator == "pym_rutas_dx_check"

    def test_register_in_registry(self):
        """Should be registered in EVALUATOR_REGISTRY."""
        from app.services.engine.evaluators import EVALUATOR_REGISTRY
        assert "pym_rutas_dx_check" in EVALUATOR_REGISTRY

    def test_not_intramural_returns_true(self, evaluator):
        """Non-Intramural tipo factura → skip validation (True)."""
        ctx = EvaluationContext(invoice_data={
            "tipo_factura_descripcion": "Urgencias",
            "codigo_entidad_cobrar": "EPSS41",
        })
        result = evaluator.evaluate({}, "906127", "999", context=ctx)
        assert result is True

    def test_none_row_value_returns_true(self, evaluator):
        """None codigo → skip validation (True)."""
        ctx = EvaluationContext(invoice_data={
            "tipo_factura_descripcion": "Intramural",
        })
        result = evaluator.evaluate({}, None, "999", context=ctx)
        assert result is True

    def test_no_context_returns_false(self, evaluator):
        """No context → cannot validate (False)."""
        result = evaluator.evaluate({}, "906127", "999", context=None)
        assert result is False

    def test_no_entity_returns_true(self, evaluator):
        """Empty codigo_entidad_cobrar → skip validation (True)."""
        ctx = EvaluationContext(invoice_data={
            "tipo_factura_descripcion": "Intramural",
            "codigo_entidad_cobrar": "",
        })
        result = evaluator.evaluate({}, "906127", "999", context=ctx)
        assert result is True

    def test_codigo_not_in_pym_rutas_returns_true(self, evaluator):
        """Codigo not in PYM_RUTAS set → skip (True)."""
        ctx = EvaluationContext(invoice_data={
            "tipo_factura_descripcion": "Intramural",
            "codigo_entidad_cobrar": "EPSS41",
        })
        result = evaluator.evaluate({}, "999999", "999", context=ctx)
        assert result is True

    def test_no_dx_principal_returns_true(self, evaluator):
        """No Dx principal → skip validation (True)."""
        ctx = EvaluationContext(invoice_data={
            "tipo_factura_descripcion": "Intramural",
            "codigo_entidad_cobrar": "EPSS41",
            "codigo_dx_principal": "",
        })
        # 903841 (Glucosa en Suero) is in CODIGOS_PYM_RUTAS
        result = evaluator.evaluate({}, "903841", "999", context=ctx)
        assert result is True

    def test_dx_not_in_necesitan_returns_true(self, evaluator):
        """Dx not in NECESITAN_DX → skip (True)."""
        ctx = EvaluationContext(invoice_data={
            "tipo_factura_descripcion": "Intramural",
            "codigo_entidad_cobrar": "EPSS41",
            "codigo_dx_principal": "A000",
        })
        result = evaluator.evaluate({}, "903841", "999", context=ctx)
        assert result is True

    def test_entity_not_in_map_returns_true(self, evaluator):
        """Entity without PYM_RUTAS mapping → skip (True)."""
        ctx = EvaluationContext(invoice_data={
            "tipo_factura_descripcion": "Intramural",
            "codigo_entidad_cobrar": "UNKNOWN",
            "codigo_dx_principal": "Z359",
        })
        result = evaluator.evaluate({}, "906127", "999", context=ctx)
        assert result is True

    def test_correct_ide_returns_true(self, evaluator):
        """Code+Entity+Dx match PYM_RUTAS and IDE is in valid set → True."""
        # 903841 (Glucosa en Suero) is in CODIGOS_PYM_RUTAS
        evaluator._pre_scan_cache = {"F001": True}  # has non-lab codes
        ctx = EvaluationContext(invoice_data={
            "tipo_factura_descripcion": "Intramural",
            "codigo_entidad_cobrar": "EPSS41",
            "codigo_dx_principal": "Z359",
            "numero_factura": "F001",
        })
        # EPSS41 PYM_RUTAS_IDE_MAP = {"955"}
        result = evaluator.evaluate({}, "903841", "955", context=ctx)
        assert result is True

    def test_wrong_ide_returns_false(self, evaluator):
        """Code+Entity+Dx match PYM_RUTAS but IDE is wrong → False."""
        evaluator._pre_scan_cache = {"F001": True}
        ctx = EvaluationContext(invoice_data={
            "tipo_factura_descripcion": "Intramural",
            "codigo_entidad_cobrar": "EPSS41",
            "codigo_dx_principal": "Z359",
            "numero_factura": "F001",
        })
        result = evaluator.evaluate({}, "903841", "999", context=ctx)
        assert result is False

    def test_solo_laboratorio_exception(self, evaluator):
        """Factura with ALL lab codes (not in pre_scan_cache) → skip (True)."""
        evaluator._load_constants()
        evaluator._pre_scan_cache = {}  # no non-lab codes for F002
        ctx = EvaluationContext(invoice_data={
            "tipo_factura_descripcion": "Intramural",
            "codigo_entidad_cobrar": "EPSS41",
            "codigo_dx_principal": "Z359",
            "numero_factura": "F002",
        })
        result = evaluator.evaluate({}, "906127", "999", context=ctx)
        assert result is True

    def test_pym_intramural_in_nueva_eps_no_capita(self, evaluator):
        """PYM_INTRAMURAL code in NUEVA_EPS_NO_CAPITA → skip (True)."""
        # 897011 (Monitoria Fetal) is in CODIGOS_PYM_INTRAMURAL
        # AND in CODIGOS_NUEVA_EPS_NO_CAPITA
        evaluator._load_constants()
        ctx = EvaluationContext(invoice_data={
            "tipo_factura_descripcion": "Intramural",
            "codigo_entidad_cobrar": "EPSS41",
        })
        result = evaluator.evaluate({}, "897011", "999", context=ctx)
        assert result is True

    def test_pre_scan_detects_non_lab(self, evaluator):
        """pre_scan_sheet correctly identifies facturas with non-lab codes."""
        from app.constants.base import CODIGOS_LABORATORIO_ENVIO
        lab_code = list(CODIGOS_LABORATORIO_ENVIO)[0]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=1, column=3, value="TIPO_FACTURA_DESCRIPCION")
        # F001: has a non-lab code
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="906127")
        ws.cell(row=2, column=3, value="Intramural")
        # F002: only lab codes
        ws.cell(row=3, column=1, value="F002")
        ws.cell(row=3, column=2, value=lab_code)
        ws.cell(row=3, column=3, value="Intramural")

        indices = {"numero_factura": 0, "codigo": 1, "tipo_factura_descripcion": 2}
        evaluator.pre_scan_sheet(ws, indices)

        assert "F001" in evaluator._pre_scan_cache  # has non-lab
        assert "F002" not in evaluator._pre_scan_cache  # all lab
