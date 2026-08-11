"""Tests for Hospitalización codes — group rules F2.

These tests verify the group rules that replace detect_hospitalizacion_codes():
- hosp_codigos_oblig_mayor24h: estancia > 24h AND missing OBLIG_MAYOR_24 codes
- hosp_codigos_oblig_menor24h: estancia <= 24h AND missing OBLIG_MENOR_24 codes
- hosp_codigos_prohibidos: prohibited codes present (includes SOAT variants)

Tests use direct condition trees (no DB) matching how test_group_evaluator.py works.
"""

from __future__ import annotations

from datetime import datetime

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ── Helpers ────────────────────────────────────────────────────────────────

def _build_ws() -> tuple[Worksheet, dict[str, int]]:
    """Create a worksheet with standard hospitalizacion columns."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="NUMERO_FACTURA")
    ws.cell(row=1, column=2, value="CODIGO")
    ws.cell(row=1, column=3, value="FEC_FACTURA")
    ws.cell(row=1, column=4, value="FECHA_CIERRE")
    ws.cell(row=1, column=5, value="TIPO_FACTURA_DESCRIPCION")
    ws.cell(row=1, column=6, value="TARIFARIO")
    ws.cell(row=1, column=7, value="CODIGO_ENTIDAD_COBRAR")
    indices = {
        "numero_factura": 0,
        "codigo": 1,
        "fec_factura": 2,
        "fecha_cierre": 3,
        "tipo_factura_descripcion": 4,
        "tarifario": 5,
        "codigo_entidad_cobrar": 6,
    }
    return ws, indices


def _ce():
    from app.services.engine.condition_evaluator import ConditionEvaluator
    return ConditionEvaluator()


def _collector():
    from app.services.engine.evidence_collector import EvidenceCollector
    return EvidenceCollector()


def _rule_info(name="test_rule", desc="Test rule"):
    return {
        "id": 999, "version": 1, "dominio": "hospitalizacion",
        "nombre": name, "descripcion": desc, "severidad": "error",
    }


def _flat_tree(conditions: list[dict]) -> dict:
    """Build a condition tree from a flat list of condition dicts."""
    return _ce().build_tree(conditions)


# Shared condition sets
OBLIG_MAYOR_24 = ["129B02", "890601H", "890601"]
OBLIG_MENOR_24 = ["890601H", "129B02"]
PROHIBIDOS = ["05DSB01", "5DSB01", "890701"]
SOAT_OBLIG = ["39133", "38114", "39131"]
SOAT_PROH = ["39145", "38915"]


# ═══════════════════════════════════════════════════════════════════════════
# Rule: hosp_codigos_oblig_mayor24h
# Condition: gt(estancia_horas, 24) AND NOT(set_contains_all(codigos, OBLIG))
# ═══════════════════════════════════════════════════════════════════════════

class TestHospObligMayor24h:
    """hosp_codigos_oblig_mayor24h: >24h must have OBLIG_MAYOR_24."""

    @pytest.fixture
    def standard_agg(self):
        return [
            {"function": "compute_horas", "field1": "fec_factura",
             "field2": "fecha_cierre", "target": "estancia_horas"},
            {"function": "collect_set", "field": "codigo", "target": "collect_set_codigo"},
        ]

    def test_mayor24h_falta_un_codigo(self, standard_agg):
        """>24h factura missing one required code → MATCH."""
        from app.services.engine.group_evaluator import GroupEvaluator

        ws, indices = _build_ws()
        # F001: 48h, has 129B02 and 890601H, missing 890601
        for r in range(2, 5):
            ws.cell(row=r, column=1, value="F001")
            ws.cell(row=r, column=3, value=datetime(2024, 1, 15, 8, 0, 0))
            ws.cell(row=r, column=4, value=datetime(2024, 1, 17, 8, 0, 0))
            ws.cell(row=r, column=5, value="Hospitalización")
        ws.cell(row=2, column=2, value="129B02")
        ws.cell(row=3, column=2, value="890601H")

        groups = GroupEvaluator.build_groups(
            ws, indices, filter_field="tipo_factura_descripcion", filter_value="Hospitalización"
        )

        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "gt",
             "fuente_datos": "invoice.estancia_horas", "valor_esperado": "24", "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_contains_all",
             "fuente_datos": "invoice.collect_set_codigo",
             "valor_esperado": OBLIG_MAYOR_24, "orden": 0},
        ]
        tree = _flat_tree(flat)
        results = GroupEvaluator.evaluate(
            groups, ws, indices, standard_agg, tree, _ce(),
            _rule_info("hosp_codigos_oblig_mayor24h"), _collector(),
        )
        assert len(results) == 1
        assert results[0]["factura"] == "F001"

    def test_mayor24h_todos_codigos(self, standard_agg):
        """>24h factura with ALL required codes → NO_MATCH."""
        from app.services.engine.group_evaluator import GroupEvaluator

        ws, indices = _build_ws()
        for r in range(2, 6):
            ws.cell(row=r, column=1, value="F001")
            ws.cell(row=r, column=3, value=datetime(2024, 1, 15, 8, 0, 0))
            ws.cell(row=r, column=4, value=datetime(2024, 1, 17, 8, 0, 0))
            ws.cell(row=r, column=5, value="Hospitalización")
        ws.cell(row=2, column=2, value="129B02")
        ws.cell(row=3, column=2, value="890601H")
        ws.cell(row=4, column=2, value="890601")

        groups = GroupEvaluator.build_groups(
            ws, indices, filter_field="tipo_factura_descripcion", filter_value="Hospitalización"
        )
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "gt",
             "fuente_datos": "invoice.estancia_horas", "valor_esperado": "24", "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_contains_all",
             "fuente_datos": "invoice.collect_set_codigo",
             "valor_esperado": OBLIG_MAYOR_24, "orden": 0},
        ]
        tree = _flat_tree(flat)
        results = GroupEvaluator.evaluate(
            groups, ws, indices, standard_agg, tree, _ce(),
            _rule_info("hosp_codigos_oblig_mayor24h"), _collector(),
        )
        assert len(results) == 0

    def test_menor24h_no_aplica_oblig_mayor(self, standard_agg):
        """<=24h factura should NOT trigger mayor24h rule."""
        from app.services.engine.group_evaluator import GroupEvaluator

        ws, indices = _build_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="999999")
        ws.cell(row=2, column=3, value=datetime(2024, 1, 15, 8, 0, 0))
        ws.cell(row=2, column=4, value=datetime(2024, 1, 15, 10, 0, 0))
        ws.cell(row=2, column=5, value="Hospitalización")

        groups = GroupEvaluator.build_groups(
            ws, indices, filter_field="tipo_factura_descripcion", filter_value="Hospitalización"
        )
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "gt",
             "fuente_datos": "invoice.estancia_horas", "valor_esperado": "24", "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_contains_all",
             "fuente_datos": "invoice.collect_set_codigo",
             "valor_esperado": OBLIG_MAYOR_24, "orden": 0},
        ]
        tree = _flat_tree(flat)
        results = GroupEvaluator.evaluate(
            groups, ws, indices, standard_agg, tree, _ce(),
            _rule_info("hosp_codigos_oblig_mayor24h"), _collector(),
        )
        # estancia_horas = 2.0, gt(24) is False → AND short-circuits → NO_MATCH
        assert len(results) == 0


# ═══════════════════════════════════════════════════════════════════════════
# Rule: hosp_codigos_oblig_menor24h
# Condition: lte(estancia_horas, 24) AND NOT(set_contains_all(codigos, OBLIG))
# ═══════════════════════════════════════════════════════════════════════════

class TestHospObligMenor24h:
    """hosp_codigos_oblig_menor24h: <=24h must have OBLIG_MENOR_24."""

    @pytest.fixture
    def standard_agg(self):
        return [
            {"function": "compute_horas", "field1": "fec_factura",
             "field2": "fecha_cierre", "target": "estancia_horas"},
            {"function": "collect_set", "field": "codigo", "target": "collect_set_codigo"},
        ]

    def test_menor24h_falta_codigo(self, standard_agg):
        """<=24h missing a required code → MATCH."""
        from app.services.engine.group_evaluator import GroupEvaluator

        ws, indices = _build_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="890601H")  # missing 129B02
        ws.cell(row=2, column=3, value=datetime(2024, 1, 15, 8, 0, 0))
        ws.cell(row=2, column=4, value=datetime(2024, 1, 15, 14, 0, 0))
        ws.cell(row=2, column=5, value="Hospitalización")

        groups = GroupEvaluator.build_groups(
            ws, indices, filter_field="tipo_factura_descripcion", filter_value="Hospitalización"
        )
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "lte",
             "fuente_datos": "invoice.estancia_horas", "valor_esperado": "24", "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_contains_all",
             "fuente_datos": "invoice.collect_set_codigo",
             "valor_esperado": OBLIG_MENOR_24, "orden": 0},
        ]
        tree = _flat_tree(flat)
        results = GroupEvaluator.evaluate(
            groups, ws, indices, standard_agg, tree, _ce(),
            _rule_info("hosp_codigos_oblig_menor24h"), _collector(),
        )
        assert len(results) == 1
        assert results[0]["factura"] == "F001"

    def test_menor24h_todos_codigos(self, standard_agg):
        """<=24h with all required codes → NO_MATCH."""
        from app.services.engine.group_evaluator import GroupEvaluator

        ws, indices = _build_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="890601H")
        ws.cell(row=2, column=3, value=datetime(2024, 1, 15, 8, 0, 0))
        ws.cell(row=2, column=4, value=datetime(2024, 1, 15, 14, 0, 0))
        ws.cell(row=2, column=5, value="Hospitalización")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="129B02")
        ws.cell(row=3, column=3, value=datetime(2024, 1, 15, 8, 0, 0))
        ws.cell(row=3, column=4, value=datetime(2024, 1, 15, 14, 0, 0))
        ws.cell(row=3, column=5, value="Hospitalización")

        groups = GroupEvaluator.build_groups(
            ws, indices, filter_field="tipo_factura_descripcion", filter_value="Hospitalización"
        )
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "lte",
             "fuente_datos": "invoice.estancia_horas", "valor_esperado": "24", "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_contains_all",
             "fuente_datos": "invoice.collect_set_codigo",
             "valor_esperado": OBLIG_MENOR_24, "orden": 0},
        ]
        tree = _flat_tree(flat)
        results = GroupEvaluator.evaluate(
            groups, ws, indices, standard_agg, tree, _ce(),
            _rule_info("hosp_codigos_oblig_menor24h"), _collector(),
        )
        assert len(results) == 0


# ═══════════════════════════════════════════════════════════════════════════
# Rule: hosp_codigos_prohibidos
# Condition: set_intersects(collect_set_codigo, PROHIBIDOS)
# ═══════════════════════════════════════════════════════════════════════════

class TestHospCodigosProhibidos:
    """hosp_codigos_prohibidos: prohibited codes detected."""

    @pytest.fixture
    def standard_agg(self):
        return [
            {"function": "collect_set", "field": "codigo", "target": "collect_set_codigo"},
        ]

    def test_prohibido_presente(self, standard_agg):
        """Factura with prohibited code → MATCH."""
        from app.services.engine.group_evaluator import GroupEvaluator

        ws, indices = _build_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="05DSB01")
        ws.cell(row=2, column=5, value="Hospitalización")

        groups = GroupEvaluator.build_groups(
            ws, indices, filter_field="tipo_factura_descripcion", filter_value="Hospitalización"
        )
        flat = [
            {"id": 1, "padre_id": None, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_codigo",
             "valor_esperado": PROHIBIDOS, "orden": 0},
        ]
        tree = _flat_tree(flat)
        results = GroupEvaluator.evaluate(
            groups, ws, indices, standard_agg, tree, _ce(),
            _rule_info("hosp_codigos_prohibidos"), _collector(),
        )
        assert len(results) == 1
        assert results[0]["factura"] == "F001"

    def test_sin_prohibidos(self, standard_agg):
        """Factura without prohibited codes → NO_MATCH."""
        from app.services.engine.group_evaluator import GroupEvaluator

        ws, indices = _build_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="129B02")
        ws.cell(row=2, column=5, value="Hospitalización")

        groups = GroupEvaluator.build_groups(
            ws, indices, filter_field="tipo_factura_descripcion", filter_value="Hospitalización"
        )
        flat = [
            {"id": 1, "padre_id": None, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_codigo",
             "valor_esperado": PROHIBIDOS, "orden": 0},
        ]
        tree = _flat_tree(flat)
        results = GroupEvaluator.evaluate(
            groups, ws, indices, standard_agg, tree, _ce(),
            _rule_info("hosp_codigos_prohibidos"), _collector(),
        )
        assert len(results) == 0

    def test_5dsb01_tambien_prohibido(self, standard_agg):
        """5DSB01 (without leading zero) is also prohibited."""
        from app.services.engine.group_evaluator import GroupEvaluator

        ws, indices = _build_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="5DSB01")
        ws.cell(row=2, column=5, value="Hospitalización")

        groups = GroupEvaluator.build_groups(
            ws, indices, filter_field="tipo_factura_descripcion", filter_value="Hospitalización"
        )
        flat = [
            {"id": 1, "padre_id": None, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_codigo",
             "valor_esperado": PROHIBIDOS, "orden": 0},
        ]
        tree = _flat_tree(flat)
        results = GroupEvaluator.evaluate(
            groups, ws, indices, standard_agg, tree, _ce(),
            _rule_info("hosp_codigos_prohibidos"), _collector(),
        )
        assert len(results) == 1


class TestHospProhibidosSOAT:
    """SOAT-specific prohibited+obligatory codes for hospitalizacion."""

    def test_soat_prohibido_39145(self):
        """SOAT factura with SOAT-prohibited 39145 → MATCH."""
        from app.services.engine.group_evaluator import GroupEvaluator

        ws, indices = _build_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="39145")
        ws.cell(row=2, column=5, value="Hospitalización")
        ws.cell(row=2, column=6, value="SOAT")

        groups = GroupEvaluator.build_groups(
            ws, indices, filter_field="tipo_factura_descripcion", filter_value="Hospitalización"
        )
        agg_configs = [
            {"function": "collect_set", "field": "codigo", "target": "collect_set_codigo"},
        ]
        flat = [
            {"id": 1, "padre_id": None, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_codigo",
             "valor_esperado": SOAT_PROH, "orden": 0},
        ]
        tree = _flat_tree(flat)
        results = GroupEvaluator.evaluate(
            groups, ws, indices, agg_configs, tree, _ce(),
            _rule_info("hosp_codigos_prohibidos"), _collector(),
        )
        assert len(results) == 1

    def test_soat_sin_prohibidos(self):
        """SOAT factura without SOAT-prohibited codes → NO_MATCH."""
        from app.services.engine.group_evaluator import GroupEvaluator

        ws, indices = _build_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="129B02")
        ws.cell(row=2, column=5, value="Hospitalización")
        ws.cell(row=2, column=6, value="SOAT")

        groups = GroupEvaluator.build_groups(
            ws, indices, filter_field="tipo_factura_descripcion", filter_value="Hospitalización"
        )
        agg_configs = [
            {"function": "collect_set", "field": "codigo", "target": "collect_set_codigo"},
        ]
        flat = [
            {"id": 1, "padre_id": None, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_codigo",
             "valor_esperado": SOAT_PROH, "orden": 0},
        ]
        tree = _flat_tree(flat)
        results = GroupEvaluator.evaluate(
            groups, ws, indices, agg_configs, tree, _ce(),
            _rule_info("hosp_codigos_prohibidos"), _collector(),
        )
        assert len(results) == 0
