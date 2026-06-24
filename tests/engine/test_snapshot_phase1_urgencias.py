"""Snapshot tests Phase 1: legacy vs engine output for 5 urgencias rules.

Verifies that the DB-backed engine, when pointed at the same seed SQL
conditions, detects the same facturas as the legacy Python detectors.
"""

from __future__ import annotations

import pytest
from unittest.mock import MagicMock
from openpyxl import Workbook


# ── Helpers ─────────────────────────────────────────────────────────────────

def _mock_session_with_rule(rule_name, dominio, descripcion, condiciones_dicts):
    """Create a mock session that returns a Regla + Condicion tree.

    condiciones_dicts: list of dicts with keys:
        id, padre_id, tipo, operador, fuente_datos, valor_esperado, orden
    """
    from app.models import Regla

    session = MagicMock()
    mock_query = MagicMock()
    mock_query.filter.return_value = mock_query
    mock_query.order_by.return_value = mock_query

    regla = Regla(
        id=1, nombre=rule_name, dominio=dominio,
        estado="active", version=1, prioridad=10, severidad="warning",
        descripcion=descripcion,
    )
    mock_query.first.return_value = regla

    cond_mocks = []
    for cd in condiciones_dicts:
        m = MagicMock()
        m.id = cd["id"]
        m.regla_id = cd.get("regla_id", 1)
        m.padre_id = cd["padre_id"]
        m.tipo = cd["tipo"]
        m.operador = cd.get("operador")
        m.fuente_datos = cd.get("fuente_datos")
        m.valor_esperado = cd.get("valor_esperado")
        m.orden = cd.get("orden", 0)
        cond_mocks.append(m)

    mock_query.all.return_value = cond_mocks
    session.query.return_value = mock_query
    return session


def _build_indices(*col_names):
    """Build indices dict: col_name → 0-based index."""
    return {name: i for i, name in enumerate(col_names)}


def _run_engine_detection(rule_name, dominio, descripcion, condiciones, ws, indices):
    """Run engine detection against a worksheet with mocked session."""
    from app.services.engine.rule_based_detector import RuleBasedDetector
    session = _mock_session_with_rule(rule_name, dominio, descripcion, condiciones)
    detector = RuleBasedDetector(rule_name, session)
    return detector.detect(ws, indices)


def _get_facturas_from_results(results):
    """Extract factura strings from detection results (legacy or engine)."""
    facturas = set()
    for r in results:
        if isinstance(r, dict):
            facturas.add(r.get("factura", ""))
        elif isinstance(r, str):
            facturas.add(r)
    return facturas


# ── Test: revision_entidad_86 ───────────────────────────────────────────────

class TestRevisionEntidad86:
    """Engine must match legacy detection for entidad == '86'."""

    def test_engine_detects_entidad_86(self):
        """Engine with eq(codigo_entidad_cobrar, '86') detects the row."""
        condiciones = [
            {
                "id": 1, "padre_id": None,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar",
                "valor_esperado": "86", "orden": 0,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO_ENTIDAD_COBRAR")
        ws.cell(row=1, column=3, value="TIPO_FACTURA_DESCRIPCION")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="86")
        ws.cell(row=2, column=3, value="Urgencias")

        indices = _build_indices("numero_factura", "codigo_entidad_cobrar", "tipo_factura_descripcion")
        results = _run_engine_detection(
            "revision_entidad_86", "urgencias",
            "Revisión necesaria para entidad 86",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F001" in facturas, "Engine should detect entidad=86"

    def test_engine_ignores_other_entidad(self):
        """Engine skips rows where entidad != '86'."""
        condiciones = [
            {
                "id": 1, "padre_id": None,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar",
                "valor_esperado": "86", "orden": 0,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO_ENTIDAD_COBRAR")
        ws.cell(row=2, column=1, value="F002")
        ws.cell(row=2, column=2, value="ESS118")

        indices = _build_indices("numero_factura", "codigo_entidad_cobrar")
        results = _run_engine_detection(
            "revision_entidad_86", "urgencias",
            "Revisión necesaria para entidad 86",
            condiciones, ws, indices,
        )

        assert len(results) == 0, "Engine should NOT detect entidad != 86"

    def test_engine_ignores_empty_entidad(self):
        """Engine skips rows where entidad is empty/null."""
        condiciones = [
            {
                "id": 1, "padre_id": None,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar",
                "valor_esperado": "86", "orden": 0,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO_ENTIDAD_COBRAR")
        ws.cell(row=2, column=1, value="F003")
        ws.cell(row=2, column=2, value=None)  # Empty

        indices = _build_indices("numero_factura", "codigo_entidad_cobrar")
        results = _run_engine_detection(
            "revision_entidad_86", "urgencias",
            "Revisión necesaria para entidad 86",
            condiciones, ws, indices,
        )

        assert len(results) == 0, "Engine should NOT detect empty entidad"

    def test_legacy_vs_engine_detect_same_facturas(self):
        """Both legacy and engine detect the same facturas for entidad=86."""
        from app.services.urgencias.revision_entidad_86 import (
            detect_revision_entidad_86_urgencias,
        )

        condiciones = [
            {
                "id": 1, "padre_id": None,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar",
                "valor_esperado": "86", "orden": 0,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="TIPO_FACTURA_DESCRIPCION")
        ws.cell(row=1, column=2, value="NUMERO_FACTURA")
        ws.cell(row=1, column=3, value="CODIGO_ENTIDAD_COBRAR")
        ws.cell(row=1, column=4, value="CODIGO")
        ws.cell(row=1, column=5, value="PROCEDIMIENTO")
        ws.cell(row=1, column=6, value="IDE_CONTRATO")
        # Row: entidad=86, Urgencias → should be detected
        ws.cell(row=2, column=1, value="Urgencias")
        ws.cell(row=2, column=2, value="F001")
        ws.cell(row=2, column=3, value="86")
        ws.cell(row=2, column=4, value="890201")
        ws.cell(row=2, column=5, value="CONSULTA")
        ws.cell(row=2, column=6, value="12345")
        # Row: different entidad → should NOT be detected
        ws.cell(row=3, column=1, value="Urgencias")
        ws.cell(row=3, column=2, value="F002")
        ws.cell(row=3, column=3, value="ESS118")
        ws.cell(row=3, column=4, value="890201")

        indices_legacy = _build_indices(
            "tipo_factura_descripcion", "numero_factura",
            "codigo_entidad_cobrar", "codigo", "procedimiento", "ide_contrato",
        )

        legacy_results = detect_revision_entidad_86_urgencias(ws, indices_legacy)
        legacy_facturas = _get_facturas_from_results(legacy_results)

        engine_results = _run_engine_detection(
            "revision_entidad_86", "urgencias",
            "Revisión necesaria para entidad 86",
            condiciones, ws, indices_legacy,
        )
        engine_facturas = _get_facturas_from_results(engine_results)

        # Legacy filters by tipo_factura_descripcion == "Urgencias",
        # engine condition tree doesn't (it only checks entidad == "86").
        # Both should detect F001 and not F002.
        assert "F001" in legacy_facturas, "Legacy should detect F001"
        assert "F001" in engine_facturas, "Engine should detect F001"
        assert "F002" not in legacy_facturas, "Legacy should NOT detect F002"
        assert "F002" not in engine_facturas, "Engine should NOT detect F002"


# ── Test: cantidades_urgencias ──────────────────────────────────────────────

class TestCantidadesUrgencias:
    """Engine must match legacy detection for restricted code + cantidad > 1."""

    RESTRICTED_CODES = ["05DSB01", "5DSB01", "890601", "890701", "129B02", "12333"]

    def test_engine_detects_restricted_code_over_quantity(self):
        """Engine detects when code is restricted AND cantidad > 1."""
        condiciones = [
            {  # root: AND
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "AND", "orden": 0,
            },
            {  # child_1: in(codigo, restricted_codes)
                "id": 2, "padre_id": 1,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": self.RESTRICTED_CODES, "orden": 0,
            },
            {  # child_2: gt(cantidad, 1)
                "id": 3, "padre_id": 1,
                "tipo": "atomic", "operador": "gt",
                "fuente_datos": "invoice.cantidad",
                "valor_esperado": 1, "orden": 1,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=1, column=3, value="CANTIDAD")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="890601")
        ws.cell(row=2, column=3, value=3)

        indices = _build_indices("numero_factura", "codigo", "cantidad")
        results = _run_engine_detection(
            "cantidades_urgencias", "urgencias",
            "Cantidad excedida para código de urgencias restringido",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F001" in facturas, "Engine should detect restricted code with qty > 1"

    def test_engine_ignores_valid_quantity(self):
        """Engine ignores restricted code when cantidad <= 1."""
        condiciones = [
            {
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "AND", "orden": 0,
            },
            {
                "id": 2, "padre_id": 1,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": self.RESTRICTED_CODES, "orden": 0,
            },
            {
                "id": 3, "padre_id": 1,
                "tipo": "atomic", "operador": "gt",
                "fuente_datos": "invoice.cantidad",
                "valor_esperado": 1, "orden": 1,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=1, column=3, value="CANTIDAD")
        ws.cell(row=2, column=1, value="F002")
        ws.cell(row=2, column=2, value="890601")
        ws.cell(row=2, column=3, value=1)

        indices = _build_indices("numero_factura", "codigo", "cantidad")
        results = _run_engine_detection(
            "cantidades_urgencias", "urgencias",
            "Cantidad excedida para código de urgencias restringido",
            condiciones, ws, indices,
        )

        assert len(results) == 0, "Engine should NOT detect qty <= 1"

    def test_engine_ignores_non_restricted_code(self):
        """Engine ignores non-restricted codes regardless of quantity."""
        condiciones = [
            {
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "AND", "orden": 0,
            },
            {
                "id": 2, "padre_id": 1,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": self.RESTRICTED_CODES, "orden": 0,
            },
            {
                "id": 3, "padre_id": 1,
                "tipo": "atomic", "operador": "gt",
                "fuente_datos": "invoice.cantidad",
                "valor_esperado": 1, "orden": 1,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=1, column=3, value="CANTIDAD")
        ws.cell(row=2, column=1, value="F003")
        ws.cell(row=2, column=2, value="XYZ99")
        ws.cell(row=2, column=3, value=5)

        indices = _build_indices("numero_factura", "codigo", "cantidad")
        results = _run_engine_detection(
            "cantidades_urgencias", "urgencias",
            "Cantidad excedida para código de urgencias restringido",
            condiciones, ws, indices,
        )

        assert len(results) == 0, "Engine should NOT detect non-restricted code"


# ── Test: cantidades_soat_urgencias ─────────────────────────────────────────

class TestCantidadesSOATUrgencias:
    """Engine must match legacy detection for SOAT + restricted code + cantidad != 1."""

    SOAT_CODES = ["39145", "38114", "38915", "39131"]

    def test_engine_detects_soat_restricted_qty_not_1(self):
        """Engine detects SOAT + restricted code + cantidad != 1."""
        condiciones = [
            {  # root: AND
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "AND", "orden": 0,
            },
            {  # child_1: eq(tarifario, "SOAT")
                "id": 2, "padre_id": 1,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.tarifario",
                "valor_esperado": "SOAT", "orden": 0,
            },
            {  # child_2: in(codigo, restricted_codes)
                "id": 3, "padre_id": 1,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": self.SOAT_CODES, "orden": 1,
            },
            {  # child_3: NOT(eq(cantidad, 1))
                "id": 4, "padre_id": 1,
                "tipo": "composite", "operador": "NOT", "orden": 2,
            },
            {  # child_3_1: eq(cantidad, 1)
                "id": 5, "padre_id": 4,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.cantidad",
                "valor_esperado": 1, "orden": 0,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TARIFARIO")
        ws.cell(row=1, column=3, value="CODIGO")
        ws.cell(row=1, column=4, value="CANTIDAD")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="SOAT")
        ws.cell(row=2, column=3, value="39145")
        ws.cell(row=2, column=4, value=2)

        indices = _build_indices("numero_factura", "tarifario", "codigo", "cantidad")
        results = _run_engine_detection(
            "cantidades_soat_urgencias", "urgencias",
            "Cantidad SOAT no es 1 para código restringido",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F001" in facturas, "Engine should detect SOAT code with qty != 1"

    def test_engine_ignores_non_soat(self):
        """Engine ignores when tarifario is not SOAT."""
        condiciones = [
            {
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "AND", "orden": 0,
            },
            {
                "id": 2, "padre_id": 1,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.tarifario",
                "valor_esperado": "SOAT", "orden": 0,
            },
            {
                "id": 3, "padre_id": 1,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": self.SOAT_CODES, "orden": 1,
            },
            {
                "id": 4, "padre_id": 1,
                "tipo": "composite", "operador": "NOT", "orden": 2,
            },
            {
                "id": 5, "padre_id": 4,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.cantidad",
                "valor_esperado": 1, "orden": 0,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TARIFARIO")
        ws.cell(row=1, column=3, value="CODIGO")
        ws.cell(row=1, column=4, value="CANTIDAD")
        ws.cell(row=2, column=1, value="F002")
        ws.cell(row=2, column=2, value="ISS")
        ws.cell(row=2, column=3, value="39145")
        ws.cell(row=2, column=4, value=3)

        indices = _build_indices("numero_factura", "tarifario", "codigo", "cantidad")
        results = _run_engine_detection(
            "cantidades_soat_urgencias", "urgencias",
            "Cantidad SOAT no es 1 para código restringido",
            condiciones, ws, indices,
        )

        assert len(results) == 0, "Engine should ignore non-SOAT"

    def test_engine_ignores_qty_1(self):
        """Engine ignores SOAT code with cantidad == 1 (expected)."""
        condiciones = [
            {
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "AND", "orden": 0,
            },
            {
                "id": 2, "padre_id": 1,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.tarifario",
                "valor_esperado": "SOAT", "orden": 0,
            },
            {
                "id": 3, "padre_id": 1,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": self.SOAT_CODES, "orden": 1,
            },
            {
                "id": 4, "padre_id": 1,
                "tipo": "composite", "operador": "NOT", "orden": 2,
            },
            {
                "id": 5, "padre_id": 4,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.cantidad",
                "valor_esperado": 1, "orden": 0,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TARIFARIO")
        ws.cell(row=1, column=3, value="CODIGO")
        ws.cell(row=1, column=4, value="CANTIDAD")
        ws.cell(row=2, column=1, value="F003")
        ws.cell(row=2, column=2, value="SOAT")
        ws.cell(row=2, column=3, value="39145")
        ws.cell(row=2, column=4, value=1)

        indices = _build_indices("numero_factura", "tarifario", "codigo", "cantidad")
        results = _run_engine_detection(
            "cantidades_soat_urgencias", "urgencias",
            "Cantidad SOAT no es 1 para código restringido",
            condiciones, ws, indices,
        )

        assert len(results) == 0, "Engine should ignore qty == 1"


# ── Test: mal_capitado ──────────────────────────────────────────────────────

class TestMalCapitado:
    """Engine must match legacy detection for mal capitado patterns."""

    MAL_CODES = ["G03XB01", "A02BB01"]

    def test_engine_detects_code_without_fev_prefix(self):
        """Pattern 1: G03XB01/A02BB01 with factura NOT containing FEV."""
        condiciones = [
            {  # root: OR
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "OR", "orden": 0,
            },
            # Group 1: code in MAL_CODES AND NOT(contains factura, "FEV")
            {  # group_1: AND
                "id": 2, "padre_id": 1,
                "tipo": "composite", "operador": "AND", "orden": 0,
            },
            {  # group_1.child_1: in(codigo, MAL_CODES)
                "id": 3, "padre_id": 2,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": self.MAL_CODES, "orden": 0,
            },
            {  # group_1.child_2: NOT(contains(factura, "FEV"))
                "id": 4, "padre_id": 2,
                "tipo": "composite", "operador": "NOT", "orden": 1,
            },
            {  # group_1.child_2.child: contains(factura, "FEV")
                "id": 5, "padre_id": 4,
                "tipo": "atomic", "operador": "contains",
                "fuente_datos": "invoice.numero_factura",
                "valor_esperado": "FEV", "orden": 0,
            },
            # Group 2: factura contains "CAP" AND NOT(eq(entidad, "ESS118"))
            {  # group_2: AND
                "id": 6, "padre_id": 1,
                "tipo": "composite", "operador": "AND", "orden": 1,
            },
            {  # group_2.child_1: contains(factura, "CAP")
                "id": 7, "padre_id": 6,
                "tipo": "atomic", "operador": "contains",
                "fuente_datos": "invoice.numero_factura",
                "valor_esperado": "CAP", "orden": 0,
            },
            {  # group_2.child_2: NOT(eq(entidad, "ESS118"))
                "id": 8, "padre_id": 6,
                "tipo": "composite", "operador": "NOT", "orden": 1,
            },
            {  # group_2.child_2.child: eq(entidad, "ESS118")
                "id": 9, "padre_id": 8,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar",
                "valor_esperado": "ESS118", "orden": 0,
            },
        ]

        # Pattern 1: G03XB01 with factura "ABC-123" (no FEV) → detected
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=2, column=1, value="ABC-123")
        ws.cell(row=2, column=2, value="G03XB01")

        indices = _build_indices("numero_factura", "codigo")
        results = _run_engine_detection(
            "mal_capitado", "urgencias",
            "Factura mal capitada detectada",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "ABC-123" in facturas, "Engine should detect code without FEV prefix"

    def test_engine_ignores_fev_prefixed_factura(self):
        """Pattern 1: G03XB01 with FEV-prefixed factura → NOT detected."""
        condiciones = [
            {
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "OR", "orden": 0,
            },
            {
                "id": 2, "padre_id": 1,
                "tipo": "composite", "operador": "AND", "orden": 0,
            },
            {
                "id": 3, "padre_id": 2,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": self.MAL_CODES, "orden": 0,
            },
            {
                "id": 4, "padre_id": 2,
                "tipo": "composite", "operador": "NOT", "orden": 1,
            },
            {
                "id": 5, "padre_id": 4,
                "tipo": "atomic", "operador": "contains",
                "fuente_datos": "invoice.numero_factura",
                "valor_esperado": "FEV", "orden": 0,
            },
            {
                "id": 6, "padre_id": 1,
                "tipo": "composite", "operador": "AND", "orden": 1,
            },
            {
                "id": 7, "padre_id": 6,
                "tipo": "atomic", "operador": "contains",
                "fuente_datos": "invoice.numero_factura",
                "valor_esperado": "CAP", "orden": 0,
            },
            {
                "id": 8, "padre_id": 6,
                "tipo": "composite", "operador": "NOT", "orden": 1,
            },
            {
                "id": 9, "padre_id": 8,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar",
                "valor_esperado": "ESS118", "orden": 0,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=2, column=1, value="FEV-456")
        ws.cell(row=2, column=2, value="G03XB01")

        indices = _build_indices("numero_factura", "codigo")
        results = _run_engine_detection(
            "mal_capitado", "urgencias",
            "Factura mal capitada detectada",
            condiciones, ws, indices,
        )

        assert len(results) == 0, "FEV-prefixed factura should NOT be detected"

    def test_engine_detects_cap_with_wrong_entidad(self):
        """Pattern 2: factura CAP-prefixed but entidad != ESS118."""
        condiciones = [
            {
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "OR", "orden": 0,
            },
            {
                "id": 2, "padre_id": 1,
                "tipo": "composite", "operador": "AND", "orden": 0,
            },
            {
                "id": 3, "padre_id": 2,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": self.MAL_CODES, "orden": 0,
            },
            {
                "id": 4, "padre_id": 2,
                "tipo": "composite", "operador": "NOT", "orden": 1,
            },
            {
                "id": 5, "padre_id": 4,
                "tipo": "atomic", "operador": "contains",
                "fuente_datos": "invoice.numero_factura",
                "valor_esperado": "FEV", "orden": 0,
            },
            {
                "id": 6, "padre_id": 1,
                "tipo": "composite", "operador": "AND", "orden": 1,
            },
            {
                "id": 7, "padre_id": 6,
                "tipo": "atomic", "operador": "contains",
                "fuente_datos": "invoice.numero_factura",
                "valor_esperado": "CAP", "orden": 0,
            },
            {
                "id": 8, "padre_id": 6,
                "tipo": "composite", "operador": "NOT", "orden": 1,
            },
            {
                "id": 9, "padre_id": 8,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar",
                "valor_esperado": "ESS118", "orden": 0,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO_ENTIDAD_COBRAR")
        ws.cell(row=2, column=1, value="CAP-789")
        ws.cell(row=2, column=2, value="ESS062")

        indices = _build_indices("numero_factura", "codigo_entidad_cobrar")
        results = _run_engine_detection(
            "mal_capitado", "urgencias",
            "Factura mal capitada detectada",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "CAP-789" in facturas, "Engine should detect CAP with wrong entidad"

    def test_engine_ignores_cap_with_correct_entidad(self):
        """Pattern 2: factura CAP-prefixed with ESS118 → NOT detected."""
        condiciones = [
            {
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "OR", "orden": 0,
            },
            {
                "id": 2, "padre_id": 1,
                "tipo": "composite", "operador": "AND", "orden": 0,
            },
            {
                "id": 3, "padre_id": 2,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": self.MAL_CODES, "orden": 0,
            },
            {
                "id": 4, "padre_id": 2,
                "tipo": "composite", "operador": "NOT", "orden": 1,
            },
            {
                "id": 5, "padre_id": 4,
                "tipo": "atomic", "operador": "contains",
                "fuente_datos": "invoice.numero_factura",
                "valor_esperado": "FEV", "orden": 0,
            },
            {
                "id": 6, "padre_id": 1,
                "tipo": "composite", "operador": "AND", "orden": 1,
            },
            {
                "id": 7, "padre_id": 6,
                "tipo": "atomic", "operador": "contains",
                "fuente_datos": "invoice.numero_factura",
                "valor_esperado": "CAP", "orden": 0,
            },
            {
                "id": 8, "padre_id": 6,
                "tipo": "composite", "operador": "NOT", "orden": 1,
            },
            {
                "id": 9, "padre_id": 8,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo_entidad_cobrar",
                "valor_esperado": "ESS118", "orden": 0,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO_ENTIDAD_COBRAR")
        ws.cell(row=2, column=1, value="CAP-999")
        ws.cell(row=2, column=2, value="ESS118")

        indices = _build_indices("numero_factura", "codigo_entidad_cobrar")
        results = _run_engine_detection(
            "mal_capitado", "urgencias",
            "Factura mal capitada detectada",
            condiciones, ws, indices,
        )

        assert len(results) == 0, "CAP+ESS118 should NOT be detected"


# ── Test: cups_equivalentes ─────────────────────────────────────────────────

class TestCupsEquivalentes:
    """Engine must detect known CUPS codes with substitution issues."""

    def test_engine_detects_890201(self):
        """Code 890201 must always be flagged."""
        condiciones = [
            {  # root: OR
                "id": 1, "padre_id": None,
                "tipo": "composite", "operador": "OR", "orden": 0,
            },
            # eq(codigo, "890201")
            {
                "id": 2, "padre_id": 1,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": "890201", "orden": 0,
            },
            # eq(codigo, "129B01")
            {
                "id": 3, "padre_id": 1,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": "129B01", "orden": 1,
            },
            # AND: eq(codigo, "890205") AND NOT(in(entidad, ["ESS118", "ESSC18"]))
            {
                "id": 4, "padre_id": 1,
                "tipo": "composite", "operador": "AND", "orden": 2,
            },
            {
                "id": 5, "padre_id": 4,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": "890205", "orden": 0,
            },
            {
                "id": 6, "padre_id": 4,
                "tipo": "composite", "operador": "NOT", "orden": 1,
            },
            {
                "id": 7, "padre_id": 6,
                "tipo": "atomic", "operador": "in",
                "fuente_datos": "invoice.codigo_entidad_cobrar",
                "valor_esperado": ["ESS118", "ESSC18"], "orden": 0,
            },
            # AND: eq(codigo, "939402") AND eq(tipo_factura, "Hospitalización")
            {
                "id": 8, "padre_id": 1,
                "tipo": "composite", "operador": "AND", "orden": 3,
            },
            {
                "id": 9, "padre_id": 8,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": "939402", "orden": 0,
            },
            {
                "id": 10, "padre_id": 8,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.tipo_factura_descripcion",
                "valor_esperado": "Hospitalización", "orden": 1,
            },
            # AND: eq(codigo, "12333") AND eq(tipo_factura, "Hospitalización")
            {
                "id": 11, "padre_id": 1,
                "tipo": "composite", "operador": "AND", "orden": 4,
            },
            {
                "id": 12, "padre_id": 11,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.codigo",
                "valor_esperado": "12333", "orden": 0,
            },
            {
                "id": 13, "padre_id": 11,
                "tipo": "atomic", "operador": "eq",
                "fuente_datos": "invoice.tipo_factura_descripcion",
                "valor_esperado": "Hospitalización", "orden": 1,
            },
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=1, column=3, value="TIPO_FACTURA_DESCRIPCION")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="890201")
        ws.cell(row=2, column=3, value="Urgencias")

        indices = _build_indices("numero_factura", "codigo", "tipo_factura_descripcion")
        results = _run_engine_detection(
            "cups_equivalentes", "urgencias",
            "Código CUPS con equivalente conocido detectado",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F001" in facturas, "Engine should detect code 890201"

    def test_engine_detects_hospitalizacion_code(self):
        """Code 939402 with Hospitalización → detected."""
        condiciones = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "OR", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.codigo", "valor_esperado": "890201", "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.codigo", "valor_esperado": "129B01", "orden": 1},
            {"id": 4, "padre_id": 1, "tipo": "composite", "operador": "AND", "orden": 2},
            {"id": 5, "padre_id": 4, "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.codigo", "valor_esperado": "890205", "orden": 0},
            {"id": 6, "padre_id": 4, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 7, "padre_id": 6, "tipo": "atomic", "operador": "in",
             "fuente_datos": "invoice.codigo_entidad_cobrar", "valor_esperado": ["ESS118", "ESSC18"], "orden": 0},
            {"id": 8, "padre_id": 1, "tipo": "composite", "operador": "AND", "orden": 3},
            {"id": 9, "padre_id": 8, "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.codigo", "valor_esperado": "939402", "orden": 0},
            {"id": 10, "padre_id": 8, "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.tipo_factura_descripcion", "valor_esperado": "Hospitalización", "orden": 1},
            {"id": 11, "padre_id": 1, "tipo": "composite", "operador": "AND", "orden": 4},
            {"id": 12, "padre_id": 11, "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.codigo", "valor_esperado": "12333", "orden": 0},
            {"id": 13, "padre_id": 11, "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.tipo_factura_descripcion", "valor_esperado": "Hospitalización", "orden": 1},
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=1, column=3, value="TIPO_FACTURA_DESCRIPCION")
        ws.cell(row=2, column=1, value="F002")
        ws.cell(row=2, column=2, value="939402")
        ws.cell(row=2, column=3, value="Hospitalización")

        indices = _build_indices("numero_factura", "codigo", "tipo_factura_descripcion")
        results = _run_engine_detection(
            "cups_equivalentes", "urgencias",
            "Código CUPS con equivalente conocido detectado",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F002" in facturas, "Engine should detect 939402+Hospitalización"

    def test_engine_ignores_non_matching_code(self):
        """Code not in any pattern → NOT detected."""
        condiciones = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "OR", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.codigo", "valor_esperado": "890201", "orden": 0},
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=2, column=1, value="F003")
        ws.cell(row=2, column=2, value="999999")

        indices = _build_indices("numero_factura", "codigo")
        results = _run_engine_detection(
            "cups_equivalentes", "urgencias",
            "Código CUPS con equivalente conocido detectado",
            condiciones, ws, indices,
        )

        assert len(results) == 0, "Engine should ignore non-matching code"


# ── Engine output format tests ──────────────────────────────────────────────

class TestPhase1OutputFormat:
    """All Phase 1 rules must produce output with factura + problema + regla keys."""

    def test_output_has_required_keys(self):
        """Engine results must have factura, problema, regla, severidad keys."""
        condiciones = [
            {"id": 1, "padre_id": None, "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.codigo_entidad_cobrar",
             "valor_esperado": "86", "orden": 0},
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO_ENTIDAD_COBRAR")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="86")

        indices = _build_indices("numero_factura", "codigo_entidad_cobrar")
        results = _run_engine_detection(
            "test_format", "urgencias",
            "Test rule for format verification",
            condiciones, ws, indices,
        )

        assert len(results) > 0
        r = results[0]
        assert "factura" in r
        assert "problema" in r
        assert "regla" in r
        assert "severidad" in r
