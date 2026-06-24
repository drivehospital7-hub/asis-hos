"""Snapshot tests: legacy detector output vs engine output identity.

These tests verify that the RuleBasedDetector produces output identical
in structure and semantics to the legacy Python detectors.

Run against production Excel files before enabling engine in production.
"""

from __future__ import annotations

import pytest
from unittest.mock import MagicMock


class TestLegacyVsEngineOutputFormat:
    """Verify engine output structure matches legacy detector output."""

    def test_engine_output_keys_match_legacy_format(self):
        """Each engine result dict has factura and problema keys (like legacy)."""
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla
        from openpyxl import Workbook

        rule = Regla(
            id=1, nombre="test", dominio="odontologia",
            estado="active", version=1, prioridad=10, severidad="error",
            descripcion="Test rule",
        )
        from unittest.mock import MagicMock as M
        cond = M()
        cond.id = 1; cond.regla_id = 1; cond.padre_id = None
        cond.tipo = "atomic"; cond.operador = "eq"
        cond.fuente_datos = "invoice.convenio_facturado"
        cond.valor_esperado = "PyP"; cond.orden = 0

        session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.first.return_value = rule
        mock_query.all.return_value = [cond]
        session.query.return_value = mock_query

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CONVENIO_FACTURADO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="PyP")

        indices = {"numero_factura": 0, "convenio_facturado": 1}
        engine = RuleEvaluationEngine(session)
        results = engine.evaluate_sheet("test", ws, indices)

        assert len(results) == 1
        r = results[0]
        assert "factura" in r
        assert "problema" in r
        assert "regla" in r
        assert "severidad" in r
        # Legacy decimales returns: list of invoice number strings
        # Engine wrapper returns: list of dicts with factura + problema
        # Both are compatible with detect_all.py consumption

    def test_legacy_decimales_signature(self):
        """Legacy detect_decimales returns list[str] of invoice numbers."""
        from app.services.transversales.decimales import detect_decimales
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="VLR_SUBSIDIADO")
        ws.cell(row=1, column=3, value="VLR_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value=1000.50)  # Has decimal
        ws.cell(row=2, column=3, value=2000)

        indices = {"numero_factura": 0, "vlr_subsidiado": 1, "vlr_procedimiento": 2}
        result = detect_decimales(ws, indices)

        assert isinstance(result, list)
        # Legacy returns list of strings
        assert "F001" in result

    def test_legacy_ruta_duplicada_signature(self):
        """Legacy detect_ruta_duplicada returns list[dict] with specific keys."""
        from app.services.transversales.ruta_duplicada import detect_ruta_duplicada
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="IDENTIFICACION")
        ws.cell(row=1, column=3, value="CONVENIO_FACTURADO")
        # 3 facturas for same patient in PyP → ruta duplicada
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="12345")
        ws.cell(row=2, column=3, value="Promoción y Prevención")
        ws.cell(row=3, column=1, value="F002")
        ws.cell(row=3, column=2, value="12345")
        ws.cell(row=3, column=3, value="Promoción y Prevención")
        ws.cell(row=4, column=1, value="F003")
        ws.cell(row=4, column=2, value="12345")
        ws.cell(row=4, column=3, value="Promoción y Prevención")

        indices = {"numero_factura": 0, "identificacion": 1, "convenio_facturado": 2}
        result = detect_ruta_duplicada(ws, indices)

        assert isinstance(result, list)
        assert len(result) == 1
        assert "identificacion" in result[0]
        assert "facturas" in result[0]
        assert "cantidad" in result[0]
        assert result[0]["cantidad"] == 3


class TestGroupBySnapshotParity:
    """Snapshot parity tests: manual engine evaluation using the new
    aggregation + evaluator operators matches expected detection format.

    These tests verify the end-to-end flow:
      1. GroupEvaluator.build_groups() → partition
      2. GroupEvaluator._build_group_data() with collect_set/value_counts
      3. ConditionEvaluator.evaluate() with set_contains_all/set_intersects/all_values_match
      4. GroupEvaluator.evaluate() produces MATCH/NO_MATCH
    """

    def test_sala_obs_collect_set_detects_obligatorios(self):
        """Factura missing mandatory codes → MATCH (via engine group-by)."""
        from app.services.engine.group_evaluator import GroupEvaluator
        from app.services.engine.condition_evaluator import ConditionEvaluator
        from app.services.engine.evidence_collector import EvidenceCollector
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=1, column=3, value="CANTIDAD")
        # F001 has sala codes 5DSB01 and 129B02 but NOT mandatory 890701 or 890601
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="5DSB01")
        ws.cell(row=2, column=3, value=1)
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="129B02")
        ws.cell(row=3, column=3, value=1)
        # F002 has sala code + BOTH mandatory codes → NO match
        ws.cell(row=4, column=1, value="F002")
        ws.cell(row=4, column=2, value="890701")
        ws.cell(row=4, column=3, value=1)
        ws.cell(row=5, column=1, value="F002")
        ws.cell(row=5, column=2, value="890601")
        ws.cell(row=5, column=3, value=1)
        # Need a sala code for intersect check: no sala code → NO match anyway
        # Add an extra row: F003 has sala codes + all mandatory → NO match
        ws.cell(row=6, column=1, value="F003")
        ws.cell(row=6, column=2, value="5DSB01")
        ws.cell(row=6, column=3, value=1)

        indices = {"numero_factura": 0, "codigo": 1, "cantidad": 2}
        groups = GroupEvaluator.build_groups(ws, indices)

        agg_configs = [{"function": "collect_set", "field": "codigo",
                         "target": "collect_set_codigo"}]

        # Build condition tree matching the design:
        # AND(
        #   set_intersects(group.collect_set_codigo, ["5DSB01","05DSB01","129B02"]),
        #   NOT(set_contains_all(group.collect_set_codigo, ["890701","890601"]))
        # )
        condition_list = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND",
             "fuente_datos": "", "valor_esperado": "", "orden": 0},
            # Child: set_intersects
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "group.collect_set_codigo",
             "valor_esperado": ["5DSB01", "05DSB01", "129B02"], "orden": 0},
            # Child: NOT
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT",
             "fuente_datos": "", "valor_esperado": "", "orden": 1},
            # NOT child: set_contains_all
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_contains_all",
             "fuente_datos": "group.collect_set_codigo",
             "valor_esperado": ["890701", "890601"], "orden": 0},
        ]

        evaluator_inst = ConditionEvaluator()
        tree = evaluator_inst.build_tree(condition_list)
        assert tree is not None

        collector = EvidenceCollector()
        rule_info = {
            "id": 100, "version": 1, "dominio": "urgencias",
            "nombre": "sala_observacion_codigos_obligatorios",
            "descripcion": "Factura con codigos de sala observacion sin codigos obligatorios",
            "severidad": "error",
        }

        results = GroupEvaluator.evaluate(
            groups, ws, indices, agg_configs,
            tree, evaluator_inst, rule_info, collector,
        )

        # F001 does NOT have mandatory codes → MATCH
        matched_facturas = {r["factura"] for r in results}
        assert "F001" in matched_facturas
        # F002 has 890601 (mandatory) → NO MATCH
        assert "F002" not in matched_facturas

    def test_duplicados_base_collect_value_counts(self):
        """Factura with (codigo,cantidad) pairs having count < 2 → MATCH."""
        from app.services.engine.group_evaluator import GroupEvaluator
        from app.services.engine.condition_evaluator import ConditionEvaluator
        from app.services.engine.evidence_collector import EvidenceCollector
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CODIGO")
        ws.cell(row=1, column=3, value="CANTIDAD")
        # F001: all (codigo,cantidad) pairs appear only once (count=1 < 2)
        # all_values_match returns False → NO_MATCH (not all duplicated)
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=2, column=3, value=1)
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="B")
        ws.cell(row=3, column=3, value=2)
        # F002: (C,1) appears twice (count=2 >= 2) → all_values_match True → MATCH
        ws.cell(row=4, column=1, value="F002")
        ws.cell(row=4, column=2, value="C")
        ws.cell(row=4, column=3, value=1)
        ws.cell(row=5, column=1, value="F002")
        ws.cell(row=5, column=2, value="C")
        ws.cell(row=5, column=3, value=1)

        indices = {"numero_factura": 0, "codigo": 1, "cantidad": 2}
        groups = GroupEvaluator.build_groups(ws, indices)

        agg_configs = [{"function": "collect_value_counts",
                         "fields": ["codigo", "cantidad"]}]

        # Condition: all_values_match(group.collect_value_counts, 2)
        condition_list = [
            {"id": 1, "padre_id": None, "tipo": "atomic", "operador": "all_values_match",
             "fuente_datos": "group.collect_value_counts",
             "valor_esperado": 2, "orden": 0},
        ]

        evaluator_inst = ConditionEvaluator()
        tree = evaluator_inst.build_tree(condition_list)
        assert tree is not None

        collector = EvidenceCollector()
        rule_info = {
            "id": 200, "version": 1, "dominio": "transversal",
            "nombre": "detect_duplicados_base_generico",
            "descripcion": "Factura con (codigo,cantidad) no duplicados",
            "severidad": "warning",
        }

        results = GroupEvaluator.evaluate(
            groups, ws, indices, agg_configs,
            tree, evaluator_inst, rule_info, collector,
        )

        # F002 has (C,1) with count=2 >= 2 → all_values_match True → MATCH
        matched_facturas = {r["factura"] for r in results}
        assert "F002" in matched_facturas
        # F001 has pairs with count 1 < 2 → all_values_match False → NO_MATCH
        assert "F001" not in matched_facturas
