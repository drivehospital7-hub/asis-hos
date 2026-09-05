"""TDD tests for intra-row engine optimizations.

Tests exception caching, lazy date.edad, and NO_MATCH evidence skip.
All tests start RED — they fail until Phase 2 optimizations are implemented.
"""

from __future__ import annotations

import pytest
from unittest.mock import MagicMock, patch, PropertyMock


# ── Helpers ────────────────────────────────────────────────────────────────

def _mock_rule(id=1, nombre="test_rule", dominio="odontologia", parametros=None):
    """Create a mock Regla with minimal attributes."""
    from app.models import Regla
    rule = Regla(
        id=id, nombre=nombre, dominio=dominio,
        estado="active", version=1, prioridad=10, severidad="error",
        descripcion="Test rule description",
    )
    if parametros is not None:
        rule.parametros = parametros
    return rule


def _mock_condition_dict(cond_dict: dict) -> MagicMock:
    """Convert a condition dict to a MagicMock with attribute access."""
    m = MagicMock()
    for key, value in cond_dict.items():
        setattr(m, key, value)
    return m


def _mock_session_with_rule(rule, conditions=None):
    """Create a mock session that returns a specific rule and its conditions."""
    session = MagicMock()
    mock_query = MagicMock()
    mock_query.filter.return_value = mock_query
    mock_query.order_by.return_value = mock_query
    mock_query.first.return_value = rule

    if conditions:
        cond_mocks = [_mock_condition_dict(c) for c in conditions]
        mock_query.all.return_value = cond_mocks
    else:
        mock_query.all.return_value = []

    session.query.return_value = mock_query
    return session


def _make_atomic_tree(fuente_datos="invoice.convenio_facturado",
                      operador="eq", valor_esperado="PyP"):
    """Build a simple atomic condition tree dict."""
    return {
        "id": 1, "regla_id": 1, "padre_id": None,
        "tipo": "atomic", "operador": operador,
        "fuente_datos": fuente_datos,
        "valor_esperado": valor_esperado, "orden": 0,
        "_children": [],
    }


def _make_simple_sheet():
    """Create an Excel worksheet with 2 data rows."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    ws.cell(row=1, column=1, value="NUMERO_FACTURA")
    ws.cell(row=1, column=2, value="CONVENIO_FACTURADO")
    ws.cell(row=2, column=1, value="F001")
    ws.cell(row=2, column=2, value="PyP")
    ws.cell(row=3, column=1, value="F002")
    ws.cell(row=3, column=2, value="Asistencial")
    return ws


# ═══════════════════════════════════════════════════════════════════════════
# Task 1.1: Exception Caching — query_exceptions() called once per rule
# ═══════════════════════════════════════════════════════════════════════════

class TestExceptionCaching:
    """Verify apply_exceptions() uses pre-queried cache (not per-row query)."""

    def test_query_exceptions_called_once_per_rule_not_per_row(self):
        """ExceptionHandler.query_exceptions() called once, NOT N times for N rows."""
        from app.services.engine.engine import RuleEvaluationEngine
        from app.services.engine.condition_evaluator import ConditionEvaluator

        rule = _mock_rule(id=1, nombre="test_rule")
        session = _mock_session_with_rule(rule, [
            {"id": 1, "regla_id": 1, "padre_id": None,
             "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.convenio_facturado",
             "valor_esperado": "PyP", "orden": 0},
        ])

        ws = _make_simple_sheet()
        indices = {"numero_factura": 0, "convenio_facturado": 1}

        engine = RuleEvaluationEngine(session)

        # Patch query_exceptions to track calls
        engine._exception_handler.query_exceptions = MagicMock(
            return_value=[]  # No exceptions
        )
        engine._exception_handler.apply_exceptions = MagicMock(
            return_value=("normal", None)
        )

        results = engine.evaluate_sheet("test_rule", ws, indices)

        # query_exceptions should be called exactly ONCE (before loop)
        assert engine._exception_handler.query_exceptions.call_count == 1, (
            f"Expected 1 query_exceptions call, got {engine._exception_handler.query_exceptions.call_count}"
        )

        # apply_exceptions should be called for each row (2 rows)
        assert engine._exception_handler.apply_exceptions.call_count == 2, (
            f"Expected 2 apply_exceptions calls, got {engine._exception_handler.apply_exceptions.call_count}"
        )

    def test_apply_exceptions_receives_cached_result(self):
        """apply_exceptions() receives cached_exc keyword from pre-queried list."""
        from app.services.engine.engine import RuleEvaluationEngine

        rule = _mock_rule(id=1, nombre="test_rule")
        session = _mock_session_with_rule(rule, [
            {"id": 1, "regla_id": 1, "padre_id": None,
             "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.convenio_facturado",
             "valor_esperado": "PyP", "orden": 0},
        ])

        ws = _make_simple_sheet()
        indices = {"numero_factura": 0, "convenio_facturado": 1}

        engine = RuleEvaluationEngine(session)

        fake_exceptions = [MagicMock(id=99, regla_id=1)]
        engine._exception_handler.query_exceptions = MagicMock(
            return_value=fake_exceptions
        )
        engine._exception_handler.apply_exceptions = MagicMock(
            return_value=("normal", None)
        )

        engine.evaluate_sheet("test_rule", ws, indices)

        # Verify apply_exceptions was called with cached_exc kwarg
        for call_args in engine._exception_handler.apply_exceptions.call_args_list:
            kwargs = call_args.kwargs
            assert "cached_exc" in kwargs, (
                "apply_exceptions missing 'cached_exc' kwarg"
            )
            assert kwargs["cached_exc"] == fake_exceptions, (
                "cached_exc should be the result of query_exceptions"
            )


# ═══════════════════════════════════════════════════════════════════════════
# Task 1.2: Lazy date.edad — resolve only when tree references it
# ═══════════════════════════════════════════════════════════════════════════

class TestLazyDateEdad:
    """Verify _resolve_computed('date.edad') only called when tree has date refs."""

    def test_date_edad_not_resolved_when_tree_has_no_date_refs(self):
        """When tree has no date.edad references, _resolve_computed is NOT called."""
        from app.services.engine.engine import RuleEvaluationEngine

        rule = _mock_rule(id=1, nombre="test_rule")
        # Tree references invoice.convenio_facturado — NOT date.edad
        session = _mock_session_with_rule(rule, [
            {"id": 1, "regla_id": 1, "padre_id": None,
             "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.convenio_facturado",
             "valor_esperado": "PyP", "orden": 0},
        ])

        ws = _make_simple_sheet()
        indices = {"numero_factura": 0, "convenio_facturado": 1}

        engine = RuleEvaluationEngine(session)

        with patch.object(engine, "_resolve_computed", wraps=engine._resolve_computed) as mock_resolve:
            engine.evaluate_sheet("test_rule", ws, indices)

            # _resolve_computed should NEVER be called with "date.edad"
            date_edad_calls = [
                c for c in mock_resolve.call_args_list
                if c.args[0] == "date.edad"
            ]
            assert len(date_edad_calls) == 0, (
                f"_resolve_computed('date.edad') should NOT be called when tree has no date refs, "
                f"but was called {len(date_edad_calls)} times"
            )

    def test_date_edad_still_resolved_when_tree_has_date_refs(self):
        """When tree HAS date.edad references, _resolve_computed IS called."""
        from app.services.engine.engine import RuleEvaluationEngine

        rule = _mock_rule(id=1, nombre="test_rule")
        # Tree references date.edad
        session = _mock_session_with_rule(rule, [
            {"id": 1, "regla_id": 1, "padre_id": None,
             "tipo": "atomic", "operador": "gt",
             "fuente_datos": "date.edad",
             "valor_esperado": "18", "orden": 0},
        ])

        ws = _make_simple_sheet()
        indices = {"numero_factura": 0, "convenio_facturado": 1}

        engine = RuleEvaluationEngine(session)

        with patch.object(engine, "_resolve_computed", wraps=engine._resolve_computed) as mock_resolve:
            results = engine.evaluate_sheet("test_rule", ws, indices)

            # _resolve_computed MUST be called with "date.edad"
            date_edad_calls = [
                c for c in mock_resolve.call_args_list
                if c.args[0] == "date.edad"
            ]
            assert len(date_edad_calls) > 0, (
                "_resolve_computed('date.edad') should be called when tree references date.edad"
            )

    def test_tree_uses_field_detects_nested_date_refs(self):
        """tree_uses_field() detects date.edad in deeply nested trees."""
        from app.services.engine.condition_evaluator import ConditionEvaluator
        evaluator = ConditionEvaluator()

        # Build a nested tree: AND(OR(eq(invoice.convenio, "A"), eq(date.edad, ...)), gt(invoice.valor, ...))
        tree = {
            "tipo": "composite",
            "operador": "AND",
            "_children": [
                {
                    "tipo": "composite",
                    "operador": "OR",
                    "_children": [
                        {"tipo": "atomic", "operador": "eq",
                         "fuente_datos": "invoice.convenio_facturado",
                         "valor_esperado": "PyP"},
                        {"tipo": "atomic", "operador": "gt",
                         "fuente_datos": "date.edad",
                         "valor_esperado": "18"},
                    ],
                },
                {
                    "tipo": "atomic", "operador": "gt",
                    "fuente_datos": "invoice.valor",
                    "valor_esperado": "1000",
                },
            ],
        }

        assert evaluator.tree_uses_field(tree, "date.edad") is True, (
            "tree_uses_field should find date.edad in nested tree"
        )
        assert evaluator.tree_uses_field(tree, "date.edad_meses") is False, (
            "tree_uses_field should NOT find date.edad_meses when absent"
        )
        assert evaluator.tree_uses_field(tree, "invoice.valor") is True, (
            "tree_uses_field should find invoice.valor"
        )

    def test_tree_uses_field_no_match(self):
        """tree_uses_field() returns False when field not in tree."""
        from app.services.engine.condition_evaluator import ConditionEvaluator
        evaluator = ConditionEvaluator()

        tree = {
            "tipo": "composite",
            "operador": "AND",
            "_children": [
                {"tipo": "atomic", "operador": "eq",
                 "fuente_datos": "invoice.convenio_facturado",
                 "valor_esperado": "PyP"},
            ],
        }

        assert evaluator.tree_uses_field(tree, "date.edad") is False
        assert evaluator.tree_uses_field(tree, "nonexistent.field") is False


# ═══════════════════════════════════════════════════════════════════════════
# Task 1.3: NO_MATCH evidence skip
# ═══════════════════════════════════════════════════════════════════════════

class TestNoMatchSkip:
    """Verify collector.record() is NOT called for NO_MATCH outcomes."""

    def test_collector_record_not_called_for_no_match(self):
        """collector.record() is NOT called when outcome is NO_MATCH."""
        from app.services.engine.engine import RuleEvaluationEngine

        rule = _mock_rule(id=1, nombre="test_rule")
        # Tree matches "PyP" only — F002 has "Asistencial" → NO_MATCH
        session = _mock_session_with_rule(rule, [
            {"id": 1, "regla_id": 1, "padre_id": None,
             "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.convenio_facturado",
             "valor_esperado": "PyP", "orden": 0},
        ])

        ws = _make_simple_sheet()
        indices = {"numero_factura": 0, "convenio_facturado": 1}

        engine = RuleEvaluationEngine(session)
        # Replace collector with a mock to track calls
        mock_collector = MagicMock()
        engine._evidence_collector = mock_collector

        results = engine.evaluate_sheet("test_rule", ws, indices)

        # Only F001 matches PyP → one MATCH; F002 is NO_MATCH
        assert len(results) == 1, "Expected 1 MATCH result"
        assert results[0]["factura"] == "F001"

        # collector.record() should only be called for MATCH (or ERROR)
        # NOT for NO_MATCH rows
        no_match_calls = [
            c for c in mock_collector.record.call_args_list
            if c.kwargs.get("outcome") == "NO_MATCH"
        ]
        assert len(no_match_calls) == 0, (
            f"collector.record() should NOT be called for NO_MATCH, "
            f"but was called {len(no_match_calls)} times"
        )

        # But MATCH rows SHOULD be recorded
        match_calls = [
            c for c in mock_collector.record.call_args_list
            if c.kwargs.get("outcome") == "MATCH"
        ]
        assert len(match_calls) == 1, (
            f"collector.record() should be called for MATCH, "
            f"but was called {len(match_calls)} times"
        )

    def test_collector_record_still_called_for_error_outcomes(self):
        """collector.record() IS called for ERROR outcomes."""
        from app.services.engine.engine import RuleEvaluationEngine

        rule = _mock_rule(id=1, nombre="test_rule")
        # Tree with invalid evaluator — produces ERROR
        session = _mock_session_with_rule(rule, [
            {"id": 1, "regla_id": 1, "padre_id": None,
             "tipo": "atomic", "operador": "this_operator_does_not_exist",
             "fuente_datos": "invoice.convenio_facturado",
             "valor_esperado": "x", "orden": 0},
        ])

        ws = _make_simple_sheet()
        indices = {"numero_factura": 0, "convenio_facturado": 1}

        engine = RuleEvaluationEngine(session)
        mock_collector = MagicMock()
        engine._evidence_collector = mock_collector

        engine.evaluate_sheet("test_rule", ws, indices)

        # ERROR outcomes should still be recorded
        error_calls = [
            c for c in mock_collector.record.call_args_list
            if c.kwargs.get("outcome") == "ERROR"
        ]
        assert len(error_calls) > 0, (
            "collector.record() should be called for ERROR outcomes"
        )


# ═══════════════════════════════════════════════════════════════════════════
# Phase 1: Pre-resolve providers/evaluators at tree-build time
# ═══════════════════════════════════════════════════════════════════════════

class TestPreResolveProvidersEvaluators:
    """Verify build_tree() pre-resolves and caches _provider/_evaluator on nodes."""

    def test_build_tree_attaches_provider_and_evaluator_to_atomic_nodes(self):
        """build_tree() stores _provider and _evaluator on each atomic node."""
        from app.services.engine.condition_evaluator import ConditionEvaluator
        from app.services.engine.providers import InvoiceProvider
        from app.services.engine.evaluators import EqEvaluator

        evaluator = ConditionEvaluator()
        conditions = [
            {"id": 1, "regla_id": 1, "padre_id": None,
             "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.convenio_facturado",
             "valor_esperado": "PyP", "orden": 0},
        ]
        tree = evaluator.build_tree(conditions)

        assert tree is not None, "build_tree should return a tree"
        # Atomic node must have cached provider and evaluator
        assert "_provider" in tree, (
            "build_tree() must attach _provider to atomic nodes"
        )
        assert "_evaluator" in tree, (
            "build_tree() must attach _evaluator to atomic nodes"
        )
        assert tree["_provider"] is not None, (
            "Cached _provider should be a real provider instance"
        )
        assert tree["_evaluator"] is not None, (
            "Cached _evaluator should be a real evaluator instance"
        )
        # Verify correct types
        assert isinstance(tree["_provider"], InvoiceProvider), (
            f"Expected InvoiceProvider, got {type(tree['_provider'])}"
        )
        assert isinstance(tree["_evaluator"], EqEvaluator), (
            f"Expected EqEvaluator, got {type(tree['_evaluator'])}"
        )

    def test_build_tree_attaches_provider_and_evaluator_to_nested_nodes(self):
        """Composite tree: ALL atomic children get cached provider/evaluator."""
        from app.services.engine.condition_evaluator import ConditionEvaluator

        evaluator = ConditionEvaluator()
        conditions = [
            {"id": 1, "regla_id": 1, "padre_id": None,
             "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "regla_id": 1, "padre_id": 1,
             "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.convenio_facturado",
             "valor_esperado": "PyP", "orden": 0},
            {"id": 3, "regla_id": 1, "padre_id": 1,
             "tipo": "atomic", "operador": "gt",
             "fuente_datos": "invoice.valor",
             "valor_esperado": "1000", "orden": 1},
        ]
        tree = evaluator.build_tree(conditions)

        assert tree is not None
        children = tree.get("_children", [])
        assert len(children) == 2, f"Expected 2 children, got {len(children)}"

        # Both atomic children must have cached provider/evaluator
        for i, child in enumerate(children):
            assert "_provider" in child, (
                f"Child {i}: missing _provider"
            )
            assert "_evaluator" in child, (
                f"Child {i}: missing _evaluator"
            )
            assert child["_provider"] is not None, (
                f"Child {i}: _provider should not be None"
            )
            assert child["_evaluator"] is not None, (
                f"Child {i}: _evaluator should not be None"
            )

    def test_cached_provider_evaluator_used_during_evaluation(self):
        """_evaluate_fast uses node._provider/_evaluator (NO dynamic lookup)."""
        from app.services.engine.condition_evaluator import ConditionEvaluator
        from app.services.engine.context import EvaluationContext
        from unittest.mock import MagicMock

        # Build a tree with real pre-resolution
        evaluator = ConditionEvaluator()
        conditions = [
            {"id": 1, "regla_id": 1, "padre_id": None,
             "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.convenio_facturado",
             "valor_esperado": "PyP", "orden": 0},
        ]
        tree = evaluator.build_tree(conditions)
        assert tree is not None

        # Replace cached provider/evaluator with mocks
        mock_provider = MagicMock()
        mock_provider.resolve.return_value = "PyP"
        mock_evaluator = MagicMock()
        mock_evaluator.evaluate.return_value = True

        tree["_provider"] = mock_provider
        tree["_evaluator"] = mock_evaluator

        ctx = EvaluationContext(invoice_data={"convenio_facturado": "PyP"})
        result = evaluator._evaluate_fast(tree, ctx)

        assert result["outcome"] is True
        # Mocks must have been called
        mock_provider.resolve.assert_called_once()
        mock_evaluator.evaluate.assert_called_once()

    def test_fallback_to_dynamic_lookup_when_cache_missing(self):
        """Node without _provider/_evaluator falls back to dynamic lookup."""
        from app.services.engine.condition_evaluator import ConditionEvaluator
        from app.services.engine.context import EvaluationContext

        evaluator = ConditionEvaluator()
        # Manually crafted tree WITHOUT _provider/_evaluator
        tree = {
            "id": 1, "regla_id": 1, "padre_id": None,
            "tipo": "atomic", "operador": "eq",
            "fuente_datos": "invoice.convenio_facturado",
            "valor_esperado": "PyP", "orden": 0,
            "_children": [],
        }

        ctx = EvaluationContext(invoice_data={"convenio_facturado": "PyP"})
        result = evaluator._evaluate_atomic(tree, ctx)

        # Should still work via dynamic lookup fallback
        assert result["outcome"] is True, (
            f"Fallback evaluation failed: {result}"
        )


# ═══════════════════════════════════════════════════════════════════════════
# Phase 2: Eliminate redundant EvaluationContext + dict copy
# ═══════════════════════════════════════════════════════════════════════════

class TestEliminateRedundantContext:
    """Verify apply_exceptions() accepts dict directly and single ctx per row."""

    def test_apply_exceptions_accepts_row_data_dict(self):
        """apply_exceptions() works with row_data dict (not EvaluationContext)."""
        from app.services.engine.exception_handler import ExceptionHandler
        from app.models import Excepcion

        rule = _mock_rule(id=1, nombre="test_rule")
        session = MagicMock()

        # Create an exception that matches "convenio_facturado == PyP"
        exc = MagicMock(spec=Excepcion)
        exc.regla_id = 1
        exc.activo = True
        exc.tipo_efecto = "skip"
        exc.condicion_json = {"convenio_facturado": "PyP"}

        handler = ExceptionHandler()
        row_data = {"convenio_facturado": "PyP", "numero_factura": "F001"}

        # Call with row_data dict directly (NOT EvaluationContext)
        effect, overrides = handler.apply_exceptions(
            rule, row_data, session, cached_exc=[exc],
        )

        assert effect == "skip", (
            f"Expected 'skip' when scope matches, got '{effect}'"
        )

    def test_apply_exceptions_row_data_does_not_match(self):
        """apply_exceptions() with dict: no match when scope differs."""
        from app.services.engine.exception_handler import ExceptionHandler
        from app.models import Excepcion

        rule = _mock_rule(id=1, nombre="test_rule")
        session = MagicMock()

        exc = MagicMock(spec=Excepcion)
        exc.regla_id = 1
        exc.activo = True
        exc.tipo_efecto = "skip"
        exc.condicion_json = {"convenio_facturado": "Asistencial"}

        handler = ExceptionHandler()
        row_data = {"convenio_facturado": "PyP", "numero_factura": "F001"}

        effect, overrides = handler.apply_exceptions(
            rule, row_data, session, cached_exc=[exc],
        )

        assert effect == "normal", (
            f"Expected 'normal' when scope doesn't match, got '{effect}'"
        )

    def test_apply_exceptions_empty_scope_matches_all(self):
        """apply_exceptions() with dict: empty scope matches any row."""
        from app.services.engine.exception_handler import ExceptionHandler
        from app.models import Excepcion

        rule = _mock_rule(id=1, nombre="test_rule")
        session = MagicMock()

        exc = MagicMock(spec=Excepcion)
        exc.regla_id = 1
        exc.activo = True
        exc.tipo_efecto = "skip"
        exc.condicion_json = {}  # Empty scope = match all

        handler = ExceptionHandler()
        row_data = {"convenio_facturado": "Anything"}

        effect, overrides = handler.apply_exceptions(
            rule, row_data, session, cached_exc=[exc],
        )

        assert effect == "skip", (
            f"Empty scope should match all rows, got '{effect}'"
        )


# ═══════════════════════════════════════════════════════════════════════════
# Phase 3: Two-phase evaluation in production
# ═══════════════════════════════════════════════════════════════════════════

class TestTwoPhaseEvaluation:
    """Verify fast path first, trace only for MATCH/ERROR when persist=True."""

    def test_fast_path_first_trace_only_for_match(self):
        """When persist=True and outcome is MATCH, trace is re-evaluated."""
        from app.services.engine.condition_evaluator import ConditionEvaluator
        from app.services.engine.context import EvaluationContext

        evaluator = ConditionEvaluator()
        conditions = [
            {"id": 1, "regla_id": 1, "padre_id": None,
             "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.convenio_facturado",
             "valor_esperado": "PyP", "orden": 0},
        ]
        tree = evaluator.build_tree(conditions)
        assert tree is not None

        ctx = EvaluationContext(invoice_data={"convenio_facturado": "PyP"})

        # Fast path: NO trace
        fast_result = evaluator._evaluate_fast(tree, ctx)
        assert fast_result["outcome"] is True
        assert "trace" not in fast_result, (
            "_evaluate_fast should NOT include trace dict"
        )

        # Full trace path: HAS trace
        full_result = evaluator.evaluate(tree, ctx, collect_trace=True)
        assert full_result["outcome"] is True
        assert "trace" in full_result, (
            "evaluate(collect_trace=True) must include trace dict"
        )

    def test_fast_path_no_trace_for_no_match(self):
        """_evaluate_fast returns compact result (no trace overhead)."""
        from app.services.engine.condition_evaluator import ConditionEvaluator
        from app.services.engine.context import EvaluationContext

        evaluator = ConditionEvaluator()
        conditions = [
            {"id": 1, "regla_id": 1, "padre_id": None,
             "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.convenio_facturado",
             "valor_esperado": "PyP", "orden": 0},
        ]
        tree = evaluator.build_tree(conditions)
        assert tree is not None

        # Data that does NOT match
        ctx = EvaluationContext(invoice_data={"convenio_facturado": "Asistencial"})

        result = evaluator._evaluate_fast(tree, ctx)
        assert result["outcome"] is False
        assert "trace" not in result, (
            "_evaluate_fast should return compact dict, no trace"
        )

    def test_collect_trace_false_uses_fast_path(self):
        """evaluate(collect_trace=False) delegates to _evaluate_fast."""
        from app.services.engine.condition_evaluator import ConditionEvaluator
        from app.services.engine.context import EvaluationContext

        evaluator = ConditionEvaluator()
        conditions = [
            {"id": 1, "regla_id": 1, "padre_id": None,
             "tipo": "atomic", "operador": "eq",
             "fuente_datos": "invoice.convenio_facturado",
             "valor_esperado": "PyP", "orden": 0},
        ]
        tree = evaluator.build_tree(conditions)
        assert tree is not None

        ctx = EvaluationContext(invoice_data={"convenio_facturado": "PyP"})

        result = evaluator.evaluate(tree, ctx, collect_trace=False)
        assert result["outcome"] is True
        assert "trace" not in result, (
            "collect_trace=False must NOT produce trace dict"
        )
