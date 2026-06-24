"""Unit tests for RuleEvaluationEngine — orchestrates full evaluation flow."""

from __future__ import annotations

import pytest
from unittest.mock import MagicMock, patch


def _make_condition_dict(cond_dict: dict) -> MagicMock:
    """Convert a condition dict to a MagicMock with attribute access."""
    m = MagicMock()
    for key, value in cond_dict.items():
        setattr(m, key, value)
    return m


def _mock_session_with_rule(rule, conditions=None):
    """Create a mock session that returns a specific rule and its conditions."""
    session = MagicMock()
    mock_query = MagicMock()
    mock_query.filter.return_value = mock_query
    mock_query.order_by.return_value = mock_query
    mock_query.first.return_value = rule

    # For condition queries, return mock objects with proper attributes
    if conditions:
        cond_mocks = [_make_condition_dict(c) for c in conditions]
        mock_query.all.return_value = cond_mocks
    else:
        mock_query.all.return_value = []

    session.query.return_value = mock_query
    return session


class TestRuleEvaluationEngine:
    """Tests for RuleEvaluationEngine — orchestrator."""

    def test_import_exists(self):
        from app.services.engine.engine import RuleEvaluationEngine
        assert RuleEvaluationEngine is not None

    def test_evaluate_sheet_returns_list(self):
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla, Condicion
        from openpyxl import Workbook

        # Build a simple rule
        rule = Regla(
            id=1, nombre="test_rule", dominio="odontologia",
            estado="active", version=1, prioridad=10, severidad="error",
        )
        # Condition: eq(convenio, "PyP")
        root_cond = {
            "id": 1, "regla_id": 1, "padre_id": None,
            "tipo": "atomic", "operador": "eq",
            "fuente_datos": "invoice.convenio_facturado",
            "valor_esperado": "PyP", "orden": 0,
        }
        session = _mock_session_with_rule(rule, [root_cond])

        # Create a test worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CONVENIO_FACTURADO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="PyP")
        ws.cell(row=3, column=1, value="F002")
        ws.cell(row=3, column=2, value="Asistencial")

        indices = {"numero_factura": 0, "convenio_facturado": 1}

        engine = RuleEvaluationEngine(session)
        results = engine.evaluate_sheet("test_rule", ws, indices)
        assert isinstance(results, list)

    def test_evaluate_sheet_returns_legacy_format(self):
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla
        from openpyxl import Workbook

        rule = Regla(
            id=1, nombre="test_rule", dominio="odontologia",
            estado="active", version=1, prioridad=10, severidad="error",
        )
        root_cond = {
            "id": 1, "regla_id": 1, "padre_id": None,
            "tipo": "atomic", "operador": "eq",
            "fuente_datos": "invoice.convenio_facturado",
            "valor_esperado": "PyP", "orden": 0,
        }
        session = _mock_session_with_rule(rule, [root_cond])

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CONVENIO_FACTURADO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="PyP")

        indices = {"numero_factura": 0, "convenio_facturado": 1}

        engine = RuleEvaluationEngine(session)
        results = engine.evaluate_sheet("test_rule", ws, indices)

        # Each result should have factura and problema keys
        for r in results:
            assert "factura" in r
            assert "problema" in r

    def test_build_row_context_extracts_cell_values(self):
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla
        from openpyxl import Workbook

        rule = Regla(
            id=1, nombre="test_rule", dominio="odontologia",
            estado="active", version=1, prioridad=10, severidad="error",
        )
        session = _mock_session_with_rule(rule, [])

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CONVENIO_FACTURADO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="PyP")

        indices = {"numero_factura": 0, "convenio_facturado": 1}

        engine = RuleEvaluationEngine(session)
        row_data, factura = engine._build_row_context(ws, 2, indices)
        assert factura == "F001"
        assert row_data["numero_factura"] == "F001"
        assert row_data["convenio_facturado"] == "PyP"


class TestEvaluateSheetDomain:
    """Tests for evaluate_sheet_domain — loads all rules for a domain and evaluates."""

    def test_method_exists(self):
        """evaluate_sheet_domain is a callable method on RuleEvaluationEngine."""
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla

        session = _mock_session_with_rule(
            Regla(id=1, nombre="r1", dominio="odontologia", estado="active", version=1, prioridad=10, severidad="error"),
            [],
        )
        engine = RuleEvaluationEngine(session)
        assert callable(engine.evaluate_sheet_domain)

    def test_resolves_domain_rules_and_returns_empty_for_no_matches(self):
        """When resolver returns empty rules list, evaluate_sheet_domain returns empty."""
        from app.services.engine.engine import RuleEvaluationEngine
        from openpyxl import Workbook
        from unittest.mock import patch

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=2, column=1, value="F001")
        indices = {"numero_factura": 0}

        session = MagicMock()
        with patch.object(RuleEvaluationEngine, "__init__", lambda self, s: setattr(self, "_session", s)):
            engine = RuleEvaluationEngine.__new__(RuleEvaluationEngine)
            engine._session = session
            engine._resolver = MagicMock()
            engine._resolver.resolve.return_value = []

            results = engine.evaluate_sheet_domain("odontologia", ws, indices)
            assert isinstance(results, list)
            assert len(results) == 0

    def test_evaluates_each_rule_and_combines_results(self):
        """evaluate_sheet_domain calls evaluate_sheet for each rule and combines."""
        from app.services.engine.engine import RuleEvaluationEngine
        from openpyxl import Workbook
        from unittest.mock import patch, MagicMock

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=2, column=1, value="F001")
        indices = {"numero_factura": 0}

        # Create 3 rules
        from app.models import Regla
        r1 = Regla(id=1, nombre="rule_a", dominio="odontologia", estado="active", version=1, prioridad=10, severidad="error")
        r2 = Regla(id=2, nombre="rule_b", dominio="transversal", estado="active", version=1, prioridad=20, severidad="warning")
        r3 = Regla(id=3, nombre="rule_c", dominio="odontologia", estado="active", version=1, prioridad=30, severidad="error")

        session = MagicMock()
        engine = RuleEvaluationEngine.__new__(RuleEvaluationEngine)
        engine._session = session
        engine._resolver = MagicMock()
        engine._resolver.resolve.return_value = [r1, r2, r3]

        # Mock evaluate_sheet to return per-rule results
        calls_log = []

        def mock_evaluate_sheet(rule_name, ds, inds):
            calls_log.append(rule_name)
            if rule_name == "rule_a":
                return [{"factura": "F001", "problema": "A", "regla": "rule_a", "severidad": "error"}]
            elif rule_name == "rule_b":
                return [{"factura": "F001", "problema": "B", "regla": "rule_b", "severidad": "warning"}]
            else:
                return []

        engine.evaluate_sheet = mock_evaluate_sheet

        results = engine.evaluate_sheet_domain("odontologia", ws, indices)
        assert len(results) == 2
        assert results[0]["regla"] == "rule_a"
        assert results[1]["regla"] == "rule_b"

    def test_passes_correct_domain_to_resolver(self):
        """evaluate_sheet_domain passes domain argument to resolver.resolve()."""
        from app.services.engine.engine import RuleEvaluationEngine
        from openpyxl import Workbook
        from unittest.mock import MagicMock

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=2, column=1, value="F001")
        indices = {"numero_factura": 0}

        session = MagicMock()
        engine = RuleEvaluationEngine.__new__(RuleEvaluationEngine)
        engine._session = session
        engine._resolver = MagicMock()
        engine._resolver.resolve.return_value = []

        engine.evaluate_sheet = MagicMock(return_value=[])

        engine.evaluate_sheet_domain("urgencias", ws, indices)
        engine._resolver.resolve.assert_called_once_with("urgencias", session)


class TestGroupByRouting:
    """Tests for engine routing to GroupEvaluator when rule has group_by parametros."""

    def test_group_by_rule_routes_to_group_evaluator(self):
        """Rule with group_by in parametros → GroupEvaluator path, not row-by-row."""
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla
        from openpyxl import Workbook

        rule = Regla(
            id=1, nombre="doble_tipo_procedimiento",
            dominio="transversal", estado="active", version=1,
            prioridad=10, severidad="error",
            descripcion="Doble tipo de procedimiento",
            parametros=[{
                "group_by": "numero_factura",
                "aggregations": [
                    {"function": "distinct_count", "field": "tipo_procedimiento",
                     "target": "distinct_count_tipo_procedimiento"},
                ],
            }],
        )
        # Simple condition: gt(invoice.distinct_count_tipo_procedimiento, 1)
        root_cond = {
            "id": 1, "regla_id": 1, "padre_id": None,
            "tipo": "atomic", "operador": "gt",
            "fuente_datos": "invoice.distinct_count_tipo_procedimiento",
            "valor_esperado": "1", "orden": 0,
        }

        from unittest.mock import MagicMock as M
        cond_mock = M()
        cond_mock.id = 1; cond_mock.regla_id = 1; cond_mock.padre_id = None
        cond_mock.tipo = "atomic"; cond_mock.operador = "gt"
        cond_mock.fuente_datos = "invoice.distinct_count_tipo_procedimiento"
        cond_mock.valor_esperado = "1"; cond_mock.orden = 0

        session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.first.return_value = rule
        mock_query.all.return_value = [cond_mock]
        session.query.return_value = mock_query

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="B")  # 2 tipos → MATCH

        indices = {"numero_factura": 0, "tipo_procedimiento": 1}

        engine = RuleEvaluationEngine(session)
        results = engine.evaluate_sheet("doble_tipo_procedimiento", ws, indices)

        assert isinstance(results, list)
        assert len(results) == 1
        assert results[0]["factura"] == "F001"
        assert "regla" in results[0]
        assert results[0]["regla"] == "doble_tipo_procedimiento"

    def test_no_group_by_uses_existing_row_by_row(self):
        """Rules without group_by still use the existing row-by-row path."""
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla
        from openpyxl import Workbook

        rule = Regla(
            id=1, nombre="test_rule", dominio="odontologia",
            estado="active", version=1, prioridad=10, severidad="error",
            parametros=[{}],  # No group_by
        )
        root_cond = {
            "id": 1, "regla_id": 1, "padre_id": None,
            "tipo": "atomic", "operador": "eq",
            "fuente_datos": "invoice.convenio_facturado",
            "valor_esperado": "PyP", "orden": 0,
        }

        from unittest.mock import MagicMock as M
        cond_mock = M()
        cond_mock.id = 1; cond_mock.regla_id = 1; cond_mock.padre_id = None
        cond_mock.tipo = "atomic"; cond_mock.operador = "eq"
        cond_mock.fuente_datos = "invoice.convenio_facturado"
        cond_mock.valor_esperado = "PyP"; cond_mock.orden = 0

        session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.first.return_value = rule
        mock_query.all.return_value = [cond_mock]
        session.query.return_value = mock_query

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CONVENIO_FACTURADO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="PyP")

        indices = {"numero_factura": 0, "convenio_facturado": 1}

        engine = RuleEvaluationEngine(session)
        results = engine.evaluate_sheet("test_rule", ws, indices)

        # Still works via row-by-row path
        assert isinstance(results, list)
        assert len(results) == 1
        assert results[0]["factura"] == "F001"
