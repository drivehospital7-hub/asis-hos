"""Tests for engine rows path — verify RowStore produces identical output to Worksheet path.

Strict TDD: These tests MUST fail initially because neither the ``rows`` nor
``evidence_collector`` parameters exist yet on ``evaluate_sheet()``.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook
from unittest.mock import MagicMock


# ── Helpers (mirrored from test_engine.py) ──────────────────────────────


def _make_condition_dict(cond_dict: dict) -> MagicMock:
    """Convert a condition dict to a MagicMock with attribute access."""
    m = MagicMock()
    for key, value in cond_dict.items():
        setattr(m, key, value)
    return m


def _mock_session_with_rule(rule, conditions=None):
    """Create a mock session that returns a specific rule and its conditions."""
    session = MagicMock()
    mock_query = MagicMock()
    mock_query.filter.return_value = mock_query
    mock_query.order_by.return_value = mock_query
    mock_query.first.return_value = rule

    if conditions:
        cond_mocks = [_make_condition_dict(c) for c in conditions]
        mock_query.all.return_value = cond_mocks
    else:
        mock_query.all.return_value = []

    session.query.return_value = mock_query
    return session


# ── Tests ───────────────────────────────────────────────────────────────


class TestEngineRowsPathEquality:
    """Rows path (list[dict]) MUST produce identical results to Worksheet path."""

    def test_simple_eq_rule_identical_results(self):
        """Simple eq rule: rows path produces identical results to worksheet path."""
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla

        rule = Regla(
            id=1, nombre="test_simple_eq", dominio="odontologia",
            estado="active", version=1, prioridad=10, severidad="error",
        )
        root_cond = {
            "id": 1, "regla_id": 1, "padre_id": None,
            "tipo": "atomic", "operador": "eq",
            "fuente_datos": "invoice.convenio_facturado",
            "valor_esperado": "PyP", "orden": 0,
        }
        session = _mock_session_with_rule(rule, [root_cond])

        # ── Build worksheet (old path) ──
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CONVENIO_FACTURADO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="PyP")        # MATCH
        ws.cell(row=3, column=1, value="F002")
        ws.cell(row=3, column=2, value="Asistencial") # NO MATCH
        ws.cell(row=4, column=1, value="F003")
        ws.cell(row=4, column=2, value="PyP")         # MATCH

        indices = {"numero_factura": 0, "convenio_facturado": 1}

        # ── Build equivalent dict rows (new path) ──
        rows = [
            {"numero_factura": "F001", "convenio_facturado": "PyP"},
            {"numero_factura": "F002", "convenio_facturado": "Asistencial"},
            {"numero_factura": "F003", "convenio_facturado": "PyP"},
        ]

        engine = RuleEvaluationEngine(session)

        # Old path — positional args unchanged
        results_old = engine.evaluate_sheet(
            "test_simple_eq", ws, indices, persist=False,
        )
        # New path — uses rows= keyword
        results_new = engine.evaluate_sheet(
            "test_simple_eq", ws, indices, persist=False, rows=rows,
        )
        # NOTE: RED phase — 'rows' kwarg does NOT exist yet, this WILL fail
        assert results_old == results_new
        assert len(results_old) == 2  # F001 and F003

    def test_group_by_rule_identical_results(self):
        """Group-by rule: rows path produces identical results to worksheet path."""
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla

        rule = Regla(
            id=2, nombre="doble_tipo_procedimiento",
            dominio="transversal", estado="active", version=1,
            prioridad=10, severidad="error",
            descripcion="Doble tipo de procedimiento",
            parametros=[{
                "group_by": "numero_factura",
                "aggregations": [
                    {"function": "distinct_count", "field": "tipo_procedimiento",
                     "target": "distinct_count_tipo_procedimiento"},
                ],
            }],
        )
        root_cond = {
            "id": 1, "regla_id": 2, "padre_id": None,
            "tipo": "atomic", "operador": "gt",
            "fuente_datos": "invoice.distinct_count_tipo_procedimiento",
            "valor_esperado": "1", "orden": 0,
        }
        session = _mock_session_with_rule(rule, [root_cond])

        # ── Build worksheet ──
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        # F001: 1 tipo → NO MATCH
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="A")
        # F002: 2 tipos → MATCH
        ws.cell(row=4, column=1, value="F002")
        ws.cell(row=4, column=2, value="A")
        ws.cell(row=5, column=1, value="F002")
        ws.cell(row=5, column=2, value="B")

        indices = {"numero_factura": 0, "tipo_procedimiento": 1}

        # ── Build equivalent dict rows ──
        rows = [
            {"numero_factura": "F001", "tipo_procedimiento": "A"},
            {"numero_factura": "F001", "tipo_procedimiento": "A"},
            {"numero_factura": "F002", "tipo_procedimiento": "A"},
            {"numero_factura": "F002", "tipo_procedimiento": "B"},
        ]

        engine = RuleEvaluationEngine(session)

        results_old = engine.evaluate_sheet(
            "doble_tipo_procedimiento", ws, indices, persist=False,
        )
        results_new = engine.evaluate_sheet(
            "doble_tipo_procedimiento", ws, indices, persist=False, rows=rows,
        )

        assert results_old == results_new
        assert len(results_old) == 1
        assert results_old[0]["factura"] == "F002"

    def test_empty_rows_produces_no_results(self):
        """Engine with empty rows list returns empty results."""
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla

        rule = Regla(
            id=3, nombre="test_empty", dominio="odontologia",
            estado="active", version=1, prioridad=10, severidad="error",
        )
        root_cond = {
            "id": 1, "regla_id": 3, "padre_id": None,
            "tipo": "atomic", "operador": "eq",
            "fuente_datos": "invoice.convenio_facturado",
            "valor_esperado": "PyP", "orden": 0,
        }
        session = _mock_session_with_rule(rule, [root_cond])

        indices = {"numero_factura": 0, "convenio_facturado": 1}

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CONVENIO_FACTURADO")

        engine = RuleEvaluationEngine(session)

        results_old = engine.evaluate_sheet(
            "test_empty", ws, indices, persist=False,
        )
        results_new = engine.evaluate_sheet(
            "test_empty", ws, indices, persist=False, rows=[],
        )

        assert results_old == results_new
        assert results_old == []

    def test_missing_column_in_rows_still_processes(self):
        """Rows missing a column continue without error (None for missing key)."""
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla

        rule = Regla(
            id=4, nombre="test_missing_col", dominio="odontologia",
            estado="active", version=1, prioridad=10, severidad="error",
        )
        root_cond = {
            "id": 1, "regla_id": 4, "padre_id": None,
            "tipo": "atomic", "operador": "eq",
            "fuente_datos": "invoice.convenio_facturado",
            "valor_esperado": "PyP", "orden": 0,
        }
        session = _mock_session_with_rule(rule, [root_cond])

        indices = {"numero_factura": 0, "convenio_facturado": 1}

        # Rows where some entries are missing the 'convenio_facturado' key
        rows = [
            {"numero_factura": "F001", "convenio_facturado": "PyP"},
            {"numero_factura": "F002"},  # missing key → None
            {"numero_factura": "F003", "convenio_facturado": "PyP"},
        ]

        engine = RuleEvaluationEngine(session)
        results = engine.evaluate_sheet(
            "test_missing_col", None, indices, persist=False, rows=rows,
        )

        # Only F001 and F003 have PyP → 2 results
        assert len(results) == 2
        assert results[0]["factura"] == "F001"
        assert results[1]["factura"] == "F003"


class TestEngineRowsPathWithEvidenceCollector:
    """When evidence_collector is provided, engine uses it instead of internal one."""

    def test_uses_external_evidence_collector(self):
        """External collector records evidence instead of internal engine collector."""
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla

        rule = Regla(
            id=5, nombre="test_ext_collector", dominio="odontologia",
            estado="active", version=1, prioridad=10, severidad="error",
        )
        root_cond = {
            "id": 1, "regla_id": 5, "padre_id": None,
            "tipo": "atomic", "operador": "eq",
            "fuente_datos": "invoice.convenio_facturado",
            "valor_esperado": "PyP", "orden": 0,
        }
        session = _mock_session_with_rule(rule, [root_cond])

        rows = [
            {"numero_factura": "F001", "convenio_facturado": "PyP"},
            {"numero_factura": "F002", "convenio_facturado": "NoMatch"},
        ]
        indices = {"numero_factura": 0, "convenio_facturado": 1}

        from app.services.engine.evidence_collector import EvidenceCollector
        external_collector = EvidenceCollector()

        engine = RuleEvaluationEngine(session)

        # Pass external collector — engine should NOT use its own
        results = engine.evaluate_sheet(
            "test_ext_collector", None, indices,
            persist=True, rows=rows, evidence_collector=external_collector,
        )

        # External collector should have records
        assert len(external_collector._buffer) >= 1
        # Engine's internal collector should be empty
        assert len(engine._evidence_collector._buffer) == 0
        assert len(results) == 1
        assert results[0]["factura"] == "F001"

    def test_external_collector_no_flush_by_engine(self):
        """When external collector is used, engine does NOT call flush_batch."""
        from app.services.engine.engine import RuleEvaluationEngine
        from app.models import Regla

        rule = Regla(
            id=6, nombre="test_no_flush", dominio="odontologia",
            estado="active", version=1, prioridad=10, severidad="error",
        )
        root_cond = {
            "id": 1, "regla_id": 6, "padre_id": None,
            "tipo": "atomic", "operador": "eq",
            "fuente_datos": "invoice.convenio_facturado",
            "valor_esperado": "PyP", "orden": 0,
        }
        session = _mock_session_with_rule(rule, [root_cond])

        rows = [
            {"numero_factura": "F001", "convenio_facturado": "PyP"},
        ]
        indices = {"numero_factura": 0, "convenio_facturado": 1}

        from app.services.engine.evidence_collector import EvidenceCollector
        external_collector = EvidenceCollector()

        # Spy on flush_batch
        original_flush = external_collector.flush_batch
        flush_called = False

        def spy_flush(session):
            nonlocal flush_called
            flush_called = True
            return original_flush(session)

        external_collector.flush_batch = spy_flush

        engine = RuleEvaluationEngine(session)
        engine.evaluate_sheet(
            "test_no_flush", None, indices,
            persist=True, rows=rows, evidence_collector=external_collector,
        )

        # Engine must NOT have called flush_batch on external collector
        assert not flush_called, (
            "Engine should not flush an external evidence collector — "
            "the caller owns the flush lifecycle"
        )
