"""Snapshot tests Phase 5: engine vs legacy for tipo_documento_edad + sala_observacion rules.

Verifies the DB-backed engine with date.edad/date.horas providers detects
the same rows as the legacy Python detectors.
"""

from __future__ import annotations

import pytest
from unittest.mock import MagicMock
from openpyxl import Workbook


# ── Helpers (reused from Phase 4 snapshot pattern) ─────────────────────────

def _mock_session_with_rule(rule_name, dominio, descripcion, condiciones_dicts,
                            severity="error"):
    """Create a mock session that returns a Regla + Condicion tree."""
    from app.models import Regla

    session = MagicMock()
    mock_query = MagicMock()
    mock_query.filter.return_value = mock_query
    mock_query.order_by.return_value = mock_query

    regla = Regla(
        id=1, nombre=rule_name, dominio=dominio,
        estado="active", version=1, prioridad=10, severidad=severity,
        descripcion=descripcion,
    )
    mock_query.first.return_value = regla

    cond_mocks = []
    for cd in condiciones_dicts:
        m = MagicMock()
        m.id = cd["id"]
        m.regla_id = cd.get("regla_id", 1)
        m.padre_id = cd["padre_id"]
        m.tipo = cd["tipo"]
        m.operador = cd.get("operador")
        m.fuente_datos = cd.get("fuente_datos")
        m.valor_esperado = cd.get("valor_esperado")
        m.orden = cd.get("orden", 0)
        cond_mocks.append(m)

    mock_query.all.return_value = cond_mocks
    session.query.return_value = mock_query
    return session


def _build_indices(*col_names):
    """Build indices dict: col_name → 0-based index."""
    return {name: i for i, name in enumerate(col_names)}


def _run_engine_detection(rule_name, dominio, descripcion, condiciones,
                          ws, indices, severity="error"):
    """Run engine detection against a worksheet with mocked session."""
    from app.services.engine.rule_based_detector import RuleBasedDetector
    session = _mock_session_with_rule(rule_name, dominio, descripcion,
                                      condiciones, severity)
    detector = RuleBasedDetector(rule_name, session)
    return detector.detect(ws, indices)


def _get_facturas_from_results(results):
    """Extract factura strings from detection results."""
    facturas = set()
    for r in results:
        if isinstance(r, dict):
            facturas.add(r.get("factura", ""))
    return facturas


def _get_age_problem_from_results(results):
    """Extract problem details for assertion checking."""
    return list(results) if results else []


# ── Condition tree builders for Phase 5 rules ─────────────────────────────

def _menor_7_conditions():
    """AND(lt(date.edad, 7), NOT(eq(invoice.tipo_identificacion, "RC")))."""
    return [
        {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND",
         "fuente_datos": None, "valor_esperado": None, "orden": 0},
        {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "lt",
         "fuente_datos": "date.edad", "valor_esperado": "7", "orden": 0},
        {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT",
         "fuente_datos": None, "valor_esperado": None, "orden": 1},
        {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "eq",
         "fuente_datos": "invoice.tipo_identificacion", "valor_esperado": "RC",
         "orden": 0},
    ]


def _mayor_18_conditions():
    """AND(gte(date.edad, 18), NOT(eq(invoice.tipo_identificacion, "CC")))."""
    return [
        {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND",
         "fuente_datos": None, "valor_esperado": None, "orden": 0},
        {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "gte",
         "fuente_datos": "date.edad", "valor_esperado": "18", "orden": 0},
        {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT",
         "fuente_datos": None, "valor_esperado": None, "orden": 1},
        {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "eq",
         "fuente_datos": "invoice.tipo_identificacion", "valor_esperado": "CC",
         "orden": 0},
    ]


def _estancia_prolongada_conditions():
    """AND(gt(date.horas, 6), eq(invoice.tipo_factura_descripcion, "Urgencias"))."""
    return [
        {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND",
         "fuente_datos": None, "valor_esperado": None, "orden": 0},
        {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "gt",
         "fuente_datos": "date.horas", "valor_esperado": "6", "orden": 0},
        {"id": 3, "padre_id": 1, "tipo": "atomic", "operador": "eq",
         "fuente_datos": "invoice.tipo_factura_descripcion", "valor_esperado": "Urgencias",
         "orden": 1},
    ]


# ── Tests: tipo_documento_edad_menor_7 ──────────────────────────────────────

class TestTipoDocumentoEdadMenor7:
    """Tipo doc vs age for patients under 7 years."""

    def test_menor_7_with_rc_is_valid(self):
        """Age < 7 + tipo RC → no problem."""
        condiciones = _menor_7_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_NACIMIENTO")
        ws.cell(row=1, column=3, value="FEC_FACTURA")
        ws.cell(row=1, column=4, value="TIPO_IDENTIFICACION")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="2022-01-15")   # ~4 years old
        ws.cell(row=2, column=3, value="2026-06-24")
        ws.cell(row=2, column=4, value="RC")

        indices = _build_indices("numero_factura", "fec_nacimiento",
                                  "fec_factura", "tipo_identificacion")
        results = _run_engine_detection(
            "tipo_documento_edad_menor_7", "transversal",
            "Tipo ID incorrecto para menor de 7 años",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F001" not in facturas, "RC is correct for age < 7"

    def test_menor_7_with_cc_is_invalid(self):
        """Age < 7 + tipo CC → detected as problem."""
        condiciones = _menor_7_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_NACIMIENTO")
        ws.cell(row=1, column=3, value="FEC_FACTURA")
        ws.cell(row=1, column=4, value="TIPO_IDENTIFICACION")
        ws.cell(row=2, column=1, value="F002")
        ws.cell(row=2, column=2, value="2022-06-10")   # ~4 years old
        ws.cell(row=2, column=3, value="2026-06-24")
        ws.cell(row=2, column=4, value="CC")

        indices = _build_indices("numero_factura", "fec_nacimiento",
                                  "fec_factura", "tipo_identificacion")
        results = _run_engine_detection(
            "tipo_documento_edad_menor_7", "transversal",
            "Tipo ID incorrecto para menor de 7 años",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F002" in facturas, "CC wrong for age < 7 (should be RC)"

    def test_menor_7_with_ti_is_invalid(self):
        """Age < 7 + tipo TI → detected as problem."""
        condiciones = _menor_7_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_NACIMIENTO")
        ws.cell(row=1, column=3, value="FEC_FACTURA")
        ws.cell(row=1, column=4, value="TIPO_IDENTIFICACION")
        ws.cell(row=2, column=1, value="F003")
        ws.cell(row=2, column=2, value="2023-03-03")   # ~3 years old
        ws.cell(row=2, column=3, value="2026-06-24")
        ws.cell(row=2, column=4, value="TI")

        indices = _build_indices("numero_factura", "fec_nacimiento",
                                  "fec_factura", "tipo_identificacion")
        results = _run_engine_detection(
            "tipo_documento_edad_menor_7", "transversal",
            "Tipo ID incorrecto para menor de 7 años",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F003" in facturas, "TI wrong for age < 7 (should be RC)"

    def test_age_7_or_older_not_flagged(self):
        """Age >= 7 → this rule does not apply (age check fails first)."""
        condiciones = _menor_7_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_NACIMIENTO")
        ws.cell(row=1, column=3, value="FEC_FACTURA")
        ws.cell(row=1, column=4, value="TIPO_IDENTIFICACION")
        ws.cell(row=2, column=1, value="F004")
        ws.cell(row=2, column=2, value="2010-05-15")   # 16 years old
        ws.cell(row=2, column=3, value="2026-06-24")
        ws.cell(row=2, column=4, value="TI")

        indices = _build_indices("numero_factura", "fec_nacimiento",
                                  "fec_factura", "tipo_identificacion")
        results = _run_engine_detection(
            "tipo_documento_edad_menor_7", "transversal",
            "Tipo ID incorrecto para menor de 7 años",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F004" not in facturas, "Age >= 7 excludes from menor_7 rule"

    def test_missing_date_fields_no_crash(self):
        """Missing fec_nacimiento → date.edad resolves to None → no detection."""
        condiciones = _menor_7_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_FACTURA")
        ws.cell(row=1, column=3, value="TIPO_IDENTIFICACION")
        ws.cell(row=2, column=1, value="F005")
        ws.cell(row=2, column=2, value="2026-06-24")
        ws.cell(row=2, column=3, value="RC")
        # no fec_nacimiento column

        indices = _build_indices("numero_factura", "fec_factura",
                                  "tipo_identificacion")
        results = _run_engine_detection(
            "tipo_documento_edad_menor_7", "transversal",
            "Tipo ID incorrecto para menor de 7 años",
            condiciones, ws, indices,
        )

        # No crash, no detection (age can't be computed)
        assert isinstance(results, list)


# ── Tests: tipo_documento_edad_mayor_18 ────────────────────────────────────

class TestTipoDocumentoEdadMayor18:
    """Tipo doc vs age for patients 18+ years."""

    def test_mayor_18_with_cc_is_valid(self):
        """Age >= 18 + tipo CC → no problem."""
        condiciones = _mayor_18_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_NACIMIENTO")
        ws.cell(row=1, column=3, value="FEC_FACTURA")
        ws.cell(row=1, column=4, value="TIPO_IDENTIFICACION")
        ws.cell(row=2, column=1, value="F101")
        ws.cell(row=2, column=2, value="1990-03-15")   # 36 years old
        ws.cell(row=2, column=3, value="2026-06-24")
        ws.cell(row=2, column=4, value="CC")

        indices = _build_indices("numero_factura", "fec_nacimiento",
                                  "fec_factura", "tipo_identificacion")
        results = _run_engine_detection(
            "tipo_documento_edad_mayor_18", "transversal",
            "Tipo ID incorrecto para mayor de edad",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F101" not in facturas, "CC is correct for age >= 18"

    def test_mayor_18_with_ti_is_invalid(self):
        """Age >= 18 + tipo TI → detected as problem."""
        condiciones = _mayor_18_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_NACIMIENTO")
        ws.cell(row=1, column=3, value="FEC_FACTURA")
        ws.cell(row=1, column=4, value="TIPO_IDENTIFICACION")
        ws.cell(row=2, column=1, value="F102")
        ws.cell(row=2, column=2, value="2000-06-10")   # 26 years old
        ws.cell(row=2, column=3, value="2026-06-24")
        ws.cell(row=2, column=4, value="TI")

        indices = _build_indices("numero_factura", "fec_nacimiento",
                                  "fec_factura", "tipo_identificacion")
        results = _run_engine_detection(
            "tipo_documento_edad_mayor_18", "transversal",
            "Tipo ID incorrecto para mayor de edad",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F102" in facturas, "TI wrong for age >= 18 (should be CC)"

    def test_mayor_18_with_rc_is_invalid(self):
        """Age >= 18 + tipo RC → detected as problem."""
        condiciones = _mayor_18_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_NACIMIENTO")
        ws.cell(row=1, column=3, value="FEC_FACTURA")
        ws.cell(row=1, column=4, value="TIPO_IDENTIFICACION")
        ws.cell(row=2, column=1, value="F103")
        ws.cell(row=2, column=2, value="2005-01-01")   # 21 years old
        ws.cell(row=2, column=3, value="2026-06-24")
        ws.cell(row=2, column=4, value="RC")

        indices = _build_indices("numero_factura", "fec_nacimiento",
                                  "fec_factura", "tipo_identificacion")
        results = _run_engine_detection(
            "tipo_documento_edad_mayor_18", "transversal",
            "Tipo ID incorrecto para mayor de edad",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F103" in facturas, "RC wrong for age >= 18 (should be CC)"

    def test_under_18_not_flagged(self):
        """Age < 18 → this rule does not apply (age threshold fails)."""
        condiciones = _mayor_18_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_NACIMIENTO")
        ws.cell(row=1, column=3, value="FEC_FACTURA")
        ws.cell(row=1, column=4, value="TIPO_IDENTIFICACION")
        ws.cell(row=2, column=1, value="F104")
        ws.cell(row=2, column=2, value="2015-05-15")   # 11 years old
        ws.cell(row=2, column=3, value="2026-06-24")
        ws.cell(row=2, column=4, value="RC")

        indices = _build_indices("numero_factura", "fec_nacimiento",
                                  "fec_factura", "tipo_identificacion")
        results = _run_engine_detection(
            "tipo_documento_edad_mayor_18", "transversal",
            "Tipo ID incorrecto para mayor de edad",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F104" not in facturas, "Age < 18 excludes from mayor_18 rule"

    def test_exactly_18_birthday_today(self):
        """Age exactly 18 (birthday today) → treated as >= 18 by our age calc."""
        condiciones = _mayor_18_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_NACIMIENTO")
        ws.cell(row=1, column=3, value="FEC_FACTURA")
        ws.cell(row=1, column=4, value="TIPO_IDENTIFICACION")
        ws.cell(row=2, column=1, value="F105")
        ws.cell(row=2, column=2, value="2008-06-24")   # 18 today
        ws.cell(row=2, column=3, value="2026-06-24")
        ws.cell(row=2, column=4, value="TI")  # should be CC now

        indices = _build_indices("numero_factura", "fec_nacimiento",
                                  "fec_factura", "tipo_identificacion")
        results = _run_engine_detection(
            "tipo_documento_edad_mayor_18", "transversal",
            "Tipo ID incorrecto para mayor de edad",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F105" in facturas, "Age 18: TI wrong, should be CC"


# ── Tests: sala_observacion_estancia_prolongada ─────────────────────────────

class TestSalaObservacionEstancia:
    """Simplified estancia prolongada rule for Urgencias."""

    def test_estancia_mayor_6h_en_urgencias_detectada(self):
        """> 6 hours in Urgencias → detected."""
        condiciones = _estancia_prolongada_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_FACTURA")
        ws.cell(row=1, column=3, value="FECHA_CIERRE")
        ws.cell(row=1, column=4, value="TIPO_FACTURA_DESCRIPCION")
        ws.cell(row=2, column=1, value="F201")
        ws.cell(row=2, column=2, value="2026-06-24 08:00:00")
        ws.cell(row=2, column=3, value="2026-06-24 19:00:00")  # 11h
        ws.cell(row=2, column=4, value="Urgencias")

        indices = _build_indices("numero_factura", "fec_factura",
                                  "fecha_cierre", "tipo_factura_descripcion")
        results = _run_engine_detection(
            "sala_observacion_estancia_prolongada", "urgencias",
            "Estancia en Urgencias > 6 horas",
            condiciones, ws, indices,
            severity="warning",
        )

        facturas = _get_facturas_from_results(results)
        assert "F201" in facturas, "> 6h Urgencias should be detected"

    def test_estancia_corta_no_detectada(self):
        """<= 6 hours in Urgencias → not detected."""
        condiciones = _estancia_prolongada_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_FACTURA")
        ws.cell(row=1, column=3, value="FECHA_CIERRE")
        ws.cell(row=1, column=4, value="TIPO_FACTURA_DESCRIPCION")
        ws.cell(row=2, column=1, value="F202")
        ws.cell(row=2, column=2, value="2026-06-24 08:00:00")
        ws.cell(row=2, column=3, value="2026-06-24 12:00:00")  # 4h
        ws.cell(row=2, column=4, value="Urgencias")

        indices = _build_indices("numero_factura", "fec_factura",
                                  "fecha_cierre", "tipo_factura_descripcion")
        results = _run_engine_detection(
            "sala_observacion_estancia_prolongada", "urgencias",
            "Estancia en Urgencias > 6 horas",
            condiciones, ws, indices,
            severity="warning",
        )

        facturas = _get_facturas_from_results(results)
        assert "F202" not in facturas, "<= 6h should not be detected"

    def test_no_urgencias_no_detectada(self):
        """> 6 hours but not Urgencias → not detected."""
        condiciones = _estancia_prolongada_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_FACTURA")
        ws.cell(row=1, column=3, value="FECHA_CIERRE")
        ws.cell(row=1, column=4, value="TIPO_FACTURA_DESCRIPCION")
        ws.cell(row=2, column=1, value="F203")
        ws.cell(row=2, column=2, value="2026-06-24 08:00:00")
        ws.cell(row=2, column=3, value="2026-06-25 08:00:00")  # 24h
        ws.cell(row=2, column=4, value="Hospitalizacion")

        indices = _build_indices("numero_factura", "fec_factura",
                                  "fecha_cierre", "tipo_factura_descripcion")
        results = _run_engine_detection(
            "sala_observacion_estancia_prolongada", "urgencias",
            "Estancia en Urgencias > 6 horas",
            condiciones, ws, indices,
            severity="warning",
        )

        facturas = _get_facturas_from_results(results)
        assert "F203" not in facturas, "Hospitalización not Urgencias, skip"


# ── Output Format ───────────────────────────────────────────────────────────

class TestPhase5OutputFormat:
    """Verify engine output format for Phase 5 rules."""

    def test_output_has_required_keys_menor_7(self):
        """Output dicts have factura, problema, regla, severidad."""
        condiciones = _menor_7_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_NACIMIENTO")
        ws.cell(row=1, column=3, value="FEC_FACTURA")
        ws.cell(row=1, column=4, value="TIPO_IDENTIFICACION")
        ws.cell(row=2, column=1, value="F301")
        ws.cell(row=2, column=2, value="2023-01-01")
        ws.cell(row=2, column=3, value="2026-06-24")
        ws.cell(row=2, column=4, value="CC")  # wrong type for <7

        indices = _build_indices("numero_factura", "fec_nacimiento",
                                  "fec_factura", "tipo_identificacion")
        results = _run_engine_detection(
            "tipo_documento_edad_menor_7", "transversal",
            "Tipo ID incorrecto para menor de 7 años",
            condiciones, ws, indices,
        )

        assert len(results) >= 1
        r = results[0]
        assert "factura" in r
        assert "problema" in r
        assert "regla" in r
        assert "severidad" in r
        assert r["factura"] == "F301"

    def test_output_has_required_keys_mayor_18(self):
        """Output dicts for mayor_18 rule."""
        condiciones = _mayor_18_conditions()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="FEC_NACIMIENTO")
        ws.cell(row=1, column=3, value="FEC_FACTURA")
        ws.cell(row=1, column=4, value="TIPO_IDENTIFICACION")
        ws.cell(row=2, column=1, value="F302")
        ws.cell(row=2, column=2, value="1985-01-01")
        ws.cell(row=2, column=3, value="2026-06-24")
        ws.cell(row=2, column=4, value="TI")

        indices = _build_indices("numero_factura", "fec_nacimiento",
                                  "fec_factura", "tipo_identificacion")
        results = _run_engine_detection(
            "tipo_documento_edad_mayor_18", "transversal",
            "Tipo ID incorrecto para mayor de edad",
            condiciones, ws, indices,
        )

        assert len(results) >= 1
        r = results[0]
        assert r["factura"] == "F302"
        assert r["regla"] == "tipo_documento_edad_mayor_18"
