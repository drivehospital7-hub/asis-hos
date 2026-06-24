"""Unit tests for GroupEvaluator — pre-scan, group-by, aggregate evaluation."""

from __future__ import annotations

import pytest
from openpyxl import Workbook


class TestBuildGroups:
    """Tests for GroupEvaluator.build_groups — pre-scan rows into factura groups."""

    def test_build_groups_multiple_facturas(self):
        """Multiple facturas produce separate groups with correct row counts."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="B")
        ws.cell(row=4, column=1, value="F002")
        ws.cell(row=4, column=2, value="C")

        indices = {"numero_factura": 0, "tipo_procedimiento": 1}
        groups = GroupEvaluator.build_groups(ws, indices)

        assert len(groups) == 2
        assert "F001" in groups
        assert "F002" in groups
        assert len(groups["F001"]) == 2  # rows 2,3
        assert len(groups["F002"]) == 1  # row 4

    def test_build_groups_single_factura(self):
        """All rows with same factura → single group with all rows."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=4, column=1, value="F001")

        indices = {"numero_factura": 0}
        groups = GroupEvaluator.build_groups(ws, indices)

        assert len(groups) == 1
        assert len(groups["F001"]) == 3

    def test_build_groups_empty_sheet(self):
        """Sheet with only header row → empty groups dict."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")

        indices = {"numero_factura": 0}
        groups = GroupEvaluator.build_groups(ws, indices)

        assert isinstance(groups, dict)
        assert len(groups) == 0

    def test_build_groups_missing_column(self):
        """Missing group_by column → empty groups dict."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="OTHER_COLUMN")
        ws.cell(row=2, column=1, value="F001")

        indices = {"numero_factura": None, "other_column": 0}
        groups = GroupEvaluator.build_groups(ws, indices)

        assert isinstance(groups, dict)
        assert len(groups) == 0

    def test_build_groups_skips_blank_factura(self):
        """Rows with empty factura value are skipped."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=3, column=1, value="")   # blank → skip
        ws.cell(row=4, column=1, value=None)  # None → skip
        ws.cell(row=5, column=1, value="F002")

        indices = {"numero_factura": 0}
        groups = GroupEvaluator.build_groups(ws, indices)

        assert len(groups) == 2
        assert "F001" in groups
        assert "F002" in groups
        assert "" not in groups

    def test_build_groups_whitespace_factura_skipped(self):
        """Rows with whitespace-only factura are skipped."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=2, column=1, value="   ")   # whitespace → skip
        ws.cell(row=3, column=1, value="F001")

        indices = {"numero_factura": 0}
        groups = GroupEvaluator.build_groups(ws, indices)

        assert len(groups) == 1
        assert "F001" in groups


class TestBuildGroupData:
    """Tests for GroupEvaluator._build_group_data — compute aggregates."""

    def test_distinct_count_multiple_values(self):
        """distinct_count of a field with multiple values → correct count."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=3, column=1, value="B")
        ws.cell(row=4, column=1, value="A")  # duplicate

        indices = {"tipo_procedimiento": 0}
        agg_configs = [{"function": "distinct_count", "field": "tipo_procedimiento",
                         "target": "distinct_count_tipo_procedimiento"}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3, 4], ws, indices, agg_configs
        )

        assert group_data["distinct_count_tipo_procedimiento"] == 2  # A, B

    def test_distinct_count_single_value(self):
        """distinct_count of a field with one value → 1."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=3, column=1, value="A")

        indices = {"tipo_procedimiento": 0}
        agg_configs = [{"function": "distinct_count", "field": "tipo_procedimiento",
                         "target": "dc_tipo"}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3], ws, indices, agg_configs
        )

        assert group_data["dc_tipo"] == 1

    def test_distinct_count_with_nulls(self):
        """distinct_count skips None values correctly."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=3, column=1, value=None)
        ws.cell(row=4, column=1, value="B")

        indices = {"tipo_procedimiento": 0}
        agg_configs = [{"function": "distinct_count", "field": "tipo_procedimiento",
                         "target": "dc_tipo"}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3, 4], ws, indices, agg_configs
        )

        assert group_data["dc_tipo"] == 2  # A, B (None skipped)

    def test_group_size(self):
        """group_size function returns row count in group."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")

        indices = {"numero_factura": 0}
        agg_configs = [{"function": "group_size", "target": "size"}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3, 4, 5], ws, indices, agg_configs
        )

        assert group_data["size"] == 4

    def test_sum_numeric_values(self):
        """sum function correctly totals numeric values."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="CANTIDAD")
        ws.cell(row=2, column=1, value=10)
        ws.cell(row=3, column=1, value=5)
        ws.cell(row=4, column=1, value=3)

        indices = {"cantidad": 0}
        agg_configs = [{"function": "sum", "field": "cantidad",
                         "target": "sum_cantidad"}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3, 4], ws, indices, agg_configs
        )

        assert group_data["sum_cantidad"] == 18

    def test_sum_with_nulls_and_non_numeric(self):
        """sum handles None and non-numeric values gracefully."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="CANTIDAD")
        ws.cell(row=2, column=1, value=5)
        ws.cell(row=3, column=1, value=None)
        ws.cell(row=4, column=1, value="N/A")

        indices = {"cantidad": 0}
        agg_configs = [{"function": "sum", "field": "cantidad",
                         "target": "sum_cantidad"}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3, 4], ws, indices, agg_configs
        )

        assert group_data["sum_cantidad"] == 5  # only the valid 5

    def test_multiple_aggregations(self):
        """Multiple aggregation configs computed in one call."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="TIPO")
        ws.cell(row=1, column=2, value="CANTIDAD")
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=2, column=2, value=10)
        ws.cell(row=3, column=1, value="B")
        ws.cell(row=3, column=2, value=20)

        indices = {"tipo_procedimiento": 0, "cantidad": 1}
        agg_configs = [
            {"function": "distinct_count", "field": "tipo_procedimiento",
             "target": "dc_tipo"},
            {"function": "sum", "field": "cantidad", "target": "total"},
            {"function": "group_size", "target": "size"},
        ]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3], ws, indices, agg_configs
        )

        assert group_data["dc_tipo"] == 2
        assert group_data["total"] == 30
        assert group_data["size"] == 2

    def test_default_target_name(self):
        """When target not specified, uses {function}_{field}."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="TIPO")
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=3, column=1, value="B")

        indices = {"tipo_procedimiento": 0}
        agg_configs = [{"function": "distinct_count", "field": "tipo_procedimiento"}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3], ws, indices, agg_configs
        )

        # default: distinct_count_tipo_procedimiento
        assert group_data["distinct_count_tipo_procedimiento"] == 2


class TestGroupEvaluatorEvaluate:
    """Tests for GroupEvaluator.evaluate — full group-by evaluation flow."""

    def _build_condition_evaluator(self):
        """Create a real ConditionEvaluator for group-by tests."""
        from app.services.engine.condition_evaluator import ConditionEvaluator
        return ConditionEvaluator()

    def _build_evidence_collector(self):
        """Create a real EvidenceCollector."""
        from app.services.engine.evidence_collector import EvidenceCollector
        return EvidenceCollector()

    def test_evaluate_detects_doble_tipo(self):
        """Group with >1 distinct tipos → MATCH detection."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="B")

        indices = {"numero_factura": 0, "tipo_procedimiento": 1}
        groups = GroupEvaluator.build_groups(ws, indices)
        agg_configs = [{"function": "distinct_count", "field": "tipo_procedimiento",
                         "target": "distinct_count_tipo_procedimiento"}]

        # Condition: gt(invoice.distinct_count_tipo_procedimiento, 1)
        condition_tree = {
            "id": 1, "padre_id": None, "tipo": "atomic", "operador": "gt",
            "fuente_datos": "invoice.distinct_count_tipo_procedimiento",
            "valor_esperado": "1", "orden": 0,
        }
        evaluator = self._build_condition_evaluator()
        tree = evaluator.build_tree([condition_tree])

        collector = self._build_evidence_collector()
        rule_info = {
            "id": 1, "version": 1, "dominio": "transversal",
            "nombre": "doble_tipo_procedimiento",
            "descripcion": "Factura con doble tipo de procedimiento",
            "severidad": "error",
        }

        results = GroupEvaluator.evaluate(
            groups, ws, indices, agg_configs,
            tree, evaluator, rule_info, collector,
        )

        assert len(results) == 1
        assert results[0]["factura"] == "F001"
        assert "doble tipo" in results[0]["problema"].lower()

    def test_evaluate_no_match_single_tipo(self):
        """Group with single tipo → no detection."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="A")  # same tipo

        indices = {"numero_factura": 0, "tipo_procedimiento": 1}
        groups = GroupEvaluator.build_groups(ws, indices)
        agg_configs = [{"function": "distinct_count", "field": "tipo_procedimiento",
                         "target": "distinct_count_tipo_procedimiento"}]

        condition_tree = {
            "id": 1, "padre_id": None, "tipo": "atomic", "operador": "gt",
            "fuente_datos": "invoice.distinct_count_tipo_procedimiento",
            "valor_esperado": "1", "orden": 0,
        }
        evaluator = self._build_condition_evaluator()
        tree = evaluator.build_tree([condition_tree])

        collector = self._build_evidence_collector()
        rule_info = {
            "id": 1, "version": 1, "dominio": "transversal",
            "nombre": "doble_tipo", "descripcion": "Doble tipo",
            "severidad": "error",
        }

        results = GroupEvaluator.evaluate(
            groups, ws, indices, agg_configs,
            tree, evaluator, rule_info, collector,
        )

        assert len(results) == 0

    def test_evaluate_multiple_groups_mixed(self):
        """Two groups: one match, one no-match → only match returned."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        # F001: 2 tipos → MATCH
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="B")
        # F002: 1 tipo → NO_MATCH
        ws.cell(row=4, column=1, value="F002")
        ws.cell(row=4, column=2, value="C")

        indices = {"numero_factura": 0, "tipo_procedimiento": 1}
        groups = GroupEvaluator.build_groups(ws, indices)
        agg_configs = [{"function": "distinct_count", "field": "tipo_procedimiento",
                         "target": "distinct_count_tipo_procedimiento"}]

        condition_tree = {
            "id": 1, "padre_id": None, "tipo": "atomic", "operador": "gt",
            "fuente_datos": "invoice.distinct_count_tipo_procedimiento",
            "valor_esperado": "1", "orden": 0,
        }
        evaluator = self._build_condition_evaluator()
        tree = evaluator.build_tree([condition_tree])

        collector = self._build_evidence_collector()
        rule_info = {
            "id": 1, "version": 1, "dominio": "transversal",
            "nombre": "doble_tipo", "descripcion": "Doble tipo",
            "severidad": "error",
        }

        results = GroupEvaluator.evaluate(
            groups, ws, indices, agg_configs,
            tree, evaluator, rule_info, collector,
        )

        assert len(results) == 1
        assert results[0]["factura"] == "F001"

    def test_evaluate_sum_quantity_detection(self):
        """Sum of quantity > 1 → detection."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CANTIDAD")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value=1)
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value=1)  # total 2 > 1

        indices = {"numero_factura": 0, "cantidad": 1}
        groups = GroupEvaluator.build_groups(ws, indices)
        agg_configs = [{"function": "sum", "field": "cantidad",
                         "target": "sum_cantidad"}]

        condition_tree = {
            "id": 1, "padre_id": None, "tipo": "atomic", "operador": "gt",
            "fuente_datos": "invoice.sum_cantidad",
            "valor_esperado": "1", "orden": 0,
        }
        evaluator = self._build_condition_evaluator()
        tree = evaluator.build_tree([condition_tree])

        collector = self._build_evidence_collector()
        rule_info = {
            "id": 2, "version": 1, "dominio": "urgencias",
            "nombre": "revision_cantidad", "descripcion": "Revisar cantidad",
            "severidad": "warning",
        }

        results = GroupEvaluator.evaluate(
            groups, ws, indices, agg_configs,
            tree, evaluator, rule_info, collector,
        )

        assert len(results) == 1
        assert results[0]["factura"] == "F001"

    def test_evaluate_empty_groups_returns_empty(self):
        """Empty groups → no results, no crash."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")

        indices = {"numero_factura": 0}
        groups = GroupEvaluator.build_groups(ws, indices)  # empty
        agg_configs = [{"function": "group_size", "target": "size"}]

        condition_tree = {
            "id": 1, "padre_id": None, "tipo": "atomic", "operador": "gt",
            "fuente_datos": "invoice.size", "valor_esperado": "0", "orden": 0,
        }
        evaluator = self._build_condition_evaluator()
        tree = evaluator.build_tree([condition_tree])

        collector = self._build_evidence_collector()
        rule_info = {
            "id": 1, "version": 1, "dominio": "transversal",
            "nombre": "test", "descripcion": "Test",
            "severidad": "error",
        }

        results = GroupEvaluator.evaluate(
            groups, ws, indices, agg_configs,
            tree, evaluator, rule_info, collector,
        )

        assert len(results) == 0

    def test_evaluate_records_evidence(self):
        """GroupEvaluator records evidence via EvidenceCollector."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="B")

        indices = {"numero_factura": 0, "tipo_procedimiento": 1}
        groups = GroupEvaluator.build_groups(ws, indices)
        agg_configs = [{"function": "distinct_count", "field": "tipo_procedimiento",
                         "target": "distinct_count_tipo_procedimiento"}]

        condition_tree = {
            "id": 1, "padre_id": None, "tipo": "atomic", "operador": "gt",
            "fuente_datos": "invoice.distinct_count_tipo_procedimiento",
            "valor_esperado": "1", "orden": 0,
        }
        evaluator = self._build_condition_evaluator()
        tree = evaluator.build_tree([condition_tree])

        collector = self._build_evidence_collector()
        rule_info = {
            "id": 1, "version": 1, "dominio": "transversal",
            "nombre": "doble_tipo", "descripcion": "Doble tipo",
            "severidad": "error",
        }

        results = GroupEvaluator.evaluate(
            groups, ws, indices, agg_configs,
            tree, evaluator, rule_info, collector,
        )

        # Verify evidence was recorded (buffer is non-empty)
        assert len(collector._buffer) >= 1
        # Verify at least one evidence record for F001
        facturas_in_evidence = {e.factura for e in collector._buffer}
        assert "F001" in facturas_in_evidence


class TestCollectSetAggregation:
    """Integration tests for collect_set aggregation.

    collect_set collects unique non-None values from a field across group rows
    and returns them as a sorted list.
    """

    def test_collect_set_multiple_values(self):
        """collect_set across multiple rows with distinct values → sorted list."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="CODIGO")
        ws.cell(row=2, column=1, value="5DSB01")
        ws.cell(row=3, column=1, value="890701")
        ws.cell(row=4, column=1, value="5DSB01")  # duplicate

        indices = {"codigo": 0}
        agg_configs = [{"function": "collect_set", "field": "codigo",
                         "target": "collect_set_codigo"}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3, 4], ws, indices, agg_configs
        )

        result = group_data["collect_set_codigo"]
        assert isinstance(result, list)
        assert "5DSB01" in result
        assert "890701" in result
        assert len(result) == 2

    def test_collect_set_with_nulls(self):
        """collect_set skips None values."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="CODIGO")
        ws.cell(row=2, column=1, value="5DSB01")
        ws.cell(row=3, column=1, value=None)
        ws.cell(row=4, column=1, value="890701")

        indices = {"codigo": 0}
        agg_configs = [{"function": "collect_set", "field": "codigo",
                         "target": "collect_set_codigo"}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3, 4], ws, indices, agg_configs
        )

        result = group_data["collect_set_codigo"]
        assert isinstance(result, list)
        assert len(result) == 2
        assert "5DSB01" in result
        assert "890701" in result

    def test_collect_set_empty_group(self):
        """collect_set with no data rows → empty list."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="CODIGO")

        indices = {"codigo": 0}
        agg_configs = [{"function": "collect_set", "field": "codigo",
                         "target": "collect_set_codigo"}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [], ws, indices, agg_configs
        )

        assert group_data["collect_set_codigo"] == []


class TestCollectValueCountsAggregation:
    """Integration tests for collect_value_counts aggregation.

    collect_value_counts counts (field1, field2) pairs across group rows
    and returns a list of dicts with codigo, cantidad, count keys.
    """

    def test_collect_value_counts_multiple_pairs(self):
        """Multiple (codigo, cantidad) pairs → list of dicts with counts."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="CODIGO")
        ws.cell(row=1, column=2, value="CANTIDAD")
        # Row 1: (A, 1) appears twice
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=2, column=2, value=1)
        ws.cell(row=3, column=1, value="A")
        ws.cell(row=3, column=2, value=1)
        # Row 2: (B, 1) appears once
        ws.cell(row=4, column=1, value="B")
        ws.cell(row=4, column=2, value=1)

        indices = {"codigo": 0, "cantidad": 1}
        agg_configs = [{"function": "collect_value_counts", "fields": ["codigo", "cantidad"]}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3, 4], ws, indices, agg_configs
        )

        result = group_data["collect_value_counts"]
        assert isinstance(result, list)
        assert len(result) == 2
        counts = {item["codigo"]: item["count"] for item in result}
        assert counts["A"] == 2
        assert counts["B"] == 1

    def test_collect_value_counts_list_of_dicts_format(self):
        """Each item has codigo, cantidad, count keys."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="CODIGO")
        ws.cell(row=1, column=2, value="CANTIDAD")
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=2, column=2, value=1)
        ws.cell(row=3, column=1, value="B")
        ws.cell(row=3, column=2, value=2)

        indices = {"codigo": 0, "cantidad": 1}
        agg_configs = [{"function": "collect_value_counts", "fields": ["codigo", "cantidad"]}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3], ws, indices, agg_configs
        )

        for item in group_data["collect_value_counts"]:
            assert "codigo" in item
            assert "cantidad" in item
            assert "count" in item

    def test_collect_value_counts_empty_group(self):
        """collect_value_counts with no data rows → empty list."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="CODIGO")
        ws.cell(row=1, column=2, value="CANTIDAD")

        indices = {"codigo": 0, "cantidad": 1}
        agg_configs = [{"function": "collect_value_counts", "fields": ["codigo", "cantidad"]}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [], ws, indices, agg_configs
        )

        assert group_data["collect_value_counts"] == []

    def test_collect_value_counts_with_nulls(self):
        """collect_value_counts handles None values by treating as empty string."""
        from app.services.engine.group_evaluator import GroupEvaluator

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="CODIGO")
        ws.cell(row=1, column=2, value="CANTIDAD")
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=2, column=2, value=None)
        ws.cell(row=3, column=1, value="A")
        ws.cell(row=3, column=2, value=None)

        indices = {"codigo": 0, "cantidad": 1}
        agg_configs = [{"function": "collect_value_counts", "fields": ["codigo", "cantidad"]}]
        group_data = GroupEvaluator._build_group_data(
            "F001", [2, 3], ws, indices, agg_configs
        )

        result = group_data["collect_value_counts"]
        assert len(result) == 1
        assert result[0]["codigo"] == "A"
        assert result[0]["count"] == 2
