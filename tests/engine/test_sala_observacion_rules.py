"""Tests for Sala Observación group rules — F4.

Group rules that complement the existing SalaObservacionEvaluator:
4.1: sala_obs_obligatorios — activadores sin OBLIG_SALA
4.2: sala_obs_ess_129b02 — entidad ESS con 129B02 (prohibido)
4.3: sala_obs_soat_completo — SOAT sala sin SOAT_OBLIG
4.4: sala_obs_soat_prohibido — SOAT con 39133
4.5a: sala_obs_890601h — 890601H prohibido
4.5b: sala_obs_05dsb01_no_ess — 05DSB01 en no-ESS
"""

from __future__ import annotations

from datetime import datetime

import pytest
from openpyxl import Workbook

SALA_ACTIV = frozenset({"5DSB01", "05DSB01", "129B02", "38114", "38915"})
OBLIG_SALA = frozenset({"890701", "890601"})
ENTIDADES_ESS = frozenset({"ESS118", "ESSC18"})
SOAT_OBLIG_SALA = frozenset({"39145", "39131"})
SOAT_SALA = frozenset({"38114", "38915"})


def _mk_ws():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="NUMERO_FACTURA")
    ws.cell(row=1, column=2, value="CODIGO")
    ws.cell(row=1, column=3, value="TIPO_FACTURA_DESCRIPCION")
    ws.cell(row=1, column=4, value="CODIGO_ENTIDAD_COBRAR")
    ws.cell(row=1, column=5, value="TARIFARIO")
    indices = {
        "numero_factura": 0, "codigo": 1,
        "tipo_factura_descripcion": 2, "codigo_entidad_cobrar": 3,
        "tarifario": 4,
    }
    return ws, indices


def _ce():
    from app.services.engine.condition_evaluator import ConditionEvaluator
    return ConditionEvaluator()


def _collector():
    from app.services.engine.evidence_collector import EvidenceCollector
    return EvidenceCollector()


def _rule_info(name="test_sala", desc="Sala test"):
    return {
        "id": 500, "version": 1, "dominio": "urgencias",
        "nombre": name, "descripcion": desc, "severidad": "error",
    }


def _agg():
    return [
        {"function": "collect_set", "field": "codigo", "target": "collect_set_codigo"},
        {"function": "collect_set", "field": "codigo_entidad_cobrar", "target": "collect_set_entidad"},
    ]


def _flat_tree(conditions):
    return _ce().build_tree(conditions)


def _ev():
    from app.services.engine.group_evaluator import GroupEvaluator
    return GroupEvaluator


# ══════════════════════════════════════════════════════════════════════════
# 4.1: sala_obs_obligatorios
# set_intersects(codigos, SALA_ACTIV) AND NOT(set_contains_all(codigos, OBLIG_SALA))
# ══════════════════════════════════════════════════════════════════════════

class TestSalaObsObligatorios:
    """Missing obligatory codes when sala activators present."""

    def test_sala_sin_obligatorios(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="5DSB01")
        ws.cell(row=2, column=3, value="Urgencias")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": list(SALA_ACTIV), "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_contains_all",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": list(OBLIG_SALA), "orden": 0},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_obligatorios"), _collector())
        assert len(results) == 1
        assert results[0]["factura"] == "F001"

    def test_sala_con_obligatorios(self):
        ws, indices = _mk_ws()
        for r in range(2, 5):
            ws.cell(row=r, column=1, value="F001")
            ws.cell(row=r, column=3, value="Urgencias")
        ws.cell(row=2, column=2, value="5DSB01")
        ws.cell(row=3, column=2, value="890701")
        ws.cell(row=4, column=2, value="890601")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": list(SALA_ACTIV), "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_contains_all",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": list(OBLIG_SALA), "orden": 0},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_obligatorios"), _collector())
        assert len(results) == 0

    def test_sin_sala_activador(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="999999")
        ws.cell(row=2, column=3, value="Urgencias")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": list(SALA_ACTIV), "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_contains_all",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": list(OBLIG_SALA), "orden": 0},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_obligatorios"), _collector())
        assert len(results) == 0


# ══════════════════════════════════════════════════════════════════════════
# 4.2: sala_obs_ess_129b02
# set_intersects(entidad, ESS) AND contains(codigos, 129B02)
# ══════════════════════════════════════════════════════════════════════════

class TestSalaObsESS129B02:
    """ESS entidad cannot have 129B02."""

    def test_ess_con_129b02(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="129B02")
        ws.cell(row=2, column=3, value="Urgencias")
        ws.cell(row=2, column=4, value="ESS118")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_entidad", "valor_esperado": list(ENTIDADES_ESS), "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "atomic", "operador": "contains",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": "129B02", "orden": 1},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_ess_129b02"), _collector())
        assert len(results) == 1

    def test_no_ess_con_129b02(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="129B02")
        ws.cell(row=2, column=3, value="Urgencias")
        ws.cell(row=2, column=4, value="EPSS41")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_entidad", "valor_esperado": list(ENTIDADES_ESS), "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "atomic", "operador": "contains",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": "129B02", "orden": 1},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_ess_129b02"), _collector())
        assert len(results) == 0


# ══════════════════════════════════════════════════════════════════════════
# 4.4: sala_obs_soat_prohibido
# contains(codigos, 39133)
# ══════════════════════════════════════════════════════════════════════════

class TestSalaObsSOATProhibido:
    """SOAT Urgencias cannot have 39133."""

    def test_con_39133(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="39133")
        ws.cell(row=2, column=3, value="Urgencias")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "atomic", "operador": "contains",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": "39133", "orden": 0},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_soat_prohibido"), _collector())
        assert len(results) == 1

    def test_sin_39133(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="38114")
        ws.cell(row=2, column=3, value="Urgencias")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "atomic", "operador": "contains",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": "39133", "orden": 0},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_soat_prohibido"), _collector())
        assert len(results) == 0


# ══════════════════════════════════════════════════════════════════════════
# 4.5a: sala_obs_890601h — contains(codigos, 890601H)
# ══════════════════════════════════════════════════════════════════════════

class TestSalaObs890601H:
    """890601H prohibido en Urgencias."""

    def test_890601h_presente(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="890601H")
        ws.cell(row=2, column=3, value="Urgencias")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [{"id": 1, "padre_id": None, "tipo": "atomic", "operador": "contains",
                 "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": "890601H", "orden": 0}]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_890601h"), _collector())
        assert len(results) == 1

    def test_sin_890601h(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="129B02")
        ws.cell(row=2, column=3, value="Urgencias")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [{"id": 1, "padre_id": None, "tipo": "atomic", "operador": "contains",
                 "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": "890601H", "orden": 0}]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_890601h"), _collector())
        assert len(results) == 0


# ══════════════════════════════════════════════════════════════════════════
# 4.5b: sala_obs_05dsb01_no_ess
# contains(codigos, 05DSB01) AND NOT(set_intersects(entidad, ESS))
# ══════════════════════════════════════════════════════════════════════════

class TestSalaObs05DSB01NoESS:
    """05DSB01 prohibido en entidades no-ESS."""

    def test_05dsb01_en_no_ess(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="05DSB01")
        ws.cell(row=2, column=3, value="Urgencias")
        ws.cell(row=2, column=4, value="EPSS41")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "contains",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": "05DSB01", "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_entidad", "valor_esperado": list(ENTIDADES_ESS), "orden": 0},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_05dsb01_no_ess"), _collector())
        assert len(results) == 1

    def test_05dsb01_en_ess(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="05DSB01")
        ws.cell(row=2, column=3, value="Urgencias")
        ws.cell(row=2, column=4, value="ESS118")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "contains",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": "05DSB01", "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_entidad", "valor_esperado": list(ENTIDADES_ESS), "orden": 0},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_05dsb01_no_ess"), _collector())
        assert len(results) == 0

    def test_sin_05dsb01(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="129B02")
        ws.cell(row=2, column=3, value="Urgencias")
        ws.cell(row=2, column=4, value="EPSS41")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "contains",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": "05DSB01", "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_entidad", "valor_esperado": list(ENTIDADES_ESS), "orden": 0},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_05dsb01_no_ess"), _collector())
        assert len(results) == 0


# ══════════════════════════════════════════════════════════════════════════
# 4.3: sala_obs_soat_completo
# set_intersects(codigos, SOAT_SALA) AND NOT(set_contains_all(codigos, SOAT_OBLIG))
# ══════════════════════════════════════════════════════════════════════════

class TestSalaObsSOATCompleto:
    """SOAT sala needs SOAT_OBLIG codes."""

    def test_soat_sala_sin_oblig(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="38114")
        ws.cell(row=2, column=3, value="Urgencias")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": list(SOAT_SALA), "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_contains_all",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": list(SOAT_OBLIG_SALA), "orden": 0},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_soat_completo"), _collector())
        assert len(results) == 1

    def test_soat_sala_con_oblig(self):
        ws, indices = _mk_ws()
        for r in range(2, 5):
            ws.cell(row=r, column=1, value="F001")
            ws.cell(row=r, column=3, value="Urgencias")
        ws.cell(row=2, column=2, value="38114")
        ws.cell(row=3, column=2, value="39145")
        ws.cell(row=4, column=2, value="39131")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": list(SOAT_SALA), "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_contains_all",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": list(SOAT_OBLIG_SALA), "orden": 0},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_soat_completo"), _collector())
        assert len(results) == 0

    def test_sin_soat_sala(self):
        ws, indices = _mk_ws()
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="129B02")
        ws.cell(row=2, column=3, value="Urgencias")

        groups = _ev().build_groups(ws, indices, filter_field="tipo_factura_descripcion", filter_value="Urgencias")
        flat = [
            {"id": 1, "padre_id": None, "tipo": "composite", "operador": "AND", "orden": 0},
            {"id": 2, "padre_id": 1, "tipo": "atomic", "operador": "set_intersects",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": list(SOAT_SALA), "orden": 0},
            {"id": 3, "padre_id": 1, "tipo": "composite", "operador": "NOT", "orden": 1},
            {"id": 4, "padre_id": 3, "tipo": "atomic", "operador": "set_contains_all",
             "fuente_datos": "invoice.collect_set_codigo", "valor_esperado": list(SOAT_OBLIG_SALA), "orden": 0},
        ]
        results = _ev().evaluate(groups, ws, indices, _agg(), _flat_tree(flat), _ce(), _rule_info("sala_obs_soat_completo"), _collector())
        assert len(results) == 0
