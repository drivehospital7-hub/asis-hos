"""Snapshot tests Phase 4: engine vs legacy for 3 centro_costo rules.

Verifies that the DB-backed engine with NOT+in(centro_costo, valid_centers)
condition tree detects the same rows as the legacy Python centro_costo detectors.
"""

from __future__ import annotations

import pytest
from unittest.mock import MagicMock
from openpyxl import Workbook


# ── Helpers ─────────────────────────────────────────────────────────────────

def _mock_session_with_rule(rule_name, dominio, descripcion, condiciones_dicts):
    """Create a mock session that returns a Regla + Condicion tree."""
    from app.models import Regla

    session = MagicMock()
    mock_query = MagicMock()
    mock_query.filter.return_value = mock_query
    mock_query.order_by.return_value = mock_query

    regla = Regla(
        id=1, nombre=rule_name, dominio=dominio,
        estado="active", version=1, prioridad=10, severidad="error",
        descripcion=descripcion,
    )
    mock_query.first.return_value = regla

    cond_mocks = []
    for cd in condiciones_dicts:
        m = MagicMock()
        m.id = cd["id"]
        m.regla_id = cd.get("regla_id", 1)
        m.padre_id = cd["padre_id"]
        m.tipo = cd["tipo"]
        m.operador = cd.get("operador")
        m.fuente_datos = cd.get("fuente_datos")
        m.valor_esperado = cd.get("valor_esperado")
        m.orden = cd.get("orden", 0)
        cond_mocks.append(m)

    mock_query.all.return_value = cond_mocks
    session.query.return_value = mock_query
    return session


def _build_indices(*col_names):
    """Build indices dict: col_name → 0-based index."""
    return {name: i for i, name in enumerate(col_names)}


def _run_engine_detection(rule_name, dominio, descripcion, condiciones, ws, indices):
    """Run engine detection against a worksheet with mocked session."""
    from app.services.engine.rule_based_detector import RuleBasedDetector
    session = _mock_session_with_rule(rule_name, dominio, descripcion, condiciones)
    detector = RuleBasedDetector(rule_name, session)
    return detector.detect(ws, indices)


def _get_facturas_from_results(results):
    """Extract factura strings from detection results."""
    facturas = set()
    for r in results:
        if isinstance(r, dict):
            facturas.add(r.get("factura", ""))
    return facturas


# ── Shared condition tree builders ──────────────────────────────────────────

VALID_ODON_CENTERS = ["ODONTOLOGIA", "SERVICIOS ODONTOLOGIA -EXTRAMURALES"]
VALID_URG_CENTERS = [
    "URGENCIAS", "APOYO TERAPEUTICO-FARMACIA E INSUMOS.",
    "APOYO DIAGNOSTICO-LABORATOR CLINICO",
    "PROCEDIMIENTO DE PROMOCIÓN Y PREVENCIÓN",
    "HOSPITALIZACIÓN - ESTANCIA GENERAL",
    "APOYO DIAGNOSTICO-IMAGENOLOGIA", "TRASLADOS",
    "QUIRÓFANOS Y SALAS DE PARTO- SALA DE PARTO",
]
VALID_EQBAS_CENTERS = ["EQUIPOS BASICOS ODONTOLOGIA"]

def _not_in_conditions(valid_list, source="invoice.centro_costo"):
    """Build NOT+in condition tree dicts. valid_list is a Python list (as DB jsonb returns)."""
    return [
        {
            "id": 1, "padre_id": None,
            "tipo": "composite", "operador": "NOT",
            "fuente_datos": None, "valor_esperado": None, "orden": 0,
        },
        {
            "id": 2, "padre_id": 1,
            "tipo": "atomic", "operador": "in",
            "fuente_datos": source,
            "valor_esperado": valid_list,  # Python list, not json.dumps (DB jsonb returns Python objects)
            "orden": 0,
        },
    ]


# ── Tests: Odontología ──────────────────────────────────────────────────────

class TestCentroCostoOdontologia:
    """Engine with NOT+in(centro_costo, odon_valid_centers) validates odontologia."""

    def test_engine_detects_invalid_centro(self):
        """Invalid centro_costo row detected."""
        condiciones = _not_in_conditions(VALID_ODON_CENTERS)

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CENTRO_COSTO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="CENTRO_INVALIDO")

        indices = _build_indices("numero_factura", "centro_costo")
        results = _run_engine_detection(
            "centro_costo_odontologia_valido", "odontologia",
            "Centro de costo no válido en Odontología",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F001" in facturas, "Engine should detect invalid centro_costo"

    def test_engine_ignores_valid_centro(self):
        """Valid centro_costo row not flagged."""
        condiciones = _not_in_conditions(VALID_ODON_CENTERS)

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CENTRO_COSTO")
        ws.cell(row=2, column=1, value="F002")
        ws.cell(row=2, column=2, value="ODONTOLOGIA")
        ws.cell(row=3, column=1, value="F003")
        ws.cell(row=3, column=2, value="SERVICIOS ODONTOLOGIA -EXTRAMURALES")

        indices = _build_indices("numero_factura", "centro_costo")
        results = _run_engine_detection(
            "centro_costo_odontologia_valido", "odontologia",
            "Centro de costo no válido en Odontología",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F002" not in facturas, "ODONTOLOGIA is valid"
        assert "F003" not in facturas, "EXTRAMURALES is valid"

    def test_engine_handles_empty_centro(self):
        """Empty centro_costo row is treated as invalid (NOT in valid list)."""
        condiciones = _not_in_conditions(VALID_ODON_CENTERS)

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CENTRO_COSTO")
        ws.cell(row=2, column=1, value="F004")
        # centro_costo is empty/None

        indices = _build_indices("numero_factura", "centro_costo")
        results = _run_engine_detection(
            "centro_costo_odontologia_valido", "odontologia",
            "Centro de costo no válido en Odontología",
            condiciones, ws, indices,
        )

        # Empty string / None is NOT in the valid list, so it should be detected
        facturas = _get_facturas_from_results(results)
        assert "F004" in facturas, "Empty centro_costo is not in valid list"


# ── Tests: Urgencias ────────────────────────────────────────────────────────

class TestCentroCostoUrgencias:
    """Engine with NOT+in(centro_costo, urg_valid_centers) validates urgencias."""

    def test_engine_detects_invalid_centro(self):
        """Invalid centro_costo row detected."""
        condiciones = _not_in_conditions(VALID_URG_CENTERS)

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CENTRO_COSTO")
        ws.cell(row=2, column=1, value="F101")
        ws.cell(row=2, column=2, value="CENTRO_FANTASMA")

        indices = _build_indices("numero_factura", "centro_costo")
        results = _run_engine_detection(
            "centro_costo_urgencias_valido", "urgencias",
            "Centro de costo no válido en Urgencias",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F101" in facturas, "Engine should detect invalid centro_costo"

    def test_engine_ignores_valid_centros(self):
        """All 8 valid urgencias centros pass."""
        condiciones = _not_in_conditions(VALID_URG_CENTERS)

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CENTRO_COSTO")
        row = 2
        for centro in VALID_URG_CENTERS:
            ws.cell(row=row, column=1, value=f"F{row:03d}")
            ws.cell(row=row, column=2, value=centro)
            row += 1

        indices = _build_indices("numero_factura", "centro_costo")
        results = _run_engine_detection(
            "centro_costo_urgencias_valido", "urgencias",
            "Centro de costo no válido en Urgencias",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert len(facturas) == 0, f"All {len(VALID_URG_CENTERS)} valid centros should pass, got: {facturas}"

    def test_multiple_rows_mixed(self):
        """Mixed valid/invalid rows — only invalid detected."""
        condiciones = _not_in_conditions(VALID_URG_CENTERS)

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CENTRO_COSTO")
        ws.cell(row=2, column=1, value="F201")
        ws.cell(row=2, column=2, value="URGENCIAS")  # valid
        ws.cell(row=3, column=1, value="F202")
        ws.cell(row=3, column=2, value="INVALIDO_X")  # invalid
        ws.cell(row=4, column=1, value="F203")
        ws.cell(row=4, column=2, value="TRASLADOS")  # valid
        ws.cell(row=5, column=1, value="F204")
        ws.cell(row=5, column=2, value="INVALIDO_Y")  # invalid

        indices = _build_indices("numero_factura", "centro_costo")
        results = _run_engine_detection(
            "centro_costo_urgencias_valido", "urgencias",
            "Centro de costo no válido en Urgencias",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F201" not in facturas, "URGENCIAS is valid"
        assert "F202" in facturas, "INVALIDO_X should be detected"
        assert "F203" not in facturas, "TRASLADOS is valid"
        assert "F204" in facturas, "INVALIDO_Y should be detected"
        assert len(facturas) == 2


# ── Tests: Equipos Básicos ──────────────────────────────────────────────────

class TestCentroCostoEquiposBasicos:
    """Engine with NOT+in(centro_costo, eqbas_valid_centers) validates equipos_basicos."""

    def test_engine_detects_invalid_centro(self):
        """Invalid centro_costo row detected."""
        condiciones = _not_in_conditions(VALID_EQBAS_CENTERS)

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CENTRO_COSTO")
        ws.cell(row=2, column=1, value="F301")
        ws.cell(row=2, column=2, value="URGENCIAS")  # not valid for eqbas

        indices = _build_indices("numero_factura", "centro_costo")
        results = _run_engine_detection(
            "centro_costo_equipos_basicos_valido", "equipos_basicos",
            "Centro de costo no válido en Equipos Básicos",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F301" in facturas, "URGENCIAS is not a valid eqbas centro"

    def test_engine_ignores_valid_centro(self):
        """Only EQUIPOS BASICOS ODONTOLOGIA is valid for eqbas."""
        condiciones = _not_in_conditions(VALID_EQBAS_CENTERS)

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CENTRO_COSTO")
        ws.cell(row=2, column=1, value="F302")
        ws.cell(row=2, column=2, value="EQUIPOS BASICOS ODONTOLOGIA")

        indices = _build_indices("numero_factura", "centro_costo")
        results = _run_engine_detection(
            "centro_costo_equipos_basicos_valido", "equipos_basicos",
            "Centro de costo no válido en Equipos Básicos",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F302" not in facturas, "EQUIPOS BASICOS ODONTOLOGIA is valid"


# ── Edge Cases ──────────────────────────────────────────────────────────────

class TestCentroCostoEdgeCases:
    """Edge cases for centro_costo NOT+in rules."""

    def test_rule_with_empty_valid_list_detects_all(self):
        """Empty valid centers list → all centros are invalid."""
        condiciones = _not_in_conditions([])

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CENTRO_COSTO")
        ws.cell(row=2, column=1, value="F401")
        ws.cell(row=2, column=2, value="ANYTHING")

        indices = _build_indices("numero_factura", "centro_costo")
        results = _run_engine_detection(
            "test_centro", "odontologia",
            "Any centro detected",
            condiciones, ws, indices,
        )

        facturas = _get_facturas_from_results(results)
        assert "F401" in facturas, "With empty valid list, all centros are invalid"

    def test_whitespace_centro_costo_handled(self):
        """Centro costo with leading/trailing whitespace."""
        condiciones = _not_in_conditions(VALID_ODON_CENTERS)

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CENTRO_COSTO")
        ws.cell(row=2, column=1, value="F501")
        ws.cell(row=2, column=2, value="  ODONTOLOGIA  ")  # whitespace

        indices = _build_indices("numero_factura", "centro_costo")
        results = _run_engine_detection(
            "centro_costo_odontologia_valido", "odontologia",
            "Centro de costo no válido en Odontología",
            condiciones, ws, indices,
        )

        # The in evaluator does exact match; whitespace not stripped
        # This test documents current behavior — may change with trim evaluator
        facturas = _get_facturas_from_results(results)
        assert "F501" in facturas, "Whitespace causes mismatch with in operator"


# ── Output Format ───────────────────────────────────────────────────────────

class TestCentroCostoOutputFormat:
    """Verify engine output format for centro_costo rules."""

    def test_output_has_required_keys(self):
        """Output dicts have factura, problema, regla, severidad."""
        condiciones = _not_in_conditions(VALID_URG_CENTERS)

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CENTRO_COSTO")
        ws.cell(row=2, column=1, value="F601")
        ws.cell(row=2, column=2, value="INVALIDO")

        indices = _build_indices("numero_factura", "centro_costo")
        results = _run_engine_detection(
            "centro_costo_urgencias_valido", "urgencias",
            "Centro de costo no válido en Urgencias",
            condiciones, ws, indices,
        )

        assert len(results) >= 1
        r = results[0]
        assert "factura" in r
        assert "problema" in r
        assert "regla" in r
        assert "severidad" in r
        assert r["regla"] == "centro_costo_urgencias_valido"
        assert r["factura"] == "F601"
