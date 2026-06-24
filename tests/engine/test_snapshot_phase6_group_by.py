"""Snapshot tests Phase 6: engine vs legacy for doble_tipo_procedimiento + revision_cantidad.

Verifies the DB-backed engine with GroupEvaluator detects the same rows
as the legacy Python detectors. Uses mocked DB session with parametros
containing group_by + aggregation configs.
"""

from __future__ import annotations

import pytest
from unittest.mock import MagicMock
from openpyxl import Workbook


# ── Helpers (same pattern as Phases 1-5) ────────────────────────────────────

def _mock_session_with_group_rule(rule_name, dominio, descripcion,
                                   parametros, condiciones_dicts, severity="error"):
    """Create a mock session with Regla that has group_by parametros."""
    from app.models import Regla

    session = MagicMock()
    mock_query = MagicMock()
    mock_query.filter.return_value = mock_query
    mock_query.order_by.return_value = mock_query

    regla = Regla(
        id=1, nombre=rule_name, dominio=dominio,
        estado="active", version=1, prioridad=10, severidad=severity,
        descripcion=descripcion, parametros=parametros,
    )
    mock_query.first.return_value = regla

    cond_mocks = []
    for cd in condiciones_dicts:
        m = MagicMock()
        m.id = cd["id"]
        m.regla_id = cd.get("regla_id", 1)
        m.padre_id = cd["padre_id"]
        m.tipo = cd["tipo"]
        m.operador = cd.get("operador")
        m.fuente_datos = cd.get("fuente_datos")
        m.valor_esperado = cd.get("valor_esperado")
        m.orden = cd.get("orden", 0)
        cond_mocks.append(m)

    mock_query.all.return_value = cond_mocks
    session.query.return_value = mock_query
    return session


def _build_indices(*col_names):
    """Build indices dict: col_name → 0-based index."""
    return {name: i for i, name in enumerate(col_names)}


def _run_engine_detection(rule_name, dominio, descripcion, parametros,
                           condiciones, ws, indices, severity="error"):
    """Run engine detection with mocked session that has group-by parametros."""
    from app.services.engine.rule_based_detector import RuleBasedDetector
    session = _mock_session_with_group_rule(
        rule_name, dominio, descripcion, parametros, condiciones, severity,
    )
    detector = RuleBasedDetector(rule_name, session)
    return detector.detect(ws, indices)


def _get_facturas_from_results(results):
    """Extract factura strings from detection results."""
    facturas = set()
    for r in results:
        if isinstance(r, dict):
            facturas.add(r.get("factura", ""))
        elif isinstance(r, str):
            facturas.add(r)  # legacy decimales returns strings
    return facturas


# ── Condition tree builders for Phase 6 rules ───────────────────────────────

def _doble_tipo_parametros():
    """Parametros for doble_tipo_procedimiento group-by rule."""
    return [{
        "group_by": "numero_factura",
        "aggregations": [
            {"function": "distinct_count", "field": "tipo_procedimiento",
             "target": "distinct_count_tipo_procedimiento"},
        ],
    }]


def _doble_tipo_conditions():
    """Condition tree: gt(invoice.distinct_count_tipo_procedimiento, 1)."""
    return [
        {"id": 1, "padre_id": None, "tipo": "atomic", "operador": "gt",
         "fuente_datos": "invoice.distinct_count_tipo_procedimiento",
         "valor_esperado": "1", "orden": 0},
    ]


def _revision_cantidad_parametros():
    """Parametros for revision_cantidad group-by rule."""
    return [{
        "group_by": "numero_factura",
        "aggregations": [
            {"function": "sum", "field": "cantidad",
             "target": "sum_cantidad"},
        ],
    }]


def _revision_cantidad_conditions():
    """Condition tree: gt(invoice.sum_cantidad, 1)."""
    return [
        {"id": 1, "padre_id": None, "tipo": "atomic", "operador": "gt",
         "fuente_datos": "invoice.sum_cantidad",
         "valor_esperado": "1", "orden": 0},
    ]


# ── Legacy vs Engine comparison ─────────────────────────────────────────────

class TestDobleTipoProcedimientoEngineVsLegacy:
    """Engine with GroupEvaluator vs legacy detect_doble_tipo_procedimiento."""

    def test_engine_detects_same_as_legacy_two_tipos(self):
        """Both detect factura with 2 distinct tipos."""
        from app.services.transversales.doble_tipo_procedimiento import (
            detect_doble_tipo_procedimiento,
        )

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="B")

        indices = _build_indices("numero_factura", "tipo_procedimiento")

        # Legacy
        legacy = detect_doble_tipo_procedimiento(ws, indices)
        legacy_facturas = {r["factura"] for r in legacy}

        # Engine
        engine_results = _run_engine_detection(
            "doble_tipo_procedimiento", "transversal",
            "Factura con doble tipo de procedimiento",
            _doble_tipo_parametros(), _doble_tipo_conditions(),
            ws, indices,
        )
        engine_facturas = _get_facturas_from_results(engine_results)

        assert legacy_facturas == engine_facturas
        assert "F001" in engine_facturas

    def test_engine_detects_same_as_legacy_single_tipo(self):
        """Both produce no results for factura with single tipo."""
        from app.services.transversales.doble_tipo_procedimiento import (
            detect_doble_tipo_procedimiento,
        )

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="A")  # same tipo

        indices = _build_indices("numero_factura", "tipo_procedimiento")

        legacy = detect_doble_tipo_procedimiento(ws, indices)
        engine_results = _run_engine_detection(
            "doble_tipo_procedimiento", "transversal",
            "Factura con doble tipo de procedimiento",
            _doble_tipo_parametros(), _doble_tipo_conditions(),
            ws, indices,
        )

        # Both should be empty
        assert len(legacy) == 0
        assert len(engine_results) == 0

    def test_engine_detects_same_as_legacy_three_tipos(self):
        """Both detect factura with 3 distinct tipos."""
        from app.services.transversales.doble_tipo_procedimiento import (
            detect_doble_tipo_procedimiento,
        )

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="B")
        ws.cell(row=4, column=1, value="F001")
        ws.cell(row=4, column=2, value="C")

        indices = _build_indices("numero_factura", "tipo_procedimiento")

        legacy = detect_doble_tipo_procedimiento(ws, indices)
        legacy_facturas = {r["factura"] for r in legacy}

        engine_results = _run_engine_detection(
            "doble_tipo_procedimiento", "transversal",
            "Factura con doble tipo de procedimiento",
            _doble_tipo_parametros(), _doble_tipo_conditions(),
            ws, indices,
        )
        engine_facturas = _get_facturas_from_results(engine_results)

        assert legacy_facturas == engine_facturas
        assert "F001" in engine_facturas

    def test_engine_matches_legacy_multiple_facturas(self):
        """Both detect the correct subset among multiple facturas."""
        from app.services.transversales.doble_tipo_procedimiento import (
            detect_doble_tipo_procedimiento,
        )

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        # F001: 2 tipos → MATCH
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="B")
        # F002: 1 tipo → NO MATCH
        ws.cell(row=4, column=1, value="F002")
        ws.cell(row=4, column=2, value="C")
        # F003: 3 tipos → MATCH
        ws.cell(row=5, column=1, value="F003")
        ws.cell(row=5, column=2, value="X")
        ws.cell(row=6, column=1, value="F003")
        ws.cell(row=6, column=2, value="Y")
        ws.cell(row=7, column=1, value="F003")
        ws.cell(row=7, column=2, value="Z")

        indices = _build_indices("numero_factura", "tipo_procedimiento")

        legacy = detect_doble_tipo_procedimiento(ws, indices)
        legacy_facturas = {r["factura"] for r in legacy}

        engine_results = _run_engine_detection(
            "doble_tipo_procedimiento", "transversal",
            "Factura con doble tipo de procedimiento",
            _doble_tipo_parametros(), _doble_tipo_conditions(),
            ws, indices,
        )
        engine_facturas = _get_facturas_from_results(engine_results)

        assert legacy_facturas == engine_facturas
        assert "F001" in engine_facturas
        assert "F003" in engine_facturas
        assert "F002" not in engine_facturas

    def test_engine_handles_null_tipo_procedimiento(self):
        """Null tipo_procedimiento values are skipped, not counted as distinct."""
        from app.services.transversales.doble_tipo_procedimiento import (
            detect_doble_tipo_procedimiento,
        )

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value=None)  # null → skip

        indices = _build_indices("numero_factura", "tipo_procedimiento")

        legacy = detect_doble_tipo_procedimiento(ws, indices)
        legacy_facturas = {r["factura"] for r in legacy}

        engine_results = _run_engine_detection(
            "doble_tipo_procedimiento", "transversal",
            "Factura con doble tipo de procedimiento",
            _doble_tipo_parametros(), _doble_tipo_conditions(),
            ws, indices,
        )
        engine_facturas = _get_facturas_from_results(engine_results)

        # Single distinct tipo → no match for both
        assert legacy_facturas == engine_facturas

    def test_engine_handles_missing_column_same_as_legacy(self):
        """Both return empty when column is missing."""
        from app.services.transversales.doble_tipo_procedimiento import (
            detect_doble_tipo_procedimiento,
        )

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=2, column=1, value="F001")

        indices = _build_indices("numero_factura")  # no tipo_procedimiento column

        legacy = detect_doble_tipo_procedimiento(ws, indices)
        engine_results = _run_engine_detection(
            "doble_tipo_procedimiento", "transversal",
            "Factura con doble tipo de procedimiento",
            _doble_tipo_parametros(), _doble_tipo_conditions(),
            ws, indices,
        )

        assert len(legacy) == 0
        assert len(engine_results) == 0


class TestRevisionCantidadEngine:
    """Engine GroupEvaluator for simplified revision_cantidad (sum per factura)."""

    def test_engine_detects_sum_gt_one(self):
        """Total quantity > 1 → detection."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CANTIDAD")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value=1)
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value=1)  # total = 2 > 1

        indices = _build_indices("numero_factura", "cantidad")

        engine_results = _run_engine_detection(
            "revision_cantidad_urgencias", "urgencias",
            "Revisión necesaria: cantidad anómala",
            _revision_cantidad_parametros(), _revision_cantidad_conditions(),
            ws, indices, severity="warning",
        )

        assert len(engine_results) == 1
        assert engine_results[0]["factura"] == "F001"

    def test_engine_no_match_sum_eq_one(self):
        """Total quantity = 1 → no detection."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CANTIDAD")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value=1)

        indices = _build_indices("numero_factura", "cantidad")

        engine_results = _run_engine_detection(
            "revision_cantidad_urgencias", "urgencias",
            "Revisión necesaria: cantidad anómala",
            _revision_cantidad_parametros(), _revision_cantidad_conditions(),
            ws, indices, severity="warning",
        )

        assert len(engine_results) == 0

    def test_engine_handles_empty_values(self):
        """Empty cantidad values default to 0, sum unaffected."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CANTIDAD")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value=5)
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value=None)  # skipped

        indices = _build_indices("numero_factura", "cantidad")

        engine_results = _run_engine_detection(
            "revision_cantidad_urgencias", "urgencias",
            "Revisión necesaria: cantidad anómala",
            _revision_cantidad_parametros(), _revision_cantidad_conditions(),
            ws, indices, severity="warning",
        )

        assert len(engine_results) == 1
        assert engine_results[0]["factura"] == "F001"


class TestPhase6OutputFormat:
    """Verify engine output format for Phase 6 group-by rules."""

    def test_output_has_required_keys_doble_tipo(self):
        """Output dicts for doble_tipo have factura, problema, regla, severidad."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="TIPO_PROCEDIMIENTO")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="A")
        ws.cell(row=3, column=1, value="F001")
        ws.cell(row=3, column=2, value="B")

        indices = _build_indices("numero_factura", "tipo_procedimiento")

        engine_results = _run_engine_detection(
            "doble_tipo_procedimiento", "transversal",
            "Factura con doble tipo de procedimiento",
            _doble_tipo_parametros(), _doble_tipo_conditions(),
            ws, indices,
        )

        assert len(engine_results) >= 1
        r = engine_results[0]
        assert "factura" in r
        assert "problema" in r
        assert "regla" in r
        assert "severidad" in r

    def test_output_has_required_keys_revision_cantidad(self):
        """Output dicts for revision_cantidad have required keys."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NUMERO_FACTURA")
        ws.cell(row=1, column=2, value="CANTIDAD")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value=5)

        indices = _build_indices("numero_factura", "cantidad")

        engine_results = _run_engine_detection(
            "revision_cantidad_urgencias", "urgencias",
            "Revisión necesaria: cantidad anómala",
            _revision_cantidad_parametros(), _revision_cantidad_conditions(),
            ws, indices, severity="warning",
        )

        assert len(engine_results) >= 1
        r = engine_results[0]
        assert "factura" in r
        assert "problema" in r
        assert "regla" in r
        assert "severidad" in r
