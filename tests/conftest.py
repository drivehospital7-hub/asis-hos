"""Configuración global de pytest para el proyecto control_system."""

from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Generator

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from werkzeug.security import generate_password_hash

from app import create_app
from app.database import Base
from app.models import User
from app.utils import users_store
import app.models  # noqa: F401  (registra los modelos en Base.metadata)


@pytest.fixture(autouse=True)
def _db_users_store():
    """Store de usuarios hermético: SQLite en memoria para toda la suite.

    Desde sdd control-errores-role-visibility la DB es la única fuente de
    verdad para usuarios. Este fixture parchea ``users_store.SessionLocal``
    con un engine SQLite en memoria sembrado con usuarios de prueba, para
    que cualquier test que haga login/gestión funcione sin PostgreSQL real.
    Los tests que necesitan comportamiento específico (DB-down, seeds
    propios) parchean SessionLocal por encima sin conflicto.
    """
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    seed_db = Session()
    try:
        for u in [
            {
                "username": "admin",
                "password_hash": generate_password_hash("admin123"),
                "rol": "admin",
                "permisos": ["*"],
                "primer_nombre": "",
                "segundo_nombre": "",
                "apellido_1": "",
                "apellido_2": "",
            },
            {
                "username": "urgencias",
                "password_hash": generate_password_hash("urgencias123"),
                "rol": "usuario",
                "permisos": ["urgencias", "control_urgencias", "facturas_abiertas"],
                "primer_nombre": "",
                "segundo_nombre": "",
                "apellido_1": "",
                "apellido_2": "",
            },
        ]:
            seed_db.add(User(**u))
        seed_db.commit()
    finally:
        seed_db.close()

    with (
        pytest.MonkeyPatch.context() as mp,
    ):
        mp.setattr(users_store, "SessionLocal", Session)
        yield


@pytest.fixture
def app_client():
    """Flask test client usando create_app()."""
    app = create_app()
    app.config["TESTING"] = True
    with app.test_client() as client:
        yield client


@pytest.fixture
def fresh_client():
    """Flask test client with clean session (no cookies from previous tests)."""
    app = create_app()
    app.config["TESTING"] = True
    with app.test_client() as client:
        yield client


@pytest.fixture
def temp_output_dir() -> Generator[Path, None, None]:
    """Directorio temporal para archivos de salida."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def sample_excel_file(temp_output_dir: Path) -> Generator[Path, None, None]:
    """Crea un archivo Excel temporal con datos de ejemplo."""
    file_path = temp_output_dir / "sample.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    
    # Headers en fila 1
    headers = ["NUMERO_FACTURA", "VALOR", "FECHA", "CONVENIO"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos de ejemplo en filas 2-4
    sample_data = [
        ["FAC-001", 15000.50, "2024-01-15", "ODONTOLOGIA"],
        ["FAC-002", 22300.00, "2024-01-16", "ODONTOLOGIA"],
        ["FAC-003", 8750.25, "2024-01-17", "GENERAL"],
    ]
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)
    yield file_path


@pytest.fixture
def empty_excel_file(temp_output_dir: Path) -> Generator[Path, None, None]:
    """Crea un archivo Excel temporal sin columnas (hoja vacía)."""
    file_path = temp_output_dir / "empty.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Vacia"
    # No agregamos nada - hoja completamente vacía
    
    wb.save(file_path)
    yield file_path
