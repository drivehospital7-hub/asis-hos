"""Tests para app/services/cruce_sheet.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.cruce_sheet import (
    get_or_create_sheet,
    find_column_letter_by_header,
    apply_cruce_headers,
    create_cruce_facturas_sheet,
    CRUCE_HEADERS,
)
from app.constants import CRUCE_FACTURAS_SHEET


@pytest.fixture
def empty_workbook() -> Workbook:
    """Crea un workbook vacío."""
    return Workbook()


@pytest.fixture
def workbook_with_cruce_sheet() -> Workbook:
    """Crea un workbook con hoja CruceFacturas existente."""
    wb = Workbook()
    ws = wb.create_sheet(title=CRUCE_FACTURAS_SHEET)
    ws["A1"] = "Contenido existente"
    return wb


@pytest.fixture
def workbook_with_headers() -> Workbook:
    """Crea un workbook con headers en fila 1."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    
    headers = ["Número Factura", "Convenio Facturado", "Centro Costo"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    return wb


class TestGetOrCreateSheet:
    """Tests para la función get_or_create_sheet."""

    def test_crea_hoja_si_no_existe(self, empty_workbook: Workbook) -> None:
        """Debe crear la hoja si no existe en el workbook."""
        wb = empty_workbook
        
        # Antes: la hoja no existe
        assert "NuevaHoja" not in wb.sheetnames
        
        sheet = get_or_create_sheet(wb, "NuevaHoja")
        
        # Después: la hoja existe
        assert sheet.title == "NuevaHoja"
        assert "NuevaHoja" in wb.sheetnames

    def test_retorna_hoja_existente(
        self, workbook_with_cruce_sheet: Workbook
    ) -> None:
        """Debe retornar la hoja existente sin crear una nueva."""
        wb = workbook_with_cruce_sheet
        
        # Antes: la hoja existe con contenido
        hojas_antes = len(wb.sheetnames)
        
        sheet = get_or_create_sheet(wb, CRUCE_FACTURAS_SHEET)
        
        # Después: misma cantidad de hojas, mismo contenido
        assert len(wb.sheetnames) == hojas_antes
        assert sheet.title == CRUCE_FACTURAS_SHEET
        assert sheet["A1"].value == "Contenido existente"

    def test_multiples_llamadas_no_duplican(
        self, empty_workbook: Workbook
    ) -> None:
        """Múltiples llamadas no deben crear hojas duplicadas."""
        wb = empty_workbook
        
        sheet1 = get_or_create_sheet(wb, "TestSheet")
        sheet1["A1"] = "Valor original"
        
        sheet2 = get_or_create_sheet(wb, "TestSheet")
        
        # Deben ser la misma hoja
        assert sheet1 is sheet2
        assert sheet2["A1"].value == "Valor original"


class TestFindColumnLetterByHeader:
    """Tests para la función find_column_letter_by_header (cruce_sheet)."""

    def test_encuentra_primera_columna(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe encontrar header en primera columna."""
        ws = workbook_with_headers.active
        
        result = find_column_letter_by_header(ws, "Número Factura")
        
        assert result == "A"

    def test_encuentra_columna_intermedia(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe encontrar header en columna intermedia."""
        ws = workbook_with_headers.active
        
        result = find_column_letter_by_header(ws, "Convenio Facturado")
        
        assert result == "B"

    def test_encuentra_ultima_columna(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe encontrar header en última columna."""
        ws = workbook_with_headers.active
        
        result = find_column_letter_by_header(ws, "Centro Costo")
        
        assert result == "C"

    def test_retorna_none_si_no_existe(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe retornar None si el header no existe."""
        ws = workbook_with_headers.active
        
        result = find_column_letter_by_header(ws, "Header Inexistente")
        
        assert result is None

    def test_busca_en_fila_especificada(self) -> None:
        """Debe buscar en la fila especificada por headers_row."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "No buscar aquí"
        ws["A3"] = "Header en fila 3"
        
        result = find_column_letter_by_header(ws, "Header en fila 3", headers_row=3)
        
        assert result == "A"


class TestApplyCruceHeaders:
    """Tests para la función apply_cruce_headers."""

    def test_aplica_headers_predefinidos(self, empty_workbook: Workbook) -> None:
        """Debe aplicar los headers predefinidos por defecto."""
        wb = empty_workbook
        ws = wb.create_sheet(title="TestSheet")
        
        result = apply_cruce_headers(ws)
        
        assert ws["B2"].value == "Cruce Facturas"
        assert ws["D2"].value == "Cruce Identificación"
        assert result["sheet"] == "TestSheet"
        assert result["headers"] == CRUCE_HEADERS

    def test_aplica_headers_custom(self, empty_workbook: Workbook) -> None:
        """Debe aplicar headers personalizados si se proporcionan."""
        wb = empty_workbook
        ws = wb.create_sheet(title="CustomSheet")
        custom_headers = {"A1": "Custom Header 1", "C1": "Custom Header 2"}
        
        result = apply_cruce_headers(ws, headers=custom_headers)
        
        assert ws["A1"].value == "Custom Header 1"
        assert ws["C1"].value == "Custom Header 2"
        assert result["headers"] == custom_headers

    def test_retorna_info_correcta(self, empty_workbook: Workbook) -> None:
        """Debe retornar dict con información del proceso."""
        wb = empty_workbook
        ws = wb.create_sheet(title="InfoSheet")
        
        result = apply_cruce_headers(ws)
        
        assert "sheet" in result
        assert "headers" in result
        assert result["sheet"] == "InfoSheet"


class TestCreateCruceFacturasSheet:
    """Tests para la función create_cruce_facturas_sheet."""

    def test_crea_hoja_con_headers(self, empty_workbook: Workbook) -> None:
        """Debe crear hoja CruceFacturas con headers aplicados."""
        wb = empty_workbook
        
        sheet, info = create_cruce_facturas_sheet(wb)
        
        assert sheet.title == CRUCE_FACTURAS_SHEET
        assert CRUCE_FACTURAS_SHEET in wb.sheetnames
        assert sheet["B2"].value == "Cruce Facturas"
        assert sheet["D2"].value == "Cruce Identificación"

    def test_retorna_tupla_con_info(self, empty_workbook: Workbook) -> None:
        """Debe retornar tupla (worksheet, info_dict)."""
        wb = empty_workbook
        
        sheet, info = create_cruce_facturas_sheet(wb)
        
        assert info["rule"] == "cruce_facturas_headers"
        assert info["sheet"] == CRUCE_FACTURAS_SHEET
        assert "B2" in info["cells"]
        assert "D2" in info["cells"]

    def test_no_duplica_si_ya_existe(
        self, workbook_with_cruce_sheet: Workbook
    ) -> None:
        """Si la hoja ya existe, no debe duplicarla."""
        wb = workbook_with_cruce_sheet
        hojas_antes = len(wb.sheetnames)
        
        sheet, info = create_cruce_facturas_sheet(wb)
        
        # No debe haber más hojas
        assert len(wb.sheetnames) == hojas_antes
        # Pero sí debe aplicar los headers (sobrescribiendo si hay)
        assert sheet["B2"].value == "Cruce Facturas"

    def test_integra_get_or_create_y_apply_headers(
        self, empty_workbook: Workbook
    ) -> None:
        """Debe integrar correctamente get_or_create_sheet y apply_cruce_headers."""
        wb = empty_workbook
        
        # Primera llamada: crea
        sheet1, info1 = create_cruce_facturas_sheet(wb)
        sheet1["Z1"] = "Marcador"
        
        # Segunda llamada: retorna existente
        sheet2, info2 = create_cruce_facturas_sheet(wb)
        
        # Deben ser la misma hoja
        assert sheet2["Z1"].value == "Marcador"
