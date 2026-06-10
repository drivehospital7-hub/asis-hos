"""Tests for app/services/transversales/cups_equivalentes.py.

Transversal detector — applies to ALL factura types.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.transversales.cups_equivalentes import (
    CODIGOS_CUPS_EQUIVALENTES,
    detect_cups_equivalentes_transversal,
)


# =============================================================================
# Helpers
# =============================================================================


def _build_workbook(
    headers: list[str],
    rows: list[list],
) -> tuple[Workbook, dict[str, int | None]]:
    """Construye un workbook con headers y datos para pruebas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    indices: dict[str, int | None] = {}
    internal_names = {
        "Número Factura": "numero_factura",
        "Código": "codigo",
        "Procedimiento": "procedimiento",
    }
    for col_idx, header in enumerate(headers):
        if header in internal_names:
            indices[internal_names[header]] = col_idx
    for name in ("numero_factura", "codigo", "procedimiento"):
        indices.setdefault(name, None)

    return wb, indices


# =============================================================================
# Constants tests
# =============================================================================


class TestCodigosCupsEquivalentesConstant:
    """Tests for CODIGOS_CUPS_EQUIVALENTES constant."""

    def test_constant_values(self) -> None:
        """Constant dict must contain both CUPS mappings."""
        assert CODIGOS_CUPS_EQUIVALENTES == {
            "906317": "1906317",   # Hepatitis B (Prueba rápida)
            "906249": "906249PR",  # VIH Prueba rápida
        }

    def test_constant_is_dict(self) -> None:
        """Constant must be of type dict."""
        assert isinstance(CODIGOS_CUPS_EQUIVALENTES, dict)

    def test_constant_has_two_entries(self) -> None:
        """Constant must have exactly 2 entries."""
        assert len(CODIGOS_CUPS_EQUIVALENTES) == 2


# =============================================================================
# Detector unit tests
# =============================================================================


class TestDetectCupsEquivalentesTransversal:
    """Tests para detect_cups_equivalentes_transversal."""

    # ── R1: Detect 906317 (Hepatitis B Rápida) ────────────────────────

    def test_detect_906317(self) -> None:
        """Código 906317 debe ser detectado como error."""
        headers = ["Número Factura", "Código", "Procedimiento"]
        rows = [["FAC-001", "906317", "Hepatitis B Antígeno Superficie"]]
        wb, indices = _build_workbook(headers, rows)
        result = detect_cups_equivalentes_transversal(wb.active, indices)

        assert len(result) == 1
        item = result[0]
        assert item["factura"] == "FAC-001"
        assert item["codigo"] == "906317"
        assert item["codigo_equiv"] == "1906317"
        assert item["accion"] == "Usar 1906317"
        assert item["procedimiento"] == "Hepatitis B Antígeno Superficie"

    # ── R2: Detect 906249 (VIH Prueba Rápida) ─────────────────────────

    def test_detect_906249(self) -> None:
        """Código 906249 debe ser detectado como error."""
        headers = ["Número Factura", "Código", "Procedimiento"]
        rows = [["FAC-002", "906249", "VIH Prueba Rápida"]]
        wb, indices = _build_workbook(headers, rows)
        result = detect_cups_equivalentes_transversal(wb.active, indices)

        assert len(result) == 1
        item = result[0]
        assert item["factura"] == "FAC-002"
        assert item["codigo"] == "906249"
        assert item["codigo_equiv"] == "906249PR"
        assert item["accion"] == "Usar 906249PR"
        assert item["procedimiento"] == "VIH Prueba Rápida"

    # ── Both codes in same sheet ──────────────────────────────────────

    def test_detect_both_codes(self) -> None:
        """Ambos códigos incorrectos deben ser detectados."""
        headers = ["Número Factura", "Código", "Procedimiento"]
        rows = [
            ["FAC-001", "906317", "Hepatitis B"],
            ["FAC-002", "906249", "VIH Prueba"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_cups_equivalentes_transversal(wb.active, indices)

        assert len(result) == 2
        codigos = {item["codigo"] for item in result}
        assert codigos == {"906317", "906249"}

    # ── Already correct codes → no flags ──────────────────────────────

    def test_correct_codes_not_flagged(self) -> None:
        """Códigos correctos (1906317, 906249PR) no deben generar error."""
        headers = ["Número Factura", "Código", "Procedimiento"]
        rows = [
            ["FAC-001", "1906317", "Hepatitis B Rápida"],
            ["FAC-002", "906249PR", "VIH Prueba Rápida"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_cups_equivalentes_transversal(wb.active, indices)

        assert result == []

    # ── Non-matching codes → no flags ─────────────────────────────────

    def test_non_matching_codes_not_flagged(self) -> None:
        """Códigos no relacionados no deben generar error."""
        headers = ["Número Factura", "Código", "Procedimiento"]
        rows = [
            ["FAC-001", "904902", "TSH"],
            ["FAC-002", "906241", "Rubeola IGG"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_cups_equivalentes_transversal(wb.active, indices)

        assert result == []

    # ── Missing columns → empty list ──────────────────────────────────

    def test_missing_codigo_column(self) -> None:
        """Sin columna Código debe retornar lista vacía."""
        headers = ["Número Factura", "Procedimiento"]
        rows = [["FAC-001", "Hepatitis B"]]
        wb, indices = _build_workbook(headers, rows)
        result = detect_cups_equivalentes_transversal(wb.active, indices)

        assert result == []

    def test_missing_numero_factura_column(self) -> None:
        """Sin columna Número Factura debe retornar lista vacía."""
        headers = ["Código", "Procedimiento"]
        rows = [["906317", "Hepatitis B"]]
        wb, indices = _build_workbook(headers, rows)
        result = detect_cups_equivalentes_transversal(wb.active, indices)

        assert result == []

    # ── Empty/non-string cells ────────────────────────────────────────

    def test_empty_codigo_cell_skipped(self) -> None:
        """Celda Código vacía debe skipearse."""
        headers = ["Número Factura", "Código", "Procedimiento"]
        rows = [["FAC-001", None, "Hepatitis B"]]
        wb, indices = _build_workbook(headers, rows)
        result = detect_cups_equivalentes_transversal(wb.active, indices)

        assert result == []

    def test_empty_factura_cell_skipped(self) -> None:
        """Celda Número Factura vacía debe skipearse."""
        headers = ["Número Factura", "Código", "Procedimiento"]
        rows = [[None, "906317", "Hepatitis B"]]
        wb, indices = _build_workbook(headers, rows)
        result = detect_cups_equivalentes_transversal(wb.active, indices)

        assert result == []

    # ── procedimiento column missing → still works ────────────────────

    def test_missing_procedimiento_column(self) -> None:
        """Sin columna Procedimiento, resultado debe tener proc vacío."""
        headers = ["Número Factura", "Código"]
        rows = [["FAC-001", "906317"]]
        wb, indices = _build_workbook(headers, rows)
        result = detect_cups_equivalentes_transversal(wb.active, indices)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"
        assert result[0]["procedimiento"] == ""

    # ── All 5 keys present ────────────────────────────────────────────

    def test_all_five_keys_present(self) -> None:
        """Cada item debe tener factura, codigo, codigo_equiv, accion, procedimiento."""
        headers = ["Número Factura", "Código", "Procedimiento"]
        rows = [["FAC-001", "906317", "Hepatitis B"]]
        wb, indices = _build_workbook(headers, rows)
        result = detect_cups_equivalentes_transversal(wb.active, indices)

        assert len(result) == 1
        assert set(result[0].keys()) == {"factura", "codigo", "codigo_equiv", "accion", "procedimiento"}
