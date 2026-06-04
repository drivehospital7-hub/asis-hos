"""Tests for app/services/intramural/bacteriologas_cronograma.py.

Strict TDD: tests written BEFORE implementation.
"""

from __future__ import annotations

# T-01: These imports will fail until constants are added (RED test)
from app.constants.intramural import (
    RESPONSABLE_CHAPUEL,
    RESPONSABLE_ORDONEZ,
    RESPONSABLE_TAPIA,
)
from app.constants.urgencias import FACTURADORES_URGENCIAS

from datetime import date

import pytest
from openpyxl import Workbook

from app.services.intramural.bacteriologas_cronograma import (
    _parse_fecha,
    detect_bacteriologas_cronograma,
)


# =============================================================================
# _parse_fecha tests (Task 3.1)
# =============================================================================


class TestParseFecha:
    """Tests para _parse_fecha helper."""

    def test_iso_string(self) -> None:
        """ISO string '2024-03-15' debe devolver date(2024, 3, 15)."""
        result = _parse_fecha("2024-03-15")
        assert result == date(2024, 3, 15)

    def test_excel_serial_int(self) -> None:
        """Serial Excel 45367 debe devolver date(2024, 3, 16)."""
        result = _parse_fecha(45367)
        assert result == date(2024, 3, 16)

    def test_excel_serial_float(self) -> None:
        """Serial Excel como float 45367.0 debe devolver date(2024, 3, 16)."""
        result = _parse_fecha(45367.0)
        assert result == date(2024, 3, 16)

    def test_local_format_dd_mm_yyyy(self) -> None:
        """String local '15/03/2024' debe devolver date(2024, 3, 15)."""
        result = _parse_fecha("15/03/2024")
        assert result == date(2024, 3, 15)

    def test_local_format_dd_mm_yyyy_dash(self) -> None:
        """String local '15-03-2024' debe devolver date(2024, 3, 15)."""
        result = _parse_fecha("15-03-2024")
        assert result == date(2024, 3, 15)

    def test_invalid_string(self) -> None:
        """String inválido 'not-a-date' debe devolver None."""
        result = _parse_fecha("not-a-date")
        assert result is None

    def test_none(self) -> None:
        """None debe devolver None."""
        result = _parse_fecha(None)
        assert result is None

    def test_empty_string(self) -> None:
        """String vacío debe devolver None."""
        result = _parse_fecha("")
        assert result is None


# =============================================================================
# Helpers for detector tests
# =============================================================================


def _build_workbook(
    rows: list[dict],
    extra_headers: list[str] | None = None,
) -> tuple[Workbook, dict[str, int | None]]:
    """Construye un workbook con headers y datos para pruebas.

    Headers fijos: Numero Factura, Tipo Factura Descripcion,
    Codigo Tipo Procedimiento, Laboratorio, Codigo,
    Codigo Profesional, Profesional Atiende, Procedimiento,
    Fec Factura.
    Nota: headers sin acentos para que el normalize del test
    produzca keys sin tilde (coincidiendo con lo que busca el detector).

    Args:
        rows: Lista de dicts con valores por columna.
        extra_headers: Headers adicionales opcionales.

    Returns:
        (Workbook, dict of indices).
    """
    headers = [
        "Numero Factura",
        "Tipo Factura Descripcion",
        "Codigo Tipo Procedimiento",
        "Laboratorio",
        "Codigo",
        "Codigo Profesional",
        "Profesional Atiende",
        "Procedimiento",
        "Fec Factura",
    ]
    if extra_headers:
        headers.extend(extra_headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_name, value in row_data.items():
            if col_name in headers:
                col_pos = headers.index(col_name) + 1
                ws.cell(row=row_idx, column=col_pos, value=value)

    indices = {h.lower().replace(" ", "_"): i for i, h in enumerate(headers)}
    return wb, indices


# =============================================================================
# detect_bacteriologas_cronograma tests (Task 3.2)
# =============================================================================


class TestDetectBacteriologasCronograma:
    """Tests para detect_bacteriologas_cronograma."""

    # ── Filtros: skip si no es Intramural ──────────────────────────────

    def test_skip_no_intramural(self) -> None:
        """Factura no Intramural debe skipearse."""
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Urgencias",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    # ── Filtros: skip si tipo no es 02/05 ──────────────────────────────

    def test_skip_wrong_tipo_procedimiento(self) -> None:
        """Tipo procedimiento diferente de 02/05 debe skipearse."""
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "14",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    # ── Filtros: skip si laboratorio != "Si" ───────────────────────────

    def test_skip_laboratorio_no(self) -> None:
        """Laboratorio distinto de 'Si' debe skipearse."""
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "No",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    def test_skip_laboratorio_si_con_acento(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Laboratorio 'SÍ' (con acento) debe tratarse como 'Si'."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia: [
                {"nombre": "PABON GARCIA ALEJANDRA", "codigo": "CE"},
            ],
        )
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "SÍ",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    def test_skip_laboratorio_si_mayusculas(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Laboratorio 'SI' (mayúsculas) debe tratarse como 'Si'."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia: [
                {"nombre": "PABON GARCIA ALEJANDRA", "codigo": "CE"},
            ],
        )
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "SI",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    # ── Filtros: skip si código en EXCEPCIONES_BACTERIOLOGA ────────────

    def test_skip_excepcion_bacteriologa_904903(self) -> None:
        """Código 904903 (excepción) debe skipearse."""
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904903",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "ALGO",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    def test_skip_excepcion_bacteriologa_903883(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Código 903883 (excepción) debe skipearse."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia: [],
        )
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "903883",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "ALGO",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    # ── PROFESIONALES_URGENCIAS: no encontrado ─────────────────────────

    def test_error_profesional_no_encontrado(self) -> None:
        """Profesional no en PROFESIONALES_URGENCIAS debe dar error."""
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "99999",
                "Profesional Atiende": "NO EXISTE",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"
        assert result[0]["codigo_profesional"] == "99999"
        assert "no está en el listado" in result[0]["problema"].lower()

    # ── PROFESIONALES_URGENCIAS: encontrado pero no BACTERIOLOGA ───────

    def test_error_profesional_no_bacteriologa(self) -> None:
        """Profesional en listado pero no BACTERIOLOGA debe dar error."""
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "01293",
                # 01293 = RODRIGUEZ MORALES JAMEZ ARLEY (MEDICO)
                "Profesional Atiende": "RODRIGUEZ MORALES JAMEZ ARLEY",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"
        assert "no es una bacterióloga" in result[0]["problema"].lower()

    # ── Cronograma: skip si get_turno_del_dia retorna [] ───────────────

    def test_skip_cronograma_vacio(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Cronograma vacío debe skipearse."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia: [],
        )
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    # ── Error: bacterióloga no en cronograma del día ───────────────────

    def test_error_fuera_de_cronograma(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Bacterióloga no en turnos del día debe dar error."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia: [
                {"nombre": "OTRA PERSONA", "codigo": "03375"},
            ],
        )
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"
        assert result[0]["codigo_profesional"] == "03730"
        assert result[0]["regla"] == "Bacterióloga debe estar en cronograma del día"
        assert "no está en el cronograma" in result[0]["problema"].lower()

    # ── Happy path: bacterióloga en cronograma ─────────────────────────

    def test_ok_en_cronograma(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Bacterióloga en turnos del día no debe generar error."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia: [
                {"nombre": "PABON GARCIA ALEJANDRA", "codigo": "03730"},
            ],
        )
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    # ── Dedup: una factura, un error ───────────────────────────────────

    def test_una_factura_un_error(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Misma factura con múltiples filas solo reporta un error."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia: [
                {"nombre": "OTRA PERSONA", "codigo": "03375"},
            ],
        )
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "05",
                "Laboratorio": "Si",
                "Codigo": "904903",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "OTRO EXAMEN",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert len(result) == 1

    # ── Fecha inválida → skip ──────────────────────────────────────────

    def test_fecha_invalida_skip(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Fecha no parseable debe skipear la fila."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia: [],
        )
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "NOT-A-DATE",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    # ── Tipo 05 también procesado ──────────────────────────────────────

    def test_tipo_05_procesado(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Tipo procedimiento 05 también debe validarse."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia: [
                {"nombre": "PABON GARCIA ALEJANDRA", "codigo": "03730"},
            ],
        )
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "05",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    # ── Indices faltantes → retorna [] ─────────────────────────────────

    def test_sin_indices_retorna_vacio(self) -> None:
        """Sin índices de columnas clave, retorna lista vacía."""
        wb, _ = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, {})
        assert result == []

    def test_sin_indice_numero_factura_retorna_vacio(self) -> None:
        """Sin índice numero_factura, retorna lista vacía."""
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        indices.pop("numero_factura", None)
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    def test_sin_indice_codigo_profesional_retorna_vacio(self) -> None:
        """Sin índice codigo_profesional, retorna lista vacía."""
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        indices.pop("codigo_profesional", None)
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []

    def test_serial_excel_fec_factura(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Fec Factura como serial Excel debe parsearse correctamente."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia: [
                {"nombre": "PABON GARCIA ALEJANDRA", "codigo": "03730"},
            ],
        )
        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": 45367,
            },
        ])
        result = detect_bacteriologas_cronograma(wb.active, indices)
        assert result == []


# =============================================================================
# Integration test (Task 3.3)
# =============================================================================


class TestDetectAllIncludesBacteriologas:
    """Integración: detect_all_problems_intramural incluye bacteriólogas."""

    def test_resultado_incluye_profesionales_key(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """detect_all_problems_intramural debe incluir key 'profesionales'."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia: [],
        )
        from app.services.intramural.detect_all import detect_all_problems_intramural

        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result, _ = detect_all_problems_intramural(wb.active, indices)
        assert "profesionales" in result["problemas"]
        assert "profesionales" in result["totales"]

    def test_profesionales_es_lista(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """La key profesionales debe ser una lista."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia: [],
        )
        from app.services.intramural.detect_all import detect_all_problems_intramural

        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result, _ = detect_all_problems_intramural(wb.active, indices)
        assert isinstance(result["problemas"]["profesionales"], list)
        assert isinstance(result["totales"]["profesionales"], int)


# =============================================================================
# T-03: get_turno_del_dia siglas_filter tests (written BEFORE implementation)
# =============================================================================


class TestGetTurnoDelDiaSiglasFilter:
    """Tests para get_turno_del_dia con siglas_filter parameter (T-03/T-06)."""

    def _mock_cronograma(self) -> dict:
        """Retorna un cronograma con turnos mixtos CE/PYM/otros."""
        return {
            "mes": 6,
            "anio": 2026,
            "dias": [
                {
                    "dia": 1,
                    "turnos": {
                        "PABON GARCIA ALEJANDRA": "03730",
                        "PEÑA PEÑA LISBETH PAOLA": "03375",
                        "MOLINA ALVAREZ KAROL DAYANNA": "PYM-03374",
                        "MARIN ZULUAGA VALENTINA": "CE-03255",
                    },
                },
            ],
        }

    def test_default_none_filters_ce_or_pym(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """siglas_filter=None debe mantener filtro actual CE|PYM."""
        from app.services.cronograma_bacteriologas_service import get_turno_del_dia

        monkeypatch.setattr(
            "app.services.cronograma_bacteriologas_service.get_cronograma",
            lambda mes, anio: self._mock_cronograma(),
        )
        turnos = get_turno_del_dia(6, 2026, 1, siglas_filter=None)
        # 03730 y 03375 no tienen CE/PYM en codigo -> filtrados
        # PYM-03374 tiene PYM -> incluido
        # CE-03255 tiene CE -> incluido
        assert len(turnos) == 2
        codigos = [t["codigo"] for t in turnos]
        assert "PYM-03374" in codigos
        assert "CE-03255" in codigos

    def test_empty_set_returns_all(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """siglas_filter=set() (vacio) debe retornar todos los turnos sin filtrar."""
        from app.services.cronograma_bacteriologas_service import get_turno_del_dia

        monkeypatch.setattr(
            "app.services.cronograma_bacteriologas_service.get_cronograma",
            lambda mes, anio: self._mock_cronograma(),
        )
        turnos = get_turno_del_dia(6, 2026, 1, siglas_filter=set())
        assert len(turnos) == 4

    def test_pym_only_filter(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """siglas_filter={'PYM'} debe incluir solo turnos con PYM en codigo."""
        from app.services.cronograma_bacteriologas_service import get_turno_del_dia

        monkeypatch.setattr(
            "app.services.cronograma_bacteriologas_service.get_cronograma",
            lambda mes, anio: self._mock_cronograma(),
        )
        turnos = get_turno_del_dia(6, 2026, 1, siglas_filter={"PYM"})
        assert len(turnos) == 1
        assert turnos[0]["codigo"] == "PYM-03374"

    def test_ce_only_filter(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """siglas_filter={'CE'} debe incluir solo turnos con CE en codigo."""
        from app.services.cronograma_bacteriologas_service import get_turno_del_dia

        monkeypatch.setattr(
            "app.services.cronograma_bacteriologas_service.get_cronograma",
            lambda mes, anio: self._mock_cronograma(),
        )
        turnos = get_turno_del_dia(6, 2026, 1, siglas_filter={"CE"})
        assert len(turnos) == 1
        assert turnos[0]["codigo"] == "CE-03255"


# =============================================================================
# T-04/T-07: detect_bacteriologas_cronograma con responsable_cierra
# =============================================================================


class TestDetectBacteriologasCronogramaConResponsable:
    """Tests para detect_bacteriologas_cronograma con responsable_cierra (T-04/T-07)."""

    # ── Chapuel + bacterióloga solo PYM -> sin error ─────────────────────

    def test_chapuel_pym_ok(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Chapuel + bacterióloga solo PYM debe ser valido."""
        call_log: list[dict] = []

        def mock_get_turno(mes, anio, dia, siglas_filter=None):
            call_log.append({"siglas_filter": siglas_filter})
            if siglas_filter == {"PYM"}:
                return [{"nombre": "PABON GARCIA ALEJANDRA", "codigo": "03730"}]
            return []

        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            mock_get_turno,
        )
        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "03730",
                    "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
            extra_headers=["Responsable Cierra"],
        )
        result = detect_bacteriologas_cronograma(
            wb.active, indices,
            responsable_cierra={"FAC-001": "CHAPUEL CASANOVA ANGIE TATIANA"},
        )
        assert result == []
        assert len(call_log) == 1
        assert call_log[0]["siglas_filter"] == {"PYM"}

    # ── Chapuel + bacterióloga solo CE -> error ─────────────────────────

    def test_chapuel_ce_error(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Chapuel + bacterióloga solo CE debe dar error (solo PYM permitido)."""
        call_log: list[dict] = []

        def mock_get_turno(mes, anio, dia, siglas_filter=None):
            call_log.append({"siglas_filter": siglas_filter})
            if siglas_filter == {"PYM"}:
                # PYM turnos existen pero NO incluyen a nuestra bacterióloga
                return [{"nombre": "OTRA PERSONA", "codigo": "PYM-09999"}]
            return []

        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            mock_get_turno,
        )
        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "03730",
                    "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
            extra_headers=["Responsable Cierra"],
        )
        result = detect_bacteriologas_cronograma(
            wb.active, indices,
            responsable_cierra={"FAC-001": "CHAPUEL CASANOVA ANGIE TATIANA"},
        )
        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"
        assert "no está en el cronograma" in result[0]["problema"].lower()
        assert call_log[0]["siglas_filter"] == {"PYM"}

    # ── Tapia + bacterióloga solo CE -> sin error ───────────────────────

    def test_tapia_ce_ok(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Tapia + bacterióloga solo CE debe ser valido."""
        call_log: list[dict] = []

        def mock_get_turno(mes, anio, dia, siglas_filter=None):
            call_log.append({"siglas_filter": siglas_filter})
            if siglas_filter == {"CE"}:
                return [{"nombre": "MARIN ZULUAGA VALENTINA", "codigo": "CE-03255"}]
            return []

        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            mock_get_turno,
        )
        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "03255",
                    "Profesional Atiende": "MARIN ZULUAGA VALENTINA",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
            extra_headers=["Responsable Cierra"],
        )
        result = detect_bacteriologas_cronograma(
            wb.active, indices,
            responsable_cierra={"FAC-001": "TAPIA PERDOMO ANYI CATALEYA"},
        )
        assert result == []
        assert call_log[0]["siglas_filter"] == {"CE"}

    # ── Tapia + bacterióloga solo PYM -> error ──────────────────────────

    def test_tapia_pym_error(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Tapia + bacterióloga solo PYM debe dar error (solo CE permitido)."""
        call_log: list[dict] = []

        def mock_get_turno(mes, anio, dia, siglas_filter=None):
            call_log.append({"siglas_filter": siglas_filter})
            if siglas_filter == {"CE"}:
                # CE turnos existen pero NO incluyen a nuestra bacterióloga
                return [{"nombre": "OTRA PERSONA", "codigo": "CE-09999"}]
            return []

        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            mock_get_turno,
        )
        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "03730",
                    "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
            extra_headers=["Responsable Cierra"],
        )
        result = detect_bacteriologas_cronograma(
            wb.active, indices,
            responsable_cierra={"FAC-001": "TAPIA PERDOMO ANYI CATALEYA"},
        )
        assert len(result) == 1
        assert call_log[0]["siglas_filter"] == {"CE"}

    # ── Ordoñez + bacterióloga solo CE -> sin error ─────────────────────

    def test_ordonez_ce_ok(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Ordonez + bacterióloga solo CE debe ser valido."""
        call_log: list[dict] = []

        def mock_get_turno(mes, anio, dia, siglas_filter=None):
            call_log.append({"siglas_filter": siglas_filter})
            if siglas_filter == {"CE"}:
                return [{"nombre": "MARIN ZULUAGA VALENTINA", "codigo": "CE-03255"}]
            return []

        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            mock_get_turno,
        )
        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "03255",
                    "Profesional Atiende": "MARIN ZULUAGA VALENTINA",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
            extra_headers=["Responsable Cierra"],
        )
        result = detect_bacteriologas_cronograma(
            wb.active, indices,
            responsable_cierra={"FAC-001": "ORDOÑEZ MEZA SILVIA ELEY"},
        )
        assert result == []
        assert call_log[0]["siglas_filter"] == {"CE"}

    # ── Facturador Urgencias + bacterióloga valida -> sin error (bypass) ─

    def test_urgencias_bypass_ok(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Facturador Urgencias + bacterióloga valida debe ser valido (sin cronograma)."""
        call_log: list[dict] = []

        def mock_get_turno(mes, anio, dia, siglas_filter=None):
            call_log.append({"siglas_filter": siglas_filter})
            return [{"nombre": "PABON GARCIA ALEJANDRA", "codigo": "03730"}]

        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            mock_get_turno,
        )
        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "03730",
                    "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
            extra_headers=["Responsable Cierra"],
        )
        result = detect_bacteriologas_cronograma(
            wb.active, indices,
            responsable_cierra={"FAC-001": "ARIAS CULCHA ANGIE CAROLINA"},
        )
        # Urgencias bypass: valida contra PROFESIONALES, no cronograma
        assert result == []
        # NO debe llamar a get_turno_del_dia (bypass)
        assert len(call_log) == 0

    # ── Facturador Urgencias + bacterióloga NO encontrada -> error ──────

    def test_urgencias_bypass_no_encontrado_error(self) -> None:
        """Facturador Urgencias + bacterióloga no en PROFESIONALES debe dar error."""
        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "99999",
                    "Profesional Atiende": "NO EXISTE",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
            extra_headers=["Responsable Cierra"],
        )
        result = detect_bacteriologas_cronograma(
            wb.active, indices,
            responsable_cierra={"FAC-001": "ESPAÑA DIAZ LORENY ALEJANDRA"},
        )
        assert len(result) == 1
        assert "no está en el listado" in result[0]["problema"].lower()

    # ── Facturador Urgencias + profesional no BACTERIOLOGA -> error ─────

    def test_urgencias_bypass_no_bacteriologa_error(self) -> None:
        """Facturador Urgencias + profesional no BACTERIOLOGA debe dar error."""
        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "01293",
                    "Profesional Atiende": "RODRIGUEZ MORALES JAMEZ ARLEY",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
            extra_headers=["Responsable Cierra"],
        )
        result = detect_bacteriologas_cronograma(
            wb.active, indices,
            responsable_cierra={"FAC-001": "MEZA FERNANDEZ CARLOS OMAR"},
        )
        assert len(result) == 1
        assert "no es una bacterióloga" in result[0]["problema"].lower()

    # ── Otro responsable -> default CE|PYM (comportamiento actual) ──────

    def test_otro_responsable_default_filter(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Otro responsable debe usar filtro default CE|PYM."""
        call_log: list[dict] = []

        def mock_get_turno(mes, anio, dia, siglas_filter=None):
            call_log.append({"siglas_filter": siglas_filter})
            return [{"nombre": "PABON GARCIA ALEJANDRA", "codigo": "03730"}]

        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            mock_get_turno,
        )
        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "03730",
                    "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
            extra_headers=["Responsable Cierra"],
        )
        result = detect_bacteriologas_cronograma(
            wb.active, indices,
            responsable_cierra={"FAC-001": "ALGUN OTRO RESPONSABLE"},
        )
        assert result == []
        assert call_log[0]["siglas_filter"] is None

    # ── responsable_cierra=None -> fallback a default CE|PYM ────────────

    def test_responsable_none_fallback(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """responsable_cierra=None debe usar comportamiento default CE|PYM."""
        call_log: list[dict] = []

        def mock_get_turno(mes, anio, dia, siglas_filter=None):
            call_log.append({"siglas_filter": siglas_filter})
            return [{"nombre": "PABON GARCIA ALEJANDRA", "codigo": "03730"}]

        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            mock_get_turno,
        )
        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "03730",
                    "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
        )
        result = detect_bacteriologas_cronograma(
            wb.active, indices,
            responsable_cierra=None,
        )
        assert result == []
        # None -> siglas_filter=None -> default CE|PYM
        assert call_log[0]["siglas_filter"] is None

    # ── Responsable no encontrado en mapa (factura sin responsable) ─────

    def test_responsable_no_en_mapa(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Factura sin entry en responsable_cierra debe usar default CE|PYM."""
        call_log: list[dict] = []

        def mock_get_turno(mes, anio, dia, siglas_filter=None):
            call_log.append({"siglas_filter": siglas_filter})
            return [{"nombre": "PABON GARCIA ALEJANDRA", "codigo": "03730"}]

        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            mock_get_turno,
        )
        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "03730",
                    "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
            extra_headers=["Responsable Cierra"],
        )
        # FAC-001 no esta en el mapa
        result = detect_bacteriologas_cronograma(
            wb.active, indices,
            responsable_cierra={"FAC-002": "CHAPUEL CASANOVA ANGIE TATIANA"},
        )
        assert result == []
        assert call_log[0]["siglas_filter"] is None

    # ── Case insensitive: espacios/tildes irregulares ───────────────────

    def test_responsable_case_insensitive(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Responsable con case/espacios irregulares debe matchear."""
        call_log: list[dict] = []

        def mock_get_turno(mes, anio, dia, siglas_filter=None):
            call_log.append({"siglas_filter": siglas_filter})
            if siglas_filter == {"PYM"}:
                return [{"nombre": "PABON GARCIA ALEJANDRA", "codigo": "03730"}]
            return []

        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            mock_get_turno,
        )
        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "03730",
                    "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
            extra_headers=["Responsable Cierra"],
        )
        # Case irregular (lowercase) + espacios extra
        result = detect_bacteriologas_cronograma(
            wb.active, indices,
            responsable_cierra={"FAC-001": "  chapuel casanova angie tatiana  "},
        )
        assert result == []
        assert call_log[0]["siglas_filter"] == {"PYM"}


# =============================================================================
# T-05/T-08: detect_all_problems_intramural pasa responsable_cierra
# =============================================================================


class TestDetectAllPassesResponsableCierra:
    """Integracion: detect_all_problems_intramural pasa responsable_cierra (T-05/T-08)."""

    def test_detect_all_passes_responsable_cierra(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """detect_all_problems_intramural debe pasar responsable_cierra al detector."""
        call_log: list[dict] = []

        def mock_get_turno(mes, anio, dia, siglas_filter=None):
            call_log.append({"siglas_filter": siglas_filter})
            return []

        # Patch the get_turno_del_dia used by the detector module
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            mock_get_turno,
        )
        from app.services.intramural.detect_all import detect_all_problems_intramural

        wb, indices = _build_workbook(
            [
                {
                    "Numero Factura": "FAC-001",
                    "Tipo Factura Descripcion": "Intramural",
                    "Codigo Tipo Procedimiento": "02",
                    "Laboratorio": "Si",
                    "Codigo": "904902",
                    "Codigo Profesional": "03730",
                    "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                    "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                    "Fec Factura": "15/03/2024",
                },
            ],
            extra_headers=["Responsable Cierra"],
        )
        # Set the responsable_cierra cell value
        ws = wb.active
        resp_idx = indices["responsable_cierra"]
        ws.cell(row=2, column=resp_idx + 1, value="CHAPUEL CASANOVA ANGIE TATIANA")

        result, _ = detect_all_problems_intramural(wb.active, indices)
        assert "profesionales" in result["problemas"]
        # Should have called get_turno_del_dia with siglas_filter={"PYM"}
        assert any(
            c.get("siglas_filter") == {"PYM"}
            for c in call_log
        )

    def test_detect_all_without_responsable_column(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """detect_all sin columna responsable_cierra debe funcionar igual."""
        monkeypatch.setattr(
            "app.services.intramural.bacteriologas_cronograma.get_turno_del_dia",
            lambda mes, anio, dia, siglas_filter=None: [],
        )
        from app.services.intramural.detect_all import detect_all_problems_intramural

        wb, indices = _build_workbook([
            {
                "Numero Factura": "FAC-001",
                "Tipo Factura Descripcion": "Intramural",
                "Codigo Tipo Procedimiento": "02",
                "Laboratorio": "Si",
                "Codigo": "904902",
                "Codigo Profesional": "03730",
                "Profesional Atiende": "PABON GARCIA ALEJANDRA",
                "Procedimiento": "Hormona Estimulante del Tiroides [TSH]",
                "Fec Factura": "15/03/2024",
            },
        ])
        result, _ = detect_all_problems_intramural(wb.active, indices)
        assert "profesionales" in result["problemas"]
