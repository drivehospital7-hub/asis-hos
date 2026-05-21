"""Tests para app/services/urgencias/cantidades_soat_urgencias.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.urgencias.cantidades_soat_urgencias import detect_cantidades_soat_urgencias


@pytest.fixture
def wb_soat_urgencias_headers() -> Workbook:
    """Crea un workbook con headers para validación SOAT Urgencias."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código")
    ws.cell(row=1, column=3, value="Procedimiento")
    ws.cell(row=1, column=4, value="Cantidad")
    ws.cell(row=1, column=5, value="Tipo Factura Descripción")
    ws.cell(row=1, column=6, value="Tarifario")
    return wb


class TestDetectCantidadesSoatUrgencias:
    """Tests para detect_cantidades_soat_urgencias."""

    def test_cantidad_distinta_1_con_codigo_soat_genera_error(
        self, wb_soat_urgencias_headers: Workbook
    ) -> None:
        """Código 39145 con cantidad != 1 en SOAT Urgencias debe generar error."""
        ws = wb_soat_urgencias_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="39145")
        ws.cell(row=2, column=3, value="PROC A")
        ws.cell(row=2, column=4, value=2)
        ws.cell(row=2, column=5, value="Urgencias")
        ws.cell(row=2, column=6, value="SOAT")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
            "tarifario": 5,
        }
        result = detect_cantidades_soat_urgencias(ws, indices)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"
        assert result[0]["codigo"] == "39145"
        assert result[0]["cantidad"] == 2

    def test_cantidad_1_no_genera_error(
        self, wb_soat_urgencias_headers: Workbook
    ) -> None:
        """Código 39145 con cantidad 1 en SOAT Urgencias NO genera error."""
        ws = wb_soat_urgencias_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="39145")
        ws.cell(row=2, column=3, value="PROC A")
        ws.cell(row=2, column=4, value=1)
        ws.cell(row=2, column=5, value="Urgencias")
        ws.cell(row=2, column=6, value="SOAT")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
            "tarifario": 5,
        }
        result = detect_cantidades_soat_urgencias(ws, indices)

        assert len(result) == 0

    def test_tarifario_no_soat_no_genera_error(
        self, wb_soat_urgencias_headers: Workbook
    ) -> None:
        """Si Tarifario no es SOAT, no se valida."""
        ws = wb_soat_urgencias_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="39145")
        ws.cell(row=2, column=3, value="PROC A")
        ws.cell(row=2, column=4, value=2)
        ws.cell(row=2, column=5, value="Urgencias")
        ws.cell(row=2, column=6, value="ISS")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
            "tarifario": 5,
        }
        result = detect_cantidades_soat_urgencias(ws, indices)

        assert len(result) == 0

    def test_codigo_no_soat_no_genera_error(
        self, wb_soat_urgencias_headers: Workbook
    ) -> None:
        """Código no en CODIGOS_SOAT_CANTIDAD_OBLIGATORIA no se valida."""
        ws = wb_soat_urgencias_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="890101")
        ws.cell(row=2, column=3, value="PROC A")
        ws.cell(row=2, column=4, value=2)
        ws.cell(row=2, column=5, value="Urgencias")
        ws.cell(row=2, column=6, value="SOAT")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
            "tarifario": 5,
        }
        result = detect_cantidades_soat_urgencias(ws, indices)

        assert len(result) == 0

    def test_sin_indices_retorna_vacio(
        self, wb_soat_urgencias_headers: Workbook
    ) -> None:
        """Si faltan índices necesarios, retorna lista vacía."""
        ws = wb_soat_urgencias_headers.active
        result = detect_cantidades_soat_urgencias(ws, {})
        assert result == []
