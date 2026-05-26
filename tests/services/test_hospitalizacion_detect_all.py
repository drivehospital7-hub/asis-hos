"""Tests for app/services/hospitalizacion/detect_all.py.

Strict TDD: tests written BEFORE implementation.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.hospitalizacion.detect_all import detect_all_problems_hospitalizacion


@pytest.fixture
def workbook_minimal() -> Workbook:
    """Crea un workbook con headers mínimos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    return wb


class TestDetectAllProblemsHospitalizacion:
    """Tests para detect_all_problems_hospitalizacion."""

    def _run(self, ws, indices):
        """Helper que corre el detector y retorna solo el dict resultado."""
        result, _ = detect_all_problems_hospitalizacion(ws, indices)
        return result

    def test_retorna_dict_con_key_problemas(self, workbook_minimal: Workbook) -> None:
        """Resultado debe contener key 'problemas'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert "problemas" in result
        assert isinstance(result["problemas"], dict)

    def test_retorna_dict_con_key_totales(self, workbook_minimal: Workbook) -> None:
        """Resultado debe contener key 'totales'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert "totales" in result
        assert isinstance(result["totales"], dict)

    def test_retorna_area_hospitalizacion(self, workbook_minimal: Workbook) -> None:
        """Resultado debe contener 'area' = 'hospitalizacion'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert result.get("area") == "hospitalizacion"

    def test_resultado_incluye_normalizados(self, workbook_minimal: Workbook) -> None:
        """Resultado debe incluir 'normalizados' en problemas."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert "normalizados" in result["problemas"]
        assert isinstance(result["problemas"]["normalizados"], list)

    def test_resultado_incluye_missing_columns(self, workbook_minimal: Workbook) -> None:
        """Resultado debe contener 'missing_columns'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert "missing_columns" in result
