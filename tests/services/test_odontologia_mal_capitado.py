"""Tests para app/services/urgencias/mal_capitado.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.urgencias.mal_capitado import detect_mal_capitado


@pytest.fixture
def workbook_with_mal_capitado_headers() -> Workbook:
    """Crea un workbook con headers para mal capitado."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código")
    ws.cell(row=1, column=3, value="Procedimiento")
    ws.cell(row=1, column=4, value="IDE Contrato")
    ws.cell(row=1, column=5, value="Cód Entidad Cobrar")
    return wb


class TestDetectMalCapitado:
    """Tests para detect_mal_capitado."""

    def test_codigo_g03xb01_con_prefijo_fev_no_genera_error(
        self, workbook_with_mal_capitado_headers: Workbook
    ) -> None:
        """G03XB01 con prefijo FEV en factura es válido."""
        ws = workbook_with_mal_capitado_headers.active
        ws.cell(row=2, column=1, value="FEV12345")
        ws.cell(row=2, column=2, value="G03XB01")
        ws.cell(row=2, column=3, value="Procedimiento G03")

        indices = {"numero_factura": 0, "codigo": 1,
                   "procedimiento": 2, "ide_contrato": 3,
                   "codigo_entidad_cobrar": 4}
        result = detect_mal_capitado(ws, indices)

        # Solo filtro FEV - debe pasar
        mal_capitados = [r for r in result if r.get("codigo") in ("G03XB01", "A02BB01")]
        assert len(mal_capitados) == 0

    def test_codigo_a02bb01_sin_prefijo_fev_genera_error(
        self, workbook_with_mal_capitado_headers: Workbook
    ) -> None:
        """A02BB01 sin prefijo FEV genera error."""
        ws = workbook_with_mal_capitado_headers.active
        ws.cell(row=2, column=1, value="CAP12345")
        ws.cell(row=2, column=2, value="A02BB01")
        ws.cell(row=2, column=3, value="Procedimiento A02")

        indices = {"numero_factura": 0, "codigo": 1,
                   "procedimiento": 2, "ide_contrato": 3,
                   "codigo_entidad_cobrar": 4}
        result = detect_mal_capitado(ws, indices)

        # Buscar mal capitado (FEV)
        fev_errors = [r for r in result if "FEV" in r.get("observacion", "")]
        assert len(fev_errors) == 1
        assert fev_errors[0]["factura"] == "CAP12345"
        assert fev_errors[0]["codigo"] == "A02BB01"

    def test_codigo_comun_no_genera_error(
        self, workbook_with_mal_capitado_headers: Workbook
    ) -> None:
        """Códigos no mal capitado no deben generar error."""
        ws = workbook_with_mal_capitado_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="890101")
        ws.cell(row=2, column=3, value="Consulta General")

        indices = {"numero_factura": 0, "codigo": 1,
                   "procedimiento": 2, "ide_contrato": 3,
                   "codigo_entidad_cobrar": 4}
        result = detect_mal_capitado(ws, indices)

        fev_errors = [r for r in result if "FEV" in r.get("observacion", "")]
        assert len(fev_errors) == 0

    def test_sin_indices_retorna_vacio(
        self, workbook_with_mal_capitado_headers: Workbook
    ) -> None:
        """Sin indices necesarios, retorna lista vacía."""
        ws = workbook_with_mal_capitado_headers.active
        result = detect_mal_capitado(ws, {})
        assert result == []

    def test_factura_con_prefijo_cap_requiere_ess118(
        self, workbook_with_mal_capitado_headers: Workbook
    ) -> None:
        """CAP-xxx requiere Cód Entidad Cobrar = ESS118."""
        ws = workbook_with_mal_capitado_headers.active
        ws.cell(row=2, column=1, value="CAP-001")
        ws.cell(row=2, column=2, value="890101")
        ws.cell(row=2, column=3, value="Consulta")
        ws.cell(row=2, column=4, value="123")
        ws.cell(row=2, column=5, value="OTRA_ENTIDAD")

        indices = {"numero_factura": 0, "codigo": 1,
                   "procedimiento": 2, "ide_contrato": 3,
                   "codigo_entidad_cobrar": 4}
        result = detect_mal_capitado(ws, indices)

        cap_errors = [r for r in result if "CAP" in r.get("observacion", "")]
        assert len(cap_errors) == 1
        assert "ESS118" in cap_errors[0]["observacion"]
