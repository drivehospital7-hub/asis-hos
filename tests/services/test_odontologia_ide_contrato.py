"""Tests para app/services/odontologia/ide_contrato.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.odontologia.ide_contrato import detect_ide_contrato_odontologia


@pytest.fixture
def workbook_with_ide_contrato_headers() -> Workbook:
    """Crea un workbook con headers para IDE Contrato."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Cód Entidad Cobrar")
    ws.cell(row=1, column=3, value="Cód. Equivalente CUPS")
    ws.cell(row=1, column=4, value="IDE Contrato")
    return wb


class TestDetectIdeContratoOdontologia:
    """Tests para detect_ide_contrato_odontologia."""

    def test_ide_correcto_ess118_pyp_no_genera_error(
        self, workbook_with_ide_contrato_headers: Workbook
    ) -> None:
        """ESS118 + PyP con IDE 970 es válido."""
        ws = workbook_with_ide_contrato_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="ESS118")
        ws.cell(row=2, column=3, value="997002")  # PyP
        ws.cell(row=2, column=4, value="970")

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_odontologia(ws, indices)

        assert len(result) == 0

    def test_ide_incorrecto_ess118_pyp_genera_error(
        self, workbook_with_ide_contrato_headers: Workbook
    ) -> None:
        """ESS118 + PyP con IDE 969 es inválido."""
        ws = workbook_with_ide_contrato_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="ESS118")
        ws.cell(row=2, column=3, value="997002")  # PyP
        ws.cell(row=2, column=4, value="969")  # Debería ser 970 o 974

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_odontologia(ws, indices)

        assert len(result) == 1
        assert result[0]["cod_entidad"] == "ESS118"
        assert "970" in result[0]["ide_deberia"] or "974" in result[0]["ide_deberia"]

    def test_ess118_no_pyp_ide_969_es_valido(
        self, workbook_with_ide_contrato_headers: Workbook
    ) -> None:
        """ESS118 + NO PyP con IDE 969 es válido."""
        ws = workbook_with_ide_contrato_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="ESS118")
        ws.cell(row=2, column=3, value="890101")  # No PyP
        ws.cell(row=2, column=4, value="969")

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_odontologia(ws, indices)

        assert len(result) == 0

    def test_entidad_sin_regla_no_genera_error(
        self, workbook_with_ide_contrato_headers: Workbook
    ) -> None:
        """Entidad sin regla definida no debe generar error."""
        ws = workbook_with_ide_contrato_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="DESCONOCIDA")
        ws.cell(row=2, column=3, value="890101")
        ws.cell(row=2, column=4, value="999")

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_odontologia(ws, indices)

        assert len(result) == 0

    def test_sin_indices_retorna_vacio(
        self, workbook_with_ide_contrato_headers: Workbook
    ) -> None:
        """Si faltan índices, retorna lista vacía."""
        ws = workbook_with_ide_contrato_headers.active
        result = detect_ide_contrato_odontologia(ws, {})
        assert result == []

    def test_essc18_pyp_ide_975_es_valido(
        self, workbook_with_ide_contrato_headers: Workbook
    ) -> None:
        """ESSC18 + PyP con IDE 975 es válido."""
        ws = workbook_with_ide_contrato_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="ESSC18")
        ws.cell(row=2, column=3, value="997002")  # PyP
        ws.cell(row=2, column=4, value="975")

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_odontologia(ws, indices)

        assert len(result) == 0


class TestDedupIdeContratoOdontologia:
    """Tests para deduplicación de errores de contrato por factura (R1)."""

    def test_misma_factura_tres_filas_un_error(
        self, workbook_with_ide_contrato_headers: Workbook
    ) -> None:
        """Misma factura con 3 filas violando la regla → exactamente 1 error.

        Spec R1 Happy path: 12 rows all violating → exactly 1 contract error.
        """
        ws = workbook_with_ide_contrato_headers.active
        for row in [2, 3, 4]:
            ws.cell(row=row, column=1, value="FAC-001")
            ws.cell(row=row, column=2, value="ESS118")
            ws.cell(row=row, column=3, value="997002")  # PyP
            ws.cell(row=row, column=4, value="969")  # Inválido

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_odontologia(ws, indices)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"

    def test_dos_facturas_distintas_dos_errores(
        self, workbook_with_ide_contrato_headers: Workbook
    ) -> None:
        """2 facturas distintas con errores → 2 errores (1 por factura).

        Spec R1 Mixed invoices: 2 invoices, 24 rows → exactly 2 errors.
        """
        ws = workbook_with_ide_contrato_headers.active
        # FAC-001: 2 filas violando
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="ESS118")
        ws.cell(row=2, column=3, value="997002")
        ws.cell(row=2, column=4, value="969")

        ws.cell(row=3, column=1, value="FAC-001")
        ws.cell(row=3, column=2, value="ESS118")
        ws.cell(row=3, column=3, value="997002")
        ws.cell(row=3, column=4, value="969")

        # FAC-002: 2 filas violando
        ws.cell(row=4, column=1, value="FAC-002")
        ws.cell(row=4, column=2, value="ESS118")
        ws.cell(row=4, column=3, value="997002")
        ws.cell(row=4, column=4, value="969")

        ws.cell(row=5, column=1, value="FAC-002")
        ws.cell(row=5, column=2, value="ESS118")
        ws.cell(row=5, column=3, value="997002")
        ws.cell(row=5, column=4, value="969")

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_odontologia(ws, indices)

        assert len(result) == 2
        facturas = {r["factura"] for r in result}
        assert facturas == {"FAC-001", "FAC-002"}

    def test_sin_violaciones_cero_errores(
        self, workbook_with_ide_contrato_headers: Workbook
    ) -> None:
        """Factura sin violaciones de contrato → 0 errores (regression).

        Spec R1 No violations: all rows correct → 0 errors.
        """
        ws = workbook_with_ide_contrato_headers.active
        for row in [2, 3, 4]:
            ws.cell(row=row, column=1, value="FAC-001")
            ws.cell(row=row, column=2, value="ESS118")
            ws.cell(row=row, column=3, value="997002")  # PyP
            ws.cell(row=row, column=4, value="970")  # Correcto

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_odontologia(ws, indices)

        assert len(result) == 0

    def test_factura_sin_error_no_contamina_otras(
        self, workbook_with_ide_contrato_headers: Workbook
    ) -> None:
        """Factura sin error + factura con error → solo 1 error.

        Triangulation: verify no cross-invoice contamination when one
        invoice has no violations and another does.
        """
        ws = workbook_with_ide_contrato_headers.active
        # FAC-001: 2 filas SIN error (IDE 970 correcto para ESS118+PyP)
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="ESS118")
        ws.cell(row=2, column=3, value="997002")
        ws.cell(row=2, column=4, value="970")

        ws.cell(row=3, column=1, value="FAC-001")
        ws.cell(row=3, column=2, value="ESS118")
        ws.cell(row=3, column=3, value="997002")
        ws.cell(row=3, column=4, value="970")

        # FAC-002: 2 filas con error
        ws.cell(row=4, column=1, value="FAC-002")
        ws.cell(row=4, column=2, value="ESS118")
        ws.cell(row=4, column=3, value="997002")
        ws.cell(row=4, column=4, value="969")

        ws.cell(row=5, column=1, value="FAC-002")
        ws.cell(row=5, column=2, value="ESS118")
        ws.cell(row=5, column=3, value="997002")
        ws.cell(row=5, column=4, value="969")

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_odontologia(ws, indices)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-002"

    def test_fila_sin_factura_se_ignora(
        self, workbook_with_ide_contrato_headers: Workbook
    ) -> None:
        """Fila sin número de factura se ignora, no crash (R3).

        Spec R3: Missing invoice → no error, no crash.
        """
        ws = workbook_with_ide_contrato_headers.active
        # Fila sin factura (None)
        ws.cell(row=2, column=1, value=None)
        ws.cell(row=2, column=2, value="ESS118")
        ws.cell(row=2, column=3, value="997002")
        ws.cell(row=2, column=4, value="969")

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_odontologia(ws, indices)

        assert len(result) == 0
