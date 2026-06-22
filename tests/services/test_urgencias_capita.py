"""Tests para app/services/urgencias/valida_capita.py — detect_capita_cups_invalidos."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.urgencias.valida_capita import detect_capita_cups_invalidos


@pytest.fixture
def workbook_with_capita_headers() -> Workbook:
    """Crea un workbook con headers para validación CAPITA."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código")
    ws.cell(row=1, column=3, value="Procedimiento")
    ws.cell(row=1, column=4, value="Código Tipo Procedimiento")
    ws.cell(row=1, column=5, value="Cód. Equivalente CUPS")
    return wb


CODIGO_EN_LISTADO = "890101"
CODIGO_NO_EN_LISTADO = "999999"


class TestDetectCapitaCupsInvalidos:
    """Tests para detect_capita_cups_invalidos, incluyendo fallback a codigo_equiv."""

    def test_cups_en_listado_sin_equiv_no_genera_error(
        self, workbook_with_capita_headers: Workbook
    ) -> None:
        """CUPS directamente en listado → 0 errores (regresión)."""
        ws = workbook_with_capita_headers.active
        ws.cell(row=2, column=1, value="CAP-001")
        ws.cell(row=2, column=2, value=CODIGO_EN_LISTADO)
        ws.cell(row=2, column=3, value="Consulta General")
        ws.cell(row=2, column=4, value="01")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "codigo_tipo_procedimiento": 3,
            "codigo_equiv": 4,
        }
        result = detect_capita_cups_invalidos(ws, indices)
        assert len(result) == 0

    def test_cups_no_listado_con_equiv_valido_no_genera_error(
        self, workbook_with_capita_headers: Workbook
    ) -> None:
        """CUPS no en listado + Cód. Equivalente CUPS en listado → 0 errores."""
        ws = workbook_with_capita_headers.active
        ws.cell(row=2, column=1, value="CAP-002")
        ws.cell(row=2, column=2, value=CODIGO_NO_EN_LISTADO)
        ws.cell(row=2, column=3, value="Procedimiento X")
        ws.cell(row=2, column=4, value="01")
        ws.cell(row=2, column=5, value=CODIGO_EN_LISTADO)

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "codigo_tipo_procedimiento": 3,
            "codigo_equiv": 4,
        }
        result = detect_capita_cups_invalidos(ws, indices)
        assert len(result) == 0

    def test_cups_no_listado_equiv_vacio_genera_error(
        self, workbook_with_capita_headers: Workbook
    ) -> None:
        """CUPS no en listado + Cód. Equivalente CUPS vacío → 1 error."""
        ws = workbook_with_capita_headers.active
        ws.cell(row=2, column=1, value="CAP-003")
        ws.cell(row=2, column=2, value=CODIGO_NO_EN_LISTADO)
        ws.cell(row=2, column=3, value="Procedimiento Y")
        ws.cell(row=2, column=4, value="01")
        ws.cell(row=2, column=5, value=None)

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "codigo_tipo_procedimiento": 3,
            "codigo_equiv": 4,
        }
        result = detect_capita_cups_invalidos(ws, indices)
        assert len(result) == 1
        assert result[0]["factura"] == "CAP-003"
        assert result[0]["codigo"] == CODIGO_NO_EN_LISTADO
        assert "CAPITA" in result[0]["observacion"]

    def test_cups_no_listado_equiv_no_listado_genera_error(
        self, workbook_with_capita_headers: Workbook
    ) -> None:
        """CUPS no en listado + Equivalente CUPS no en listado → 1 error."""
        ws = workbook_with_capita_headers.active
        ws.cell(row=2, column=1, value="CAP-004")
        ws.cell(row=2, column=2, value=CODIGO_NO_EN_LISTADO)
        ws.cell(row=2, column=3, value="Procedimiento Z")
        ws.cell(row=2, column=4, value="01")
        ws.cell(row=2, column=5, value="111111")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "codigo_tipo_procedimiento": 3,
            "codigo_equiv": 4,
        }
        result = detect_capita_cups_invalidos(ws, indices)
        assert len(result) == 1
        assert result[0]["factura"] == "CAP-004"

    def test_cups_no_listado_columna_equiv_ausente_genera_error(
        self, workbook_with_capita_headers: Workbook
    ) -> None:
        """CUPS no en listado + columna codigo_equiv ausente → 1 error."""
        ws = workbook_with_capita_headers.active
        ws.cell(row=2, column=1, value="CAP-005")
        ws.cell(row=2, column=2, value=CODIGO_NO_EN_LISTADO)
        ws.cell(row=2, column=3, value="Procedimiento W")
        ws.cell(row=2, column=4, value="01")

        # indices SIN codigo_equiv
        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "codigo_tipo_procedimiento": 3,
        }
        result = detect_capita_cups_invalidos(ws, indices)
        assert len(result) == 1
        assert result[0]["factura"] == "CAP-005"

    def test_tipo_proc_09_excluido_no_genera_error(
        self, workbook_with_capita_headers: Workbook
    ) -> None:
        """Tipo Procedimiento 09 excluido → 0 errores (regresión)."""
        ws = workbook_with_capita_headers.active
        ws.cell(row=2, column=1, value="CAP-006")
        ws.cell(row=2, column=2, value=CODIGO_NO_EN_LISTADO)
        ws.cell(row=2, column=3, value="Farmacia")
        ws.cell(row=2, column=4, value="09")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "codigo_tipo_procedimiento": 3,
            "codigo_equiv": 4,
        }
        result = detect_capita_cups_invalidos(ws, indices)
        assert len(result) == 0

    def test_normalizacion_equiv_con_espacios(
        self, workbook_with_capita_headers: Workbook
    ) -> None:
        """Equivalente con espacios se normaliza via strip().upper()."""
        ws = workbook_with_capita_headers.active
        ws.cell(row=2, column=1, value="CAP-007")
        ws.cell(row=2, column=2, value=CODIGO_NO_EN_LISTADO)
        ws.cell(row=2, column=3, value="Procedimiento V")
        ws.cell(row=2, column=4, value="01")
        ws.cell(row=2, column=5, value="  890101  ")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "codigo_tipo_procedimiento": 3,
            "codigo_equiv": 4,
        }
        result = detect_capita_cups_invalidos(ws, indices)
        assert len(result) == 0

    def test_cups_no_en_listado_factura_cap_sin_equiv(
        self, workbook_with_capita_headers: Workbook
    ) -> None:
        """CUPS no listado, sin codigo_equiv, con codigo_tipo_procedimiento normal."""
        ws = workbook_with_capita_headers.active
        ws.cell(row=2, column=1, value="CAP-008")
        ws.cell(row=2, column=2, value=CODIGO_NO_EN_LISTADO)
        ws.cell(row=2, column=3, value="Procedimiento U")
        ws.cell(row=2, column=4, value="01")
        # codigo_equiv en blanco
        ws.cell(row=2, column=5, value="")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "codigo_tipo_procedimiento": 3,
            "codigo_equiv": 4,
        }
        result = detect_capita_cups_invalidos(ws, indices)
        assert len(result) == 1
        assert result[0]["factura"] == "CAP-008"
