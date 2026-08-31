"""Integration tests for /api/examenes, /api/listado and /api/examenes/export
(EX-2/EX-3/EX-19/EX-34).

Covers: 401 anonymous JSON, read-only save/edit POST → 200 (flipped gate),
403 read-only DELETE/export writes with file unchanged, envelope shape
(data.examenes / data.listado nesting), 400 non-array bodies, POST→GET
round-trips, CAS (409 on stale base_hash) for POST and DELETE, and the xlsx
export route (401 / read-200 / 400 empty / filename labels).
"""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.constants.examenes import CSV_HEADERS
from app.utils import examenes_store

CATALOG = [
    {"cod": "903859", "nom": "Potasio En Suero U Otros Fluidos", "neps": "X", "mall": "X", "emss": "X"},
    {"cod": "904921", "nom": "Tiroxina Libre", "neps": "AUTH", "mall": "X", "emss": ""},
]

PREFACTURAS = [
    {
        "id": "pf-1720000000000-abc",
        "paciente": "Paciente Uno",
        "cedula": "1001",
        "facturador": "ANGIE ARIAS",
        "hora": "01/01/2026 08:00",
        "items": [
            {
                "cod": "903859",
                "nom": "Potasio En Suero U Otros Fluidos",
                "neps": "X",
                "mall": "X",
                "emss": "X",
                "cantidad": 2,
            }
        ],
    },
    {
        "id": "pf-1720000000001-xyz",
        "paciente": "Paciente Dos",
        "cedula": "1002",
        "facturador": "CATALEYA TAPIA",
        "hora": "02/01/2026 09:30",
        "items": [
            # legacy 5-field item — no cantidad key, must stay valid (EX-21/EX-28)
            {"cod": "904921", "nom": "Tiroxina Libre", "neps": "AUTH", "mall": "X", "emss": ""}
        ],
    },
]


@pytest.fixture(autouse=True)
def _tmp_data_dir(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """Redirect store IO to tmp so tests never touch app/data."""
    monkeypatch.setattr(examenes_store, "DATA_DIR", tmp_path)


def _authenticate(client, permisos: list[str], username: str = "test") -> None:
    """Establece sesión autenticada con los permisos dados."""
    with client.session_transaction() as sess:
        sess["ce_authenticated"] = True
        sess["username"] = username
        sess["permisos"] = permisos


def _write_listado(tmp_path: Path, data: list) -> None:
    (tmp_path / "listado.json").write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


class TestAnonymous401:
    """EX-1/EX-2: no session → 401 error envelope for JSON/XHR requests."""

    def test_anonymous_get_catalog_401(self, app_client) -> None:
        """GET /api/examenes without session → 401 {status:error, errors:['No autenticado']}."""
        response = app_client.get(
            "/api/examenes", headers={"X-Requested-With": "XMLHttpRequest"}
        )

        assert response.status_code == 401
        body = response.get_json()
        assert body["status"] == "error"
        assert body["data"] == {}
        assert body["errors"] == ["No autenticado"]

    def test_anonymous_post_listado_401(self, app_client) -> None:
        """POST /api/listado without session (JSON) → 401 envelope."""
        response = app_client.post("/api/listado", json=PREFACTURAS)

        assert response.status_code == 401
        body = response.get_json()
        assert body["status"] == "error"
        assert body["data"] == {}
        assert body["errors"] == ["No autenticado"]


class TestReadOnlyGating:
    """EX-2: read-only user can GET; writes → 403 with file unchanged."""

    def test_read_only_get_catalog_envelope(self, app_client, tmp_path: Path) -> None:
        """GET /api/examenes → 200, data.examenes holds the catalog array."""
        (tmp_path / "examenes.json").write_text(
            json.dumps(CATALOG, ensure_ascii=False), encoding="utf-8"
        )
        _authenticate(app_client, ["examenes"])

        response = app_client.get("/api/examenes")

        assert response.status_code == 200
        body = response.get_json()
        assert body["status"] == "success"
        assert body["errors"] == []
        assert body["data"]["examenes"] == CATALOG

    def test_read_only_get_listado_envelope(self, app_client, tmp_path: Path) -> None:
        """GET /api/listado → 200, data.listado holds the array."""
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes"])

        response = app_client.get("/api/listado")

        assert response.status_code == 200
        body = response.get_json()
        assert body["status"] == "success"
        assert body["data"]["listado"] == PREFACTURAS

    def test_read_only_post_listado_200_persists(self, app_client, tmp_path: Path) -> None:
        """POST /api/listado by read-only → 200 + persisted (save/edit read-allowed, EX-2)."""
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes"])

        response = app_client.post("/api/listado", json=[{"id": "pf-evil"}])

        assert response.status_code == 200
        assert response.get_json() == {"status": "success", "data": {}, "errors": []}
        on_disk = json.loads((tmp_path / "listado.json").read_text(encoding="utf-8"))
        assert on_disk == [{"id": "pf-evil"}]

    def test_read_only_post_catalog_403(self, app_client) -> None:
        """POST /api/examenes by read-only → 403 envelope."""
        _authenticate(app_client, ["examenes"])

        response = app_client.post("/api/examenes", json=CATALOG)

        assert response.status_code == 403
        assert response.get_json()["status"] == "error"


class TestWriteOps:
    """EX-2/EX-3: write user persists arrays; returns success envelope."""

    def test_write_post_listado_persists(self, app_client, tmp_path: Path) -> None:
        """POST /api/listado (array) → 200 success envelope + persisted file."""
        _authenticate(app_client, ["examenes:write"])

        response = app_client.post("/api/listado", json=PREFACTURAS)

        assert response.status_code == 200
        body = response.get_json()
        assert body == {"status": "success", "data": {}, "errors": []}
        assert json.loads((tmp_path / "listado.json").read_text(encoding="utf-8")) == PREFACTURAS

    def test_write_post_catalog_persists(self, app_client, tmp_path: Path) -> None:
        """POST /api/examenes (array) → 200 success envelope + persisted file."""
        _authenticate(app_client, ["examenes:write"])

        response = app_client.post("/api/examenes", json=CATALOG)

        assert response.status_code == 200
        assert response.get_json() == {"status": "success", "data": {}, "errors": []}
        assert json.loads((tmp_path / "examenes.json").read_text(encoding="utf-8")) == CATALOG


class TestRoundTrip:
    """EX-3 round-trip: GET returns exactly what was posted (fields+order)."""

    def test_listado_round_trip(self, app_client, tmp_path: Path) -> None:
        _authenticate(app_client, ["examenes:write"])
        assert app_client.post("/api/listado", json=PREFACTURAS).status_code == 200

        response = app_client.get("/api/listado")

        assert response.status_code == 200
        assert response.get_json()["data"]["listado"] == PREFACTURAS

    def test_cantidad_persists_round_trip_legacy_valid(self, app_client, tmp_path: Path) -> None:
        """EX-28: item cantidad persists POST→GET; legacy 5-field items valid."""
        _authenticate(app_client, ["examenes:write"])
        assert app_client.post("/api/listado", json=PREFACTURAS).status_code == 200

        response = app_client.get("/api/listado")

        assert response.status_code == 200
        listado = response.get_json()["data"]["listado"]
        assert listado[0]["items"][0]["cantidad"] == 2
        # legacy item is returned verbatim: cantidad key stays ABSENT
        assert "cantidad" not in listado[1]["items"][0]

    def test_catalog_round_trip(self, app_client, tmp_path: Path) -> None:
        _authenticate(app_client, ["examenes:write"])
        assert app_client.post("/api/examenes", json=CATALOG).status_code == 200

        response = app_client.get("/api/examenes")

        assert response.status_code == 200
        assert response.get_json()["data"]["examenes"] == CATALOG


class TestInvalidBody400:
    """EX-3 deviation: non-array POST bodies → 400 envelope, file unchanged."""

    def test_listado_object_body_400(self, app_client, tmp_path: Path) -> None:
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes:write"])

        response = app_client.post("/api/listado", json={"id": "not-an-array"})

        assert response.status_code == 400
        body = response.get_json()
        assert body["status"] == "error"
        assert body["data"] == {}
        assert isinstance(body["errors"], list)
        assert body["errors"]
        assert json.loads((tmp_path / "listado.json").read_text(encoding="utf-8")) == PREFACTURAS

    def test_catalog_object_body_400(self, app_client, tmp_path: Path) -> None:
        _authenticate(app_client, ["examenes:write"])

        response = app_client.post("/api/examenes", json={"cod": "x"})

        assert response.status_code == 400
        assert response.get_json()["status"] == "error"
        assert not (tmp_path / "examenes.json").exists()

    def test_listado_malformed_json_400(self, app_client, tmp_path: Path) -> None:
        """Unparseable JSON body → 400 envelope, file unchanged."""
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes:write"])

        response = app_client.post(
            "/api/listado",
            data="{not json",
            content_type="application/json",
        )

        assert response.status_code == 400
        assert response.get_json()["status"] == "error"
        assert json.loads((tmp_path / "listado.json").read_text(encoding="utf-8")) == PREFACTURAS


EXTRA = {
    "id": "pf-99",
    "paciente": "Paciente Nuevo",
    "cedula": "1099",
    "facturador": "ANGIE ARIAS",
    "hora": "03/01/2026 10:00",
    "items": [
        {"cod": "903859", "nom": "Potasio En Suero U Otros Fluidos", "neps": "X", "mall": "X", "emss": "X"}
    ],
}


class TestOptimisticConcurrency:
    """R4-001: base_hash opcional → 409 sin escribir si el estado cambió."""

    def test_sequential_same_base_second_write_409(self, app_client, tmp_path: Path) -> None:
        """Dos POSTs con el mismo base_hash: el segundo → 409 sin pisar el primero."""
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes:write"])
        base = examenes_store.file_hash("listado.json")
        primera = PREFACTURAS + [EXTRA]
        segunda = PREFACTURAS + [EXTRA, {**EXTRA, "id": "pf-100"}]

        assert (
            app_client.post("/api/listado", json={"data": primera, "base_hash": base}).status_code
            == 200
        )
        res = app_client.post("/api/listado", json={"data": segunda, "base_hash": base})

        assert res.status_code == 409
        assert json.loads((tmp_path / "listado.json").read_text(encoding="utf-8")) == primera

    def test_matching_base_hash_200_and_persisted(self, app_client, tmp_path: Path) -> None:
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes:write"])
        base = examenes_store.file_hash("listado.json")
        nueva = PREFACTURAS + [EXTRA]

        res = app_client.post("/api/listado", json={"data": nueva, "base_hash": base})

        assert res.status_code == 200
        assert res.get_json() == {"status": "success", "data": {}, "errors": []}
        assert json.loads((tmp_path / "listado.json").read_text(encoding="utf-8")) == nueva

    def test_stale_base_hash_409_file_unchanged(self, app_client, tmp_path: Path) -> None:
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes:write"])

        res = app_client.post("/api/listado", json={"data": [EXTRA], "base_hash": "0" * 64})

        assert res.status_code == 409
        body = res.get_json()
        assert body["status"] == "error" and body["data"] == {} and body["errors"]
        assert json.loads((tmp_path / "listado.json").read_text(encoding="utf-8")) == PREFACTURAS

    def test_stale_base_hash_catalog_409(self, app_client, tmp_path: Path) -> None:
        (tmp_path / "examenes.json").write_text(json.dumps(CATALOG, ensure_ascii=False), encoding="utf-8")
        _authenticate(app_client, ["examenes:write"])

        res = app_client.post("/api/examenes", json={"data": CATALOG, "base_hash": "0" * 64})

        assert res.status_code == 409 and res.get_json()["status"] == "error"
        assert json.loads((tmp_path / "examenes.json").read_text(encoding="utf-8")) == CATALOG

    def test_absent_base_hash_legacy_replace(self, app_client, tmp_path: Path) -> None:
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes:write"])

        res = app_client.post("/api/listado", json=[EXTRA])

        assert res.status_code == 200
        assert res.get_json()["status"] == "success"
        assert json.loads((tmp_path / "listado.json").read_text(encoding="utf-8")) == [EXTRA]


class TestDeleteListado:
    """NEW DELETE /api/listado/<prefactura_id> (EX-2): write-gated, CAS, 404/409."""

    FIRST_ID = "pf-1720000000000-abc"
    SECOND_ID = "pf-1720000000001-xyz"

    def test_anonymous_delete_401(self, app_client, tmp_path: Path) -> None:
        """DELETE without session → 401 error envelope."""
        _write_listado(tmp_path, PREFACTURAS)

        response = app_client.delete(f"/api/listado/{self.FIRST_ID}", json={"base_hash": "x"})

        assert response.status_code == 401
        body = response.get_json()
        assert body["status"] == "error"
        assert body["data"] == {}
        assert body["errors"] == ["No autenticado"]

    def test_read_only_delete_403_file_unchanged(self, app_client, tmp_path: Path) -> None:
        """Whole-record delete by read-only → 403 AND the file is untouched (EX-2)."""
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes"])

        response = app_client.delete(
            f"/api/listado/{self.FIRST_ID}",
            headers={"X-Requested-With": "XMLHttpRequest"},
        )

        assert response.status_code == 403
        assert response.get_json()["status"] == "error"
        assert json.loads((tmp_path / "listado.json").read_text(encoding="utf-8")) == PREFACTURAS

    def test_write_delete_200_record_removed(self, app_client, tmp_path: Path) -> None:
        """Write user deletes one prefactura → 200, record gone, rest persisted."""
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes:write"])
        base = examenes_store.file_hash("listado.json")

        response = app_client.delete(
            f"/api/listado/{self.FIRST_ID}", json={"base_hash": base}
        )

        assert response.status_code == 200
        assert response.get_json() == {"status": "success", "data": {}, "errors": []}
        on_disk = json.loads((tmp_path / "listado.json").read_text(encoding="utf-8"))
        assert [p["id"] for p in on_disk] == [self.SECOND_ID]

    def test_write_delete_unknown_id_404_file_unchanged(self, app_client, tmp_path: Path) -> None:
        """Unknown prefactura id → 404 envelope, file untouched."""
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes:write"])

        response = app_client.delete("/api/listado/pf-unknown")

        assert response.status_code == 404
        body = response.get_json()
        assert body["status"] == "error"
        assert body["errors"] == ["Prefactura no encontrada"]
        assert json.loads((tmp_path / "listado.json").read_text(encoding="utf-8")) == PREFACTURAS

    def test_write_delete_stale_base_hash_409_file_unchanged(self, app_client, tmp_path: Path) -> None:
        """Stale base_hash → 409 sin escribir (CAS, R4-001)."""
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes:write"])

        response = app_client.delete(
            f"/api/listado/{self.FIRST_ID}", json={"base_hash": "0" * 64}
        )

        assert response.status_code == 409
        assert response.get_json()["status"] == "error"
        assert json.loads((tmp_path / "listado.json").read_text(encoding="utf-8")) == PREFACTURAS

    def test_write_delete_absent_base_hash_legacy_200(self, app_client, tmp_path: Path) -> None:
        """Sin base_hash → reemplazo legacy (misma semántica que POST)."""
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes:write"])

        response = app_client.delete(f"/api/listado/{self.FIRST_ID}")

        assert response.status_code == 200
        assert [p["id"] for p in json.loads((tmp_path / "listado.json").read_text(encoding="utf-8"))] == [
            self.SECOND_ID
        ]


class TestExportRoute:
    """NEW GET /api/examenes/export (EX-34): read-gated, styled xlsx, 400 empty."""

    def test_anonymous_export_401(self, app_client, tmp_path: Path) -> None:
        """Export without session → 401 error envelope."""
        _write_listado(tmp_path, PREFACTURAS)

        response = app_client.get(
            "/api/examenes/export",
            headers={"X-Requested-With": "XMLHttpRequest"},
        )

        assert response.status_code == 401
        body = response.get_json()
        assert body["status"] == "error"
        assert body["data"] == {}
        assert body["errors"] == ["No autenticado"]

    def test_read_only_export_200_styled_xlsx(self, app_client, tmp_path: Path) -> None:
        """Read-only export → 200 xlsx with control-novedades styling (EX-34)."""
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes"])

        response = app_client.get("/api/examenes/export")

        assert response.status_code == 200
        assert response.content_type.startswith("application/vnd.openxmlformats")
        wb = load_workbook(BytesIO(response.data))
        ws = wb.active
        assert ws.title == "Listado"
        assert [c.value for c in ws[1]] == CSV_HEADERS
        assert ws["A1"].fill.fgColor.rgb == "001B5E20"
        assert ws["A1"].font.bold is True
        assert ws.freeze_panes == "A2"
        pacientes = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        # EX-33: date-desc → Paciente Dos (02/01) antes que Paciente Uno (01/01).
        assert pacientes == ["Paciente Dos", "Paciente Uno"]
        # Exports never mutate listado.json.
        assert json.loads((tmp_path / "listado.json").read_text(encoding="utf-8")) == PREFACTURAS

    def test_export_filename_labels(self, app_client, tmp_path: Path) -> None:
        """Download name: {from}_{to}, {from}_hasta o Todos_los_meses (EX-34)."""
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes"])

        ranged = app_client.get("/api/examenes/export?from=2026-01-01&to=2026-01-31")
        assert "Listado_Lab_HospitalOrito_2026-01-01_2026-01-31.xlsx" in ranged.headers["Content-Disposition"]

        from_only = app_client.get("/api/examenes/export?from=2026-01-01")
        assert "Listado_Lab_HospitalOrito_2026-01-01_hasta.xlsx" in from_only.headers["Content-Disposition"]

        all_meses = app_client.get("/api/examenes/export")
        assert "Listado_Lab_HospitalOrito_Todos_los_meses.xlsx" in all_meses.headers["Content-Disposition"]

    def test_export_empty_filtered_set_400(self, app_client, tmp_path: Path) -> None:
        """Rango sin registros → 400 envelope, sin archivo (EX-34)."""
        _write_listado(tmp_path, PREFACTURAS)
        _authenticate(app_client, ["examenes"])

        response = app_client.get("/api/examenes/export?from=2025-01-01&to=2025-01-31")

        assert response.status_code == 400
        body = response.get_json()
        assert body["status"] == "error"
        assert body["data"] == {}
        assert body["errors"] == ["No hay datos para exportar"]

    def test_export_q_ignores_range(self, app_client, tmp_path: Path) -> None:
        """Búsqueda global fuera del rango activo incluye el registro (D1)."""
        listado = PREFACTURAS + [
            {
                "id": "pf-out",
                "paciente": "Beto",
                "cedula": "1003",
                "facturador": "ANGIE ARIAS",
                "hora": "20/08/2026 08:00",
                "items": [{"cod": "903810", "nom": "Calcio", "neps": "", "mall": "", "emss": ""}],
            }
        ]
        _write_listado(tmp_path, listado)
        _authenticate(app_client, ["examenes"])

        response = app_client.get(
            "/api/examenes/export?from=2026-01-01&to=2026-01-31&q=beto"
        )

        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.data))
        ws = wb.active
        pacientes = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        assert pacientes == ["Beto"]