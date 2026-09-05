"""Tests for the Listado Excel export service (EX-14/EX-34) — control-novedades parity.

Unit layer: builds the workbook in memory and asserts styling + rows (one row
per item, sequential N°, CSV-flavor tc mapping, cantidad defaults), filter
parity vs ``composeListadoView`` (accent folding, sin-fecha range exclusion,
q-ignores-range, desc sort) and filename label variants.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.constants.examenes import CSV_HEADERS
from app.services.examenes_export import (
    build_listado_export_workbook,
    compose_listado_view,
    filename_export,
)


def _pf(pf_id: str, paciente: str, hora: str, items: list[dict], facturador: str = "ANGIE ARIAS") -> dict:
    return {
        "id": pf_id,
        "paciente": paciente,
        "cedula": "111",
        "facturador": facturador,
        "hora": hora,
        "items": [dict(it) for it in items],
    }


def _item(cod: str, nom: str, **flags) -> dict:
    item = {"cod": cod, "nom": nom, "neps": "", "mall": "", "emss": ""}
    item.update(flags)
    return item


def _open(buffer: BytesIO):
    wb = load_workbook(BytesIO(buffer.read()))
    return wb, wb.active


class TestBuildWorkbook:
    """EX-34 styling + one-row-per-item content contract."""

    def test_headers_estilos_y_freeze(self) -> None:
        buffer = build_listado_export_workbook([_pf("pf-1", "Juan Perez", "15/01/2026 08:30", [_item("903859", "Potasio")])])
        wb, ws = _open(buffer)

        assert ws.title == "Listado"
        assert [c.value for c in ws[1]] == CSV_HEADERS
        assert len(CSV_HEADERS) == 11
        assert ws["A1"].fill.fgColor.rgb == "001B5E20"
        assert ws["A1"].font.color.rgb == "00FFFFFF"
        assert ws["A1"].font.bold is True
        assert ws.freeze_panes == "A2"
        for col in "ABCDEFGHIJK":
            assert ws.column_dimensions[col].width == 20

    def test_una_fila_por_item_numero_secuencial(self) -> None:
        listado = [
            _pf(
                "pf-1",
                "Juan Perez",
                "15/01/2026 08:30",
                [
                    _item("903859", "Potasio", neps="X", mall="X", emss="X", cantidad=2),
                    _item("903016", "Ferritina", neps="AUTH", mall="AUTH", emss="AUTH"),
                ],
                facturador="Angie Chapuel",
            ),
            _pf(
                "pf-2",
                "Maria Lopez",
                "28/08/2026 09:00",
                [_item("906131", "Trypanosoma")],
                facturador="Cataleya Tapia",
            ),
        ]
        buffer = build_listado_export_workbook(listado)
        wb, ws = _open(buffer)

        assert ws.max_row == 4  # header + 2 items + 1 item
        assert [c.value for c in ws[2]] == [
            1, "Juan Perez", "111", "903859", "Potasio", 2, "SI", "SI", "SI", "Angie Chapuel", "15/01/2026 08:30",
        ]
        assert [c.value for c in ws[3]] == [
            2, "Juan Perez", "111", "903016", "Ferritina", 1, "AUTH", "AUTH", "AUTH", "Angie Chapuel", "15/01/2026 08:30",
        ]
        assert [c.value for c in ws[4]] == [
            3, "Maria Lopez", "111", "906131", "Trypanosoma", 1, None, None, None, "Cataleya Tapia", "28/08/2026 09:00",
        ]

    def test_cantidad_ausente_y_fraccionaria(self) -> None:
        listado = [
            _pf("pf-a", "A", "01/01/2026", [_item("903859", "Potasio")]),
            _pf("pf-b", "B", "01/01/2026", [_item("903016", "Ferritina", cantidad=2.9)]),
            _pf("pf-c", "C", "01/01/2026", [_item("903810", "Calcio", cantidad=0)]),
        ]
        buffer = build_listado_export_workbook(listado)
        wb, ws = _open(buffer)

        assert ws.cell(row=2, column=6).value == 1   # legacy item sin cantidad
        assert ws.cell(row=3, column=6).value == 2   # 2.9 → trunc 2
        assert ws.cell(row=4, column=6).value == 1   # 0 → clamp 1

    def test_estilos_alternados_y_bordes(self) -> None:
        listado = [
            _pf("pf-1", "A", "01/01/2026", [_item("903859", "Potasio")]),
            _pf("pf-2", "B", "02/01/2026", [_item("903016", "Ferritina")]),
        ]
        buffer = build_listado_export_workbook(listado)
        wb, ws = _open(buffer)

        assert ws["A2"].fill.fgColor.rgb == "00E8F5E9"
        assert ws["A3"].fill.fgColor.rgb == "00FFFFFF"
        assert ws["A2"].fill.fgColor.rgb != ws["A3"].fill.fgColor.rgb
        for cell in (ws["A2"], ws["K3"], ws["A1"]):
            for side in ("left", "right", "top", "bottom"):
                assert getattr(cell.border, side).style == "thin"

    def test_formula_injection_sanitized(self) -> None:
        """OWASP: strings starting with = + - @ are stored as text, not live formulas."""
        evil_paciente = "=WEBSERVICE(\"http://evil\")"
        evil_cedula = "@SUM(A1:A10)"
        evil_cod = "+2+3"
        evil_nom = "-2+3"
        evil_facturador = "=HYPERLINK(\"http://evil\",\"click\")"
        listado = [
            _pf(
                "pf-evil",
                evil_paciente,
                "01/01/2026 08:00",
                [_item(evil_cod, evil_nom)],
                facturador=evil_facturador,
            )
        ]
        # use evil cedula via direct dict (helper defaults to 111)
        listado[0]["cedula"] = evil_cedula
        buffer = build_listado_export_workbook(listado)
        wb, ws = _open(buffer)

        # header styling untouched
        assert ws["A1"].fill.fgColor.rgb == "001B5E20"
        row = [c for c in ws[2]]
        # columns: N, paciente(2), cedula(3), cod(4), nom(5), ..., facturador(10), hora(11)
        for col_idx in (2, 3, 4, 5, 10):
            cell = row[col_idx - 1]
            assert cell.data_type == "s", f"col {col_idx} should be text, got {cell.data_type}"
            assert str(cell.value).startswith("'"), f"col {col_idx} should be prefixed with single quote"
            assert cell.value.startswith("'= ") is False  # sanity: no stray space injection
        assert row[1].value == "'" + evil_paciente
        assert row[2].value == "'" + evil_cedula
        assert row[3].value == "'" + evil_cod
        assert row[4].value == "'" + evil_nom
        assert row[9].value == "'" + evil_facturador
        # hora is also sanitized when it starts with trigger char (edge)
        listado2 = [_pf("pf-2", "Normal", "=cmd|'/C calc'!A0", [_item("903859", "Potasio")])]
        listado2[0]["hora"] = "=2+2"
        buffer2 = build_listado_export_workbook(listado2)
        _, ws2 = _open(buffer2)
        assert ws2.cell(row=2, column=11).value == "'=2+2"
        assert ws2.cell(row=2, column=11).data_type == "s"


class TestComposeListadoView:
    """Filter parity vs the frontend composeListadoView (D4)."""

    def test_range_excluye_sin_fecha(self) -> None:
        listado = [
            _pf("dated", "Ana", "10/08/2026 09:00", [_item("903859", "Potasio")]),
            _pf("sin", "Dora", "n/a", [_item("906131", "Trypanosoma")]),
        ]
        view = compose_listado_view(listado, "2026-08-01", "2026-08-15", "")

        assert [p["id"] for p in view] == ["dated"]
        assert all(p["id"] != "sin" for p in view)

    def test_q_ignora_el_rango(self) -> None:
        listado = [
            _pf("in", "Ana", "10/08/2026 09:00", [_item("903859", "Potasio")]),
            _pf("out", "Beto", "20/08/2026 08:00", [_item("903016", "Ferritina")]),
            _pf("sin", "Dora", "n/a", [_item("906131", "Trypanosoma")]),
        ]
        # Record dated OUTSIDE the range still matches a global search (D1).
        view = compose_listado_view(listado, "2026-08-01", "2026-08-15", "beto")
        assert [p["id"] for p in view] == ["out"]

        # sin-fecha records match a global search even with a range active.
        view_sin = compose_listado_view(listado, "2026-08-01", "2026-08-15", "dora")
        assert [p["id"] for p in view_sin] == ["sin"]

    def test_desc_order_hora_y_sin_fecha_al_final(self) -> None:
        listado = [
            _pf("jul", "Caro", "30/07/2026 08:00", [_item("903810", "Calcio")]),
            _pf("aug1-m", "M", "01/08/2026 09:00", [_item("903016", "Ferritina")]),
            _pf("aug3", "N", "03/08/2026 10:00", [_item("903859", "Potasio")]),
            _pf("sin", "Dora", "n/a", [_item("906131", "Trypanosoma")]),
        ]
        view = compose_listado_view(listado, None, None, "")

        assert [p["id"] for p in view] == ["aug3", "aug1-m", "jul", "sin"]

    def test_hora_desc_dentro_del_dia_y_sin_hora_al_final(self) -> None:
        listado = [
            _pf("morning", "A", "03/08/2026 08:30", [_item("903859", "Potasio")]),
            _pf("afternoon", "B", "03/08/2026 14:05", [_item("903016", "Ferritina")]),
            _pf("untimed", "C", "03/08/2026", [_item("903810", "Calcio")]),
        ]
        view = compose_listado_view(listado, None, None, "")

        assert [p["id"] for p in view] == ["afternoon", "morning", "untimed"]

    def test_rango_vacio_incluye_sin_fecha(self) -> None:
        listado = [
            _pf("dated", "Ana", "10/08/2026 09:00", [_item("903859", "Potasio")]),
            _pf("sin", "Dora", "n/a", [_item("906131", "Trypanosoma")]),
        ]
        view = compose_listado_view(listado, None, None, "")

        assert [p["id"] for p in view] == ["dated", "sin"]

    def test_folding_de_acentos(self) -> None:
        listado = [
            _pf("nunez", "Álvaro Ñúñez", "10/08/2026 09:00", [_item("903859", "Potasio")]),
            _pf("otro", "Ana", "10/08/2026 09:00", [_item("903016", "Ferritina")]),
        ]
        view = compose_listado_view(listado, None, None, "alvaro nunez")

        assert [p["id"] for p in view] == ["nunez"]

    def test_listado_vacio(self) -> None:
        assert compose_listado_view([], None, None, "") == []


class TestFilenameExport:
    """EX-34 label variants: ${from}_${to}, ${from}_hasta, Todos_los_meses."""

    def test_rango_completo(self) -> None:
        assert filename_export("2026-08-01", "2026-08-31") == (
            "Listado_Lab_HospitalOrito_2026-08-01_2026-08-31.xlsx"
        )

    def test_solo_desde_hasta(self) -> None:
        assert filename_export("2026-08-01", None) == (
            "Listado_Lab_HospitalOrito_2026-08-01_hasta.xlsx"
        )

    def test_solo_hasta_etiqueta_con_guion_inicial(self) -> None:
        # Mirror del frontend: solo `to` produce una etiqueta que empieza con "_".
        assert filename_export(None, "2026-08-31") == (
            "Listado_Lab_HospitalOrito__2026-08-31.xlsx"
        )

    def test_sin_rango_todos_los_meses(self) -> None:
        assert filename_export(None, None) == (
            "Listado_Lab_HospitalOrito_Todos_los_meses.xlsx"
        )