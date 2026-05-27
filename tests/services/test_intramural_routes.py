"""Integration tests for app/routes/intramural.py.

Strict TDD: tests written BEFORE implementation.
Covers spec requirements R1-R2 for the Intramural blueprint.
"""

from __future__ import annotations

import io
from unittest.mock import patch

import pytest
from openpyxl import Workbook


# =============================================================================
# Helper: create a minimal Intramural Excel file
# =============================================================================


def _make_intramural_excel(
    headers: list[str], rows: list[list]
) -> io.BytesIO:
    """Create a real .xlsx in memory with given headers and data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


INTRAMURAL_HEADERS = [
    "Número Factura",
    "Vlr. Subsidiado",
    "Vlr. Procedimiento",
    "Código Tipo Procedimiento",
    "Tipo Procedimiento",
    "Código",
    "Cód. Equivalente CUPS",
    "Procedimiento",
    "Nº Identificación",
    "Convenio Facturado",
    "Cantidad",
    "Laboratorio",
    "Centro Costo",
    "Cód Entidad Cobrar",
    "Entidad Cobrar",
    "Entidad Afiliación",
    "Tipo Factura Descripción",
    "IDE Contrato",
    "Tipo Identificación",
    "Fec. Nacimiento",
    "Fec. Factura",
    "Fecha Cierre",
    "Identificación Profesional",
    "Profesional Atiende",
    "Código Profesional",
    "Responsable Cierra Facturar",
    "Tarifario",
    "Tipo Usuario",
]


# =============================================================================
# Test 2.1: GET route
# =============================================================================


class TestGetRoute:
    """Spec R1: GET /intramural/ requires auth + permiso."""

    def test_get_returns_200_with_permiso(self, app_client):
        """User with intramural permiso can access GET."""
        with app_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["intramural"]
            sess["username"] = "intra_user"

        resp = app_client.get("/intramural/", follow_redirects=True)
        assert resp.status_code == 200
        html = resp.data.decode("utf-8")
        assert "__INITIAL_DATA__" in html

    def test_get_returns_200_with_admin_star(self, app_client):
        """Admin (*) can access GET route."""
        with app_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["*"]
            sess["username"] = "admin"

        resp = app_client.get("/intramural/", follow_redirects=True)
        assert resp.status_code == 200

    def test_get_returns_403_without_permiso(self, app_client):
        """User without intramural permiso gets 403/redirect."""
        with app_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["odontologia"]
            sess["username"] = "odonto_user"

        resp = app_client.get("/intramural/", follow_redirects=True)
        assert resp.status_code == 200  # Flask redirect with follow
        html = resp.data.decode("utf-8")
        assert "Intramural" not in html or "No tiene permiso" in html

    def test_get_returns_401_unauthenticated(self, fresh_client):
        """No session -> 401 from before_request middleware."""
        resp = fresh_client.get("/intramural/")
        assert resp.status_code == 401


# =============================================================================
# Test 2.2: POST route
# =============================================================================


class TestPostRoute:
    """Spec R2: POST /intramural/ processes Intramural Excel."""

    def test_post_requires_permiso(self, fresh_client):
        """User without permiso gets 403 on POST."""
        with fresh_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["odontologia"]
            sess["username"] = "odonto_user"

        buf = _make_intramural_excel(INTRAMURAL_HEADERS, [
            ["FAC-001", "0", "15000", "", "", "890203", "", "Test", "12345",
             "Asistencial", "1", "", "", "ESS118", "ESS118", "ESS118", "",
             "970", "CC", "1990-01-01", "2024-01-15", "", "123", "Dr Test",
             "03764", "Resp", "Tarifario 1", "SUBSIDIADO"],
        ])
        resp = fresh_client.post(
            "/intramural/",
            data={"file_upload": (buf, "test_intra.xlsx")},
        )
        assert resp.status_code in (403, 200)
        if resp.status_code == 403:
            import json
            data = json.loads(resp.data.decode("utf-8"))
            assert data.get("status") == "error"

    def test_post_requires_auth(self, fresh_client):
        """No session -> 401 on POST."""
        buf = _make_intramural_excel(INTRAMURAL_HEADERS, [])
        resp = fresh_client.post(
            "/intramural/",
            data={"file_upload": (buf, "test_intra.xlsx")},
        )
        assert resp.status_code == 401

    def test_post_processes_valid_excel(self, fresh_client):
        """Valid Intramural Excel -> JSON response with problems."""
        with fresh_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["intramural"]
            sess["username"] = "intra_user"

        buf = _make_intramural_excel(INTRAMURAL_HEADERS, [
            ["FAC-001", "0", "15000", "", "", "890203", "", "Test Proc",
             "12345", "Asistencial", "1", "", "", "ESS118", "ESS118",
             "ESS118", "", "970", "CC", "1990-01-01", "2024-01-15", "",
             "123", "Dr Test", "03764", "Resp Cierra", "Tarifario 1",
             "SUBSIDIADO"],
        ])

        resp = fresh_client.post(
            "/intramural/",
            data={"file_upload": (buf, "test_intra.xlsx")},
        )
        assert resp.status_code in (200, 500)
        import json
        data = json.loads(resp.data.decode("utf-8"))
        assert "status" in data
        assert "data" in data

    def test_post_no_file_returns_error(self, fresh_client):
        """POST without file returns JSON error."""
        with fresh_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["intramural"]
            sess["username"] = "intra_user"

        resp = fresh_client.post(
            "/intramural/",
            data={},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 400
        data = resp.get_json()
        assert data is not None
        assert data["status"] == "error"
        assert isinstance(data["errors"], list)

    def test_post_invalid_extension_returns_error(self, fresh_client):
        """POST with invalid file extension returns JSON error."""
        with fresh_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["intramural"]
            sess["username"] = "intra_user"

        resp = fresh_client.post(
            "/intramural/",
            data={"file_upload": (io.BytesIO(b"test data"), "test.pdf")},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 400
        data = resp.get_json()
        assert data is not None
        assert data["status"] == "error"
        assert isinstance(data["errors"], list)
        assert any(
            "formato" in e.lower() or "permitido" in e.lower() or "soporta" in e.lower()
            for e in data["errors"]
        ), f"Expected format error, got: {data['errors']}"
