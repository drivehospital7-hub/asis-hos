"""Tests para app/services/exporter.py."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.services.exporter import (
    export_excel_with_cruce_facturas,
    _copy_file,
)
from app.constants import CRUCE_FACTURAS_SHEET, REVISION_SHEET


@pytest.fixture
def mock_input_output_dirs(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> tuple[Path, Path]:
    """Configura directorios input/output temporales."""
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Monkeypatch las funciones de directorio
    # input_data_directory no tiene argumentos
    monkeypatch.setattr(
        "app.utils.input_data.input_data_directory",
        lambda: input_dir,
    )
    # output_data_directory acepta create=bool keyword argument
    monkeypatch.setattr(
        "app.utils.input_data.output_data_directory",
        lambda *, create=False: output_dir,
    )
    
    return input_dir, output_dir


def _create_sample_excel(file_path: Path) -> None:
    """Helper para crear archivo Excel de prueba."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    
    # Headers con espacios adicionales simulando estructura real
    # Fila 1 y 2: headers combinados que se eliminarán
    ws.cell(row=1, column=1, value="TITULO REPORTE")
    ws.cell(row=2, column=1, value="Subtitulo")
    
    # Fila 3: Headers reales
    headers = [
        "Número Factura",
        "Entidad Cobrar",
        "Convenio Facturado",
        "Centro Costo",
        "Vlr. Subsidiado",
        "Vlr. Procedimiento",
        "Tipo Procedimiento",
        "Procedimiento",
        "Nº Identificación",
        "Cantidad",
        "Tipo Identificación",
        "Fec. Nacimiento",
        "Fec. Factura",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=header)
    
    # Datos de ejemplo
    data = [
        ["FAC-001", "MALLAMAS", "Asistencial", "ODONTOLOGIA", 1000, 500, "Consultas", "Consulta", "123", 1, "CC", "1990-01-01", "2024-01-15"],
        ["FAC-002", "MALLAMAS", "Asistencial", "ODONTOLOGIA", 2000, 1000, "Procedimientos", "Extracción", "456", 1, "CC", "1985-05-10", "2024-01-16"],
    ]
    for row_idx, row_data in enumerate(data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)


class TestCopyFile:
    """Tests para la función _copy_file."""

    def test_copia_archivo_correctamente(self, tmp_path: Path) -> None:
        """Debe copiar el archivo manteniendo contenido."""
        source = tmp_path / "source.txt"
        dest = tmp_path / "dest.txt"
        source.write_text("Contenido de prueba")
        
        _copy_file(source, dest)
        
        assert dest.exists()
        assert dest.read_text() == "Contenido de prueba"

    def test_sobrescribe_archivo_existente(self, tmp_path: Path) -> None:
        """Debe sobrescribir si el destino ya existe."""
        source = tmp_path / "source.txt"
        dest = tmp_path / "dest.txt"
        source.write_text("Contenido nuevo")
        dest.write_text("Contenido viejo")
        
        _copy_file(source, dest)
        
        assert dest.read_text() == "Contenido nuevo"


class TestExportExcelWithCruceFacturas:
    """Tests para export_excel_with_cruce_facturas."""

    def test_archivo_no_existe_retorna_error(
        self, mock_input_output_dirs: tuple[Path, Path]
    ) -> None:
        """Debe retornar error si el archivo no existe."""
        result = export_excel_with_cruce_facturas(filename="no_existe.xlsx")
        
        assert result["status"] == "error"
        assert result["data"] == {}
        assert len(result["errors"]) > 0

    def test_exporta_archivo_valido_exitosamente(
        self, 
        mock_input_output_dirs: tuple[Path, Path],
    ) -> None:
        """Debe exportar correctamente un archivo válido."""
        input_dir, output_dir = mock_input_output_dirs
        
        # Crear el archivo de prueba directamente en el input mockeado
        test_file = input_dir / "test_facturas.xlsx"
        _create_sample_excel(test_file)
        
        result = export_excel_with_cruce_facturas(filename="test_facturas.xlsx")
        
        assert result["status"] == "success"
        assert result["errors"] == []
        assert "output_file" in result["data"]
        assert "applied_rules" in result["data"]

    def test_crea_hoja_cruce_facturas(
        self,
        mock_input_output_dirs: tuple[Path, Path],
    ) -> None:
        """Debe crear la hoja CruceFacturas en el archivo de salida."""
        input_dir, output_dir = mock_input_output_dirs
        
        test_file = input_dir / "test_facturas.xlsx"
        _create_sample_excel(test_file)
        
        result = export_excel_with_cruce_facturas(filename="test_facturas.xlsx")
        
        # Verificar que se creó el archivo de salida con la hoja
        output_file = output_dir / "test_facturas.xlsx"
        assert output_file.exists()
        
        wb = load_workbook(output_file)
        assert CRUCE_FACTURAS_SHEET in wb.sheetnames

    def test_crea_hoja_revision(
        self,
        mock_input_output_dirs: tuple[Path, Path],
    ) -> None:
        """Debe crear la hoja Revision en el archivo de salida."""
        input_dir, output_dir = mock_input_output_dirs
        
        test_file = input_dir / "test_facturas.xlsx"
        _create_sample_excel(test_file)
        
        result = export_excel_with_cruce_facturas(filename="test_facturas.xlsx")
        
        output_file = output_dir / "test_facturas.xlsx"
        wb = load_workbook(output_file)
        assert REVISION_SHEET in wb.sheetnames

    def test_retorna_formato_respuesta_estandar(
        self,
        mock_input_output_dirs: tuple[Path, Path],
    ) -> None:
        """Debe retornar respuesta con formato estándar."""
        input_dir, _ = mock_input_output_dirs
        
        test_file = input_dir / "test_facturas.xlsx"
        _create_sample_excel(test_file)
        
        result = export_excel_with_cruce_facturas(filename="test_facturas.xlsx")
        
        # Verificar estructura estándar
        assert "status" in result
        assert "data" in result
        assert "errors" in result
        assert isinstance(result["data"], dict)
        assert isinstance(result["errors"], list)

    def test_filename_vacio_retorna_error(
        self, mock_input_output_dirs: tuple[Path, Path]
    ) -> None:
        """Debe retornar error si filename está vacío."""
        result = export_excel_with_cruce_facturas(filename="")
        
        assert result["status"] == "error"
        assert len(result["errors"]) > 0

    def test_path_traversal_retorna_error(
        self, mock_input_output_dirs: tuple[Path, Path]
    ) -> None:
        """Debe retornar error si se intenta path traversal."""
        result = export_excel_with_cruce_facturas(filename="../../../etc/passwd")
        
        assert result["status"] == "error"
        assert len(result["errors"]) > 0
