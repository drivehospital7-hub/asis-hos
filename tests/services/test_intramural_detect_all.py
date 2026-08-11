"""Tests for app/services/intramural/detect_all.py.

Strict TDD: tests written BEFORE implementation.
"""

from __future__ import annotations

from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from app.services.intramural.detect_all import detect_all_problems_intramural


@pytest.fixture
def workbook_minimal() -> Workbook:
    """Crea un workbook con headers mínimos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    return wb


class TestDetectAllProblemsIntramural:
    """Tests para detect_all_problems_intramural."""

    def _run(self, ws, indices):
        with patch("app.services.intramural.detect_all.is_rule_engine_enabled", return_value=False):
            result, _ = detect_all_problems_intramural(ws, indices)
        return result

    def _run_with_data(self, ws, indices, revision_data=None):
        """Helper que corre con engine activo pero mockeado."""
        def _mock_detector(name, session):
            d = MagicMock()
            if name == "revision_cantidad_intramural" and revision_data is not None:
                d.detect.return_value = revision_data
            else:
                d.detect.return_value = []
            return d

        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = MagicMock()
                m_dc.side_effect = _mock_detector
                result, _ = detect_all_problems_intramural(ws, indices)
        return result

    def test_retorna_dict_con_key_problemas(self, workbook_minimal: Workbook) -> None:
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert "problemas" in result

    def test_retorna_area_intramural(self, workbook_minimal: Workbook) -> None:
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert result.get("area") == "intramural"

    def test_resultado_incluye_normalizados(self, workbook_minimal: Workbook) -> None:
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert "normalizados" in result["problemas"]
        assert isinstance(result["problemas"]["normalizados"], list)

    def test_missing_columns_present(self, workbook_minimal: Workbook) -> None:
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert "missing_columns" in result

    def test_revision_cantidad_in_resultado(self) -> None:
        """resultado['problemas'] debe incluir 'revision_cantidad'."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        headers = ["Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
                    "Cantidad", "Código Tipo Procedimiento", "Laboratorio"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="X001")
        ws.cell(row=2, column=3, value="Proc A")
        ws.cell(row=2, column=4, value=3)
        ws.cell(row=2, column=5, value="06")
        ws.cell(row=2, column=6, value="Si")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "codigo_tipo_procedimiento": 4,
            "laboratorio": 5,
        }
        revision_data = [{"factura": "F001", "tipo_error": "⚠️ Revisión Necesaria", "cantidad": 3}]
        result = self._run_with_data(ws, indices, revision_data=revision_data)
        assert "revision_cantidad" in result["problemas"]
        assert len(result["problemas"]["revision_cantidad"]) == 1

    def test_revision_cantidad_in_totales(self) -> None:
        """resultado['totales'] debe incluir 'revision_cantidad'."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        headers = ["Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
                    "Cantidad", "Código Tipo Procedimiento", "Laboratorio"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="X001")
        ws.cell(row=2, column=3, value="Proc A")
        ws.cell(row=2, column=4, value=3)
        ws.cell(row=2, column=5, value="06")
        ws.cell(row=2, column=6, value="Si")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "codigo_tipo_procedimiento": 4,
            "laboratorio": 5,
        }
        revision_data = [{"factura": "F001", "tipo_error": "⚠️ Revisión Necesaria", "cantidad": 3}]
        result = self._run_with_data(ws, indices, revision_data=revision_data)
        assert "revision_cantidad" in result["totales"]
        assert result["totales"]["revision_cantidad"] == 1

    def test_revision_cantidad_in_normalized_rows(self) -> None:
        """revision_cantidad items aparecen en normalizados como ⚠️ Revisión."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        headers = ["Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
                    "Cantidad", "Código Tipo Procedimiento", "Laboratorio"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="X001")
        ws.cell(row=2, column=3, value="Proc A")
        ws.cell(row=2, column=4, value=3)
        ws.cell(row=2, column=5, value="06")
        ws.cell(row=2, column=6, value="Si")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "codigo_tipo_procedimiento": 4,
            "laboratorio": 5,
        }
        revision_data = [{"factura": "F001", "tipo_error": "⚠️ Revisión Necesaria",
                          "cantidad": 3, "detalle": "Cant: 3"}]
        result = self._run_with_data(ws, indices, revision_data=revision_data)
        normalizados = result["problemas"]["normalizados"]
        revision_rows = [
            r for r in normalizados if r["tipo_error"] == "⚠️ Revisión Necesaria"
        ]
        assert len(revision_rows) == 1
        assert revision_rows[0]["factura"] == "F001"
        assert "Cant:" in revision_rows[0]["detalle"]

