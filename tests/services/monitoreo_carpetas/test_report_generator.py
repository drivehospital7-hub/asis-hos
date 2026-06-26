"""Tests for app/services/monitoreo_carpetas/report_generator.py."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.monitoreo_carpetas import InvoiceRecord, ScanResult
from app.services.monitoreo_carpetas.report_generator import generate_excel


class TestGenerateExcel:
    """Tests for generate_excel()."""

    @pytest.fixture
    def sample_scan_result(self) -> ScanResult:
        """Sample ScanResult with known data."""
        facturas = [
            InvoiceRecord(
                filename="FEV123.pdf", facturador="Juan",
                full_path="/ruta/Juan/FEV123.pdf", status="Verificada",
                invoice_type="FEV", invoice_code="FEV123",
            ),
            InvoiceRecord(
                filename="CAP001_ABC002.pdf", facturador="Maria",
                full_path="/ruta/Maria/CAP001_ABC002.pdf", status="Por corregir",
                invoice_type="CAP", invoice_code="CAP001_ABC002",
            ),
            InvoiceRecord(
                filename="FEV999.pdf", facturador="Luis",
                full_path="/ruta/Luis/FEV999.pdf", status="En revisión",
                invoice_type="FEV", invoice_code="FEV999",
            ),
        ]
        return ScanResult(
            facturas=facturas,
            duplicados=[{"filename": "FEV123.pdf", "facturadores": ["Juan", "Otro"]}],
            vacias=[{"facturador": "Luis", "folder": "/ruta/Luis", "reason": "empty"}],
            indicadores={
                "total_facturas": 3,
                "total_facturadores": 3,
                "total_vacias": 1,
                "total_duplicados": 1,
                "status_Verificada": 1,
                "status_Por corregir": 1,
                "status_En revisión": 1,
                "type_FEV": 2,
                "type_CAP": 1,
            },
            errores_scan=[],
            excel_path=None,
        )

    def test_generates_excel_file(self, sample_scan_result: ScanResult, tmp_path: Path) -> None:
        """generate_excel creates an xlsx file at the given path."""
        output_path = tmp_path / "monitoreo_test.xlsx"
        result = generate_excel(sample_scan_result, str(output_path))
        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_excel_has_facturas_sheet(self, sample_scan_result: ScanResult, tmp_path: Path) -> None:
        """Excel file has Facturas sheet."""
        output_path = tmp_path / "test.xlsx"
        result = generate_excel(sample_scan_result, str(output_path))
        wb = load_workbook(result)
        assert "Facturas" in wb.sheetnames

    def test_excel_has_indicadores_sheet(self, sample_scan_result: ScanResult, tmp_path: Path) -> None:
        """Excel file has Indicadores sheet."""
        output_path = tmp_path / "test.xlsx"
        result = generate_excel(sample_scan_result, str(output_path))
        wb = load_workbook(result)
        assert "Indicadores" in wb.sheetnames

    def test_facturas_has_header_row(self, sample_scan_result: ScanResult, tmp_path: Path) -> None:
        """Facturas sheet has header row with expected columns."""
        output_path = tmp_path / "test.xlsx"
        result = generate_excel(sample_scan_result, str(output_path))
        wb = load_workbook(result)
        ws = wb["Facturas"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Código Factura" in headers
        assert "Tipo" in headers
        assert "Estado" in headers
        assert "Ruta Completa" in headers
        assert "Facturador" in headers

    def test_facturas_has_data_rows(self, sample_scan_result: ScanResult, tmp_path: Path) -> None:
        """Facturas sheet has data rows matching invoice count."""
        output_path = tmp_path / "test.xlsx"
        result = generate_excel(sample_scan_result, str(output_path))
        wb = load_workbook(result)
        ws = wb["Facturas"]
        # Row 1 is header, so data rows start from 2
        data_rows = ws.max_row - 1
        assert data_rows == 3

    def test_indicadores_has_data(self, sample_scan_result: ScanResult, tmp_path: Path) -> None:
        """Indicadores sheet has status and type counts."""
        output_path = tmp_path / "test.xlsx"
        result = generate_excel(sample_scan_result, str(output_path))
        wb = load_workbook(result)
        ws = wb["Indicadores"]
        # Read all values
        values = {}
        for row in range(1, ws.max_row + 1):
            key = ws.cell(row=row, column=1).value
            val = ws.cell(row=row, column=2).value
            if key:
                values[key] = val
        assert values.get("total_facturas") == 3
        assert values.get("total_vacias") == 1

    def test_empty_scan_result_has_header_only(self, tmp_path: Path) -> None:
        """Empty ScanResult produces header-only Facturas sheet."""
        empty = ScanResult()
        output_path = tmp_path / "empty.xlsx"
        result = generate_excel(empty, str(output_path))
        wb = load_workbook(result)
        ws = wb["Facturas"]
        assert ws.max_row == 1  # header only
