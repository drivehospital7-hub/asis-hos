"""Tests para app/services/transversales/doble_tipo_procedimiento.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.transversales.doble_tipo_procedimiento import (
    detect_doble_tipo_procedimiento,
)


@pytest.fixture
def workbook_with_headers() -> Workbook:
    """Crea un workbook con headers mínimos para doble tipo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Tipo Procedimiento")
    return wb


class TestDetectDobleTipoProcedimiento:
    """Tests para detect_doble_tipo_procedimiento."""

    def test_detecta_factura_con_multiples_tipos(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe detectar facturas con más de un tipo de procedimiento."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="Consultas")
        ws.cell(row=3, column=1, value="FAC-001")
        ws.cell(row=3, column=2, value="Procedimientos")
        ws.cell(row=4, column=1, value="FAC-002")
        ws.cell(row=4, column=2, value="Consultas")
        ws.cell(row=5, column=1, value="FAC-002")
        ws.cell(row=5, column=2, value="Consultas")

        indices = {
            "numero_factura": 0,
            "tipo_procedimiento": 1,
        }
        result = detect_doble_tipo_procedimiento(ws, indices)

        facturas = [r["factura"] for r in result]
        assert "FAC-001" in facturas
        assert "FAC-002" not in facturas

    def test_factura_con_un_solo_tipo_no_aparece(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Factura con un solo tipo no debe aparecer en resultados."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="Consultas")
        ws.cell(row=3, column=1, value="FAC-001")
        ws.cell(row=3, column=2, value="Consultas")

        indices = {
            "numero_factura": 0,
            "tipo_procedimiento": 1,
        }
        result = detect_doble_tipo_procedimiento(ws, indices)

        assert len(result) == 0

    def test_sin_indices_retorna_vacio(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Sin indices necesarios debe retornar lista vacía."""
        result = detect_doble_tipo_procedimiento(
            workbook_with_headers.active,
            {"numero_factura": None, "tipo_procedimiento": None},
        )

        assert result == []

    def test_valores_none_en_celda_no_causan_error(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Celdas vacías no deben causar error."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value=None)  # Tipo vacío
        ws.cell(row=3, column=1, value="FAC-001")
        ws.cell(row=3, column=2, value="Consultas")
        ws.cell(row=4, column=1, value="FAC-002")
        ws.cell(row=4, column=2, value="Procedimientos")

        indices = {
            "numero_factura": 0,
            "tipo_procedimiento": 1,
        }
        result = detect_doble_tipo_procedimiento(ws, indices)

        # FAC-001 tiene Consultas (fila 3) y None (fila 2) -> solo 1 tipo real
        # FAC-002 tiene 1 tipo -> no debe aparecer
        assert len(result) == 0

    def test_multiples_tipos_aparecen_en_resultado(
        self, workbook_with_headers: Workbook
    ) -> None:
        """El resultado debe listar los tipos separados por coma."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="Consultas")
        ws.cell(row=3, column=1, value="FAC-001")
        ws.cell(row=3, column=2, value="Procedimientos")
        ws.cell(row=4, column=1, value="FAC-001")
        ws.cell(row=4, column=2, value="Cirugía")

        indices = {
            "numero_factura": 0,
            "tipo_procedimiento": 1,
        }
        result = detect_doble_tipo_procedimiento(ws, indices)

        assert len(result) == 1
        tipos = result[0]["tipos"].split(", ")
        assert "Cirugía" in tipos
        assert "Consultas" in tipos
        assert "Procedimientos" in tipos
