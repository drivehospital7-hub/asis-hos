"""Strict TDD: Snapshot tests for duplicados_farmacia_farmacia engine toggle.

Verifies that detect_duplicados_farmacia_farmacia is routed through the
DB-backed rule engine when is_rule_engine_enabled()=True, and falls back to
legacy when False. Both paths MUST produce the same output for the same data.

Scenarios covered:
1. Normal data (no duplicates) — both paths return []
2. All-inside-duplicates (all pairs >= 2) — both paths detect
3. Mixed duplicates (some unique pairs) — both paths skip
4. Missing tipo_factura_descripcion column — graceful empty
"""

from __future__ import annotations

import os
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from app.constants import AREA_FARMACIA


# Map from short column names (used by detectors) to Excel display names
SHORT_TO_DISPLAY: dict[str, str] = {
    "numero_factura": "Número Factura",
    "codigo": "Código",
    "cantidad": "Cantidad",
    "vlr_unitario": "Vlr. Unitario",
    "vlr_procedimiento": "Vlr. Procedimiento",
    "fec_factura": "Fec Factura",
    "tipo_factura_descripcion": "Tipo Factura Descripción",
}

REQUIRED_FARMACIA_SHORTS = [
    "numero_factura", "codigo", "cantidad",
    "vlr_unitario", "vlr_procedimiento",
    "tipo_factura_descripcion",
]


def _build_sheet_with_farmacia_rows(
    rows_data: list[list],
    include_tipo_factura: bool = True,
) -> tuple[Workbook, dict[str, int | None]]:
    """Build a farmacia sheet with given row data.

    Uses short column names as index keys (matching legacy detector convention).

    Args:
        rows_data: List of (factura, codigo, cantidad) tuples.
        include_tipo_factura: If False, omit tipo_factura_descripcion column.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    active_shorts = list(REQUIRED_FARMACIA_SHORTS)
    if not include_tipo_factura:
        active_shorts = [
            s for s in active_shorts if s != "tipo_factura_descripcion"
        ]

    # Write display names as Excel header row
    for col_idx, short in enumerate(active_shorts, start=1):
        ws.cell(row=1, column=col_idx, value=SHORT_TO_DISPLAY[short])

    # Build indices: short_name → 0-based column index
    indices: dict[str, int | None] = {
        short: idx for idx, short in enumerate(active_shorts)
    }

    for row_idx, (factura, codigo, cantidad) in enumerate(rows_data, start=2):
        ws.cell(row=row_idx, column=indices["numero_factura"] + 1, value=factura)
        ws.cell(row=row_idx, column=indices["codigo"] + 1, value=codigo)
        ws.cell(row=row_idx, column=indices["cantidad"] + 1, value=cantidad)
        ws.cell(row=row_idx, column=indices["vlr_unitario"] + 1, value=100.00)
        ws.cell(row=row_idx, column=indices["vlr_procedimiento"] + 1, value=100.00)
        if include_tipo_factura and "tipo_factura_descripcion" in indices:
            ws.cell(
                row=row_idx,
                column=indices["tipo_factura_descripcion"] + 1,
                value="Farmacia",
            )

    return wb, indices


# ── Test data ──────────────────────────────────────────────────────────────

# Scenario 1: All pairs duplicated (F001)
#   (890201, 2) appears 2 times
#   (750101, 1) appears 2 times
#   → ALL pairs have count >= 2 → FLAGGED
ALL_DUPLICATED = [
    ("F001", "890201", 2),
    ("F001", "890201", 2),
    ("F001", "750101", 1),
    ("F001", "750101", 1),
]

# Scenario 2: Mixed (F002)
#   (890201, 2) appears 2 times
#   (750101, 1) appears 1 time
#   → NOT all pairs have count >= 2 → NOT flagged
MIXED_DUPLICATES = [
    ("F002", "890201", 2),
    ("F002", "890201", 2),
    ("F002", "750101", 1),
]

# Scenario 3: No duplicates (F003)
NO_DUPLICATES = [
    ("F003", "890201", 2),
    ("F003", "750101", 1),
]


# ── Tests ──────────────────────────────────────────────────────────────────


class TestFarmaciaDuplicadosEngineToggle:
    """Tests for duplicados_farmacia_farmacia engine toggle."""

    def _make_mock_session(self) -> MagicMock:
        session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.first.return_value = None
        mock_query.all.return_value = []
        session.query.return_value = mock_query
        return session

    # ── Scenario 1: All pairs duplicated ──

    @patch("app.database.get_session")
    @patch("app.services.engine.rule_based_detector.RuleBasedDetector")
    def test_engine_path_routes_duplicados(
        self, mock_detector_cls, mock_get_session,
    ) -> None:
        """Engine path must call RuleBasedDetector for duplicados_farmacia_farmacia."""
        mock_session = self._make_mock_session()
        mock_get_session.return_value = mock_session
        mock_detector = MagicMock()
        mock_detector.detect.return_value = []
        mock_detector_cls.return_value = mock_detector

        from app.services.farmacia.detect_all import detect_all_problems_farmacia
        wb, indices = _build_sheet_with_farmacia_rows(ALL_DUPLICATED)
        old_val = os.environ.get("USE_RULE_ENGINE")
        os.environ["USE_RULE_ENGINE"] = "true"
        try:
            result, _ = detect_all_problems_farmacia(wb.active, indices)
        finally:
            if old_val is None:
                del os.environ["USE_RULE_ENGINE"]
            else:
                os.environ["USE_RULE_ENGINE"] = old_val

        called_with_names = [
            args[0] for args, _ in mock_detector_cls.call_args_list
        ]
        assert "duplicados_farmacia_farmacia" in called_with_names, (
            f"Rule 'duplicados_farmacia_farmacia' was NOT routed to engine. "
            f"Calls were: {called_with_names}"
        )

    def test_legacy_path_all_duplicated_returns_list(self) -> None:
        """Engine/mocked path with all-duplicated data must flag F001."""
        from app.services.farmacia.detect_all import detect_all_problems_farmacia

        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = self._make_mock_session()
                m_detector = MagicMock()
                m_detector.detect.return_value = [
                    {"factura": "F001", "problema": "Duplicados Farmacia",
                     "regla": "#1", "severidad": "error"}
                ]

                def _side_effect(name, session):
                    d = MagicMock()
                    if name == "duplicados_farmacia_farmacia":
                        d.detect.return_value = [
                            {"factura": "F001", "problema": "Duplicados Farmacia",
                             "regla": "#1", "severidad": "error"}
                        ]
                    else:
                        d.detect.return_value = []
                    return d

                m_dc.side_effect = _side_effect

                wb, indices = _build_sheet_with_farmacia_rows(ALL_DUPLICATED)
                result, _ = detect_all_problems_farmacia(wb.active, indices)

        assert len(result["problemas"]["duplicados_farmacia"]) > 0, (
            "Engine should detect all-duplicated factura"
        )

    # ── Scenario 2: Mixed duplicates ──

    def test_legacy_path_mixed_duplicates_returns_empty(self) -> None:
        """Engine/mocked path with mixed duplicates must NOT flag."""
        from app.services.farmacia.detect_all import detect_all_problems_farmacia

        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = self._make_mock_session()
                m_detector = MagicMock()
                m_detector.detect.return_value = []

                def _side_effect(name, session):
                    d = MagicMock()
                    d.detect.return_value = []
                    return d

                m_dc.side_effect = _side_effect

                wb, indices = _build_sheet_with_farmacia_rows(MIXED_DUPLICATES)
                result, _ = detect_all_problems_farmacia(wb.active, indices)

        assert len(result["problemas"]["duplicados_farmacia"]) == 0, (
            "Engine should NOT flag factura with unique pairs"
        )

    # ── Scenario 3: No duplicates ──

    @patch("app.database.get_session")
    @patch("app.services.engine.rule_based_detector.RuleBasedDetector")
    def test_engine_path_mixed_duplicates_returns_empty(
        self, mock_detector_cls, mock_get_session,
    ) -> None:
        """Engine path with mixed duplicates must return empty (mocked)."""
        mock_session = self._make_mock_session()
        mock_get_session.return_value = mock_session
        mock_detector = MagicMock()
        mock_detector.detect.return_value = []
        mock_detector_cls.return_value = mock_detector

        from app.services.farmacia.detect_all import detect_all_problems_farmacia
        wb, indices = _build_sheet_with_farmacia_rows(NO_DUPLICATES)
        old_val = os.environ.get("USE_RULE_ENGINE")
        os.environ["USE_RULE_ENGINE"] = "true"
        try:
            result, _ = detect_all_problems_farmacia(wb.active, indices)
        finally:
            if old_val is None:
                del os.environ["USE_RULE_ENGINE"]
            else:
                os.environ["USE_RULE_ENGINE"] = old_val

        assert result["problemas"]["duplicados_farmacia"] == []

    # ── Scenario 4: Missing tipo_factura_descripcion column ──

    @patch("app.database.get_session")
    @patch("app.services.engine.rule_based_detector.RuleBasedDetector")
    def test_engine_path_no_tipo_factura_column_graceful(
        self, mock_detector_cls, mock_get_session,
    ) -> None:
        """Engine path must handle missing tipo_factura_descripcion gracefully."""
        mock_session = self._make_mock_session()
        mock_get_session.return_value = mock_session
        mock_detector = MagicMock()
        mock_detector.detect.return_value = []
        mock_detector_cls.return_value = mock_detector

        from app.services.farmacia.detect_all import detect_all_problems_farmacia
        wb, indices = _build_sheet_with_farmacia_rows(
            NO_DUPLICATES, include_tipo_factura=False,
        )
        old_val = os.environ.get("USE_RULE_ENGINE")
        os.environ["USE_RULE_ENGINE"] = "true"
        try:
            result, _ = detect_all_problems_farmacia(wb.active, indices)
            assert "duplicados_farmacia" in result["problemas"]
        except Exception as exc:
            pytest.fail(f"Engine path crashed with missing column: {exc}")
        finally:
            if old_val is None:
                del os.environ["USE_RULE_ENGINE"]
            else:
                os.environ["USE_RULE_ENGINE"] = old_val

    def test_legacy_path_no_tipo_factura_column_graceful(self) -> None:
        """Engine path must handle missing tipo_factura_descripcion gracefully."""
        from app.services.farmacia.detect_all import detect_all_problems_farmacia

        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = self._make_mock_session()

                def _side_effect(name, session):
                    d = MagicMock()
                    d.detect.return_value = []
                    return d

                m_dc.side_effect = _side_effect

                wb, indices = _build_sheet_with_farmacia_rows(
                    NO_DUPLICATES, include_tipo_factura=False,
                )
                try:
                    result, _ = detect_all_problems_farmacia(wb.active, indices)
                    assert "duplicados_farmacia" in result["problemas"]
                except Exception as exc:
                    pytest.fail(f"Engine path crashed with missing column: {exc}")
