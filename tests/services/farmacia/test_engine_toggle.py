"""Strict TDD: Tests for engine toggle in farmacia/detect_all.py.

Verifies that is_rule_engine_enabled() toggle routes to RuleBasedDetector
when True, and falls back to legacy detectors when False, without crashing.
"""

from __future__ import annotations

from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from app.constants import AREA_FARMACIA


def _build_simple_sheet() -> tuple[Workbook, dict[str, int | None]]:
    """Build a workbook with minimal columns for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    headers = [
        "Número Factura", "Código", "Cantidad", "Vlr. Unitario",
        "Vlr. Procedimiento", "Tipo Doc.", "Edad", "Tipo Identificación",
        "Código Entidad Cobrar", "Entidad Afiliación", "Tipo Usuario",
        "Vlr. Copago", "Código CUPS", "Fec Factura", "Fecha Cierre",
        "Responsable Cierra",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    ws.cell(row=2, column=1, value="FAC-001")

    indices = {h: i for i, h in enumerate(headers)}
    return wb, indices


class TestFarmaciaEngineToggle:
    """Tests for the engine toggle in detect_all_problems_farmacia."""

    def _make_mock_session(self) -> MagicMock:
        """Create a mock DB session that returns empty results."""
        session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.first.return_value = None
        mock_query.all.return_value = []
        session.query.return_value = mock_query
        return session

    @patch("app.database.get_session")
    @patch("app.services.engine.rule_based_detector.RuleBasedDetector")
    @patch("app.constants.base.is_rule_engine_enabled", return_value=True)
    def test_engine_path_routes_to_rule_based_detector(
        self, mock_enabled: MagicMock, mock_detector_cls: MagicMock,
        mock_get_session: MagicMock,
    ) -> None:
        """Engine path must instantiate RuleBasedDetector for transversal rules.

        This test FAILS if the toggle is not implemented.
        """
        mock_session = self._make_mock_session()
        mock_get_session.return_value = mock_session
        mock_detector = MagicMock()
        mock_detector.detect.return_value = []
        mock_detector_cls.return_value = mock_detector

        from app.services.farmacia.detect_all import (
            detect_all_problems_farmacia,
        )
        wb, indices = _build_simple_sheet()
        result, responsables = detect_all_problems_farmacia(
            wb.active, indices,
        )

        assert mock_detector_cls.call_count >= 3, (
            f"RuleBasedDetector called only {mock_detector_cls.call_count}x "
            f"— toggle likely not implemented"
        )
        assert "problemas" in result
        assert isinstance(result["problemas"], dict)
        assert "totales" in result
        assert result["area"] == AREA_FARMACIA
        assert responsables == {}

    @patch("app.constants.base.is_rule_engine_enabled", return_value=False)
    def test_legacy_path_returns_valid_structure(
        self, mock_enabled: MagicMock,
    ) -> None:
        """Legacy path must return a valid result dict without crashing."""
        from app.services.farmacia.detect_all import (
            detect_all_problems_farmacia,
        )
        wb, indices = _build_simple_sheet()
        result, responsables = detect_all_problems_farmacia(
            wb.active, indices,
        )

        assert "problemas" in result
        assert isinstance(result["problemas"], dict)
        assert "totales" in result
        assert result["area"] == AREA_FARMACIA
        assert responsables == {}

    @patch("app.database.get_session")
    @patch("app.services.engine.rule_based_detector.RuleBasedDetector")
    @patch("app.constants.base.is_rule_engine_enabled", return_value=True)
    def test_engine_path_with_all_detectors(
        self, mock_enabled: MagicMock, mock_detector_cls: MagicMock,
        mock_get_session: MagicMock,
    ) -> None:
        """Engine path must produce problems dict with all keys present."""
        mock_session = self._make_mock_session()
        mock_get_session.return_value = mock_session
        mock_detector = MagicMock()
        mock_detector.detect.return_value = []
        mock_detector_cls.return_value = mock_detector

        from app.services.farmacia.detect_all import (
            detect_all_problems_farmacia,
        )
        wb, indices = _build_simple_sheet()
        result, _ = detect_all_problems_farmacia(wb.active, indices)

        assert mock_detector_cls.call_count >= 3

        problemas = result["problemas"]
        expected_keys = {
            "normalizados", "centros_de_costos", "ide_contrato",
            "cups_equivalentes", "decimales", "tipo_identificacion_edad",
            "tipo_identificacion_entidad", "codigo_entidad_vs_afiliacion",
            "tipo_usuario", "copago_entidad", "cups_sin_contrato",
            "duplicados_farmacia",
        }
        for key in expected_keys:
            assert key in problemas, f"Missing key: {key}"

    @patch("app.constants.base.is_rule_engine_enabled", return_value=False)
    def test_legacy_path_with_all_detectors(
        self, mock_enabled: MagicMock,
    ) -> None:
        """Legacy path must produce problems dict with all keys present."""
        from app.services.farmacia.detect_all import (
            detect_all_problems_farmacia,
        )
        wb, indices = _build_simple_sheet()
        result, _ = detect_all_problems_farmacia(wb.active, indices)

        problemas = result["problemas"]
        expected_keys = {
            "normalizados", "centros_de_costos", "ide_contrato",
            "cups_equivalentes", "decimales", "tipo_identificacion_edad",
            "tipo_identificacion_entidad", "codigo_entidad_vs_afiliacion",
            "tipo_usuario", "copago_entidad", "cups_sin_contrato",
            "duplicados_farmacia",
        }
        for key in expected_keys:
            assert key in problemas, f"Missing key: {key}"
