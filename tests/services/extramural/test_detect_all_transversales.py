"""F1 Snapshot tests: engine vs legacy structure match."""
from __future__ import annotations
import os
from unittest.mock import MagicMock, patch
from openpyxl import Workbook

from app.constants import AREA_EXTRAMURAL

TRANSVERSAL_KEYS = frozenset({
    "decimales",
    "tipo_identificacion_edad",
    "tipo_identificacion_entidad",
    "codigo_entidad_vs_afiliacion",
    "tipo_usuario",
    "copago_entidad",
    "cups_sin_contrato",
})

ALL_PROBLEM_KEYS = TRANSVERSAL_KEYS | {
    "normalizados",
    "centros_de_costos",
    "ide_contrato",
    "cups_equivalentes",

}


def _env(val):
    old = os.environ.pop("USE_RULE_ENGINE", None)
    os.environ["USE_RULE_ENGINE"] = val
    def restore():
        if old is not None:
            os.environ["USE_RULE_ENGINE"] = old
        else:
            os.environ.pop("USE_RULE_ENGINE", None)
    return restore


def _build_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    headers = [
        "N\u00famero Factura", "C\u00f3digo", "Cantidad", "Vlr. Unitario",
        "Vlr. Procedimiento", "Vlr. Subsidiado", "Tipo Doc.", "Edad",
        "Tipo Identificaci\u00f3n", "C\u00f3digo Entidad Cobrar", "Entidad Afiliaci\u00f3n",
        "Tipo Usuario", "Vlr. Copago", "C\u00f3digo CUPS", "Fec Factura",
        "Fecha Cierre", "Responsable Cierra", "Tarifario",
        "Tipo Factura Descripci\u00f3n",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    ws.cell(row=2, column=1, value="FAC-001")
    ws.cell(row=2, column=3, value=1)
    ws.cell(row=2, column=4, value=150.50)
    ws.cell(row=2, column=5, value=150.50)
    ws.cell(row=2, column=7, value="CC")
    ws.cell(row=2, column=8, value=30)
    ws.cell(row=2, column=9, value="CC")
    ws.cell(row=2, column=10, value="EPS001")
    ws.cell(row=2, column=11, value="EPS001")
    ws.cell(row=2, column=12, value="CONTRIBUTIVO")
    ws.cell(row=2, column=13, value=0)
    ws.cell(row=2, column=15, value="2024-06-01")
    ws.cell(row=2, column=17, value="")
    ws.cell(row=2, column=19, value="Farmacia")
    indices = {h: i for i, h in enumerate(headers)}
    return wb, indices


def _mock_session():
    s = MagicMock()
    q = MagicMock()
    q.filter.return_value = q
    q.order_by.return_value = q
    q.first.return_value = None
    q.all.return_value = []
    s.query.return_value = q
    return s


class TestSnapshot:
    def test_engine_path_structure(self):
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = _mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md
                mod = __import__("app.services.extramural.detect_all", fromlist=["detect_all_problems_extramural"])
                func = getattr(mod, "detect_all_problems_extramural")
                wb, idx = _build_sheet()
                restore = _env("true")
                try:
                    r, _ = func(wb.active, idx)
                finally:
                    restore()
        for k in ALL_PROBLEM_KEYS:
            assert k in r["problemas"], f"Missing {k}"
        for k in TRANSVERSAL_KEYS:
            assert r["problemas"][k] == [], f"Expected empty {k}"
        assert r["area"] == AREA_EXTRAMURAL

    def test_legacy_path_structure(self):
        mod = __import__("app.services.extramural.detect_all", fromlist=["detect_all_problems_extramural"])
        func = getattr(mod, "detect_all_problems_extramural")
        wb, idx = _build_sheet()
        restore = _env("false")
        try:
            r, _ = func(wb.active, idx)
        finally:
            restore()
        for k in ALL_PROBLEM_KEYS:
            assert k in r["problemas"], f"Missing {k}"
        assert r["area"] == AREA_EXTRAMURAL

    def test_keys_match_across_paths(self):
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = _mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md
                mod = __import__("app.services.extramural.detect_all", fromlist=["detect_all_problems_extramural"])
                func = getattr(mod, "detect_all_problems_extramural")
                wb, idx = _build_sheet()
                restore = _env("true")
                try:
                    er, _ = func(wb.active, idx)
                finally:
                    restore()
                restore2 = _env("false")
                try:
                    lr, _ = func(wb.active, idx)
                finally:
                    restore2()
        assert er["problemas"].keys() == lr["problemas"].keys()
