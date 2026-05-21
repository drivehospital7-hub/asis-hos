"""Tests para app/services/odontologia/centro_costo.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.odontologia.centro_costo import detect_centro_costo_odontologia


@pytest.fixture
def workbook_with_centro_costo_headers() -> Workbook:
    """Crea un workbook con headers para centro de costo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Centro Costo")
    ws.cell(row=1, column=3, value="Fec. Nacimiento")
    ws.cell(row=1, column=4, value="Identificación Profesional")
    return wb


class TestDetectCentroCostoOdontologia:
    """Tests para detect_centro_costo_odontologia."""

    def test_centro_costo_valido_no_genera_error(
        self, workbook_with_centro_costo_headers: Workbook
    ) -> None:
        """Centro ODONTOLOGIA válido no genera error."""
        ws = workbook_with_centro_costo_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="ODONTOLOGIA")

        indices = {"numero_factura": 0, "centro_costo": 1,
                   "fec_factura": 2, "profesional_identificacion": 3}
        result = detect_centro_costo_odontologia(ws, indices,
                                                  permitir_todos_centros=True)

        assert len(result) == 0

    def test_centro_costo_invalido_genera_error(
        self, workbook_with_centro_costo_headers: Workbook
    ) -> None:
        """Centro de costo no válido genera error."""
        ws = workbook_with_centro_costo_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="CIRUGIA")

        indices = {"numero_factura": 0, "centro_costo": 1,
                   "fec_factura": 2, "profesional_identificacion": 3}
        result = detect_centro_costo_odontologia(ws, indices,
                                                  permitir_todos_centros=True)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"
        assert result[0]["centro_actual"] == "CIRUGIA"
        assert "ODONTOLOGIA" in result[0]["centro_deberia"]

    def test_centro_extramural_valido(
        self, workbook_with_centro_costo_headers: Workbook
    ) -> None:
        """Centro EXTRAMURAL también es válido con permitir_todos_centros."""
        ws = workbook_with_centro_costo_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="SERVICIOS ODONTOLOGIA -EXTRAMURALES")

        indices = {"numero_factura": 0, "centro_costo": 1,
                   "fec_factura": 2, "profesional_identificacion": 3}
        result = detect_centro_costo_odontologia(ws, indices,
                                                  permitir_todos_centros=True)

        assert len(result) == 0

    def test_sin_indices_no_genera_error(
        self, workbook_with_centro_costo_headers: Workbook
    ) -> None:
        """Sin numero_factura o centro_costo, retorna vacío."""
        ws = workbook_with_centro_costo_headers.active
        result = detect_centro_costo_odontologia(ws, {})
        assert result == []

    def test_centros_validos_personalizados(
        self, workbook_with_centro_costo_headers: Workbook
    ) -> None:
        """centros_validos personalizado funciona correctamente."""
        ws = workbook_with_centro_costo_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="MI_CENTRO")

        indices = {"numero_factura": 0, "centro_costo": 1,
                   "fec_factura": 2, "profesional_identificacion": 3}
        result = detect_centro_costo_odontologia(
            ws, indices, permitir_todos_centros=True,
            centros_validos=["MI_CENTRO", "OTRO_CENTRO"]
        )

        assert len(result) == 0
