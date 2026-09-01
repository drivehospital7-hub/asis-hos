"""R14 corrected: Factura + FURIPS split — RED suite (Strict TDD).

Expected to FAIL against the pre-fix vocabulary (5 entries with
"Factura y Furips" combined). After the fix to
app/constants/urgencias.py:438-444 the suite must turn GREEN.

Covers spec S1-S6: constant, opciones, create/update verbatim,
exact case-sensitive filtering, export verbatim.
"""

from __future__ import annotations

from io import BytesIO
from unittest.mock import patch
from contextlib import ExitStack

from openpyxl import load_workbook

from app import create_app
from app.services.control_errores_export import build_errores_export_workbook
from app.services.control_errores_service import get_errores, get_opciones

_APP = create_app({"TESTING": True, "SECRET_KEY": "test-secret-key"})

EXPECTED = ["Otros", "Soportes de Carpeta", "Factura Abierta", "Carpeta no entregada", "Factura", "FURIPS"]


def _error_fixture(error_id="err-1", tipo_error="Otros", factura="FAC-001", creado_en="2026-05-15T10:30:00"):
    return {
        "id": error_id,
        "factura": factura,
        "refactura": "",
        "creado_en": creado_en,
        "tipo_error": tipo_error,
        "observacion": "Desc",
        "responsable": "JUAN PEREZ",
        "observacion_facturador": "Revisar",
        "estado": "S",
        "validador": "MARIA GOMEZ",
    }


# ---------------------------------------------------------------------------
# 1.1 — Constant vocabulary
# ---------------------------------------------------------------------------

class TestR14SplitConstant:
    """R14 S1/S5: ERROR_TIPO_URGENCIAS is 6 entries, split, no combined."""

    def test_error_tipo_urgencias_is_six(self):
        from app.constants import ERROR_TIPO_URGENCIAS

        assert len(ERROR_TIPO_URGENCIAS) == 6

    def test_error_tipo_urgencias_exact_order(self):
        from app.constants import ERROR_TIPO_URGENCIAS

        assert ERROR_TIPO_URGENCIAS == EXPECTED

    def test_contains_factura_title_case(self):
        from app.constants import ERROR_TIPO_URGENCIAS

        assert "Factura" in ERROR_TIPO_URGENCIAS

    def test_contains_furips_upper(self):
        from app.constants import ERROR_TIPO_URGENCIAS

        assert "FURIPS" in ERROR_TIPO_URGENCIAS

    def test_not_contains_combined(self):
        from app.constants import ERROR_TIPO_URGENCIAS

        assert "Factura y Furips" not in ERROR_TIPO_URGENCIAS

    def test_legacy_four_preserved_order(self):
        from app.constants import ERROR_TIPO_URGENCIAS

        assert ERROR_TIPO_URGENCIAS[:4] == ["Otros", "Soportes de Carpeta", "Factura Abierta", "Carpeta no entregada"]


# ---------------------------------------------------------------------------
# 1.2 — get_opciones()
# ---------------------------------------------------------------------------

class TestR14SplitOpciones:
    """R14 S1: get_opciones().tipos_error exposes Factura + FURIPS, omits combined."""

    def test_opciones_contains_both_separate(self):
        with _APP.test_request_context(), patch(
            "app.services.control_errores_service.users_store.get_facturadores", return_value=[]
        ):
            data = get_opciones()["data"]
        assert "Factura" in data["tipos_error"]
        assert "FURIPS" in data["tipos_error"]
        assert "Factura y Furips" not in data["tipos_error"]

    def test_opciones_exact_six_ordered(self):
        with _APP.test_request_context(), patch(
            "app.services.control_errores_service.users_store.get_facturadores", return_value=[]
        ):
            tipos = get_opciones()["data"]["tipos_error"]
        assert tipos == EXPECTED
        assert len(tipos) == 6

    def test_opciones_keeps_legacy_four(self):
        with _APP.test_request_context(), patch(
            "app.services.control_errores_service.users_store.get_facturadores", return_value=[]
        ):
            tipos = get_opciones()["data"]["tipos_error"]
        assert tipos[:4] == ["Otros", "Soportes de Carpeta", "Factura Abierta", "Carpeta no entregada"]


# ---------------------------------------------------------------------------
# 1.2b — route GET /api/control-errores/opciones
# ---------------------------------------------------------------------------

class TestR14SplitOpcionesRoute:
    def test_route_opciones_contains_split(self, app_client):
        with app_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["rol"] = "validador"
            sess["username"] = "val1"
            sess["permisos"] = ["control_urgencias", "control_urgencias:write"]
        with patch("app.services.control_errores_service.users_store.get_facturadores", return_value=[]):
            resp = app_client.get("/api/control-errores/opciones")
        assert resp.status_code == 200
        tipos = resp.get_json()["data"]["tipos_error"]
        assert "Factura" in tipos
        assert "FURIPS" in tipos
        assert "Factura y Furips" not in tipos
        assert tipos == EXPECTED


# ---------------------------------------------------------------------------
# 1.3 — POST/PUT verbatim
# ---------------------------------------------------------------------------

class TestR14SplitCreateUpdate:
    """R14 S2/S3: POST and PUT persist tipo_error verbatim (casing preserved)."""

    def test_post_factura_persists_verbatim(self, app_client):
        with app_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["control_urgencias:write"]
            sess["username"] = "val1"
            sess["primer_nombre"] = "Juan"
            sess["apellido_1"] = "Perez"
        with patch("app.services.control_errores_service.crear_error") as mock_crear:
            mock_crear.return_value = {"id": "new-1", "tipo_error": "Factura", "factura": "FEV-001"}
            resp = app_client.post(
                "/api/control-errores",
                json={"tipo_error": "Factura", "factura": "FEV-001", "responsable": "LORENY ESPAÑA", "observacion": "x"},
            )
        assert resp.status_code == 200
        assert resp.get_json()["data"]["error"]["tipo_error"] == "Factura"
        assert mock_crear.call_args.args[0] == "Factura"
        # vocabulary must contain it
        from app.constants import ERROR_TIPO_URGENCIAS

        assert "Factura" in ERROR_TIPO_URGENCIAS

    def test_post_furips_persists_upper_verbatim(self, app_client):
        with app_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["control_urgencias:write"]
            sess["username"] = "val1"
            sess["primer_nombre"] = "Juan"
            sess["apellido_1"] = "Perez"
        with patch("app.services.control_errores_service.crear_error") as mock_crear:
            mock_crear.return_value = {"id": "new-2", "tipo_error": "FURIPS", "factura": "FEV-002"}
            resp = app_client.post(
                "/api/control-errores",
                json={"tipo_error": "FURIPS", "factura": "FEV-002", "responsable": "LORENY ESPAÑA", "observacion": "x"},
            )
        assert resp.status_code == 200
        assert resp.get_json()["data"]["error"]["tipo_error"] == "FURIPS"
        assert mock_crear.call_args.args[0] == "FURIPS"
        from app.constants import ERROR_TIPO_URGENCIAS

        assert "FURIPS" in ERROR_TIPO_URGENCIAS

    def test_put_otros_to_furips_verbatim(self, app_client):
        def _fake():
            return {"id": "test-i1", "estado": "S", "tipo_error": "Otros", "observacion": "pac", "observacion_facturador": "", "factura": "FAC-001", "responsable": ""}

        with app_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["control_urgencias:write"]
            sess["username"] = "val1"
        with (
            patch("app.services.control_errores_service.obtener_error", return_value=_fake()),
            patch("app.services.control_errores_service.actualizar_error", return_value={"id": "test-i1", "tipo_error": "FURIPS"}) as mock_upd,
        ):
            resp = app_client.put("/api/control-errores/test-i1", json={"tipo_error": "FURIPS"})
        assert resp.status_code == 200
        assert resp.get_json()["data"]["error"]["tipo_error"] == "FURIPS"
        assert mock_upd.call_args.kwargs["tipo_error"] == "FURIPS"


# ---------------------------------------------------------------------------
# 1.4 — Filter exact case-sensitive
# ---------------------------------------------------------------------------

class TestR14SplitFilter:
    """R14 S4/S5: ?tipo_error= exact, case-sensitive per literal."""

    def _fixture(self):
        return [
            {"id": "factura-1", "tipo_error": "Factura", "estado": "S", "responsable": "A", "creado_en": "2026-08-10T10:00:00"},
            {"id": "furips-1", "tipo_error": "FURIPS", "estado": "S", "responsable": "A", "creado_en": "2026-08-10T10:00:00"},
            {"id": "otros-1", "tipo_error": "Otros", "estado": "S", "responsable": "A", "creado_en": "2026-08-09T10:00:00"},
            {"id": "abierta-1", "tipo_error": "Factura Abierta", "estado": "S", "responsable": "A", "creado_en": "2026-08-08T10:00:00"},
        ]

    def test_filter_factura_exact(self):
        with _APP.test_request_context(), patch(
            "app.utils.errores_storage._leer_datos", return_value={"errores": self._fixture()}
        ), patch("app.utils.errores_storage.obtener_imagenes_count", return_value=0):
            result = get_errores(tipo_error="Factura", session={"rol": "validador"})
        assert [e["id"] for e in result["data"]["errores"]] == ["factura-1"]

    def test_filter_furips_exact(self):
        with _APP.test_request_context(), patch(
            "app.utils.errores_storage._leer_datos", return_value={"errores": self._fixture()}
        ), patch("app.utils.errores_storage.obtener_imagenes_count", return_value=0):
            result = get_errores(tipo_error="FURIPS", session={"rol": "validador"})
        assert [e["id"] for e in result["data"]["errores"]] == ["furips-1"]

    def test_filter_factura_lowercase_zero(self):
        with _APP.test_request_context(), patch(
            "app.utils.errores_storage._leer_datos", return_value={"errores": self._fixture()}
        ), patch("app.utils.errores_storage.obtener_imagenes_count", return_value=0):
            result = get_errores(tipo_error="factura", session={"rol": "validador"})
        assert result["data"]["errores"] == []

    def test_filter_furips_lowercase_zero(self):
        with _APP.test_request_context(), patch(
            "app.utils.errores_storage._leer_datos", return_value={"errores": self._fixture()}
        ), patch("app.utils.errores_storage.obtener_imagenes_count", return_value=0):
            result = get_errores(tipo_error="furips", session={"rol": "validador"})
        assert result["data"]["errores"] == []

    def test_filter_combined_zero(self):
        with _APP.test_request_context(), patch(
            "app.utils.errores_storage._leer_datos", return_value={"errores": self._fixture()}
        ), patch("app.utils.errores_storage.obtener_imagenes_count", return_value=0):
            result = get_errores(tipo_error="Factura y Furips", session={"rol": "validador"})
        assert result["data"]["errores"] == []

    def test_filter_route_factura_and_furips(self, app_client):
        with app_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["rol"] = "validador"
            sess["username"] = "val1"
            sess["permisos"] = ["control_urgencias", "control_urgencias:write"]
        with patch("app.utils.errores_storage._leer_datos", return_value={"errores": self._fixture()}), patch(
            "app.utils.errores_storage.obtener_imagenes_count", return_value=0
        ):
            resp = app_client.get("/api/control-errores?tipo_error=Factura")
            assert [e["id"] for e in resp.get_json()["data"]["errores"]] == ["factura-1"]
            resp2 = app_client.get("/api/control-errores?tipo_error=FURIPS")
            assert [e["id"] for e in resp2.get_json()["data"]["errores"]] == ["furips-1"]
            resp3 = app_client.get("/api/control-errores?tipo_error=factura")
            assert resp3.get_json()["data"]["errores"] == []
            resp4 = app_client.get("/api/control-errores?tipo_error=Factura y Furips")
            assert resp4.get_json()["data"]["errores"] == []


# ---------------------------------------------------------------------------
# 1.5 — Export verbatim
# ---------------------------------------------------------------------------

class TestR14SplitExport:
    """R14 S6: export writes tipo_error verbatim, no normalization."""

    def _patch_images(self):
        return patch("app.services.control_errores_export.listar_imagenes", return_value=[])

    def test_export_factura_verbatim(self):
        with self._patch_images(), _APP.app_context():
            buf = build_errores_export_workbook([_error_fixture(tipo_error="Factura")], "http://testserver/")
        wb = load_workbook(BytesIO(buf.read()))
        ws = wb.active
        assert ws.cell(row=2, column=5).value == "Factura"
        assert [c.value for c in ws[1]][4] == "Categoría"

    def test_export_furips_verbatim_upper(self):
        with self._patch_images(), _APP.app_context():
            buf = build_errores_export_workbook([_error_fixture(tipo_error="FURIPS")], "http://testserver/")
        wb = load_workbook(BytesIO(buf.read()))
        ws = wb.active
        assert ws.cell(row=2, column=5).value == "FURIPS"

    def test_export_both_rows_verbatim(self):
        with self._patch_images(), _APP.app_context():
            buf = build_errores_export_workbook(
                [_error_fixture(error_id="e1", tipo_error="Factura"), _error_fixture(error_id="e2", tipo_error="FURIPS")],
                "http://testserver/",
            )
        wb = load_workbook(BytesIO(buf.read()))
        ws = wb.active
        assert ws.cell(row=2, column=5).value == "Factura"
        assert ws.cell(row=3, column=5).value == "FURIPS"

    def test_export_does_not_normalize_casing(self):
        """lowercase inputs stay lowercase — export is verbatim, not Title/upper."""
        with self._patch_images(), _APP.app_context():
            buf = build_errores_export_workbook([_error_fixture(tipo_error="factura")], "http://testserver/")
        wb = load_workbook(BytesIO(buf.read()))
        ws = wb.active
        assert ws.cell(row=2, column=5).value == "factura"
