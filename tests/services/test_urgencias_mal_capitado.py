"""Tests for tipo_factura filter on app/services/urgencias/mal_capitado.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.constants.odontologia import CODIGOS_MAL_CAPITADO
from app.services.urgencias.mal_capitado import detect_mal_capitado


@pytest.fixture
def wb_with_tipo_factura() -> Workbook:
    """Workbook with headers including tipo_factura_descripcion."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código")
    ws.cell(row=1, column=3, value="Procedimiento")
    ws.cell(row=1, column=4, value="IDE Contrato")
    ws.cell(row=1, column=5, value="Cód Entidad Cobrar")
    ws.cell(row=1, column=6, value="Tipo Factura Descripción")
    return wb


_INDICES_FULL = {
    "numero_factura": 0,
    "codigo": 1,
    "procedimiento": 2,
    "ide_contrato": 3,
    "codigo_entidad_cobrar": 4,
    "tipo_factura_descripcion": 5,
}


class TestMalCapitadoTipoFacturaFilter:
    """Verify detect_mal_capitado filters by tipo_factura_descripcion."""

    def test_filtra_solo_urgencias(
        self, wb_with_tipo_factura: Workbook
    ) -> None:
        """Urgencias rows with bad codes → detected; Hospitalización rows → skipped."""
        ws = wb_with_tipo_factura.active
        bad_code = next(iter(CODIGOS_MAL_CAPITADO))  # e.g. "G03XB01"

        # Row 2: Urgencias, bad code, no FEV prefix → should be detected
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value=bad_code)
        ws.cell(row=2, column=3, value="Procedimiento X")
        ws.cell(row=2, column=4, value="IDE001")
        ws.cell(row=2, column=5, value="ESS118")
        ws.cell(row=2, column=6, value="Urgencias")

        # Row 3: Hospitalización, same bad code, no FEV prefix → should be SKIPPED
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value=bad_code)
        ws.cell(row=3, column=3, value="Procedimiento Y")
        ws.cell(row=3, column=4, value="IDE002")
        ws.cell(row=3, column=5, value="ESS118")
        ws.cell(row=3, column=6, value="Hospitalización")

        result = detect_mal_capitado(ws, _INDICES_FULL)

        # Only the Urgencias row should be detected
        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"

    def test_sin_urgencias_retorna_vacio(
        self, wb_with_tipo_factura: Workbook
    ) -> None:
        """No Urgencias rows → empty result even if bad codes exist."""
        ws = wb_with_tipo_factura.active
        bad_code = next(iter(CODIGOS_MAL_CAPITADO))

        ws.cell(row=2, column=1, value="FAC-003")
        ws.cell(row=2, column=2, value=bad_code)
        ws.cell(row=2, column=3, value="Proc Z")
        ws.cell(row=2, column=4, value="IDE003")
        ws.cell(row=2, column=5, value="ESS118")
        ws.cell(row=2, column=6, value="Intramural")

        result = detect_mal_capitado(ws, _INDICES_FULL)
        assert result == []

    def test_missing_tipo_factura_col_returns_empty(
        self, wb_with_tipo_factura: Workbook
    ) -> None:
        """Missing tipo_factura_descripcion column → return [].

        Since tipo_factura_descripcion is now REQUIRED for filtering,
        the detector should return empty when it's missing.
        """
        ws = wb_with_tipo_factura.active
        bad_code = next(iter(CODIGOS_MAL_CAPITADO))

        ws.cell(row=2, column=1, value="FAC-004")
        ws.cell(row=2, column=2, value=bad_code)

        indices_no_tipo = {
            "numero_factura": 0,
            "codigo": 1,
        }
        result = detect_mal_capitado(ws, indices_no_tipo)
        assert result == []

    def test_urgencias_con_fev_prefix_no_error(
        self, wb_with_tipo_factura: Workbook
    ) -> None:
        """Urgencias with FEV prefix → no error (code is fine)."""
        ws = wb_with_tipo_factura.active
        bad_code = next(iter(CODIGOS_MAL_CAPITADO))

        ws.cell(row=2, column=1, value="FEV12345")
        ws.cell(row=2, column=2, value=bad_code)
        ws.cell(row=2, column=3, value="Proc OK")
        ws.cell(row=2, column=4, value="IDE004")
        ws.cell(row=2, column=5, value="ESS118")
        ws.cell(row=2, column=6, value="Urgencias")

        result = detect_mal_capitado(ws, _INDICES_FULL)
        assert result == []

    def test_cap_prefix_rule_with_urgencias(
        self, wb_with_tipo_factura: Workbook
    ) -> None:
        """CAP prefix without ESS118 → detected only for Urgencias rows."""
        ws = wb_with_tipo_factura.active

        # Row 2: Urgencias, CAP prefix, wrong entidad → should be detected
        ws.cell(row=2, column=1, value="CAP12345")
        ws.cell(row=2, column=2, value="890101")
        ws.cell(row=2, column=3, value="Proc CAP")
        ws.cell(row=2, column=4, value="IDE005")
        ws.cell(row=2, column=5, value="ESS999")
        ws.cell(row=2, column=6, value="Urgencias")

        # Row 3: Hospitalización, CAP prefix, wrong entidad → should be SKIPPED
        ws.cell(row=3, column=1, value="CAP67890")
        ws.cell(row=3, column=2, value="890101")
        ws.cell(row=3, column=3, value="Proc CAP Hosp")
        ws.cell(row=3, column=4, value="IDE006")
        ws.cell(row=3, column=5, value="ESS999")
        ws.cell(row=3, column=6, value="Hospitalización")

        result = detect_mal_capitado(ws, _INDICES_FULL)

        # Only the Urgencias CAP row should be detected
        assert len(result) == 1
        assert result[0]["factura"] == "CAP12345"
