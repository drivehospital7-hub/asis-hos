"""Tests para app/services/urgencias/ide_contrato_urgencias.py.

Strict TDD: tests written BEFORE implementation.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.urgencias.ide_contrato_urgencias import detect_ide_contrato_urgencias


@pytest.fixture
def workbook_urgencias_ide_contrato() -> Workbook:
    """Crea un workbook con headers para IDE Contrato de Urgencias."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código Entidad Cobrar")
    ws.cell(row=1, column=3, value="Código")
    ws.cell(row=1, column=4, value="IDE Contrato")
    return wb


# Rule used in tests: codigo 906340 + entidad EPSI05 → IDE debe ser 986


class TestDetectIdeContratoUrgencias:
    """Tests para detect_ide_contrato_urgencias — reglas básicas."""

    def test_ide_correcto_no_genera_error(
        self, workbook_urgencias_ide_contrato: Workbook
    ) -> None:
        """906340 + EPSI05 con IDE 986 es válido → 0 errores."""
        ws = workbook_urgencias_ide_contrato.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="EPSI05")
        ws.cell(row=2, column=3, value="906340")
        ws.cell(row=2, column=4, value="986")

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_urgencias(ws, indices)

        assert len(result) == 0

    def test_ide_incorrecto_genera_error(
        self, workbook_urgencias_ide_contrato: Workbook
    ) -> None:
        """906340 + EPSI05 con IDE 999 es inválido → 1 error."""
        ws = workbook_urgencias_ide_contrato.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="EPSI05")
        ws.cell(row=2, column=3, value="906340")
        ws.cell(row=2, column=4, value="999")

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_urgencias(ws, indices)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"
        assert result[0]["codigo"] == "906340"
        assert result[0]["entidad"] == "EPSI05"
        assert result[0]["ide_contrato_deberia"] == "986"

    def test_sin_indices_retorna_vacio(
        self, workbook_urgencias_ide_contrato: Workbook
    ) -> None:
        """Si faltan índices obligatorios, retorna lista vacía."""
        ws = workbook_urgencias_ide_contrato.active
        result = detect_ide_contrato_urgencias(ws, {})
        assert result == []

    def test_con_indices_parciales_retorna_vacio(
        self, workbook_urgencias_ide_contrato: Workbook
    ) -> None:
        """Si falta codigo_entidad_cobrar (None), retorna lista vacía."""
        ws = workbook_urgencias_ide_contrato.active
        indices = {"numero_factura": 0, "codigo_entidad_cobrar": None,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_urgencias(ws, indices)
        assert result == []


class TestDedupIdeContratoUrgencias:
    """Tests para deduplicación de errores de contrato por factura (R1) en urgencias."""

    def test_misma_factura_tres_filas_un_error(
        self, workbook_urgencias_ide_contrato: Workbook
    ) -> None:
        """Misma factura con 3 filas violando la regla → exactamente 1 error.

        Spec R1 Happy path: 12 rows all violating → exactly 1 contract error.
        """
        ws = workbook_urgencias_ide_contrato.active
        for row in [2, 3, 4]:
            ws.cell(row=row, column=1, value="FAC-001")
            ws.cell(row=row, column=2, value="EPSI05")
            ws.cell(row=row, column=3, value="906340")
            ws.cell(row=row, column=4, value="999")  # Inválido

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_urgencias(ws, indices)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"

    def test_dos_facturas_distintas_dos_errores(
        self, workbook_urgencias_ide_contrato: Workbook
    ) -> None:
        """2 facturas distintas con errores → 2 errores (1 por factura).

        Spec R1 Mixed invoices: 2 invoices → exactly 2 contract errors.
        """
        ws = workbook_urgencias_ide_contrato.active
        # FAC-001: 2 filas violando
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="EPSI05")
        ws.cell(row=2, column=3, value="906340")
        ws.cell(row=2, column=4, value="999")

        ws.cell(row=3, column=1, value="FAC-001")
        ws.cell(row=3, column=2, value="EPSI05")
        ws.cell(row=3, column=3, value="906340")
        ws.cell(row=3, column=4, value="999")

        # FAC-002: 2 filas violando
        ws.cell(row=4, column=1, value="FAC-002")
        ws.cell(row=4, column=2, value="EPSI05")
        ws.cell(row=4, column=3, value="906340")
        ws.cell(row=4, column=4, value="999")

        ws.cell(row=5, column=1, value="FAC-002")
        ws.cell(row=5, column=2, value="EPSI05")
        ws.cell(row=5, column=3, value="906340")
        ws.cell(row=5, column=4, value="999")

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_urgencias(ws, indices)

        assert len(result) == 2
        facturas = {r["factura"] for r in result}
        assert facturas == {"FAC-001", "FAC-002"}

    def test_sin_violaciones_cero_errores(
        self, workbook_urgencias_ide_contrato: Workbook
    ) -> None:
        """Factura sin violaciones de contrato → 0 errores (regression).

        Spec R1 No violations: all rows correct → 0 errors.
        """
        ws = workbook_urgencias_ide_contrato.active
        for row in [2, 3, 4]:
            ws.cell(row=row, column=1, value="FAC-001")
            ws.cell(row=row, column=2, value="EPSI05")
            ws.cell(row=row, column=3, value="906340")
            ws.cell(row=row, column=4, value="986")  # Correcto

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_urgencias(ws, indices)

        assert len(result) == 0

    def test_factura_sin_error_no_contamina_otras(
        self, workbook_urgencias_ide_contrato: Workbook
    ) -> None:
        """Factura sin error + factura con error → solo 1 error (triangulation).

        Spec R1: verify no cross-invoice contamination when one
        invoice has no violations and another does.
        """
        ws = workbook_urgencias_ide_contrato.active
        # FAC-001: 2 filas SIN error
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="EPSI05")
        ws.cell(row=2, column=3, value="906340")
        ws.cell(row=2, column=4, value="986")

        ws.cell(row=3, column=1, value="FAC-001")
        ws.cell(row=3, column=2, value="EPSI05")
        ws.cell(row=3, column=3, value="906340")
        ws.cell(row=3, column=4, value="986")

        # FAC-002: 2 filas con error
        ws.cell(row=4, column=1, value="FAC-002")
        ws.cell(row=4, column=2, value="EPSI05")
        ws.cell(row=4, column=3, value="906340")
        ws.cell(row=4, column=4, value="999")

        ws.cell(row=5, column=1, value="FAC-002")
        ws.cell(row=5, column=2, value="EPSI05")
        ws.cell(row=5, column=3, value="906340")
        ws.cell(row=5, column=4, value="999")

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_urgencias(ws, indices)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-002"

    def test_fila_sin_factura_se_ignora(
        self, workbook_urgencias_ide_contrato: Workbook
    ) -> None:
        """Fila sin número de factura se ignora, no crash (R3).

        Spec R3: Missing invoice → no error, no crash.
        """
        ws = workbook_urgencias_ide_contrato.active
        # Fila sin factura (None)
        ws.cell(row=2, column=1, value=None)
        ws.cell(row=2, column=2, value="EPSI05")
        ws.cell(row=2, column=3, value="906340")
        ws.cell(row=2, column=4, value="999")

        indices = {"numero_factura": 0, "codigo_entidad_cobrar": 1,
                   "codigo": 2, "ide_contrato": 3}
        result = detect_ide_contrato_urgencias(ws, indices)

        assert len(result) == 0
