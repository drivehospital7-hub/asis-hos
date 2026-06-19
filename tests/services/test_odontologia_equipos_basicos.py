"""Tests for EB constants and exporter behavior.

The old /odontologia-equipos-basicos blueprint was replaced by the unified
/procesar route. These tests cover the remaining backend behavior:
EB constants importability and exporter reject of equipos_basicos kwarg.
"""

from __future__ import annotations

import io
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook


# =============================================================================
# Test 4.5: Constants importable
# =============================================================================


class TestConstantsImportable:
    """Spec R4: EB constants MUST reside in app/constants/equipos_basicos.py."""

    def test_import_profesionales_equipos_basicos(self):
        """PROFESIONALES_EQUIPOS_BASICOS importable from app.constants."""
        from app.constants import PROFESIONALES_EQUIPOS_BASICOS

        assert len(PROFESIONALES_EQUIPOS_BASICOS) > 0
        assert "03764" in PROFESIONALES_EQUIPOS_BASICOS
        assert PROFESIONALES_EQUIPOS_BASICOS["03764"]["tipo"] == "ODONTOLOGO"

    def test_import_centro_costo_equipos_basicos(self):
        """CENTRO_COSTO_EQUIPOS_BASICOS importable from app.constants."""
        from app.constants import CENTRO_COSTO_EQUIPOS_BASICOS

        assert CENTRO_COSTO_EQUIPOS_BASICOS == "EQUIPOS BASICOS ODONTOLOGIA"

    def test_import_equipos_basicos_thresholds(self):
        """EB-specific thresholds importable from app.constants."""
        from app.constants import (
            EQUIPOS_BASICOS_RUTA_DUPLICADA_THRESHOLD,
            EQUIPOS_BASICOS_CANTIDAD_CONSULTAS_MIN,
            EQUIPOS_BASICOS_CANTIDAD_MAX,
            EQUIPOS_BASICOS_CANTIDAD_PYP_MIN,
        )

        assert EQUIPOS_BASICOS_RUTA_DUPLICADA_THRESHOLD == 3
        assert EQUIPOS_BASICOS_CANTIDAD_CONSULTAS_MIN == 2
        assert EQUIPOS_BASICOS_CANTIDAD_MAX == 10

    def test_import_odontologia_does_not_have_eb_constants(self):
        """Odontologia module no longer contains EB-specific constants."""
        from app.constants import odontologia

        assert not hasattr(odontologia, "PROFESIONALES_EQUIPOS_BASICOS")

    def test_import_columnas_does_not_have_eb_constants(self):
        """Columnas module no longer contains EB-specific constants."""
        from app.constants import columnas

        assert not hasattr(columnas, "EQUIPOS_BASICOS_COLUMNS_TO_KEEP")
        assert not hasattr(columnas, "EQUIPOS_BASICOS_REVISION_HEADERS")
        assert not hasattr(columnas, "CENTRO_COSTO_EQUIPOS_BASICOS")


# =============================================================================
# Test 4.4: exporter.py rejects equipos_basicos kwarg
# =============================================================================


class TestExporterRejectsEquiposBasicosKwarg:
    """Spec R3: equipos_basicos param removed from exporter signature."""

    def test_detect_problems_only_no_equipos_basicos_param(self):
        """detect_problems_only raises TypeError if equipos_basicos is passed."""
        from app.services.exporter import detect_problems_only

        with pytest.raises(TypeError):
            detect_problems_only(
                filename="test.xlsx",
                equipos_basicos=True,  # type: ignore
            )

    def test_detect_problems_only_works_without_equipos_basicos(self):
        """detect_problems_only works when called without equipos_basicos."""
        from app.services.exporter import detect_problems_only

        with patch(
            "app.services.exporter.acquire_semaphore",
            return_value=True,
        ):
            with patch(
                "app.services.exporter.release_semaphore",
            ):
                result, status = detect_problems_only(
                    filename="nonexistent.xlsx",
                    area="equipos_basicos",
                )
                # Should get error about path resolution, not TypeError
                assert status == 500 or result["status"] == "error"

    def test_do_detect_problems_rejects_equipos_basicos(self):
        """_do_detect_problems raises TypeError if equipos_basicos is passed."""
        from app.services.exporter import _do_detect_problems

        with pytest.raises(TypeError):
            _do_detect_problems(
                filename="test.xlsx",
                equipos_basicos=True,  # type: ignore
            )

    def test_do_detect_problems_works_with_area_param(self):
        """_do_detect_problems works when area=AREA_EQUIPOS_BASICOS is passed instead."""
        from app.services.exporter import _do_detect_problems

        with patch(
            "app.services.exporter.resolve_safe_excel_absolute",
            return_value=(Path("/tmp/test.xlsx"), None),
        ):
            with patch(
                "app.services.exporter.validate_excel_path",
                return_value=None,
            ):
                import polars as pl
                df = pl.DataFrame({"A": ["header"]})
                with patch(
                    "app.services.exporter.pl.read_excel",
                    return_value=df,
                ):
                    result = _do_detect_problems(
                        filename="test.xlsx",
                        area="equipos_basicos",
                    )
                    # Should proceed to detection (will fail with KeyError since no real indices)
                    assert result["status"] in ("error", "success")


# =============================================================================
# Helper: create a minimal EB Excel file
# =============================================================================


def _make_eb_excel(headers: list[str], rows: list[list]) -> io.BytesIO:
    """Create a real .xlsx in memory with given headers and data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


EB_HEADERS = [
    "Número Factura",
    "Vlr. Subsidiado",
    "Vlr. Procedimiento",
    "Código Tipo Procedimiento",
    "Tipo Procedimiento",
    "Código",
    "Cód. Equivalente CUPS",
    "Procedimiento",
    "Nº Identificación",
    "Convenio Facturado",
    "Cantidad",
    "Laboratorio",
    "Centro Costo",
    "Cód Entidad Cobrar",
    "Entidad Cobrar",
    "Entidad Afiliación",
    "Tipo Factura Descripción",
    "IDE Contrato",
    "Tipo Identificación",
    "Fec. Nacimiento",
    "Fec. Factura",
    "Fecha Cierre",
    "Identificación Profesional",
    "Profesional Atiende",
    "Código Profesional",
    "Responsable Cierra Facturar",
    "Tarifario",
    "Tipo Usuario",
]



