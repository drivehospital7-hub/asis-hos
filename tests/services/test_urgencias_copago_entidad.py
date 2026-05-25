"""Tests para app/services/urgencias/detect_copago_entidad.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.urgencias.detect_copago_entidad import (
    detect_copago_entidad_urgencias,
)


@pytest.fixture
def workbook_with_copago_headers() -> Workbook:
    """Crea un workbook con headers para copago vs entidad."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código")
    ws.cell(row=1, column=3, value="Procedimiento")
    ws.cell(row=1, column=4, value="Cód Entidad Cobrar")
    ws.cell(row=1, column=5, value="Vlr. Copago")
    return wb


class TestDetectCopagoEntidad:
    """Tests para detect_copago_entidad_urgencias."""

    def _build_indices(self, has_vlr_copago: bool = True) -> dict[str, int | None]:
        """Construye dict de índices para las pruebas."""
        indices: dict[str, int | None] = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "codigo_entidad_cobrar": 3,
        }
        if has_vlr_copago:
            indices["vlr_copago"] = 4
        return indices

    # ── Entidad default "1" — nunca genera error ──

    def test_entidad_1_con_copago_0_no_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Entidad='1', copago=0 → sin error."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="890101")
        ws.cell(row=2, column=3, value="Consulta")
        ws.cell(row=2, column=4, value="1")
        ws.cell(row=2, column=5, value=0)

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert result == []

    def test_entidad_1_con_copago_500_no_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Entidad='1', copago=500 → sin error (entidad default)."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="890101")
        ws.cell(row=2, column=3, value="Consulta")
        ws.cell(row=2, column=4, value="1")
        ws.cell(row=2, column=5, value=500)

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert result == []

    def test_entidad_0001_con_copago_500_no_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Entidad='0001', copago=500 → sin error (entidad default)."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-002")
        ws.cell(row=2, column=2, value="890102")
        ws.cell(row=2, column=3, value="Procedimiento")
        ws.cell(row=2, column=4, value="0001")
        ws.cell(row=2, column=5, value=500)

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert result == []

    # ── Entidad no default con copago 0 — sin error ──

    def test_entidad_86_con_copago_0_no_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Entidad='86', copago=0 → sin error."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-003")
        ws.cell(row=2, column=2, value="890103")
        ws.cell(row=2, column=3, value="Otro")
        ws.cell(row=2, column=4, value="86")
        ws.cell(row=2, column=5, value=0)

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert result == []

    def test_entidad_ess118_con_copago_0_no_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Entidad='ESS118', copago=0 → sin error."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-004")
        ws.cell(row=2, column=2, value="890104")
        ws.cell(row=2, column=3, value="Otro")
        ws.cell(row=2, column=4, value="ESS118")
        ws.cell(row=2, column=5, value=0)

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert result == []

    # ── Entidad no default con copago != 0 → error ──

    def test_entidad_86_con_copago_500_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Entidad='86', copago=500 → error: entidad no default y copago no es 0."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-005")
        ws.cell(row=2, column=2, value="890105")
        ws.cell(row=2, column=3, value="Procedimiento 86")
        ws.cell(row=2, column=4, value="86")
        ws.cell(row=2, column=5, value=500)

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert len(result) == 1
        error = result[0]
        assert error["factura"] == "FAC-005"
        assert error["codigo"] == "890105"
        assert error["procedimiento"] == "Procedimiento 86"
        assert error["entidad_cobrar"] == "86"
        assert error["vlr_copago"] == 500.0

    def test_entidad_ess118_con_copago_100_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Entidad='ESS118', copago=100.0 → error."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-006")
        ws.cell(row=2, column=2, value="890106")
        ws.cell(row=2, column=3, value="Procedimiento ESS")
        ws.cell(row=2, column=4, value="ESS118")
        ws.cell(row=2, column=5, value=100.0)

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert len(result) == 1
        assert result[0]["entidad_cobrar"] == "ESS118"
        assert result[0]["vlr_copago"] == 100.0

    # ── Type normalization ──

    def test_entidad_vacia_con_copago_500_no_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Entidad vacía/None, copago=500 → sin error (no aplica regla)."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-007")
        ws.cell(row=2, column=2, value="890107")
        ws.cell(row=2, column=3, value="Proc")
        ws.cell(row=2, column=4, value=None)
        ws.cell(row=2, column=5, value=500)

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert result == []

    def test_entidad_86_con_copago_none_no_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Entidad='86', copago=None → sin error (None tratado como 0)."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-008")
        ws.cell(row=2, column=2, value="890108")
        ws.cell(row=2, column=3, value="Proc")
        ws.cell(row=2, column=4, value="86")
        ws.cell(row=2, column=5, value=None)

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert result == []

    def test_entidad_int_3_con_copago_string_500_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Entidad=3 (int), copago='500' (string) → error (type normalization)."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-009")
        ws.cell(row=2, column=2, value="890109")
        ws.cell(row=2, column=3, value="Proc")
        ws.cell(row=2, column=4, value=3)
        ws.cell(row=2, column=5, value="500")

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert len(result) == 1
        assert result[0]["entidad_cobrar"] == "3"
        assert result[0]["vlr_copago"] == 500.0

    def test_entidad_86_copago_string_0_no_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Entidad='86', copago='0' (string) → sin error."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-010")
        ws.cell(row=2, column=2, value="890110")
        ws.cell(row=2, column=3, value="Proc")
        ws.cell(row=2, column=4, value="86")
        ws.cell(row=2, column=5, value="0")

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert result == []

    # ── Missing column ──

    def test_sin_indices_vlr_copago_retorna_vacio(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Sin índice vlr_copago → retorna vacío, no crash."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-011")
        ws.cell(row=2, column=2, value="890111")
        ws.cell(row=2, column=3, value="Proc")
        ws.cell(row=2, column=4, value="86")
        ws.cell(row=2, column=5, value=500)

        result = detect_copago_entidad_urgencias(ws, self._build_indices(has_vlr_copago=False))
        assert result == []

    # ── Per-row: misma factura, múltiples filas ──

    def test_misma_factura_dos_filas_un_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Misma factura, row1=error, row2=ok → exactamente 1 error."""
        ws = workbook_with_copago_headers.active
        # Fila 2: error
        ws.cell(row=2, column=1, value="FAC-020")
        ws.cell(row=2, column=2, value="890120")
        ws.cell(row=2, column=3, value="Proc A")
        ws.cell(row=2, column=4, value="86")
        ws.cell(row=2, column=5, value=500)
        # Fila 3: ok (entidad default)
        ws.cell(row=3, column=1, value="FAC-020")
        ws.cell(row=3, column=2, value="890121")
        ws.cell(row=3, column=3, value="Proc B")
        ws.cell(row=3, column=4, value="1")
        ws.cell(row=3, column=5, value=300)

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert len(result) == 1
        assert result[0]["codigo"] == "890120"

    def test_misma_factura_dos_filas_dos_errores(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Misma factura, 2 filas con error → 2 errores (per-row)."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-021")
        ws.cell(row=2, column=2, value="890122")
        ws.cell(row=2, column=3, value="Proc A")
        ws.cell(row=2, column=4, value="86")
        ws.cell(row=2, column=5, value=500)
        ws.cell(row=3, column=1, value="FAC-021")
        ws.cell(row=3, column=2, value="890123")
        ws.cell(row=3, column=3, value="Proc B")
        ws.cell(row=3, column=4, value="ESS118")
        ws.cell(row=3, column=5, value=200)

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert len(result) == 2

    # ── Sin datos (solo headers) ──

    def test_sin_datos_retorna_vacio(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Sin filas de datos → retorna vacío."""
        ws = workbook_with_copago_headers.active
        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert result == []

    # ── Entidad "" (string vacío) con copago ≠ 0 ──

    def test_entidad_vacia_string_con_copago_500_no_error(
        self, workbook_with_copago_headers: Workbook
    ) -> None:
        """Entidad='' (string vacío), copago=500 → sin error."""
        ws = workbook_with_copago_headers.active
        ws.cell(row=2, column=1, value="FAC-022")
        ws.cell(row=2, column=2, value="890122")
        ws.cell(row=2, column=3, value="Proc")
        ws.cell(row=2, column=4, value="")
        ws.cell(row=2, column=5, value=500)

        result = detect_copago_entidad_urgencias(ws, self._build_indices())
        assert result == []
