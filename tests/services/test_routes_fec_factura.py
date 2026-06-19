"""Integration tests: routes include fec_factura in JSON response.

Strict TDD: tests written BEFORE implementation.
Covers Spec R3: Backend Response — Column Consistency.
"""

from __future__ import annotations

import io
import json

from openpyxl import Workbook


# =============================================================================
# Helpers
# =============================================================================


def _make_excel(headers: list[str], rows: list[list]) -> io.BytesIO:
    """Create a minimal .xlsx in memory."""
    wb = Workbook()
    ws = wb.active
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# Odontología headers (complete set required for processing)
ODONTO_HEADERS = [
    "Número Factura",
    "Fec. Factura",
    "Vlr. Procedimiento",
    "Vlr. Subsidiado",
    "Procedimiento",
    "Nº Identificación",
    "Convenio Facturado",
    "Cantidad",
    "Centro Costo",
    "Cód Entidad Cobrar",
    "Entidad Cobrar",
    "Entidad Afiliación",
    "IDE Contrato",
    "Tipo Identificación",
    "Fec. Nacimiento",
    "Fecha Cierre",
    "Identificación Profesional",
    "Profesional Atiende",
    "Código Profesional",
    "Responsable Cierra Facturar",
    "Código Tipo Procedimiento",
    "Tipo Procedimiento",
    "Código",
    "Cód. Equivalente CUPS",
    "Laboratorio",
    "Tipo Factura Descripción",
    "Tarifario",
    "Tipo Usuario",
    "Vlr. Copago",
    "Nº Reingreso",
    "Cód. Dx Principal",
]

URGENCIAS_HEADERS = ODONTO_HEADERS[:]  # Same column set for urgencias


# =============================================================================
# Test: Odontología route
# =============================================================================


class TestUnifiedRouteFecFactura:
    """Spec R3: POST /procesar/ MUST include fec_factura in JSON."""

    def _make_data_row(self, fac_num="FAC-001", fec_fact="2024-01-15", centro="ODONTOLOGIA", resp_cierra="Resp"):
        """Build a full data row matching ODONTO_HEADERS length."""
        row = [""] * len(ODONTO_HEADERS)
        header_map = {h: i for i, h in enumerate(ODONTO_HEADERS)}
        row[header_map["Número Factura"]] = fac_num
        row[header_map["Fec. Factura"]] = fec_fact
        row[header_map["Vlr. Procedimiento"]] = "15000"
        row[header_map["Vlr. Subsidiado"]] = "0"
        row[header_map["Procedimiento"]] = "Consulta"
        row[header_map["Nº Identificación"]] = "12345"
        row[header_map["Convenio Facturado"]] = "Asistencial"
        row[header_map["Cantidad"]] = "1"
        row[header_map["Centro Costo"]] = centro
        row[header_map["Cód Entidad Cobrar"]] = "ESS118"
        row[header_map["Entidad Cobrar"]] = "ESS118"
        row[header_map["Entidad Afiliación"]] = "ESS118"
        row[header_map["IDE Contrato"]] = "970"
        row[header_map["Tipo Identificación"]] = "CC"
        row[header_map["Fec. Nacimiento"]] = "1990-01-01"
        row[header_map["Identificación Profesional"]] = "123"
        row[header_map["Profesional Atiende"]] = "Dr Test"
        row[header_map["Código Profesional"]] = "03764"
        row[header_map["Responsable Cierra Facturar"]] = resp_cierra
        row[header_map["Código Tipo Procedimiento"]] = "ODT"
        row[header_map["Tipo Procedimiento"]] = "ODT"
        row[header_map["Código"]] = "890203"
        row[header_map["Tarifario"]] = "Tarifario 1"
        row[header_map["Tipo Usuario"]] = "SUBSIDIADO"
        return row

    def _make_urgencias_row(self):
        """Build a full data row for urgencias context."""
        row = [""] * len(ODONTO_HEADERS)
        header_map = {h: i for i, h in enumerate(ODONTO_HEADERS)}
        row[header_map["Número Factura"]] = "FAC-001"
        row[header_map["Fec. Factura"]] = "2024-01-15"
        row[header_map["Vlr. Procedimiento"]] = "15000"
        row[header_map["Vlr. Subsidiado"]] = "0"
        row[header_map["Procedimiento"]] = "Consulta"
        row[header_map["Nº Identificación"]] = "12345"
        row[header_map["Convenio Facturado"]] = "Asistencial"
        row[header_map["Cantidad"]] = "1"
        row[header_map["Centro Costo"]] = "ODONTOLOGIA"
        row[header_map["Cód Entidad Cobrar"]] = "ESS118"
        row[header_map["Entidad Cobrar"]] = "ESS118"
        row[header_map["Entidad Afiliación"]] = "ESS118"
        row[header_map["IDE Contrato"]] = "970"
        row[header_map["Tipo Identificación"]] = "CC"
        row[header_map["Fec. Nacimiento"]] = "1990-01-01"
        row[header_map["Identificación Profesional"]] = "123"
        row[header_map["Profesional Atiende"]] = "Dr Test"
        row[header_map["Código Profesional"]] = "03764"
        row[header_map["Responsable Cierra Facturar"]] = "Resp"
        row[header_map["Código Tipo Procedimiento"]] = "ODT"
        row[header_map["Tipo Procedimiento"]] = "ODT"
        row[header_map["Código"]] = "890203"
        row[header_map["Tarifario"]] = "Tarifario 1"
        row[header_map["Tipo Usuario"]] = "SUBSIDIADO"
        return row

    def _make_eb_row(self):
        """Build a full data row for equipos basicos context."""
        row = [""] * len(ODONTO_HEADERS)
        header_map = {h: i for i, h in enumerate(ODONTO_HEADERS)}
        row[header_map["Número Factura"]] = "FAC-001"
        row[header_map["Fec. Factura"]] = "2024-01-15"
        row[header_map["Vlr. Procedimiento"]] = "15000"
        row[header_map["Vlr. Subsidiado"]] = "0"
        row[header_map["Procedimiento"]] = "Consulta"
        row[header_map["Nº Identificación"]] = "12345"
        row[header_map["Convenio Facturado"]] = "Asistencial"
        row[header_map["Cantidad"]] = "1"
        row[header_map["Centro Costo"]] = "EQUIPOS BASICOS ODONTOLOGIA"
        row[header_map["Cód Entidad Cobrar"]] = "ESS118"
        row[header_map["Entidad Cobrar"]] = "ESS118"
        row[header_map["Entidad Afiliación"]] = "ESS118"
        row[header_map["IDE Contrato"]] = "970"
        row[header_map["Tipo Identificación"]] = "CC"
        row[header_map["Fec. Nacimiento"]] = "1990-01-01"
        row[header_map["Identificación Profesional"]] = "123"
        row[header_map["Profesional Atiende"]] = "Dr Test"
        row[header_map["Código Profesional"]] = "03764"
        row[header_map["Responsable Cierra Facturar"]] = "Resp"
        row[header_map["Código Tipo Procedimiento"]] = "ODT"
        row[header_map["Tipo Procedimiento"]] = "ODT"
        row[header_map["Código"]] = "890203"
        row[header_map["Tarifario"]] = "Tarifario 1"
        row[header_map["Tipo Usuario"]] = "SUBSIDIADO"
        return row

    def test_columnas_starts_with_fec_factura_odonto(self, fresh_client):
        """columnas[0] MUST be 'Fec. Factura' (odontologia context)."""
        with fresh_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["*"]
            sess["username"] = "admin"

        buf = _make_excel(ODONTO_HEADERS, [self._make_data_row(centro="ODONTOLOGIA")])
        resp = fresh_client.post(
            "/procesar/",
            data={"file_upload": (buf, "test_odonto.xlsx")},
        )
        assert resp.status_code == 200
        data = json.loads(resp.data.decode("utf-8"))
        assert data.get("status") == "success", f"Expected success, got: {data.get('errors', 'no errors')}"
        columnas = data["data"].get("columnas", [])
        assert len(columnas) > 0
        assert columnas[0] == "Fec. Factura"

    def test_all_items_include_fec_factura_odonto(self, fresh_client):
        """Every item MUST have 'fec_factura' key (odontologia context)."""
        with fresh_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["*"]
            sess["username"] = "admin"

        buf = _make_excel(ODONTO_HEADERS, [self._make_data_row(centro="ODONTOLOGIA")])
        resp = fresh_client.post(
            "/procesar/",
            data={"file_upload": (buf, "test_odonto.xlsx")},
        )
        assert resp.status_code == 200
        data = json.loads(resp.data.decode("utf-8"))
        assert data.get("status") == "success", f"Expected success, got: {data.get('errors', 'no errors')}"
        for grupo in data["data"].get("errores", []):
            for item in grupo.get("facturas", []):
                assert "fec_factura" in item

    def test_columnas_starts_with_fec_factura_urgencias(self, fresh_client):
        """columnas[0] MUST be 'Fec. Factura' (urgencias context)."""
        with fresh_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["*"]
            sess["username"] = "admin"

        buf = _make_excel(ODONTO_HEADERS, [self._make_urgencias_row()])
        resp = fresh_client.post(
            "/procesar/",
            data={"file_upload": (buf, "test_urgencias.xlsx")},
        )
        assert resp.status_code == 200
        data = json.loads(resp.data.decode("utf-8"))
        assert data.get("status") == "success", f"Expected success, got: {data.get('errors', 'no errors')}"
        columnas = data["data"].get("columnas", [])
        assert len(columnas) > 0
        assert columnas[0] == "Fec. Factura"

    def test_all_items_include_fec_factura_urgencias(self, fresh_client):
        """Every item MUST have 'fec_factura' key (urgencias context)."""
        with fresh_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["*"]
            sess["username"] = "admin"

        buf = _make_excel(ODONTO_HEADERS, [self._make_urgencias_row()])
        resp = fresh_client.post(
            "/procesar/",
            data={"file_upload": (buf, "test_urgencias.xlsx")},
        )
        assert resp.status_code == 200
        data = json.loads(resp.data.decode("utf-8"))
        assert data.get("status") == "success", f"Expected success, got: {data.get('errors', 'no errors')}"
        for grupo in data["data"].get("errores", []):
            for item in grupo.get("facturas", []):
                assert "fec_factura" in item

    def test_columnas_starts_with_fec_factura_eb(self, fresh_client):
        """columnas[0] MUST be 'Fec. Factura' (equipos basicos context)."""
        with fresh_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["*"]
            sess["username"] = "admin"

        buf = _make_excel(ODONTO_HEADERS, [self._make_eb_row()])
        resp = fresh_client.post(
            "/procesar/",
            data={"file_upload": (buf, "test_eb.xlsx")},
        )
        assert resp.status_code == 200
        data = json.loads(resp.data.decode("utf-8"))
        assert data.get("status") == "success", f"Expected success, got: {data.get('errors', 'no errors')}"
        columnas = data["data"].get("columnas", [])
        assert len(columnas) > 0
        assert columnas[0] == "Fec. Factura"

    def test_all_items_include_fec_factura_eb(self, fresh_client):
        """Every item MUST have 'fec_factura' key (equipos basicos context)."""
        with fresh_client.session_transaction() as sess:
            sess["ce_authenticated"] = True
            sess["permisos"] = ["*"]
            sess["username"] = "admin"

        buf = _make_excel(ODONTO_HEADERS, [self._make_eb_row()])
        resp = fresh_client.post(
            "/procesar/",
            data={"file_upload": (buf, "test_eb.xlsx")},
        )
        assert resp.status_code == 200
        data = json.loads(resp.data.decode("utf-8"))
        assert data.get("status") == "success", f"Expected success, got: {data.get('errors', 'no errors')}"
        for grupo in data["data"].get("errores", []):
            for item in grupo.get("facturas", []):
                assert "fec_factura" in item
