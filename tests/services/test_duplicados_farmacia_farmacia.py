"""Tests for app/services/farmacia/duplicados_farmacia_farmacia.py.

Cubre escenarios de duplicados para tipo factura Farmacia sin filtros
de tarifario ni codigo_tipo_procedimiento.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.farmacia.duplicados_farmacia_farmacia import detect_duplicados_farmacia_farmacia


@pytest.fixture
def workbook_with_headers() -> Workbook:
    """Crea un workbook con headers mínimos para Farmacia."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código")
    ws.cell(row=1, column=3, value="Cantidad")
    ws.cell(row=1, column=4, value="Tarifario")
    ws.cell(row=1, column=5, value="Procedimiento")
    ws.cell(row=1, column=6, value="Código Tipo Procedimiento")
    ws.cell(row=1, column=7, value="Tipo Factura Descripción")
    return wb


_INDICES_FARMACIA = {
    "numero_factura": 0,
    "codigo": 1,
    "cantidad": 2,
    "tarifario": 3,
    "procedimiento": 4,
    "codigo_tipo_procedimiento": 5,
    "tipo_factura_descripcion": 6,
}


def _write_farmacia_row(ws, row, factura, codigo, cantidad, tipo_factura="Farmacia"):
    """Helper: escribe una fila con tipo_factura Farmacia."""
    ws.cell(row=row, column=1, value=factura)
    ws.cell(row=row, column=2, value=codigo)
    ws.cell(row=row, column=3, value=cantidad)
    ws.cell(row=row, column=4, value="")
    ws.cell(row=row, column=5, value="Med")
    ws.cell(row=row, column=6, value="12")
    ws.cell(row=row, column=7, value=tipo_factura)


class TestDetectDuplicadosFarmaciaFarmacia:
    """Tests para detect_duplicados_farmacia_farmacia."""

    def test_duplicidad_total_retorna_flag(self, workbook_with_headers: Workbook):
        """2 pares, cada uno x2 → 1 flag."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 4, "FAC-001", "A002", 2)
        _write_farmacia_row(ws, 5, "FAC-001", "A002", 2)

        result = detect_duplicados_farmacia_farmacia(ws, _INDICES_FARMACIA)
        assert len(result) == 1
        item = result[0]
        assert item["factura"] == "FAC-001"
        assert item["total_pares"] == 2
        assert len(item["pares_duplicados"]) == 2

    def test_grupo_con_mezcla_no_flag(self, workbook_with_headers: Workbook):
        """Par A x2, par B x1 → NO flag."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 4, "FAC-001", "A002", 2)

        result = detect_duplicados_farmacia_farmacia(ws, _INDICES_FARMACIA)
        assert result == []

    def test_sin_filas_farmacia_retorna_vacio(self, workbook_with_headers: Workbook):
        """Filas tipo Urgencias → []."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1, tipo_factura="Urgencias")
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1, tipo_factura="Urgencias")

        result = detect_duplicados_farmacia_farmacia(ws, _INDICES_FARMACIA)
        assert result == []

    def test_multiples_facturas_independientes(self, workbook_with_headers: Workbook):
        """Dos facturas, solo una con duplicidad total → 1 flag."""
        ws = workbook_with_headers.active
        # FAC-001: duplicidad total
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1)
        # FAC-002: mezcla
        _write_farmacia_row(ws, 4, "FAC-002", "B001", 1)
        _write_farmacia_row(ws, 5, "FAC-002", "B001", 1)
        _write_farmacia_row(ws, 6, "FAC-002", "B002", 2)

        result = detect_duplicados_farmacia_farmacia(ws, _INDICES_FARMACIA)
        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"

    def test_output_no_incluye_codigo_tipo_procedimiento(self, workbook_with_headers: Workbook):
        """Output NO incluye key 'codigo_tipo_procedimiento'."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1)

        result = detect_duplicados_farmacia_farmacia(ws, _INDICES_FARMACIA)
        assert "codigo_tipo_procedimiento" not in result[0]

    def test_missing_numero_factura_retorna_vacio(self, workbook_with_headers: Workbook):
        """Sin columna numero_factura → []."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1)

        indices_no_fact = dict(_INDICES_FARMACIA)
        del indices_no_fact["numero_factura"]
        result = detect_duplicados_farmacia_farmacia(ws, indices_no_fact)
        assert result == []

    def test_sin_datos_retorna_vacio(self, workbook_with_headers: Workbook):
        """Sin filas de datos → []."""
        ws = workbook_with_headers.active
        result = detect_duplicados_farmacia_farmacia(ws, _INDICES_FARMACIA)
        assert result == []

    def test_cantidad_none_tratado_como_cero(self, workbook_with_headers: Workbook):
        """Cantidad None se trata como 0."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", None)
        _write_farmacia_row(ws, 3, "FAC-001", "A001", None)

        result = detect_duplicados_farmacia_farmacia(ws, _INDICES_FARMACIA)
        assert len(result) == 1
        assert result[0]["pares_duplicados"][0]["cantidad"] == 0

    def test_factura_none_se_salta(self, workbook_with_headers: Workbook):
        """Factura None se salta."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, None, "A001", 1)
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 4, "FAC-001", "A001", 1)

        result = detect_duplicados_farmacia_farmacia(ws, _INDICES_FARMACIA)
        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"

    def test_no_filtra_por_tarifario(self, workbook_with_headers: Workbook):
        """Filas Farmacia se evalúan sin importar tarifario."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="A001")
        ws.cell(row=2, column=3, value=1)
        ws.cell(row=2, column=4, value="Honorarios")  # distinto tarifario
        ws.cell(row=2, column=5, value="Med")
        ws.cell(row=2, column=6, value="12")
        ws.cell(row=2, column=7, value="Farmacia")

        ws.cell(row=3, column=1, value="FAC-001")
        ws.cell(row=3, column=2, value="A001")
        ws.cell(row=3, column=3, value=1)
        ws.cell(row=3, column=4, value="Honorarios")
        ws.cell(row=3, column=5, value="Med")
        ws.cell(row=3, column=6, value="12")
        ws.cell(row=3, column=7, value="Farmacia")

        result = detect_duplicados_farmacia_farmacia(ws, _INDICES_FARMACIA)
        assert len(result) == 1

    def test_no_filtra_por_tipo_proc(self, workbook_with_headers: Workbook):
        """Distintos tipo_proc en misma factura cuentan como mismo par."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 4, "FAC-001", "A002", 2, "Farmacia")
        ws.cell(row=4, column=6, value="09")  # mismo par con distinto tipo_proc
        _write_farmacia_row(ws, 5, "FAC-001", "A002", 2, "Farmacia")
        ws.cell(row=5, column=6, value="09")

        result = detect_duplicados_farmacia_farmacia(ws, _INDICES_FARMACIA)
        # Como no se agrupa por tipo_proc, tenemos 2 pares: (A001,1) y (A002,2), ambos x2
        assert len(result) == 1
        assert result[0]["total_pares"] == 2
