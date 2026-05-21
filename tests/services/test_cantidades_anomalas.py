"""Tests para app/services/transversales/cantidades_anomalas.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.constants import CONVENIO_PYP, CONVENIO_ASISTENCIAL
from app.services.transversales.cantidades_anomalas import (
    detect_cantidades_anomalas,
)


@pytest.fixture
def workbook_with_headers() -> Workbook:
    """Crea un workbook con headers mínimos para cantidades anómalas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Tipo Procedimiento")
    ws.cell(row=1, column=3, value="Cantidad")
    ws.cell(row=1, column=4, value="Convenio Facturado")
    return wb


class TestDetectCantidadesAnomalas:
    """Tests para detect_cantidades_anomalas."""

    def test_detecta_consultas_con_cantidad_mayor_igual_2(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe detectar consultas con cantidad >= 2 (default)."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="Consultas")
        ws.cell(row=2, column=3, value=2)
        ws.cell(row=2, column=4, value=CONVENIO_ASISTENCIAL)
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value="Consultas")
        ws.cell(row=3, column=3, value=1)
        ws.cell(row=3, column=4, value=CONVENIO_ASISTENCIAL)

        indices = {
            "numero_factura": 0,
            "tipo_procedimiento": 1,
            "cantidad": 2,
            "convenio_facturado": 3,
        }
        result = detect_cantidades_anomalas(ws, indices)

        facturas = [r["factura"] for r in result]
        assert "FAC-001" in facturas
        assert "FAC-002" not in facturas

    def test_detecta_cantidad_mayor_10(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe detectar cualquier cantidad > 10 (default)."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="Otros")
        ws.cell(row=2, column=3, value=11)
        ws.cell(row=2, column=4, value=CONVENIO_ASISTENCIAL)
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value="Otros")
        ws.cell(row=3, column=3, value=10)
        ws.cell(row=3, column=4, value=CONVENIO_ASISTENCIAL)

        indices = {
            "numero_factura": 0,
            "tipo_procedimiento": 1,
            "cantidad": 2,
            "convenio_facturado": 3,
        }
        result = detect_cantidades_anomalas(ws, indices)

        facturas = [r["factura"] for r in result]
        assert "FAC-001" in facturas
        assert "FAC-002" not in facturas

    def test_detecta_pyp_con_cantidad_mayor_igual_3(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe detectar PyP con cantidad >= 3 (default)."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="Procedimientos")
        ws.cell(row=2, column=3, value=3)
        ws.cell(row=2, column=4, value=CONVENIO_PYP)
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value="Procedimientos")
        ws.cell(row=3, column=3, value=2)
        ws.cell(row=3, column=4, value=CONVENIO_PYP)

        indices = {
            "numero_factura": 0,
            "tipo_procedimiento": 1,
            "cantidad": 2,
            "convenio_facturado": 3,
        }
        result = detect_cantidades_anomalas(ws, indices)

        facturas = [r["factura"] for r in result]
        assert "FAC-001" in facturas
        assert "FAC-002" not in facturas

    def test_parametros_personalizados(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe usar parámetros personalizados correctamente."""
        ws = workbook_with_headers.active
        # Consultas con cantidad 1 -> no anómalo con threshold=2 default
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="Consultas")
        ws.cell(row=2, column=3, value=1)
        ws.cell(row=2, column=4, value=CONVENIO_ASISTENCIAL)
        # Cantidad 6 -> anómalo con cantidad_max_general=5
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value="Otros")
        ws.cell(row=3, column=3, value=6)
        ws.cell(row=3, column=4, value=CONVENIO_ASISTENCIAL)
        # PyP con cantidad 2 -> anómalo con cantidad_pyp_min=2
        ws.cell(row=4, column=1, value="FAC-003")
        ws.cell(row=4, column=2, value="Procedimientos")
        ws.cell(row=4, column=3, value=2)
        ws.cell(row=4, column=4, value=CONVENIO_PYP)

        indices = {
            "numero_factura": 0,
            "tipo_procedimiento": 1,
            "cantidad": 2,
            "convenio_facturado": 3,
        }
        result = detect_cantidades_anomalas(
            ws,
            indices,
            cantidad_consultas_min=5,
            cantidad_max_general=5,
            cantidad_pyp_min=2,
        )

        facturas = [r["factura"] for r in result]
        assert "FAC-001" not in facturas  # Consultas con 1 < 5
        assert "FAC-002" in facturas       # 6 > 5
        assert "FAC-003" in facturas       # PyP con 2 >= 2

    def test_sin_indices_retorna_vacio(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Sin indices necesarios debe retornar lista vacía."""
        indices = {
            "numero_factura": None,
            "tipo_procedimiento": None,
            "cantidad": None,
            "convenio_facturado": None,
        }
        result = detect_cantidades_anomalas(
            workbook_with_headers.active, indices
        )

        assert result == []

    def test_no_duplica_facturas(
        self, workbook_with_headers: Workbook
    ) -> None:
        """No debe duplicar facturas con múltiples filas anómalas."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="Consultas")
        ws.cell(row=2, column=3, value=5)
        ws.cell(row=2, column=4, value=CONVENIO_ASISTENCIAL)
        ws.cell(row=3, column=1, value="FAC-001")
        ws.cell(row=3, column=2, value="Consultas")
        ws.cell(row=3, column=3, value=8)
        ws.cell(row=3, column=4, value=CONVENIO_ASISTENCIAL)

        indices = {
            "numero_factura": 0,
            "tipo_procedimiento": 1,
            "cantidad": 2,
            "convenio_facturado": 3,
        }
        result = detect_cantidades_anomalas(ws, indices)

        assert len(result) == 1

    def test_valores_no_numericos_ignorados(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Cantidades no numéricas deben ser ignoradas."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="Consultas")
        ws.cell(row=2, column=3, value="No numerico")
        ws.cell(row=2, column=4, value=CONVENIO_ASISTENCIAL)

        indices = {
            "numero_factura": 0,
            "tipo_procedimiento": 1,
            "cantidad": 2,
            "convenio_facturado": 3,
        }
        result = detect_cantidades_anomalas(ws, indices)

        assert len(result) == 0
