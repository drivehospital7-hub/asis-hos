"""Tests for tipo_factura filter on app/services/urgencias/revision_entidad_86.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.urgencias.revision_entidad_86 import (
    detect_revision_entidad_86_urgencias,
)


@pytest.fixture
def wb_with_tipo_factura() -> Workbook:
    """Workbook with headers including tipo_factura_descripcion."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Cód Entidad Cobrar")
    ws.cell(row=1, column=3, value="Código")
    ws.cell(row=1, column=4, value="Procedimiento")
    ws.cell(row=1, column=5, value="IDE Contrato")
    ws.cell(row=1, column=6, value="Tipo Factura Descripción")
    return wb


_INDICES_FULL = {
    "numero_factura": 0,
    "codigo_entidad_cobrar": 1,
    "codigo": 2,
    "procedimiento": 3,
    "ide_contrato": 4,
    "tipo_factura_descripcion": 5,
}


class TestRevisionEntidad86TipoFacturaFilter:
    """Verify detect_revision_entidad_86_urgencias filters by tipo_factura_descripcion."""

    def test_filtra_solo_urgencias_entidad_86(
        self, wb_with_tipo_factura: Workbook
    ) -> None:
        """Urgencias rows with entidad=86 → detected; Hospitalización → skipped."""
        ws = wb_with_tipo_factura.active

        # Row 2: Urgencias, entidad=86 → should be detected
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="86")
        ws.cell(row=2, column=3, value="890101")
        ws.cell(row=2, column=4, value="Procedimiento A")
        ws.cell(row=2, column=5, value="IDE001")
        ws.cell(row=2, column=6, value="Urgencias")

        # Row 3: Hospitalización, entidad=86 → should be SKIPPED
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value="86")
        ws.cell(row=3, column=3, value="890102")
        ws.cell(row=3, column=4, value="Procedimiento B")
        ws.cell(row=3, column=5, value="IDE002")
        ws.cell(row=3, column=6, value="Hospitalización")

        result = detect_revision_entidad_86_urgencias(ws, _INDICES_FULL)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"
        assert result[0]["entidad"] == "86"

    def test_sin_urgencias_entidad_86_retorna_vacio(
        self, wb_with_tipo_factura: Workbook
    ) -> None:
        """No Urgencias rows → empty even if entidad=86 exists."""
        ws = wb_with_tipo_factura.active

        ws.cell(row=2, column=1, value="FAC-003")
        ws.cell(row=2, column=2, value="86")
        ws.cell(row=2, column=3, value="890103")
        ws.cell(row=2, column=4, value="Proc C")
        ws.cell(row=2, column=5, value="IDE003")
        ws.cell(row=2, column=6, value="Intramural")

        result = detect_revision_entidad_86_urgencias(ws, _INDICES_FULL)
        assert result == []

    def test_urgencias_sin_entidad_86_no_flag(
        self, wb_with_tipo_factura: Workbook
    ) -> None:
        """Urgencias rows with entidad != 86 → not flagged."""
        ws = wb_with_tipo_factura.active

        ws.cell(row=2, column=1, value="FAC-004")
        ws.cell(row=2, column=2, value="99")
        ws.cell(row=2, column=3, value="890104")
        ws.cell(row=2, column=4, value="Proc D")
        ws.cell(row=2, column=5, value="IDE004")
        ws.cell(row=2, column=6, value="Urgencias")

        result = detect_revision_entidad_86_urgencias(ws, _INDICES_FULL)
        assert result == []

    def test_missing_tipo_factura_col_entidad_86(
        self, wb_with_tipo_factura: Workbook
    ) -> None:
        """Missing tipo_factura_descripcion column → return []."""
        ws = wb_with_tipo_factura.active

        ws.cell(row=2, column=1, value="FAC-005")
        ws.cell(row=2, column=2, value="86")

        indices_no_tipo = {
            "numero_factura": 0,
            "codigo_entidad_cobrar": 1,
        }
        result = detect_revision_entidad_86_urgencias(ws, indices_no_tipo)
        assert result == []
