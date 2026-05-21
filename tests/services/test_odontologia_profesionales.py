"""Tests para app/services/odontologia/profesionales.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.constants import PROFESIONALES_ODONTOLOGIA_VALIDACION
from app.services.odontologia.profesionales import detect_profesionales_odontologia


@pytest.fixture
def workbook_with_odontologia_headers() -> Workbook:
    """Crea un workbook con headers para odontología."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código Profesional")
    ws.cell(row=1, column=3, value="Cód. Equivalente CUPS")
    return wb


class TestDetectProfesionalesOdontologia:
    """Tests para detect_profesionales_odontologia."""

    def test_profesional_valido_no_genera_error(
        self, workbook_with_odontologia_headers: Workbook
    ) -> None:
        """Un profesional válido en el listado no debe generar error."""
        ws = workbook_with_odontologia_headers.active
        # "03424" = ARIAS MOREANO LAURA MELISSA (ODONTOLOGO)
        # 890101 no está en PYP_CODES_HIGIENISTA, ODONTOLOGO puede usarlo
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="03424")
        ws.cell(row=2, column=3, value="890101")

        indices = {"numero_factura": 0, "codigo_profesional": 1, "codigo": 2}
        result = detect_profesionales_odontologia(ws, indices)

        assert len(result) == 0

    def test_profesional_no_en_listado_genera_error(
        self, workbook_with_odontologia_headers: Workbook
    ) -> None:
        """Un profesional NO en el listado debe generar error."""
        ws = workbook_with_odontologia_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="99999")  # No existe
        ws.cell(row=2, column=3, value="890101")

        indices = {"numero_factura": 0, "codigo_profesional": 1, "codigo": 2}
        result = detect_profesionales_odontologia(ws, indices)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"
        assert result[0]["regla"] == "Profesional debe estar en listado"

    def test_higienista_con_codigo_no_pyp_genera_error(
        self, workbook_with_odontologia_headers: Workbook
    ) -> None:
        """HIGIENISTA solo puede usar códigos PYP."""
        ws = workbook_with_odontologia_headers.active
        # "01329" = CASTILLO DUQUE NOHORA ELENA (HIGIENISTA)
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="01329")
        ws.cell(row=2, column=3, value="890101")  # No PyP

        indices = {"numero_factura": 0, "codigo_profesional": 1, "codigo": 2}
        result = detect_profesionales_odontologia(ws, indices)

        assert len(result) == 1
        assert result[0]["tipo"] == "HIGIENISTA"
        assert "Solo códigos PYP" in result[0]["regla"]

    def test_odontologo_con_codigo_higienista_genera_error(
        self, workbook_with_odontologia_headers: Workbook
    ) -> None:
        """ODONTOLOGO no puede usar códigos del set PYP_CODES_HIGIENISTA."""
        ws = workbook_with_odontologia_headers.active
        # "03424" = ARIAS MOREANO LAURA MELISSA (ODONTOLOGO)
        # 997002 está en PYP_CODES_HIGIENISTA -> ODONTOLOGO no puede usarlo
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="03424")
        ws.cell(row=2, column=3, value="997002")

        indices = {"numero_factura": 0, "codigo_profesional": 1, "codigo": 2}
        result = detect_profesionales_odontologia(ws, indices)

        assert len(result) == 1
        assert result[0]["tipo"] == "ODONTOLOGO"

    def test_sin_indices_retorna_vacio(
        self, workbook_with_odontologia_headers: Workbook
    ) -> None:
        """Si faltan índices necesarios, retorna lista vacía."""
        ws = workbook_with_odontologia_headers.active
        result = detect_profesionales_odontologia(ws, {})
        assert result == []

    def test_no_duplica_facturas(
        self, workbook_with_odontologia_headers: Workbook
    ) -> None:
        """Misma factura con múltiples filas se reporta una sola vez."""
        ws = workbook_with_odontologia_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="99999")
        ws.cell(row=2, column=3, value="890101")
        ws.cell(row=3, column=1, value="FAC-001")
        ws.cell(row=3, column=2, value="99999")
        ws.cell(row=3, column=3, value="890102")

        indices = {"numero_factura": 0, "codigo_profesional": 1, "codigo": 2}
        result = detect_profesionales_odontologia(ws, indices)

        assert len(result) == 1
