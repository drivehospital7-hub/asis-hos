"""Tests para app/services/urgencias/hospitalizacion.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.urgencias.hospitalizacion import detect_cantidades_hospitalizacion


@pytest.fixture
def wb_hosp_headers() -> Workbook:
    """Crea un workbook con headers para validación cantidades hospitalización."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código")
    ws.cell(row=1, column=3, value="Procedimiento")
    ws.cell(row=1, column=4, value="Cantidad")
    ws.cell(row=1, column=5, value="Tipo Factura Descripción")
    ws.cell(row=1, column=6, value="Tarifario")
    ws.cell(row=1, column=7, value="Fec. Factura")
    ws.cell(row=1, column=8, value="Fecha Cierre")
    return wb


class TestDetectCantidadesHospitalizacion:
    """Tests para detect_cantidades_hospitalizacion."""

    def test_codigo_129b02_cantidad_incorrecta_genera_error(
        self, wb_hosp_headers: Workbook
    ) -> None:
        """Código 129B02 (Estancia) con cantidad incorrecta en Hospitalización genera error.
        
        24h (1 día completo) → cantidad esperada = 1 + 1 = 2
        """
        ws = wb_hosp_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="129B02")
        ws.cell(row=2, column=3, value="ESTANCIA")
        ws.cell(row=2, column=4, value=1)  # incorrecto, debe ser 2
        ws.cell(row=2, column=5, value="Hospitalización")
        ws.cell(row=2, column=6, value="ISS")
        ws.cell(row=2, column=7, value="2024-01-01 08:00:00")
        ws.cell(row=2, column=8, value="2024-01-02 08:00:00")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
            "tarifario": 5,
            "fec_factura": 6,
            "fecha_cierre": 7,
        }
        result = detect_cantidades_hospitalizacion(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "129B02"
        assert result[0]["cantidad_esperada"] == 2

    def test_codigo_129b02_cantidad_correcta_no_genera_error(
        self, wb_hosp_headers: Workbook
    ) -> None:
        """Código 129B02 con cantidad correcta no genera error."""
        ws = wb_hosp_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="129B02")
        ws.cell(row=2, column=3, value="ESTANCIA")
        ws.cell(row=2, column=4, value=2)  # correcto para 24h
        ws.cell(row=2, column=5, value="Hospitalización")
        ws.cell(row=2, column=6, value="ISS")
        ws.cell(row=2, column=7, value="2024-01-01 08:00:00")
        ws.cell(row=2, column=8, value="2024-01-02 08:00:00")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
            "tarifario": 5,
            "fec_factura": 6,
            "fecha_cierre": 7,
        }
        result = detect_cantidades_hospitalizacion(ws, indices)

        assert len(result) == 0

    def test_codigo_890601_cantidad_incorrecta_genera_error(
        self, wb_hosp_headers: Workbook
    ) -> None:
        """Código 890601 (Camas) con cantidad incorrecta genera error.
        
        48h (2 días completos) → cantidad esperada = 2
        """
        ws = wb_hosp_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="890601")
        ws.cell(row=2, column=3, value="CAMA")
        ws.cell(row=2, column=4, value=1)  # incorrecto, debe ser 2
        ws.cell(row=2, column=5, value="Hospitalización")
        ws.cell(row=2, column=6, value="ISS")
        ws.cell(row=2, column=7, value="2024-01-01 08:00:00")
        ws.cell(row=2, column=8, value="2024-01-03 08:00:00")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
            "tarifario": 5,
            "fec_factura": 6,
            "fecha_cierre": 7,
        }
        result = detect_cantidades_hospitalizacion(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "890601"
        assert result[0]["cantidad_esperada"] == 2

    def test_codigo_890601h_cantidad_mayor_1_no_soat_genera_error(
        self, wb_hosp_headers: Workbook
    ) -> None:
        """Código 890601H con cantidad > 1 y NO SOAT genera error."""
        ws = wb_hosp_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="890601H")
        ws.cell(row=2, column=3, value="PROC")
        ws.cell(row=2, column=4, value=2)
        ws.cell(row=2, column=5, value="Hospitalización")
        ws.cell(row=2, column=6, value="ISS")
        ws.cell(row=2, column=7, value="2024-01-01 08:00:00")
        ws.cell(row=2, column=8, value="2024-01-02 08:00:00")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
            "tarifario": 5,
            "fec_factura": 6,
            "fecha_cierre": 7,
        }
        result = detect_cantidades_hospitalizacion(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "890601H"
        assert result[0]["cantidad_esperada"] == 1

    def test_codigo_890601_menos_24h_genera_error(
        self, wb_hosp_headers: Workbook
    ) -> None:
        """Código 890601 con estancia < 24h genera error (no debe existir)."""
        ws = wb_hosp_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="890601")
        ws.cell(row=2, column=3, value="CAMA")
        ws.cell(row=2, column=4, value=0)
        ws.cell(row=2, column=5, value="Hospitalización")
        ws.cell(row=2, column=6, value="ISS")
        ws.cell(row=2, column=7, value="2024-01-01 08:00:00")
        ws.cell(row=2, column=8, value="2024-01-01 12:00:00")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
            "tarifario": 5,
            "fec_factura": 6,
            "fecha_cierre": 7,
        }
        result = detect_cantidades_hospitalizacion(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "890601"
        assert result[0]["cantidad_esperada"] == 0

    def test_tipo_factura_no_hospitalizacion_no_genera_error(
        self, wb_hosp_headers: Workbook
    ) -> None:
        """Si Tipo Factura no es Hospitalización, no se valida."""
        ws = wb_hosp_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="129B02")
        ws.cell(row=2, column=3, value="ESTANCIA")
        ws.cell(row=2, column=4, value=1)
        ws.cell(row=2, column=5, value="Urgencias")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
        }
        result = detect_cantidades_hospitalizacion(ws, indices)

        assert len(result) == 0

    def test_sin_indices_retorna_vacio(
        self, wb_hosp_headers: Workbook
    ) -> None:
        """Si faltan índices necesarios, retorna lista vacía."""
        ws = wb_hosp_headers.active
        result = detect_cantidades_hospitalizacion(ws, {})
        assert result == []
