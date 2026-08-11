"""Tests para app/services/odontologia/detect_all.py."""

from __future__ import annotations

from unittest.mock import patch

import pytest
from openpyxl import Workbook

from app.constants import CONVENIO_PYP
from app.services.odontologia.detect_all import detect_all_problems_odontologia


@pytest.fixture
def workbook_minimal() -> Workbook:
    """Crea un workbook con headers mínimos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    return wb


class TestDetectAllProblemsOdontologia:
    """Tests para detect_all_problems_odontologia."""

    def _run(self, ws, indices):
        """Helper que corre el detector y retorna solo el dict resultado."""
        with patch("app.services.odontologia.detect_all.is_rule_engine_enabled", return_value=False):
            result, _ = detect_all_problems_odontologia(ws, indices)
        return result

    def test_retorna_dict_con_key_problemas(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe contener key 'problemas'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = {"numero_factura": 0}
        result = self._run(ws, indices)

        assert "problemas" in result
        assert isinstance(result["problemas"], dict)

    def test_retorna_dict_con_key_totales(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe contener key 'totales'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = {"numero_factura": 0}
        result = self._run(ws, indices)

        assert "totales" in result
        assert isinstance(result["totales"], dict)

    def test_retorna_dict_con_key_area(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe contener 'area' = 'odontologia'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = {"numero_factura": 0}
        result = self._run(ws, indices)

        assert result.get("area") == "odontologia"

    def test_resultado_incluye_normalizados(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe incluir 'normalizados' en problemas."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = {"numero_factura": 0}
        result = self._run(ws, indices)

        assert "normalizados" in result["problemas"]
        assert isinstance(result["problemas"]["normalizados"], list)

    def test_resultado_incluye_missing_columns(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe contener 'missing_columns'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = {"numero_factura": 0}
        result = self._run(ws, indices)

        assert "missing_columns" in result
        assert isinstance(result["missing_columns"], list)

    def test_normalizados_incluyen_fec_factura(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado normalizados MUST include 'fec_factura' in every row."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="2024-01-15")
        ws.cell(row=1, column=2, value="Fec. Factura")

        indices = {"numero_factura": 0, "fec_factura": 1}
        result = self._run(ws, indices)
        norm = result["problemas"]["normalizados"]
        for row in norm:
            assert "fec_factura" in row

    def _run_legacy(self, ws, indices):
        """Helper que corre legacy path sin mockear engine (usa else branch con [])."""
        with patch("app.services.odontologia.detect_all.is_rule_engine_enabled", return_value=False):
            result, _ = detect_all_problems_odontologia(ws, indices)
        return result

    def test_ruta_duplicada_excluye_3_facturas_con_codigo_exento(
        self, workbook_minimal: Workbook
    ) -> None:
        """3 facturas PyP con código 990203, P0000011 o 990212 NO se reportan."""
        ws = workbook_minimal.active
        ws.cell(row=1, column=1, value="Número Factura")
        ws.cell(row=1, column=2, value="Nº Identificación")
        ws.cell(row=1, column=3, value="Convenio Facturado")
        ws.cell(row=1, column=4, value="Código")

        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="PAC-001")
        ws.cell(row=2, column=3, value=CONVENIO_PYP)
        ws.cell(row=2, column=4, value="990203")
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value="PAC-001")
        ws.cell(row=3, column=3, value=CONVENIO_PYP)
        ws.cell(row=3, column=4, value="997002")
        ws.cell(row=4, column=1, value="FAC-003")
        ws.cell(row=4, column=2, value="PAC-001")
        ws.cell(row=4, column=3, value=CONVENIO_PYP)
        ws.cell(row=4, column=4, value="997106")
        ws.cell(row=5, column=1, value="FAC-004")
        ws.cell(row=5, column=2, value="PAC-002")
        ws.cell(row=5, column=3, value=CONVENIO_PYP)
        ws.cell(row=5, column=4, value="997002")
        ws.cell(row=6, column=1, value="FAC-005")
        ws.cell(row=6, column=2, value="PAC-002")
        ws.cell(row=6, column=3, value=CONVENIO_PYP)
        ws.cell(row=6, column=4, value="997106")
        ws.cell(row=7, column=1, value="FAC-006")
        ws.cell(row=7, column=2, value="PAC-002")
        ws.cell(row=7, column=3, value=CONVENIO_PYP)
        ws.cell(row=7, column=4, value="997301")

        indices = {
            "numero_factura": 0,
            "identificacion": 1,
            "convenio_facturado": 2,
            "codigo": 3,
            "codigo_equiv": None,
            "tipo_procedimiento": None,
            "codigo_tipo_procedimiento": None,
            "procedimiento": None,
            "centro_costo": None,
            "codigo_entidad_cobrar": None,
            "entidad_cobrar": None,
            "entidad_afiliacion": None,
            "fec_nacimiento": None,
            "fec_factura": None,
            "fecha_cierre": None,
            "profesional_identificacion": None,
            "profesional_atiende": None,
            "codigo_profesional": None,
            "responsable_cierra": None,
            "vlr_subsidiado": None,
            "vlr_procedimiento": None,
            "laboratorio": None,
            "tarifario": None,
            "tipo_factura_descripcion": None,
            "ide_contrato": None,
            "tipo_identificacion": None,
            "tipo_usuario": None,
            "vlr_copago": None,
            "numero_reingreso": None,
            "codigo_dx_principal": None,
            "cantidad": None,
        }
        # Mock RuleBasedDetector to let engine return ruta_dup data
        from unittest.mock import MagicMock

        def _mock_detector(name, session):
            d = MagicMock()
            if name == "ruta_duplicada":
                # Return data matching the PAC-001/PAC-002 setup
                d.detect.return_value = [
                    {"identificacion": "PAC-001", "factura": "FAC-001",
                     "cantidad": 3, "codigo": "990203"},
                    {"identificacion": "PAC-001", "factura": "FAC-002",
                     "cantidad": 3, "codigo": "997002"},
                    {"identificacion": "PAC-001", "factura": "FAC-003",
                     "cantidad": 3, "codigo": "997106"},
                    {"identificacion": "PAC-002", "factura": "FAC-004",
                     "cantidad": 3, "codigo": "997002"},
                    {"identificacion": "PAC-002", "factura": "FAC-005",
                     "cantidad": 3, "codigo": "997106"},
                    {"identificacion": "PAC-002", "factura": "FAC-006",
                     "cantidad": 3, "codigo": "997301"},
                ]
            else:
                d.detect.return_value = []
            return d

        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = MagicMock()
                m_dc.side_effect = _mock_detector
                result, _ = detect_all_problems_odontologia(ws, indices)

        ruta_dup = result["problemas"]["ruta_duplicada"]
        identificaciones = [r["identificacion"] for r in ruta_dup]

        assert "PAC-001" not in identificaciones, (
            "PAC-001 tiene 3 facturas con código 990203 => debe excluirse"
        )
        assert "PAC-002" in identificaciones, (
            "PAC-002 tiene 3 facturas sin código exento => debe reportarse"
        )

    def test_ruta_duplicada_no_excluye_4_facturas_con_codigo_exento(
        self, workbook_minimal: Workbook
    ) -> None:
        """4+ facturas PyP se reportan aunque tengan códigos exentos."""
        ws = workbook_minimal.active
        ws.cell(row=1, column=1, value="Número Factura")
        ws.cell(row=1, column=2, value="Nº Identificación")
        ws.cell(row=1, column=3, value="Convenio Facturado")
        ws.cell(row=1, column=4, value="Código")

        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="PAC-001")
        ws.cell(row=2, column=3, value=CONVENIO_PYP)
        ws.cell(row=2, column=4, value="990203")
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value="PAC-001")
        ws.cell(row=3, column=3, value=CONVENIO_PYP)
        ws.cell(row=3, column=4, value="997002")
        ws.cell(row=4, column=1, value="FAC-003")
        ws.cell(row=4, column=2, value="PAC-001")
        ws.cell(row=4, column=3, value=CONVENIO_PYP)
        ws.cell(row=4, column=4, value="997106")
        ws.cell(row=5, column=1, value="FAC-004")
        ws.cell(row=5, column=2, value="PAC-001")
        ws.cell(row=5, column=3, value=CONVENIO_PYP)
        ws.cell(row=5, column=4, value="997301")

        indices = {
            "numero_factura": 0,
            "identificacion": 1,
            "convenio_facturado": 2,
            "codigo": 3,
            "codigo_equiv": None,
            "tipo_procedimiento": None,
            "codigo_tipo_procedimiento": None,
            "procedimiento": None,
            "centro_costo": None,
            "codigo_entidad_cobrar": None,
            "entidad_cobrar": None,
            "entidad_afiliacion": None,
            "fec_nacimiento": None,
            "fec_factura": None,
            "fecha_cierre": None,
            "profesional_identificacion": None,
            "profesional_atiende": None,
            "codigo_profesional": None,
            "responsable_cierra": None,
            "vlr_subsidiado": None,
            "vlr_procedimiento": None,
            "laboratorio": None,
            "tarifario": None,
            "tipo_factura_descripcion": None,
            "ide_contrato": None,
            "tipo_identificacion": None,
            "tipo_usuario": None,
            "vlr_copago": None,
            "numero_reingreso": None,
            "codigo_dx_principal": None,
            "cantidad": None,
        }
        from unittest.mock import MagicMock

        def _mock_detector(name, session):
            d = MagicMock()
            if name == "ruta_duplicada":
                d.detect.return_value = [
                    {"identificacion": "PAC-001", "factura": "FAC-001",
                     "cantidad": 4, "codigo": "990203"},
                    {"identificacion": "PAC-001", "factura": "FAC-002",
                     "cantidad": 4, "codigo": "997002"},
                    {"identificacion": "PAC-001", "factura": "FAC-003",
                     "cantidad": 4, "codigo": "997106"},
                    {"identificacion": "PAC-001", "factura": "FAC-004",
                     "cantidad": 4, "codigo": "997301"},
                ]
            else:
                d.detect.return_value = []
            return d

        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = MagicMock()
                m_dc.side_effect = _mock_detector
                result, _ = detect_all_problems_odontologia(ws, indices)

        ruta_dup = result["problemas"]["ruta_duplicada"]
        identificaciones = [r["identificacion"] for r in ruta_dup]

        assert "PAC-001" in identificaciones, (
            "PAC-001 tiene 4 facturas aunque con código exento => debe reportarse"
        )
