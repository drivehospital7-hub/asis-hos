"""Tests para app/services/equipos_basicos/detect_all.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.constants import AREA_EQUIPOS_BASICOS


@pytest.fixture
def workbook_minimal() -> Workbook:
    """Crea un workbook con headers mínimos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    return wb


class TestDetectAllProblemsEquiposBasicos:
    """Tests para detect_all_problems_equipos_basicos."""

    @staticmethod
    def _make_indices(extra: dict[str, int | None] | None = None) -> dict[str, int | None]:
        """Construye un dict de índices con todas las claves esperadas."""
        base: dict[str, int | None] = {
            "numero_factura": None,
            "codigo_profesional": None,
            "codigo": None,
            "codigo_entidad_cobrar": None,
            "entidad_afiliacion": None,
            "tipo_usuario": None,
            "procedimiento": None,
            "ide_contrato": None,
            "responsable_cierra": None,
            "tipo_identificacion": None,
            "fec_nacimiento": None,
            "fec_factura": None,
            "cantidad": None,
            "codigo_tipo_procedimiento": None,
            "laboratorio": None,
            "tipo_procedimiento": None,
            "convenio_facturado": None,
            "entidad_cobrar": None,
            "vlr_subsidiado": None,
            "vlr_procedimiento": None,
            "centro_costo": None,
        }
        if extra:
            base.update(extra)
        return base

    def _run(self, ws, indices):
        """Helper que corre el detector y retorna solo el dict resultado."""
        from app.services.equipos_basicos.detect_all import (
            detect_all_problems_equipos_basicos,
        )

        result, _ = detect_all_problems_equipos_basicos(ws, indices)
        return result

    def test_retorna_dict_con_key_problemas(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe contener key 'problemas'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = self._make_indices({"numero_factura": 0})
        result = self._run(ws, indices)

        assert "problemas" in result
        assert isinstance(result["problemas"], dict)

    def test_retorna_dict_con_key_totales(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe contener key 'totales'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = self._make_indices({"numero_factura": 0})
        result = self._run(ws, indices)

        assert "totales" in result
        assert isinstance(result["totales"], dict)

    def test_retorna_dict_con_key_area(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe contener 'area' = 'equipos_basicos'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = self._make_indices({"numero_factura": 0})
        result = self._run(ws, indices)

        assert result.get("area") == AREA_EQUIPOS_BASICOS

    def test_resultado_incluye_normalizados(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe incluir 'normalizados' en problemas."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = self._make_indices({"numero_factura": 0})
        result = self._run(ws, indices)

        assert "normalizados" in result["problemas"]
        assert isinstance(result["problemas"]["normalizados"], list)

    def test_resultado_incluye_profesionales(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe incluir 'profesionales' en problemas."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = self._make_indices({"numero_factura": 0})
        result = self._run(ws, indices)

        assert "profesionales" in result["problemas"]

    def test_missing_columns_incluido(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe contener 'missing_columns'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = self._make_indices({"numero_factura": 0})
        result = self._run(ws, indices)

        assert "missing_columns" in result
        assert isinstance(result["missing_columns"], list)

    def test_es_equipos_basicos_flag(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe tener flag 'es_equipos_basicos' = True."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = self._make_indices({"numero_factura": 0})
        result = self._run(ws, indices)

        assert result.get("es_equipos_basicos") is True
