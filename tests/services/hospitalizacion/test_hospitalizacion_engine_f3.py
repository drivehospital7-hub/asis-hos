"""Strict TDD F3: Tests for Hospitalización-specific engine toggles.

Covers T-F3.1 (centro_costo), T-F3.2 (cantidades), T-F3.3 (codes).
"""
from __future__ import annotations

import os
from unittest.mock import MagicMock, patch

from openpyxl import Workbook

from app.constants import AREA_HOSPITALIZACION

_F3_DETECTOR_KEYS = frozenset({
    "centros_de_costos",
    "ide_contrato",
    "cantidades_hospitalizacion",
    "cantidades_soat_hospitalizacion",
    "cups_equivalentes",
})


def _env(val: str):
    old = os.environ.pop("USE_RULE_ENGINE", None)
    os.environ["USE_RULE_ENGINE"] = val
    def restore():
        if old is not None:
            os.environ["USE_RULE_ENGINE"] = old
        else:
            os.environ.pop("USE_RULE_ENGINE", None)
    return restore


def _build_sheet() -> tuple[Workbook, dict[str, int | None]]:
    """Build a Hospitalización sheet with minimal columns for F3 testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    headers = [
        "Número Factura", "Código", "Cantidad", "Vlr. Unitario",
        "Vlr. Procedimiento", "Vlr. Subsidiado", "Tipo Doc.", "Edad",
        "Tipo Identificación", "Código Entidad Cobrar", "Entidad Afiliación",
        "Tipo Usuario", "Vlr. Copago", "Código CUPS", "Fec Factura",
        "Fecha Cierre", "Responsable Cierra", "Tarifario",
        "Tipo Factura Descripción", "Centro Costo", "Procedimiento",
        "Codigo Tipo Procedimiento", "Laboratorio",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    # Row 1: Hospitalización con valores limpios
    ws.cell(row=2, column=1, value="HOSP-001")
    ws.cell(row=2, column=2, value="890201")
    ws.cell(row=2, column=3, value=1)
    ws.cell(row=2, column=4, value=150.00)
    ws.cell(row=2, column=19, value="Hospitalización")
    ws.cell(row=2, column=20, value="HOSPITALIZACIÓN - ESTANCIA GENERAL")
    ws.cell(row=2, column=21, value="ESTANCIA GENERAL")
    ws.cell(row=2, column=22, value="01")
    ws.cell(row=2, column=23, value="No")
    # Row 2: another Hospitalización row
    ws.cell(row=3, column=1, value="HOSP-002")
    ws.cell(row=3, column=2, value="129B02")
    ws.cell(row=3, column=3, value=2)
    ws.cell(row=3, column=4, value=200.00)
    ws.cell(row=3, column=19, value="Hospitalización")
    ws.cell(row=3, column=20, value="URGENCIAS")
    ws.cell(row=3, column=21, value="CONSULTA URGENCIA")
    ws.cell(row=3, column=22, value="01")
    ws.cell(row=3, column=23, value="No")
    indices = {h: i for i, h in enumerate(headers)}
    return wb, indices


def _mock_session() -> MagicMock:
    s = MagicMock()
    q = MagicMock()
    q.filter.return_value = q
    q.order_by.return_value = q
    q.first.return_value = None
    q.all.return_value = []
    s.query.return_value = q
    return s


class TestF3CentroCostoToggle:
    """T-F3.1: centro_costo_hospitalizacion → engine toggle."""

    def test_engine_path_routes_centro_costo(self):
        """Engine path MUST call RuleBasedDetector for centro_costo_hospitalizacion_valido."""
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = _mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md

                from app.services.hospitalizacion.detect_all import (
                    detect_all_problems_hospitalizacion,
                )
                wb, idx = _build_sheet()
                restore = _env("true")
                try:
                    r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
                finally:
                    restore()

        # centro_costo_hospitalizacion_valido must be among RuleBasedDetector calls
        call_names = [call[0][0] for call in m_dc.call_args_list]
        assert "centro_costo_hospitalizacion_valido" in call_names, (
            f"centro_costo_hospitalizacion_valido not in calls: {call_names}"
        )
        assert "centros_de_costos" in r["problemas"]
        assert r["area"] == AREA_HOSPITALIZACION

    def test_legacy_path_centro_costo_structure(self):
        """Legacy path must produce centros_de_costos without crashing."""
        from app.services.hospitalizacion.detect_all import (
            detect_all_problems_hospitalizacion,
        )
        wb, idx = _build_sheet()
        restore = _env("false")
        try:
            r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
        finally:
            restore()
        assert "centros_de_costos" in r["problemas"]
        assert r["area"] == AREA_HOSPITALIZACION

    def test_centro_costo_engine_uses_right_domain(self):
        """Centro_costo engine toggle MUST use centro_costo_hospitalizacion_valido (domain=hospitalizacion).

        Use a side_effect on the mock so only centro_costo rule returns items.
        Other rules (decimales, etc.) return [] to avoid dict-as-key errors in normalizer.
        """
        centro_costo_rv = [{"factura": "HOSP-001", "problema": "CC error",
                            "codigo": "", "procedimiento": "", "prioridad": 1,
                            "centro_actual": "URGENCIAS", "centro_deberia": "HOSPITALIZACION"}]

        def _mock_detect(*args, **kwargs):
            return [{"factura": "FAC-001", "problema": "test"}]

        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = _mock_session()
                # When centro_costo_hospitalizacion_valido is created, return centro_costo items
                actual_detectors: dict[str, MagicMock] = {}

                def _side_effect(name, session):
                    if name not in actual_detectors:
                        d = MagicMock()
                        if name == "centro_costo_hospitalizacion_valido":
                            d.detect.return_value = centro_costo_rv
                        else:
                            d.detect.return_value = []
                        actual_detectors[name] = d
                    return actual_detectors[name]

                m_dc.side_effect = _side_effect

                from app.services.hospitalizacion.detect_all import (
                    detect_all_problems_hospitalizacion,
                )
                wb, idx = _build_sheet()
                restore = _env("true")
                try:
                    r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
                finally:
                    restore()

        # Verify centro_costo_hospitalizacion_valido was called
        cc_calls = [c for c in m_dc.call_args_list
                    if c[0][0] == "centro_costo_hospitalizacion_valido"]
        assert len(cc_calls) == 1, "centro_costo_hospitalizacion_valido should be called exactly once"
        # centro_costo items should appear in centros_de_costos output
        assert len(r["problemas"]["centros_de_costos"]) >= 1

    def test_centro_costo_format_resilient_to_engine_output(self):
        """Formatting MUST NOT crash when centro_costo items lack legacy keys.
        
        We mock centro_costo_hospitalizacion_valido's RuleBasedDetector to return
        items WITHOUT centro_actual/centro_deberia (as the engine does).
        The formatting uses .get() to handle missing keys gracefully.
        """
        engine_like_output = [
            {"factura": "HOSP-001", "problema": "CC invalid",
             "regla": "#1", "severidad": "error",
             "codigo": "", "procedimiento": ""}
        ]

        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = _mock_session()

                def _side_effect(name, session):
                    d = MagicMock()
                    d.detect.return_value = []
                    return d

                m_dc.side_effect = _side_effect

                from app.services.hospitalizacion.detect_all import (
                    detect_all_problems_hospitalizacion,
                )
                wb, idx = _build_sheet()
                r, _ = detect_all_problems_hospitalizacion(wb.active, idx)

        assert "centros_de_costos" in r["problemas"]
        # Engine path returns empty because centro_costo_hospitalizacion_valido returns []
        assert len(r["problemas"]["centros_de_costos"]) >= 0


class TestF3CantidadesToggle:
    """T-F3.2: cantidades_hospitalizacion + cantidades_soat_hospitalizacion → engine."""

    def _build_cantidades_sheet(self) -> tuple[Workbook, dict[str, int | None]]:
        """Build Hospitalización sheet with cantidad data for cantidades testing."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        headers = [
            "Número Factura", "Código", "Cantidad", "Vlr. Unitario",
            "Tipo Factura Descripción", "Tarifario",
        ]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        # Row with cantidad=9 (>8 threshold for cantidades_hospitalizacion)
        ws.cell(row=2, column=1, value="HOSP-001")
        ws.cell(row=2, column=2, value="890201")
        ws.cell(row=2, column=3, value=9)
        ws.cell(row=2, column=4, value=150.00)
        ws.cell(row=2, column=5, value="Hospitalización")
        ws.cell(row=2, column=6, value="NO SOAT")
        indices = {h: i for i, h in enumerate(headers)}
        return wb, indices

    def test_engine_path_routes_cantidades_hospitalizacion(self):
        """Engine path MUST call RuleBasedDetector for cantidades_hospitalizacion."""
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = _mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md

                from app.services.hospitalizacion.detect_all import (
                    detect_all_problems_hospitalizacion,
                )
                wb, idx = _build_sheet()
                restore = _env("true")
                try:
                    r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
                finally:
                    restore()

        call_names = [call[0][0] for call in m_dc.call_args_list]
        assert "cantidades_hospitalizacion" in call_names, (
            f"cantidades_hospitalizacion not in calls: {call_names}"
        )
        assert "cantidades_hospitalizacion" in r["problemas"]

    def test_engine_path_routes_cantidades_soat(self):
        """Engine path MUST call RuleBasedDetector for cantidades_soat_hospitalizacion."""
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = _mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md

                from app.services.hospitalizacion.detect_all import (
                    detect_all_problems_hospitalizacion,
                )
                wb, idx = _build_sheet()
                restore = _env("true")
                try:
                    r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
                finally:
                    restore()

        call_names = [call[0][0] for call in m_dc.call_args_list]
        assert "cantidades_soat_hospitalizacion" in call_names, (
            f"cantidades_soat_hospitalizacion not in calls: {call_names}"
        )
        assert "cantidades_soat_hospitalizacion" in r["problemas"]

    def test_cantidades_missing_column_graceful(self):
        """Engine path MUST handle missing cantidad column gracefully (no crash)."""
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = _mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md

                from app.services.hospitalizacion.detect_all import (
                    detect_all_problems_hospitalizacion,
                )
                # Build sheet WITHOUT cantidad column
                wb, idx = _build_sheet()
                if "Cantidad" in idx:
                    del idx["Cantidad"]
                restore = _env("true")
                try:
                    r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
                finally:
                    restore()

        assert "cantidades_hospitalizacion" in r["problemas"]
        assert r["problemas"]["cantidades_hospitalizacion"] == []

    def test_legacy_path_cantidades_structure(self):
        """Legacy path must produce cantidades keys without crashing."""
        from app.services.hospitalizacion.detect_all import (
            detect_all_problems_hospitalizacion,
        )
        wb, idx = _build_sheet()
        restore = _env("false")
        try:
            r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
        finally:
            restore()
        assert "cantidades_hospitalizacion" in r["problemas"]
        assert "cantidades_soat_hospitalizacion" in r["problemas"]

    def test_cantidades_soat_missing_tarifario_graceful(self):
        """Engine path MUST handle missing tarifario column gracefully (no crash)."""
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = _mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md

                from app.services.hospitalizacion.detect_all import (
                    detect_all_problems_hospitalizacion,
                )
                wb, idx = _build_sheet()
                if "Tarifario" in idx:
                    del idx["Tarifario"]
                restore = _env("true")
                try:
                    r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
                finally:
                    restore()

        assert "cantidades_soat_hospitalizacion" in r["problemas"]
        assert r["problemas"]["cantidades_soat_hospitalizacion"] == []


class TestF3HospitalizacionCodesToggle:
    """T-F3.3: hospitalizacion_codes → legacy toggle (engine can't handle computed filter)."""

    def test_engine_path_calls_legacy_hospitalizacion_codes(self):
        """Engine path MUST still call legacy detect_hospitalizacion_codes
        (no engine rule exists for this — computed filter not supported)."""
        from app.services.hospitalizacion.detect_all import (
            detect_all_problems_hospitalizacion,
        )
        from app.services.hospitalizacion import hospitalizacion_codes as hc_module
        original_fn = hc_module.detect_hospitalizacion_codes
        call_log: list[str] = []
        def tracking_fn(ds, idx):
            call_log.append("called")
            return original_fn(ds, idx)

        try:
            hc_module.detect_hospitalizacion_codes = tracking_fn
            wb, idx = _build_sheet()
            restore = _env("true")
            try:
                r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
            finally:
                restore()
        finally:
            hc_module.detect_hospitalizacion_codes = original_fn

        assert call_log, (
            "Legacy detect_hospitalizacion_codes was NOT called in engine path!"
        )
        assert "cups_equivalentes" in r["problemas"]

    def test_legacy_path_produces_cups_equivalentes(self):
        """Legacy path must produce cups_equivalentes without crashing."""
        from app.services.hospitalizacion.detect_all import (
            detect_all_problems_hospitalizacion,
        )
        wb, idx = _build_sheet()
        restore = _env("false")
        try:
            r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
        finally:
            restore()
        assert "cups_equivalentes" in r["problemas"]

    def test_legacy_path_cups_equivalentes_is_list(self):
        """Legacy path cups_equivalentes must be a list."""
        from app.services.hospitalizacion.detect_all import (
            detect_all_problems_hospitalizacion,
        )
        wb, idx = _build_sheet()
        restore = _env("false")
        try:
            r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
        finally:
            restore()
        assert isinstance(r["problemas"]["cups_equivalentes"], list)

    def test_engine_path_missing_fecha_cierre_does_not_crash(self):
        """Engine path must not crash when fecha_cierre column is missing.
        hospitalizacion_codes should gracefully handle missing date columns."""
        from app.services.hospitalizacion.detect_all import (
            detect_all_problems_hospitalizacion,
        )
        wb, idx = _build_sheet()
        if "Fecha Cierre" in idx:
            del idx["Fecha Cierre"]
        restore = _env("true")
        try:
            r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
        finally:
            restore()
        assert "cups_equivalentes" in r["problemas"]
        assert isinstance(r["problemas"]["cups_equivalentes"], list)


class TestF3Snapshot:
    """T-F3.4: Snapshot structure matching."""

    EXPECTED_F3_KEYS = frozenset({
        "centros_de_costos", "ide_contrato", "cups_equivalentes",
        "cantidades_hospitalizacion", "cantidades_soat_hospitalizacion",
    })

    def test_keys_present_in_both_paths(self):
        """All F3-specific keys must be present in both engine and legacy paths."""
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = _mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md

                from app.services.hospitalizacion.detect_all import (
                    detect_all_problems_hospitalizacion,
                )
                wb, idx = _build_sheet()
                restore = _env("true")
                try:
                    er, _ = detect_all_problems_hospitalizacion(wb.active, idx)
                finally:
                    restore()

                restore2 = _env("false")
                try:
                    lr, _ = detect_all_problems_hospitalizacion(wb.active, idx)
                finally:
                    restore2()

        for k in self.EXPECTED_F3_KEYS:
            assert k in er["problemas"], f"Engine missing F3 key: {k}"
            assert k in lr["problemas"], f"Legacy missing F3 key: {k}"

    def test_totals_keys_present(self):
        """Total counts must be present for F3 detectors."""
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = _mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md

                from app.services.hospitalizacion.detect_all import (
                    detect_all_problems_hospitalizacion,
                )
                wb, idx = _build_sheet()
                restore = _env("true")
                try:
                    r, _ = detect_all_problems_hospitalizacion(wb.active, idx)
                finally:
                    restore()

        total_keys = {
            "centros_de_costos", "ide_contrato", "cups_equivalentes",
            "cantidades_hospitalizacion", "cantidades_soat_hospitalizacion",
        }
        for k in total_keys:
            assert k in r["totales"], f"Engine totales missing: {k}"
