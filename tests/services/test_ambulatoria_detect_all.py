"""Tests for app/services/ambulatoria/detect_all.py.

Strict TDD: tests written BEFORE implementation.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.ambulatoria.detect_all import detect_all_problems_ambulatoria


@pytest.fixture
def workbook_minimal() -> Workbook:
    """Crea un workbook con headers mínimos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    return wb


class TestDetectAllProblemsAmbulatoria:
    """Tests para detect_all_problems_ambulatoria."""

    def _run(self, ws, indices):
        result, _ = detect_all_problems_ambulatoria(ws, indices)
        return result

    def test_retorna_dict_con_key_problemas(self, workbook_minimal: Workbook) -> None:
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert "problemas" in result

    def test_retorna_area_ambulatoria(self, workbook_minimal: Workbook) -> None:
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert result.get("area") == "ambulatoria"

    def test_resultado_incluye_normalizados(self, workbook_minimal: Workbook) -> None:
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert "normalizados" in result["problemas"]
        assert isinstance(result["problemas"]["normalizados"], list)

    def test_missing_columns_present(self, workbook_minimal: Workbook) -> None:
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        indices = {"numero_factura": 0}
        result = self._run(ws, indices)
        assert "missing_columns" in result
