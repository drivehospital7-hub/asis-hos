"""Tests para app/services/transversales/procedimiento_contratado.py.

STRICT TDD — Test escrito antes que la implementación.
"""

from __future__ import annotations

from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from app.services.transversales.normalize import normalize_invoice


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_ws(headers: list[str], data_rows: list[list]) -> Workbook.active:
    """Crea una Worksheet con headers en fila 1 y datos desde fila 2."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    return ws


def _build_indices(headers: list[str]) -> dict[str, int]:
    """Construye índices 0-based desde lista de headers."""
    return {h: i for i, h in enumerate(headers)}


REQUIRED = ["numero_factura", "codigo_entidad_cobrar", "codigo"]


def _mock_eps_instance(cod_contrato: str, eps: str) -> MagicMock:
    ec = MagicMock()
    ec.cod_contrato = cod_contrato
    ec.eps = eps
    return ec


def _mock_proc_instance(cups: str) -> MagicMock:
    p = MagicMock()
    p.cups = cups
    return p


def _make_mock_session(
    pairs: list[tuple[str, str]],
    eps_names: dict[str, str],
    nota_urgencias_cups: list[str] | None = None,
    cap_cups: dict[int, list[str]] | None = None,
) -> MagicMock:
    """Crea un mock de sesión SQLAlchemy que retorna los datos dados.

    Args:
        pairs: Lista de (cod_contrato, cups) — resultados del primer query
        eps_names: Dict {cod_contrato: eps} — resultados del segundo query
        nota_urgencias_cups: Optional lista de CUPS para nota_hoja id=1 y 27 (tercer .all())
        cap_cups: Optional dict {id_nota_hoja: [cups]} para nota_hoja 2 y 3 (4to .all())
    """
    nota_urgencias_cups = nota_urgencias_cups or []
    cap_cups = cap_cups or {}
    mock_session = MagicMock()
    mock_query = MagicMock()
    mock_query.join.return_value = mock_query
    mock_query.filter.return_value = mock_query

    # First .all() → join results
    join_results = [
        (_mock_eps_instance(cod, eps_names.get(cod, cod)), _mock_proc_instance(cups))
        for cod, cups in pairs
    ]
    # Second .all() → EPS list
    eps_objects = [
        _mock_eps_instance(cod, name) for cod, name in eps_names.items()
    ]
    # Third .all() → nota_hoja id=1 y 27 procedimientos
    nota_urgencias_objects = [_mock_proc_instance(cups) for cups in nota_urgencias_cups]
    # Fourth .all() → CAP nota_hoja id=2 y 3 (tuples of id_nota_hoja, cups)
    cap_objects = [
        (nt_id, cups) for nt_id, cups_list in cap_cups.items() for cups in cups_list
    ]

    mock_query.all.side_effect = [join_results, eps_objects, nota_urgencias_objects, cap_objects]
    mock_session.query.return_value = mock_query
    return mock_session


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestDetectCupsSinContrato:
    """Tests para detect_cups_sin_contrato."""

    # ── 1. Columnas faltantes ──────────────────────────────────────────────

    def test_missing_numero_factura_returns_empty(self):
        """Cuando falta numero_factura, retorna []."""
        ws = _make_ws(
            headers=["codigo_entidad_cobrar", "codigo"],
            data_rows=[["ESS118", "CUPS001"]],
        )
        indices = _build_indices(["codigo_entidad_cobrar", "codigo"])

        from app.services.transversales.procedimiento_contratado import (
            detect_cups_sin_contrato,
        )
        result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    def test_missing_codigo_entidad_cobrar_returns_empty(self):
        """Cuando falta codigo_entidad_cobrar, retorna []."""
        ws = _make_ws(
            headers=["numero_factura", "codigo"],
            data_rows=[["FAC-001", "CUPS001"]],
        )
        indices = _build_indices(["numero_factura", "codigo"])

        from app.services.transversales.procedimiento_contratado import (
            detect_cups_sin_contrato,
        )
        result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    def test_missing_codigo_returns_empty(self):
        """Cuando falta codigo, retorna []."""
        ws = _make_ws(
            headers=["numero_factura", "codigo_entidad_cobrar"],
            data_rows=[["FAC-001", "ESS118"]],
        )
        indices = _build_indices(["numero_factura", "codigo_entidad_cobrar"])

        from app.services.transversales.procedimiento_contratado import (
            detect_cups_sin_contrato,
        )
        result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    # ── 2. DB no disponible ────────────────────────────────────────────────

    def test_db_not_available_returns_empty(self):
        """Cuando SessionLocal lanza excepción, retorna []."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["FAC-001", "ESS118", "CUPS001"]],
        )
        indices = _build_indices(REQUIRED)

        with patch("app.database.SessionLocal", side_effect=RuntimeError("DB down")):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    # ── 3. Happy path: todos contratados ───────────────────────────────────

    def test_happy_path_all_contracted(self):
        """Todos los pares (entidad, CUPS) están contratados → []."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["FAC-001", "ESS118", "CUPS001"]],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR ESS118"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    # ── 4. CUPS no contratado ──────────────────────────────────────────────

    def test_cups_not_contratado_detected(self):
        """CUPS no contratado → error con todos los campos."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["FAC-001", "ESS118", "CUPS999"]],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR ESS118"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        err = result[0]
        assert err["factura"] == "FAC-001"
        assert err["codigo"] == "CUPS999"
        assert err["codigo_entidad_cobrar"] == "ESS118"
        assert err["entidad"] == "EMSSANAR ESS118"
        assert "CUPS999" in err["problema"]
        assert "ESS118" in err["problema"]

    def test_multiple_rows_some_errors(self):
        """Cuando algunas filas tienen error y otras no, solo errores."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[
                ["FAC-001", "ESS118", "CUPS001"],   # OK
                ["FAC-002", "ESS118", "CUPS999"],   # ERROR
                ["FAC-003", "ESS119", "CUPS003"],   # OK
            ],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001"), ("ESS119", "CUPS003")],
            eps_names={"ESS118": "EMSSANAR", "ESS119": "OTRA EPS"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-002"

    # ── 5. Normalización ──────────────────────────────────────────────────

    def test_normalization_case_and_whitespace(self):
        """Normalización: .strip().upper() en ambos códigos."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["FAC-001", "  ess118  ", "  cups001  "]],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    # ── 6. Columna procedimiento opcional ──────────────────────────────────

    def test_includes_procedimiento_when_available(self):
        """Cuando la columna procedimiento existe, se incluye en el error."""
        headers = REQUIRED + ["procedimiento"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "CUPS999", "PROC-X"]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["procedimiento"] == "PROC-X"

    # ── 7. Filas sin factura se ignoran ────────────────────────────────────

    def test_empty_invoice_skipped(self):
        """Filas sin número de factura se ignoran silenciosamente."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[
                [None, "ESS118", "CUPS999"],
                ["", "ESS118", "CUPS001"],
            ],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    # ── 9. Tarifario farmacia se ignora ─────────────────────────────────────

    def test_tarifario_farmacia_skipped(self):
        """Filas con tarifario=Suminstros, Medicamentos se ignoran."""
        headers = REQUIRED + ["tarifario"]
        ws = _make_ws(
            headers=headers,
            data_rows=[
                # FAC-001 tiene una fila de farmacia y otra normal
                ["FAC-001", "ESS118", "CUPS999", "Suminstros, Medicamentos"],
                ["FAC-001", "ESS118", "CUPS001", "Consultas"],
                # FAC-002 solo tiene farmacia
                ["FAC-002", "ESS118", "CUPS999", "Suminstros, Medicamentos"],
            ],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        # Solo la fila normal no contratada debería aparecer
        # FAC-001 fila 1 (farmacia) → skip
        # FAC-001 fila 2 (CUPS001 contratado) → OK
        # FAC-002 (farmacia) → skip
        assert len(result) == 0

    def test_tarifario_farmacia_skip_still_detects_errors(self):
        """Otras filas de la misma factura se siguen procesando."""
        headers = REQUIRED + ["tarifario"]
        ws = _make_ws(
            headers=headers,
            data_rows=[
                ["FAC-001", "ESS118", "CUPS001", "Suminstros, Medicamentos"],
                ["FAC-001", "ESS118", "CUPS999", "Consultas"],
            ],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "CUPS999"

    def test_unknown_entity_without_data_skipped(self):
        """Entidad sin procedimientos en DB se salta (no se valida)."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["FAC-001", "UNKNOWN_ENT", "CUPS999"]],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    # ── 11. Código Equivalente CUPS ─────────────────────────────────────────

    def test_codigo_equiv_fallback_avoids_error(self):
        """Código principal no encontrado pero el equivalente sí → sin error."""
        headers = REQUIRED + ["codigo_equiv"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "CUPS999", "CUPS001"]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    def test_codigo_equiv_not_found_still_errors(self):
        """Ni el código principal ni el equivalente están contratados → error."""
        headers = REQUIRED + ["codigo_equiv"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "CUPS999", "CUPS888"]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "CUPS999"

    def test_codigo_equiv_column_missing_fallback_old_behavior(self):
        """Sin columna codigo_equiv, el comportamiento es el de siempre → error."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["FAC-001", "ESS118", "CUPS999"]],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1

    def test_codigo_equiv_empty_cell_ignored(self):
        """Columna codigo_equiv existe pero celda vacía → solo usa código principal."""
        headers = REQUIRED + ["codigo_equiv"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "CUPS999", None]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "CUPS999"

    # ── 12. Excepción FEV con autorización ──────────────────────────────────

    def test_fev_eps037_autorizada_no_error(self):
        """Factura FEV + EPS037 → sin error aunque CUPS no esté contratado."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["FEV-001", "EPS037", "CUPS999"]],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("EPS037", "CUPS001")],
            eps_names={"EPS037": "EPS 037"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    def test_fev_epss41_autorizada_no_error(self):
        """Factura FEV + EPSS41 → sin error aunque CUPS no esté contratado."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["FEV-002", "EPSS41", "CUPS888"]],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("EPSS41", "CUPS001")],
            eps_names={"EPSS41": "EPS S41"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    def test_fev_otra_entidad_si_error(self):
        """Factura FEV pero con otra entidad → error (no aplica excepción)."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["FEV-003", "ESS118", "CUPS999"]],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "CUPS001")],
            eps_names={"ESS118": "EMSSANAR"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1

    def test_no_fev_eps037_si_error(self):
        """Factura normal con EPS037 → error (no es FEV)."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["FAC-001", "EPS037", "CUPS999"]],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("EPS037", "CUPS001")],
            eps_names={"EPS037": "EPS 037"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1

    # ── 13. Entidad sin procedimientos en DB se ignora ──────────────────────

    def test_entity_without_procedures_skipped(self):
        """Entidades que no tienen procedimientos cargados se ignoran."""
        headers = REQUIRED
        ws = _make_ws(
            headers=headers,
            data_rows=[
                # ASMET SALUD no tiene procedimientos en DB
                ["FAC-001", "ESS062", "CUPS001"],
                # MALLAMAS sí tiene
                ["FAC-002", "EPSI05", "CUPS999"],
            ],
        )
        indices = _build_indices(headers)
        # Solo MALLAMAS aparece en los resultados del JOIN
        mock_session = _make_mock_session(
            pairs=[("EPSI05", "CUPS001")],
            eps_names={"EPSI05": "MALLAMAS", "ESS062": "ASMET SALUD"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        # FAC-001 (ASMET sin datos) → skip
        # FAC-002 (MALLAMAS, CUPS999 no contratado) → error
        assert len(result) == 1
        assert result[0]["factura"] == "FAC-002"

    # ── 14. Excepción responsable urgencias ──────────────────────────────────

    def test_urgencias_facturador_cups_in_nota1_no_error(self):
        """Responsable urgencias + CUPS en nota1/27 → sin error (excepción)."""
        headers = REQUIRED + ["responsable_cierra"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "965201", "ARIAS CULCHA ANGIE CAROLINA"]],
        )
        indices = _build_indices(headers)
        # "965201" NOT in pares_validos para ESS118
        mock_session = _make_mock_session(
            pairs=[("ESS118", "878001")],
            eps_names={"ESS118": "EMSSANAR ESS E.S.S."},
            nota_urgencias_cups=["965201"],
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    def test_urgencias_facturador_cups_not_in_nota1_errors(self):
        """Responsable urgencias + CUPS no en nota1/27 → error."""
        headers = REQUIRED + ["responsable_cierra"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "999999", "ESPAÑA DIAZ LORENY ALEJANDRA"]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "878001")],
            eps_names={"ESS118": "EMSSANAR ESS E.S.S."},
            nota_urgencias_cups=["965201"],
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "999999"

    def test_urgencias_facturador_codigo_equiv_in_nota1_no_error(self):
        """Responsable urgencias + codigo_equiv en nota1/27 → sin error."""
        headers = REQUIRED + ["codigo_equiv", "responsable_cierra"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "CUPS999", "965201", "ARIAS CULCHA ANGIE CAROLINA"]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "878001")],
            eps_names={"ESS118": "EMSSANAR ESS E.S.S."},
            nota_urgencias_cups=["965201"],
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    def test_urgencias_facturador_nota1_empty_errors(self):
        """Responsable urgencias + nota1/27 vacío → error (fails closed)."""
        headers = REQUIRED + ["responsable_cierra"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "965201", "MEZA FERNANDEZ CARLOS OMAR"]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "878001")],
            eps_names={"ESS118": "EMSSANAR ESS E.S.S."},
            nota_urgencias_cups=[],
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "965201"

    def test_urgencias_facturador_column_missing_normal_validation(self):
        """Columna responsable_cierra ausente → validación normal."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["FAC-001", "ESS118", "965201"]],
        )
        indices = _build_indices(REQUIRED)
        # No incluir "responsable_cierra" en indices
        mock_session = _make_mock_session(
            pairs=[("ESS118", "878001")],
            eps_names={"ESS118": "EMSSANAR ESS E.S.S."},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        # 965201 no está contratado para ESS118 → error
        assert len(result) == 1
        assert result[0]["codigo"] == "965201"

    def test_urgencias_facturador_empty_cell_normal_validation(self):
        """Celda responsable_cierra vacía → validación normal."""
        headers = REQUIRED + ["responsable_cierra"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "965201", None]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "878001")],
            eps_names={"ESS118": "EMSSANAR ESS E.S.S."},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        # Celda vacía → no aplica excepción → error normal
        assert len(result) == 1
        assert result[0]["codigo"] == "965201"

    def test_urgencias_facturador_double_space_matches_norm(self):
        """Doble espacio en nombre del facturador -> colapsado -> coincide con FACTURADORES_URGENCIAS."""
        headers = REQUIRED + ["responsable_cierra"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "878001", "PAEZ  YULIETH DANIELA"]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "OTHER")],
            eps_names={"ESS118": "EMSSANAR ESS E.S.S."},
            nota_urgencias_cups=["878001"],
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        # Doble espacio colapsado -> coincide con "PAEZ YULIETH DANIELA" -> sin error
        assert result == []

    # ── 15. CAP exception — ESS118 / EPSS41 ───────────────────────────────────

    def test_cap_ess118_cups_in_nota3_no_error(self):
        """CAP + ESS118 + CUPS en nota_hoja id=3 → sin error (excepción CAP)."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["CAP-2024-001", "ESS118", "878001"]],
        )
        indices = _build_indices(REQUIRED)
        # ESS118 tiene datos contractuales pero 878001 NO está en pares_validos
        # Sin CAP exception → error. Con CAP exception → no error (capturado antes de entidades_con_datos)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "OTHER_CUPS")],
            eps_names={"ESS118": "EMSSANAR ESS118"},
            cap_cups={3: ["878001"]},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    def test_cap_epss41_cups_in_nota2_no_error(self):
        """CAP + EPSS41 + CUPS en nota_hoja id=2 → sin error (excepción CAP)."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["CAP-2024-002", "EPSS41", "965201"]],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("EPSS41", "OTHER_CUPS")],
            eps_names={"EPSS41": "NUEVA EPS EPSS41"},
            cap_cups={2: ["965201"]},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    def test_cap_ess118_cups_not_in_nota3_errors(self):
        """CAP + ESS118 + CUPS NO en nota_hoja id=3 → error."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["CAP-2024-003", "ESS118", "999999"]],
        )
        indices = _build_indices(REQUIRED)
        # ESS118 tiene datos contractuales para que el fall-through valide
        mock_session = _make_mock_session(
            pairs=[("ESS118", "878001")],
            eps_names={"ESS118": "EMSSANAR ESS118"},
            cap_cups={3: ["878001"]},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "999999"

    def test_cap_epss41_cups_not_in_nota2_errors(self):
        """CAP + EPSS41 + CUPS NO en nota_hoja id=2 → error."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["CAP-2024-004", "EPSS41", "888888"]],
        )
        indices = _build_indices(REQUIRED)
        mock_session = _make_mock_session(
            pairs=[("EPSS41", "965201")],
            eps_names={"EPSS41": "NUEVA EPS EPSS41"},
            cap_cups={2: ["965201"]},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "888888"

    def test_cap_ess118_nota3_empty_errors(self):
        """CAP + ESS118 + nota_hoja id=3 vacía → error (fails closed)."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["CAP-2024-005", "ESS118", "878001"]],
        )
        indices = _build_indices(REQUIRED)
        # ESS118 tiene datos contractuales pero 878001 no está contratado
        # nota_hoja id=3 está vacía → falls through → error
        mock_session = _make_mock_session(
            pairs=[("ESS118", "333333")],
            eps_names={"ESS118": "EMSSANAR ESS118"},
            cap_cups={3: []},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "878001"

    def test_non_cap_ess118_standard_validation(self):
        """Factura NO-CAP + ESS118 → validación normal (sin excepción)."""
        ws = _make_ws(
            headers=REQUIRED,
            data_rows=[["FAC-2024-001", "ESS118", "878001"]],
        )
        indices = _build_indices(REQUIRED)
        # ESS118 tiene datos pero 878001 no está contratado
        mock_session = _make_mock_session(
            pairs=[("ESS118", "333333")],
            eps_names={"ESS118": "EMSSANAR ESS118"},
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "878001"

    # ── 16. Urgencias bypass — entidades EN _ENTIDADES_NOTA_URGENCIAS ─────────

    def test_urgencias_facturador_entity_en_lista_cups_in_nota(self):
        """EPSS08 (en _ENTIDADES_NOTA_URGENCIAS) + urgencias + CUPS en nota → no error (regression)."""
        headers = REQUIRED + ["responsable_cierra"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "EPSS08", "965201", "MEZA FERNANDEZ CARLOS OMAR"]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("EPSS08", "878001")],
            eps_names={"EPSS08": "EPS S08"},
            nota_urgencias_cups=["965201"],
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    def test_urgencias_facturador_entity_en_lista_cups_not_in_nota(self):
        """EPSS08 + urgencias + CUPS ni nota ni pares → error."""
        headers = REQUIRED + ["responsable_cierra"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "EPSS08", "999999", "MEZA FERNANDEZ CARLOS OMAR"]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("EPSS08", "878001")],
            eps_names={"EPSS08": "EPS S08"},
            nota_urgencias_cups=["965201"],
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "999999"

    # ── 17. Urgencias bypass — entidades FUERA de _ENTIDADES_NOTA_URGENCIAS ───

    def test_urgencias_entity_no_lista_cups_in_pares_validos(self):
        """ESS118 (FUERA de lista) + urgencias + CUPS en pares_validos → no error (fallback)."""
        headers = REQUIRED + ["responsable_cierra"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "878001", "ARIAS CULCHA ANGIE CAROLINA"]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "878001")],
            eps_names={"ESS118": "EMSSANAR ESS E.S.S."},
            nota_urgencias_cups=["965201"],
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    def test_urgencias_bug_scenario(self):
        """ESS118 + CUPS 903437 + MEZA FERNANDEZ CARLOS OMAR → no error (original bug)."""
        headers = REQUIRED + ["responsable_cierra"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "903437", "MEZA FERNANDEZ CARLOS OMAR"]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "878001")],
            eps_names={"ESS118": "EMSSANAR ESS E.S.S."},
            nota_urgencias_cups=["903437"],
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert result == []

    def test_non_urgencias_biller_unaffected(self):
        """ESS118 + CUPS 965201 + facturador NO urgencias → error (normal validation)."""
        headers = REQUIRED + ["responsable_cierra"]
        ws = _make_ws(
            headers=headers,
            data_rows=[["FAC-001", "ESS118", "965201", "UN NOMBRE CUALQUIERA"]],
        )
        indices = _build_indices(headers)
        mock_session = _make_mock_session(
            pairs=[("ESS118", "878001")],
            eps_names={"ESS118": "EMSSANAR ESS E.S.S."},
            nota_urgencias_cups=["965201"],
        )

        with patch("app.database.SessionLocal", return_value=mock_session):
            from app.services.transversales.procedimiento_contratado import (
                detect_cups_sin_contrato,
            )
            result = detect_cups_sin_contrato(ws, indices)

        assert len(result) == 1
        assert result[0]["codigo"] == "965201"
