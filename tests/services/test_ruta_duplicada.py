"""Tests para app/services/transversales/ruta_duplicada.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.constants import CONVENIO_PYP, CONVENIO_ASISTENCIAL
from app.services.transversales.ruta_duplicada import detect_ruta_duplicada


@pytest.fixture
def workbook_with_headers() -> Workbook:
    """Crea un workbook con headers mínimos para ruta duplicada."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Nº Identificación")
    ws.cell(row=1, column=3, value="Convenio Facturado")
    return wb


class TestDetectRutaDuplicada:
    """Tests para detect_ruta_duplicada."""

    def test_detecta_paciente_con_multiples_facturas_pyp_threshold_3(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe detectar pacientes con >= 3 facturas en PyP (threshold=3)."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="PAC-001")
        ws.cell(row=2, column=3, value=CONVENIO_PYP)
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value="PAC-001")
        ws.cell(row=3, column=3, value=CONVENIO_PYP)
        ws.cell(row=4, column=1, value="FAC-003")
        ws.cell(row=4, column=2, value="PAC-001")
        ws.cell(row=4, column=3, value=CONVENIO_PYP)
        ws.cell(row=5, column=1, value="FAC-004")
        ws.cell(row=5, column=2, value="PAC-002")
        ws.cell(row=5, column=3, value=CONVENIO_PYP)
        ws.cell(row=6, column=1, value="FAC-005")
        ws.cell(row=6, column=2, value="PAC-002")
        ws.cell(row=6, column=3, value=CONVENIO_PYP)

        indices = {
            "numero_factura": 0,
            "identificacion": 1,
            "convenio_facturado": 2,
        }
        result = detect_ruta_duplicada(ws, indices, threshold=3)

        identificaciones = [r["identificacion"] for r in result]
        assert "PAC-001" in identificaciones
        assert "PAC-002" not in identificaciones

    def test_threshold_2_detecta_menos_facturas(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Con threshold=2 debe detectar pacientes con >= 2 facturas."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="PAC-001")
        ws.cell(row=2, column=3, value=CONVENIO_PYP)
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value="PAC-001")
        ws.cell(row=3, column=3, value=CONVENIO_PYP)
        ws.cell(row=4, column=1, value="FAC-003")
        ws.cell(row=4, column=2, value="PAC-002")
        ws.cell(row=4, column=3, value=CONVENIO_PYP)
        ws.cell(row=5, column=1, value="FAC-004")
        ws.cell(row=5, column=2, value="PAC-002")
        ws.cell(row=5, column=3, value=CONVENIO_PYP)

        indices = {
            "numero_factura": 0,
            "identificacion": 1,
            "convenio_facturado": 2,
        }
        result = detect_ruta_duplicada(ws, indices, threshold=2)

        identificaciones = [r["identificacion"] for r in result]
        assert "PAC-001" in identificaciones
        assert "PAC-002" in identificaciones

    def test_ignora_facturas_no_pyp(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe ignorar facturas con convenio diferente a PyP."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="PAC-001")
        ws.cell(row=2, column=3, value=CONVENIO_ASISTENCIAL)
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value="PAC-001")
        ws.cell(row=3, column=3, value=CONVENIO_ASISTENCIAL)
        ws.cell(row=4, column=1, value="FAC-003")
        ws.cell(row=4, column=2, value="PAC-001")
        ws.cell(row=4, column=3, value=CONVENIO_ASISTENCIAL)

        indices = {
            "numero_factura": 0,
            "identificacion": 1,
            "convenio_facturado": 2,
        }
        result = detect_ruta_duplicada(ws, indices, threshold=3)

        assert len(result) == 0

    def test_sin_indices_retorna_vacio(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Sin indices necesarios debe retornar lista vacía."""
        indices = {
            "numero_factura": None,
            "identificacion": None,
            "convenio_facturado": None,
        }
        result = detect_ruta_duplicada(
            workbook_with_headers.active, indices, threshold=3
        )

        assert result == []

    def test_default_threshold_es_3(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Threshold por defecto debe ser 3."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="PAC-001")
        ws.cell(row=2, column=3, value=CONVENIO_PYP)
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value="PAC-001")
        ws.cell(row=3, column=3, value=CONVENIO_PYP)
        # PAC-001 solo tiene 2 facturas -> no debe aparecer con threshold=3 default

        indices = {
            "numero_factura": 0,
            "identificacion": 1,
            "convenio_facturado": 2,
        }
        result = detect_ruta_duplicada(ws, indices)

        assert len(result) == 0

    def test_resultado_incluye_cantidad(
        self, workbook_with_headers: Workbook
    ) -> None:
        """El resultado debe incluir cantidad de facturas."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="PAC-001")
        ws.cell(row=2, column=3, value=CONVENIO_PYP)
        ws.cell(row=3, column=1, value="FAC-002")
        ws.cell(row=3, column=2, value="PAC-001")
        ws.cell(row=3, column=3, value=CONVENIO_PYP)
        ws.cell(row=4, column=1, value="FAC-003")
        ws.cell(row=4, column=2, value="PAC-001")
        ws.cell(row=4, column=3, value=CONVENIO_PYP)

        indices = {
            "numero_factura": 0,
            "identificacion": 1,
            "convenio_facturado": 2,
        }
        result = detect_ruta_duplicada(ws, indices, threshold=3)

        assert len(result) == 1
        assert result[0]["cantidad"] == 3
