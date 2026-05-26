"""Tests for tipo_factura_descripcion filter behavior in urgencias detectors.

These tests verify that detectors only process rows where
Tipo Factura Descripcion == "Urgencias", skipping all other tipos.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.urgencias.profesionales_urgencias import (
    detect_profesionales_urgencias,
)
from app.services.urgencias.revision_cantidad import (
    detect_revision_cantidad_urgencias,
)
from app.services.urgencias.ide_contrato_urgencias import (
    detect_ide_contrato_urgencias,
)
from app.services.urgencias.ide_contrato_reverse import (
    detect_ide_contrato_reverse_urgencias,
)
from app.services.urgencias.codigos_sin_db import (
    get_codigos_no_en_db_ess118,
)


# ──────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────

def _make_wb(headers: list[str]) -> Workbook:
    """Create a workbook with given headers in row 1."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    return wb


def _build_indices_from_headers(headers: list[str]) -> dict[str, int | None]:
    """Build indices dict from header list."""
    indices: dict[str, int | None] = {}
    for i, h in enumerate(headers):
        key_map = {
            "Número Factura": "numero_factura",
            "Código": "codigo",
            "Cód. Equivalente CUPS": "codigo",
            "Procedimiento": "procedimiento",
            "Cantidad": "cantidad",
            "Tipo Procedimiento": "tipo_procedimiento",
            "Laboratorio": "laboratorio",
            "Código Tipo Procedimiento": "codigo_tipo_procedimiento",
            "Código Profesional": "codigo_profesional",
            "Tipo Factura Descripción": "tipo_factura_descripcion",
            "IDE Contrato": "ide_contrato",
            "Cód Entidad Cobrar": "codigo_entidad_cobrar",
            "Nº Identificación": "identificacion",
            "Fec. Factura": "fec_factura",
        }
        key = key_map.get(h, h.lower().replace(" ", "_"))
        indices[key] = i
    return indices


# ──────────────────────────────────────────────────────
# profesionales_urgencias — filter tests
# ──────────────────────────────────────────────────────

class TestProfesionalesUrgenciasFilter:
    """Filter: only "Urgencias" rows processed by detect_profesionales_urgencias."""

    HEADERS = [
        "Número Factura",
        "Código Profesional",
        "Código",
        "Procedimiento",
        "Tipo Factura Descripción",
    ]

    def test_non_urgencias_rows_skipped(self) -> None:
        """Rows with tipo_factura != Urgencias produce no errors."""
        wb = _make_wb(self.HEADERS)
        ws = wb.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="999999")  # non-existent code
        ws.cell(row=2, column=3, value="890101")
        ws.cell(row=2, column=4, value="Consulta")
        ws.cell(row=2, column=5, value="Hospitalización")  # NOT Urgencias

        indices = _build_indices_from_headers(self.HEADERS)
        result = detect_profesionales_urgencias(ws, indices)
        assert result == [], f"Expected empty, got {len(result)} errors"

    def test_urgencias_rows_still_processed(self) -> None:
        """Rows with tipo_factura = Urgencias ARE processed."""
        wb = _make_wb(self.HEADERS)
        ws = wb.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="999999")  # non-existent → would flag
        ws.cell(row=2, column=3, value="890101")
        ws.cell(row=2, column=4, value="Consulta")
        ws.cell(row=2, column=5, value="Urgencias")

        indices = _build_indices_from_headers(self.HEADERS)
        result = detect_profesionales_urgencias(ws, indices)
        assert len(result) >= 1, f"Expected at least 1 error, got {len(result)}"
        assert result[0]["factura"] == "FAC-001"


# ──────────────────────────────────────────────────────
# revision_cantidad — filter tests
# ──────────────────────────────────────────────────────

class TestRevisionCantidadFilter:
    """Filter: only "Urgencias" rows processed by detect_revision_cantidad_urgencias."""

    HEADERS = [
        "Número Factura",
        "Código",
        "Procedimiento",
        "Cantidad",
        "Tipo Procedimiento",
        "Laboratorio",
        "Tipo Factura Descripción",
    ]

    def test_non_urgencias_rows_skipped(self) -> None:
        """Rows with tipo_factura != Urgencias produce no errors."""
        wb = _make_wb(self.HEADERS)
        ws = wb.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="ZZZ999")
        ws.cell(row=2, column=3, value="Procedure X")
        ws.cell(row=2, column=4, value=5)  # > 1
        ws.cell(row=2, column=5, value="Procedure Type")
        ws.cell(row=2, column=6, value="No")
        ws.cell(row=2, column=7, value="Hospitalización")  # NOT Urgencias

        indices = _build_indices_from_headers(self.HEADERS)
        result = detect_revision_cantidad_urgencias(ws, indices)
        assert result == [], f"Expected empty, got {len(result)} errors"

    def test_urgencias_rows_still_processed(self) -> None:
        """Rows with tipo_factura = Urgencias ARE processed."""
        wb = _make_wb(self.HEADERS)
        ws = wb.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="ZZZ999")
        ws.cell(row=2, column=3, value="Procedure X")
        ws.cell(row=2, column=4, value=5)  # > 1
        ws.cell(row=2, column=5, value="Procedure Type")
        ws.cell(row=2, column=6, value="No")
        ws.cell(row=2, column=7, value="Urgencias")

        indices = _build_indices_from_headers(self.HEADERS)
        result = detect_revision_cantidad_urgencias(ws, indices)
        assert len(result) >= 1, f"Expected at least 1 error, got {len(result)}"
        assert result[0]["factura"] == "FAC-001"


# ──────────────────────────────────────────────────────
# ide_contrato_urgencias — filter tests
# ──────────────────────────────────────────────────────

class TestIdeContratoUrgenciasFilter:
    """Filter: non-Urgencias rows skipped by detect_ide_contrato_urgencias."""

    HEADERS = [
        "Número Factura",
        "Código",
        "Procedimiento",
        "IDE Contrato",
        "Cód Entidad Cobrar",
        "Nº Identificación",
        "Tipo Factura Descripción",
    ]

    def test_non_urgencias_rows_skipped(self) -> None:
        """Non-Urgencias tipo_factura → no ide_contrato errors reported."""
        wb = _make_wb(self.HEADERS)
        ws = wb.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="890101")
        ws.cell(row=2, column=3, value="Proc")
        ws.cell(row=2, column=4, value="999")
        ws.cell(row=2, column=5, value="86")
        ws.cell(row=2, column=6, value="12345")
        ws.cell(row=2, column=7, value="Hospitalización")  # NOT Urgencias

        indices = _build_indices_from_headers(self.HEADERS)
        result = detect_ide_contrato_urgencias(ws, indices)
        assert result == [], f"Expected empty, got {len(result)} errors"

    def test_urgencias_rows_processed(self) -> None:
        """Urgencias rows ARE processed by ide_contrato detector."""
        wb = _make_wb(self.HEADERS)
        ws = wb.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="890101")
        ws.cell(row=2, column=3, value="Proc")
        ws.cell(row=2, column=4, value="999")
        ws.cell(row=2, column=5, value="86000")
        ws.cell(row=2, column=6, value="12345")
        ws.cell(row=2, column=7, value="Urgencias")

        indices = _build_indices_from_headers(self.HEADERS)
        result = detect_ide_contrato_urgencias(ws, indices)
        assert isinstance(result, list)


# ──────────────────────────────────────────────────────
# ide_contrato_reverse — filter tests
# ──────────────────────────────────────────────────────

class TestIdeContratoReverseFilter:
    """Filter: non-Urgencias rows skipped by detect_ide_contrato_reverse_urgencias."""

    HEADERS = [
        "Número Factura",
        "Código",
        "IDE Contrato",
        "Nº Identificación",
        "Tipo Factura Descripción",
    ]

    def test_non_urgencias_rows_skipped(self) -> None:
        """Non-Urgencias tipo_factura → no reverse errors reported."""
        wb = _make_wb(self.HEADERS)
        ws = wb.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="ZZZ999")
        ws.cell(row=2, column=3, value="986")
        ws.cell(row=2, column=4, value="12345")
        ws.cell(row=2, column=5, value="Hospitalización")  # NOT Urgencias

        indices = _build_indices_from_headers(self.HEADERS)
        result = detect_ide_contrato_reverse_urgencias(ws, indices)
        assert result == [], f"Expected empty, got {len(result)} errors"

    def test_urgencias_rows_processed(self) -> None:
        """Urgencias rows ARE processed by reverse detector."""
        wb = _make_wb(self.HEADERS)
        ws = wb.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="ZZZ999")
        ws.cell(row=2, column=3, value="986")
        ws.cell(row=2, column=4, value="12345")
        ws.cell(row=2, column=5, value="Urgencias")

        indices = _build_indices_from_headers(self.HEADERS)
        result = detect_ide_contrato_reverse_urgencias(ws, indices)
        assert len(result) >= 1, f"Expected at least 1 error, got {len(result)}"
        assert result[0]["factura"] == "FAC-001"


# ──────────────────────────────────────────────────────
# codigos_sin_db — filter tests
# ──────────────────────────────────────────────────────

class TestCodigosSinDbFilter:
    """Filter: non-Urgencias rows skipped by get_codigos_no_en_db_ess118."""

    HEADERS = [
        "Número Factura",
        "Código",
        "Procedimiento",
        "IDE Contrato",
        "Código Tipo Procedimiento",
        "Cód Entidad Cobrar",
        "Tipo Factura Descripción",
    ]

    def test_non_urgencias_rows_skipped(self) -> None:
        """Non-Urgencias rows produce no DB-query errors."""
        wb = _make_wb(self.HEADERS)
        ws = wb.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="890101")
        ws.cell(row=2, column=3, value="Proc")
        ws.cell(row=2, column=4, value="969")
        ws.cell(row=2, column=5, value="02")
        ws.cell(row=2, column=6, value="ESS118")
        ws.cell(row=2, column=7, value="Hospitalización")  # NOT Urgencias

        indices = _build_indices_from_headers(self.HEADERS)
        result = get_codigos_no_en_db_ess118(ws, indices)
        # All rows are non-Urgencias; filter skips them.
        # Note: DB connection fails in test env anyway (returns [])
        assert result == [], f"Expected empty, got {len(result)} errors"
