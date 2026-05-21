"""Tests para app/services/urgencias/cantidades_urgencias.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.services.urgencias.cantidades_urgencias import detect_cantidades_urgencias


@pytest.fixture
def wb_urgencias_headers() -> Workbook:
    """Crea un workbook con headers para validación de cantidades urgencias."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código")
    ws.cell(row=1, column=3, value="Procedimiento")
    ws.cell(row=1, column=4, value="Cantidad")
    ws.cell(row=1, column=5, value="Tipo Factura Descripción")
    return wb


class TestDetectCantidadesUrgencias:
    """Tests para detect_cantidades_urgencias."""

    def test_cantidad_mayor_1_con_codigo_restringido_genera_error(
        self, wb_urgencias_headers: Workbook
    ) -> None:
        """Código 05DSB01 con cantidad > 1 en Urgencias debe generar error."""
        ws = wb_urgencias_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="05DSB01")
        ws.cell(row=2, column=3, value="PROC A")
        ws.cell(row=2, column=4, value=3)
        ws.cell(row=2, column=5, value="Urgencias")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
        }
        result = detect_cantidades_urgencias(ws, indices)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"
        assert result[0]["codigo"] == "05DSB01"
        assert result[0]["cantidad"] == 3

    def test_cantidad_1_no_genera_error(
        self, wb_urgencias_headers: Workbook
    ) -> None:
        """Código 05DSB01 con cantidad 1 en Urgencias NO debe generar error."""
        ws = wb_urgencias_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="05DSB01")
        ws.cell(row=2, column=3, value="PROC A")
        ws.cell(row=2, column=4, value=1)
        ws.cell(row=2, column=5, value="Urgencias")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
        }
        result = detect_cantidades_urgencias(ws, indices)

        assert len(result) == 0

    def test_tipo_factura_no_urgencias_no_genera_error(
        self, wb_urgencias_headers: Workbook
    ) -> None:
        """Si Tipo Factura no es Urgencias, no se valida."""
        ws = wb_urgencias_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="05DSB01")
        ws.cell(row=2, column=3, value="PROC A")
        ws.cell(row=2, column=4, value=3)
        ws.cell(row=2, column=5, value="Odontología")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
        }
        result = detect_cantidades_urgencias(ws, indices)

        assert len(result) == 0

    def test_codigo_no_restringido_no_genera_error(
        self, wb_urgencias_headers: Workbook
    ) -> None:
        """Código no en URGENCIAS_CODIGOS_CANTIDAD_MAX_1 no se valida."""
        ws = wb_urgencias_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="890101")  # no está en lista restringida
        ws.cell(row=2, column=3, value="PROC A")
        ws.cell(row=2, column=4, value=5)
        ws.cell(row=2, column=5, value="Urgencias")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
        }
        result = detect_cantidades_urgencias(ws, indices)

        assert len(result) == 0

    def test_sin_indices_retorna_vacio(
        self, wb_urgencias_headers: Workbook
    ) -> None:
        """Si faltan índices necesarios, retorna lista vacía."""
        ws = wb_urgencias_headers.active
        result = detect_cantidades_urgencias(ws, {})
        assert result == []

    def test_no_duplica_facturas(
        self, wb_urgencias_headers: Workbook
    ) -> None:
        """Misma factura con múltiples filas se reporta una sola vez."""
        ws = wb_urgencias_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="05DSB01")
        ws.cell(row=2, column=3, value="PROC A")
        ws.cell(row=2, column=4, value=3)
        ws.cell(row=2, column=5, value="Urgencias")
        ws.cell(row=3, column=1, value="FAC-001")
        ws.cell(row=3, column=2, value="890601")
        ws.cell(row=3, column=3, value="PROC B")
        ws.cell(row=3, column=4, value=2)
        ws.cell(row=3, column=5, value="Urgencias")

        indices = {
            "numero_factura": 0,
            "codigo": 1,
            "procedimiento": 2,
            "cantidad": 3,
            "tipo_factura_descripcion": 4,
        }
        result = detect_cantidades_urgencias(ws, indices)

        assert len(result) == 1
