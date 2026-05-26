"""Tests for app/services/urgencias/duplicados_farmacia.py.

Cubre escenarios originales + filtro por tipo_factura_descripcion.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.constants.urgencias import VALOR_TARIFARIO_FARMACIA
from app.services.urgencias.duplicados_farmacia import detect_duplicados_farmacia


@pytest.fixture
def workbook_with_headers() -> Workbook:
    """Crea un workbook con headers mínimos para el detector."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código")
    ws.cell(row=1, column=3, value="Cantidad")
    ws.cell(row=1, column=4, value="Tarifario")
    ws.cell(row=1, column=5, value="Procedimiento")
    ws.cell(row=1, column=6, value="Código Tipo Procedimiento")
    ws.cell(row=1, column=7, value="Tipo Factura Descripción")
    return wb


_INDICES = {
    "numero_factura": 0,
    "codigo": 1,
    "cantidad": 2,
    "tarifario": 3,
    "procedimiento": 4,
    "codigo_tipo_procedimiento": 5,
    "tipo_factura_descripcion": 6,
}


def _write_farmacia_row(
    ws, row, factura, codigo, cantidad, tipo_proc, tipo_factura="Urgencias"
) -> None:
    """Helper: escribe una fila con tarifario farmacia y tipo_factura."""
    ws.cell(row=row, column=1, value=factura)
    ws.cell(row=row, column=2, value=codigo)
    ws.cell(row=row, column=3, value=cantidad)
    ws.cell(row=row, column=4, value=VALOR_TARIFARIO_FARMACIA)
    ws.cell(row=row, column=5, value="Med")
    ws.cell(row=row, column=6, value=tipo_proc)
    ws.cell(row=row, column=7, value=tipo_factura)


class TestDetectDuplicadosFarmacia:
    """Tests para detect_duplicados_farmacia (algoritmo grupal + filtro tipo_factura)."""

    # ── Escenario 1: Grupo 12 con duplicidad total → flag ──

    def test_grupo_12_duplicidad_total_retorna_flag(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Grupo 12 con 2 pares, cada uno x2 → 1 flag."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "890101", 1, "12")
        _write_farmacia_row(ws, 3, "FAC-001", "890101", 1, "12")
        _write_farmacia_row(ws, 4, "FAC-001", "890102", 2, "12")
        _write_farmacia_row(ws, 5, "FAC-001", "890102", 2, "12")

        result = detect_duplicados_farmacia(ws, _INDICES)
        assert len(result) == 1
        item = result[0]
        assert item["factura"] == "FAC-001"
        assert item["codigo_tipo_procedimiento"] == "12"
        assert item["total_pares"] == 2
        assert len(item["pares_duplicados"]) == 2

    # ── Escenario 2: Grupo 09 con duplicidad total → flag ──

    def test_grupo_09_duplicidad_total_retorna_flag(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Grupo 09 con 3 pares distintos, todos x2 → 1 flag."""
        ws = workbook_with_headers.active
        pares = [("890101", 1), ("890102", 2), ("890103", 3)]
        for i, (codigo, cantidad) in enumerate(pares):
            base_row = 2 + i * 2
            for offset in range(2):
                _write_farmacia_row(ws, base_row + offset, "FAC-001", codigo, cantidad, "09")

        result = detect_duplicados_farmacia(ws, _INDICES)
        assert len(result) == 1
        assert result[0]["codigo_tipo_procedimiento"] == "09"
        assert result[0]["total_pares"] == 3
        assert len(result[0]["pares_duplicados"]) == 3

    # ── Escenario 3: Grupo con mezcla (duplicados y únicos) → NO flag ──

    def test_grupo_con_mezcla_no_flag(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Par A x2, par B x1 → NO flag (no todos duplicados)."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "890101", 1, "09")
        _write_farmacia_row(ws, 3, "FAC-001", "890101", 1, "09")
        _write_farmacia_row(ws, 4, "FAC-001", "890102", 2, "09")

        result = detect_duplicados_farmacia(ws, _INDICES)
        assert result == []

    # ── Escenario 4: Múltiples grupos independientes (09 flag, 12 no) ──

    def test_multiples_grupos_independientes_solo_09_flag(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Grupo 09 todo duplicado, grupo 12 con mezcla → solo 09 flag."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "890101", 1, "09")
        _write_farmacia_row(ws, 3, "FAC-001", "890101", 1, "09")
        _write_farmacia_row(ws, 4, "FAC-001", "890103", 3, "12")
        _write_farmacia_row(ws, 5, "FAC-001", "890103", 3, "12")
        _write_farmacia_row(ws, 6, "FAC-001", "890104", 4, "12")

        result = detect_duplicados_farmacia(ws, _INDICES)
        assert len(result) == 1
        assert result[0]["codigo_tipo_procedimiento"] == "09"

    # ── Escenario 5: Sin filas de farmacia → [] ──

    def test_sin_filas_farmacia_retorna_vacio(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Filas con tarifario no farmacia → []."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="890101")
        ws.cell(row=2, column=3, value=1)
        ws.cell(row=2, column=4, value="Honorarios")
        ws.cell(row=2, column=5, value="Consulta")
        ws.cell(row=2, column=6, value="09")
        ws.cell(row=2, column=7, value="Urgencias")

        result = detect_duplicados_farmacia(ws, _INDICES)
        assert result == []

    # ── Escenario 6: Columna codigo_tipo_procedimiento faltante → [] ──

    def test_columna_tipo_proc_faltante_retorna_vacio(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Sin índice codigo_tipo_procedimiento → []."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "890101", 1, "09")

        indices_sin_tipo_proc = {
            "numero_factura": 0,
            "codigo": 1,
            "cantidad": 2,
            "tarifario": 3,
            "procedimiento": 4,
            "tipo_factura_descripcion": 6,
        }
        result = detect_duplicados_farmacia(ws, indices_sin_tipo_proc)
        assert result == []

    # ── Escenario 7: Columna tarifario faltante → [] ──

    def test_columna_tarifario_faltante_retorna_vacio(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Sin índice tarifario → []."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "890101", 1, "09")

        indices_sin_tarifario = {
            "numero_factura": 0,
            "codigo": 1,
            "cantidad": 2,
            "procedimiento": 4,
            "codigo_tipo_procedimiento": 5,
            "tipo_factura_descripcion": 6,
        }
        result = detect_duplicados_farmacia(ws, indices_sin_tarifario)
        assert result == []

    # ── Escenario 8: Tipo procedimiento distinto de 09/12 → ignorado ──

    def test_tipo_proc_distinto_ignorado(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Tipo 02 con farmacia → ignorado (no 09/12)."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "890101", 1, "02")
        _write_farmacia_row(ws, 3, "FAC-001", "890101", 1, "02")

        result = detect_duplicados_farmacia(ws, _INDICES)
        assert result == []

    # ── Sin datos (solo headers) ──

    def test_sin_datos_retorna_vacio(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Sin filas de datos → retorna []."""
        ws = workbook_with_headers.active
        result = detect_duplicados_farmacia(ws, _INDICES)
        assert result == []

    # ── Grupo con 3 pares distintos todos duplicados ──

    def test_grupo_3_pares_todos_duplicados(
        self, workbook_with_headers: Workbook
    ) -> None:
        """3 pares distintos, cada uno x2 → total_pares=3, flag."""
        ws = workbook_with_headers.active
        pares = [("A", 1), ("B", 2), ("C", 3)]
        for i, (codigo, cantidad) in enumerate(pares):
            base = 2 + i * 2
            for offset in range(2):
                _write_farmacia_row(ws, base + offset, "FAC-001", codigo, cantidad, "12")

        result = detect_duplicados_farmacia(ws, _INDICES)
        assert len(result) == 1
        assert result[0]["total_pares"] == 3
        assert result[0]["codigo_tipo_procedimiento"] == "12"

    # ── Cantidad None tratado como 0 ──

    def test_cantidad_none_tratado_como_cero(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Cantidad None debe tratarse como 0."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "890101", None, "09")
        _write_farmacia_row(ws, 3, "FAC-001", "890101", None, "09")

        result = detect_duplicados_farmacia(ws, _INDICES)
        assert len(result) == 1
        assert result[0]["total_pares"] == 1
        assert result[0]["pares_duplicados"][0]["cantidad"] == 0

    # ── NEW: tipo_factura filter ──

    def test_filtra_solo_urgencias_duplicados(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Urgencias rows → detected; Hospitalización rows with same data → skipped."""
        ws = workbook_with_headers.active
        # Urgencias: duplicated pair → should be detected
        _write_farmacia_row(ws, 2, "FAC-001", "890101", 1, "12", "Urgencias")
        _write_farmacia_row(ws, 3, "FAC-001", "890101", 1, "12", "Urgencias")
        # Hospitalización: same duplicated pair → should be SKIPPED
        _write_farmacia_row(ws, 4, "FAC-002", "890101", 1, "12", "Hospitalización")
        _write_farmacia_row(ws, 5, "FAC-002", "890101", 1, "12", "Hospitalización")

        result = detect_duplicados_farmacia(ws, _INDICES)
        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"

    def test_sin_urgencias_duplicados_retorna_vacio(
        self, workbook_with_headers: Workbook
    ) -> None:
        """All rows are Hospitalización → empty result."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-003", "890101", 1, "09", "Hospitalización")
        _write_farmacia_row(ws, 3, "FAC-003", "890101", 1, "09", "Hospitalización")

        result = detect_duplicados_farmacia(ws, _INDICES)
        assert result == []

    def test_missing_tipo_factura_col_duplicados(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Missing tipo_factura_descripcion column → return []."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-004", "890101", 1, "12")

        indices_no_tipo = {
            "numero_factura": 0,
            "codigo": 1,
            "cantidad": 2,
            "tarifario": 3,
            "procedimiento": 4,
            "codigo_tipo_procedimiento": 5,
        }
        result = detect_duplicados_farmacia(ws, indices_no_tipo)
        assert result == []
