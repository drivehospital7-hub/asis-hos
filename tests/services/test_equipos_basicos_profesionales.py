"""Tests para app/services/equipos_basicos/profesionales.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.constants import PROFESIONALES_EQUIPOS_BASICOS


@pytest.fixture
def workbook_with_eb_headers() -> Workbook:
    """Crea un workbook con headers para equipos básicos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código Profesional")
    ws.cell(row=1, column=3, value="Cód. Equivalente CUPS")
    return wb


class TestDetectProfesionalesEquiposBasicos:
    """Tests para detect_profesionales_equipos_basicos."""

    def test_indices_none_retorna_vacio(
        self, workbook_with_eb_headers: Workbook
    ) -> None:
        """Si numero_factura es None, retorna lista vacía."""
        from app.services.equipos_basicos.profesionales import (
            detect_profesionales_equipos_basicos,
        )

        ws = workbook_with_eb_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="03764")
        ws.cell(row=2, column=3, value="997002")

        indices = {"numero_factura": None, "codigo_profesional": 1, "codigo": 2}
        result = detect_profesionales_equipos_basicos(ws, indices)
        assert result == []

    def test_profesional_no_en_listado_genera_error(
        self, workbook_with_eb_headers: Workbook
    ) -> None:
        """Un profesional NO en el listado debe generar error."""
        from app.services.equipos_basicos.profesionales import (
            detect_profesionales_equipos_basicos,
        )

        ws = workbook_with_eb_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="99999")  # No existe
        ws.cell(row=2, column=3, value="997002")

        indices = {"numero_factura": 0, "codigo_profesional": 1, "codigo": 2}
        result = detect_profesionales_equipos_basicos(ws, indices)

        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"
        assert result[0]["codigo_profesional"] == "99999"
        assert result[0]["problema"] == "Profesional no existe en el listado de Equipos Básicos"

    def test_odontologo_valido_no_genera_error(
        self, workbook_with_eb_headers: Workbook
    ) -> None:
        """ODONTOLOGO con código NO de HIGIENISTA no debe generar error."""
        from app.services.equipos_basicos.profesionales import (
            detect_profesionales_equipos_basicos,
        )

        ws = workbook_with_eb_headers.active
        # 03764 = JARAMILLO HERNANDEZ YAMILE LORENA (ODONTOLOGO)
        # 890101 no está en PYP_CODES_HIGIENISTA → OK para ODONTOLOGO
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="03764")
        ws.cell(row=2, column=3, value="890101")

        indices = {"numero_factura": 0, "codigo_profesional": 1, "codigo": 2}
        result = detect_profesionales_equipos_basicos(ws, indices)

        assert len(result) == 0

    def test_odontologo_con_codigo_higienista_genera_error(
        self, workbook_with_eb_headers: Workbook
    ) -> None:
        """ODONTOLOGO no puede usar códigos HIGIENISTA (PYP_CODES_HIGIENISTA)."""
        from app.services.equipos_basicos.profesionales import (
            detect_profesionales_equipos_basicos,
        )

        ws = workbook_with_eb_headers.active
        # 03764 = ODONTOLOGO
        # 997002 = Control de Placa Bacteriana (en PYP_CODES_HIGIENISTA)
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="03764")
        ws.cell(row=2, column=3, value="997002")

        indices = {"numero_factura": 0, "codigo_profesional": 1, "codigo": 2}
        result = detect_profesionales_equipos_basicos(ws, indices)

        assert len(result) == 1
        assert result[0]["tipo"] == "ODONTOLOGO"
        assert "no permitido" in result[0]["problema"]

    def test_higienista_valido_no_genera_error(
        self, workbook_with_eb_headers: Workbook
    ) -> None:
        """HIGIENISTA con código en PYP_CODES_HIGIENISTA no debe generar error."""
        from app.services.equipos_basicos.profesionales import (
            detect_profesionales_equipos_basicos,
        )

        ws = workbook_with_eb_headers.active
        # 03762 = CHAVES GONZALEZ NURY ADRIANA (HIGIENISTA)
        # 997002 = Control de Placa Bacteriana (en PYP_CODES_HIGIENISTA)
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="03762")
        ws.cell(row=2, column=3, value="997002")

        indices = {"numero_factura": 0, "codigo_profesional": 1, "codigo": 2}
        result = detect_profesionales_equipos_basicos(ws, indices)

        assert len(result) == 0

    def test_higienista_con_codigo_no_permitido_genera_error(
        self, workbook_with_eb_headers: Workbook
    ) -> None:
        """HIGIENISTA con código fuera de PYP_CODES_HIGIENISTA debe generar error."""
        from app.services.equipos_basicos.profesionales import (
            detect_profesionales_equipos_basicos,
        )

        ws = workbook_with_eb_headers.active
        # 03762 = CHAVES GONZALEZ NURY ADRIANA (HIGIENISTA)
        # 890101 no es PYP → error
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="03762")
        ws.cell(row=2, column=3, value="890101")

        indices = {"numero_factura": 0, "codigo_profesional": 1, "codigo": 2}
        result = detect_profesionales_equipos_basicos(ws, indices)

        assert len(result) == 1
        assert result[0]["tipo"] == "HIGIENISTA"
        assert "no permitido" in result[0]["problema"]

    def test_codigo_profesional_vacio_no_genera_error(
        self, workbook_with_eb_headers: Workbook
    ) -> None:
        """Si codigo_profesional está vacío, no se genera error."""
        from app.services.equipos_basicos.profesionales import (
            detect_profesionales_equipos_basicos,
        )

        ws = workbook_with_eb_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="")  # vacío
        ws.cell(row=2, column=3, value="997002")

        indices = {"numero_factura": 0, "codigo_profesional": 1, "codigo": 2}
        result = detect_profesionales_equipos_basicos(ws, indices)

        assert len(result) == 0

    def test_indices_codigo_none_retorna_vacio(
        self, workbook_with_eb_headers: Workbook
    ) -> None:
        """Si codigo (Cód. Equivalente CUPS) es None, retorna lista vacía."""
        from app.services.equipos_basicos.profesionales import (
            detect_profesionales_equipos_basicos,
        )

        ws = workbook_with_eb_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="03764")
        ws.cell(row=2, column=3, value="997002")

        indices = {"numero_factura": 0, "codigo_profesional": 1, "codigo": None}
        result = detect_profesionales_equipos_basicos(ws, indices)
        assert result == []
