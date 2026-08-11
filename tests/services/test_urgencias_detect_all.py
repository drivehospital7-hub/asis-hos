"""Tests para app/services/urgencias/detect_all.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

import app.services.urgencias.detect_all as urgencias_detect_all
from app.services.urgencias.detect_all import detect_all_problems_urgencias


@pytest.fixture
def workbook_minimal() -> Workbook:
    """Crea un workbook con headers mínimos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    return wb


class TestDetectAllProblemsUrgencias:
    """Tests para detect_all_problems_urgencias."""

    def _run(self, ws, indices):
        """Helper que corre el detector y retorna solo el dict resultado."""
        result, _ = detect_all_problems_urgencias(ws, indices)
        return result

    def test_retorna_dict_con_key_problemas(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe contener key 'problemas'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = {"numero_factura": 0}
        result = self._run(ws, indices)

        assert "problemas" in result
        assert isinstance(result["problemas"], dict)

    def test_retorna_dict_con_key_totales(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe contener key 'totales'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = {"numero_factura": 0}
        result = self._run(ws, indices)

        assert "totales" in result
        assert isinstance(result["totales"], dict)

    def test_retorna_dict_con_key_area(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe contener 'area' = 'urgencias'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = {"numero_factura": 0}
        result = self._run(ws, indices)

        assert result.get("area") == "urgencias"

    def test_resultado_incluye_normalizados(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe incluir 'normalizados' en problemas."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = {"numero_factura": 0}
        result = self._run(ws, indices)

        assert "normalizados" in result["problemas"]
        assert isinstance(result["problemas"]["normalizados"], list)

    def test_resultado_incluye_missing_columns(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado debe contener 'missing_columns'."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")

        indices = {"numero_factura": 0}
        result = self._run(ws, indices)

        assert "missing_columns" in result
        assert isinstance(result["missing_columns"], list)

    def test_normalizados_incluyen_fec_factura(
        self, workbook_minimal: Workbook
    ) -> None:
        """Resultado normalizados MUST include 'fec_factura' in every row."""
        ws = workbook_minimal.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="2024-01-15")
        ws.cell(row=1, column=2, value="Fec. Factura")

        indices = {"numero_factura": 0, "fec_factura": 1}
        result = self._run(ws, indices)
        norm = result["problemas"]["normalizados"]
        for row in norm:
            assert "fec_factura" in row

    def test_engine_result_without_accion_uses_problema(self, workbook_minimal, monkeypatch):
        """Engine-style results without accion must remain in Urgencias output."""
        class FakeSessionManager:
            def __init__(self, domain):
                self.domain = domain

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc_value, traceback):
                return False

        class FakeEvidenceCollector:
            def __init__(self, domain):
                self.domain = domain

        class FakeRuleBasedDetector:
            def __init__(self, rule_name, session):
                self.rule_name = rule_name

            def detect(self, *args, **kwargs):
                if self.rule_name == "cups_equivalentes":
                    return [{
                        "factura": "FEV437512",
                        "codigo": "890201",
                        "codigo_equiv": "890201",
                        "problema": "Usar codigo equivalente 890201",
                    }]
                return []

        monkeypatch.setattr(urgencias_detect_all, "is_rule_engine_enabled", lambda: True)
        monkeypatch.setattr(urgencias_detect_all, "_PERSIST", False)
        monkeypatch.setattr(
            "app.services.engine.session_manager.SessionManager", FakeSessionManager
        )
        monkeypatch.setattr(
            "app.services.engine.evidence_collector.EvidenceCollector",
            FakeEvidenceCollector,
        )
        monkeypatch.setattr(
            "app.services.engine.rule_based_detector.RuleBasedDetector",
            FakeRuleBasedDetector,
        )

        result, _ = detect_all_problems_urgencias(
            workbook_minimal.active, {"numero_factura": 0}
        )

        assert result["problemas"]["cups_equivalentes"] == [{
            "factura": "FEV437512",
            "codigo": "890201",
            "codigo_equiv": "890201",
            "accion": "Usar codigo equivalente 890201",
            "responsable": "",
        }]
