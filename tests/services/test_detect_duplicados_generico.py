"""Tests for app/services/transversales/detect_duplicados_base.py.

Cubre ambos modos de output: con y sin codigos_tipo_proc.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.constants.urgencias import VALOR_TARIFARIO_FARMACIA, CODIGOS_TIPO_PROC_09_12
from app.services.transversales.detect_duplicados_base import detect_duplicados_generico


@pytest.fixture
def workbook_with_headers() -> Workbook:
    """Crea un workbook con headers mínimos para ambos modos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.cell(row=1, column=1, value="Número Factura")
    ws.cell(row=1, column=2, value="Código")
    ws.cell(row=1, column=3, value="Cantidad")
    ws.cell(row=1, column=4, value="Tarifario")
    ws.cell(row=1, column=5, value="Procedimiento")
    ws.cell(row=1, column=6, value="Código Tipo Procedimiento")
    ws.cell(row=1, column=7, value="Tipo Factura Descripción")
    return wb


_INDICES_FULL = {
    "numero_factura": 0,
    "codigo": 1,
    "cantidad": 2,
    "tarifario": 3,
    "procedimiento": 4,
    "codigo_tipo_procedimiento": 5,
    "tipo_factura_descripcion": 6,
}

_INDICES_BASIC = {
    "numero_factura": 0,
    "codigo": 1,
    "cantidad": 2,
    "tipo_factura_descripcion": 6,
}


def _write_row(ws, row, factura, codigo, cantidad, tipo_proc="12", tarifario=VALOR_TARIFARIO_FARMACIA, tipo_factura="Urgencias"):
    """Helper: escribe una fila completa."""
    ws.cell(row=row, column=1, value=factura)
    ws.cell(row=row, column=2, value=codigo)
    ws.cell(row=row, column=3, value=cantidad)
    ws.cell(row=row, column=4, value=tarifario)
    ws.cell(row=row, column=5, value="Med")
    ws.cell(row=row, column=6, value=tipo_proc)
    ws.cell(row=row, column=7, value=tipo_factura)


def _write_farmacia_row(ws, row, factura, codigo, cantidad, tipo_proc="12", tipo_factura="Farmacia"):
    """Helper for Farmacia rows (no tarifario filtering needed)."""
    ws.cell(row=row, column=1, value=factura)
    ws.cell(row=row, column=2, value=codigo)
    ws.cell(row=row, column=3, value=cantidad)
    ws.cell(row=row, column=4, value="")
    ws.cell(row=row, column=5, value="Med")
    ws.cell(row=row, column=6, value=tipo_proc)
    ws.cell(row=row, column=7, value=tipo_factura)


class TestDetectDuplicadosGenericoUrgencias:
    """Modo con codigos_tipo_proc — mismo comportamiento que el detector de Urgencias."""

    def test_grupo_duplicidad_total_retorna_flag(self, workbook_with_headers: Workbook):
        """2 pares, cada uno x2 → 1 flag con codigo_tipo_procedimiento."""
        ws = workbook_with_headers.active
        _write_row(ws, 2, "FAC-001", "890101", 1, "12")
        _write_row(ws, 3, "FAC-001", "890101", 1, "12")
        _write_row(ws, 4, "FAC-001", "890102", 2, "12")
        _write_row(ws, 5, "FAC-001", "890102", 2, "12")

        result = detect_duplicados_generico(
            ws, _INDICES_FULL,
            tipo_factura="Urgencias",
            tarifario_val=VALOR_TARIFARIO_FARMACIA,
            codigos_tipo_proc=CODIGOS_TIPO_PROC_09_12,
        )
        assert len(result) == 1
        item = result[0]
        assert item["factura"] == "FAC-001"
        assert item["codigo_tipo_procedimiento"] == "12"
        assert item["total_pares"] == 2
        assert len(item["pares_duplicados"]) == 2

    def test_grupo_con_mezcla_no_flag(self, workbook_with_headers: Workbook):
        """Par A x2, par B x1 → NO flag."""
        ws = workbook_with_headers.active
        _write_row(ws, 2, "FAC-001", "890101", 1, "09")
        _write_row(ws, 3, "FAC-001", "890101", 1, "09")
        _write_row(ws, 4, "FAC-001", "890102", 2, "09")

        result = detect_duplicados_generico(
            ws, _INDICES_FULL,
            tipo_factura="Urgencias",
            tarifario_val=VALOR_TARIFARIO_FARMACIA,
            codigos_tipo_proc=CODIGOS_TIPO_PROC_09_12,
        )
        assert result == []

    def test_sin_filas_urgencias_retorna_vacio(self, workbook_with_headers: Workbook):
        """Filas con tipo_factura distinto a Urgencias → []."""
        ws = workbook_with_headers.active
        _write_row(ws, 2, "FAC-001", "890101", 1, "12", tipo_factura="Hospitalización")
        _write_row(ws, 3, "FAC-001", "890101", 1, "12", tipo_factura="Hospitalización")

        result = detect_duplicados_generico(
            ws, _INDICES_FULL,
            tipo_factura="Urgencias",
            tarifario_val=VALOR_TARIFARIO_FARMACIA,
            codigos_tipo_proc=CODIGOS_TIPO_PROC_09_12,
        )
        assert result == []

    def test_output_incluye_codigo_tipo_procedimiento(self, workbook_with_headers: Workbook):
        """Output dict incluye key 'codigo_tipo_procedimiento'."""
        ws = workbook_with_headers.active
        _write_row(ws, 2, "FAC-001", "890101", 1, "09")
        _write_row(ws, 3, "FAC-001", "890101", 1, "09")

        result = detect_duplicados_generico(
            ws, _INDICES_FULL,
            tipo_factura="Urgencias",
            tarifario_val=VALOR_TARIFARIO_FARMACIA,
            codigos_tipo_proc=CODIGOS_TIPO_PROC_09_12,
        )
        assert "codigo_tipo_procedimiento" in result[0]

    def test_missing_required_cols_retorna_vacio(self, workbook_with_headers: Workbook):
        """Falta tarifario → retorna []."""
        ws = workbook_with_headers.active
        _write_row(ws, 2, "FAC-001", "890101", 1, "12")

        indices_no_tarifario = dict(_INDICES_FULL)
        del indices_no_tarifario["tarifario"]
        result = detect_duplicados_generico(
            ws, indices_no_tarifario,
            tipo_factura="Urgencias",
            tarifario_val=VALOR_TARIFARIO_FARMACIA,
            codigos_tipo_proc=CODIGOS_TIPO_PROC_09_12,
        )
        assert result == []

    def test_cantidad_none_tratado_como_cero(self, workbook_with_headers: Workbook):
        """Cantidad None debe tratarse como 0."""
        ws = workbook_with_headers.active
        _write_row(ws, 2, "FAC-001", "890101", None, "09")
        _write_row(ws, 3, "FAC-001", "890101", None, "09")

        result = detect_duplicados_generico(
            ws, _INDICES_FULL,
            tipo_factura="Urgencias",
            tarifario_val=VALOR_TARIFARIO_FARMACIA,
            codigos_tipo_proc=CODIGOS_TIPO_PROC_09_12,
        )
        assert len(result) == 1
        assert result[0]["pares_duplicados"][0]["cantidad"] == 0


class TestDetectDuplicadosGenericoFarmacia:
    """Modo sin codigos_tipo_proc — para tipo factura Farmacia."""

    def test_duplicidad_total_retorna_flag(self, workbook_with_headers: Workbook):
        """2 pares, cada uno x2 → 1 flag sin codigo_tipo_procedimiento."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 4, "FAC-001", "A002", 2)
        _write_farmacia_row(ws, 5, "FAC-001", "A002", 2)

        result = detect_duplicados_generico(
            ws, _INDICES_BASIC,
            tipo_factura="Farmacia",
        )
        assert len(result) == 1
        item = result[0]
        assert item["factura"] == "FAC-001"
        assert item["total_pares"] == 2
        assert len(item["pares_duplicados"]) == 2

    def test_output_no_incluye_codigo_tipo_procedimiento(self, workbook_with_headers: Workbook):
        """Output NO incluye key 'codigo_tipo_procedimiento'."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1)

        result = detect_duplicados_generico(
            ws, _INDICES_BASIC,
            tipo_factura="Farmacia",
        )
        assert "codigo_tipo_procedimiento" not in result[0]

    def test_grupo_con_mezcla_no_flag(self, workbook_with_headers: Workbook):
        """Par A x2, par B x1 → NO flag."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 4, "FAC-001", "A002", 2)

        result = detect_duplicados_generico(
            ws, _INDICES_BASIC,
            tipo_factura="Farmacia",
        )
        assert result == []

    def test_sin_filas_farmacia_retorna_vacio(self, workbook_with_headers: Workbook):
        """Ninguna fila tipo Farmacia → []."""
        ws = workbook_with_headers.active
        ws.cell(row=2, column=1, value="FAC-001")
        ws.cell(row=2, column=2, value="A001")
        ws.cell(row=2, column=3, value=1)
        ws.cell(row=2, column=7, value="Urgencias")

        result = detect_duplicados_generico(
            ws, _INDICES_BASIC,
            tipo_factura="Farmacia",
        )
        assert result == []

    def test_multiples_facturas_independientes(self, workbook_with_headers: Workbook):
        """Dos facturas, solo una con duplicidad total → 1 flag."""
        ws = workbook_with_headers.active
        # FAC-001: duplicidad total
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1)
        # FAC-002: mezcla
        _write_farmacia_row(ws, 4, "FAC-002", "B001", 1)
        _write_farmacia_row(ws, 5, "FAC-002", "B001", 1)
        _write_farmacia_row(ws, 6, "FAC-002", "B002", 2)

        result = detect_duplicados_generico(
            ws, _INDICES_BASIC,
            tipo_factura="Farmacia",
        )
        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"

    def test_missing_numero_factura_retorna_vacio(self, workbook_with_headers: Workbook):
        """Falta columna numero_factura → []."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1)

        indices_no_fact = dict(_INDICES_BASIC)
        del indices_no_fact["numero_factura"]
        result = detect_duplicados_generico(
            ws, indices_no_fact,
            tipo_factura="Farmacia",
        )
        assert result == []

    def test_missing_tipo_factura_col_retorna_vacio(self, workbook_with_headers: Workbook):
        """Falta columna tipo_factura_descripcion → []."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1)

        indices_no_tipo = dict(_INDICES_BASIC)
        del indices_no_tipo["tipo_factura_descripcion"]
        result = detect_duplicados_generico(
            ws, indices_no_tipo,
            tipo_factura="Farmacia",
        )
        assert result == []

    def test_output_no_agrupa_por_tipo_proc(self, workbook_with_headers: Workbook):
        """Misma factura, distintos tipo_proc → agrupado como una sola."""
        ws = workbook_with_headers.active
        # Misma factura, mismo par, distintos tipo_proc
        _write_farmacia_row(ws, 2, "FAC-001", "A001", 1, "09")
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1, "09")
        _write_farmacia_row(ws, 4, "FAC-001", "A001", 1, "12")
        _write_farmacia_row(ws, 5, "FAC-001", "A001", 1, "12")

        result = detect_duplicados_generico(
            ws, _INDICES_BASIC,
            tipo_factura="Farmacia",
        )
        # 2 pares distintos (A001,1 con tipo_proc=09 vs A001,1 con tipo_proc=12 son iguales para Farmacia)
        # pero en los datos, son 2 pares (A001,1) con count=4, más (A002,...) - no hay más
        # Actually: todas tienen codigo=A001, cantidad=1 → es 1 par con count=4
        # total_pares=1, todos duplicados → 1 flag
        assert len(result) == 1
        assert result[0]["total_pares"] == 1
        assert result[0]["pares_duplicados"][0]["count"] == 4

    def test_factura_filtra_correctamente(self, workbook_with_headers: Workbook):
        """Factura inválida (None) se salta."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, None, "A001", 1)
        _write_farmacia_row(ws, 3, "FAC-001", "A001", 1)
        _write_farmacia_row(ws, 4, "FAC-001", "A001", 1)

        result = detect_duplicados_generico(
            ws, _INDICES_BASIC,
            tipo_factura="Farmacia",
        )
        assert len(result) == 1
        assert result[0]["factura"] == "FAC-001"

    def test_cantidad_str_se_convierte(self, workbook_with_headers: Workbook):
        """Cantidad como string se convierte a int."""
        ws = workbook_with_headers.active
        _write_farmacia_row(ws, 2, "FAC-001", "A001", "1")
        _write_farmacia_row(ws, 3, "FAC-001", "A001", "1")

        result = detect_duplicados_generico(
            ws, _INDICES_BASIC,
            tipo_factura="Farmacia",
        )
        assert len(result) == 1
        assert result[0]["pares_duplicados"][0]["cantidad"] == 1
