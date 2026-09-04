"""Tests para el export a Excel de Control de Errores (servicio + ruta)."""

import contextlib
import shutil
from contextlib import ExitStack
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

from app import create_app
from app.services.control_errores_export import (
    HEADERS,
    build_errores_export_workbook,
    filename_export,
)

_APP = create_app()
_APP.config.update({"TESTING": True, "SECRET_KEY": "test-secret-key"})

_SERVIDOR_PRUEBA = "http://servidor-test:5001"


def _error_fixture(error_id="err-1", factura="FAC-001", creado_en="2026-05-15T10:30:00",
                   estado="S", validador="MARIA GOMEZ", refactura="", tipo_error="Otros"):
    """Un error mínimo con los campos que el export lee."""
    return {
        "id": error_id,
        "factura": factura,
        "refactura": refactura,
        "creado_en": creado_en,
        "tipo_error": tipo_error,
        "observacion": "Descripción con acentos",
        "responsable": "JUAN PEREZ",
        "observacion_facturador": "Revisar",
        "estado": estado,
        "validador": validador,
    }


def _adjuntos_por_scope(obs, fac):
    """Side effect para listar_imagenes: scope "" → obs, scope "facturador" → fac."""

    def _impl(error_id, scope=""):
        return obs if scope == "" else fac

    return _impl


class TestFilenameExport:
    def test_mes_valido_deriva_nombre(self):
        assert filename_export("2026-05") == "control-errores-May-2026.xlsx"

    def test_mes_invalido_usa_fecha_actual(self):
        name = filename_export(None)
        assert name.startswith("control-errores-")
        assert name.endswith(".xlsx")


class TestBuildWorkbook:
    def test_headers_y_fila_con_hipervinculos(self):
        with patch(
            "app.services.control_errores_export.listar_imagenes",
            side_effect=_adjuntos_por_scope(
                ["file_1.pdf", "image_2.jpg"], ["f_1.pdf"]
            ),
        ):
            with _APP.app_context():
                buffer = build_errores_export_workbook([_error_fixture()], "http://testserver/")

        wb = load_workbook(BytesIO(buffer.read()))
        ws = wb.active

        assert [c.value for c in ws[1]] == HEADERS
        assert ws["A1"].font.bold is True
        assert ws["L1"].font.bold is True

        row = [c.value for c in ws[2]]
        assert row[0] == "Maria Gomez"       # Validador (Title Case)
        assert row[1] == "15/05/2026"        # Creado dd/mm/yyyy
        assert row[2] == "FAC-001"           # Factura
        assert row[3] in (None, "")         # ReFactura (opcional, vacío)
        assert row[4] == "Otros"             # Categoría
        assert row[6] == "Juan Perez"        # Responsables (Title Case)
        assert row[7] == "Pendiente"         # Estado
        assert row[8] == "Abrir PDF"         # Adjunto 1
        assert row[9] == "Abrir imagen"      # Adjunto 2
        assert row[10] is None               # Adjunto 3 vacío
        assert row[11] == "Revisar"          # Observación del Facturador
        assert row[12] == "Abrir PDF"        # Adjunto 4 (facturador)
        assert row[13] is None               # Adjunto 5 vacío
        assert row[14] is None               # Adjunto 6 vacío

        cell9 = ws.cell(row=2, column=9)
        assert cell9.hyperlink.target == (
            "http://testserver/api/control-errores/err-1/imagenes/file_1.pdf"
        )
        assert "?token=" not in cell9.hyperlink.target
        assert cell9.style != "Hyperlink"
        assert cell9.font.color.rgb == "000563C1"
        assert cell9.font.underline == "single"

        cell10 = ws.cell(row=2, column=10)
        assert cell10.hyperlink.target == (
            "http://testserver/api/control-errores/err-1/imagenes/image_2.jpg"
        )
        assert "?token=" not in cell10.hyperlink.target

        assert ws.cell(row=2, column=11).hyperlink is None

        cell13 = ws.cell(row=2, column=13)
        assert cell13.hyperlink.target == (
            "http://testserver/api/control-errores/err-1/imagenes/f_1.pdf?scope=facturador"
        )
        assert "?token=" not in cell13.hyperlink.target
        assert cell13.style != "Hyperlink"
        assert cell13.font.color.rgb == "000563C1"
        assert cell13.font.underline == "single"

        assert ws.cell(row=2, column=14).hyperlink is None
        assert ws.cell(row=2, column=15).hyperlink is None

    def test_estado_resuelto_y_fecha_bruta(self):
        with patch(
            "app.services.control_errores_export.listar_imagenes",
            side_effect=_adjuntos_por_scope([], []),
        ):
            with _APP.app_context():
                buffer = build_errores_export_workbook(
                    [_error_fixture(estado="N", creado_en="fecha-rara")],
                    "http://testserver/",
                )

        wb = load_workbook(BytesIO(buffer.read()))
        ws = wb.active
        row = [c.value for c in ws[2]]
        assert row[7] == "Resuelto"
        assert row[1] == "fecha-rara"

    def test_hipervinculo_escapada_nombre_con_espacios(self):
        with patch(
            "app.services.control_errores_export.listar_imagenes",
            side_effect=_adjuntos_por_scope(["mi archivo.pdf"], []),
        ):
            with _APP.app_context():
                buffer = build_errores_export_workbook([_error_fixture()], "http://testserver/")

        wb = load_workbook(BytesIO(buffer.read()))
        ws = wb.active
        cell = ws.cell(row=2, column=9)
        assert cell.hyperlink.target == (
            "http://testserver/api/control-errores/err-1/imagenes/mi%20archivo.pdf"
        )
        assert "?token=" not in cell.hyperlink.target

    def test_hipervinculo_sin_token(self):
        with patch(
            "app.services.control_errores_export.listar_imagenes",
            side_effect=_adjuntos_por_scope(["file_1.pdf"], []),
        ):
            with _APP.app_context():
                buffer = build_errores_export_workbook([_error_fixture()], "http://testserver/")

        wb = load_workbook(BytesIO(buffer.read()))
        ws = wb.active
        target = ws.cell(row=2, column=9).hyperlink.target
        assert "?token=" not in target
        assert target == "http://testserver/api/control-errores/err-1/imagenes/file_1.pdf"

    def test_estilos_tabla(self):
        with patch(
            "app.services.control_errores_export.listar_imagenes",
            side_effect=_adjuntos_por_scope([], []),
        ):
            with _APP.app_context():
                buffer = build_errores_export_workbook(
                    [_error_fixture(error_id="e1"), _error_fixture(error_id="e2")],
                    "http://testserver/",
                )

        wb = load_workbook(BytesIO(buffer.read()))
        ws = wb.active

        assert ws["A1"].fill.fgColor.rgb == "001B5E20"
        assert ws["A1"].font.color.rgb == "00FFFFFF"
        assert ws["A1"].font.bold is True

        assert ws["A2"].fill.fgColor.rgb == "00E8F5E9"
        assert ws["A3"].fill.fgColor.rgb == "00FFFFFF"
        assert ws["A2"].fill.fgColor.rgb != ws["A3"].fill.fgColor.rgb

        for cell in (ws["A2"], ws["L2"], ws["O2"], ws["A3"], ws["A1"]):
            assert cell.font.color.rgb == "00000000" or cell.font.bold is True
            for side in ("left", "right", "top", "bottom"):
                assert getattr(cell.border, side).style == "thin"

        assert ws.column_dimensions["A"].width == 20
        assert ws.column_dimensions["A"].width == ws.column_dimensions["L"].width
        assert ws.column_dimensions["A"].width == ws.column_dimensions["O"].width

    def test_hipervinculo_estilo_fuente_azul_subrayado(self):
        with patch(
            "app.services.control_errores_export.listar_imagenes",
            side_effect=_adjuntos_por_scope([], ["file_1.pdf"]),
        ):
            with _APP.app_context():
                buffer = build_errores_export_workbook([_error_fixture()], "http://testserver/")

        wb = load_workbook(BytesIO(buffer.read()))
        ws = wb.active
        cell = ws.cell(row=2, column=13)
        assert cell.hyperlink is not None
        assert cell.hyperlink.target.endswith("file_1.pdf?scope=facturador")
        assert cell.font.color.rgb == "000563C1"
        assert cell.font.underline == "single"
        assert cell.style != "Hyperlink"
        assert cell.fill.fgColor.rgb == "00E8F5E9"
        assert cell.border.left.style == "thin"

    def test_facturador_adjuntos_en_columnas_13_15_con_scope(self):
        with patch(
            "app.services.control_errores_export.listar_imagenes",
            side_effect=_adjuntos_por_scope(
                ["obs_1.png"], ["fac_1.pdf", "fac_2.xlsx", "fac_3.jpg"]
            ),
        ):
            with _APP.app_context():
                buffer = build_errores_export_workbook(
                    [_error_fixture()], "http://testserver/"
                )

        wb = load_workbook(BytesIO(buffer.read()))
        ws = wb.active
        row = [c.value for c in ws[2]]

        assert row[12] == "Abrir PDF"        # Adjunto 4 (fac_1.pdf)
        assert row[13] == "Abrir Excel"      # Adjunto 5 (fac_2.xlsx)
        assert row[14] == "Abrir imagen"     # Adjunto 6 (fac_3.jpg)

        for col, name in ((13, "fac_1.pdf"), (14, "fac_2.xlsx"), (15, "fac_3.jpg")):
            cell = ws.cell(row=2, column=col)
            assert cell.hyperlink is not None
            assert cell.hyperlink.target == (
                f"http://testserver/api/control-errores/err-1/imagenes/{name}?scope=facturador"
            )
            assert "?token=" not in cell.hyperlink.target
            assert cell.font.color.rgb == "000563C1"
            assert cell.font.underline == "single"

        obs_cell = ws.cell(row=2, column=9)
        assert obs_cell.hyperlink.target == (
            "http://testserver/api/control-errores/err-1/imagenes/obs_1.png"
        )
        assert "?scope=" not in obs_cell.hyperlink.target

    def test_facturador_sin_adjuntos_deja_vacias_13_15(self):
        with patch(
            "app.services.control_errores_export.listar_imagenes",
            side_effect=_adjuntos_por_scope([], []),
        ):
            with _APP.app_context():
                buffer = build_errores_export_workbook(
                    [_error_fixture()], "http://testserver/"
                )

        wb = load_workbook(BytesIO(buffer.read()))
        ws = wb.active
        assert ws.cell(row=2, column=13).value in (None, "")
        assert ws.cell(row=2, column=14).value in (None, "")
        assert ws.cell(row=2, column=15).value in (None, "")
        assert ws.cell(row=2, column=13).hyperlink is None

    def test_headers_incluye_refactura_despues_de_factura(self):
        """R13: el header ReFactura va inmediatamente después de Factura."""
        idx_factura = HEADERS.index("Factura")
        assert HEADERS[idx_factura + 1] == "ReFactura"

    def test_fila_exporta_refactura_en_columna_4(self):
        """R12: el valor refactura se exporta en la columna D (col 4)."""
        with patch(
            "app.services.control_errores_export.listar_imagenes",
            side_effect=_adjuntos_por_scope([], []),
        ):
            with _APP.app_context():
                buffer = build_errores_export_workbook(
                    [_error_fixture(refactura="R-42")], "http://testserver/"
                )

        wb = load_workbook(BytesIO(buffer.read()))
        ws = wb.active
        assert ws["D2"].value == "R-42"
        assert ws["C2"].value == "FAC-001"   # Factura sigue en col 3
        assert ws["E2"].value == "Otros"     # Categoría desplazada a col 5

    def test_fila_legacy_sin_refactura_exporta_vacio(self):
        """Legacy: registro sin key refactura → celda vacía (get con default)."""
        error = _error_fixture()
        del error["refactura"]
        with patch(
            "app.services.control_errores_export.listar_imagenes",
            side_effect=_adjuntos_por_scope([], []),
        ):
            with _APP.app_context():
                buffer = build_errores_export_workbook([error], "http://testserver/")

        wb = load_workbook(BytesIO(buffer.read()))
        ws = wb.active
        assert ws["D2"].value in (None, "")


def _login(app_client):
    with app_client.session_transaction() as sess:
        sess["ce_authenticated"] = True
        sess["rol"] = "validador"
        sess["username"] = "val1"
        sess["permisos"] = ["control_urgencias", "control_urgencias:write"]


def _patch_export_pipeline(fixture, stack: ExitStack, adjuntos=None, adjuntos_facturador=None):
    """Parchea el pipeline completo que la ruta usa (storage + adjuntos)."""
    stack.enter_context(patch("app.utils.errores_storage._leer_datos",
                              return_value={"errores": fixture}))
    stack.enter_context(patch("app.utils.errores_storage.obtener_imagenes_count",
                              return_value=0))
    stack.enter_context(patch(
        "app.services.control_errores_export.listar_imagenes",
        side_effect=_adjuntos_por_scope(
            adjuntos if adjuntos is not None else [],
            adjuntos_facturador if adjuntos_facturador is not None else [],
        ),
    ))


def _client(**config_overrides):
    """Test client con config controlable (por defecto la misma clave de prueba)."""
    app = create_app()
    app.config.update({"TESTING": True, "SECRET_KEY": "test-secret-key"})
    app.config.update(config_overrides)
    return app.test_client()


def _hipervinculo_de_respuesta(resp, cell="I2"):
    wb = load_workbook(BytesIO(resp.data))
    ws = wb.active
    return ws[cell].hyperlink.target


class TestExportRoute:
    def test_requiere_autenticacion(self, app_client):
        resp = app_client.get(
            "/api/control-errores/export",
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        assert resp.status_code == 401
        data = resp.get_json()
        assert data["status"] == "error"
        assert data["data"] == {}
        assert len(data["errors"]) > 0

    def test_devuelve_xlsx_valido(self, app_client):
        _login(app_client)
        fixture = [_error_fixture()]
        with ExitStack() as stack:
            _patch_export_pipeline(fixture, stack, adjuntos=["file_1.pdf"])
            resp = app_client.get("/api/control-errores/export?mes=2026-05")

        assert resp.status_code == 200
        assert resp.content_type.startswith("application/vnd.openxmlformats")
        assert "control-errores-May-2026.xlsx" in resp.headers["Content-Disposition"]

        wb = load_workbook(BytesIO(resp.data))
        ws = wb.active
        assert [c.value for c in ws[1]] == HEADERS
        assert ws["A2"].value == "Maria Gomez"
        assert ws["B2"].value == "15/05/2026"
        assert ws["C2"].value == "FAC-001"
        target = ws["I2"].hyperlink.target
        # Sin EXPORT_BASE_URL → usa el host del request (comportamiento previo)
        assert target == "http://localhost/api/control-errores/err-1/imagenes/file_1.pdf"
        assert "?token=" not in target

    def test_filtro_por_mes(self, app_client):
        _login(app_client)
        fixture = [
            _error_fixture(error_id="may", factura="FAC-MAY", creado_en="2026-05-15T10:00:00"),
            _error_fixture(error_id="jun", factura="FAC-JUN", creado_en="2026-06-02T10:00:00"),
        ]
        with ExitStack() as stack:
            _patch_export_pipeline(fixture, stack)
            resp = app_client.get("/api/control-errores/export?mes=2026-05")

        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.data))
        ws = wb.active
        facturas = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
        assert facturas == ["FAC-MAY"]

    def test_sin_datos_devuelve_error_400(self, app_client):
        _login(app_client)
        fixture = [_error_fixture(creado_en="2026-06-02T10:00:00")]
        with ExitStack() as stack:
            _patch_export_pipeline(fixture, stack)
            resp = app_client.get("/api/control-errores/export?mes=2026-05")

        assert resp.status_code == 400
        data = resp.get_json()
        assert data["status"] == "error"
        assert len(data["errors"]) > 0

    def test_export_usa_base_url_configurada(self):
        client = _client(EXPORT_BASE_URL="http://servidor-test:5001/")
        _login(client)
        fixture = [_error_fixture()]
        with ExitStack() as stack:
            _patch_export_pipeline(fixture, stack, adjuntos=["file_1.pdf"])
            resp = client.get("/api/control-errores/export")

        assert resp.status_code == 200
        target = _hipervinculo_de_respuesta(resp)
        assert target == (
            f"{_SERVIDOR_PRUEBA}/api/control-errores/err-1/imagenes/file_1.pdf"
        )
        assert "?token=" not in target
        assert "localhost" not in target

    def test_export_base_url_sin_barra_final(self):
        client = _client(EXPORT_BASE_URL="http://servidor-test:5001")
        _login(client)
        fixture = [_error_fixture()]
        with ExitStack() as stack:
            _patch_export_pipeline(fixture, stack, adjuntos=["file_1.pdf"])
            resp = client.get("/api/control-errores/export")

        assert resp.status_code == 200
        target = _hipervinculo_de_respuesta(resp)
        assert target == (
            f"{_SERVIDOR_PRUEBA}/api/control-errores/err-1/imagenes/file_1.pdf"
        )
        assert "?token=" not in target


@contextlib.contextmanager
def _archivo_adjunto_real(error_id: str, files: dict[str, bytes]):
    """Crea archivos reales bajo app/data/imagenes/{error_id} y limpia al salir."""
    from app.utils.errores_storage import IMAGENES_PATH

    d = IMAGENES_PATH / error_id
    d.mkdir(parents=True, exist_ok=True)
    try:
        for name, content in files.items():
            (d / name).write_bytes(content)
        yield
    finally:
        shutil.rmtree(d, ignore_errors=True)


class TestServirImagen:
    ERROR_ID = "err-servir"
    FILENAME = "file_1.pdf"
    PDF_BYTES = b"%PDF-1.4 test-fake-pdf-content"

    def _url(self, filename=None):
        filename = filename or self.FILENAME
        return f"/api/control-errores/{self.ERROR_ID}/imagenes/{filename}"

    def test_sin_sesion_200(self):
        client = _client()
        with _archivo_adjunto_real(self.ERROR_ID, {self.FILENAME: self.PDF_BYTES}):
            resp = client.get(self._url())
            assert resp.status_code == 200
            assert resp.data == self.PDF_BYTES
            resp.close()

    def test_con_sesion_200(self):
        client = _client()
        _login(client)
        with _archivo_adjunto_real(self.ERROR_ID, {self.FILENAME: self.PDF_BYTES}):
            resp = client.get(self._url())
            assert resp.status_code == 200
            assert resp.data == self.PDF_BYTES
            resp.close()

    def test_archivo_no_listado_404(self):
        client = _client()
        with _archivo_adjunto_real(self.ERROR_ID, {"otro.pdf": self.PDF_BYTES}):
            resp = client.get(self._url())
            assert resp.status_code == 404
            data = resp.get_json()
            assert data["status"] == "error"

    def test_path_trick_fuera_de_la_carpeta_404(self):
        from app.utils.errores_storage import IMAGENES_PATH

        client = _client()
        with _archivo_adjunto_real(self.ERROR_ID, {self.FILENAME: self.PDF_BYTES}):
            secret = IMAGENES_PATH / "secret.txt"
            secret.write_bytes(b"top-secret")
            try:
                resp = client.get(self._url("../secret.txt"))
                assert resp.status_code == 404
                data = resp.get_json()
                assert data["status"] == "error"
            finally:
                secret.unlink(missing_ok=True)


# =============================================================================
# R14 S6 — Export verbatim Factura and FURIPS (split)
# =============================================================================

class TestR14ExportFacturaYFurips:
    """R14 S6: export tipo_error column is verbatim Factura and FURIPS (split)."""

    def test_export_factura_y_furips_verbatim(self):
        """build_errores_export_workbook writes tipo_error Factura and FURIPS verbatim."""
        with patch(
            "app.services.control_errores_export.listar_imagenes",
            side_effect=_adjuntos_por_scope([], []),
        ):
            with _APP.app_context():
                buffer = build_errores_export_workbook(
                    [_error_fixture(tipo_error="Factura"), _error_fixture(error_id="err-2", tipo_error="FURIPS")],
                    "http://testserver/",
                )
        wb = load_workbook(BytesIO(buffer.read()))
        ws = wb.active
        # Categoría is column 5 (E) = index 4 after ReFactura insertion
        assert ws.cell(row=2, column=5).value == "Factura"
        assert ws.cell(row=3, column=5).value == "FURIPS"
        assert [c.value for c in ws[1]][4] == "Categoría"
        # row list index 4 is also Categoría
        assert [c.value for c in ws[2]][4] == "Factura"
        assert [c.value for c in ws[3]][4] == "FURIPS"

    def test_export_filter_factura_y_furips_via_route(self, app_client):
        """GET /api/control-errores/export with Factura and FURIPS records exports verbatim."""
        _login(app_client)
        fixture = [
            _error_fixture(error_id="factura-1", factura="FEV-001", tipo_error="Factura", creado_en="2026-05-15T10:30:00"),
            _error_fixture(error_id="furips-1", factura="FEV-002", tipo_error="FURIPS", creado_en="2026-05-15T10:30:00"),
            _error_fixture(error_id="otros-1", factura="FAC-002", tipo_error="Otros", creado_en="2026-05-16T10:30:00"),
        ]
        with ExitStack() as stack:
            _patch_export_pipeline(fixture, stack)
            resp = app_client.get("/api/control-errores/export?mes=2026-05")
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.data))
        ws = wb.active
        categorias = [ws.cell(row=r, column=5).value for r in range(2, ws.max_row + 1)]
        assert "Factura" in categorias
        assert "FURIPS" in categorias
        assert "Factura y Furips" not in categorias