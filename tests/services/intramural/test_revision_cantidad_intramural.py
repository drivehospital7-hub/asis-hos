"""Tests for app/services/intramural/revision_cantidad_intramural.py.

Strict TDD: tests written BEFORE implementation.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook


# =============================================================================
# Helpers
# =============================================================================


def _build_workbook(
    headers: list[str],
    rows: list[list],
) -> tuple[Workbook, dict[str, int | None]]:
    """Construye un workbook con headers y datos para pruebas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Build indices matching the internal names used by the detector
    indices: dict[str, int | None] = {}
    internal_names = {
        "Número Factura": "numero_factura",
        "Cód. Equivalente CUPS": "codigo",
        "Procedimiento": "procedimiento",
        "Cantidad": "cantidad",
        "Código Tipo Procedimiento": "codigo_tipo_procedimiento",
        "Laboratorio": "laboratorio",
    }
    for col_idx, header in enumerate(headers):
        if header in internal_names:
            indices[internal_names[header]] = col_idx
    for name in (
        "numero_factura", "codigo", "procedimiento", "cantidad",
        "codigo_tipo_procedimiento", "laboratorio",
    ):
        indices.setdefault(name, None)

    return wb, indices


# =============================================================================
# Unit tests: detect_revision_cantidad_intramural
# =============================================================================


class TestDetectRevisionCantidadIntramural:
    """Tests para detect_revision_cantidad_intramural."""

    # ------------------------------------------------------------------
    # R2: 02 + Lab=No → Cantidad ≤ 2
    # ------------------------------------------------------------------

    def test_r2_flagged_cantidad_5(self) -> None:
        """02 + Lab=No, Cantidad=5 (>2) → flagged as revision."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", 5, "02", "No"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert len(result) == 1
        assert result[0]["factura"] == "F001"
        assert result[0]["cantidad"] == 5

    def test_r2_not_flagged_cantidad_2(self) -> None:
        """02 + Lab=No, Cantidad=2 (≤2) → NOT flagged."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", 2, "02", "No"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert result == []

    # ------------------------------------------------------------------
    # R3: 03/04 → Cantidad ≤ 12
    # ------------------------------------------------------------------

    def test_r3_flagged_cantidad_15(self) -> None:
        """Tipo 03, Cantidad=15 (>12) → flagged."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", 15, "03", "Si"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert len(result) == 1
        assert result[0]["cantidad"] == 15

    def test_r3_not_flagged_cantidad_12(self) -> None:
        """Tipo 04, Cantidad=12 (≤12) → NOT flagged."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", 12, "04", "Si"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert result == []

    # ------------------------------------------------------------------
    # R4: General → Cantidad ≤ 1
    # ------------------------------------------------------------------

    def test_r4_flagged_cantidad_3(self) -> None:
        """Tipo 06, Cantidad=3 (>1) → flagged (general rule)."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", 3, "06", "Si"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert len(result) == 1
        assert result[0]["cantidad"] == 3

    def test_r4_not_flagged_cantidad_1(self) -> None:
        """Tipo 06, Cantidad=1 (≤1) → NOT flagged."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", 1, "06", "Si"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert result == []

    # ------------------------------------------------------------------
    # R5: No tipo_factura filter — all rows evaluated
    # ------------------------------------------------------------------

    def test_r5_all_rows_evaluated(self) -> None:
        """Rows with any tipo_factura value are evaluated — no filter."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [
            ["F001", "X001", "Proc A", 5, "02", "No"],   # flagged
            ["F002", "X002", "Proc B", 3, "06", "Si"],   # flagged (general)
            ["F003", "X003", "Proc C", 1, "06", "Si"],   # not flagged
        ]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert len(result) == 2

    # ------------------------------------------------------------------
    # R6: Flagged item has all 7 keys
    # ------------------------------------------------------------------

    def test_r6_flagged_item_has_all_keys(self) -> None:
        """Flagged item must have factura, codigo, procedimiento, cantidad,
        codigo_tipo_procedimiento, laboratorio, detalle."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", 5, "02", "No"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert len(result) == 1
        item = result[0]
        expected_keys = {
            "factura", "codigo", "procedimiento", "cantidad",
            "codigo_tipo_procedimiento", "laboratorio", "detalle",
        }
        assert set(item.keys()) == expected_keys

    # ------------------------------------------------------------------
    # R7: Graceful degradation — missing columns
    # ------------------------------------------------------------------

    def test_r7_missing_cantidad_returns_empty(self) -> None:
        """Missing Cantidad column → empty list."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", "02", "No"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert result == []

    def test_r7_missing_laboratorio_falls_through_to_general(self) -> None:
        """Missing Laboratorio column → 02+Lab rule can't match,
        falls through to general rule."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento",
        ]
        rows = [["F001", "X001", "Proc A", 3, "02"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        # 02 with missing Laboratorio → general rule → Cantidad 3 > 1 → flagged
        assert len(result) == 1
        assert result[0]["cantidad"] == 3

    def test_r7_missing_codigo_tipo_procedimiento_general_rule(self) -> None:
        """Missing Código Tipo Procedimiento → all rows evaluated
        with general rule."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Laboratorio",
        ]
        rows = [
            ["F001", "X001", "Proc A", 2, "No"],   # 2 > 1 → flagged
            ["F002", "X002", "Proc B", 1, "Si"],   # 1 ≤ 1 → not flagged
        ]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert len(result) == 1
        assert result[0]["factura"] == "F001"

    # ------------------------------------------------------------------
    # Cascade: first-match wins
    # ------------------------------------------------------------------

    def test_cascade_02_beats_03_04(self) -> None:
        """02+Lab=No with Cantidad=15 → uses 02 threshold (≤2),
        flagged as 02 rule, NOT as 03/04 rule."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", 15, "02", "No"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert len(result) == 1
        # Detalle should reference 02+Lab=No rule
        assert "02" in result[0]["detalle"]

    def test_cascade_03_04_beats_general(self) -> None:
        """Tipo 03 with Cantidad=15 → uses 03/04 threshold (≤12),
        flagged as 03/04 rule, NOT as general rule."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", 15, "03", "Si"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert len(result) == 1
        # Detalle should reference 03/04 rule
        assert "03" in result[0]["detalle"] or "04" in result[0]["detalle"]

    def test_cantidad_as_string_flagged(self) -> None:
        """Cantidad as string '2' with general rule → flagged."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", "2", "05", "No"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert len(result) == 1
        assert result[0]["factura"] == "F001"

    def test_cantidad_as_string_not_flagged(self) -> None:
        """Cantidad as string '1' with general rule → NOT flagged."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", "1", "05", "No"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert result == []

    def test_codigo_901101_cantidad_3_not_flagged(self) -> None:
        """Código 901101 con Cant=3 → límite específico (≤3) → NOT flagged."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "901101", "Proc A", 3, "05", "No"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert result == []

    def test_codigo_901101_cantidad_4_flagged(self) -> None:
        """Código 901101 con Cant=4 → excede límite específico → flagged por cascade."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "901101", "Proc A", 4, "05", "No"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert len(result) == 1

    def test_cantidad_as_empty_string_skipped(self) -> None:
        """Cantidad as empty string '' → skipped gracefully."""
        headers = [
            "Número Factura", "Cód. Equivalente CUPS", "Procedimiento",
            "Cantidad", "Código Tipo Procedimiento", "Laboratorio",
        ]
        rows = [["F001", "X001", "Proc A", "", "05", "No"]]
        wb, indices = _build_workbook(headers, rows)
        from app.services.intramural.revision_cantidad_intramural import (
            detect_revision_cantidad_intramural,
        )
        result = detect_revision_cantidad_intramural(wb.active, indices)
        assert result == []
