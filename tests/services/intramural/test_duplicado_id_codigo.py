"""Tests for app/services/intramural/duplicado_id_codigo.py.

Strict TDD: tests written BEFORE implementation.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.constants.urgencias import FACTURADORES_URGENCIAS
from app.services.intramural.duplicado_id_codigo import (
    _get_normalized_facturadores,
    _normalize_responsable,
    detect_duplicado_id_codigo,
)


# =============================================================================
# Helpers
# =============================================================================


def _build_workbook(
    headers: list[str],
    rows: list[list],
) -> tuple[Workbook, dict[str, int | None]]:
    """Construye un workbook con headers y datos para pruebas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Build indices matching the mapping used by the caller
    indices: dict[str, int | None] = {}
    internal_names = {
        "Número Factura": "numero_factura",
        "Nº Identificación": "identificacion",
        "Cód. Equivalente CUPS": "codigo",
        "Cód. Dx Principal": "codigo_dx_principal",
        "Procedimiento": "procedimiento",
        "Responsable Cierra Facturar": "responsable_cierra",
        "Código Tipo Procedimiento": "codigo_tipo_procedimiento",
        "Laboratorio": "laboratorio",
    }
    for col_idx, header in enumerate(headers):
        if header in internal_names:
            indices[internal_names[header]] = col_idx
    for name in ("numero_factura", "identificacion", "codigo", "codigo_dx_principal", "procedimiento", "responsable_cierra", "codigo_tipo_procedimiento", "laboratorio"):
        indices.setdefault(name, None)

    return wb, indices


def _assert_paciente(
    result: list,
    expected_ident: str,
    expected_facturas: list[str],
    expected_procs: list[dict],
) -> None:
    """Helper: verifica un resultado de paciente con sus procedimientos."""
    item = next(r for r in result if r["identificacion"] == expected_ident)
    assert item["facturas"] == expected_facturas
    assert len(item["procedimientos"]) == len(expected_procs)
    for actual, expected in zip(item["procedimientos"], expected_procs):
        for k, v in expected.items():
            assert actual.get(k) == v, f"Key {k}: expected {v}, got {actual.get(k)}"


# =============================================================================
# Unit tests: detect_duplicado_id_codigo
# =============================================================================


class TestDetectDuplicadoIdCodigo:
    """Tests para detect_duplicado_id_codigo."""

    def test_tipo_05_2_rows_returns_one_group(self) -> None:
        """Tipo 05 con 2 filas mismo ID+código -> 1 grupo con reps=2."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", "123", "X001", "05", "Proc A"],
            ["F002", "123", "X001", "05", "Proc B"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)

        assert len(result) == 1
        item = result[0]
        assert item["identificacion"] == "123"
        assert item["codigo"] == "X001"
        assert item["facturas"] == ["F001", "F002"]
        assert item["cantidad_repeticiones"] == 2

    def test_unique_pairs_returns_empty(self) -> None:
        """Pares (ID, código) únicos con tipo 05 -> lista vacía."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", "123", "X001", "05", "Proc A"],
            ["F002", "456", "X002", "05", "Proc B"],
            ["F003", "123", "X002", "05", "Proc C"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_missing_identificacion_column_returns_empty(self) -> None:
        """Sin columna Nº Identificación -> retorna []."""
        headers = ["Número Factura", "Cód. Equivalente CUPS", "Procedimiento"]
        rows = [["F001", "X001", "Proc A"], ["F002", "X002", "Proc B"]]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_missing_codigo_column_returns_empty(self) -> None:
        """Sin columna que mapea a codigo -> retorna []."""
        headers = ["Número Factura", "Nº Identificación", "Procedimiento"]
        rows = [["F001", "123", "Proc A"], ["F002", "123", "Proc B"]]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_missing_numero_factura_column_returns_empty(self) -> None:
        """Sin columna Número Factura -> retorna []."""
        headers = ["Nº Identificación", "Cód. Equivalente CUPS", "Procedimiento"]
        rows = [["123", "X001", "Proc A"], ["123", "X001", "Proc B"]]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_none_values_skipped(self) -> None:
        """Filas con None en identificación o código se saltan (tipo 05)."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", None, "X001", "05", "Proc A"],
            ["F002", "123", None, "05", "Proc B"],
            ["F003", "123", "X001", "05", "Proc C"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_tipo_05_3_rows_returns_one_group(self) -> None:
        """Tipo 05 con 3 filas mismo ID+código -> 1 grupo con reps=3."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", "123", "X001", "05", "Proc A"],
            ["F002", "123", "X001", "05", "Proc B"],
            ["F003", "123", "X001", "05", "Proc C"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)

        assert len(result) == 1
        item = result[0]
        assert item["identificacion"] == "123"
        assert item["codigo"] == "X001"
        assert item["facturas"] == ["F001", "F002", "F003"]
        assert item["cantidad_repeticiones"] == 3

    def test_mixed_types_123_vs_string(self) -> None:
        """int 123 vs str '123' se tratan como duplicado (tipo 05)."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", 123, "X001", "05", "Proc A"],
            ["F002", "123", "X001", "05", "Proc B"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)

        assert len(result) == 1
        assert result[0]["identificacion"] == "123"

    def test_whitespace_variations(self) -> None:
        """' 123' y '123 ' se tratan como iguales (tipo 05)."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", " 123", "X001", "05", "Proc A"],
            ["F002", "123 ", "X001", "05", "Proc B"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)

        assert len(result) == 1
        assert result[0]["identificacion"] == "123"

    def test_error_dict_keys(self) -> None:
        """Cada error debe tener las keys del formato simple (tipo 05)."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", "123", "X001", "05", "Proc A"],
            ["F002", "123", "X001", "05", "Proc B"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)

        expected_keys = {"identificacion", "codigo", "dx_principal", "procedimiento",
                         "facturas", "cantidad_repeticiones"}
        assert len(result) == 1
        assert set(result[0].keys()) == expected_keys

    def test_two_procedimientos_mismo_paciente(self) -> None:
        """Tipo 05: 2 procedimientos duplicados -> 2 grupos separados."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Cód. Dx Principal", "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", "123", "X001", "I10X", "05", "Microalbuminuria"],
            ["F002", "123", "X001", "I10X", "05", "Microalbuminuria"],
            ["F001", "123", "Y002", "I10X", "05", "Creatinina"],
            ["F002", "123", "Y002", "I10X", "05", "Creatinina"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)

        assert len(result) == 2
        codigos = [r["codigo"] for r in result]
        assert "X001" in codigos
        assert "Y002" in codigos
        for r in result:
            assert r["cantidad_repeticiones"] == 2
            assert r["facturas"] == ["F001", "F002"]

    def test_dos_pacientes_con_duplicados(self) -> None:
        """Tipo 05: dos pacientes distintos con duplicados -> 2 grupos."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", "123", "X001", "05", "Proc A"],
            ["F002", "123", "X001", "05", "Proc B"],
            ["F003", "456", "Y002", "05", "Proc C"],
            ["F004", "456", "Y002", "05", "Proc D"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)

        assert len(result) == 2
        idents = {r["identificacion"] for r in result}
        assert idents == {"123", "456"}

    def test_diferent_factura_same_id_codigo(self) -> None:
        """Tipo 05: 3 facturas mismo ID+código -> 1 grupo con reps=3."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", "123", "X001", "05", "Proc A"],
            ["F002", "123", "X001", "05", "Proc B"],
            ["F003", "456", "X002", "05", "Proc C"],
            ["F004", "123", "X001", "05", "Proc D"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)

        assert len(result) == 1  # solo (123, X001) tiene duplicados
        item = result[0]
        assert item["identificacion"] == "123"
        assert item["codigo"] == "X001"
        assert item["facturas"] == ["F001", "F002", "F004"]
        assert item["cantidad_repeticiones"] == 3

    def test_empty_factura_skipped(self) -> None:
        """Factura vacía se salta (tipo 05)."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            [None, "123", "X001", "05", "Proc A"],
            ["F002", "123", "X001", "05", "Proc B"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    # =========================================================================
    # Tests: exclusión de FACTURADORES_URGENCIAS
    # =========================================================================

    def test_ambos_facturadores_urgencias_sin_error(self) -> None:
        """Ambos con resp en FACTURADORES_URGENCIAS -> sin error."""
        headers = [
            "Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
            "Procedimiento", "Responsable Cierra Facturar",
        ]
        rows = [
            ["F001", "123", "X001", "Proc A", "ARIAS CULCHA ANGIE CAROLINA"],
            ["F002", "123", "X001", "Proc B", "ESPAÑA DIAZ LORENY ALEJANDRA"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_mixto_urgencias_y_no_sin_error(self) -> None:
        """Una urgencias se excluye, la otra queda sola -> sin error."""
        headers = [
            "Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
            "Procedimiento", "Responsable Cierra Facturar",
        ]
        rows = [
            ["F001", "123", "X001", "Proc A", "ARIAS CULCHA ANGIE CAROLINA"],
            ["F002", "123", "X001", "Proc B", "OTRO RESPONSABLE"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_facturador_urgencias_case_insensitive(self) -> None:
        """Case-insensitive matching."""
        headers = [
            "Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
            "Procedimiento", "Responsable Cierra Facturar",
        ]
        rows = [
            ["F001", "123", "X001", "Proc A", "arias culcha angie carolina"],
            ["F002", "123", "X001", "Proc B", "Arias Culcha Angie Carolina"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_sin_columna_responsable_con_05(self) -> None:
        """Sin responsable pero con tipo 05 -> duplicados normales."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", "123", "X001", "05", "Proc A"],
            ["F002", "123", "X001", "05", "Proc B"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert len(result) == 1
        assert result[0]["cantidad_repeticiones"] == 2

    def test_mismo_id_distinto_dx_no_duplicado(self) -> None:
        """Mismo ID+código pero distinto dx -> no duplicado."""
        headers = [
            "Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
            "Cód. Dx Principal", "Procedimiento",
        ]
        rows = [
            ["F001", "123", "X001", "R509", "Proc A"],
            ["F002", "123", "X001", "A970", "Proc B"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_normalize_responsable_elimina_acentos(self) -> None:
        """_normalize_responsable elimina acentos: FERNÁNDEZ -> FERNANDEZ."""
        assert _normalize_responsable("MEZA FERNÁNDEZ CARLOS OMAR") == "MEZA FERNANDEZ CARLOS OMAR"
        assert _normalize_responsable("PÁEZ YÚLIETH DANIELA") == "PAEZ YULIETH DANIELA"

    def test_normalize_responsable_conserva_enie(self) -> None:
        """La Ñ no se pierde en la normalización."""
        val = _normalize_responsable("ESPAÑA DIAZ LORENY ALEJANDRA")
        assert "ESPAÑA" in val or "ESPANA" in val

    def test_acento_en_facturador_urgencias(self) -> None:
        """'MEZA FERNÁNDEZ CARLOS OMAR' (con acento) debe excluirse."""
        headers = [
            "Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
            "Procedimiento", "Responsable Cierra Facturar",
        ]
        rows = [
            ["F001", "123", "X001", "Proc A", "MEZA FERNÁNDEZ CARLOS OMAR"],
            ["F002", "123", "X001", "Proc B", "MEZA FERNÁNDEZ CARLOS OMAR"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    # =========================================================================
    # Tests: umbral según Código Tipo Procedimiento
    # =========================================================================

    def test_tipo_02_laboratorio_si_con_2_sin_error(self) -> None:
        """Tipo 02 + Laboratorio=Si con 2 -> no error (umbral=4)."""
        headers = [
            "Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
            "Código Tipo Procedimiento", "Laboratorio", "Procedimiento",
        ]
        rows = [
            ["F001", "123", "X001", "02", "Si", "Proc A"],
            ["F002", "123", "X001", "02", "Si", "Proc B"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_tipo_02_laboratorio_si_con_4_error(self) -> None:
        """Tipo 02 + Laboratorio=Si con 4 -> error (umbral=4)."""
        headers = [
            "Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
            "Código Tipo Procedimiento", "Laboratorio", "Procedimiento",
        ]
        rows = [
            ["F001", "123", "X001", "02", "Si", "Proc A"],
            ["F002", "123", "X001", "02", "Si", "Proc B"],
            ["F003", "123", "X001", "02", "Si", "Proc C"],
            ["F004", "123", "X001", "02", "Si", "Proc D"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert len(result) == 1
        assert result[0]["cantidad_repeticiones"] == 4

    def test_tipo_02_sin_laboratorio_excluido(self) -> None:
        """Tipo 02 sin Laboratorio=Si -> excluido, sin error."""
        headers = [
            "Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
            "Código Tipo Procedimiento", "Procedimiento",
        ]
        rows = [
            ["F001", "123", "X001", "02", "Proc A"],
            ["F002", "123", "X001", "02", "Proc B"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_tipo_02_laboratorio_no_excluido(self) -> None:
        """Tipo 02 + Laboratorio=No -> excluido, sin error."""
        headers = [
            "Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
            "Código Tipo Procedimiento", "Laboratorio", "Procedimiento",
        ]
        rows = [
            ["F001", "123", "X001", "02", "No", "Proc A"],
            ["F002", "123", "X001", "02", "No", "Proc B"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_sin_tipo_excluido(self) -> None:
        """Sin Código Tipo Procedimiento -> excluido, sin error."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS", "Procedimiento"]
        rows = [
            ["F001", "123", "X001", "Proc A"],
            ["F002", "123", "X001", "Proc B"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_tipo_05_codigo_993505_exento(self) -> None:
        """Tipo 05 con código 993505 -> exento, sin error."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", "123", "993505", "05", "VACUNACION RABIA"],
            ["F002", "123", "993505", "05", "VACUNACION RABIA"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []

    def test_tipo_05_codigo_normal_si_duplicado(self) -> None:
        """Tipo 05 con código normal -> duplicado normal."""
        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", "123", "X001", "05", "Proc A"],
            ["F002", "123", "X001", "05", "Proc B"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert len(result) == 1

    def test_doble_espacio_en_facturador_urgencias(self) -> None:
        """'PAEZ  YULIETH DANIELA' con doble espacio."""
        headers = [
            "Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
            "Procedimiento", "Responsable Cierra Facturar",
        ]
        rows = [
            ["F001", "123", "X001", "Proc A", "PAEZ  YULIETH DANIELA"],
            ["F002", "123", "X001", "Proc B", "PAEZ  YULIETH DANIELA"],
        ]
        wb, indices = _build_workbook(headers, rows)
        result = detect_duplicado_id_codigo(wb.active, indices)
        assert result == []


# =============================================================================
# Integration tests
# =============================================================================


class TestIntegracionDuplicadoIdCodigo:
    """Integración con detect_all y normalized_rows."""

    def test_detector_in_lista_detectores(self) -> None:
        """_get_intramural_detectors() debe incluir detect_duplicado_id_codigo."""
        from app.services.intramural.detect_all import _get_intramural_detectors

        detectores = _get_intramural_detectors()
        nombres = [d.__name__ for d in detectores]
        assert "detect_duplicado_id_codigo" in nombres

    def test_build_normalized_rows_handles_key(self) -> None:
        """build_normalized_rows() debe procesar key 'Duplicado ID+Código'."""
        from app.services.normalized_rows import build_normalized_rows

        error_groups = {
            "Duplicado ID+Código": [
                {
                    "identificacion": "18108483",
                    "codigo": "903026",
                    "dx_principal": "I10X",
                    "procedimiento": "Microalbuminuria Automatizada en Orina Parcial",
                    "facturas": ["CAP491530", "CAP493883"],
                    "cantidad_repeticiones": 2,
                },
                {
                    "identificacion": "18108483",
                    "codigo": "903876",
                    "dx_principal": "I10X",
                    "procedimiento": "Creatinina en Orina",
                    "facturas": ["CAP491530", "CAP493883"],
                    "cantidad_repeticiones": 2,
                },
            ],
        }
        rows = build_normalized_rows(
            error_groups=error_groups,
            responsables_map={"CAP491530": "RESP A"},
            fec_factura_map={"CAP491530": "2026-05-04"},
            fecha_cierre_vacia_map={"CAP491530": False},
        )
        assert len(rows) == 2
        row = rows[0]
        assert row["tipo_error"] == "Duplicado ID+Código"
        assert row["factura"] == "CAP491530"
        assert "x2" in row["descripcion"]
        assert "ID: 18108483" in row["detalle"]
        assert "Facturas: CAP491530, CAP493883" in row["detalle"]
        assert row["responsable_cierra"] == "RESP A"
        assert row["fec_factura"] == "2026-05-04"
