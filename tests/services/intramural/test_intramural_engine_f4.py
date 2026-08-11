"""Strict TDD F4: Tests for Intramural-specific engine toggles.

Covers T-F4.1 (CentroCostoIntramuralEvaluator), T-F4.2 (ide_contrato legacy),
T-F4.3 (RevisionCantidadIntramuralEvaluator), T-F4.4 (detect_all integration),
T-F4.5 (snapshot structure).
"""
from __future__ import annotations

from unittest.mock import MagicMock, patch

from app.services.engine.context import EvaluationContext
from app.services.engine.evaluators import (
    CentroCostoIntramuralEvaluator,
    EVALUATOR_REGISTRY,
    RevisionCantidadIntramuralEvaluator,
)


# ═══════════════════════════════════════════════════════════════════════════
# T-F4.1: CentroCostoIntramuralEvaluator
# ═══════════════════════════════════════════════════════════════════════════

class TestCentroCostoIntramuralEvaluator:
    """Unit tests for CentroCostoIntramuralEvaluator (operator: centro_costo_intramural).

    Each test creates an EvaluationContext with the needed invoice_data fields
    and asserts the evaluator returns True (violation) or False (valid).
    """

    def _make_context(self, overrides: dict | None = None) -> EvaluationContext:
        """Build a default-valid Intramural context.

        Defaults: centro=SERVICIOS AMBULATORIOS- CONSULTA EXTERNA Y PROCEDIMIENTOS,
        codigo=890201, codigo_tipo=01, laboratorio=No, tarifario=NO SOAT,
        responsable_cierra=''.
        """
        data = {
            "centro_costo": "SERVICIOS AMBULATORIOS- CONSULTA EXTERNA Y PROCEDIMIENTOS",
            "codigo": "890201",
            "codigo_tipo_procedimiento": "01",
            "laboratorio": "No",
            "tarifario": "NO SOAT",
            "responsable_cierra": "",
        }
        if overrides:
            data.update(overrides)
        return EvaluationContext(invoice_data=data)

    # ── Common rules (same as CentroCostoCheck except REGLA3) ──

    def test_regla9_tarifario_farmacia_fuera_de_farmacia(self):
        """REGLA9: tarifario=farmacia, centro!=FARMACIA → violation."""
        ctx = self._make_context({
            "tarifario": "Suminstros, Medicamentos",
            "centro_costo": "URGENCIAS",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_regla9_tarifario_farmacia_en_farmacia(self):
        """REGLA9: tarifario=farmacia, centro=FARMACIA → OK."""
        ctx = self._make_context({
            "tarifario": "Suminstros, Medicamentos",
            "centro_costo": "APOYO TERAPEUTICO-FARMACIA E INSUMOS.",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    def test_regla1_diagnostico_lab_no_fuera_diag(self):
        """REGLA1: cod_tipo=02, lab=No, codigo not exceptuado, centro!=APOYO_DIAG → violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "02",
            "laboratorio": "No",
            "codigo": "901210",
            "centro_costo": "URGENCIAS",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_regla1_diagnostico_lab_no_en_diag(self):
        """REGLA1: cod_tipo=02, lab=No, centro=APOYO_DIAG → OK."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "02",
            "laboratorio": "No",
            "codigo": "901210",
            "centro_costo": "APOYO DIAGNOSTICO-IMAGENOLOGIA",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    def test_reverse1_centro_apoyo_diag_sin_tipo_02(self):
        """REVERSE1: centro=APOYO_DIAG, tipo!=02 → violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "01",
            "laboratorio": "No",
            "centro_costo": "APOYO DIAGNOSTICO-IMAGENOLOGIA",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_regla2_traslados_fuera_traslados(self):
        """REGLA2: tipo=14, centro!=TRASLADOS → violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "14",
            "centro_costo": "URGENCIAS",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_reverse2_centro_traslados_sin_tipo_14(self):
        """REVERSE2: centro=TRASLADOS, tipo!=14 → violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "01",
            "centro_costo": "TRASLADOS",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_regla8_cod_hospitalizacion_fuera_hosp(self):
        """REGLA8: codigo in CODIGOS_HOSPITALIZACION_ESTANCIA, centro!=HOSP → violation."""
        ctx = self._make_context({
            "codigo": "890601H",
            "centro_costo": "URGENCIAS",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_regla8_cod_hospitalizacion_en_hosp(self):
        """REGLA8: codigo=890601H, centro=HOSPITALIZACIÓN → OK."""
        ctx = self._make_context({
            "codigo": "890601H",
            "centro_costo": "HOSPITALIZACIÓN - ESTANCIA GENERAL",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    def test_regla4_cod_quirofano_fuera_quirofano(self):
        """REGLA4: codigo in CODIGOS_QUIROFANO_URGENCIAS, centro!=QUIROFANO → violation."""
        ctx = self._make_context({
            "codigo": "735301",
            "centro_costo": "URGENCIAS",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_reverse4_centro_quirofano_sin_cod_quirofano(self):
        """REVERSE4: centro=QUIROFANO, codigo not in CODIGOS_QUIROFANO → violation."""
        ctx = self._make_context({
            "codigo": "890201",
            "centro_costo": "QUIRÓFANOS Y SALAS DE PARTO- SALA DE PARTO",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_reverse9_centro_farmacia_sin_tarifario_farmacia(self):
        """REVERSE9: centro=FARMACIA, tarifario!=farmacia → violation."""
        ctx = self._make_context({
            "centro_costo": "APOYO TERAPEUTICO-FARMACIA E INSUMOS.",
            "tarifario": "NO SOAT",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    # ── REGLA3 is NOT applied (Urgencias-specific, replaced by REGLA3-INTRAMURAL) ──

    def test_regla3_urgencias_not_applied(self):
        """REGLA3 (Urgencias centro PYP) must NOT be applied in Intramural evaluator.
        codigo in CODIGOS_PYP_URGENCIAS but centro != CC_PYP_URGENCIAS → IS a violation
        in Intramural because REGLA3-INTRAMURAL also checks codigo in CODIGOS_PYP_URGENCIAS
        but maps to CENTROS_COSTO_PYP_INTRAMURAL instead.
        """
        ctx = self._make_context({
            "codigo": "990211",  # in CODIGOS_PYP_URGENCIAS → triggers REGLA3-INTRAMURAL
            "centro_costo": "URGENCIAS",  # NOT in CENTROS_COSTO_PYP_INTRAMURAL
        })
        evaluator = CentroCostoIntramuralEvaluator()
        # REGLA3-INTRAMURAL triggers: codigo PyP + centro not in CC_PYP_INTRA
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    # ── REGLA3-INTRAMURAL: codigo PyP → CC PyP Intramural ──

    def test_regla3_intramural_codigo_pyp_fuera_pyp(self):
        """REGLA3-INTRAMURAL: codigo in CODIGOS_PYP_URGENCIAS, centro not in
        CENTROS_COSTO_PYP_INTRAMURAL → violation."""
        ctx = self._make_context({
            "codigo": "990211",
            "centro_costo": "URGENCIAS",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_regla3_intramural_codigo_pyp_en_pyp(self):
        """REGLA3-INTRAMURAL: codigo PyP, centro in CENTROS_COSTO_PYP_INTRAMURAL → OK."""
        ctx = self._make_context({
            "codigo": "990211",
            "centro_costo": "SERVICIOS AMBULATORIOS- PROMOCION Y PREVENCION",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    def test_reverse3_intramural_centro_pyp_sin_codigo_pyp(self):
        """REVERSE3-INTRAMURAL: centro PyP Intramural, codigo not in CODIGOS_PYP → violation."""
        ctx = self._make_context({
            "codigo": "890201",
            "centro_costo": "SERVICIOS AMBULATORIOS- PROMOCION Y PREVENCION",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    # ── REGLA6: tipo=05 + vacunación → SALUD PUBLICA ──

    def test_regla6_tipo_05_vacunacion_fuera_salud_publica(self):
        """REGLA6: tipo=05, codigo not in excluidos, not PyP, centro!=SALUD PUBLICA → violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "05",
            "codigo": "993505",  # not in excluidos
            "centro_costo": "FARMACIA",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_regla6_tipo_05_en_salud_publica(self):
        """REGLA6: tipo=05, centro=SALUD PUBLICA → OK."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "05",
            "codigo": "993505",
            "centro_costo": "SALUD PUBLICA-VACUNACION  REGULAR",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    def test_regla6_tipo_05_codigo_excluido_vacunacion(self):
        """REGLA6: tipo=05, codigo in CODIGOS_EXCLUIDOS_VACUNACION → NOT violation
        (código excluido no exige SALUD PUBLICA)."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "05",
            "codigo": "906249",  # in CODIGOS_EXCLUIDOS_VACUNACION
            "centro_costo": "FARMACIA",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    def test_reverse6_centro_salud_publica_sin_tipo_05(self):
        """REVERSE6: centro=SALUD PUBLICA, tipo!=05 → violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "01",
            "centro_costo": "SALUD PUBLICA-VACUNACION  REGULAR",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    # ── REGLA7: tipo=03/04 → SERVICIOS AMBULATORIOS ──

    def test_regla7_tipo_03_fuera_ambulatorio(self):
        """REGLA7: tipo=03, codigo not exceptuado, centro!=AMBULATORIO → violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "03",
            "codigo": "890201",
            "centro_costo": "URGENCIAS",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_regla7_tipo_04_en_ambulatorio(self):
        """REGLA7: tipo=04, centro=AMBULATORIO → OK."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "04",
            "codigo": "890201",
            "centro_costo": "SERVICIOS AMBULATORIOS- CONSULTA EXTERNA Y PROCEDIMIENTOS",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    def test_regla7_tipo_03_codigo_exceptuado(self):
        """REGLA7: tipo=03, codigo in CODIGOS_EXCEPTUADOS_AMBULATORIO → NOT violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "03",
            "codigo": "735301",  # exceptuado (QUIROFANO)
            "centro_costo": "QUIRÓFANOS Y SALAS DE PARTO- SALA DE PARTO",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    def test_reverse7_centro_ambulatorio_sin_tipo_03_04(self):
        """REVERSE7: centro=AMBULATORIO, tipo not in {03,04} → violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "01",
            "codigo": "890201",
            "centro_costo": "SERVICIOS AMBULATORIOS- CONSULTA EXTERNA Y PROCEDIMIENTOS",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    # ── REGLA10: tipo=02/05 + Lab=Si → LABORATORIO CLINICO ──

    def test_regla10_tipo_02_lab_si_fuera_lab(self):
        """REGLA10: tipo=02, lab=Si, centro not in CENTROS_COSTO_LABORATORIO → violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "02",
            "laboratorio": "Si",
            "codigo": "901210",
            "centro_costo": "URGENCIAS",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_regla10_tipo_05_lab_si_en_lab(self):
        """REGLA10: tipo=05, lab=Si, centro in CENTROS_COSTO_LABORATORIO → OK."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "05",
            "laboratorio": "Si",
            "codigo": "901210",
            "centro_costo": "APOYO DIAGNOSTICO-LABORATOR CLINICO",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    def test_reverse10_centro_lab_sin_tipo_02_05(self):
        """REVERSE10: centro=LAB, tipo not in {02,05} → violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "01",
            "laboratorio": "Si",
            "codigo": "901210",
            "centro_costo": "APOYO DIAGNOSTICO-LABORATOR CLINICO",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_reverse10_centro_lab_tipo_02_lab_no(self):
        """REVERSE10: centro=LAB, tipo=02, lab=No, not exceptuado → violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "02",
            "laboratorio": "No",
            "codigo": "901210",
            "centro_costo": "APOYO DIAGNOSTICO-LABORATOR CLINICO",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_reverse10_centro_lab_codigo_exceptuado(self):
        """REVERSE10: centro=LAB, codigo in CODIGOS_EXCEPTUADOS → OK
        (exceptuados tienen Lab=No pero son laboratorio válido)."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "02",
            "laboratorio": "No",
            "codigo": "903883",  # in CODIGOS_EXCEPTUADOS
            "centro_costo": "APOYO DIAGNOSTICO-LABORATOR CLINICO",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    # ── REGLA_RESPONSABLE_URGENCIAS ──

    def test_responsable_urgencias_tipo_01(self):
        """RESPONSABLE: responsable in FACTURADORES_URGENCIAS, tipo=01,
        centro not in {URGENCIAS, HOSP} → violation."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "01",
            "responsable_cierra": "ESPAÑA DIAZ LORENY ALEJANDRA",
            "codigo": "890201",
            "centro_costo": "FARMACIA",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is True

    def test_responsable_urgencias_ya_en_urgencias(self):
        """RESPONSABLE: responsable in FACTURADORES, tipo=01, centro=URGENCIAS → OK."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "01",
            "responsable_cierra": "ESPAÑA DIAZ LORENY ALEJANDRA",
            "codigo": "890201",
            "centro_costo": "URGENCIAS",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    def test_responsable_urgencias_codigo_exceptuado(self):
        """RESPONSABLE: codigo in CODIGOS_EXCEPTUADOS_RESPONSABLE_URGENCIAS → OK
        (códigos exceptuados como 735301=QUIRÓFANO)."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "04",
            "responsable_cierra": "ESPAÑA DIAZ LORENY ALEJANDRA",
            "codigo": "735301",  # exceptuado
            "centro_costo": "QUIRÓFANOS Y SALAS DE PARTO- SALA DE PARTO",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    def test_responsable_urgencias_no_es_facturador(self):
        """RESPONSABLE: responsable NOT in FACTURADORES_URGENCIAS → OK."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "01",
            "responsable_cierra": "PEREZ JUAN",
            "codigo": "890201",
            "centro_costo": "FARMACIA",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    # ── Edge cases ──

    def test_empty_centro_returns_false(self):
        """Empty centro_costo must not cause violation (invalid state)."""
        ctx = self._make_context({
            "centro_costo": "",
        })
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=ctx) is False

    def test_none_context_returns_false(self):
        """No context must return False (cannot evaluate)."""
        evaluator = CentroCostoIntramuralEvaluator()
        assert evaluator.evaluate({}, "", None, context=None) is False


# ═══════════════════════════════════════════════════════════════════════════
# T-F4.3: RevisionCantidadIntramuralEvaluator
# ═══════════════════════════════════════════════════════════════════════════

class TestRevisionCantidadIntramuralEvaluator:
    """Unit tests for RevisionCantidadIntramuralEvaluator.

    Returns True when quantity exceeds threshold (detection = problem found).
    """

    def _make_context(self, overrides: dict | None = None) -> EvaluationContext:
        data = {
            "codigo_tipo_procedimiento": "01",
            "laboratorio": "No",
            "codigo": "890201",
        }
        if overrides:
            data.update(overrides)
        return EvaluationContext(invoice_data=data)

    # Rule 1: tipo=02 + Lab=No → Cant > 2

    def test_rule1_tipo_02_lab_no_cant_5_flagged(self):
        """tipo=02, Lab=No, cantidad=5 (>2) → True."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "02",
            "laboratorio": "No",
        })
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, 5, None, context=ctx) is True

    def test_rule1_tipo_02_lab_no_cant_2_not_flagged(self):
        """tipo=02, Lab=No, cantidad=2 (=2) → False (threshold 2, not >2)."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "02",
            "laboratorio": "No",
        })
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, 2, None, context=ctx) is False

    # Rule 2: tipo=03/04 → Cant > 12

    def test_rule2_tipo_03_cant_15_flagged(self):
        """tipo=03, cantidad=15 (>12) → True."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "03",
            "laboratorio": "No",
        })
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, 15, None, context=ctx) is True

    def test_rule2_tipo_04_cant_14_flagged(self):
        """tipo=04, cantidad=14 (>13 threshold) → True."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "04",
            "laboratorio": "No",
        })
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, 14, None, context=ctx) is True

    def test_rule2_tipo_03_cant_12_not_flagged(self):
        """tipo=03, cantidad=12 (=12) → False (threshold 12, not >12)."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "03",
            "laboratorio": "No",
        })
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, 12, None, context=ctx) is False

    # Rule 3 (general): any other → Cant > 1

    def test_rule3_general_cant_3_flagged(self):
        """tipo=01, cantidad=3 (>1) → True."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "01",
            "laboratorio": "No",
        })
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, 3, None, context=ctx) is True

    def test_rule3_general_cant_1_not_flagged(self):
        """tipo=01, cantidad=1 (=1) → False."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "01",
            "laboratorio": "No",
        })
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, 1, None, context=ctx) is False

    # Specific code 901101 → max 3 (CODIGOS_LIMITE_ESPECIFICO_INTRAMURAL)

    def test_codigo_901101_cant_3_not_flagged(self):
        """código=901101, cantidad=3 → False (specific limit is 3)."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "03",
            "laboratorio": "No",
            "codigo": "901101",
        })
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, 3, None, context=ctx) is False

    def test_codigo_901101_cant_4_falls_to_cascade(self):
        """código=901101, cantidad=4, tipo=03 → falls through specific limit
        (4 > 3) to cascade, then tipo=03 checks Cant>12 → False (4 <= 12)."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "03",
            "laboratorio": "No",
            "codigo": "901101",
        })
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, 4, None, context=ctx) is False

    def test_codigo_901101_cant_15_tipo_03_flagged(self):
        """código=901101, cantidad=15, tipo=03 → specific limit exceeded,
        falls to cascade, then Cant>12 → True."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "03",
            "laboratorio": "No",
            "codigo": "901101",
        })
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, 15, None, context=ctx) is True

    # Edge cases

    def test_cascade_02_beats_03_04(self):
        """tipo=02 + lab=No → only rule1 applies (does not fall to rule2/3)."""
        ctx = self._make_context({
            "codigo_tipo_procedimiento": "02",
            "laboratorio": "No",
        })
        evaluator = RevisionCantidadIntramuralEvaluator()
        # cantidad=2 → =2 (not >2) → False (rule1 rejects, does NOT fall through to general)
        assert evaluator.evaluate({}, 2, None, context=ctx) is False
        # cantidad=3 → >2 (rule1) → True
        assert evaluator.evaluate({}, 3, None, context=ctx) is True

    def test_none_context_returns_false(self):
        """No context → False."""
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, 5, None, context=None) is False

    def test_empty_context_returns_false(self):
        """Context without invoice_data → False."""
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, 5, None, context=EvaluationContext()) is False

    def test_null_cantidad_returns_false(self):
        """Null cantidad → False."""
        ctx = self._make_context()
        evaluator = RevisionCantidadIntramuralEvaluator()
        assert evaluator.evaluate({}, None, None, context=ctx) is False


# ═══════════════════════════════════════════════════════════════════════════
# T-F4.4 & T-F4.5: detect_all.py integration
# ═══════════════════════════════════════════════════════════════════════════

class TestIntramuralF4Integration:
    """Integration tests: detect_all_problems_intramural routes to engine when activated.

    T-F4.4: toggles route centro_costo_intramural and revision_cantidad_intramural
    to engine. ide_contrato_intramural stays legacy (too complex for row-by-row engine).

    T-F4.5: snapshot structure matching between engine and legacy paths.
    """

    AREA = "intramural"
    EXPECTED_KEYS = frozenset({
        "centros_de_costos", "ide_contrato", "revision_cantidad",
        "decimales", "tipo_identificacion_edad",
        "tipo_identificacion_entidad", "codigo_entidad_vs_afiliacion",
        "tipo_usuario", "copago_entidad", "cups_sin_contrato",
        "profesionales", "duplicado_id_codigo",
    })

    def _build_sheet(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        headers = [
            "Número Factura", "Código", "Cantidad", "Vlr. Unitario",
            "Vlr. Procedimiento", "Tipo Doc.", "Edad", "Tipo Identificación",
            "Código Entidad Cobrar", "Entidad Afiliación", "Tipo Usuario",
            "Vlr. Copago", "Código CUPS", "Fec Factura", "Fecha Cierre",
            "Responsable Cierra", "Tarifario",
            "Tipo Factura Descripción", "Centro Costo", "Procedimiento",
            "Codigo Tipo Procedimiento", "Laboratorio",
        ]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        # Row 1: Intramural with clean data (centro correct for tipo=01)
        ws.cell(row=2, column=1, value="INTRA-001")
        ws.cell(row=2, column=2, value="890201")
        ws.cell(row=2, column=3, value=1)
        ws.cell(row=2, column=4, value=100.00)
        ws.cell(row=2, column=18, value="Intramural")
        ws.cell(row=2, column=19, value="SERVICIOS AMBULATORIOS- CONSULTA EXTERNA Y PROCEDIMIENTOS")
        ws.cell(row=2, column=20, value="CONSULTA GENERAL")
        ws.cell(row=2, column=21, value="01")
        ws.cell(row=2, column=22, value="No")
        # Row 2: Another Intramural row
        ws.cell(row=3, column=1, value="INTRA-002")
        ws.cell(row=3, column=2, value="990211")
        ws.cell(row=3, column=3, value=2)
        ws.cell(row=3, column=4, value=200.00)
        ws.cell(row=3, column=18, value="Intramural")
        ws.cell(row=3, column=19, value="URGENCIAS")  # Wrong for PyP code!
        ws.cell(row=3, column=20, value="CONSEJERIA VIH")
        ws.cell(row=3, column=21, value="01")
        ws.cell(row=3, column=22, value="No")
        indices = {h: i for i, h in enumerate(headers)}
        return wb, indices

    def _mock_session(self):
        s = MagicMock()
        q = MagicMock()
        q.filter.return_value = q
        q.order_by.return_value = q
        q.first.return_value = None
        q.all.return_value = []
        s.query.return_value = q
        return s

    # ── Engine path routes centro_costo_intramural ──

    def test_engine_path_routes_centro_costo_intramural(self):
        """Engine path MUST call RuleBasedDetector for centro_costo_intramural_valido."""
        import os
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = self._mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md

                from app.services.intramural.detect_all import (
                    detect_all_problems_intramural,
                )
                wb, idx = self._build_sheet()
                old = os.environ.pop("USE_RULE_ENGINE", None)
                os.environ["USE_RULE_ENGINE"] = "true"
                try:
                    r, _ = detect_all_problems_intramural(wb.active, idx)
                finally:
                    if old is not None:
                        os.environ["USE_RULE_ENGINE"] = old
                    else:
                        os.environ.pop("USE_RULE_ENGINE", None)

        call_names = [call[0][0] for call in m_dc.call_args_list]
        assert "centro_costo_intramural_valido" in call_names, (
            f"centro_costo_intramural_valido not in calls: {call_names}"
        )
        assert "centros_de_costos" in r["problemas"]

    # ── Engine path routes revision_cantidad_intramural ──

    def test_engine_path_routes_revision_cantidad(self):
        """Engine path MUST call RuleBasedDetector for revision_cantidad_intramural."""
        import os
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = self._mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md

                from app.services.intramural.detect_all import (
                    detect_all_problems_intramural,
                )
                wb, idx = self._build_sheet()
                old = os.environ.pop("USE_RULE_ENGINE", None)
                os.environ["USE_RULE_ENGINE"] = "true"
                try:
                    r, _ = detect_all_problems_intramural(wb.active, idx)
                finally:
                    if old is not None:
                        os.environ["USE_RULE_ENGINE"] = old
                    else:
                        os.environ.pop("USE_RULE_ENGINE", None)

        call_names = [call[0][0] for call in m_dc.call_args_list]
        assert "revision_cantidad_intramural" in call_names, (
            f"revision_cantidad_intramural not in calls: {call_names}"
        )
        assert "revision_cantidad" in r["problemas"]

    # ── IDE Contrato stays legacy (documented) ──

    def test_ide_contrato_stays_legacy(self):
        """IDE Contrato must still call legacy detector even in engine path
        (too complex for row-by-row engine — pre-scans sheet for laboratorio)."""
        import os
        from app.services.intramural import ide_contrato_intramural as ide_module
        original_fn = ide_module.detect_ide_contrato_intramural
        call_log: list[str] = []

        def tracking_fn(ds, idx):
            call_log.append("called")
            return original_fn(ds, idx)

        try:
            ide_module.detect_ide_contrato_intramural = tracking_fn
            from app.services.intramural.detect_all import (
                detect_all_problems_intramural,
            )
            wb, idx = self._build_sheet()
            old = os.environ.pop("USE_RULE_ENGINE", None)
            os.environ["USE_RULE_ENGINE"] = "true"
            try:
                r, _ = detect_all_problems_intramural(wb.active, idx)
            finally:
                if old is not None:
                    os.environ["USE_RULE_ENGINE"] = old
                else:
                    os.environ.pop("USE_RULE_ENGINE", None)
        finally:
            ide_module.detect_ide_contrato_intramural = original_fn

        assert call_log, (
            "Legacy detect_ide_contrato_intramural was NOT called in engine path!"
        )
        assert "ide_contrato" in r["problemas"]

    # ── Snapshot: keys present in both paths ──

    def test_keys_present_in_both_paths(self):
        """All F4-specific keys must be present in both engine and legacy paths."""
        import os
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = self._mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md

                from app.services.intramural.detect_all import (
                    detect_all_problems_intramural,
                )
                wb, idx = self._build_sheet()

                # Engine path
                old = os.environ.pop("USE_RULE_ENGINE", None)
                os.environ["USE_RULE_ENGINE"] = "true"
                try:
                    er, _ = detect_all_problems_intramural(wb.active, idx)
                finally:
                    if old is not None:
                        os.environ["USE_RULE_ENGINE"] = old
                    else:
                        os.environ.pop("USE_RULE_ENGINE", None)

                # Legacy path
                old2 = os.environ.pop("USE_RULE_ENGINE", None)
                os.environ["USE_RULE_ENGINE"] = "false"
                try:
                    lr, _ = detect_all_problems_intramural(wb.active, idx)
                finally:
                    if old2 is not None:
                        os.environ["USE_RULE_ENGINE"] = old2
                    else:
                        os.environ.pop("USE_RULE_ENGINE", None)

        for k in self.EXPECTED_KEYS:
            assert k in er["problemas"], f"Engine missing key: {k}"
            assert k in lr["problemas"], f"Legacy missing key: {k}"

    def test_totals_keys_present(self):
        """Total counts must be present for F4 detectors."""
        import os
        with patch("app.database.get_session") as m_gs:
            with patch("app.services.engine.rule_based_detector.RuleBasedDetector") as m_dc:
                m_gs.return_value = self._mock_session()
                md = MagicMock()
                md.detect.return_value = []
                m_dc.return_value = md

                from app.services.intramural.detect_all import (
                    detect_all_problems_intramural,
                )
                wb, idx = self._build_sheet()
                old = os.environ.pop("USE_RULE_ENGINE", None)
                os.environ["USE_RULE_ENGINE"] = "true"
                try:
                    r, _ = detect_all_problems_intramural(wb.active, idx)
                finally:
                    if old is not None:
                        os.environ["USE_RULE_ENGINE"] = old
                    else:
                        os.environ.pop("USE_RULE_ENGINE", None)

        total_keys = {
            "centros_de_costos", "ide_contrato", "revision_cantidad",
        }
        for k in total_keys:
            assert k in r["totales"], f"Engine totales missing: {k}"

    def test_centros_de_costos_format_resilient(self):
        """centros_de_costos formatter must not crash when engine items lack
        centro_actual/centro_deberia (engine output doesn't have these keys)."""
        import os
        from app.services.intramural import centro_costo_intramural as cc_module
        engine_like_output = [
            {"factura": "INTRA-001", "problema": "CC invalid",
             "regla": "#1", "severidad": "error",
             "codigo": "", "procedimiento": ""}
        ]
        original_fn = cc_module.detect_centro_costo_intramural
        cc_module.detect_centro_costo_intramural = lambda ds, idx: engine_like_output

        try:
            from app.services.intramural.detect_all import (
                detect_all_problems_intramural,
            )
            wb, idx = self._build_sheet()
            old = os.environ.pop("USE_RULE_ENGINE", None)
            os.environ["USE_RULE_ENGINE"] = "false"
            try:
                r, _ = detect_all_problems_intramural(wb.active, idx)
            finally:
                if old is not None:
                    os.environ["USE_RULE_ENGINE"] = old
                else:
                    os.environ.pop("USE_RULE_ENGINE", None)
        finally:
            cc_module.detect_centro_costo_intramural = original_fn

        assert "centros_de_costos" in r["problemas"]
        for item in r["problemas"]["centros_de_costos"]:
            assert "centro_actual" in item
            assert "centro_deberia" in item


# ═══════════════════════════════════════════════════════════════════════════
# Registry tests
# ═══════════════════════════════════════════════════════════════════════════

class TestF4Registry:
    """Verify new evaluators are registered in EVALUATOR_REGISTRY."""

    def test_centro_costo_intramural_registered(self):
        assert "centro_costo_intramural" in EVALUATOR_REGISTRY

    def test_revision_cantidad_intramural_registered(self):
        assert "revision_cantidad_intramural" in EVALUATOR_REGISTRY
