"""Strict TDD F5.2 + F5.3 + F5.4 + F5.5: Tests for duplicado engine + bacteriologas.

T-F5.2: collect_group_keys aggregation
T-F5.3: duplicado_id_codigo → engine via GroupEvaluator
T-F5.4: bacteriologas_cronograma toggle integration
T-F5.5: Snapshot tests
"""
from __future__ import annotations

from unittest.mock import MagicMock, patch

from openpyxl import Workbook

from app.services.engine.group_evaluator import GroupEvaluator


# =============================================================================
# T-F5.2: collect_group_keys
# =============================================================================

class TestCollectGroupKeys:
    """Unit tests for _agg_collect_group_keys."""

    def _build_wb(self, headers: list[str], rows: list[list]) -> tuple[Workbook, dict[str, int | None]]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        indices: dict[str, int | None] = {}
        for col_idx, header in enumerate(headers):
            key = header.lower().replace(" ", "_")
            indices[key] = col_idx
        return wb, indices

    def test_collect_group_keys_returns_unique_facturas(self):
        """Collect unique facturas from group rows."""
        headers = ["numero_factura", "codigo"]
        rows = [
            ["F001", "A"],
            ["F002", "A"],
            ["F003", "B"],
        ]
        wb, indices = self._build_wb(headers, rows)
        result = GroupEvaluator._agg_collect_group_keys(
            [2, 3], wb.active, indices, "numero_factura"
        )
        assert sorted(result) == ["F001", "F002"]

    def test_collect_group_keys_dedup_facturas(self):
        """Same factura appearing multiple times → unique only."""
        headers = ["numero_factura", "codigo"]
        rows = [
            ["F001", "A"],
            ["F001", "B"],
            ["F001", "C"],
        ]
        wb, indices = self._build_wb(headers, rows)
        result = GroupEvaluator._agg_collect_group_keys(
            [2, 3, 4], wb.active, indices, "numero_factura"
        )
        assert result == ["F001"]

    def test_collect_group_keys_empty_if_no_field(self):
        """Missing field index → empty list."""
        headers = ["codigo"]
        rows = [["A"], ["B"]]
        wb, indices = self._build_wb(headers, rows)
        result = GroupEvaluator._agg_collect_group_keys(
            [2, 3], wb.active, indices, "numero_factura"
        )
        assert result == []

    def test_collect_group_keys_via_build_group_data(self):
        """Integration test via _build_group_data."""
        headers = ["numero_factura", "codigo"]
        rows = [
            ["F001", "A"],
            ["F002", "A"],
            ["F001", "B"],
        ]
        wb, indices = self._build_wb(headers, rows)
        agg_configs = [
            {"function": "collect_group_keys", "field": "numero_factura", "target": "facturas"},
            {"function": "group_size", "target": "count"},
        ]
        result = GroupEvaluator._build_group_data(
            "F001", [2, 4], wb.active, indices, agg_configs
        )
        assert result["facturas"] == ["F001"]
        assert result["count"] == 2


# =============================================================================
# T-F5.3 + T-F5.4: Engine routing tests
# =============================================================================

class TestEngineRoutingF5:
    """Verify detect_all.py engine toggle routes to correct rules."""

    @patch("app.constants.base.is_rule_engine_enabled", return_value=True)
    @patch("app.services.engine.rule_based_detector.RuleBasedDetector")
    def test_engine_path_calls_duplicado_05(self, mock_rbd, mock_enabled):
        """Engine path should call duplicado_id_codigo_05 rule."""
        from app.services.intramural.detect_all import detect_all_problems_intramural

        mock_results = {
            "duplicado_id_codigo_05": [
                {"factura": "123\t890201\tI10X", "count": 2, "facturas": ["F001", "F002"]}
            ],
            "duplicado_id_codigo_02_lab": [],
            "bacteriologas_cronograma": [],
            "centro_costo_intramural_valido": [],
            "revision_cantidad_intramural": [],
            "valores_decimales": [],
        }
        def _mock_rbd(name, session):
            m = MagicMock()
            m.detect.return_value = mock_results.get(name, [])
            return m
        mock_rbd.side_effect = _mock_rbd

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="numero_factura")
        ws.cell(row=1, column=2, value="codigo_tipo_procedimiento")
        ws.cell(row=1, column=3, value="identificacion")
        ws.cell(row=2, column=1, value="F001")
        ws.cell(row=2, column=2, value="05")
        ws.cell(row=2, column=3, value="123")
        ws.cell(row=3, column=1, value="F002")
        ws.cell(row=3, column=2, value="05")
        ws.cell(row=3, column=3, value="123")
        indices = {
            "numero_factura": 0,
            "codigo_tipo_procedimiento": 1,
            "identificacion": 2,
            "codigo": None,
            "codigo_dx_principal": None,
            "procedimiento": None,
            "responsable_cierra": None,
            "fecha_cierre": None,
            "fec_factura": None,
        }

        result, _ = detect_all_problems_intramural(ws, indices)

        # Verify duplicado_id_codigo_05 was called
        call_names = [c[0][0] for c in mock_rbd.call_args_list]
        assert "duplicado_id_codigo_05" in call_names
        assert "duplicado_id_codigo_02_lab" in call_names

    @patch("app.constants.base.is_rule_engine_enabled", return_value=True)
    @patch("app.services.engine.rule_based_detector.RuleBasedDetector")
    def test_engine_path_calls_bacteriologas_cronograma(self, mock_rbd, mock_enabled):
        """Engine path should call bacteriologas_cronograma rule."""
        from app.services.intramural.detect_all import detect_all_problems_intramural

        mock_rbd_instance = MagicMock()
        mock_rbd_instance.detect.return_value = []
        mock_rbd.side_effect = lambda name, session: mock_rbd_instance

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="numero_factura")
        ws.cell(row=2, column=1, value="F001")
        indices = {
            "numero_factura": 0,
            "codigo_tipo_procedimiento": None,
            "identificacion": None,
            "codigo": None,
            "codigo_dx_principal": None,
            "procedimiento": None,
            "responsable_cierra": None,
            "fecha_cierre": None,
            "fec_factura": None,
        }

        result, _ = detect_all_problems_intramural(ws, indices)

        call_names = [c[0][0] for c in mock_rbd.call_args_list]
        assert "bacteriologas_cronograma" in call_names

    @patch("app.constants.base.is_rule_engine_enabled", return_value=False)
    def test_legacy_path_called_when_engine_disabled(self, mock_enabled):
        """Legacy path uses detect_duplicado_id_codigo."""
        from app.services.intramural.duplicado_id_codigo import detect_duplicado_id_codigo

        headers = ["Número Factura", "Nº Identificación", "Cód. Equivalente CUPS",
                    "Código Tipo Procedimiento", "Procedimiento"]
        rows = [
            ["F001", "123", "X001", "05", "Proc A"],
            ["F002", "123", "X001", "05", "Proc B"],
        ]
        wb = Workbook()
        ws = wb.active
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        for ri, row_data in enumerate(rows, 2):
            for ci, val in enumerate(row_data, 1):
                ws.cell(row=ri, column=ci, value=val)
        indices = {
            "numero_factura": 0,
            "identificacion": 1,
            "codigo": 2,
            "codigo_tipo_procedimiento": 3,
            "procedimiento": 4,
            "codigo_dx_principal": None,
            "responsable_cierra": None,
            "fecha_cierre": None,
            "fec_factura": None,
        }

        result = detect_duplicado_id_codigo(ws, indices)
        assert len(result) == 1
        assert result[0]["identificacion"] == "123"

    def test_keys_present_in_engine_output(self):
        """Engine path produces resultado with all expected keys."""
        from app.services.intramural.detect_all import detect_all_problems_intramural

        with patch("app.constants.base.is_rule_engine_enabled", return_value=True), \
             patch("app.services.engine.rule_based_detector.RuleBasedDetector") as mock_rbd:

            mock_instance = MagicMock()
            mock_instance.detect.return_value = []
            mock_rbd.side_effect = lambda name, session: mock_instance

            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="numero_factura")
            ws.cell(row=1, column=2, value="codigo_tipo_procedimiento")
            ws.cell(row=2, column=1, value="F001")
            ws.cell(row=2, column=2, value="05")
            indices = {
                "numero_factura": 0,
                "codigo_tipo_procedimiento": 1,
                "identificacion": None,
                "codigo": None,
                "codigo_dx_principal": None,
                "procedimiento": None,
                "responsable_cierra": None,
                "fecha_cierre": None,
                "fec_factura": None,
            }

            result, _ = detect_all_problems_intramural(ws, indices)
            problemas = result.get("problemas", {})
            assert "profesionales" in problemas
            assert "duplicado_id_codigo" in problemas
            totales = result.get("totales", {})
            assert "profesionales" in totales
            assert "duplicado_id_codigo" in totales


# =============================================================================
# T-F5.5: Snapshot tests — engine vs legacy structure matching
# =============================================================================

class TestSnapshotF5:
    """Snapshot comparison: engine and legacy produce same structure."""

    def _build_wb(self) -> tuple[Workbook, dict[str, int | None]]:
        """Build a minimal Intramural workbook with all needed columns."""
        wb = Workbook()
        ws = wb.active
        headers = [
            "numero_factura", "codigo_tipo_procedimiento", "identificacion",
            "codigo", "codigo_dx_principal", "procedimiento",
            "responsable_cierra", "laboratorio", "fec_factura",
            "codigo_profesional", "fecha_cierre",
        ]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        indices = {h: i for i, h in enumerate(headers)}
        return wb, indices

    def test_snapshot_engine_path_keys(self):
        """Engine path produces same keys in resultado."""
        from app.services.intramural.detect_all import detect_all_problems_intramural

        with patch("app.constants.base.is_rule_engine_enabled", return_value=True), \
             patch("app.services.engine.rule_based_detector.RuleBasedDetector") as mock_rbd:

            mock_instance = MagicMock()
            mock_instance.detect.return_value = []
            mock_rbd.side_effect = lambda name, session: mock_instance

            wb, indices = self._build_wb()
            result_engine, _ = detect_all_problems_intramural(wb.active, indices)

        with patch("app.constants.base.is_rule_engine_enabled", return_value=False):
            wb2, _ = self._build_wb()
            result_legacy, _ = detect_all_problems_intramural(wb2.active, indices)

        # Both should have same top-level keys
        assert result_engine.get("area") == result_legacy.get("area")
        assert set(result_engine.get("problemas", {}).keys()) == set(result_legacy.get("problemas", {}).keys())
        assert set(result_engine.get("totales", {}).keys()) == set(result_legacy.get("totales", {}).keys())

    def test_snapshot_vacio_engine_vs_legacy(self):
        """Empty data: both paths produce empty lists for migrated detectors."""
        from app.services.intramural.detect_all import detect_all_problems_intramural

        with patch("app.constants.base.is_rule_engine_enabled", return_value=True), \
             patch("app.services.engine.rule_based_detector.RuleBasedDetector") as mock_rbd:

            mock_instance = MagicMock()
            mock_instance.detect.return_value = []
            mock_rbd.side_effect = lambda name, session: mock_instance

            wb, indices = self._build_wb()
            result_engine, _ = detect_all_problems_intramural(wb.active, indices)

        with patch("app.constants.base.is_rule_engine_enabled", return_value=False):
            wb2, _ = self._build_wb()
            result_legacy, _ = detect_all_problems_intramural(wb2.active, indices)

        engine_prof = result_engine["problemas"].get("profesionales", [])
        legacy_prof = result_legacy["problemas"].get("profesionales", [])
        assert len(engine_prof) == len(legacy_prof)

        engine_dup = result_engine["problemas"].get("duplicado_id_codigo", [])
        legacy_dup = result_legacy["problemas"].get("duplicado_id_codigo", [])
        assert len(engine_dup) == len(legacy_dup)
