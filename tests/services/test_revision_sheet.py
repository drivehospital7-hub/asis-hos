"""Tests para servicios de detección de problemas."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.constants import (
    CONVENIO_ASISTENCIAL,
    CONVENIO_PYP,
    PYP_CUPS_CODES,
    REVISION_HEADERS,
    REVISION_SHEET,
    TARGET_PROCEDURES,
)
from app.services.transversales.column_indices import get_column_indices
from app.services.transversales.create_revision_sheet import (
    _write_column,
    create_revision_sheet,
)
from app.services.transversales.decimales import detect_decimales
from app.services.transversales.doble_tipo_procedimiento import detect_doble_tipo_procedimiento
from app.services.transversales.normalize import normalize_header, normalize_invoice
from app.services.transversales.cantidades_anomalas import detect_cantidades_anomalas
from app.services.transversales.ruta_duplicada import detect_ruta_duplicada
from app.services.normalized_rows import build_urgencias_normalized_rows


@pytest.fixture
def workbook_with_invoice_data() -> Workbook:
    """Crea un workbook con datos de facturas para pruebas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Headers (incluye Código para validar convenio)
    headers = [
        "Número Factura",
        "Vlr. Subsidiado",
        "Vlr. Procedimiento",
        "Tipo Procedimiento",
        "Código",
        "Procedimiento",
        "Nº Identificación",
        "Convenio Facturado",
        "Cantidad",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    return wb


def add_invoice_row(
    ws,
    row: int,
    numero: str,
    vlr_sub: float = 1000,
    vlr_proc: float = 500,
    tipo_proc: str = "Consultas",
    codigo: str = "890101",
    procedimiento: str = "Consulta General",
    identificacion: str = "123456",
    convenio: str = CONVENIO_ASISTENCIAL,
    cantidad: int = 1,
) -> None:
    """Helper para agregar una fila de factura."""
    ws.cell(row=row, column=1, value=numero)
    ws.cell(row=row, column=2, value=vlr_sub)
    ws.cell(row=row, column=3, value=vlr_proc)
    ws.cell(row=row, column=4, value=tipo_proc)
    ws.cell(row=row, column=5, value=codigo)
    ws.cell(row=row, column=6, value=procedimiento)
    ws.cell(row=row, column=7, value=identificacion)
    ws.cell(row=row, column=8, value=convenio)
    ws.cell(row=row, column=9, value=cantidad)


class TestNormalizeHeader:
    """Tests para normalize_header."""

    def test_normaliza_a_minusculas(self) -> None:
        """Debe convertir a minúsculas."""
        assert normalize_header("NUMERO FACTURA") == "numero factura"

    def test_elimina_espacios_extra(self) -> None:
        """Debe eliminar espacios al inicio y final."""
        assert normalize_header("  header  ") == "header"

    def test_none_retorna_string_vacio(self) -> None:
        """None debe retornar string vacío."""
        assert normalize_header(None) == ""


class TestNormalizeInvoice:
    """Tests para normalize_invoice."""

    def test_convierte_float_entero_a_string(self) -> None:
        """Float sin decimales debe convertirse a string entero."""
        assert normalize_invoice(12345.0) == "12345"

    def test_mantiene_string(self) -> None:
        """String debe mantenerse (strippeado)."""
        assert normalize_invoice("FAC-001") == "FAC-001"
        assert normalize_invoice("  FAC-002  ") == "FAC-002"

    def test_none_retorna_none(self) -> None:
        """None debe retornar None."""
        assert normalize_invoice(None) is None

    def test_string_vacio_retorna_none(self) -> None:
        """String vacío debe retornar None."""
        assert normalize_invoice("") is None


class TestGetColumnIndices:
    """Tests para get_column_indices."""

    def test_mapea_headers_conocidos(self) -> None:
        """Debe mapear headers conocidos a sus índices."""
        headers = [
            "Número Factura",
            "Vlr. Subsidiado",
            "Vlr. Procedimiento",
            "Tipo Procedimiento",
        ]
        required = {
            "numero_factura": "Número Factura",
            "vlr_subsidiado": "Vlr. Subsidiado",
            "vlr_procedimiento": "Vlr. Procedimiento",
            "tipo_procedimiento": "Tipo Procedimiento",
        }

        indices, _ = get_column_indices(headers, required)

        assert indices["numero_factura"] == 0
        assert indices["vlr_subsidiado"] == 1
        assert indices["vlr_procedimiento"] == 2
        assert indices["tipo_procedimiento"] == 3

    def test_headers_no_encontrados_son_none(self) -> None:
        """Headers no encontrados deben ser None."""
        headers = ["Columna Rara", "Otra Columna"]
        required = {
            "numero_factura": "Número Factura",
            "vlr_subsidiado": "Vlr. Subsidiado",
        }

        indices, _ = get_column_indices(headers, required)

        assert indices["numero_factura"] is None
        assert indices["vlr_subsidiado"] is None

    def test_requiere_coincidencia_exacta(self) -> None:
        """Requiere coincidencia EXACTA, no infiere."""
        headers = ["Número Factura", "Nº Identificación"]
        required = {
            "numero_factura": "Número Factura",
            "identificacion": "Nº Identificación",
        }

        indices, _ = get_column_indices(headers, required)

        assert indices["numero_factura"] == 0
        assert indices["identificacion"] == 1


class TestDetectDecimals:
    """Tests para detect_decimales."""

    def test_detecta_facturas_con_decimales(
        self, workbook_with_invoice_data: Workbook
    ) -> None:
        """Debe detectar facturas con valores decimales."""
        ws = workbook_with_invoice_data.active
        add_invoice_row(ws, 2, "FAC-001", vlr_sub=1000.50)  # Decimal
        add_invoice_row(ws, 3, "FAC-002", vlr_sub=1000.00)  # Entero
        add_invoice_row(ws, 4, "FAC-003", vlr_proc=500.25)  # Decimal en vlr_proc

        required = {"numero_factura": "Número Factura", "vlr_subsidiado": "Vlr. Subsidiado", "vlr_procedimiento": "Vlr. Procedimiento"}
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        indices, _ = get_column_indices(headers, required)

        result = detect_decimales(ws, indices)

        assert len(result) == 2
        assert "FAC-001" in result
        assert "FAC-002" not in result
        assert "FAC-003" in result

    def test_no_duplica_facturas(
        self, workbook_with_invoice_data: Workbook
    ) -> None:
        """No debe duplicar facturas con decimales en ambos campos."""
        ws = workbook_with_invoice_data.active
        add_invoice_row(ws, 2, "FAC-001", vlr_sub=1000.50, vlr_proc=500.25)

        required = {"numero_factura": "Número Factura", "vlr_subsidiado": "Vlr. Subsidiado", "vlr_procedimiento": "Vlr. Procedimiento"}
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        indices, _ = get_column_indices(headers, required)

        result = detect_decimales(ws, indices)

        assert len(result) == 1
        assert result[0] == "FAC-001"


class TestDetectDobleTipoProcedimiento:
    """Tests para detect_doble_tipo_procedimiento."""

    def test_detecta_factura_con_multiples_tipos(
        self, workbook_with_invoice_data: Workbook
    ) -> None:
        """Debe detectar facturas con más de un tipo de procedimiento."""
        ws = workbook_with_invoice_data.active
        add_invoice_row(ws, 2, "FAC-001", tipo_proc="Consultas")
        add_invoice_row(ws, 3, "FAC-001", tipo_proc="Procedimientos")
        add_invoice_row(ws, 4, "FAC-002", tipo_proc="Consultas")
        add_invoice_row(ws, 5, "FAC-002", tipo_proc="Consultas")

        required = {"numero_factura": "Número Factura", "tipo_procedimiento": "Tipo Procedimiento"}
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        indices, _ = get_column_indices(headers, required)

        result = detect_doble_tipo_procedimiento(ws, indices)

        facturas = [r["factura"] for r in result]
        assert "FAC-001" in facturas
        assert "FAC-002" not in facturas


class TestDetectRutaDuplicada:
    """Tests para detect_ruta_duplicada."""

    def test_detecta_paciente_con_multiples_facturas_pyp(
        self, workbook_with_invoice_data: Workbook
    ) -> None:
        """Debe detectar pacientes con >= 3 facturas en PyP."""
        ws = workbook_with_invoice_data.active
        add_invoice_row(ws, 2, "FAC-001", identificacion="PAC-001", convenio=CONVENIO_PYP)
        add_invoice_row(ws, 3, "FAC-002", identificacion="PAC-001", convenio=CONVENIO_PYP)
        add_invoice_row(ws, 4, "FAC-003", identificacion="PAC-001", convenio=CONVENIO_PYP)
        add_invoice_row(ws, 5, "FAC-004", identificacion="PAC-002", convenio=CONVENIO_PYP)
        add_invoice_row(ws, 6, "FAC-005", identificacion="PAC-002", convenio=CONVENIO_PYP)

        required = {
            "numero_factura": "Número Factura",
            "identificacion": "Nº Identificación",
            "convenio_facturado": "Convenio Facturado",
        }
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        indices, _ = get_column_indices(headers, required)

        result = detect_ruta_duplicada(ws, indices)

        identificaciones = [r["identificacion"] for r in result]
        assert "PAC-001" in identificaciones
        assert "PAC-002" not in identificaciones

    def test_ignora_facturas_no_pyp(
        self, workbook_with_invoice_data: Workbook
    ) -> None:
        """Debe ignorar facturas con convenio diferente a PyP."""
        ws = workbook_with_invoice_data.active
        add_invoice_row(ws, 2, "FAC-001", identificacion="PAC-001", convenio=CONVENIO_ASISTENCIAL)
        add_invoice_row(ws, 3, "FAC-002", identificacion="PAC-001", convenio=CONVENIO_ASISTENCIAL)
        add_invoice_row(ws, 4, "FAC-003", identificacion="PAC-001", convenio=CONVENIO_ASISTENCIAL)

        required = {
            "numero_factura": "Número Factura",
            "identificacion": "Nº Identificación",
            "convenio_facturado": "Convenio Facturado",
        }
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        indices, _ = get_column_indices(headers, required)

        result = detect_ruta_duplicada(ws, indices)

        assert len(result) == 0


class TestDetectCantidadesAnomalas:
    """Tests para detect_cantidades_anomalas."""

    def test_detecta_consultas_con_cantidad_mayor_igual_2(
        self, workbook_with_invoice_data: Workbook
    ) -> None:
        """Debe detectar consultas con cantidad >= 2."""
        ws = workbook_with_invoice_data.active
        add_invoice_row(ws, 2, "FAC-001", tipo_proc="Consultas", cantidad=2)
        add_invoice_row(ws, 3, "FAC-002", tipo_proc="Consultas", cantidad=1)

        required = {
            "numero_factura": "Número Factura",
            "tipo_procedimiento": "Tipo Procedimiento",
            "cantidad": "Cantidad",
            "convenio_facturado": "Convenio Facturado",
        }
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        indices, _ = get_column_indices(headers, required)

        result = detect_cantidades_anomalas(ws, indices)

        facturas = [r["factura"] for r in result]
        assert "FAC-001" in facturas
        assert "FAC-002" not in facturas

    def test_detecta_cantidad_mayor_10(
        self, workbook_with_invoice_data: Workbook
    ) -> None:
        """Debe detectar cualquier cantidad > 10."""
        ws = workbook_with_invoice_data.active
        add_invoice_row(ws, 2, "FAC-001", tipo_proc="Otros", cantidad=11)
        add_invoice_row(ws, 3, "FAC-002", tipo_proc="Otros", cantidad=10)

        required = {
            "numero_factura": "Número Factura",
            "tipo_procedimiento": "Tipo Procedimiento",
            "cantidad": "Cantidad",
            "convenio_facturado": "Convenio Facturado",
        }
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        indices, _ = get_column_indices(headers, required)

        result = detect_cantidades_anomalas(ws, indices)

        facturas = [r["factura"] for r in result]
        assert "FAC-001" in facturas
        assert "FAC-002" not in facturas

    def test_detecta_pyp_con_cantidad_mayor_igual_3(
        self, workbook_with_invoice_data: Workbook
    ) -> None:
        """Debe detectar PyP con cantidad >= 3."""
        ws = workbook_with_invoice_data.active
        add_invoice_row(ws, 2, "FAC-001", tipo_proc="Procedimientos", convenio=CONVENIO_PYP, cantidad=3)
        add_invoice_row(ws, 3, "FAC-002", tipo_proc="Procedimientos", convenio=CONVENIO_PYP, cantidad=2)

        required = {
            "numero_factura": "Número Factura",
            "tipo_procedimiento": "Tipo Procedimiento",
            "cantidad": "Cantidad",
            "convenio_facturado": "Convenio Facturado",
        }
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        indices, _ = get_column_indices(headers, required)

        result = detect_cantidades_anomalas(ws, indices)

        facturas = [r["factura"] for r in result]
        assert "FAC-001" in facturas
        assert "FAC-002" not in facturas


class TestWriteColumn:
    """Tests para _write_column."""

    def test_escribe_valores_en_columna(self) -> None:
        """Debe escribir valores en una columna empezando en la fila indicada."""
        wb = Workbook()
        ws = wb.active
        values = ["A", "B", "C"]

        _write_column(ws, 2, values, start_row=3)

        assert ws.cell(row=3, column=2).value == "A"
        assert ws.cell(row=4, column=2).value == "B"
        assert ws.cell(row=5, column=2).value == "C"


class TestBuildUrgenciasNormalizedRows:
    """Tests para build_urgencias_normalized_rows."""

    def test_centros_de_costo(self) -> None:
        """Debe normalizar centros de costo."""
        rows = build_urgencias_normalized_rows(
            problemas_centros=[{
                "factura": "F001",
                "codigo": "890405",
                "procedimiento": "CONSULTA",
                "centro_actual": "URGENCIAS",
                "centro_deberia": "APOYO DIAGNOSTICO",
            }],
            problemas_ide_contrato=[],
            problemas_cups_equivalentes=[],
            mal_capitado=[],
            cantidades_urgencias=[],
            cantidades_soat_urgencias=[],
            cantidades_hospitalizacion=[],
            cantidades_soat_hospitalizacion=[],
            responsables_map={"F001": "JUAN"},
        )

        assert len(rows) == 1
        r = rows[0]
        assert r["tipo_error"] == "Centros de Costo"
        assert r["factura"] == "F001"
        assert r["responsable_cierra"] == "JUAN"
        assert r["descripcion"] == "Centro de costo debería ser APOYO DIAGNOSTICO"
        assert r["procedimiento"] == "890405 - CONSULTA"
        assert r["detalle"] == "URGENCIAS"

    def test_ide_contrato(self) -> None:
        """Debe normalizar IDE Contrato."""
        rows = build_urgencias_normalized_rows(
            problemas_centros=[],
            problemas_ide_contrato=[{
                "factura": "F002",
                "codigo": "861801",
                "procedimiento": "CONSULTA ENFERMERIA",
                "ide_contrato_actual": "900",
                "ide_contrato_deberia": "977",
            }],
            problemas_cups_equivalentes=[],
            mal_capitado=[],
            cantidades_urgencias=[],
            cantidades_soat_urgencias=[],
            cantidades_hospitalizacion=[],
            cantidades_soat_hospitalizacion=[],
            responsables_map={},
        )

        assert len(rows) == 1
        r = rows[0]
        assert r["tipo_error"] == "IDE Contrato"
        assert r["factura"] == "F002"
        assert r["descripcion"] == "IDE Contrato debería ser 977"
        assert r["procedimiento"] == "861801 - CONSULTA ENFERMERIA"
        assert r["detalle"] == "900"

    def test_cantidades_urgencias(self) -> None:
        """Debe normalizar cantidades de Urgencias."""
        rows = build_urgencias_normalized_rows(
            problemas_centros=[],
            problemas_ide_contrato=[],
            problemas_cups_equivalentes=[],
            mal_capitado=[],
            cantidades_urgencias=[{
                "factura": "F003",
                "codigo": "890601",
                "procedimiento": "MANEJO INTRAHOSP",
                "cantidad": 3,
            }],
            cantidades_soat_urgencias=[],
            cantidades_hospitalizacion=[],
            cantidades_soat_hospitalizacion=[],
            responsables_map={},
        )

        assert len(rows) == 1
        r = rows[0]
        assert r["tipo_error"] == "Cantidades"
        assert r["factura"] == "F003"
        assert r["descripcion"] == "Cantidad 3 debe ser ≤ 1 en Urgencias"
        assert r["procedimiento"] == "890601 - MANEJO INTRAHOSP"
        assert r["detalle"] == "3"


class TestCreateRevisionSheetOdontologia:
    """Tests para create_revision_sheet con área Odontología."""

    def test_crea_hoja_revision(self, workbook_with_invoice_data: Workbook) -> None:
        """Debe crear hoja Revision con headers correctos."""
        wb = workbook_with_invoice_data
        ws = wb.active
        add_invoice_row(ws, 2, "FAC-001")

        result = create_revision_sheet(wb, ws)

        assert result["rule"] == "create_revision_sheet"
        assert result["sheet"] == REVISION_SHEET
        assert REVISION_SHEET in wb.sheetnames

    def test_headers_de_odontologia(self, workbook_with_invoice_data: Workbook) -> None:
        """Headers deben ser los de Odontología."""
        wb = workbook_with_invoice_data
        ws = wb.active
        add_invoice_row(ws, 2, "FAC-001")

        result = create_revision_sheet(wb, ws)

        assert result["headers"] == list(REVISION_HEADERS.values())

    def test_resultado_tiene_key_problemas(self, workbook_with_invoice_data: Workbook) -> None:
        """Debe retornar dict con key problemas."""
        wb = workbook_with_invoice_data
        ws = wb.active
        add_invoice_row(ws, 2, "FAC-001")

        result = create_revision_sheet(wb, ws)

        assert "problemas" in result
