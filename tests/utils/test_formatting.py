"""Tests para app/utils/formatting.py."""

from __future__ import annotations

import pytest
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule

from app.utils.formatting import (
    find_column_letter_by_header,
    create_fill,
    apply_conditional_convenio_facturado,
    apply_conditional_tipo_identificacion,
    apply_conditional_cruce_facturas,
    apply_all_conditional_formatting,
)
from app.constants import (
    COLOR_GREEN,
    COLOR_RED,
    COLOR_YELLOW,
    CRUCE_FACTURAS_SHEET,
)


@pytest.fixture
def workbook_with_data_headers() -> Workbook:
    """Crea un workbook con headers de datos para formato condicional."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    
    # Headers necesarios para las reglas de formato
    headers = [
        "Número Factura",
        "Entidad Cobrar",
        "Convenio Facturado",
        "Centro Costo",
        "Tipo Identificación",
        "Fec. Nacimiento",
        "Fec. Factura",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos de ejemplo
    for row in range(2, 5):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col, value=f"Dato-{row}-{col}")
    
    return wb


@pytest.fixture
def workbook_with_cruce_and_data() -> Workbook:
    """Crea un workbook con hoja de datos y hoja CruceFacturas."""
    wb = Workbook()
    
    # Hoja de datos
    data_sheet = wb.active
    data_sheet.title = "Datos"
    headers = ["Número Factura", "Convenio Facturado", "Centro Costo"]
    for col, header in enumerate(headers, start=1):
        data_sheet.cell(row=1, column=col, value=header)
    
    # Datos
    data_sheet.cell(row=2, column=1, value="FAC-001")
    data_sheet.cell(row=3, column=1, value="FAC-002")
    data_sheet.cell(row=4, column=1, value="FAC-003")
    
    # Hoja CruceFacturas
    cruce_sheet = wb.create_sheet(title=CRUCE_FACTURAS_SHEET)
    cruce_sheet["B1"] = "Facturas Ok"
    cruce_sheet["D1"] = "Facturas Pendientes"
    cruce_sheet["F1"] = "PDFs de Facturas"
    
    return wb


class TestFindColumnLetterByHeader:
    """Tests para la función find_column_letter_by_header."""

    def test_encuentra_columna_existente(
        self, workbook_with_data_headers: Workbook
    ) -> None:
        """Debe encontrar la letra de columna para un header existente."""
        ws = workbook_with_data_headers.active
        
        result = find_column_letter_by_header(ws, "Número Factura")
        
        assert result == "A"

    def test_encuentra_columna_en_posicion_media(
        self, workbook_with_data_headers: Workbook
    ) -> None:
        """Debe encontrar columna en cualquier posición."""
        ws = workbook_with_data_headers.active
        
        result = find_column_letter_by_header(ws, "Convenio Facturado")
        
        assert result == "C"

    def test_retorna_none_si_no_existe(
        self, workbook_with_data_headers: Workbook
    ) -> None:
        """Debe retornar None si el header no existe."""
        ws = workbook_with_data_headers.active
        
        result = find_column_letter_by_header(ws, "Header Inexistente")
        
        assert result is None

    def test_respeta_headers_row_custom(self) -> None:
        """Debe buscar en la fila especificada por headers_row."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "No es el header"
        ws["A2"] = "Header Real"
        
        result = find_column_letter_by_header(ws, "Header Real", headers_row=2)
        
        assert result == "A"

    def test_hoja_vacia_retorna_none(self) -> None:
        """Hoja vacía debe retornar None."""
        wb = Workbook()
        ws = wb.active
        
        result = find_column_letter_by_header(ws, "Cualquiera")
        
        assert result is None


class TestCreateFill:
    """Tests para la función create_fill."""

    def test_crea_fill_con_color_correcto(self) -> None:
        """Debe crear un PatternFill con el color especificado."""
        fill = create_fill(COLOR_GREEN)
        
        assert fill.start_color.rgb == f"00{COLOR_GREEN}"
        assert fill.end_color.rgb == f"00{COLOR_GREEN}"
        assert fill.fill_type == "solid"

    @pytest.mark.parametrize("color", [COLOR_GREEN, COLOR_YELLOW, COLOR_RED])
    def test_crea_fill_para_todos_los_colores(self, color: str) -> None:
        """Debe crear fills para todos los colores del sistema."""
        fill = create_fill(color)
        
        assert fill.fill_type == "solid"
        assert color in fill.start_color.rgb


class TestApplyConditionalConvenioFacturado:
    """Tests para apply_conditional_convenio_facturado."""

    def test_aplica_regla_cuando_columnas_existen(
        self, workbook_with_data_headers: Workbook
    ) -> None:
        """Debe aplicar la regla cuando todas las columnas requeridas existen."""
        ws = workbook_with_data_headers.active
        
        result = apply_conditional_convenio_facturado(ws)
        
        assert result["rule"] == "convenio_facturado_conditional"
        assert result["applied"] is True
        # Verificar que se agregó formato condicional
        assert len(ws.conditional_formatting._cf_rules) > 0

    def test_no_aplica_si_faltan_columnas(self) -> None:
        """Debe retornar applied=False si faltan columnas requeridas."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Columna Incompleta"
        
        result = apply_conditional_convenio_facturado(ws)
        
        assert result["rule"] == "convenio_facturado_conditional"
        assert result["applied"] is False


class TestApplyConditionalTipoIdentificacion:
    """Tests para apply_conditional_tipo_identificacion."""

    def test_aplica_regla_cuando_columnas_existen(
        self, workbook_with_data_headers: Workbook
    ) -> None:
        """Debe aplicar la regla cuando todas las columnas requeridas existen."""
        ws = workbook_with_data_headers.active
        
        result = apply_conditional_tipo_identificacion(ws)
        
        assert result["rule"] == "tipo_identificacion_conditional"
        assert result["applied"] is True

    def test_no_aplica_si_faltan_columnas(self) -> None:
        """Debe retornar applied=False si faltan columnas."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Solo un header"
        
        result = apply_conditional_tipo_identificacion(ws)
        
        assert result["rule"] == "tipo_identificacion_conditional"
        assert result["applied"] is False


class TestApplyConditionalCruceFacturas:
    """Tests para apply_conditional_cruce_facturas."""

    def test_aplica_formato_a_ambas_hojas(
        self, workbook_with_cruce_and_data: Workbook
    ) -> None:
        """Debe aplicar formato condicional a CruceFacturas y hoja de datos."""
        wb = workbook_with_cruce_and_data
        cruce_sheet = wb[CRUCE_FACTURAS_SHEET]
        data_sheet = wb["Datos"]
        
        result = apply_conditional_cruce_facturas(
            cruce_sheet, data_sheet, numero_factura_col="A"
        )
        
        assert result["rule"] == "cruce_facturas_conditional"
        assert result["applied"] is True
        assert result["cruce_columns"] == ["B", "D", "F"]
        
        # Verificar que se aplicaron reglas a ambas hojas
        assert len(cruce_sheet.conditional_formatting._cf_rules) > 0
        assert len(data_sheet.conditional_formatting._cf_rules) > 0


class TestApplyAllConditionalFormatting:
    """Tests para apply_all_conditional_formatting."""

    def test_aplica_todas_las_reglas(
        self, workbook_with_cruce_and_data: Workbook
    ) -> None:
        """Debe aplicar todas las reglas de formato condicional."""
        wb = workbook_with_cruce_and_data
        cruce_sheet = wb[CRUCE_FACTURAS_SHEET]
        data_sheet = wb["Datos"]
        
        # Agregar headers necesarios para otras reglas
        data_sheet["D1"] = "Entidad Cobrar"
        data_sheet["E1"] = "Tipo Identificación"
        data_sheet["F1"] = "Fec. Nacimiento"
        data_sheet["G1"] = "Fec. Factura"
        
        results = apply_all_conditional_formatting(cruce_sheet, data_sheet)
        
        assert isinstance(results, list)
        assert len(results) >= 1  # Al menos cruce_facturas

    def test_retorna_error_si_falta_numero_factura(self) -> None:
        """Debe retornar error si no existe columna Número Factura."""
        wb = Workbook()
        cruce_sheet = wb.create_sheet(title=CRUCE_FACTURAS_SHEET)
        data_sheet = wb.active
        data_sheet["A1"] = "Otra Columna"
        
        results = apply_all_conditional_formatting(cruce_sheet, data_sheet)
        
        assert len(results) == 1
        assert results[0]["applied"] is False
        assert results[0]["reason"] == "missing_column"
