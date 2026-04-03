"""Tests para app/utils/column_filter.py."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.utils.column_filter import (
    get_column_headers,
    hide_non_relevant_columns,
    filter_columns,
    unmerge_header_rows,
    delete_header_rows,
)


@pytest.fixture
def workbook_with_headers() -> Workbook:
    """Crea un workbook con headers en fila 1."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # Headers en fila 1
    headers = ["Número Factura", "Vlr. Procedimiento", "Convenio Facturado", "Otra Columna"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos de ejemplo en filas 2-3
    for row in range(2, 4):
        for col in range(1, 5):
            ws.cell(row=row, column=col, value=f"Dato-{row}-{col}")
    
    return wb


@pytest.fixture
def workbook_with_merged_cells() -> Workbook:
    """Crea un workbook con celdas combinadas en las primeras filas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MergedSheet"
    
    # Celdas combinadas en filas 1-2
    ws.merge_cells("A1:B1")
    ws.merge_cells("C1:D1")
    ws["A1"] = "Header Combinado 1"
    ws["C1"] = "Header Combinado 2"
    
    # Headers reales en fila 3
    headers = ["Número Factura", "Vlr. Procedimiento", "Convenio Facturado", "Extra"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=header)
    
    # Datos en filas 4+
    for row in range(4, 6):
        for col in range(1, 5):
            ws.cell(row=row, column=col, value=f"Dato-{row}-{col}")
    
    return wb


class TestGetColumnHeaders:
    """Tests para la función get_column_headers."""

    def test_extrae_headers_de_primera_fila(self, workbook_with_headers: Workbook) -> None:
        """Debe extraer todos los headers de la fila 1."""
        ws = workbook_with_headers.active
        
        headers = get_column_headers(ws)
        
        assert len(headers) == 4
        assert headers[0] == "Número Factura"
        assert headers[1] == "Vlr. Procedimiento"
        assert headers[2] == "Convenio Facturado"
        assert headers[3] == "Otra Columna"

    def test_hoja_vacia_retorna_lista_vacia(self) -> None:
        """Hoja sin datos debe retornar lista vacía."""
        wb = Workbook()
        ws = wb.active
        
        headers = get_column_headers(ws)
        
        # max_column es 1 para hojas vacías, pero la celda A1 es None
        assert headers == [None]

    def test_headers_con_none_se_incluyen(self) -> None:
        """Headers con celdas None deben incluirse en el resultado."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Header1"
        # B1 está vacío (None)
        ws["C1"] = "Header3"
        
        headers = get_column_headers(ws)
        
        assert headers[0] == "Header1"
        assert headers[1] is None
        assert headers[2] == "Header3"


class TestUnmergeHeaderRows:
    """Tests para la función unmerge_header_rows."""

    def test_desune_celdas_combinadas_en_primeras_filas(
        self, workbook_with_merged_cells: Workbook
    ) -> None:
        """Debe desunir celdas combinadas en las primeras N filas."""
        ws = workbook_with_merged_cells.active
        
        # Antes: hay 2 rangos combinados
        assert len(list(ws.merged_cells)) == 2
        
        unmerged_count = unmerge_header_rows(ws, rows_to_check=2)
        
        # Después: deben estar desunidas
        assert unmerged_count == 2
        assert len(list(ws.merged_cells)) == 0

    def test_no_desune_celdas_fuera_del_rango(self) -> None:
        """No debe desunir celdas combinadas fuera del rango especificado."""
        wb = Workbook()
        ws = wb.active
        
        # Celda combinada en fila 5 (fuera del rango)
        ws.merge_cells("A5:B5")
        ws["A5"] = "Combinado fuera de rango"
        
        unmerged_count = unmerge_header_rows(ws, rows_to_check=2)
        
        assert unmerged_count == 0
        assert len(list(ws.merged_cells)) == 1

    def test_hoja_sin_celdas_combinadas_retorna_cero(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Hoja sin celdas combinadas debe retornar 0."""
        ws = workbook_with_headers.active
        
        unmerged_count = unmerge_header_rows(ws)
        
        assert unmerged_count == 0


class TestDeleteHeaderRows:
    """Tests para la función delete_header_rows."""

    def test_elimina_primeras_n_filas(self, workbook_with_headers: Workbook) -> None:
        """Debe eliminar las primeras N filas."""
        ws = workbook_with_headers.active
        
        # Antes: fila 1 tiene headers
        assert ws.cell(row=1, column=1).value == "Número Factura"
        
        delete_header_rows(ws, rows_to_delete=1)
        
        # Después: fila 1 tiene los datos que estaban en fila 2
        assert ws.cell(row=1, column=1).value == "Dato-2-1"

    def test_elimina_dos_filas_por_defecto(self) -> None:
        """Por defecto debe eliminar 2 filas."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Fila 1"
        ws["A2"] = "Fila 2"
        ws["A3"] = "Fila 3"
        
        delete_header_rows(ws)  # default: 2
        
        # Fila 3 ahora es fila 1
        assert ws.cell(row=1, column=1).value == "Fila 3"


class TestHideNonRelevantColumns:
    """Tests para la función hide_non_relevant_columns."""

    def test_oculta_columnas_no_en_columns_to_keep(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe ocultar columnas que no están en columns_to_keep."""
        ws = workbook_with_headers.active
        
        # Solo mantener "Número Factura"
        result = hide_non_relevant_columns(
            ws, columns_to_keep=frozenset({"Número Factura"})
        )
        
        assert result["kept_count"] == 1
        assert result["hidden_count"] == 3
        assert "Número Factura" in result["kept_columns"]
        
        # Verificar que columna A no está oculta y las demás sí
        assert ws.column_dimensions["A"].hidden is False
        assert ws.column_dimensions["B"].hidden is True
        assert ws.column_dimensions["C"].hidden is True
        assert ws.column_dimensions["D"].hidden is True

    def test_usa_columns_to_keep_por_defecto(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Sin parámetro, debe usar COLUMNS_TO_KEEP de constants."""
        ws = workbook_with_headers.active
        
        result = hide_non_relevant_columns(ws)
        
        # Solo las columnas en COLUMNS_TO_KEEP deben mantenerse
        # "Número Factura", "Vlr. Procedimiento", "Convenio Facturado" están en COLUMNS_TO_KEEP
        # "Otra Columna" no está
        assert result["kept_count"] == 3
        assert result["hidden_count"] == 1

    def test_sin_columnas_coincidentes_warning(self) -> None:
        """Si ninguna columna coincide, debe retornar kept_count=0."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "ColumnaRara"
        ws["B1"] = "OtraRara"
        
        result = hide_non_relevant_columns(
            ws, columns_to_keep=frozenset({"NoExiste"})
        )
        
        assert result["kept_count"] == 0
        assert result["hidden_count"] == 2


class TestFilterColumns:
    """Tests para la función principal filter_columns."""

    def test_orquesta_filtrado_completo(
        self, workbook_with_merged_cells: Workbook
    ) -> None:
        """Debe orquestar desunir, eliminar filas y ocultar columnas."""
        ws = workbook_with_merged_cells.active
        
        # Antes: hay celdas combinadas y headers en fila 3
        assert len(list(ws.merged_cells)) == 2
        
        result = filter_columns(
            ws, 
            columns_to_keep=frozenset({"Número Factura"}),
            delete_first_rows=2,
        )
        
        # Después de procesar:
        assert result["sheet"] == "MergedSheet"
        assert result["rows_deleted"] == 2
        assert result["kept_count"] == 1
        # Headers ahora están en fila 1 (después de eliminar 2 filas)
        assert ws.cell(row=1, column=1).value == "Número Factura"

    def test_sin_eliminar_filas_si_cero(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Si delete_first_rows=0, no debe eliminar filas."""
        ws = workbook_with_headers.active
        header_original = ws.cell(row=1, column=1).value
        
        result = filter_columns(
            ws,
            columns_to_keep=frozenset({"Número Factura"}),
            delete_first_rows=0,
        )
        
        assert result["rows_deleted"] == 0
        assert ws.cell(row=1, column=1).value == header_original

    def test_retorna_informacion_completa(
        self, workbook_with_headers: Workbook
    ) -> None:
        """Debe retornar dict con toda la información del procesamiento."""
        ws = workbook_with_headers.active
        
        result = filter_columns(ws, delete_first_rows=0)
        
        assert "sheet" in result
        assert "rows_deleted" in result
        assert "kept_count" in result
        assert "hidden_count" in result
        assert "kept_columns" in result
