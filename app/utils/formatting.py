"""Estilos para hojas Excel (headers, filas de datos, ajuste de ancho).

Actualmente solo contiene helpers de estilo usados por create_revision_sheet.py.
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    HEADER_BACKGROUND_COLOR,
    HEADER_BORDER_COLOR,
    DATA_ROW_BACKGROUND_COLOR,
    URGENCIA_HEADER_BACKGROUND_COLOR,
    URGENCIA_HEADER_BORDER_COLOR,
    URGENCIA_DATA_ROW_BACKGROUND_COLOR,
)

logger = logging.getLogger(__name__)


def create_header_style() -> dict:
    """
    Crea un diccionario de estilos para encabezados.
    
    Returns:
        Dict con Font, PatternFill y Border configurados
    """
    from openpyxl.styles import Alignment
    
    # Negrita
    font = Font(bold=True)
    
    # Color de fondo azulado
    fill = PatternFill(
        start_color=HEADER_BACKGROUND_COLOR,
        end_color=HEADER_BACKGROUND_COLOR,
        fill_type="solid",
    )
    
    # Borde azulado
    side = Side(color=HEADER_BORDER_COLOR, style="thin")
    border = Border(left=side, right=side, top=side, bottom=side)
    
    # Alineación a la izquierda
    alignment = Alignment(horizontal="left", vertical="center")
    
    return {
        "font": font,
        "fill": fill,
        "border": border,
        "alignment": alignment,
    }


def create_data_row_style() -> dict:
    """
    Crea un diccionario de estilos para filas de datos (sin negrita).
    
    Returns:
        Dict con PatternFill y Border configurados
    """
    from openpyxl.styles import Alignment
    
    # Color de fondo azulado muy claro (sin negrita)
    fill = PatternFill(
        start_color=DATA_ROW_BACKGROUND_COLOR,
        end_color=DATA_ROW_BACKGROUND_COLOR,
        fill_type="solid",
    )
    
    # Borde azulado claro
    side = Side(color="B4C7E7", style="thin")
    border = Border(left=side, right=side, top=side, bottom=side)
    
    # Alineación a la izquierda
    alignment = Alignment(horizontal="left", vertical="center")
    
    return {
        "fill": fill,
        "border": border,
        "alignment": alignment,
    }


def create_urgencia_header_style() -> dict:
    """
    Crea un diccionario de estilos para headers de Revision Urgencias.
    
    Características:
    - Negrita
    - Fondo rojo claro
    - Bordes rojos
    
    Returns:
        Dict con Font, PatternFill y Border configurados
    """
    from openpyxl.styles import Alignment
    
    # Negrita
    font = Font(bold=True)
    
    # Color de fondo rojo claro
    fill = PatternFill(
        start_color=URGENCIA_HEADER_BACKGROUND_COLOR,
        end_color=URGENCIA_HEADER_BACKGROUND_COLOR,
        fill_type="solid",
    )
    
    # Borde rojo
    side = Side(color=URGENCIA_HEADER_BORDER_COLOR, style="thin")
    border = Border(left=side, right=side, top=side, bottom=side)
    
    # Alineación a la izquierda
    alignment = Alignment(horizontal="left", vertical="center")
    
    return {
        "font": font,
        "fill": fill,
        "border": border,
        "alignment": alignment,
    }


def create_urgencia_data_row_style() -> dict:
    """
    Crea un diccionario de estilos para filas de datos de Revision Urgencias.
    
    Características:
    - Fondo rojo claro
    - Bordes rojos
    
    Returns:
        Dict con PatternFill y Border configurados
    """
    from openpyxl.styles import Alignment
    
    # Color de fondo rojo muy claro
    fill = PatternFill(
        start_color=URGENCIA_DATA_ROW_BACKGROUND_COLOR,
        end_color=URGENCIA_DATA_ROW_BACKGROUND_COLOR,
        fill_type="solid",
    )
    
    # Borde rojo
    side = Side(color=URGENCIA_HEADER_BORDER_COLOR, style="thin")
    border = Border(left=side, right=side, top=side, bottom=side)
    
    # Alineación a la izquierda
    alignment = Alignment(horizontal="left", vertical="center")
    
    return {
        "fill": fill,
        "border": border,
        "alignment": alignment,
    }


def auto_adjust_column_width(sheet: Worksheet, max_rows: int = 10) -> dict[str, int]:
    """
    Ajusta el ancho de las columnas según el contenido de las celdas.
    
    Args:
        sheet: Hoja de Excel
        max_rows: Número de filas a evaluar para calcular el ancho máximo
    
    Returns:
        Dict con letra de columna -> ancho ajustado
    """
    from openpyxl.utils import get_column_letter
    
    column_widths = {}
    
    for col in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col)
        max_length = 0
        
        for row in range(1, min(max_rows + 1, sheet.max_row + 1)):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                cell_length = len(str(cell.value))
                # Ajustar por caracteres chinos/unicode
                max_length = max(max_length, cell_length)
        
        # Ancho con padding (mínimo 8, máximo 50)
        if max_length > 0:
            adjusted_width = min(max(max_length + 2, 8), 50)
            sheet.column_dimensions[col_letter].width = adjusted_width
            column_widths[col_letter] = adjusted_width
    
    logger.debug("Anchos de columnas ajustados: %s", column_widths)
    return column_widths



