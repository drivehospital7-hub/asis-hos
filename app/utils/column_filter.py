"""Filtrado de columnas para hojas Excel.

Este módulo se encarga de filtrar/ocultar columnas que no están en la lista
de columnas a mantener visibles (COLUMNS_TO_KEEP).
"""

from __future__ import annotations

import logging
import unicodedata
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import COLUMNS_TO_KEEP

logger = logging.getLogger(__name__)


def normalize_column_name(name: str) -> str:
    """Normaliza un nombre de columna para comparación flexible.
    
    Convierte a minúsculas, elimina acentos y espacios extras.
    Ej: "NÚMERO FACTURA" -> "numero factura"
    """
    if name is None:
        return ""
    # Convertir a minúsculas
    name = str(name).strip().lower()
    # Eliminar acentos (ñ -> n, ú -> u, etc.)
    normalized = unicodedata.normalize('NFD', name)
    return ''.join(c for c in normalized if not unicodedata.combining(c))


def unmerge_header_rows(sheet: Worksheet, rows_to_check: int = 2) -> int:
    """
    Desune celdas combinadas en las primeras filas para permitir su eliminación.
    """
    unmerged_count = 0
    for merged_range in list(sheet.merged_cells):
        if merged_range.min_row <= rows_to_check:
            sheet.unmerge_cells(str(merged_range))
            unmerged_count += 1
    return unmerged_count


def get_column_headers(sheet: Worksheet) -> list[Any]:
    """Extrae los headers de la primera fila de la hoja."""
    return [
        sheet.cell(row=1, column=col).value
        for col in range(1, sheet.max_column + 1)
    ]


def delete_header_rows(sheet: Worksheet, rows_to_delete: int = 2) -> None:
    """Elimina las primeras N filas de la hoja."""
    sheet.delete_rows(1, rows_to_delete)


def hide_non_relevant_columns(
    sheet: Worksheet,
    columns_to_keep: frozenset[str] | None = None,
) -> dict[str, Any]:
    """
    Oculta columnas que no están en la lista de columnas a mantener.
    
    Args:
        sheet: Hoja de Excel con headers en fila 1
        columns_to_keep: Set de nombres de columnas a mantener visibles.
                        Si es None, NO oculta ninguna columna.
    
    Returns:
        Dict con información sobre columnas procesadas
    """
    # Si es None, no ocultar columnas
    if columns_to_keep is None:
        logger.info("No se ocultan columnas (columns_to_keep es None)")
        return {
            "kept_count": sheet.max_column,
            "hidden_count": 0,
            "kept_columns": [],
        }
    
    # Normalizar columns_to_keep para comparación flexible
    normalized_keep = {normalize_column_name(col) for col in columns_to_keep}
    
    headers = get_column_headers(sheet)
    logger.info("Headers encontrados (primeros 20): %s", headers[:20])
    logger.info("Columnas a mantener (normalizadas): %s", normalized_keep)
    
    # Encontrar índices de columnas a mantener (0-based) - comparación flexible
    indices_to_keep = {
        i for i, header in enumerate(headers)
        if normalize_column_name(header) in normalized_keep
    }
    
    logger.info("Índices de columnas a mantener: %s", sorted(indices_to_keep))
    
    hidden_columns = []
    kept_columns = []
    
    for col_idx, header in enumerate(headers):
        col_letter = sheet.cell(row=1, column=col_idx + 1).column_letter
        
        if col_idx not in indices_to_keep:
            sheet.column_dimensions[col_letter].hidden = True
            hidden_columns.append(header)
            logger.debug("Columna ocultada: %s (%s)", col_letter, header)
        else:
            kept_columns.append(header)
    
    if not indices_to_keep:
        logger.warning(
            "No se encontraron columnas coincidentes en hoja %s",
            sheet.title,
        )
    else:
        logger.info(
            "Columnas mantenidas: %d, ocultadas: %d",
            len(kept_columns),
            len(hidden_columns),
        )
    
    return {
        "kept_count": len(kept_columns),
        "hidden_count": len(hidden_columns),
        "kept_columns": kept_columns,
    }


def filter_columns(
    sheet: Worksheet,
    columns_to_keep: frozenset[str] | None = None,
    delete_first_rows: int = 2,
) -> dict[str, Any]:
    """
    Filtra una hoja Excel: elimina filas de encabezado y oculta columnas no relevantes.
    
    Esta es la función principal que orquesta:
    1. Desunir celdas combinadas en las primeras filas
    2. Eliminar las primeras N filas
    3. Ocultar columnas que no están en la lista
    
    Args:
        sheet: Hoja de Excel a procesar
        columns_to_keep: Set de nombres de columnas a mantener visibles.
                        Si es None, usa COLUMNS_TO_KEEP de constants.
        delete_first_rows: Número de filas iniciales a eliminar (por defecto 2)
    
    Returns:
        Dict con información del procesamiento
    """
    # 1. Desunir celdas combinadas
    unmerge_header_rows(sheet, rows_to_check=delete_first_rows)
    
    # 2. Eliminar filas
    if delete_first_rows > 0:
        delete_header_rows(sheet, rows_to_delete=delete_first_rows)
    
    # 3. Ocultar columnas no relevantes
    result = hide_non_relevant_columns(sheet, columns_to_keep)
    
    return {
        "sheet": sheet.title,
        "rows_deleted": delete_first_rows,
        **result,
    }
