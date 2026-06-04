"""Detector de profesionales de equipos básicos no válidos.

Extraído de revision_sheet.py._detect_profesionales_equipos_basicos.

Reglas (Equipos Básicos):
- "Código Profesional" DEBE estar en PROFESIONALES_EQUIPOS_BASICOS
- HIGIENISTA: Solo puede usar códigos en PYP_CODES_HIGIENISTA
- ODONTOLOGO: Puede usar cualquier código EXCEPTO los de PYP_CODES_HIGIENISTA
"""

from __future__ import annotations

import logging

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    PROFESIONALES_EQUIPOS_BASICOS,
    PYP_CODES_HIGIENISTA,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_profesionales_equipos_basicos(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con profesionales no válidos o procedimientos no permitidos.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo_profesional", "nombre",
        "tipo", "problema"
    """
    num_fact_idx = indices["numero_factura"]
    cod_prof_idx = indices["codigo_profesional"]
    codigo_idx = indices["codigo"]

    if None in (num_fact_idx, cod_prof_idx) or codigo_idx is None:
        return []

    problemas: list[dict[str, str]] = []
    facturas_procesadas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_procesadas:
            continue

        cod_profesional = data_sheet.cell(row=row, column=cod_prof_idx + 1).value
        cod_profesional_str = str(cod_profesional).strip() if cod_profesional else ""

        if not cod_profesional_str:
            continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip() if codigo else ""

        # Buscar profesional en el diccionario
        profesional_info = PROFESIONALES_EQUIPOS_BASICOS.get(cod_profesional_str)

        if profesional_info is None:
            problemas.append({
                "factura": factura_str,
                "codigo_profesional": cod_profesional_str,
                "nombre": "",
                "tipo": "",
                "problema": "Profesional no existe en el listado de Equipos Básicos",
            })
            facturas_procesadas.add(factura_str)
        elif profesional_info.get("tipo") == "HIGIENISTA":
            # Higienista: solo puede usar códigos de PYP_CODES_HIGIENISTA
            if codigo_str and codigo_str not in PYP_CODES_HIGIENISTA:
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "HIGIENISTA",
                    "problema": f"Higienista con código no permitido ({codigo_str})",
                })
                facturas_procesadas.add(factura_str)
        elif profesional_info.get("tipo") == "ODONTOLOGO":
            # Odontólogo: no puede usar códigos de PYP_CODES_HIGIENISTA (excepto P0000011)
            if codigo_str and codigo_str in PYP_CODES_HIGIENISTA and codigo_str != "P0000011":
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "ODONTOLOGO",
                    "problema": f"Odontólogo con código PYP no permitido ({codigo_str})",
                })
                facturas_procesadas.add(factura_str)

    return problemas
