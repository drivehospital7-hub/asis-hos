"""Orquestador de detección de problemas para Equipos Básicos.

Agrupa detectores transversales y específicos de equipos básicos.
Reutiliza detectores de odontología cuando aplican (IDE Contrato, Centro Costo).
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    AREA_EQUIPOS_BASICOS,
    CENTRO_COSTO_EQUIPOS_BASICOS,
    EQUIPOS_BASICOS_CANTIDAD_CONSULTAS_MIN,
    EQUIPOS_BASICOS_CANTIDAD_MAX,
    EQUIPOS_BASICOS_CANTIDAD_PYP_MIN,
    EQUIPOS_BASICOS_RUTA_DUPLICADA_THRESHOLD,
)
from app.constants.base import is_evidence_audit_enabled, is_rule_engine_enabled

# Module-level flag: skip evidence/audit DB writes when testing
_PERSIST = is_evidence_audit_enabled()
from app.services.transversales import (
    detect_cantidades_anomalas,
    detect_codigo_entidad_vs_entidad_afiliacion,
    detect_decimales,
    detect_doble_tipo_procedimiento,
    detect_ruta_duplicada,
    detect_tipo_documento_edad,
    detect_tipo_identificacion_entidad,
    detect_tipo_usuario,
    normalize_invoice,
)
from app.services.equipos_basicos.profesionales import (
    detect_profesionales_equipos_basicos,
)
from app.services.odontologia.centro_costo import (
    detect_centro_costo_odontologia,
)
from app.services.odontologia.ide_contrato import (
    detect_ide_contrato_odontologia,
)
from app.services.transversales.procedimiento_contratado import detect_cups_sin_contrato

logger = logging.getLogger(__name__)


def detect_all_problems_equipos_basicos(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
    rows: list[dict[str, Any]] | None = None,
    profesional_dias: dict[str, list[int]] | None = None,
    permitir_todos_centros: bool = True,
) -> tuple[dict[str, Any], dict[str, str]]:
    """
    Detecta TODOS los problemas en facturas de equipos básicos.

    Incluye detectores transversales y específicos de equipos básicos.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas
        profesional_dias: Dict {identificacion: [dias]} con días seleccionados
        permitir_todos_centros: Si True, solo permite centros válidos

    Returns:
        (resultado_dict, responsables_map)
    """
    # ── Consolidated engine rule evaluation (single session + single collector) ──
    if is_rule_engine_enabled():
        from app.services.engine.session_manager import SessionManager
        from app.services.engine.evidence_collector import EvidenceCollector
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.models import Regla, ResultadoAuditoria

        with SessionManager("equipos_basicos") as session:
            collector = EvidenceCollector(domain="equipos_basicos")

            decimales = RuleBasedDetector("valores_decimales", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            doble_tipo = RuleBasedDetector("doble_tipo_procedimiento", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            ruta_dup = RuleBasedDetector("ruta_duplicada", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )

            # tipo_documento_edad rules
            r1 = RuleBasedDetector("tipo_documento_edad_menor_7", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            r2 = RuleBasedDetector("tipo_documento_edad_mayor_18", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            r3 = RuleBasedDetector("tipo_documento_edad_7_17", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            r4 = RuleBasedDetector("tipo_documento_edad_as_menor", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            r5 = RuleBasedDetector("tipo_documento_edad_ms_mayor", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            r6 = RuleBasedDetector("tipo_documento_edad_cn_invalido", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            r7 = RuleBasedDetector("tipo_documento_edad_ce_invalido", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            tipo_id_edad = r1 + r2 + r3 + r4 + r5 + r6 + r7

            # tipo_identificacion_entidad rules
            r1_ent = RuleBasedDetector("tipo_id_requiere_entidad_86000", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            r2_ent = RuleBasedDetector("entidad_86000_requiere_as_ms", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            tipo_id_entidad = r1_ent + r2_ent

            # Cantidades anomalas
            cantidades = RuleBasedDetector("cantidades_anomalas", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )

            # codigo_entidad
            entidad_afiliacion_comparison = RuleBasedDetector("codigo_entidad", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )

            # tipo_usuario
            tipo_usuario_eb = RuleBasedDetector("tipo_usuario_valido", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )

            # IDE Contrato equipos básicos
            logger.info("detect_all_problems_equipos_basicos - Llamando detect_ide_contrato_odontologia")
            ide_contrato = RuleBasedDetector("ide_contrato_equipos_basicos_valido", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            logger.info("detect_all_problems_equipos_basicos - IDE Contrato encontrados: %d", len(ide_contrato))

            # Profesionales equipos básicos
            logger.info("detect_all_problems_equipos_basicos - Llamando detect_profesionales_equipos_basicos")
            profesionales = RuleBasedDetector("profesional_equipos_validos", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            logger.info("detect_all_problems_equipos_basicos - Profesionales encontrados: %d", len(profesionales))

            # Centro Costo
            centro_costo = RuleBasedDetector("centro_costo_equipos_basicos_valido", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )

            # CUPS sin contrato
            cups_sin_contrato = RuleBasedDetector("cups_sin_contrato", session).detect(
                data_sheet, indices, persist=_PERSIST,
                evidence_collector=collector, rows=rows,
            )
            logger.info(
                "detect_all_problems_equipos_basicos - Cups Sin Contrato encontrados: %d",
                len(cups_sin_contrato),
            )

            # ── Flush all evidence + create ResultadoAuditoria rows ──
            if _PERSIST:
                evidencias = collector.flush_batch(session)
                if evidencias:
                    regla_ids = {e.regla_id for e in evidencias}
                    reglas_map = {
                        r.id: r
                        for r in session.query(Regla).filter(Regla.id.in_(regla_ids))
                    }
                    for ev in evidencias:
                        if ev.outcome == "MATCH":
                            resultado_str = "FAIL"
                        elif ev.outcome == "ERROR":
                            resultado_str = "ERROR"
                        else:
                            resultado_str = "PASS"
                        rule = reglas_map.get(ev.regla_id)
                        ra = ResultadoAuditoria(
                            evidencia_id=ev.id,
                            regla_id=ev.regla_id,
                            regla_version=ev.regla_version,
                            factura=ev.factura,
                            param_config_id=ev.param_config_id,
                            resultado=resultado_str,
                            severidad=rule.severidad if rule else "error",
                            mensaje=ev.error_mensaje or (rule.descripcion if rule else ""),
                            detalles={"outcome": ev.outcome},
                        )
                        session.add(ra)
                    session.flush()
    else:
        decimales = detect_decimales(data_sheet, indices)
        doble_tipo = []
        ruta_dup = detect_ruta_duplicada(
            data_sheet, indices, threshold=EQUIPOS_BASICOS_RUTA_DUPLICADA_THRESHOLD
        )
        tipo_id_edad = []
        tipo_id_entidad = detect_tipo_identificacion_entidad(data_sheet, indices)
        cantidades = detect_cantidades_anomalas(
            data_sheet, indices,
            cantidad_consultas_min=EQUIPOS_BASICOS_CANTIDAD_CONSULTAS_MIN,
            cantidad_max_general=EQUIPOS_BASICOS_CANTIDAD_MAX,
            cantidad_pyp_min=EQUIPOS_BASICOS_CANTIDAD_PYP_MIN,
        ) if indices.get("procedimiento") is not None else []
        entidad_afiliacion_comparison = detect_codigo_entidad_vs_entidad_afiliacion(
            data_sheet, indices, limit_log=5
        )
        tipo_usuario_eb = detect_tipo_usuario(data_sheet, indices)
        ide_contrato = detect_ide_contrato_odontologia(data_sheet, indices)
        profesionales = []
        centro_costo = []
        cups_sin_contrato = detect_cups_sin_contrato(data_sheet, indices)

    # Build responsable_cierra mapping
    responsable_cierra: dict[str, str] = {}
    responsable_cierra_idx = indices.get("responsable_cierra")
    num_fact_idx = indices.get("numero_factura")
    if responsable_cierra_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            raw = data_sheet.cell(row=row, column=responsable_cierra_idx + 1).value
            resp = str(raw).strip() if raw else ""
            if resp and factura not in responsable_cierra:
                responsable_cierra[factura] = resp

    # Build fec_factura_map
    fec_factura_map: dict[str, str] = {}
    fec_factura_idx = indices.get("fec_factura")
    if fec_factura_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            raw = data_sheet.cell(row=row, column=fec_factura_idx + 1).value
            val = str(raw).strip() if raw else ""
            if val and factura not in fec_factura_map:
                fec_factura_map[factura] = val

    # Build normalized rows for unified 6-column display
    from app.services.odontologia.normalized_rows import build_odontologia_normalized_rows

    normalized_rows_eb = build_odontologia_normalized_rows(
        decimales=decimales,
        doble_tipo=doble_tipo,
        ruta_dup=ruta_dup,
        profesionales=profesionales,
        cantidades=cantidades,
        tipo_id_edad=tipo_id_edad,
        tipo_id_entidad=tipo_id_entidad,
        centro_costo=centro_costo,
        ide_contrato=ide_contrato,
        responsable_cierra=responsable_cierra,
        entidad_afiliacion_comparison=entidad_afiliacion_comparison,
        tipo_usuario=tipo_usuario_eb,
        fec_factura_map=fec_factura_map,
        cups_sin_contrato=cups_sin_contrato,
    )

    resultado: dict[str, Any] = {
        "area": AREA_EQUIPOS_BASICOS,
        "problemas": {
            "normalizados": normalized_rows_eb,
            "decimales": decimales,
            "doble_tipo_procedimiento": doble_tipo,
            "ruta_duplicada": ruta_dup,
            "profesionales": profesionales,
            "cantidades_anomalas": cantidades,
            "tipo_identificacion_edad": tipo_id_edad,
            "tipo_identificacion_entidad": tipo_id_entidad,
            "codigo_entidad_vs_afiliacion": entidad_afiliacion_comparison,
            "tipo_usuario": tipo_usuario_eb,
            "centro_costo": centro_costo,
            "ide_contrato": ide_contrato,
            "cups_sin_contrato": cups_sin_contrato,
        },
        "totales": {
            "decimales": len(decimales),
            "doble_tipo_procedimiento": len(doble_tipo),
            "ruta_duplicada": len(ruta_dup),
            "profesionales": len(profesionales),
            "cantidades_anomalas": len(cantidades),
            "tipo_identificacion_edad": len(tipo_id_edad),
            "tipo_identificacion_entidad": len(tipo_id_entidad),
            "centro_costo": len(centro_costo),
            "ide_contrato": len(ide_contrato),
            "codigo_entidad_vs_afiliacion": len(entidad_afiliacion_comparison),
            "tipo_usuario": len(tipo_usuario_eb),
            "cups_sin_contrato": len(cups_sin_contrato),
        },
        "es_equipos_basicos": True,
        "missing_columns": [],
    }

    # Enrich errors with responsable from mapping
    if responsable_cierra:
        for problem_type, problems in resultado["problemas"].items():
            if not isinstance(problems, list):
                continue
            for p in problems:
                if not isinstance(p, dict):
                    continue
                factura = p.get("factura")
                if factura and factura in responsable_cierra:
                    p["responsable"] = responsable_cierra[factura]
                elif "responsable" not in p:
                    p["responsable"] = ""
    else:
        for problem_type, problems in resultado["problemas"].items():
            if not isinstance(problems, list):
                continue
            for p in problems:
                if not isinstance(p, dict):
                    continue
                if "responsable" not in p:
                    p["responsable"] = ""

    return resultado, responsable_cierra
