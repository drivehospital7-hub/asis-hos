"""Servicio orquestador de exportación Excel.

Este módulo es el punto de entrada principal para la exportación de archivos
Excel con hoja de cruce de facturas. Coordina los demás módulos:
- validators: Validación de paths
- column_filter: Filtrado de columnas
- cruce_sheet: Creación de hoja CruceFacturas
- formatting: Formato condicional
"""

from __future__ import annotations

import logging
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.constants import (
    CRUCE_FACTURAS_SHEET,
    AREA_ODONTOLOGIA,
    AREA_URGENCIAS,
    AREA_EQUIPOS_BASICOS,
    COLUMNS_TO_KEEP,
    URGENCIA_COLUMNS_TO_KEEP,
    PROFESIONALES_ODONTOLOGIA,
)
from app.services.cruce_sheet import create_cruce_facturas_sheet
from app.services.revision_sheet import detect_all_problems
from app.services.transversales.estructura_excel import detectar_estructura_excel
from app.utils.column_filter import filter_columns
from app.utils.formatting import apply_all_conditional_formatting
from app.utils.input_data import (
    resolve_safe_excel_absolute,
    resolve_safe_excel_in_input,
    resolve_safe_excel_in_output,
)
from app.utils.validators import validate_excel_path

logger = logging.getLogger(__name__)


def _copy_file(source: Path, destination: Path) -> None:
    """Copia el archivo fuente al destino."""
    shutil.copy2(source, destination)
    logger.info("Archivo copiado: %s -> %s", source.name, destination.name)


def export_excel_with_cruce_facturas(
    *,
    filename: str,
    sheet_name: str | None = None,
    header_row: int = 0,
    area: str = AREA_ODONTOLOGIA,
    profesional: str = "",
    dias: list[int] | None = None,
    todos_profesionales_dias: dict[str, list[int]] | None = None,
    validar_centro_costo: bool = False,
    equipos_basicos: bool = False,
) -> dict[str, Any]:
    """
    Exporta un archivo Excel con hoja de cruce de facturas.
    
    Este es el orquestador principal que:
    1. Valida el archivo de entrada
    2. Copia el archivo a output
    3. Filtra columnas de la hoja de datos
    4. Crea hoja CruceFacturas con headers
    5. Aplica formato condicional
    6. Guarda el archivo
    
    Args:
        filename: Nombre del archivo en input/
        sheet_name: Nombre de la hoja a procesar (None = hoja activa)
        header_row: Fila de headers (no usado actualmente, reservado para futuro)
        area: Área del sistema ("odontologia" o "urgencias")
        profesional: Código del profesional seleccionado (para validación centro costo)
        dias: Lista de días seleccionados para el profesional (para validación centro costo)
        todos_profesionales_dias: Dict {codigo: [dias]} con todos los profesionales y sus días
        validar_centro_costo: Si True, valida centros de costo según días
    
    Returns:
        Dict con formato estándar:
        {
            "status": "success" | "error",
            "data": {...},
            "errors": [...]
        }
    """
    logger.info("Iniciando exportación: %s", filename)
    
    # Determinar el área efectiva (si equipos_basicos está activo, usar reglas independientes)
    area_effective = AREA_EQUIPOS_BASICOS if equipos_basicos else area
    logger.info("Área efectiva: %s (equipos_basicos: %s)", area_effective, equipos_basicos)
    
    # Construir datos para validación de centro costo según el área
    profesional_dias = {}
    permitir_todos_centros = False  # Por defecto: solo ODONTOLOGIA y EXTRAMURAL
    
    if area_effective == AREA_ODONTOLOGIA or area_effective == AREA_EQUIPOS_BASICOS:
        if validar_centro_costo and todos_profesionales_dias:
            # Activado: usar todos los profesionales y sus días desde localStorage
            for prof_codigo, dias_list in todos_profesionales_dias.items():
                if dias_list:  # Solo incluir profesionales con días seleccionados
                    profesional_info = PROFESIONALES_ODONTOLOGIA.get(prof_codigo)
                    if profesional_info:
                        profesional_id = profesional_info.get("identificacion")
                        if profesional_id:
                            profesional_dias[profesional_id] = dias_list
                            logger.info("Validación centro costo ACTIVADA - Profesional %s (%s), días: %s",
                                       prof_codigo, profesional_id, dias_list)
            
            if not profesional_dias:
                # Si no hay días para ningún profesional, permitir todos los centros
                permitir_todos_centros = True
                logger.info("No hay días seleccionados para ningún profesional - Solo se permiten ODONTOLOGIA y EXTRAMURAL")
        elif validar_centro_costo and profesional and dias:
            # Fallback: solo el profesional seleccionado (para compatibilidad)
            profesional_info = PROFESIONALES_ODONTOLOGIA.get(profesional)
            if profesional_info:
                profesional_id = profesional_info.get("identificacion")
                if profesional_id:
                    profesional_dias[profesional_id] = dias
                    logger.info("Validación centro costo ACTIVADA (fallback) - Profesional %s (%s), días: %s",
                               profesional, profesional_id, dias)
        else:
            # No activado: permitir solo ODONTOLOGIA y EXTRAMURAL (cualquier profesional)
            permitir_todos_centros = True
            logger.info("Validación centro costo NO activada - Solo se permiten ODONTOLOGIA y EXTRAMURAL")
    
    # 1. Resolver y validar path de entrada (soporta repo o archivo subido)
    source_path, source_error = resolve_safe_excel_absolute(filename)
    if source_error:
        logger.error("Error resolviendo archivo de entrada: %s", source_error)
        return {"status": "error", "data": {}, "errors": [source_error]}
    assert source_path is not None
    
    validation_error = validate_excel_path(source_path)
    if validation_error:
        logger.error("Error de validación: %s", validation_error)
        return {"status": "error", "data": {}, "errors": [validation_error]}
    
    # 2. Resolver path de salida
    # Usar nombre original del archivo (sin el UUID del temp upload)
    original_filename = Path(filename).name
    if "_" in original_filename:
        # Es un archivo temporal: extraer nombre original despues del UUID
        parts = original_filename.split("_", 1)
        if len(parts) == 2 and len(parts[0]) == 32:  # UUID es 32 hex chars
            original_filename = parts[1]
    
    output_path, output_error = resolve_safe_excel_in_output(original_filename)
    if output_error:
        logger.error("Error resolviendo archivo de salida: %s", output_error)
        return {"status": "error", "data": {}, "errors": [output_error]}
    assert output_path is not None
    
    try:
        # 3. Copiar archivo a output
        _copy_file(source_path, output_path)
        
        # 4. Cargar workbook
        workbook = load_workbook(output_path)
        
        # 5. Obtener hoja de datos
        if sheet_name and sheet_name in workbook.sheetnames:
            data_sheet = workbook[sheet_name]
        else:
            data_sheet = workbook.active
        
        # 6. Detectar estructura del Excel (cuántas filas eliminar)
        estructura_result = detectar_estructura_excel(output_path)
        if estructura_result["status"] == "success":
            filas_a_eliminar = estructura_result["data"]["filas_a_eliminar"]
            logger.info(
                "Estructura detectada: %s - filas a eliminar: %d",
                estructura_result["data"]["estructura"],
                filas_a_eliminar,
            )
        else:
            # Si falla la detección, asumir comportamiento por defecto (2 filas)
            filas_a_eliminar = 2
            logger.warning("Error detectando estructura, usando默认值: %d", filas_a_eliminar)
        
        # 7. Filtrar columnas (según el área)
        if area_effective == AREA_URGENCIAS:
            # Urgencias: no ocultar columnas (None = mantener todas)
            columns_to_keep = None
        else:
            columns_to_keep = COLUMNS_TO_KEEP
        
        filter_result = filter_columns(
            data_sheet, 
            columns_to_keep=columns_to_keep,
            delete_first_rows=filas_a_eliminar,
        )
        logger.info("Columnas filtradas: %s", filter_result)
        
        # 7. Detectar problemas para mostrar en HTML (sin crear hoja)
        problemas_detectados, responsables_map = detect_all_problems(
            data_sheet, 
            area=area_effective,
            profesional_dias=profesional_dias if area_effective in (AREA_ODONTOLOGIA, AREA_EQUIPOS_BASICOS) else None,
            permitir_todos_centros=(permitir_todos_centros or equipos_basicos) if area_effective in (AREA_ODONTOLOGIA, AREA_EQUIPOS_BASICOS) else False,
        )
        
        # 8. Crear hoja CruceFacturas
        cruce_sheet, cruce_info = create_cruce_facturas_sheet(workbook)
        
        # 9. Aplicar formato condicional
        formatting_results = apply_all_conditional_formatting(cruce_sheet, data_sheet)
        
        # 10. Guardar
        workbook.save(output_path)
        logger.info("Archivo guardado: %s", output_path.name)
        
    except Exception as exc:
        logger.exception("Error exportando Excel")
        return {"status": "error", "data": {}, "errors": [str(exc)]}
    
    logger.info("Exportación completada: %s", output_path.name)
    
    return {
        "status": "success",
        "data": {
            "input_file": source_path.name,
            "output_file": output_path.name,
            "output_path": str(output_path),
            "sheet": CRUCE_FACTURAS_SHEET,
            "headers_written": ["B1", "D1", "F1"],
            "estructura_excel": estructura_result.get("data", {}),
            "filter_result": filter_result,
            "problemas": problemas_detectados,
            "responsables_map": responsables_map,
            "applied_rules": [
                cruce_info,
                *formatting_results,
            ],
        },
        "errors": [],
    }
