"""Detector de problemas de centro de costo para Ambulatoria.

Aplica reglas comunes (1-4, 8, 9) + regla específica de Ambulatoria:
- AMBULATORIA_PYP: Tipo Factura=Ambulatoria → Centro debe ser PYP
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    CENTRO_COSTO_PYP_URGENCIAS,
)
from app.services.transversales.centro_costo_rules import apply_common_centro_costo_rules
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_centro_costo_ambulatoria(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """Detecta facturas de Ambulatoria con problemas de centro de costo.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: factura, tipo_factura, centro_actual,
        centro_deberia, codigo, procedimiento, prioridad, regla
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")
    codigo_idx = indices.get("codigo")
    laboratorio_idx = indices.get("laboratorio")
    centro_costo_idx = indices.get("centro_costo")
    codigo_entidad_cobrar_idx = indices.get("codigo_entidad_cobrar")
    tipo_factura_descripcion_idx = indices.get("tipo_factura_descripcion")
    proc_idx = indices.get("procedimiento")
    tarifario_idx = indices.get("tarifario")

    if num_fact_idx is None or centro_costo_idx is None:
        logger.warning("Centro Costo Ambulatoria - Columnas necesarias no encontradas")
        return []

    problemas_centros: list[dict[str, Any]] = []

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        tipo_factura_descripcion = (
            data_sheet.cell(row=row, column=tipo_factura_descripcion_idx + 1).value
            if tipo_factura_descripcion_idx is not None else None
        )
        tipo_factura_str = str(tipo_factura_descripcion).strip() if tipo_factura_descripcion else ""

        # Only process Ambulatoria rows
        if tipo_factura_str != "Ambulatoria":
            continue

        # Read row values
        codigo_tipo_proc = (
            data_sheet.cell(row=row, column=codigo_tipo_proc_idx + 1).value
            if codigo_tipo_proc_idx is not None else None
        )
        codigo = (
            data_sheet.cell(row=row, column=codigo_idx + 1).value
            if codigo_idx is not None else None
        )
        laboratorio = (
            data_sheet.cell(row=row, column=laboratorio_idx + 1).value
            if laboratorio_idx is not None else None
        )
        centro_costo = data_sheet.cell(row=row, column=centro_costo_idx + 1).value
        codigo_entidad_cobrar = (
            data_sheet.cell(row=row, column=codigo_entidad_cobrar_idx + 1).value
            if codigo_entidad_cobrar_idx is not None else None
        )
        procedimiento = (
            data_sheet.cell(row=row, column=proc_idx + 1).value
            if proc_idx is not None else None
        )
        tarifario = (
            data_sheet.cell(row=row, column=tarifario_idx + 1).value
            if tarifario_idx is not None else None
        )

        codigo_str = str(codigo_tipo_proc).strip() if codigo_tipo_proc else ""
        codigo_excluir = str(codigo).strip() if codigo else ""
        laboratorio_str = str(laboratorio).strip() if laboratorio else ""
        centro_costo_str = str(centro_costo).strip() if centro_costo else ""
        codigo_entidad_str = str(codigo_entidad_cobrar).strip() if codigo_entidad_cobrar else ""
        proc_str = str(procedimiento).strip() if procedimiento else ""
        tarifario_str = str(tarifario).strip() if tarifario else ""

        # Apply common rules
        errors = apply_common_centro_costo_rules(
            centro_costo_str=centro_costo_str,
            codigo_str=codigo_str,
            codigo_excluir=codigo_excluir,
            laboratorio_str=laboratorio_str,
            tarifario_str=tarifario_str,
            codigo_entidad_str=codigo_entidad_str,
            factura_str=factura_str,
            proc_str=proc_str,
        )

        for e in errors:
            e["tipo_factura"] = tipo_factura_str
        problemas_centros.extend(errors)

        # --- AMBULATORIA_PYP: Ambulatoria → Centro debe ser PYP ---
        if tipo_factura_str == "Ambulatoria":
            if centro_costo_str != CENTRO_COSTO_PYP_URGENCIAS:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": CENTRO_COSTO_PYP_URGENCIAS,
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "AMBULATORIA_PYP",
                })

    if problemas_centros:
        logger.info("Centro Costo Ambulatoria - Problemas encontrados: %d", len(problemas_centros))

    return problemas_centros
