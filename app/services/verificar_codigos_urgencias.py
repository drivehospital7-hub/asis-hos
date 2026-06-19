"""Verificación de códigos CUPS contra la base de datos.

Para Urgencias - Entidad ESS118:
Busca códigos en la columna "Código" que no existen en la cadena
eps_contratado → eps_nota → nota_hoja → notas_tecnicas → procedimiento.

Migrado a SQLAlchemy directo (antes usaba procedimientos_db + psycopg2).
La vista v_procedimientos no expone cod_contrato, por lo que se usa
la cadena completa vía modelos SQLAlchemy.

Uso:
    python -m app.services.verificar_codigos_urgencias <ruta_excel>
"""

import sys
import logging
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

from app.database import SessionLocal
from app.models import EpsContratado, Procedimiento, NotasTecnicas, NotaHoja, EpsNota

logger = logging.getLogger(__name__)

# Constantes
EPS_DB = "EMSSANAR_CAPITA"

# Mapping de nombre de EPS a cod_contrato en la cadena SQLAlchemy
EPS_NAME_TO_COD_CONTRATO = {
    "EMSSANAR_CAPITA": "ESS118",
}


def get_indices(sheet) -> dict:
    """Detector índices de columnas."""
    headers = [cell.value for cell in sheet[1]]

    indices = {}
    for idx, header in enumerate(headers):
        if header:
            header_lower = str(header).strip().lower()
            indices[header_lower] = idx

    return indices


def verificar_excel(excel_path: str | Path) -> dict:
    """Procesa el Excel y verifica códigos contra DB.

    Returns:
        Dict con:
        - codigos_no_encontrados: list of codigos
        - codigos_encontrados: list of codigos
        - total ESS118: int
    """
    excel_path = Path(excel_path)

    if not excel_path.exists():
        logger.error("Archivo no encontrado: %s", excel_path)
        return {}

    # Cargar workbook
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active

    indices = get_indices(sheet)
    logger.info("Índices detectados: %s", indices)

    # Buscar columnas relevantes
    codigo_idx = indices.get("codigo")  # columna Código
    entidad_cobrar_idx = indices.get("entidad cobrar") or indices.get("entidad_cobrar")
    codigo_entidad_idx = indices.get("cód entidad cobrar") or indices.get("cod_entidad_cobrar")

    if codigo_idx is None:
        logger.error("Columna 'Código' no encontrada")
        return {}

    logger.info("Buscando códigos para ESS118...")

    # Collect códigos únicos para ESS118
    codigos_ess118 = set()

    for row in range(2, sheet.max_row + 1):
        # Leer valores
        entidad_str = None
        if entidad_cobrar_idx is not None:
            entidad_str = sheet.cell(row=row, column=entidad_cobrar_idx + 1).value
        if codigo_entidad_idx is not None:
            codigo_entidad = sheet.cell(row=row, column=codigo_entidad_idx + 1).value
            if codigo_entidad:
                entidad_str = str(codigo_entidad).strip()

        codigo = sheet.cell(row=row, column=codigo_idx + 1).value

        # Skip empty
        if not codigo:
            continue

        codigo_str = str(codigo).strip()

        # Filtrar solo ESS118 (por código entidad o nombre entidad)
        es_ess118 = False
        if entidad_str:
            entidad_normalizada = str(entidad_str).strip().upper()
            if "ESS118" in entidad_normalizada:
                es_ess118 = True

        if es_ess118:
            codigos_ess118.add(codigo_str)

    logger.info("Total códigos únicos para ESS118: %d", len(codigos_ess118))

    # Obtener cod_contrato para la EPS
    cod_contrato = EPS_NAME_TO_COD_CONTRATO.get("EMSSANAR_CAPITA")
    if not cod_contrato:
        logger.error("No hay cod_contrato mapeado para EMSSANAR_CAPITA")
        return {
            "codigos_no_encontrados": sorted(codigos_ess118),
            "codigos_encontrados": [],
            "total_ess118": len(codigos_ess118),
        }

    # Verificar cada código contra la DB usando SQLAlchemy
    session = SessionLocal()
    codigos_no_encontrados = []
    codigos_encontrados = []

    try:
        for codigo in sorted(codigos_ess118):
            result = (
                session.query(Procedimiento)
                .join(NotasTecnicas, Procedimiento.id == NotasTecnicas.id_procedimiento)
                .join(NotaHoja, NotasTecnicas.id_nota_hoja == NotaHoja.id)
                .join(EpsNota, NotaHoja.id == EpsNota.id_nota_hoja)
                .join(EpsContratado, EpsNota.id_eps_contratado == EpsContratado.id)
                .filter(EpsContratado.cod_contrato == cod_contrato)
                .filter(Procedimiento.cups == codigo)
                .first()
            )

            if result:
                codigos_encontrados.append(codigo)
                logger.debug("Encontrado: %s - %s", codigo, result.procedimiento)
            else:
                codigos_no_encontrados.append(codigo)
                logger.warning("NO ENCONTRADO en DB: %s (cod_contrato=%s)", codigo, cod_contrato)

        # Resumen
        logger.info("=" * 60)
        logger.info("RESUMEN - Verificación de Códigos para ESS118")
        logger.info("=" * 60)
        logger.info("Total códigos ESS118 únicos: %d", len(codigos_ess118))
        logger.info("Encontrados en DB (%s): %d", cod_contrato, len(codigos_encontrados))
        logger.info("NO encontrados en DB: %d", len(codigos_no_encontrados))

        if codigos_no_encontrados:
            logger.warning("Códigos NO encontrados en la DB:")
            for codigo in codigos_no_encontrados:
                logger.warning("  - %s", codigo)

    finally:
        session.close()

    return {
        "codigos_no_encontrados": codigos_no_encontrados,
        "codigos_encontrados": codigos_encontrados,
        "total_ess118": len(codigos_ess118),
    }


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.WARNING,
        format="%(levelname)s: %(message)s"
    )

    if len(sys.argv) < 2:
        print("Uso: python -m app.services.verificar_codigos_urgencias <archivo_excel>")
        sys.exit(1)

    excel_path = sys.argv[1]
    resultado = verificar_excel(excel_path)
