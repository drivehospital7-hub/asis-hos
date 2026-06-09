"""Orquestador de detección de problemas para Farmacia.

Agrupa detectores transversales + específicos de Farmacia.
"""

from __future__ import annotations

import logging
from typing import Any, Callable

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import AREA_FARMACIA
from app.services.transversales import (
    normalize_invoice,
)
from app.services.normalized_rows import build_normalized_rows

logger = logging.getLogger(__name__)


def _get_farmacia_detectors() -> list[Callable]:
    """Returns list of Farmacia-specific detector callables.

    Used by tipo_factura_registry for lazy import.
    """
    return []


def detect_all_problems_farmacia(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> tuple[dict[str, Any], dict[str, str]]:
    """Detecta TODOS los problemas en facturas de Farmacia.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        (resultado_dict, responsables_map)
    """
    from app.services.transversales import (
        detect_decimales,
        detect_tipo_documento_edad,
        detect_tipo_identificacion_entidad,
        detect_codigo_entidad_vs_entidad_afiliacion,
        detect_tipo_usuario,
    )
    from app.services.transversales.detect_copago_entidad import (
        detect_copago_entidad_urgencias,
    )
    from app.services.transversales.procedimiento_contratado import detect_cups_sin_contrato

    # 1. Detectores transversales
    decimales = detect_decimales(data_sheet, indices)
    tipo_identificacion_edad = detect_tipo_documento_edad(data_sheet, indices)
    tipo_identificacion_entidad = detect_tipo_identificacion_entidad(data_sheet, indices)
    entidad_afiliacion_comparison = detect_codigo_entidad_vs_entidad_afiliacion(
        data_sheet, indices, limit_log=5
    )
    tipo_usuario = detect_tipo_usuario(data_sheet, indices)
    copago_entidad = detect_copago_entidad_urgencias(data_sheet, indices)
    cups_sin_contrato = detect_cups_sin_contrato(data_sheet, indices)

    # 2. Build responsable_cierra mapping
    responsable_cierra: dict[str, str] = {}
    responsable_cierra_idx = indices.get("responsable_cierra")
    num_fact_idx = indices.get("numero_factura")
    if responsable_cierra_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            raw = data_sheet.cell(row=row, column=responsable_cierra_idx + 1).value
            resp = str(raw).strip() if raw else ""
            if resp and factura not in responsable_cierra:
                responsable_cierra[factura] = resp

    # 3. Build fecha_cierre_vacia mapping
    fecha_cierre_vacia: dict[str, bool] = {}
    fecha_cierre_idx = indices.get("fecha_cierre")
    if fecha_cierre_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            fecha_cierre_val = data_sheet.cell(row=row, column=fecha_cierre_idx + 1).value
            if not fecha_cierre_val or str(fecha_cierre_val).strip() == "":
                fecha_cierre_vacia[factura] = True
            elif factura not in fecha_cierre_vacia:
                fecha_cierre_vacia[factura] = False

    # 4. Build fec_factura_map
    fec_factura_map: dict[str, str] = {}
    fec_factura_idx = indices.get("fec_factura")
    if fec_factura_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            raw = data_sheet.cell(row=row, column=fec_factura_idx + 1).value
            val = str(raw).strip() if raw else ""
            if val and factura not in fec_factura_map:
                fec_factura_map[factura] = val

    # 5. Build normalized rows
    error_groups = {
        "Decimales": decimales,
        "Tipo Identificación / Edad": tipo_identificacion_edad,
        "Código Entidad vs Afiliación": entidad_afiliacion_comparison + tipo_identificacion_entidad,
        "Tipo Usuario": tipo_usuario,
        "Copago vs Entidad": copago_entidad,
        "Cups Sin Contrato": cups_sin_contrato,
    }
    normalized_rows = build_normalized_rows(
        error_groups=error_groups,
        responsables_map=responsable_cierra,
        fec_factura_map=fec_factura_map,
        fecha_cierre_vacia_map=fecha_cierre_vacia,
    )

    # 6. Build resultado
    resultado: dict[str, Any] = {
        "area": AREA_FARMACIA,
        "problemas": {
            "normalizados": normalized_rows,
            "centros_de_costos": [],
            "ide_contrato": [],
            "cups_equivalentes": [],
            "decimales": decimales,
            "tipo_identificacion_edad": tipo_identificacion_edad,
            "tipo_identificacion_entidad": tipo_identificacion_entidad,
            "codigo_entidad_vs_afiliacion": entidad_afiliacion_comparison,
            "tipo_usuario": tipo_usuario,
            "copago_entidad": copago_entidad,
            "cups_sin_contrato": cups_sin_contrato,
        },
        "totales": {
            "centros_de_costos": 0,
            "ide_contrato": 0,
            "cups_equivalentes": 0,
            "decimales": len(decimales),
            "tipo_identificacion_edad": len(tipo_identificacion_edad),
            "tipo_identificacion_entidad": len(tipo_identificacion_entidad),
            "codigo_entidad_vs_afiliacion": len(entidad_afiliacion_comparison),
            "tipo_usuario": len(tipo_usuario),
            "copago_entidad": len(copago_entidad),
            "cups_sin_contrato": len(cups_sin_contrato),
        },
        "missing_columns": [],
    }

    # 7. Enrich errors with responsable
    for problem_type, problems in resultado["problemas"].items():
        for p in problems:
            if not isinstance(p, dict):
                continue
            factura = p.get("factura")
            if factura and factura in responsable_cierra:
                p["responsable"] = responsable_cierra[factura]
            elif "responsable" not in p:
                p["responsable"] = ""

    return resultado, responsable_cierra
