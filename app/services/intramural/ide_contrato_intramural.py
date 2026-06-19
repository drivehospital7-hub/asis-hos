"""Detector de problemas de IDE Contrato en facturas de Intramural.

Solo aplica cuando Tipo Factura Descripción = "Intramural".
Reglas definidas en ide_contrato_rules.py y constantes de intramural.
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import CODIGOS_LABORATORIO_ENVIO
from app.constants.intramural import (
    TIPO_FACTURA_INTRAMURAL,
    CODIGOS_PYM_RUTAS,
    CODIGOS_PYM_NECESITAN_DX,
    CODIGOS_PYM_INTRAMURAL,
    CODIGOS_NUEVA_EPS_NO_CAPITA,
)
from app.services.transversales.normalize import normalize_invoice
from app.services.intramural.ide_contrato_rules import (
    IDE_SIMPLE_RULES,
    IDE_INSERTION_RULES,
    IDE_MULTIPLE_RULES,
    IDE_ESSC62_890405_RULES,
)

logger = logging.getLogger(__name__)

# PYM_RUTAS + Dx: según prefijo factura
_PYM_RUTAS_IDE_MAP: dict[str, set[str]] = {
    "EPSS41": {"955"},
    "EPS037": {"961"},
    "RES001": {"993"},
    "ESSC62": {"863"},
    "ESS062": {"922"},
    "RES004": {"908"},
    "EPSI04": {"901"},
    "EPSI03": {"965"},
    "EPS025": {"902"},
    "RES002": {"952"},
    "5177": {"913"},
    "86000": {"920"},
    "CCF033": {"937"},
    "CCF050": {"914"},
    "CCF055": {"868"},
    "CCF102": {"888"},
    "CCFC33": {"990"},
    "EPS001": {"950"},
    "EPS002": {"936"},
    "EPS008": {"870"},
    "EPS010": {"925"},
    "EPS017": {"892"},
    "EPS018": {"891"},
    "EPS025": {"902"},
    "EPS037": {"961"},
    "EPS040": {"947"},
    "EPS048": {"943"},
    "EPSC005": {"932"},
    "EPSC34": {"991"},
    "EPSI03": {"965"},
    "EPSI05": {"977"},
    "EPSI06": {"896"},
    "EPSIC5": {"979"},
    "EPSS005": {"933"},
    "EPSS018": {"927"},
    "EPSS02": {"903"},
    "EPSS08": {"945"},
    "EPSS10": {"904"},
    "EPSS17": {"893"},
    "EPSS34": {"881"},
    "EPSS40": {"898"},
    "ESS062": {"922"},
    "ESS207": {"864"},
    "ESSC24": {"894"},
    "ESSC62": {"863"},
    "RES001": {"993"},
    "RES002": {"952"},
    "RES004": {"908"},
    "ESSC18": {"975"},
}
_PYM_RUTAS_FEV_MAP: dict[str, set[str]] = {
    # "EPSS41": {"958", "959"},
}
_PYM_RUTAS_EXCLUIDOS: frozenset[str] = frozenset()


def _check_pym_ruta_con_dx_ides(
    codigo: str,
    entidad: str,
    dx_principal: str,
    factura: str,
) -> set[str] | None:
    """Si código está en PYM_RUTAS, entidad tiene mapeo y Dx está en
    NECESITAN_DX, retorna el SET de IDEs válidos. Sino None.
    
    Si la factura empieza con FEV y la entidad está en _PYM_RUTAS_FEV_MAP,
    usa esos IDEs alternativos."""
    if entidad not in _PYM_RUTAS_IDE_MAP:
        return None
    if codigo in _PYM_RUTAS_EXCLUIDOS:
        return None
    if codigo not in CODIGOS_PYM_RUTAS:
        return None
    if not dx_principal or dx_principal not in CODIGOS_PYM_NECESITAN_DX:
        return None

    factura_up = factura.upper().strip()
    if factura_up.startswith("FEV") and entidad in _PYM_RUTAS_FEV_MAP:
        return _PYM_RUTAS_FEV_MAP[entidad]
    return _PYM_RUTAS_IDE_MAP[entidad]


def detect_ide_contrato_intramural(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta facturas con problemas de IDE Contrato en Intramural.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con: factura, codigo, entidad,
        ide_contrato_actual, ide_contrato_deberia
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    ide_contrato_idx = indices.get("ide_contrato")
    entidad_idx = indices.get("codigo_entidad_cobrar")
    tipo_fact_desc_idx = indices.get("tipo_factura_descripcion")
    dx_principal_idx = indices.get("codigo_dx_principal")
    tarifario_idx = indices.get("tarifario")

    if any(idx is None for idx in (num_fact_idx, codigo_idx, ide_contrato_idx, entidad_idx)):
        logger.warning(
            "IDE Contrato Intramural - Columnas necesarias no encontradas"
        )
        return []

    # =========================================================================
    # Pre-scan: detectar facturas donde TODOS los códigos son laboratorio envío
    # (excepción: si toda la factura son exámenes de laboratorio derivados,
    #  no se exige el IDE Contrato específico de PyM rutas)
    # =========================================================================
    _facturas_con_no_lab: set[str] = set()
    for row in range(2, data_sheet.max_row + 1):
        if tipo_fact_desc_idx is not None:
            tipo_fact_val = data_sheet.cell(row=row, column=tipo_fact_desc_idx + 1).value
            if str(tipo_fact_val or "").strip().upper() != TIPO_FACTURA_INTRAMURAL.upper():
                continue

        numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura = normalize_invoice(numero)
        if not factura:
            continue

        codigo = str(data_sheet.cell(row=row, column=codigo_idx + 1).value or "").strip()
        if not codigo:
            continue

        if codigo not in CODIGOS_LABORATORIO_ENVIO:
            _facturas_con_no_lab.add(factura)
    # Una factura es "solo laboratorio de envío" si aparece en Intramural
    # y NO está en _facturas_con_no_lab.

    problemas: list[dict[str, Any]] = []

    for row in range(2, data_sheet.max_row + 1):
        # Solo aplicar si Tipo Factura Descripción = "Intramural"
        if tipo_fact_desc_idx is not None:
            tipo_fact_val = data_sheet.cell(row=row, column=tipo_fact_desc_idx + 1).value
            if str(tipo_fact_val or "").strip().upper() != TIPO_FACTURA_INTRAMURAL.upper():
                continue

        numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura = normalize_invoice(numero)
        if not factura:
            continue

        codigo = str(data_sheet.cell(row=row, column=codigo_idx + 1).value or "").strip()
        ide_actual_raw = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value
        entidad = str(data_sheet.cell(row=row, column=entidad_idx + 1).value or "").strip()
        ide_actual = str(ide_actual_raw or "").strip()

        if not codigo or not entidad:
            continue

        procedimiento = str(
            data_sheet.cell(row=row, column=indices.get("procedimiento", 0) + 1).value or ""
        ).strip()

        # --- Reglas exactas (codigo + entidad → IDE esperado) ---
        matched = False
        for rule in IDE_SIMPLE_RULES:
            if codigo == rule["codigo"] and entidad == rule["entidad"]:
                # Exclusión: PYM_INTRAMURAL en NUEVA_EPS_NO_CAPITA no se valida
                if codigo in CODIGOS_PYM_INTRAMURAL and codigo in CODIGOS_NUEVA_EPS_NO_CAPITA:
                    continue  # no marca error, no hace match
                if ide_actual != rule["expected"]:
                    problemas.append({
                        "factura": factura,
                        "codigo": codigo,
                        "entidad": entidad,
                        "procedimiento": procedimiento,
                        "ide_contrato_actual": ide_actual,
                        "ide_contrato_deberia": rule["expected"],
                    })
                matched = True
                break

        if matched:
            continue

        # --- Regla: PYM_RUTAS + Dx Principal en NECESITAN_DX ---
        dx_principal = ""
        if dx_principal_idx is not None:
            dx_principal = str(
                data_sheet.cell(row=row, column=dx_principal_idx + 1).value or ""
            ).strip().upper()

        ides_validos = _check_pym_ruta_con_dx_ides(codigo, entidad, dx_principal, factura)
        if ides_validos is not None and ide_actual not in ides_validos:
            # Excepción: si TODOS los códigos de la factura son laboratorio de
            # envío (CODIGOS_LABORATORIO_ENVIO), no se exige el IDE específico.
            if factura not in _facturas_con_no_lab:
                logger.info(
                    "Excepción IDE para factura %s: todos los códigos son "
                    "laboratorio de envío, se omite validación PyM ruta",
                    factura,
                )
                continue
            problemas.append({
                "factura": factura,
                "codigo": codigo,
                "entidad": entidad,
                "procedimiento": procedimiento,
                "ide_contrato_actual": ide_actual,
                "ide_contrato_deberia": "/".join(sorted(ides_validos)),
            })

        # --- Regla: Tarifario (COMENTADO - era ESS118) ---
        # if entidad == "ESS118" and tarifario_idx is not None:
        #     ...

        # --- Regla: IDE 971/972 (COMENTADO - era ESS118) ---
        # if entidad == "ESS118" and ide_actual in ("971", "972"):
        #     ...

    # --- Deduplicar por factura: una factura tiene un solo contrato ---
    # Si múltiples filas de la misma factura tienen el mismo IDE actual y
    # debería ser el mismo, reportarlo una sola vez.
    agrupados: dict[tuple[str, str, str], dict[str, Any]] = {}
    for p in problemas:
        key = (p["factura"], p["ide_contrato_actual"], p["ide_contrato_deberia"])
        if key in agrupados:
            item = agrupados[key]
            item.setdefault("codigos_afectados", []).append(p["codigo"])
            # Mantener el procedimiento del primer hallazgo (es representativo)
        else:
            item = dict(p)
            item["codigos_afectados"] = [p["codigo"]]
            agrupados[key] = item

    resultado = list(agrupados.values())
    logger.info(
        "IDE Contrato Intramural: %d problemas crudos -> %d deduplicados por factura",
        len(problemas), len(resultado),
    )
    return resultado
