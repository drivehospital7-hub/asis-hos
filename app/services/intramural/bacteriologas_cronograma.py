"""Detector de bacteriólogas vs cronograma en facturación Intramural.

Valida que facturas Intramural con Tipo=02/05 + Laboratorio=Si tengan una
bacterióloga programada en el cronograma del día según Fec. Factura.
"""

from __future__ import annotations

import logging
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants.intramural import (
    PROFESIONALES_EXCEPTUADOS_CRONOGRAMA,
    RESPONSABLE_CHAPUEL,
    RESPONSABLE_ORDONEZ,
    RESPONSABLE_TAPIA,
)
from app.constants.urgencias import (
    EXCEPCIONES_BACTERIOLOGA,
    FACTURADORES_URGENCIAS,
    PROFESIONALES_URGENCIAS,
)
from app.services.cronograma_bacteriologas_service import get_turno_del_dia
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)

# Formatos de fecha aceptados para parseo local
_LOCAL_DATE_FORMATS = ("%d/%m/%Y", "%d-%m-%Y")


def _build_nombre_a_codigo() -> dict[str, str]:
    """Construye reverse lookup: nombre normalizado → código profesional.

    Recorre PROFESIONALES_URGENCIAS y mapea cada nombre (upper) a su
    código, para resolver nombres del cronograma a códigos de profesional.

    Además del nombre completo, indexa cada palabra individual para soportar
    matching parcial (el cronograma guarda solo el primer nombre, ej. "KAREN"
    mientras que el registro completo es "MADROÑERO BURBANO KAREN LIZETH").
    La primera coincidencia encontrada gana (evita sobrescritura).
    """
    lookup: dict[str, str] = {}
    for codigo, info in PROFESIONALES_URGENCIAS.items():
        nombre = info.get("nombre", "")
        if nombre:
            nombre_up = nombre.upper().strip()
            lookup[nombre_up] = codigo
            for word in nombre_up.split():
                if word not in lookup:
                    lookup[word] = codigo
    return lookup


# Reverse lookup construido UNA vez a nivel módulo (inmutable en runtime)
_NOMBRE_A_CODIGO: dict[str, str] = _build_nombre_a_codigo()


def _parse_fecha(val: Any) -> date | None:
    """Parsea un valor de fecha desde múltiples formatos.

    Args:
        val: Valor a parsear. Puede ser:
            - ISO string: "2024-03-15"
            - Excel serial: 45367 (int o float)
            - Local string: "15/03/2024" o "15-03-2024"
            - datetime/date object (Polars/calamine lo retorna como datetime)

    Returns:
        date object, o None si no se pudo parsear (log warning).
    """
    if val is None:
        return None
    # datetime o date object (Python nativo)
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val_stripped = val.strip()
        if not val_stripped:
            return None
        # ISO date format (sin hora)
        try:
            return datetime.strptime(val_stripped, "%Y-%m-%d").date()
        except ValueError:
            pass
        # ISO datetime format (con hora: "2026-06-01 06:15:00")
        try:
            return datetime.fromisoformat(val_stripped).date()
        except ValueError:
            pass
        # Local format
        for fmt in _LOCAL_DATE_FORMATS:
            try:
                return datetime.strptime(val_stripped, fmt).date()
            except ValueError:
                continue
        logger.warning("Fecha inválida (string no reconocido): %s", val)
        return None
    if isinstance(val, (int, float)):
        try:
            excel_epoch = datetime(1899, 12, 30)
            return (excel_epoch + timedelta(days=int(val))).date()
        except (ValueError, OverflowError):
            logger.warning("Fecha inválida (serial Excel inválido): %s", val)
            return None
    logger.warning("Fecha inválida (tipo no soportado): %s", type(val).__name__)
    return None


def _normalizar_laboratorio(val: Any) -> str:
    """Normaliza el valor de laboratorio a 'Si' o 'No'."""
    if val is None:
        return ""
    raw = str(val).strip().upper()
    # "SÍ" (con acento) y "SI" → "Si"
    if raw in ("SI", "SÍ"):
        return "Si"
    return raw.capitalize()


def detect_bacteriologas_cronograma(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
    responsable_cierra: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Detecta facturas Intramural con bacterióloga fuera del cronograma del día.

    Args:
        data_sheet: Hoja de Excel activa (openpyxl Worksheet).
        indices: Mapeo nombre_columna → índice 0-based (None si ausente).
        responsable_cierra: Mapa factura → nombre responsable cierra.
            Si es None o dict vacío, usa comportamiento default (CE|PYM).
            Chapuel → solo PYM. Tapia/Ordoñez → solo CE.
            Facturadores Urgencias → bypass de cronograma.

    Returns:
        Lista de errores con formato:
        {
            "factura": str,
            "codigo_profesional": str,
            "nombre_profesional": str,
            "procedimiento": str,
            "codigo": str,
            "regla": "Bacterióloga debe estar en cronograma del día",
            "problema": str,
            "fec_factura": str,
        }
    """
    num_fact_idx = indices.get("numero_factura")
    tipo_factura_idx = indices.get("tipo_factura_descripcion")
    codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")
    laboratorio_idx = indices.get("laboratorio")
    codigo_idx = indices.get("codigo")
    codigo_profesional_idx = indices.get("codigo_profesional")
    profesional_nombre_idx = indices.get("profesional_atiende")
    procedimiento_idx = indices.get("procedimiento")
    fec_factura_idx = indices.get("fec_factura")

    # Columnas indispensables
    if num_fact_idx is None or codigo_profesional_idx is None:
        logger.warning(
            "Bacteriólogas Cronograma - Columnas necesarias no encontradas"
        )
        return []

    errores: list[dict[str, Any]] = []
    facturas_con_error: set[str] = set()
    tipos_permitidos = frozenset({"02", "05"})

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura = normalize_invoice(numero_factura)
        if not factura:
            continue

        # If this factura already has an error, skip (dedup)
        if factura in facturas_con_error:
            continue

        # Read tipo_factura_descripcion
        if tipo_factura_idx is not None:
            tipo_factura_raw = data_sheet.cell(
                row=row, column=tipo_factura_idx + 1
            ).value
        else:
            tipo_factura_raw = None
        tipo_factura = str(tipo_factura_raw).strip() if tipo_factura_raw else ""

        # Filter: only Intramural
        if tipo_factura != "Intramural":
            continue

        # Read codigo_tipo_procedimiento
        if codigo_tipo_proc_idx is not None:
            tipo_proc_raw = data_sheet.cell(
                row=row, column=codigo_tipo_proc_idx + 1
            ).value
        else:
            tipo_proc_raw = None
        tipo_proc = str(tipo_proc_raw).strip() if tipo_proc_raw else ""

        # Filter: only 02 or 05
        if tipo_proc not in tipos_permitidos:
            continue

        # Read and normalize laboratorio
        if laboratorio_idx is not None:
            laboratorio_raw = data_sheet.cell(
                row=row, column=laboratorio_idx + 1
            ).value
        else:
            laboratorio_raw = None
        laboratorio = _normalizar_laboratorio(laboratorio_raw)

        # Filter: only Laboratorio == "Si"
        if laboratorio != "Si":
            continue

        # Read codigo (CUPS procedure code)
        if codigo_idx is not None:
            codigo_raw = data_sheet.cell(row=row, column=codigo_idx + 1).value
        else:
            codigo_raw = None
        codigo_str = str(codigo_raw).strip() if codigo_raw else ""

        # Check EXCEPCIONES_BACTERIOLOGA
        if codigo_str in EXCEPCIONES_BACTERIOLOGA:
            continue

        # Read codigo_profesional
        codigo_prof_val = data_sheet.cell(
            row=row, column=codigo_profesional_idx + 1
        ).value
        codigo_prof = str(codigo_prof_val).strip() if codigo_prof_val else ""

        if not codigo_prof:
            continue

        # Read nombre profesional
        if profesional_nombre_idx is not None:
            nombre_prof_val = data_sheet.cell(
                row=row, column=profesional_nombre_idx + 1
            ).value
        else:
            nombre_prof_val = None
        nombre_prof = str(nombre_prof_val).strip() if nombre_prof_val else codigo_prof

        # Read procedimiento description
        if procedimiento_idx is not None:
            proc_val = data_sheet.cell(
                row=row, column=procedimiento_idx + 1
            ).value
        else:
            proc_val = None
        procedimiento = str(proc_val).strip() if proc_val else ""

        # Read fec_factura (as raw string)
        if fec_factura_idx is not None:
            fec_raw = data_sheet.cell(row=row, column=fec_factura_idx + 1).value
        else:
            fec_raw = None
        fec_factura_str = str(fec_raw).strip() if fec_raw else ""

        # ── Validate against PROFESIONALES_URGENCIAS ──────────────
        profesional_info = PROFESIONALES_URGENCIAS.get(codigo_prof)

        if profesional_info is None:
            facturas_con_error.add(factura)
            errores.append({
                "factura": factura,
                "codigo_profesional": codigo_prof,
                "nombre_profesional": nombre_prof,
                "procedimiento": procedimiento,
                "codigo": codigo_str,
                "regla": "Bacterióloga debe estar en cronograma del día",
                "problema": (
                    f"El profesional {codigo_prof} no está en el listado "
                    "de Urgencias"
                ),
                "fec_factura": fec_factura_str,
            })
            continue

        tipo_profesional = profesional_info.get("tipo", "")
        nombre_profesional = profesional_info.get("nombre", nombre_prof)

        if tipo_profesional != "BACTERIOLOGA":
            facturas_con_error.add(factura)
            errores.append({
                "factura": factura,
                "codigo_profesional": codigo_prof,
                "nombre_profesional": nombre_profesional,
                "procedimiento": procedimiento,
                "codigo": codigo_str,
                "regla": "Bacterióloga debe estar en cronograma del día",
                "problema": (
                    f"El profesional {nombre_profesional} ({codigo_prof}) "
                    "no es una bacterióloga"
                ),
                "fec_factura": fec_factura_str,
            })
            continue

        # ★ PROFESIONALES_EXCEPTUADOS_CRONOGRAMA — bypass total de cronograma
        if codigo_prof in PROFESIONALES_EXCEPTUADOS_CRONOGRAMA:
            continue

        # ── Determine siglas_filter based on responsable_cierra ──
        resp = " ".join((responsable_cierra or {}).get(factura, "").upper().split())

        if resp in FACTURADORES_URGENCIAS:
            # Bypass total de cronograma para Urgencias
            continue

        if resp == RESPONSABLE_CHAPUEL:
            siglas_filter = {"PYM"}
        elif resp in {RESPONSABLE_TAPIA, RESPONSABLE_ORDONEZ}:
            siglas_filter = {"CE"}
        else:
            siglas_filter = None  # default: CE|PYM

        # ── Parse fec_factura and validate cronograma ─────────────
        fecha = _parse_fecha(fec_raw)
        if fecha is None:
            logger.warning(
                "Bacteriólogas Cronograma - Fecha inválida para factura %s: %s",
                factura,
                fec_factura_str,
            )
            continue

        if siglas_filter is not None:
            turnos = get_turno_del_dia(
                fecha.month, fecha.year, fecha.day,
                siglas_filter=siglas_filter,
            )
        else:
            turnos = get_turno_del_dia(fecha.month, fecha.year, fecha.day)
        if not turnos:
            # No hay cronograma o no hay turnos ese día → skip sin error
            continue

        # Resolve cronograma nombres → códigos profesionales via reverse lookup
        codigos_en_turno: set[str] = set()
        for t in turnos:
            nombre_turno = t.get("nombre", "").strip()
            if nombre_turno:
                cod = _NOMBRE_A_CODIGO.get(nombre_turno.upper())
                if cod:
                    codigos_en_turno.add(cod)

        if codigo_prof not in codigos_en_turno:
            facturas_con_error.add(factura)
            errores.append({
                "factura": factura,
                "codigo_profesional": codigo_prof,
                "nombre_profesional": nombre_profesional,
                "procedimiento": procedimiento,
                "codigo": codigo_str,
                "regla": "Bacterióloga debe estar en cronograma del día",
                "problema": (
                    f"Bacterióloga {nombre_profesional} ({codigo_prof}) "
                    f"no está en el cronograma del día "
                    f"{fecha.day}/{fecha.month}/{fecha.year}"
                ),
                "fec_factura": fec_factura_str,
            })
            continue

    if errores:
        logger.info(
            "Bacteriólogas Cronograma - Problemas encontrados: %d", len(errores)
        )

    return errores
