"""Detector de duplicados por identificación + código en Intramural.

Retorna un error por cada grupo duplicado (identificacion, codigo, dx).
Dos o más filas con mismo Nº Identificación + Código + Dx Principal se
consideran duplicado, EXCEPTO cuando el Responsable Cierra Facturar
es un facturador de urgencias (FACTURADORES_URGENCIAS).
"""

from __future__ import annotations

import logging
from collections import defaultdict
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

import unicodedata

from app.constants.urgencias import FACTURADORES_URGENCIAS
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)

_NORMALIZED_FACTURADORES: set[str] | None = None


def _get_normalized_facturadores() -> set[str]:
    """Retorna FACTURADORES_URGENCIAS sin acentos (lazy init)."""
    global _NORMALIZED_FACTURADORES
    if _NORMALIZED_FACTURADORES is None:
        result: set[str] = set()
        for name in FACTURADORES_URGENCIAS:
            s = unicodedata.normalize("NFKD", name.upper())
            s = s.encode("ascii", "ignore").decode("ascii")
            s = " ".join(s.split())
            result.add(s)
        _NORMALIZED_FACTURADORES = result
    return _NORMALIZED_FACTURADORES


def _normalize_responsable(raw: Any) -> str:
    """Normaliza nombre de responsable: mayúsculas, sin dobles espacios, sin acentos."""
    if not raw:
        return ""
    s = str(raw).strip().upper()
    s = " ".join(s.split())
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    return s


def detect_duplicado_id_codigo(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """Detecta grupos duplicados por (identificacion, codigo, dx_principal).

    Retorna un error por cada combinacion (ident, codigo, dx) que
    aparezca 2+ veces.

    Args:
        data_sheet: Hoja activa del Excel.
        indices: Mapeo nombre_columna → índice 0-based.

    Returns:
        Lista de dicts, uno por grupo duplicado. Vacía si no hay.
    """
    num_fact_idx = indices.get("numero_factura")
    ident_idx = indices.get("identificacion")
    codigo_idx = indices.get("codigo")
    dx_idx = indices.get("codigo_dx_principal")
    proc_idx = indices.get("procedimiento")
    resp_idx = indices.get("responsable_cierra")
    tipo_proc_idx = indices.get("codigo_tipo_procedimiento")
    laboratorio_idx = indices.get("laboratorio")

    if None in (num_fact_idx, ident_idx, codigo_idx):
        logger.warning("[BACK] Duplicado ID+Código - Columnas necesarias no encontradas")
        return []

    grupos: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    total_excluidas = 0

    for row in range(2, data_sheet.max_row + 1):
        numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura = normalize_invoice(numero)
        if not factura:
            continue

        ident_raw = data_sheet.cell(row=row, column=ident_idx + 1).value
        codigo_raw = data_sheet.cell(row=row, column=codigo_idx + 1).value

        ident_str = str(ident_raw).strip() if ident_raw is not None else ""
        codigo_str = str(codigo_raw).strip() if codigo_raw is not None else ""

        if not ident_str or not codigo_str:
            continue

        # Excluir filas facturadas por URGENCIAS
        if resp_idx is not None:
            resp_raw = data_sheet.cell(row=row, column=resp_idx + 1).value
            resp_norm = _normalize_responsable(resp_raw)
            if resp_norm in _get_normalized_facturadores():
                total_excluidas += 1
                continue

        procedimiento = (
            str(data_sheet.cell(row=row, column=proc_idx + 1).value or "").strip()
            if proc_idx is not None
            else ""
        )

        # Solo revisar duplicados si tipo=02+Lab=Si o tipo=05
        tipo_proc_raw = data_sheet.cell(row=row, column=tipo_proc_idx + 1).value if tipo_proc_idx is not None else None
        tipo_proc_str = str(tipo_proc_raw).strip() if tipo_proc_raw else ""

        lab_raw = data_sheet.cell(row=row, column=laboratorio_idx + 1).value if laboratorio_idx is not None else None
        lab_str = str(lab_raw).strip().upper() if lab_raw else ""

        if tipo_proc_str not in ("02", "05"):
            total_excluidas += 1
            continue
        if tipo_proc_str == "02" and lab_str != "SI":
            total_excluidas += 1
            continue

        dx_raw = data_sheet.cell(row=row, column=dx_idx + 1).value if dx_idx is not None else None
        dx_str = str(dx_raw).strip() if dx_raw else ""

        key = (ident_str, codigo_str, dx_str)
        grupos[key].append({
            "factura": factura,
            "identificacion": ident_str,
            "codigo": codigo_str,
            "dx_principal": dx_str,
            "procedimiento": procedimiento,
            "tipo_procedimiento": tipo_proc_str,
        })

    resultado: list[dict[str, Any]] = []
    total_grupos_dup = 0

    # Códigos exentos de duplicado para tipo 05
    CODIGOS_EXENTOS_05: set[str] = {"993505"}

    for key, filas in grupos.items():
        # Exentar si es tipo 05 con código en la lista
        if filas[0].get("tipo_procedimiento") == "05":
            if any(f.get("codigo") in CODIGOS_EXENTOS_05 for f in filas):
                continue
            umbral = 2
        else:
            umbral = 4

        if len(filas) < umbral:
            continue
        total_grupos_dup += 1
        facturas = sorted(set(f["factura"] for f in filas))

        resultado.append({
            "identificacion": filas[0]["identificacion"],
            "codigo": filas[0]["codigo"],
            "dx_principal": filas[0]["dx_principal"],
            "procedimiento": filas[0]["procedimiento"],
            "facturas": facturas,
            "cantidad_repeticiones": len(filas),
        })

    if resultado or total_excluidas:
        logger.info(
            "[BACK] Duplicado ID+Código - %d grupos, %d excluidas por ser URGENCIAS",
            total_grupos_dup, total_excluidas,
        )

    return resultado
