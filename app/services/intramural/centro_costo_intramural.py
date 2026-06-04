"""Detector de problemas de centro de costo para Intramural.

Aplica reglas comunes (1-4, 8, 9) + reglas específicas de Intramural:
- REGLA3-INTRAMURAL: Código PyP → SERVICIOS AMBULATORIOS- PROMOCION Y PREVENCION
- REVERSE3-INTRAMURAL: Centro=PyP Intramural → Código en lista PyP
- REGLA6: Código Tipo Procedimiento=05 + código != 906249PR → SALUD PUBLICA-VACUNACION
- REVERSE6: Centro=SALUD PUBLICA-VACUNACION → Tipo=05 + código != 906249PR
- REGLA7: Código Tipo Procedimiento=03 o 04 → SERVICIOS AMBULATORIOS
- REVERSE7: Centro=SERVICIOS AMBULATORIOS → Tipo=03 o 04
- REGLA10: Código Tipo Procedimiento=02 o 05 + Lab=Si → LABORATORIO CLINICO
- REVERSE10: Centro=LABORATORIO CLINICO → Tipo=02 o 05 + Lab=Si
- REGLA_RESPONSABLE_URGENCIAS: Responsable en FACTURADORES_URGENCIAS + Tipo 01/04 → URGENCIAS o HOSPITALIZACIÓN

Se filtra la REGLA3 común que usa "PROCEDIMIENTO DE PROMOCIÓN Y PREVENCIÓN"
(centro de Urgencias) y se reemplaza con la regla local de Intramural.
"""

from __future__ import annotations

import logging
import re
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    CENTRO_COSTO_AMBULATORIO,
    CENTRO_COSTO_HOSPITALIZACION_ESTANCIA,
    CENTRO_COSTO_SALUD_PUBLICA,
    CENTRO_COSTO_URGENCIAS,
    CENTROS_COSTO_LABORATORIO_VALIDOS,
    CENTROS_COSTO_PYP_INTRAMURAL,
    CODIGO_TIPO_PROCEDIMIENTO_VACUNACION,
    CODIGOS_EXCEPTUADOS,
    CODIGOS_EXCEPTUADOS_AMBULATORIO,
    CODIGOS_EXCEPTUADOS_RESPONSABLE_URGENCIAS,
    CODIGOS_EXCLUIDOS_VACUNACION,
    CODIGOS_PYP_URGENCIAS,
    CODIGOS_TIPO_PROCEDIMIENTO_AMBULATORIO,
    CODIGOS_TIPO_PROCEDIMIENTO_LABORATORIO,
    FACTURADORES_URGENCIAS,
    INTRAMURAL_CENTROS_COSTO_VALIDOS,
    LABORATORIO_SI,
)
from app.services.transversales.centro_costo_rules import apply_common_centro_costo_rules
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)

# REGLA3 común usa "PROCEDIMIENTO DE PROMOCIÓN Y PREVENCIÓN" (Urgencias),
# en Intramural se reemplaza con una regla local con centros PyP propios.
_INTRAMURAL_RULES_SKIP: frozenset[str] = frozenset({
    "REGLA3", "REVERSE3",
})


def detect_centro_costo_intramural(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """Detecta facturas de Intramural con problemas de centro de costo.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: factura, tipo_factura, centro_actual,
        centro_deberia, codigo, procedimiento, prioridad, regla
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")
    codigo_idx = indices.get("codigo")
    laboratorio_idx = indices.get("laboratorio")
    centro_costo_idx = indices.get("centro_costo")
    codigo_entidad_cobrar_idx = indices.get("codigo_entidad_cobrar")
    tipo_factura_descripcion_idx = indices.get("tipo_factura_descripcion")
    proc_idx = indices.get("procedimiento")
    tarifario_idx = indices.get("tarifario")
    responsable_cierra_idx = indices.get("responsable_cierra")

    if num_fact_idx is None or centro_costo_idx is None:
        logger.warning("Centro Costo Intramural - Columnas necesarias no encontradas")
        return []

    problemas_centros: list[dict[str, Any]] = []

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        tipo_factura_descripcion = (
            data_sheet.cell(row=row, column=tipo_factura_descripcion_idx + 1).value
            if tipo_factura_descripcion_idx is not None else None
        )
        tipo_factura_str = str(tipo_factura_descripcion).strip() if tipo_factura_descripcion else ""

        # Only process Intramural rows
        if tipo_factura_str != "Intramural":
            continue

        # Read row values
        codigo_tipo_proc = (
            data_sheet.cell(row=row, column=codigo_tipo_proc_idx + 1).value
            if codigo_tipo_proc_idx is not None else None
        )
        codigo = (
            data_sheet.cell(row=row, column=codigo_idx + 1).value
            if codigo_idx is not None else None
        )
        laboratorio = (
            data_sheet.cell(row=row, column=laboratorio_idx + 1).value
            if laboratorio_idx is not None else None
        )
        centro_costo = data_sheet.cell(row=row, column=centro_costo_idx + 1).value
        codigo_entidad_cobrar = (
            data_sheet.cell(row=row, column=codigo_entidad_cobrar_idx + 1).value
            if codigo_entidad_cobrar_idx is not None else None
        )
        procedimiento = (
            data_sheet.cell(row=row, column=proc_idx + 1).value
            if proc_idx is not None else None
        )
        tarifario = (
            data_sheet.cell(row=row, column=tarifario_idx + 1).value
            if tarifario_idx is not None else None
        )
        responsable_cierra = (
            data_sheet.cell(row=row, column=responsable_cierra_idx + 1).value
            if responsable_cierra_idx is not None else None
        )
        codigo_str = str(codigo_tipo_proc).strip() if codigo_tipo_proc else ""
        codigo_excluir = str(codigo).strip() if codigo else ""
        laboratorio_str = str(laboratorio).strip() if laboratorio else ""
        centro_costo_str = str(centro_costo).strip() if centro_costo else ""
        codigo_entidad_str = str(codigo_entidad_cobrar).strip() if codigo_entidad_cobrar else ""
        proc_str = str(procedimiento).strip() if procedimiento else ""
        tarifario_str = str(tarifario).strip() if tarifario else ""
        responsable_cierra_str = re.sub(r'\s+', ' ', str(responsable_cierra)).strip().upper() if responsable_cierra else ""

        # Apply common rules with Intramural valid centers
        errors = apply_common_centro_costo_rules(
            centro_costo_str=centro_costo_str,
            codigo_str=codigo_str,
            codigo_excluir=codigo_excluir,
            laboratorio_str=laboratorio_str,
            tarifario_str=tarifario_str,
            codigo_entidad_str=codigo_entidad_str,
            factura_str=factura_str,
            proc_str=proc_str,
            centros_validos=INTRAMURAL_CENTROS_COSTO_VALIDOS,
        )

        for e in errors:
            # Filter out REGLA3 común (usa centro Urgencias, no Intramural)
            if e.get("regla") in _INTRAMURAL_RULES_SKIP:
                continue
            e["tipo_factura"] = tipo_factura_str
            problemas_centros.append(e)

        # --- REGLA3 INTRAMURAL: Código PyP → SERVICIOS AMBULATORIOS- PROMOCION Y PREVENCION ---
        if codigo_excluir in CODIGOS_PYP_URGENCIAS:
            if centro_costo_str not in CENTROS_COSTO_PYP_INTRAMURAL:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": " o ".join(sorted(CENTROS_COSTO_PYP_INTRAMURAL)),
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "REGLA3-INTRAMURAL",
                })

        # --- REVERSE3 INTRAMURAL: Centro=SERVICIOS AMBULATORIOS- PROMOCION → Código PyP ---
        if centro_costo_str in CENTROS_COSTO_PYP_INTRAMURAL:
            if codigo_excluir not in CODIGOS_PYP_URGENCIAS:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": "Código CUPS debe estar en lista de Promoción y Prevención",
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "REVERSE3-INTRAMURAL",
                })

        # --- REGLA6: Tipo=05 + código != 906249PR + NO PyP → SALUD PÚBLICA ---
        if (
            codigo_str == CODIGO_TIPO_PROCEDIMIENTO_VACUNACION
            and codigo_excluir not in CODIGOS_EXCLUIDOS_VACUNACION
            and codigo_excluir not in CODIGOS_PYP_URGENCIAS
            and centro_costo_str != CENTRO_COSTO_SALUD_PUBLICA
        ):
            problemas_centros.append({
                "factura": factura_str,
                "tipo_factura": tipo_factura_str,
                "centro_actual": centro_costo_str,
                "centro_deberia": CENTRO_COSTO_SALUD_PUBLICA,
                "codigo": codigo_excluir,
                "procedimiento": proc_str,
                "prioridad": 1,
                "regla": "REGLA6",
            })

        # --- REVERSE6: Centro=SALUD PÚBLICA → Tipo=05 + código != 906249PR ---
        if centro_costo_str == CENTRO_COSTO_SALUD_PUBLICA:
            if codigo_str != CODIGO_TIPO_PROCEDIMIENTO_VACUNACION or codigo_excluir in CODIGOS_EXCLUIDOS_VACUNACION:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": (
                        f"Código Tipo Procedimiento={CODIGO_TIPO_PROCEDIMIENTO_VACUNACION} "
                        f"y código no en {sorted(CODIGOS_EXCLUIDOS_VACUNACION)}"
                    ),
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "REVERSE6",
                })

        # --- REGLA7: Código Tipo Procedimiento=03 o 04 → SERVICIOS AMBULATORIOS ---
        # Códigos exceptuados (735301=QUIRÓFANO, 861101=URGENCIAS) no aplican REGLA7
        if (
            codigo_str in CODIGOS_TIPO_PROCEDIMIENTO_AMBULATORIO
            and codigo_excluir not in CODIGOS_EXCEPTUADOS_AMBULATORIO
            and centro_costo_str != CENTRO_COSTO_AMBULATORIO
        ):
            problemas_centros.append({
                "factura": factura_str,
                "tipo_factura": tipo_factura_str,
                "centro_actual": centro_costo_str,
                "centro_deberia": CENTRO_COSTO_AMBULATORIO,
                "codigo": codigo_excluir,
                "procedimiento": proc_str,
                "prioridad": 1,
                "regla": "REGLA7",
            })

        # --- REVERSE7: Centro=SERVICIOS AMBULATORIOS → Tipo=03 o 04 ---
        if centro_costo_str == CENTRO_COSTO_AMBULATORIO:
            if codigo_str not in CODIGOS_TIPO_PROCEDIMIENTO_AMBULATORIO:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": (
                        f"Código Tipo Procedimiento en "
                        f"{sorted(CODIGOS_TIPO_PROCEDIMIENTO_AMBULATORIO)}"
                    ),
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "REVERSE7",
                })

        # --- REGLA10: Tipo=02 o 05 + Lab=Si → LABORATORIO CLINICO ---
        if (
            codigo_str in CODIGOS_TIPO_PROCEDIMIENTO_LABORATORIO
            and laboratorio_str == LABORATORIO_SI
            and centro_costo_str not in CENTROS_COSTO_LABORATORIO_VALIDOS
        ):
            problemas_centros.append({
                "factura": factura_str,
                "tipo_factura": tipo_factura_str,
                "centro_actual": centro_costo_str,
                "centro_deberia": " o ".join(sorted(CENTROS_COSTO_LABORATORIO_VALIDOS)),
                "codigo": codigo_excluir,
                "procedimiento": proc_str,
                "prioridad": 1,
                "regla": "REGLA10",
            })

        # --- REVERSE10: Centro=LABORATORIO CLINICO → Tipo=02 o 05 + Lab=Si ---
        # Excepción: códigos en CODIGOS_EXCEPTUADOS son laboratorio pero tienen Lab=No
        if centro_costo_str in CENTROS_COSTO_LABORATORIO_VALIDOS:
            es_exceptuado_lab = codigo_excluir in CODIGOS_EXCEPTUADOS
            if codigo_str not in CODIGOS_TIPO_PROCEDIMIENTO_LABORATORIO or (
                not es_exceptuado_lab and laboratorio_str != LABORATORIO_SI
            ):
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": (
                        f"Código Tipo Procedimiento en "
                        f"{sorted(CODIGOS_TIPO_PROCEDIMIENTO_LABORATORIO)} "
                        f"y Laboratorio={LABORATORIO_SI}"
                    ),
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "REVERSE10",
                })

        # Responsable Facturador URGENCIAS + Tipo 01/04 → Centro URGENCIAS o HOSPITALIZACIÓN
        # Excepción: códigos en CODIGOS_EXCEPTUADOS_RESPONSABLE_URGENCIAS (735301=QUIRÓFANO)
        CENTROS_VALIDOS_PARA_RESPONSABLE = {CENTRO_COSTO_URGENCIAS, CENTRO_COSTO_HOSPITALIZACION_ESTANCIA}
        if (
            responsable_cierra_str in FACTURADORES_URGENCIAS
            and codigo_str in ("01", "04")
            and codigo_excluir not in CODIGOS_EXCEPTUADOS_RESPONSABLE_URGENCIAS
            and centro_costo_str not in CENTROS_VALIDOS_PARA_RESPONSABLE
        ):
            problemas_centros.append({
                "factura": factura_str,
                "tipo_factura": tipo_factura_str,
                "centro_actual": centro_costo_str,
                "centro_deberia": "URGENCIAS o HOSPITALIZACIÓN - ESTANCIA GENERAL",
                "codigo": codigo_excluir,
                "procedimiento": proc_str,
                "prioridad": 2,
                "regla": "REGLA_RESPONSABLE_URGENCIAS",
            })

    if problemas_centros:
        logger.info("Centro Costo Intramural - Problemas encontrados: %d", len(problemas_centros))

    return problemas_centros
