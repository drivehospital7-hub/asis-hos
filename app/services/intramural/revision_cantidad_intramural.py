"""Detección de revisión necesaria por cantidad anómala en Intramural.

Sigue el patrón de Urgencias (app/services/urgencias/revision_cantidad.py)
pero simplificado: sin filtro tipo_factura, sin tablas exento/límite-específico.
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants.intramural import (
    CANTIDAD_MAX_02_NO_LAB,
    CANTIDAD_MAX_03_04,
    CANTIDAD_MAX_GENERAL_INTRAMURAL,
    CODIGOS_LIMITE_ESPECIFICO_INTRAMURAL,
    CODIGOS_TIPO_PROC_03_04,
    CODIGO_TIPO_PROC_02,
    LABORATORIO_NO,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def _read_cell_str(
    data_sheet: Worksheet,
    row: int,
    col_idx: int | None,
) -> str:
    """Lee una celda y retorna string normalizado, o vacío si no hay columna."""
    if col_idx is None:
        return ""
    value = data_sheet.cell(row=row, column=col_idx + 1).value
    return str(value).strip() if value else ""


def _apply_cascade(
    cantidad: int | float,
    codigo_tipo_proc_str: str,
    laboratorio_str: str,
) -> str | None:
    """Aplica la cascade de reglas y retorna el detalle si se excede el umbral.

    Returns:
        String detalle si debe flaggearse, None si pasa sin marcar.
    """
    # Rule 1: 02 + Lab=No → max 2
    if codigo_tipo_proc_str == CODIGO_TIPO_PROC_02 and laboratorio_str == LABORATORIO_NO:
        if cantidad > CANTIDAD_MAX_02_NO_LAB:
            return (
                f"Cant: {cantidad} — Cód.Tipo=02, Lab=No, "
                f"máx={CANTIDAD_MAX_02_NO_LAB}"
            )
        return None

    # Rule 2: 03/04 → max 12
    if codigo_tipo_proc_str in CODIGOS_TIPO_PROC_03_04:
        if cantidad > CANTIDAD_MAX_03_04:
            return (
                f"Cant: {cantidad} — Cód.Tipo={codigo_tipo_proc_str}, "
                f"máx={CANTIDAD_MAX_03_04}"
            )
        return None

    # Rule 3: General → max 1
    if cantidad > CANTIDAD_MAX_GENERAL_INTRAMURAL:
        return (
            f"Cant: {cantidad} — Regla general, "
            f"máx={CANTIDAD_MAX_GENERAL_INTRAMURAL}"
        )
    return None


def _evaluate_row(
    data_sheet: Worksheet,
    row: int,
    indices: dict[str, int | None],
) -> dict[str, Any] | None:
    """Evalúa una fila contra la cascade de reglas.
    Returns dict flagged o None si pasa sin marcar.
    """
    idx = indices.get
    num_idx = idx("numero_factura")
    can_idx = idx("cantidad")
    cod_idx = idx("codigo")
    proc_idx = idx("procedimiento")
    tipo_idx = idx("codigo_tipo_procedimiento")
    lab_idx = idx("laboratorio")

    factura_str = normalize_invoice(
        data_sheet.cell(row=row, column=num_idx + 1).value if num_idx is not None else None
    )
    if not factura_str:
        return None

    cantidad = data_sheet.cell(row=row, column=can_idx + 1).value  # type: ignore[arg-type]
    if isinstance(cantidad, str):
        cantidad = cantidad.strip()
        if cantidad == "":
            return None
        try:
            cantidad = float(cantidad)
        except ValueError:
            return None
    elif not isinstance(cantidad, (int, float)):
        return None

    codigo_str = ""
    if cod_idx is not None:
        v = data_sheet.cell(row=row, column=cod_idx + 1).value
        codigo_str = str(v).strip().upper() if v else ""

    # Verificar si el código tiene un límite específico (evalúa antes de la cascade)
    if codigo_str in CODIGOS_LIMITE_ESPECIFICO_INTRAMURAL:
        max_cant = CODIGOS_LIMITE_ESPECIFICO_INTRAMURAL[codigo_str]
        if cantidad <= max_cant:
            return None
        # Si excede el límite específico, cae a la cascade

    procedimiento = _read_cell_str(data_sheet, row, proc_idx)
    codigo_tipo_proc_str = _read_cell_str(data_sheet, row, tipo_idx)
    laboratorio_str = _read_cell_str(data_sheet, row, lab_idx)

    result = _apply_cascade(cantidad, codigo_tipo_proc_str, laboratorio_str)
    if result is None:
        return None

    return {
        "factura": factura_str,
        "codigo": codigo_str,
        "procedimiento": procedimiento,
        "cantidad": cantidad,
        "codigo_tipo_procedimiento": codigo_tipo_proc_str,
        "laboratorio": laboratorio_str,
        "detalle": result,
    }


def detect_revision_cantidad_intramural(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """Detecta filas con cantidad anómala que requieren revisión manual.

    Reglas (cascade, first-match wins):
    1. Código Tipo Procedimiento = "02" AND Laboratorio = "No" → Cantidad ≤ 2
    2. Código Tipo Procedimiento = "03" o "04" → Cantidad ≤ 12
    3. General → Cantidad ≤ 1

    Returns:
        Lista de dicts con keys: 'factura', 'codigo', 'procedimiento',
        'cantidad', 'codigo_tipo_procedimiento', 'laboratorio', 'detalle'.
        Vacía si falta la columna 'Cantidad'.
    """
    cantidad_idx = indices.get("cantidad")
    if cantidad_idx is None:
        logger.warning(
            "[BACK] Revision Cantidad Intramural - "
            "Columna 'Cantidad' no encontrada, retornando []"
        )
        return []

    revision_items: list[dict[str, Any]] = []

    for row in range(2, data_sheet.max_row + 1):
        item = _evaluate_row(data_sheet, row, indices)
        if item is not None:
            revision_items.append(item)

    logger.info(
        "[BACK] Revision Cantidad Intramural - Filas procesadas: %d, "
        "Items encontrados: %d",
        data_sheet.max_row - 1,
        len(revision_items),
    )
    return revision_items
