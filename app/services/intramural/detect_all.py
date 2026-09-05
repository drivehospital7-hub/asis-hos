"""Orquestador de detección de problemas para Intramural.

Agrupa detectores transversales + específicos de Intramural.
"""

from __future__ import annotations

import logging
from typing import Any, Callable

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import AREA_INTRAMURAL
from app.constants.base import is_evidence_audit_enabled, is_rule_engine_enabled
from app.services.transversales import (
    normalize_invoice,
)
from app.services.normalized_rows import build_normalized_rows

# Module-level flag: skip evidence/audit DB writes when testing
_PERSIST = is_evidence_audit_enabled()
logger = logging.getLogger(__name__)


def _get_intramural_detectors() -> list[Callable]:
    """Returns list of Intramural-specific detector callables.
    
    Used by tipo_factura_registry for lazy import.
    """
    from app.services.intramural.bacteriologas_cronograma import (
        detect_bacteriologas_cronograma,
    )
    from app.services.intramural.centro_costo_intramural import (
        detect_centro_costo_intramural,
    )
    from app.services.intramural.duplicado_id_codigo import (
        detect_duplicado_id_codigo,
    )
    from app.services.intramural.ide_contrato_intramural import (
        detect_ide_contrato_intramural,
    )
    return [detect_bacteriologas_cronograma, detect_centro_costo_intramural, detect_ide_contrato_intramural, detect_duplicado_id_codigo]


def detect_all_problems_intramural(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> tuple[dict[str, Any], dict[str, str]]:
    """Detecta TODOS los problemas en facturas de Intramural.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        (resultado_dict, responsables_map)
    """
    from app.services.transversales import (
        detect_decimales,
        detect_tipo_documento_edad,
        detect_tipo_identificacion_entidad,
        detect_codigo_entidad_vs_entidad_afiliacion,
        detect_tipo_usuario,
    )
    from app.services.transversales.detect_copago_entidad import (
        detect_copago_entidad_urgencias,
    )
    from app.services.transversales.procedimiento_contratado import detect_cups_sin_contrato

    # 1. Detectores transversales (con toggle engine)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            decimales = RuleBasedDetector("valores_decimales", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    else:
        decimales = []
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            r1 = RuleBasedDetector("tipo_documento_edad_menor_7", session).detect(data_sheet, indices, persist=_PERSIST)
            r2 = RuleBasedDetector("tipo_documento_edad_mayor_18", session).detect(data_sheet, indices, persist=_PERSIST)
            r3 = RuleBasedDetector("tipo_documento_edad_7_17", session).detect(data_sheet, indices, persist=_PERSIST)
            r4 = RuleBasedDetector("tipo_documento_edad_as_menor", session).detect(data_sheet, indices, persist=_PERSIST)
            r5 = RuleBasedDetector("tipo_documento_edad_ms_mayor", session).detect(data_sheet, indices, persist=_PERSIST)
            r6 = RuleBasedDetector("tipo_documento_edad_cn_invalido", session).detect(data_sheet, indices, persist=_PERSIST)
            r7 = RuleBasedDetector("tipo_documento_edad_ce_invalido", session).detect(data_sheet, indices, persist=_PERSIST)
            tipo_identificacion_edad = r1 + r2 + r3 + r4 + r5 + r6 + r7
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    else:
        tipo_identificacion_edad = []
    tipo_identificacion_entidad = detect_tipo_identificacion_entidad(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            r1 = RuleBasedDetector("tipo_id_requiere_entidad_86000", session).detect(data_sheet, indices, persist=_PERSIST)
            r2 = RuleBasedDetector("entidad_86000_requiere_as_ms", session).detect(data_sheet, indices, persist=_PERSIST)
            tipo_identificacion_entidad = r1 + r2
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    entidad_afiliacion_comparison = detect_codigo_entidad_vs_entidad_afiliacion(
        data_sheet, indices, limit_log=5
    )
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            entidad_afiliacion_comparison = RuleBasedDetector("codigo_entidad", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    tipo_usuario = detect_tipo_usuario(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            tipo_usuario = RuleBasedDetector("tipo_usuario_valido", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    copago_entidad = detect_copago_entidad_urgencias(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            copago_entidad = RuleBasedDetector("copago_entidad_valido", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    cups_sin_contrato = detect_cups_sin_contrato(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            cups_sin_contrato = RuleBasedDetector("cups_sin_contrato", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()

    # 2. Build responsable_cierra mapping
    responsable_cierra: dict[str, str] = {}
    responsable_cierra_idx = indices.get("responsable_cierra")
    num_fact_idx = indices.get("numero_factura")
    if responsable_cierra_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            raw = data_sheet.cell(row=row, column=responsable_cierra_idx + 1).value
            resp = str(raw).strip() if raw else ""
            if resp and factura not in responsable_cierra:
                responsable_cierra[factura] = resp

    # 3. Build fecha_cierre_vacia mapping
    fecha_cierre_vacia: dict[str, bool] = {}
    fecha_cierre_idx = indices.get("fecha_cierre")
    if fecha_cierre_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            fecha_cierre_val = data_sheet.cell(row=row, column=fecha_cierre_idx + 1).value
            if not fecha_cierre_val or str(fecha_cierre_val).strip() == "":
                fecha_cierre_vacia[factura] = True
            elif factura not in fecha_cierre_vacia:
                fecha_cierre_vacia[factura] = False

    # 4. Build fec_factura_map
    fec_factura_map: dict[str, str] = {}
    fec_factura_idx = indices.get("fec_factura")
    if fec_factura_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            raw = data_sheet.cell(row=row, column=fec_factura_idx + 1).value
            val = str(raw).strip() if raw else ""
            if val and factura not in fec_factura_map:
                fec_factura_map[factura] = val

    # 5. Bacteriólogas Cronograma (con toggle engine)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            bacteriologas = RuleBasedDetector("bacteriologas_cronograma", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    else:
        bacteriologas = []

    # 6. Centro de Costo (con toggle engine)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            problemas_centros = RuleBasedDetector("centro_costo_intramural_valido", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    else:
        problemas_centros = []
    logger.info(
        "Centros de Costo Intramural - Problemas encontrados: %d",
        len(problemas_centros),
    )

    # Filtrar centros de costo por prioridad
    errores_por_factura_codigo: dict[tuple[str, str], list[tuple[dict, int]]] = {}
    for item in problemas_centros:
        key = (item.get("factura", ""), item.get("codigo", ""))
        prioridad = item.get("prioridad", 1)
        if key not in errores_por_factura_codigo:
            errores_por_factura_codigo[key] = []
        errores_por_factura_codigo[key].append((item, prioridad))

    problemas_centros_filtrados = []
    for key, items in errores_por_factura_codigo.items():
        prioridades = [p for _, p in items]
        if 1 in prioridades:
            for item, p in items:
                if p == 1:
                    problemas_centros_filtrados.append(item)
        else:
            for item, _ in items:
                problemas_centros_filtrados.append(item)

    logger.info(
        "FILTRO centros_de_costos Intramural: %d -> %d",
        len(problemas_centros),
        len(problemas_centros_filtrados),
    )

    # 7. IDE Contrato (con toggle engine)
    # Nuevos evaluadores: IdeContratoSimpleEvaluator + PymRutasDxEvaluator
    # El pre-scan de laboratorio envío se maneja via PymRutasDxEvaluator.pre_scan_sheet()
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        from app.services.engine.evaluators import PymRutasDxEvaluator
        session = get_session()
        try:
            # Pre-scan: detectar facturas con solo laboratorio de envío
            pym_ev = PymRutasDxEvaluator()
            pym_ev.pre_scan_sheet(data_sheet, indices)

            # Evaluar reglas vía engine
            problemas_ide_contrato = []
            problemas_ide_contrato.extend(
                RuleBasedDetector("ide_contrato_simple", session).detect(data_sheet, indices, persist=_PERSIST)
            )
            problemas_ide_contrato.extend(
                RuleBasedDetector("pym_rutas_dx", session).detect(data_sheet, indices, persist=_PERSIST)
            )
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    else:
        from app.services.intramural.ide_contrato_intramural import (
            detect_ide_contrato_intramural,
        )
        try:
            problemas_ide_contrato = detect_ide_contrato_intramural(data_sheet, indices)
        except Exception:
            logger.exception("Error en detect_ide_contrato_intramural")
            problemas_ide_contrato = []

    # 8. Duplicado ID+Código (con toggle engine)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        from app.constants.urgencias import FACTURADORES_URGENCIAS
        session = get_session()
        try:
            r1 = RuleBasedDetector("duplicado_id_codigo_05", session).detect(data_sheet, indices, persist=_PERSIST)
            r2 = RuleBasedDetector("duplicado_id_codigo_02_lab", session).detect(data_sheet, indices, persist=_PERSIST)
            raw_results = r1 + r2
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    else:
        duplicado_id_codigo = []

    if is_rule_engine_enabled():
        # Post-process engine results to match legacy format
        # Engine output for group-by rules has factura=composite_key
        # We need to split it back to (identificacion, codigo, dx) and
        # apply FACTURADORES_URGENCIAS + CODIGOS_EXENTOS_05 filters
        from app.constants.urgencias import FACTURADORES_URGENCIAS
        _FACTURADORES_NORM = frozenset(
            " ".join(f.upper().split()) for f in FACTURADORES_URGENCIAS
        )
        CODIGOS_EXENTOS_05: set[str] = {"993505"}
        duplicado_id_codigo = []
        for item in raw_results:
            # Parse composite key: "ident\tcodigo\tdx"
            key_str = item.get("factura", "")
            if "\t" in key_str:
                parts = key_str.split("\t")
                ident = parts[0] if len(parts) > 0 else ""
                codigo = parts[1] if len(parts) > 1 else ""
                dx = parts[2] if len(parts) > 2 else ""
            else:
                # Fallback for single-field group_by (legacy format)
                continue

            facturas_list = item.get("facturas", [])
            if not isinstance(facturas_list, list):
                continue

            # Check CODIGOS_EXENTOS_05 for tipo=05 rules
            rule_name = item.get("regla", "")
            if "05" in rule_name and codigo in CODIGOS_EXENTOS_05:
                continue

            # Check FACTURADORES_URGENCIAS: if ALL facturas have facturador → skip
            all_urgencias = True
            for f in facturas_list:
                resp = responsable_cierra.get(f, "")
                resp_norm = " ".join(resp.upper().split()) if resp else ""
                if resp_norm not in _FACTURADORES_NORM:
                    all_urgencias = False
                    break
            if all_urgencias and facturas_list:
                continue

            # Build legacy-style output
            procedimiento = item.get("procedimiento", "")
            duplicado_id_codigo.append({
                "identificacion": ident,
                "codigo": codigo,
                "dx_principal": dx,
                "procedimiento": procedimiento,
                "facturas": sorted(set(facturas_list)),
                "cantidad_repeticiones": item.get("count", len(facturas_list)),
            })
        logger.info(
            "[ENGINE] Duplicado ID+Código: %d problemas (raw: %d)",
            len(duplicado_id_codigo), len(raw_results),
        )

    # 9. Revision Cantidad Intramural (con toggle engine)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            revision_cantidad = RuleBasedDetector("revision_cantidad_intramural", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    else:
        revision_cantidad = []
    logger.info(
        "[BACK] Revision Cantidad Intramural: %d items",
        len(revision_cantidad),
    )

    # 10. Build normalized rows
    error_groups = {
        "Centros de Costo": problemas_centros_filtrados,
        "Decimales": decimales,
        "Tipo Identificación / Edad": tipo_identificacion_edad,
        "Código Entidad vs Afiliación": entidad_afiliacion_comparison + tipo_identificacion_entidad,
        "Tipo Usuario": tipo_usuario,
        "Copago vs Entidad": copago_entidad,
        "IDE Contrato": problemas_ide_contrato,
        "Cups Sin Contrato": cups_sin_contrato,
        "Profesionales": bacteriologas,
        "Duplicado ID+Código": duplicado_id_codigo,
        "⚠️ Revisión Necesaria": revision_cantidad,
    }
    normalized_rows = build_normalized_rows(
        error_groups=error_groups,
        responsables_map=responsable_cierra,
        fec_factura_map=fec_factura_map,
        fecha_cierre_vacia_map=fecha_cierre_vacia,
    )

    # 11. Build resultado
    resultado: dict[str, Any] = {
        "area": AREA_INTRAMURAL,
        "problemas": {
            "normalizados": normalized_rows,
            "centros_de_costos": [
                {
                    "tipo_factura": item.get("tipo_factura") or "-",
                    "factura": item["factura"],
                    "codigo": item.get("codigo", ""),
                    "procedimiento": item.get("procedimiento", ""),
                    "centro_actual": item.get("centro_actual", item.get("centro_costo", "")),
                    "centro_deberia": item.get("centro_deberia", ""),
                    "prioridad": item.get("prioridad", 1),
                }
                for item in problemas_centros_filtrados
            ],
            "ide_contrato": problemas_ide_contrato,
            "decimales": decimales,
            "tipo_identificacion_edad": tipo_identificacion_edad,
            "tipo_identificacion_entidad": tipo_identificacion_entidad,
            "codigo_entidad_vs_afiliacion": entidad_afiliacion_comparison,
            "tipo_usuario": tipo_usuario,
            "copago_entidad": copago_entidad,
            "cups_sin_contrato": cups_sin_contrato,
            "profesionales": bacteriologas,
            "duplicado_id_codigo": duplicado_id_codigo,
            "revision_cantidad": revision_cantidad,
        },
        "totales": {
            "centros_de_costos": len(problemas_centros),
            "ide_contrato": len(problemas_ide_contrato),
            "decimales": len(decimales),
            "tipo_identificacion_edad": len(tipo_identificacion_edad),
            "tipo_identificacion_entidad": len(tipo_identificacion_entidad),
            "codigo_entidad_vs_afiliacion": len(entidad_afiliacion_comparison),
            "tipo_usuario": len(tipo_usuario),
            "copago_entidad": len(copago_entidad),
            "cups_sin_contrato": len(cups_sin_contrato),
            "profesionales": len(bacteriologas),
            "duplicado_id_codigo": len(duplicado_id_codigo),
            "revision_cantidad": len(revision_cantidad),
        },
        "missing_columns": [],
    }

    # 12. Enrich errors with responsable
    for problem_type, problems in resultado["problemas"].items():
        for p in problems:
            if not isinstance(p, dict):
                continue
            factura = p.get("factura")
            if factura and factura in responsable_cierra:
                p["responsable"] = responsable_cierra[factura]
            elif "responsable" not in p:
                p["responsable"] = ""

    return resultado, responsable_cierra
