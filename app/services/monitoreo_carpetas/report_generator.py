"""Generador de reportes Excel para Monitoreo de Carpetas.

Usa openpyxl con estilos de app/utils/formatting.py.
Genera un Workbook con 2 hojas: Facturas (detalle) e Indicadores (resumen).
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from app.services.monitoreo_carpetas import InvoiceRecord, ScanResult
from app.utils.formatting import (
    auto_adjust_column_width,
    create_data_row_style,
    create_header_style,
)

logger = logging.getLogger(__name__)

# Yellow highlight for rows with anomalies
_ANOMALY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _write_facturas_sheet(wb: Workbook, result: ScanResult) -> None:
    """Escribe la hoja Facturas con datos detallados por factura."""
    ws = wb.active
    ws.title = "Facturas"

    headers = [
        "Código Factura",
        "Tipo",
        "Estado",
        "Ruta Completa",
        "Facturador",
        "Fecha Escaneo",
        "Duplicado",
        "Carpeta Vacía",
        "Nombre Inválido",
    ]

    header_style = create_header_style()
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_style["font"]
        cell.fill = header_style["fill"]
        cell.border = header_style["border"]
        cell.alignment = header_style["alignment"]

    # Build sets for anomaly lookup
    duplicate_filenames: set[str] = set()
    for dup in result.duplicados:
        fn = dup.get("filename", "")
        if fn:
            duplicate_filenames.add(fn)

    empty_folders: set[str] = set()
    for vacia in result.vacias:
        folder = vacia.get("folder", "")
        if folder:
            empty_folders.add(folder)

    data_style = create_data_row_style()
    scan_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row_idx, inv in enumerate(result.facturas, start=2):
        has_anomaly = False

        is_duplicate = inv.filename in duplicate_filenames
        is_empty_folder = inv.full_path.rsplit("/", 1)[0] in empty_folders or \
                          inv.full_path.rsplit("\\", 1)[0] in empty_folders
        is_invalid = inv.invoice_type == "Unknown" or \
                     (inv.invoice_type in ("FEV", "CAP") and not _is_valid_filename(inv.filename))

        if is_duplicate or is_empty_folder or is_invalid:
            has_anomaly = True

        row_data = [
            inv.invoice_code,
            inv.invoice_type,
            inv.status,
            inv.full_path,
            inv.facturador,
            scan_timestamp,
            "Sí" if is_duplicate else "No",
            "Sí" if is_empty_folder else "No",
            "Sí" if is_invalid else "No",
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_style.get("font")
            cell.fill = data_style["fill"]
            cell.border = data_style["border"]
            cell.alignment = data_style.get("alignment")

            if has_anomaly:
                cell.fill = _ANOMALY_FILL

    auto_adjust_column_width(ws)


def _is_valid_filename(filename: str) -> bool:
    """Check if a filename is valid FEV or CAP."""
    from app.services.monitoreo_carpetas.name_validator import validate_name
    inv_type, is_valid = validate_name(filename)
    return is_valid


def _write_indicadores_sheet(wb: Workbook, result: ScanResult) -> None:
    """Escribe la hoja Indicadores con métricas agregadas."""
    ws = wb.create_sheet("Indicadores")

    header_style = create_header_style()
    for col, header in enumerate(["Indicador", "Valor"], start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_style["font"]
        cell.fill = header_style["fill"]
        cell.border = header_style["border"]
        cell.alignment = header_style["alignment"]

    data_style = create_data_row_style()

    # Write indicators
    row = 2
    for key, val in result.indicadores.items():
        cell_key = ws.cell(row=row, column=1, value=str(key))
        cell_val = ws.cell(row=row, column=2, value=val)
        for c in (cell_key, cell_val):
            c.fill = data_style["fill"]
            c.border = data_style["border"]
        row += 1

    # Write anomalies summary
    row += 1
    ws.cell(row=row, column=1, value="Top Anomalías").font = header_style["font"]
    row += 1

    anomalies = []
    if result.duplicados:
        anomalies.append(("Duplicados", len(result.duplicados)))
    if result.vacias:
        anomalies.append(("Carpetas Vacías", len(result.vacias)))

    # Count invalid names
    invalid_count = sum(
        1 for inv in result.facturas
        if inv.invoice_type == "Unknown" or
        (inv.invoice_type in ("FEV", "CAP") and not _is_valid_filename(inv.filename))
    )
    if invalid_count:
        anomalies.append(("Nombres Inválidos", invalid_count))

    anomalies.sort(key=lambda x: x[1], reverse=True)

    for anomaly_name, count in anomalies:
        c1 = ws.cell(row=row, column=1, value=anomaly_name)
        c2 = ws.cell(row=row, column=2, value=count)
        for c in (c1, c2):
            c.fill = data_style["fill"]
            c.border = data_style["border"]
        row += 1

    auto_adjust_column_width(ws)


def generate_excel(result: ScanResult, output_path: str | Path | None = None) -> Path:
    """Genera el reporte Excel de monitoreo de carpetas.

    Args:
        result: ScanResult con los datos del escaneo.
        output_path: Ruta donde guardar el archivo. Si es None,
            se genera una ruta en app/data/output/ con timestamp.

    Returns:
        Path al archivo generado.
    """
    wb = Workbook()

    _write_facturas_sheet(wb, result)
    _write_indicadores_sheet(wb, result)

    # Determine output path
    if output_path is None:
        from app.utils.input_data import output_data_directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = output_data_directory(create=True)
        output_path = out_dir / f"monitoreo_{timestamp}.xlsx"
    else:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(output_path))
    logger.info("Reporte generado: %s", output_path)
    return output_path
