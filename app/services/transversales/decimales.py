"""Validación de valores decimales en Vlr. Procedimiento y Vlr. Subsidiado."""

import logging

from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def detect_decimales(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[str]:
    """
    Detecta facturas con valores decimales en Vlr. Subsidiado o Vlr. Procedimiento.
    
    Returns:
        Lista de números de factura con valores decimales.
    """
    num_fact_idx = indices.get("numero_factura")
    vlr_sub_idx = indices.get("vlr_subsidiado")
    vlr_proc_idx = indices.get("vlr_procedimiento")
    
    if num_fact_idx is None:
        return []
    
    decimal_invoices: list[str] = []
    
    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = _normalize_invoice(numero_factura)
        if not factura_str:
            continue
        
        has_decimals = False
        
        if vlr_sub_idx is not None:
            vlr = data_sheet.cell(row=row, column=vlr_sub_idx + 1).value
            if isinstance(vlr, float) and vlr % 1 != 0:
                has_decimals = True
        
        if not has_decimals and vlr_proc_idx is not None:
            vlr = data_sheet.cell(row=row, column=vlr_proc_idx + 1).value
            if isinstance(vlr, float) and vlr % 1 != 0:
                has_decimals = True
        
        if has_decimals and factura_str not in decimal_invoices:
            decimal_invoices.append(factura_str)
            logger.debug("Factura %s con decimales detectada", factura_str)
    
    return decimal_invoices


def _normalize_invoice(value) -> str:
    """Normaliza número de factura a string."""
    if value is None:
        return ""
    return str(value).strip()