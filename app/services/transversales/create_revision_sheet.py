"""Creación de la hoja Revision en Excel con problemas detectados.

Extraído de app/services/revision_sheet.py para la Fase 7 (cleanup).
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    AREA_ODONTOLOGIA,
    AREA_URGENCIAS,
    REVISION_HEADERS,
    REVISION_SHEET,
    URGENCIA_REVISION_HEADERS,
)
from app.services.odontologia.normalized_rows import build_odontologia_normalized_rows
from app.services.transversales import (
    detect_decimales,
    detect_tipo_documento_edad,
    detect_doble_tipo_procedimiento,
    detect_ruta_duplicada,
    detect_cantidades_anomalas,
    detect_tipo_identificacion_entidad,
    normalize_invoice,
)
from app.services.transversales.column_indices import get_column_indices
from app.services.urgencias.codigos_sin_db import get_codigos_no_en_db_ess118
from app.services.urgencias.normalized_rows import build_urgencias_normalized_rows
from app.services.urgencias.revision_cantidad import detect_revision_cantidad_urgencias
from app.services.urgencias.revision_entidad_86 import detect_revision_entidad_86_urgencias
from app.utils.formatting import (
    auto_adjust_column_width,
    create_data_row_style,
    create_header_style,
    create_urgencia_data_row_style,
    create_urgencia_header_style,
)

logger = logging.getLogger(__name__)


def _write_column(sheet: Worksheet, column: int, values: list[str], start_row: int = 2) -> None:
    """Escribe una lista de valores en una columna."""
    for i, value in enumerate(values, start=start_row):
        sheet.cell(row=i, column=column, value=value)


def create_revision_sheet(
    workbook: Workbook,
    data_sheet: Worksheet,
    area: str = AREA_ODONTOLOGIA,
    profesional_dias: dict[str, list[int]] | None = None,
    permitir_todos_centros: bool = False,
) -> dict[str, Any]:
    """
    Crea la hoja Revision con los problemas detectados.

    Args:
        workbook: Libro de Excel (debe tener una hoja activa con datos)
        data_sheet: Hoja de datos a analizar
        area: Área del sistema ("odontologia" o "urgencias")
        profesional_dias: Dict {identificacion: [dias]} con días seleccionados por profesional
        permitir_todos_centros: Si True, solo permite ODONTOLOGIA y EXTRAMURAL

    Returns:
        Dict con información de los problemas encontrados
    """
    sheet = workbook.create_sheet(title=REVISION_SHEET)

    # Insertar fila vacía arriba
    sheet.insert_rows(1)

    # Obtener índices de columnas (coincidencia exacta - reporta faltantes)
    headers = [
        data_sheet.cell(row=1, column=col).value
        for col in range(1, data_sheet.max_column + 1)
    ]

    required_headers: dict[str, str] = {
        "numero_factura": "Número Factura",
        "vlr_subsidiado": "Vlr. Subsidiado",
        "vlr_procedimiento": "Vlr. Procedimiento",
        "codigo_tipo_procedimiento": "Código Tipo Procedimiento",
        "tipo_procedimiento": "Tipo Procedimiento",
        "codigo": "Código",
        "codigo_equiv": "Cód. Equivalente CUPS",
        "procedimiento": "Procedimiento",
        "identificacion": "Nº Identificación",
        "convenio_facturado": "Convenio Facturado",
        "cantidad": "Cantidad",
        "laboratorio": "Laboratorio",
        "centro_costo": "Centro Costo",
        "codigo_entidad_cobrar": "Cód Entidad Cobrar",
        "entidad_cobrar": "Entidad Cobrar",
        "entidad_afiliacion": "Entidad Afiliación",
        "tipo_factura_descripcion": "Tipo Factura Descripción",
        "ide_contrato": "IDE Contrato",
        "tipo_identificacion": "Tipo Identificación",
        "fec_nacimiento": "Fec. Nacimiento",
        "fec_factura": "Fec. Factura",
        "fecha_cierre": "Fecha Cierre",
        "profesional_identificacion": "Identificación Profesional",
        "profesional_atiende": "Profesional Atiende",
        "codigo_profesional": "Código Profesional",
        "responsable_cierra": "Responsable Cierra Facturar",
        "tarifario": "Tarifario",
        "tipo_usuario": "Tipo Usuario",
    }
    indices, missing_columns = get_column_indices(headers, required_headers)

    # Si hay columnas faltantes, incluir en el resultado para mostrar al usuario
    if missing_columns:
        logger.error("Columnas faltantes en el Excel: %s", missing_columns)

    # Seleccionar headers según el área
    if area == AREA_URGENCIAS:
        revision_headers = URGENCIA_REVISION_HEADERS
        header_style = create_urgencia_header_style()
    else:
        revision_headers = REVISION_HEADERS
        header_style = create_header_style()

    # Aplicar headers con estilo en fila 2
    for col, header in revision_headers.items():
        cell = sheet.cell(row=2, column=col, value=header)
        cell.font = header_style["font"]
        cell.fill = header_style["fill"]
        cell.border = header_style["border"]
        cell.alignment = header_style["alignment"]

    # Detectar problemas según el área
    if area == AREA_URGENCIAS:
        # --- Construir mapa responsable_cierra ---
        responsable_cierra: dict[str, str] = {}
        responsable_cierra_idx = indices.get("responsable_cierra")
        num_fact_idx = indices.get("numero_factura")
        if responsable_cierra_idx is not None and num_fact_idx is not None:
            for row in range(2, data_sheet.max_row + 1):
                numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
                factura = normalize_invoice(numero)
                if not factura:
                    continue
                raw = data_sheet.cell(row=row, column=responsable_cierra_idx + 1).value
                resp = str(raw).strip() if raw else ""
                if resp and factura not in responsable_cierra:
                    responsable_cierra[factura] = resp

        # --- Detectar todos los problemas ---
        logger.warning("=== VERIFICANDO CÓDIGOS ESS118 CONTRA DB ===")
        problemas_codigos_no_en_db = get_codigos_no_en_db_ess118(data_sheet, indices)
        codigos_no_en_db_set = {item["codigo"] for item in problemas_codigos_no_en_db}

        if problemas_codigos_no_en_db:
            logger.warning(
                "Procedimientos NO encontrados en DB para ESS118 (%d errores): %s",
                len(problemas_codigos_no_en_db), sorted(codigos_no_en_db_set),
            )
        else:
            logger.warning("Todos los códigos de ESS118 están en DB")

        from app.services.urgencias import (
            detect_centro_costo_urgencias,
            detect_cups_equivalentes,
            detect_hospitalizacion_codes,
            detect_ide_contrato_urgencias,
            detect_sala_observacion,
        )

        try:
            problemas_centros = detect_centro_costo_urgencias(data_sheet, indices)
            problemas_ide_contrato = detect_ide_contrato_urgencias(data_sheet, indices)
            problemas_cups_equivalentes: list[dict[str, str]] = []
            problemas_cups_equivalentes.extend(detect_cups_equivalentes(data_sheet, indices))
            problemas_cups_equivalentes.extend(detect_sala_observacion(data_sheet, indices))
            problemas_cups_equivalentes.extend(detect_hospitalizacion_codes(data_sheet, indices))
        except Exception as exc:
            logger.exception("Error detectando problemas de urgencias: %s", exc)
            problemas_centros, problemas_ide_contrato, problemas_cups_equivalentes = [], [], []

        from app.services.odontologia.mal_capitado import detect_mal_capitado
        from app.services.urgencias.cantidades_urgencias import detect_cantidades_urgencias
        from app.services.urgencias.hospitalizacion import detect_cantidades_hospitalizacion

        mal_capitado = detect_mal_capitado(data_sheet, indices)
        cantidades_urgencias = detect_cantidades_urgencias(data_sheet, indices)
        cantidades_hospitalizacion = detect_cantidades_hospitalizacion(data_sheet, indices)

        # --- Tipo Identificación AS/MS vs Cód Entidad Cobrar ---
        tipo_id_entidad = detect_tipo_identificacion_entidad(data_sheet, indices)
        logger.info(
            "create_revision_sheet - Tipo Identificación / Entidad encontrados: %d",
            len(tipo_id_entidad),
        )

        # --- ⚠️ Revisión Necesaria: Entidad 86 ---
        revision_entidad_86 = detect_revision_entidad_86_urgencias(data_sheet, indices)
        logger.info(
            "create_revision_sheet - Revision Entidad 86 encontradas: %d",
            len(revision_entidad_86),
        )

        # --- ⚠️ Revisión Necesaria: Cantidad > 1 ---
        revision_cantidad = detect_revision_cantidad_urgencias(data_sheet, indices)
        logger.info(
            "create_revision_sheet - Revision Cantidad encontradas: %d",
            len(revision_cantidad),
        )

        # --- Normalizar todos los errores a filas de 6 columnas ---
        normalized_rows = build_urgencias_normalized_rows(
            problemas_centros=problemas_centros,
            problemas_ide_contrato=problemas_ide_contrato,
            problemas_cups_equivalentes=problemas_cups_equivalentes,
            mal_capitado=mal_capitado,
            cantidades_urgencias=cantidades_urgencias,
            cantidades_soat_urgencias=[],
            cantidades_hospitalizacion=cantidades_hospitalizacion,
            cantidades_soat_hospitalizacion=[],
            responsables_map=responsable_cierra,
            revision_entidad_86=revision_entidad_86,
            revision_cantidad=revision_cantidad,
            tipo_identificacion_entidad=tipo_id_entidad,
        )

        # --- Escribir filas normalizadas en Excel ---
        has_header_override = any("_header_override" in row for row in normalized_rows)
        for i, row_data in enumerate(normalized_rows, start=3):
            sheet.cell(row=i, column=1, value=row_data["tipo_error"])
            sheet.cell(row=i, column=2, value=row_data["factura"])
            sheet.cell(row=i, column=3, value=row_data["responsable_cierra"])
            sheet.cell(row=i, column=4, value=row_data["descripcion"])
            col5_value = row_data.get("_header_override", row_data.get("procedimiento", ""))
            sheet.cell(row=i, column=5, value=col5_value)
            sheet.cell(row=i, column=6, value=row_data["detalle"])

        problemas_encontrados = {
            "normalizados": normalized_rows,
            "totales_por_tipo": {
                "Centros de Costo": len(problemas_centros),
                "IDE Contrato": len(problemas_ide_contrato),
                "Cups Equivalentes": len(problemas_cups_equivalentes),
                "MAL CAPITADO": len(mal_capitado),
                "Cantidades": len(cantidades_urgencias),
                "Cantidades Hospitalización": len(cantidades_hospitalizacion),
                "Tipo Identificación / Entidad": len(tipo_id_entidad),
                "⚠️ Revisión Necesaria": len(revision_entidad_86) + len(revision_cantidad),
            },
        }
    else:
        # Odontología / Equipos Básicos: detectar todos los problemas
        from app.services.odontologia.profesionales import detect_profesionales_odontologia
        from app.services.odontologia.centro_costo import detect_centro_costo_odontologia
        from app.services.odontologia.ide_contrato import detect_ide_contrato_odontologia

        decimales = detect_decimales(data_sheet, indices)
        doble_tipo = detect_doble_tipo_procedimiento(data_sheet, indices)
        ruta_dup = detect_ruta_duplicada(data_sheet, indices)
        cantidades = detect_cantidades_anomalas(data_sheet, indices)
        tipo_id_edad = detect_tipo_documento_edad(data_sheet, indices)
        tipo_id_entidad = detect_tipo_identificacion_entidad(data_sheet, indices)
        profesionales = detect_profesionales_odontologia(data_sheet, indices)
        centro_costo = detect_centro_costo_odontologia(
            data_sheet, indices,
            profesional_dias=profesional_dias,
            permitir_todos_centros=permitir_todos_centros,
        )
        ide_contrato = detect_ide_contrato_odontologia(data_sheet, indices)

        # Construir mapa responsable_cierra
        responsable_cierra_map: dict[str, str] = {}
        rci = indices.get("responsable_cierra")
        nfi = indices.get("numero_factura")
        if rci is not None and nfi is not None:
            for row in range(2, data_sheet.max_row + 1):
                num = data_sheet.cell(row=row, column=nfi + 1).value
                fac = normalize_invoice(num)
                if not fac:
                    continue
                raw = data_sheet.cell(row=row, column=rci + 1).value
                resp = str(raw).strip() if raw else ""
                if resp and fac not in responsable_cierra_map:
                    responsable_cierra_map[fac] = resp

        # Normalizar a filas de 6 columnas
        normalized_rows = build_odontologia_normalized_rows(
            decimales=decimales,
            doble_tipo=doble_tipo,
            ruta_dup=ruta_dup,
            profesionales=profesionales,
            cantidades=cantidades,
            tipo_id_edad=tipo_id_edad,
            tipo_id_entidad=tipo_id_entidad,
            centro_costo=centro_costo,
            ide_contrato=ide_contrato,
            responsable_cierra=responsable_cierra_map,
        )

        # Escribir filas normalizadas en Excel
        for i, row_data in enumerate(normalized_rows, start=3):
            sheet.cell(row=i, column=1, value=row_data["tipo_error"])
            sheet.cell(row=i, column=2, value=row_data["factura"])
            sheet.cell(row=i, column=3, value=row_data["responsable_cierra"])
            sheet.cell(row=i, column=4, value=row_data["descripcion"])
            sheet.cell(row=i, column=5, value=row_data["procedimiento"])
            sheet.cell(row=i, column=6, value=row_data["detalle"])

        problemas_encontrados = {
            "normalizados": normalized_rows,
            "totales_por_tipo": {
                "Decimales": len(decimales),
                "Doble tipo procedimiento": len(doble_tipo),
                "Ruta Duplicada": len(ruta_dup),
                "Convenio de procedimiento": len(profesionales),
                "Cantidades": len(cantidades),
                "Tipo Identificación": len(tipo_id_edad),
                "Tipo Identificación / Entidad": len(tipo_id_entidad),
                "Centro Costo": len(centro_costo),
                "IDE Contrato": len(ide_contrato),
            },
        }

    # Aplicar estilo a filas de datos (fila 3+) según el área
    if area == AREA_URGENCIAS:
        data_style = create_urgencia_data_row_style()
    else:
        data_style = create_data_row_style()

    for row in range(3, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = data_style["fill"]
            cell.border = data_style["border"]
            cell.alignment = data_style["alignment"]

    # Ajustar ancho de columnas automáticamente
    column_widths = auto_adjust_column_width(sheet)

    # Logging según el área
    if area == AREA_URGENCIAS:
        logger.info(
            "Hoja Revision Urgencias creada - Total filas normalizadas: %d",
            len(normalized_rows),
        )
    else:
        logger.info(
            "Hoja Revision Odontología/EB creada - Total filas normalizadas: %d",
            len(normalized_rows),
        )

    # Build resultado según el área
    if area == AREA_URGENCIAS:
        return {
            "rule": "create_revision_sheet",
            "sheet": REVISION_SHEET,
            "area": area,
            "headers": list(URGENCIA_REVISION_HEADERS.values()),
            "normalized_rows_count": len(normalized_rows),
            "problemas": problemas_encontrados,
            "column_widths": column_widths,
            "missing_columns": missing_columns,
        }
    else:
        return {
            "rule": "create_revision_sheet",
            "sheet": REVISION_SHEET,
            "area": area,
            "headers": list(REVISION_HEADERS.values()),
            "normalized_rows_count": len(normalized_rows),
            "problemas": problemas_encontrados,
            "column_widths": column_widths,
            "missing_columns": missing_columns,
        }
