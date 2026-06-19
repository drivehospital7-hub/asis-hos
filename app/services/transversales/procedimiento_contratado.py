"""Detector transversal: CUPS sin contrato para la entidad.

Verifica que cada par (codigo_entidad_cobrar, codigo) exista en la base
de datos como una combinación contratada, mediante el recorrido:

    eps_contratado → eps_nota → nota_hoja → notas_tecnicas → procedimiento
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants.urgencias import FACTURADORES_URGENCIAS, VALOR_TARIFARIO_FARMACIA
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)

# Facturadores de urgencias — cuando responsable_cierra coincide, validar
# contra nota_hoja id=1 o 27 en vez de contra la cadena contractual de la entidad.
_FACTURADORES_URGENCIAS_NORM: frozenset[str] = frozenset(
    " ".join(f.upper().split()) for f in FACTURADORES_URGENCIAS
)

# Entidades que pueden usar nota_hoja 1/27 como pase directo cuando
# el responsable es un facturador de urgencias.
# Si la entidad NO está en esta lista, cae a validación normal contra
# los procedimientos contratados en DB.
_ENTIDADES_NOTA_URGENCIAS: frozenset[str] = frozenset({
    "EPSS12", "ESSC62", "ESS207", "EPS048", "CCF055", "EPS008",
    "EPSS08", "EPSS34", "CCF102", "EPS010", "EPS018", "EPS017",
    "EPSS17", "ESSC24", "EPS042", "EPSI06", "EPSS40", "EPSI04",
    "EPS025", "EPSS02", "EPSS10", "CCF050", "ESS062", "EPS047",
    "EPSS018", "EPSC005", "EPSS005", "EPS002", "CCF033", "EPSI01",
    "EPS001", "EPS012", "EPS040", "EPSI03", "CCFC33", "EPSC34",
})


def detect_cups_sin_contrato(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """Detecta CUPS que no están contratados para la entidad facturadora.

    Para cada fila del Excel, normaliza (strip, upper) el par
    (codigo_entidad_cobrar, codigo) y lo compara contra los pares válidos
    obtenidos de la base de datos.

    Si el código principal no está contratado y existe la columna
    "Cód. Equivalente CUPS" (indice "codigo_equiv"), también prueba
    con ese código equivalente antes de marcar error.

    Args:
        data_sheet: Hoja de Excel con los datos.
        indices: Diccionario con los índices 0-based de las columnas.

    Returns:
        Lista de errores. Cada error contiene:
            factura, codigo, procedimiento, codigo_entidad_cobrar,
            entidad, problema.
        Si faltan columnas o la DB no está disponible, retorna [].
    """
    # ── 1. Validar columnas requeridas ──────────────────────────────────
    num_fact_idx = indices.get("numero_factura")
    cod_ent_idx = indices.get("codigo_entidad_cobrar")
    codigo_idx = indices.get("codigo")

    if any(idx is None for idx in [num_fact_idx, cod_ent_idx, codigo_idx]):
        logger.warning(
            "Columnas requeridas faltantes para detect_cups_sin_contrato: "
            "numero_factura=%s, codigo_entidad_cobrar=%s, codigo=%s",
            num_fact_idx,
            cod_ent_idx,
            codigo_idx,
        )
        return []

    proc_idx = indices.get("procedimiento")
    tarifario_idx = indices.get("tarifario")
    codigo_equiv_idx = indices.get("codigo_equiv")

    # ── 2. Pre-load desde DB ───────────────────────────────────────────
    try:
        from app.database import SessionLocal
        from app.models import (
            EpsContratado,
            EpsNota,
            NotaHoja,
            NotasTecnicas,
            Procedimiento,
        )

        session = SessionLocal()
        try:
            # Construir set de pares válidos (cod_contrato, cups)
            results = (
                session.query(EpsContratado, Procedimiento)
                .join(EpsNota, EpsNota.id_eps_contratado == EpsContratado.id)
                .join(NotaHoja, NotaHoja.id == EpsNota.id_nota_hoja)
                .join(NotasTecnicas, NotasTecnicas.id_nota_hoja == NotaHoja.id)
                .join(Procedimiento, Procedimiento.id == NotasTecnicas.id_procedimiento)
                .all()
            )

            pares_validos: set[tuple[str, str]] = set()
            entidades_con_datos: set[str] = set()
            for ec, proc in results:
                cod_key = ec.cod_contrato.strip().upper()
                cups_key = proc.cups.strip().upper()
                pares_validos.add((cod_key, cups_key))
                entidades_con_datos.add(cod_key)

            # Construir mapa de nombre de EPS por código de contrato
            eps_list = session.query(EpsContratado).all()
            eps_map: dict[str, str] = {}
            for ec in eps_list:
                eps_map[ec.cod_contrato.strip().upper()] = ec.eps

            # Pre-load procedimientos de nota_hoja id=1 y 27 (urgencias exception)
            nota_urgencias_results = (
                session.query(Procedimiento)
                .join(NotasTecnicas, NotasTecnicas.id_procedimiento == Procedimiento.id)
                .filter(NotasTecnicas.id_nota_hoja.in_([1, 27]))
                .all()
            )
            nota_urgencias_cups: set[str] = {p.cups.strip().upper() for p in nota_urgencias_results}

            # Pre-load procedimientos de nota_hoja id=2 y 3 (CAP exception)
            cap_results = (
                session.query(NotasTecnicas.id_nota_hoja, Procedimiento.cups)
                .join(Procedimiento, Procedimiento.id == NotasTecnicas.id_procedimiento)
                .filter(NotasTecnicas.id_nota_hoja.in_([2, 3]))
                .all()
            )
            nota_cap_cups: dict[int, set[str]] = {2: set(), 3: set()}
            for nt_id, cups_val in cap_results:
                nota_cap_cups[nt_id].add(cups_val.strip().upper())

        finally:
            session.close()

    except Exception as exc:
        logger.exception(
            "Error al consultar DB para detect_cups_sin_contrato: %s", exc
        )
        return []

    # ── 3. Iterar filas del Excel ───────────────────────────────────────
    errores: list[dict[str, Any]] = []

    def _append_error(msg_codigo: str, msg_entidad: str) -> None:
        """Agrega un error con el formato estándar."""
        nonlocal errores
        procedimiento = ""
        if proc_idx is not None:
            proc_raw = data_sheet.cell(row=row, column=proc_idx + 1).value
            if proc_raw:
                procedimiento = str(proc_raw).strip()
        entidad_nombre = eps_map.get(msg_entidad, msg_entidad)
        errores.append({
            "factura": factura,
            "codigo": msg_codigo,
            "procedimiento": procedimiento,
            "codigo_entidad_cobrar": msg_entidad,
            "entidad": entidad_nombre,
            "problema": (
                f"CUPS {msg_codigo} no contratado para "
                f"{msg_entidad}, {entidad_nombre}"
            ),
        })

    for row in range(2, data_sheet.max_row + 1):
        # Normalizar factura
        numero_raw = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura = normalize_invoice(numero_raw)
        if not factura:
            continue

        # Saltar filas de farmacia/medicamentos (no cargados en DB)
        if tarifario_idx is not None:
            tarifario_val = data_sheet.cell(row=row, column=tarifario_idx + 1).value
            if tarifario_val and str(tarifario_val).strip() == VALOR_TARIFARIO_FARMACIA:
                continue

        # Leer códigos
        cod_entidad_raw = data_sheet.cell(row=row, column=cod_ent_idx + 1).value
        codigo_raw = data_sheet.cell(row=row, column=codigo_idx + 1).value

        if not cod_entidad_raw or not codigo_raw:
            continue

        cod_entidad = str(cod_entidad_raw).strip().upper()
        codigo = str(codigo_raw).strip().upper()

        if not cod_entidad or not codigo:
            continue

        # Leer código equivalente si la columna existe
        codigo_equiv = ""
        if codigo_equiv_idx is not None:
            codigo_equiv_raw = data_sheet.cell(row=row, column=codigo_equiv_idx + 1).value
            if codigo_equiv_raw:
                codigo_equiv = str(codigo_equiv_raw).strip().upper()

        # Excepción: responsable urgencias valida contra nota_hoja id=1 o 27
        # como fuente adicional, independientemente de si la entidad está
        # o no en _ENTIDADES_NOTA_URGENCIAS. Si el CUPS está en nota_hoja
        # 1/27, se acepta; si no, cae a validación normal contra pares_validos.
        responsable_idx = indices.get("responsable_cierra")
        if responsable_idx is not None:
            raw_resp = data_sheet.cell(row=row, column=responsable_idx + 1).value
            resp_name = " ".join(str(raw_resp).upper().split()) if raw_resp else ""
            if resp_name in _FACTURADORES_URGENCIAS_NORM:
                if codigo in nota_urgencias_cups:
                    continue
                if codigo_equiv and codigo_equiv in nota_urgencias_cups:
                    continue

        # CAP invoice exceptions: facturas CAP de ESS118/EPSS41 validan
        # SOLO contra su nota_hoja capitada específica.
        # Si no matchea el set capitado, error directo — NO cae a pares_validos
        # (porque la entidad puede tener procedimientos contractuales en otra
        # nota_hoja que no aplican a facturas CAP).
        if factura and factura.upper().startswith("CAP"):
            if cod_entidad == "ESS118":
                if codigo in nota_cap_cups[3]:
                    continue
                if codigo_equiv and codigo_equiv in nota_cap_cups[3]:
                    continue
                # CAP + ESS118 + CUPS no capitado → error directo
                _append_error(codigo, cod_entidad)
                continue
            elif cod_entidad == "EPSS41":
                if codigo in nota_cap_cups[2]:
                    continue
                if codigo_equiv and codigo_equiv in nota_cap_cups[2]:
                    continue
                # CAP + EPSS41 + CUPS no capitado → error directo
                _append_error(codigo, cod_entidad)
                continue

        # Saltar entidades sin procedimientos cargados en DB
        if cod_entidad not in entidades_con_datos:
            continue

        # Verificar si el código principal está contratado
        if (cod_entidad, codigo) in pares_validos:
            continue

        # Si el código principal no está, probar con el equivalente
        if codigo_equiv and (cod_entidad, codigo_equiv) in pares_validos:
            continue

        # Excepción: facturas FEV de EPS037/EPSS41 son con autorización
        if factura.upper().startswith("FEV") and cod_entidad in ("EPS037", "EPSS41"):
            logger.info(
                "FEV autorizada: factura=%s, entidad=%s, codigo=%s",
                factura, cod_entidad, codigo,
            )
            continue

        # Ninguno de los dos está contratado → reportar error
        _append_error(codigo, cod_entidad)

    return errores
