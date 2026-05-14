from typing import TypedDict

import logging

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import TIPO_USUARIO_VALORES

logger = logging.getLogger(__name__)


class TipoUsuarioProblema(TypedDict):
    """Problema encontrado en tipo de usuario."""
    factura: str
    tipo_actual: str
    responsable: str


def detect_tipo_usuario(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[TipoUsuarioProblema]:
    """
    Detecta facturas donde el tipo de usuario no es válido.

    Valores válidos:
    - SUBSIDIADO
    - CONTRIBUTIVO
    - OTROS (REGÍMENES ESPECIALES, EOC)
    - VINCULADO
    - PARTICULAR

    Returns:
        Lista de dicts con keys: "factura", "tipo_actual", "responsable"
    """
    tipo_usuario_idx = indices.get("tipo_usuario")
    num_fact_idx = indices.get("numero_factura")
    responsable_idx = indices.get("responsable_cierra")

    if tipo_usuario_idx is None or num_fact_idx is None:
        logger.warning(
            "No se pueden detectar errores de tipo usuario: "
            "columnas requeridas no encontradas. "
            "tipo_usuario=%s, numero_factura=%s",
            tipo_usuario_idx, num_fact_idx
        )
        return []

    problemas: list[TipoUsuarioProblema] = []
    facturas_ya_procesadas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = _normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_ya_procesadas:
            continue

        tipo_usuario = data_sheet.cell(row=row, column=tipo_usuario_idx + 1).value
        if not tipo_usuario:
            continue

        tipo_usuario_str = str(tipo_usuario).strip().upper()

        logger.debug(
            "Fila %s: tipo_usuario=%s",
            row, repr(tipo_usuario_str)
        )

        # Validar contra valores permitidos
        if tipo_usuario_str not in TIPO_USUARIO_VALORES:
            responsable = ""
            if responsable_idx is not None:
                resp = data_sheet.cell(row=row, column=responsable_idx + 1).value
                responsable = str(resp).strip() if resp else ""

            problemas.append({
                "factura": factura_str,
                "tipo_actual": tipo_usuario_str,
                "responsable": responsable,
            })
            facturas_ya_procesadas.add(factura_str)
            logger.debug(
                "Fila %s: Tipo usuario inválido (Actual: %s, Esperado uno de: %s)",
                row,
                tipo_usuario_str,
                TIPO_USUARIO_VALORES,
            )

    return problemas


def _normalize_invoice(value) -> str:
    """Normaliza número de factura a string."""
    if value is None:
        return ""
    if isinstance(value, (int, float)) and value == int(value):
        return str(int(value))
    return str(value).strip()