"""Detección de ruta duplicada (pacientes con múltiples facturas en PyP).

Extraído de revision_sheet.py._detect_ruta_duplicada y _detect_ruta_duplicada_equipos_basicos.
Unificada en UNA función parametrizada.

Migration note (2026-06): This detector has been migrated to the DB-backed
rule engine as rule 'ruta_duplicada'. When USE_RULE_ENGINE=true, the engine
version is used instead via RuleBasedDetector in detect_all.py.
This legacy Python implementation is preserved for rollback (USE_RULE_ENGINE=false).
"""

from __future__ import annotations

import logging
from collections import defaultdict

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import CONVENIO_PYP
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_ruta_duplicada(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
    threshold: int = 3,
) -> list[dict]:
    """
    Detecta pacientes con múltiples facturas en Promoción y Prevención (PyP).

    Filtra solo facturas con Convenio Facturado = PyP y cuenta cuántas
    facturas diferentes tiene cada paciente. Si supera el threshold,
    se reporta como ruta duplicada.

    Args:
        data_sheet: Hoja de Excel con los datos.
        indices: Índices de columnas (numero_factura, identificacion, convenio_facturado).
        threshold: Número mínimo de facturas para considerar duplicada (default: 3).

    Returns:
        Lista de dicts con keys: "identificacion", "facturas", "cantidad".
    """
    num_fact_idx = indices.get("numero_factura")
    ident_idx = indices.get("identificacion")
    contrato_idx = indices.get("convenio_facturado")

    if None in (num_fact_idx, ident_idx, contrato_idx):
        return []

    conteo_ident: dict[str, set[str]] = defaultdict(set)

    for row in range(2, data_sheet.max_row + 1):
        contrato = data_sheet.cell(row=row, column=contrato_idx + 1).value
        if contrato != CONVENIO_PYP:
            continue

        ident = data_sheet.cell(row=row, column=ident_idx + 1).value
        factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value

        if ident is not None and factura is not None:
            ident_str = str(ident).strip()
            factura_str = str(factura).strip()
            if ident_str and factura_str:
                conteo_ident[ident_str].add(factura_str)

    result = []
    for ident, facturas in conteo_ident.items():
        if len(facturas) >= threshold:
            result.append({
                "identificacion": ident,
                "facturas": ", ".join(sorted(facturas)),
                "cantidad": len(facturas),
            })
    return result
