"""Detector transversal de CUPS equivalentes — aplica a TODOS los tipos de factura.

Reglas:
- Código 906317 -> debe usarse 1906317 (Hepatitis B Prueba rápida)
- Código 906249 -> debe usarse 906249PR (VIH Prueba rápida)
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)

# Mapeo de códigos incorrectos a su equivalente correcto (transversal)
CODIGOS_CUPS_EQUIVALENTES: dict[str, str] = {
    "906317": "1906317",   # Hepatitis B (Prueba rápida)
    "906249": "906249PR",  # VIH Prueba rápida
}


def detect_cups_equivalentes_transversal(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """Detecta facturas con códigos CUPS incorrectos (transversal a todos los tipos).

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo", "codigo_equiv",
        "accion", "procedimiento"
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    proc_idx = indices.get("procedimiento")

    if num_fact_idx is None or codigo_idx is None:
        logger.warning("CUPS Equivalentes - Columnas necesarias no encontradas")
        return []

    problemas: list[dict[str, Any]] = []

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        codigo = (
            data_sheet.cell(row=row, column=codigo_idx + 1).value
            if codigo_idx is not None else None
        )
        codigo_str = str(codigo).strip() if codigo else ""

        if not codigo_str:
            continue

        equiv = CODIGOS_CUPS_EQUIVALENTES.get(codigo_str)
        if equiv is None:
            continue

        proc_str = ""
        if proc_idx is not None:
            proc_val = data_sheet.cell(row=row, column=proc_idx + 1).value
            proc_str = str(proc_val).strip() if proc_val else ""

        problemas.append({
            "factura": factura_str,
            "codigo": codigo_str,
            "codigo_equiv": equiv,
            "accion": f"Usar {equiv}",
            "procedimiento": proc_str,
        })
        logger.info(
            "CUPS Equivalente - Fila %s: Código=%s -> debe usarse %s",
            row, codigo_str, equiv,
        )

    if problemas:
        logger.info("CUPS Equivalentes - Problemas encontrados: %d", len(problemas))

    return problemas
