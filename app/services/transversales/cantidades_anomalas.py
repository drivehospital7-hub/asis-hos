"""Detección de cantidades anómalas en facturación.

Extraído de revision_sheet.py._detect_cantidades_anomalas y
_detect_cantidades_anomalas_equipos_basicos.
Unificada en UNA función parametrizada.
"""

from __future__ import annotations

import logging

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import CONVENIO_PYP
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_cantidades_anomalas(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
    cantidad_consultas_min: int = 2,
    cantidad_max_general: int = 10,
    cantidad_pyp_min: int = 3,
) -> list[dict]:
    """
    Detecta facturas con cantidades anómalas.

    Reglas:
    - Consultas con cantidad >= cantidad_consultas_min
    - Cualquier tipo con cantidad > cantidad_max_general
    - Convenio PyP con cantidad >= cantidad_pyp_min

    Args:
        data_sheet: Hoja de Excel con los datos.
        indices: Índices de columnas (numero_factura, tipo_procedimiento,
                 cantidad, convenio_facturado).
        cantidad_consultas_min: Mínimo de cantidad para considerar consultas anómalas.
        cantidad_max_general: Máximo general antes de considerar anómalo.
        cantidad_pyp_min: Mínimo de cantidad para PyP.

    Returns:
        Lista de dicts con keys: "factura", "tipo_procedimiento", "cantidad",
                                 "convenio", "problema".
    """
    num_fact_idx = indices.get("numero_factura")
    tipo_proc_idx = indices.get("tipo_procedimiento")
    cantidad_idx = indices.get("cantidad")
    conveniencia_idx = indices.get("convenio_facturado")

    if None in (num_fact_idx, tipo_proc_idx, cantidad_idx, conveniencia_idx):
        return []

    problemas = []

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        tipo_value = data_sheet.cell(row=row, column=tipo_proc_idx + 1).value
        cantidad = data_sheet.cell(row=row, column=cantidad_idx + 1).value
        conveniencia = data_sheet.cell(row=row, column=conveniencia_idx + 1).value

        if not isinstance(cantidad, (int, float)):
            continue

        # Reglas de cantidad anómala
        is_anomaly = (
            (tipo_value == "Consultas" and cantidad >= cantidad_consultas_min)
            or cantidad > cantidad_max_general
            or (conveniencia == CONVENIO_PYP and cantidad >= cantidad_pyp_min)
        )

        if is_anomaly and factura_str not in [p.get("factura") for p in problemas]:
            problema_tipo = (
                f"Consultas con cantidad {cantidad}"
                if tipo_value == "Consultas"
                else f"Cantidad {cantidad}"
            )
            problemas.append({
                "factura": factura_str,
                "tipo_procedimiento": str(tipo_value) if tipo_value else "",
                "cantidad": cantidad,
                "convenio": str(conveniencia) if conveniencia else "",
                "problema": problema_tipo,
            })

    return problemas
