"""Detección de errores: Cód Entidad Cobrar ≠ 1/0001 con Vlr. Copago ≠ 0.

Si el código de entidad a cobrar no es '1' ni '0001', el valor de
'Vlr. Copago' debe ser 0. Cualquier valor distinto de 0 en esa
combinación es un error de datos.
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)

# Entidades default que están exentas de la regla
ENTIDADES_DEFAULT = frozenset({"1", "0001"})


def detect_copago_entidad_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta filas donde Cód Entidad Cobrar no es default y Vlr. Copago ≠ 0.

    La regla aplica por fila (no por factura): cada fila se evalúa
    independientemente según su propio valor de Vlr. Copago.

    Args:
        data_sheet: Hoja de Excel con los datos de urgencias
        indices: Diccionario con índices de columnas

    Returns:
        Lista de dicts con keys: factura, codigo, procedimiento,
        entidad_cobrar, vlr_copago
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")
    codigo_entidad_idx = indices.get("codigo_entidad_cobrar")
    vlr_copago_idx = indices.get("vlr_copago")

    # Si falta alguna columna requerida, no podemos evaluar
    if None in (num_fact_idx, codigo_entidad_idx, vlr_copago_idx):
        logger.warning(
            "Copago Entidad - Columnas necesarias no encontradas: "
            "numero_factura=%s, codigo_entidad_cobrar=%s, vlr_copago=%s",
            num_fact_idx, codigo_entidad_idx, vlr_copago_idx,
        )
        return []

    errores: list[dict[str, str]] = []

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        # Normalizar entidad
        codigo_entidad = data_sheet.cell(row=row, column=codigo_entidad_idx + 1).value
        if codigo_entidad is None:
            continue  # Sin entidad, no aplica la regla

        entidad_str = str(codigo_entidad).strip().upper()
        if not entidad_str:
            continue  # Entidad vacía, no aplica la regla

        # Si es entidad default, no se valida copago
        if entidad_str in ENTIDADES_DEFAULT:
            continue

        # Normalizar copago
        copago_raw = data_sheet.cell(row=row, column=vlr_copago_idx + 1).value
        try:
            copago_val = float(copago_raw or 0)
        except (ValueError, TypeError):
            copago_val = 0.0

        # Si copago es 0, no hay error
        if copago_val == 0.0:
            continue

        # Es un error: entidad no default con copago != 0
        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value if codigo_idx is not None else ""
        procedimiento = data_sheet.cell(row=row, column=procedimiento_idx + 1).value if procedimiento_idx is not None else ""

        errores.append({
            "factura": factura_str,
            "codigo": str(codigo).strip() if codigo else "",
            "procedimiento": str(procedimiento).strip() if procedimiento else "",
            "entidad_cobrar": entidad_str,
            "vlr_copago": copago_val,
        })

    logger.info(
        "Copago Entidad - Filas procesadas: %d, Errores encontrados: %d",
        data_sheet.max_row - 1,
        len(errores),
    )
    return errores
