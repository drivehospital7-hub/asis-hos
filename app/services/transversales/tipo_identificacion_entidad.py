"""Regla transversal: Tipo Identificación AS/MS y Cód Entidad Cobrar 86000/MIN001 son exclusivos.

Reglas:
1. Si Tipo Identificación es AS (Adulto Sin identificación) o MS (Menor Sin identificación)
   → Cód Entidad Cobrar debe ser 86000 o MIN001. Si no es ninguna, error.
2. Si Cód Entidad Cobrar es 86000
   → Tipo Identificación debe ser AS o MS. Si es cualquier otro (CC, DE, TI, RC, etc.), error.
"""

from __future__ import annotations

import logging
from typing import TypedDict

from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

TIPO_ID_AS_MS = frozenset({"AS", "MS"})
COD_ENTIDADES_AS_MS = frozenset({"86000", "MIN001"})


class TipoIdentificacionEntidadProblema(TypedDict):
    """Problema encontrado: incompatibilidad entre tipo identificación y código entidad."""
    factura: str
    tipo_identificacion: str
    cod_entidad_actual: str
    cod_entidad_esperado: str
    problema: str


def detect_tipo_identificacion_entidad(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[TipoIdentificacionEntidadProblema]:
    """
    Detecta incompatibilidades entre Tipo Identificación y Cód Entidad Cobrar.

    Reglas:
    - AS o MS requieren Cód Entidad Cobrar = 86000 o MIN001.
    - Cód Entidad Cobrar = 86000 solo es válido para AS o MS.
      Si el tipo es CC, DE, TI, RC, NIT, etc. con 86000, es error.
    - Cód Entidad Cobrar ≠ 86000/MIN001 no puede tener AS o MS.

    Returns:
        Lista de dicts con keys: "factura", "tipo_identificacion",
        "cod_entidad_actual", "cod_entidad_esperado", "problema"
    """
    tipo_id_idx = indices.get("tipo_identificacion")
    cod_entidad_idx = indices.get("codigo_entidad_cobrar")
    num_fact_idx = indices.get("numero_factura")

    if tipo_id_idx is None or cod_entidad_idx is None:
        logger.warning(
            "No se pueden detectar errores de tipo identificación vs entidad: "
            "columnas requeridas no encontradas. "
            "tipo_identificacion=%s, codigo_entidad_cobrar=%s",
            tipo_id_idx, cod_entidad_idx,
        )
        return []

    problemas: list[TipoIdentificacionEntidadProblema] = []
    facturas_ya_procesadas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        # Número de factura
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = _normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_ya_procesadas:
            continue

        # Tipo identificación
        tipo_id = data_sheet.cell(row=row, column=tipo_id_idx + 1).value
        if not tipo_id:
            continue
        tipo_id_str = str(tipo_id).strip().upper()

        # Cód entidad cobrar
        cod_entidad = data_sheet.cell(row=row, column=cod_entidad_idx + 1).value
        cod_entidad_str = str(cod_entidad).strip() if cod_entidad is not None else ""

        entidades_esperadas_str = ", ".join(sorted(COD_ENTIDADES_AS_MS))

        # --- Regla 1: AS/MS requiere 86000 o MIN001 ---
        if tipo_id_str in TIPO_ID_AS_MS and cod_entidad_str not in COD_ENTIDADES_AS_MS:
            problemas.append({
                "factura": factura_str,
                "tipo_identificacion": tipo_id_str,
                "cod_entidad_actual": cod_entidad_str,
                "cod_entidad_esperado": entidades_esperadas_str,
                "problema": "as_ms_requiere_86000_o_MIN001",
            })
            facturas_ya_procesadas.add(factura_str)
            logger.debug(
                "Fila %s: %s requiere Cód Entidad Cobrar en [%s] (actual: %s)",
                row, tipo_id_str, entidades_esperadas_str, cod_entidad_str,
            )
            continue

        # --- Regla 2: 86000 solo para AS/MS ---
        if cod_entidad_str == "86000" and tipo_id_str not in TIPO_ID_AS_MS:
            problemas.append({
                "factura": factura_str,
                "tipo_identificacion": tipo_id_str,
                "cod_entidad_actual": cod_entidad_str,
                "cod_entidad_esperado": "86000",
                "problema": "86000_solo_para_as_ms",
            })
            facturas_ya_procesadas.add(factura_str)
            logger.debug(
                "Fila %s: Cód Entidad Cobrar = 86000 solo válido para AS/MS (actual: %s)",
                row, tipo_id_str,
            )

    return problemas


def _normalize_invoice(value) -> str:
    """Normaliza número de factura a string."""
    if value is None:
        return ""
    if isinstance(value, (int, float)) and value == int(value):
        return str(int(value))
    return str(value).strip()
