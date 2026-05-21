"""Detección de facturas con doble tipo de procedimiento.

Extraído de revision_sheet.py._detect_doble_tipo_procedimiento.
Función autónoma, sin dependencia de área.
"""

from __future__ import annotations

import logging
from collections import defaultdict

from openpyxl.worksheet.worksheet import Worksheet

from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_doble_tipo_procedimiento(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict]:
    """
    Detecta facturas con más de un tipo de procedimiento.

    Args:
        data_sheet: Hoja de Excel con los datos.
        indices: Índices de columnas (debe incluir 'numero_factura' y 'tipo_procedimiento').

    Returns:
        Lista de dicts con keys: "factura", "tipos".
    """
    num_fact_idx = indices.get("numero_factura")
    tipo_proc_idx = indices.get("tipo_procedimiento")

    if num_fact_idx is None or tipo_proc_idx is None:
        return []

    tipo_por_factura: dict[str, set[str]] = {}

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        tipo_value = data_sheet.cell(row=row, column=tipo_proc_idx + 1).value
        if tipo_value is not None:
            tipo_str = str(tipo_value).strip()
            if tipo_str:
                tipo_por_factura.setdefault(factura_str, set()).add(tipo_str)

    result = []
    for fact, tipos in tipo_por_factura.items():
        if len(tipos) > 1:
            result.append({
                "factura": fact,
                "tipos": ", ".join(sorted(tipos)),
            })
    return result
