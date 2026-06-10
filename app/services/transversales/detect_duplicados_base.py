"""Función base parametrizada para detección de duplicados en facturas de farmacia.

Agrupa filas por factura (y opcionalmente codigo_tipo_procedimiento).
Dentro de cada grupo, verifica que TODOS los pares (código, cantidad)
aparezcan al menos 2 veces.
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_duplicados_generico(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
    *,
    tipo_factura: str,
    tarifario_val: str | None = None,
    codigos_tipo_proc: set[str] | None = None,
) -> list[dict[str, Any]]:
    """Detecta grupos de farmacia donde todos los pares están duplicados.

    Filtra filas por tipo_factura. Opcionalmente filtra por tarifario
    y codigo_tipo_procedimiento. Agrupa por factura (y opcionalmente
    codigo_tipo_procedimiento). Dentro de cada grupo, cuenta ocurrencias
    de cada par (codigo, cantidad). Si TODOS los pares tienen count >= 2,
    el grupo completo se marca.

    Cuando ``codigos_tipo_proc`` no es ``None``:
        - Requiere columna ``tarifario`` y ``codigo_tipo_procedimiento``
        - Agrupa por ``(factura, codigo_tipo_procedimiento)``
        - Output incluye ``codigo_tipo_procedimiento``

    Cuando ``codigos_tipo_proc`` es ``None``:
        - No requiere ``tarifario`` ni ``codigo_tipo_procedimiento``
        - Agrupa solo por ``factura``
        - Output NO incluye ``codigo_tipo_procedimiento``

    Args:
        data_sheet: Hoja de Excel con los datos.
        indices: Índices de columnas.
        tipo_factura: Valor de ``tipo_factura_descripcion`` a filtrar.
        tarifario_val: Si se provee, filtra por este valor de tarifario.
        codigos_tipo_proc: Si se provee, filtra por estos códigos y agrupa
            incluyendo ``codigo_tipo_procedimiento``.

    Returns:
        Lista de dicts con keys: ``factura``, [``codigo_tipo_procedimiento``],
        ``pares_duplicados``, ``total_pares``.
    """
    tipo_factura_idx = indices.get("tipo_factura_descripcion")
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    cantidad_idx = indices.get("cantidad")
    tarifario_idx = indices.get("tarifario") if tarifario_val else None
    tipo_proc_idx = indices.get("codigo_tipo_procedimiento") if codigos_tipo_proc else None

    # Guard: columnas requeridas faltantes
    required = [tipo_factura_idx, num_fact_idx]
    if tarifario_val:
        required.append(tarifario_idx)
    if codigos_tipo_proc:
        required.append(tipo_proc_idx)

    if None in required:
        missing = []
        if tipo_factura_idx is None:
            missing.append("tipo_factura_descripcion")
        if num_fact_idx is None:
            missing.append("numero_factura")
        if tarifario_val and tarifario_idx is None:
            missing.append("tarifario")
        if codigos_tipo_proc and tipo_proc_idx is None:
            missing.append("codigo_tipo_procedimiento")
        logger.warning(
            "Duplicados Genérico (%s) - Columnas necesarias no encontradas: %s",
            tipo_factura,
            ", ".join(missing),
        )
        return []

    # Primer pase: agrupar filas filtradas por clave de grupo
    grupos: dict[tuple, dict[tuple, int]] = {}

    for row in range(2, data_sheet.max_row + 1):
        # Filtrar por tipo_factura_descripcion
        tipo_factura_cell = data_sheet.cell(row=row, column=tipo_factura_idx + 1).value
        tipo_factura_str = str(tipo_factura_cell).strip() if tipo_factura_cell else ""
        if tipo_factura_str != tipo_factura:
            continue

        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        # Filtrar por tarifario (si aplica)
        if tarifario_val and tarifario_idx is not None:
            tarifario_cell = data_sheet.cell(row=row, column=tarifario_idx + 1).value
            tarifario_str = str(tarifario_cell).strip() if tarifario_cell else ""
            if tarifario_str != tarifario_val:
                continue

        # Leer y filtrar por codigo_tipo_procedimiento (si aplica)
        tipo_proc_str: str | None = None
        if codigos_tipo_proc and tipo_proc_idx is not None:
            tipo_proc_cell = data_sheet.cell(row=row, column=tipo_proc_idx + 1).value
            tipo_proc_str = str(tipo_proc_cell).strip() if tipo_proc_cell else ""
            if tipo_proc_str not in codigos_tipo_proc:
                continue

        # Leer código (saltar si no hay)
        codigo = None
        if codigo_idx is not None:
            codigo_val = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo = str(codigo_val).strip() if codigo_val else ""
        if not codigo:
            continue

        # Leer cantidad (None → 0)
        cantidad = 0
        if cantidad_idx is not None:
            cantidad_val = data_sheet.cell(row=row, column=cantidad_idx + 1).value
            if cantidad_val is not None:
                try:
                    cantidad = int(cantidad_val)
                except (ValueError, TypeError):
                    cantidad = 0

        # Agrupar
        if codigos_tipo_proc:
            grupo_key = (factura_str, tipo_proc_str)
        else:
            grupo_key = (factura_str,)
        par_key = (codigo, cantidad)
        par_counts = grupos.setdefault(grupo_key, {})
        par_counts[par_key] = par_counts.get(par_key, 0) + 1

    # Segundo pase: emitir un item por grupo donde TODOS los pares tengan count >= 2
    resultados: list[dict[str, Any]] = []
    for grupo_key, par_counts in grupos.items():
        total_pares = len(par_counts)
        pares_duplicados = [
            {"codigo": codigo, "cantidad": cantidad, "count": count}
            for (codigo, cantidad), count in par_counts.items()
            if count >= 2
        ]
        # Flag solo si todos los pares tienen count >= 2
        if len(pares_duplicados) == total_pares:
            item: dict[str, Any] = {
                "factura": grupo_key[0],
                "pares_duplicados": pares_duplicados,
                "total_pares": total_pares,
            }
            if codigos_tipo_proc:
                item["codigo_tipo_procedimiento"] = grupo_key[1]
            resultados.append(item)

    if resultados:
        logger.info(
            "Duplicados Genérico (%s) - %d grupos con duplicidad total encontrados",
            tipo_factura,
            len(resultados),
        )
    else:
        logger.info(
            "Duplicados Genérico (%s) - No se encontraron grupos duplicados",
            tipo_factura,
        )

    return resultados
