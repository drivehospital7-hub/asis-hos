"""Validación de tipo de documento según edad."""

import logging
from datetime import datetime
from typing import TypedDict

from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class TipoDocumentoEdadProblema(TypedDict):
    """Problema encontrado en tipo de documento vs edad."""
    factura: str
    tipo_actual: str
    tipo_deberia: str
    edad: str


def detect_tipo_documento_edad(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[TipoDocumentoEdadProblema]:
    """
    Detecta facturas donde el tipo de identificación no coincide con la edad.
    
    Reglas:
    - < 7 años: RC (Registro Civil)
    - 7-17 años: TI (Tarjeta de Identidad)
    - >= 18 años: CC (Cédula de Ciudadanía)
    - Extranjeros < 18 años: MS
    - Extranjeros >= 18 años: AS
    - CN (Certificado de Nacimiento): solo válido si edad < 2 meses
    - CE (Cédula Extranjería): solo válido si edad > 7 años
    
    Returns:
        Lista de dicts con keys: "factura", "tipo_actual", "tipo_deberia", "edad"
    """
    tipo_id_idx = indices.get("tipo_identificacion")
    fec_nac_idx = indices.get("fec_nacimiento")
    fec_fact_idx = indices.get("fec_factura")
    num_fact_idx = indices.get("numero_factura")
    
    if None in (tipo_id_idx, fec_nac_idx, fec_fact_idx, num_fact_idx):
        logger.warning(
            "No se pueden detectar errores de tipo identificación: "
            "columnas requeridas no encontradas. "
            "tipo_id=%s, fec_nac=%s, fec_fact=%s, num_fact=%s",
            tipo_id_idx, fec_nac_idx, fec_fact_idx, num_fact_idx
        )
        return []
    
    problemas: list[TipoDocumentoEdadProblema] = []
    facturas_ya_procesadas: set[str] = set()
    
    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = _normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_ya_procesadas:
            continue
        
        tipo_id = data_sheet.cell(row=row, column=tipo_id_idx + 1).value
        fec_nac = data_sheet.cell(row=row, column=fec_nac_idx + 1).value
        fec_fact = data_sheet.cell(row=row, column=fec_fact_idx + 1).value
        
        logger.debug(
            "Fila %s: tipo_id=%s, fec_nac=%s, fec_fact=%s",
            row, repr(tipo_id), repr(fec_nac), repr(fec_fact)
        )
        
        if not tipo_id or not fec_nac or not fec_fact:
            continue
        
        tipo_id_str = str(tipo_id).strip().upper()
        
        # Calcular edad
        try:
            fecha_nac = _parse_date(fec_nac)
            fecha_fact = _parse_date(fec_fact)
            
            if fecha_nac is None or fecha_fact is None:
                continue
            
            # Calcular edad en años
            edad = fecha_fact.year - fecha_nac.year
            if (fecha_fact.month, fecha_fact.day) < (fecha_nac.month, fecha_nac.day):
                edad -= 1
            
            # Calcular edad en meses para validaciones especiales (CN)
            edad_meses = (fecha_fact.year - fecha_nac.year) * 12 + (fecha_fact.month - fecha_nac.month)
            
            logger.debug(
                "Fila %s: FechaNac=%s, FechaFact=%s, Edad calculada=%d años, %d meses",
                row, fecha_nac.date(), fecha_fact.date(), edad, edad_meses
            )
        except (ValueError, TypeError) as e:
            logger.debug("Fila %s: Error calculando edad: %s", row, e)
            continue
        
        # Determinar tipo correcto según edad
        tipo_correcto = _determinar_tipo_correcto(tipo_id_str, edad, edad_meses)
        
        logger.debug(
            "Fila %s: Edad=%d, Tipo actual=%s, Tipo correcto=%s",
            row, edad, tipo_id_str, tipo_correcto
        )
        
        # Si hay error, registrar ( deduplicar por factura)
        if tipo_correcto and tipo_id_str != tipo_correcto:
            problemas.append({
                "factura": factura_str,
                "tipo_actual": tipo_id_str,
                "tipo_deberia": tipo_correcto,
                "edad": str(edad),
            })
            facturas_ya_procesadas.add(factura_str)
            logger.debug(
                "Fila %s: Tipo identificación incorrecto (Edad: %d, Tipo: %s, Debería: %s)",
                row,
                edad,
                tipo_id_str,
                tipo_correcto,
            )
    
    return problemas


def _normalize_invoice(value) -> str:
    """Normaliza número de factura a string."""
    if value is None:
        return ""
    return str(value).strip()


def _parse_date(date_value) -> datetime | None:
    """Parsea fecha de varios formatos."""
    if isinstance(date_value, datetime):
        return date_value
    
    date_str = str(date_value).strip()
    
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    
    return None


def _determinar_tipo_correcto(
    tipo_id_str: str,
    edad: int,
    edad_meses: int,
) -> str | None:
    """Determina el tipo de documento correcto según la edad."""
    tipo_correcto: str | None = None
    
    if tipo_id_str in ("RC", "TI", "CC"):
        if edad < 7:
            tipo_correcto = "RC"
        elif edad < 18:
            tipo_correcto = "TI"
        else:
            tipo_correcto = "CC"
    elif tipo_id_str in ("MS", "AS"):
        if edad < 18:
            tipo_correcto = "MS"
        else:
            tipo_correcto = "AS"
    elif tipo_id_str == "CN":
        # CN solo válido si edad < 2 meses
        if edad_meses >= 2:
            tipo_correcto = "ERROR"  # CN no válido para >= 2 meses
    elif tipo_id_str == "CE":
        # CE solo válido si edad > 7 años
        if edad <= 7:
            tipo_correcto = "ERROR"  # CE no válido para <= 7 años
    # Tipos no válidos siempre son error
    elif tipo_id_str in ("NIP", "NIT", "PAS", "PE", "SC"):
        tipo_correcto = "ERROR"  # Tipos no permitidos
    
    return tipo_correcto