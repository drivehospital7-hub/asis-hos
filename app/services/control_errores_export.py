"""Export de Control de Errores a Excel (.xlsx) en memoria."""

import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.constants import IMAGENES_FACTURADOR_SCOPE
from app.utils.errores_storage import listar_imagenes

logger = logging.getLogger(__name__)

HEADERS = [
    "Validador",
    "Creado",
    "Factura",
    "ReFactura",
    "Categoría",
    "Descripción",
    "Responsables",
    "Estado",
    "Adjunto 1",
    "Adjunto 2",
    "Adjunto 3",
    "Observación del Facturador",
    "Adjunto 4",
    "Adjunto 5",
    "Adjunto 6",
]

MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

COLUMN_WIDTH = 20

HEADER_FILL = PatternFill("solid", fgColor="1B5E20")
HEADER_FONT = Font(bold=True, color="FFFFFF")

ROW_FILL_LIGHT = PatternFill("solid", fgColor="E8F5E9")
ROW_FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
DATA_FONT = Font(color="000000")
LINK_FONT = Font(color="0563C1", underline="single")

THIN_BORDER = Border(
    left=Side(style="thin", color="A5D6A7"),
    right=Side(style="thin", color="A5D6A7"),
    top=Side(style="thin", color="A5D6A7"),
    bottom=Side(style="thin", color="A5D6A7"),
)

EXCEL_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".xlsb"}


def filename_export(mes: str | None) -> str:
    """Deriva el nombre del archivo de exportación desde ``mes`` (YYYY-MM).

    Si ``mes`` falta o no es válido, usa la fecha actual.
    """
    if mes and len(mes) == 7 and mes[:4].isdigit() and mes[5:7].isdigit():
        try:
            month_name = MESES[int(mes[5:7]) - 1]
        except IndexError:
            month_name = MESES[datetime.now().month - 1]
        year = mes[:4]
    else:
        now = datetime.now()
        month_name = MESES[now.month - 1]
        year = str(now.year)
    return f"control-errores-{month_name}-{year}.xlsx"


def _formatear_fecha(creado_en: str | None) -> str:
    """Formatea ``creado_en`` ISO como dd/mm/yyyy; si falla, devuelve el raw."""
    try:
        return datetime.fromisoformat(creado_en).strftime("%d/%m/%Y")
    except (TypeError, ValueError):
        return creado_en or ""


def _label_adjunto(filename: str) -> str:
    """Etiqueta amigable para un adjunto según su extensión."""
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return "Abrir PDF"
    if ext in EXCEL_EXTENSIONS:
        return "Abrir Excel"
    return "Abrir imagen"


def _adjunto_url(base_url: str, error_id: str, filename: str, scope: str = "") -> str:
    """URL absoluta para descargar un adjunto sin token ni sesión.

    La ruta de servicio es pública a propósito: los links del Excel deben
    seguir abriendo indefinidamente. El error_id (UUID) hace la URL difícil
    de adivinar y la ruta valida que el archivo pertenezca al registro.
    El ``scope`` aísla los adjuntos de observación (``""``) de los del
    facturador (``"facturador"``) y se añade como query param cuando existe.
    """
    url = f"{base_url}api/control-errores/{error_id}/imagenes/{quote(filename)}"
    if scope:
        url += f"?scope={quote(scope)}"
    return url


def build_errores_export_workbook(errores: list[dict], base_url: str) -> BytesIO:
    """Construye un workbook .xlsx en memoria con los errores dados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Novedades"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTH

    for row_index, error in enumerate(errores):
        error_id = error.get("id", "")
        ws.append([
            error.get("validador", ""),
            _formatear_fecha(error.get("creado_en", "")),
            error.get("factura", ""),
            error.get("refactura", ""),
            error.get("tipo_error", ""),
            error.get("observacion", ""),
            error.get("responsable", ""),
            _label_estado(error.get("estado", "")),
        ])

        row_idx = ws.max_row
        adjuntos = listar_imagenes(error_id)
        link_cols: set[int] = set()
        for i in range(3):
            cell = ws.cell(row=row_idx, column=9 + i)
            if i < len(adjuntos):
                filename = adjuntos[i]
                cell.value = _label_adjunto(filename)
                cell.hyperlink = _adjunto_url(base_url, error_id, filename)
                cell.font = LINK_FONT
                link_cols.add(9 + i)
            else:
                cell.value = ""

        ws.cell(row=row_idx, column=12).value = error.get("observacion_facturador", "")

        adjuntos_facturador = listar_imagenes(error_id, IMAGENES_FACTURADOR_SCOPE)
        for i in range(3):
            cell = ws.cell(row=row_idx, column=13 + i)
            if i < len(adjuntos_facturador):
                filename = adjuntos_facturador[i]
                cell.value = _label_adjunto(filename)
                cell.hyperlink = _adjunto_url(
                    base_url, error_id, filename, IMAGENES_FACTURADOR_SCOPE
                )
                cell.font = LINK_FONT
                link_cols.add(13 + i)
            else:
                cell.value = ""

        row_fill = ROW_FILL_LIGHT if row_index % 2 == 0 else ROW_FILL_WHITE
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            if col not in link_cols:
                cell.font = DATA_FONT

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _label_estado(estado: str) -> str:
    """Traduce el código de estado a etiqueta legible."""
    if estado == "S":
        return "Pendiente"
    if estado == "N":
        return "Resuelto"
    return estado