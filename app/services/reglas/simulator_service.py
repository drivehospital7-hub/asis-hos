"""Simulator service — dry-run comparison between engine and legacy detectors.

Parses an uploaded Excel file (max 100 rows), runs both the DB-backed
RuleBasedDetector and legacy Python detectors, and returns a diff comparison.
"""

from __future__ import annotations

import logging
from io import BytesIO
from typing import Any

import polars as pl
from openpyxl import load_workbook

from app.services.engine.rule_based_detector import RuleBasedDetector
from app.services.transversales.decimales import detect_decimales
from app.services.transversales.ruta_duplicada import detect_ruta_duplicada

logger = logging.getLogger(__name__)

_MAX_ROWS = 100
_ALLOWED_EXTENSIONS = {".xlsx", ".xls"}


def _excel_to_sheet(file_bytes: bytes) -> tuple[Any, dict[str, int | None]]:
    """Convert uploaded Excel bytes to openpyxl Worksheet + column indices.

    Returns:
        (worksheet, indices_dict)
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("El archivo Excel no tiene hojas activas")

    # Build indices from header row
    indices: dict[str, int | None] = {}
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    for col_idx, cell_value in enumerate(header_row):
        if cell_value is not None:
            indices[str(cell_value)] = col_idx

    return ws, indices


def _build_diff(
    engine_results: list[dict],
    legacy_results: list[dict],
) -> dict:
    """Build a comparison diff between engine and legacy results.

    Compares results by (factura, problema) tuple.
    Returns matched, engine-only, and legacy-only counts with details.
    """
    engine_set = {
        (r.get("factura"), r.get("problema"))
        for r in engine_results
    }
    legacy_set = {
        (r.get("factura"), r.get("problema"))
        for r in legacy_results
    }

    matched = engine_set & legacy_set
    engine_only = engine_set - legacy_set
    legacy_only = legacy_set - engine_set

    return {
        "matched": sorted([{"factura": f, "problema": p} for f, p in matched], key=lambda x: x["factura"]),
        "engine_only": sorted([{"factura": f, "problema": p} for f, p in engine_only], key=lambda x: x["factura"]),
        "legacy_only": sorted([{"factura": f, "problema": p} for f, p in legacy_only], key=lambda x: x["factura"]),
        "matched_count": len(matched),
        "engine_only_count": len(engine_only),
        "legacy_only_count": len(legacy_only),
        "engine_total": len(engine_results),
        "legacy_total": len(legacy_results),
    }


def simulate(
    db_session,
    file_storage,
    rule_name: str | None = None,
) -> dict:
    """Run a dry-run simulation comparing engine vs legacy detectors.

    Args:
        db_session: SQLAlchemy Session for DB access
        file_storage: werkzeug FileStorage (uploaded Excel)
        rule_name: Optional rule name to filter engine evaluation

    Returns:
        dict with engine_results, legacy_results, diff, and metadata

    Raises:
        ValueError: If file format is invalid
    """
    # Validate file extension
    filename = file_storage.filename or ""
    ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in _ALLOWED_EXTENSIONS:
        raise ValueError(
            "Formato no válido. Seleccioná un archivo Excel (.xlsx o .xls)."
        )

    # Read Excel with Polars (limit to first 100 rows)
    file_bytes = file_storage.read()
    try:
        df = pl.read_excel(BytesIO(file_bytes))
    except Exception as exc:
        raise ValueError(f"No se pudo leer el archivo Excel: {exc}") from exc

    total_rows = len(df)
    df_limited = df.head(_MAX_ROWS)
    truncated = total_rows > _MAX_ROWS

    # Convert to openpyxl worksheet for legacy detectors
    ws, indices = _excel_to_sheet(file_bytes)

    # Run engine detector
    rule_name_to_use = rule_name or "valores_decimales"
    detector = RuleBasedDetector(rule_name_to_use, db_session)
    engine_results = detector.detect(ws, indices)

    # Run legacy detectors
    legacy_results: list[dict] = []
    legacy_results.extend(detect_decimales(ws, indices) or [])
    legacy_results.extend(detect_ruta_duplicada(ws, indices) or [])

    # Build diff
    diff = _build_diff(engine_results, legacy_results)

    return {
        "engine_results": engine_results,
        "legacy_results": legacy_results,
        "diff": diff,
        "total_rows": total_rows,
        "rows_processed": min(total_rows, _MAX_ROWS),
        "truncated": truncated,
    }
