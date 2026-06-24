"""Orquestador de detección de problemas para Odontología.

Agrupa todos los detectores específicos de odontología más los
detectores transversales aplicables a esta área.
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import AREA_ODONTOLOGIA, CONVENIO_PYP
from app.constants.base import is_rule_engine_enabled
from app.services.transversales import (
    detect_decimales,
    detect_tipo_documento_edad,
    detect_tipo_identificacion_entidad,
    detect_codigo_entidad_vs_entidad_afiliacion,
    detect_tipo_usuario,
    detect_doble_tipo_procedimiento,
    detect_ruta_duplicada,
    detect_cantidades_anomalas,
    normalize_invoice,
)
from app.services.odontologia.profesionales import detect_profesionales_odontologia
from app.services.odontologia.centro_costo import detect_centro_costo_odontologia
from app.services.odontologia.ide_contrato import detect_ide_contrato_odontologia
from app.services.urgencias.mal_capitado import detect_mal_capitado
from app.services.transversales.procedimiento_contratado import detect_cups_sin_contrato

logger = logging.getLogger(__name__)


def detect_all_problems_odontologia(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
    profesional_dias: dict[str, list[int]] | None = None,
    permitir_todos_centros: bool = True,
    centros_validos: list[str] | None = None,
) -> tuple[dict[str, Any], dict[str, str]]:
    """
    Detecta TODOS los problemas en facturas de odontología.

    Incluye detectores transversales y específicos de odontología.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas
        profesional_dias: Dict {identificacion: [dias]} con días seleccionados
        permitir_todos_centros: Si True, solo permite ODONTOLOGIA y EXTRAMURAL
        centros_validos: Lista personalizada de centros válidos (opcional)

    Returns:
        (resultado_dict, responsables_map)
    """
    # Detectores transversales
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            decimales = RuleBasedDetector("valores_decimales", session).detect(data_sheet, indices)
            session.commit()
        finally:
            session.close()
    else:
        decimales = detect_decimales(data_sheet, indices)
    doble_tipo = detect_doble_tipo_procedimiento(data_sheet, indices)

    # Excepción odontología: código 990203 puede tener múltiples tipos de procedimiento
    codigo_idx = indices.get("codigo")
    num_fact_idx = indices.get("numero_factura")
    if codigo_idx is not None and num_fact_idx is not None:
        facturas_con_990203: set[str] = set()
        for row in range(2, data_sheet.max_row + 1):
            codigo_val = data_sheet.cell(row=row, column=codigo_idx + 1).value
            if codigo_val is not None and str(codigo_val).strip() == "990203":
                numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
                factura = normalize_invoice(numero)
                if factura:
                    facturas_con_990203.add(factura)
        if facturas_con_990203:
            antes = len(doble_tipo)
            doble_tipo = [
                item for item in doble_tipo
                if item.get("factura") not in facturas_con_990203
            ]
            despues = len(doble_tipo)
            if despues < antes:
                logger.info(
                    "Excepción código 990203: %d facturas excluidas de doble tipo procedimiento",
                    antes - despues,
                )

    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            ruta_dup = RuleBasedDetector("ruta_duplicada", session).detect(data_sheet, indices)
            session.commit()
        finally:
            session.close()
    else:
        ruta_dup = detect_ruta_duplicada(data_sheet, indices)

    # Excepción odontología: si ruta duplicada es exactamente 3 facturas y
    # alguna tiene código 990203, P0000011 o 990212, se excluye
    RUTA_DUP_EXEMPT_CODES = frozenset({"990203", "P0000011", "990212"})
    codigo_idx = indices.get("codigo")
    ident_idx = indices.get("identificacion")
    if ruta_dup and codigo_idx is not None and ident_idx is not None:
        # Construir mapa de códigos por paciente (solo PyP)
        codigos_por_paciente: dict[str, set[str]] = {}
        contrato_idx = indices.get("convenio_facturado")
        for row in range(2, data_sheet.max_row + 1):
            contrato_val = (
                data_sheet.cell(row=row, column=contrato_idx + 1).value
                if contrato_idx is not None else None
            )
            if contrato_idx is not None and contrato_val != CONVENIO_PYP:
                continue
            ident_val = data_sheet.cell(row=row, column=ident_idx + 1).value
            codigo_val = data_sheet.cell(row=row, column=codigo_idx + 1).value
            if ident_val is not None and codigo_val is not None:
                ident_str = str(ident_val).strip()
                codigo_str = str(codigo_val).strip()
                if ident_str and codigo_str:
                    if ident_str not in codigos_por_paciente:
                        codigos_por_paciente[ident_str] = set()
                    codigos_por_paciente[ident_str].add(codigo_str)

        if codigos_por_paciente:
            antes = len(ruta_dup)
            ruta_dup = [
                item for item in ruta_dup
                if not (
                    item["cantidad"] == 3
                    and RUTA_DUP_EXEMPT_CODES & codigos_por_paciente.get(item["identificacion"], set())
                )
            ]
            despues = len(ruta_dup)
            if despues < antes:
                logger.info(
                    "Excepción códigos PyP (990203, P0000011, 990212): %d rutas duplicadas excluidas",
                    antes - despues,
                )

    tipo_id_edad = detect_tipo_documento_edad(data_sheet, indices)
    tipo_id_entidad = detect_tipo_identificacion_entidad(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            r1 = RuleBasedDetector("tipo_id_requiere_entidad_86000", session).detect(data_sheet, indices)
            r2 = RuleBasedDetector("entidad_86000_requiere_as_ms", session).detect(data_sheet, indices)
            tipo_id_entidad = r1 + r2
            session.commit()
        finally:
            session.close()
    cantidades = detect_cantidades_anomalas(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            c1 = RuleBasedDetector("cantidad_consultas_anomalas", session).detect(data_sheet, indices)
            c2 = RuleBasedDetector("cantidad_general_anomalas", session).detect(data_sheet, indices)
            c3 = RuleBasedDetector("cantidad_pyp_anomalas", session).detect(data_sheet, indices)
            cantidades = c1 + c2 + c3
            session.commit()
        finally:
            session.close()
    entidad_afiliacion_comparison = detect_codigo_entidad_vs_entidad_afiliacion(
        data_sheet, indices, limit_log=5
    )
    tipo_usuario_od = detect_tipo_usuario(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            tipo_usuario_od = RuleBasedDetector("tipo_usuario_valido", session).detect(data_sheet, indices)
            session.commit()
        finally:
            session.close()

    # Detectores específicos de odontología
    logger.info("detect_all_problems_odontologia - Llamando detect_ide_contrato_odontologia")
    ide_contrato = detect_ide_contrato_odontologia(data_sheet, indices)
    logger.info("detect_all_problems_odontologia - IDE Contrato encontrados: %d", len(ide_contrato))

    logger.info("detect_all_problems_odontologia - Llamando detect_profesionales_odontologia")
    profesionales = detect_profesionales_odontologia(data_sheet, indices)
    logger.info("detect_all_problems_odontologia - Profesionales encontrados: %d", len(profesionales))

    centro_costo = detect_centro_costo_odontologia(
        data_sheet,
        indices,
        profesional_dias=profesional_dias,
        permitir_todos_centros=permitir_todos_centros,
        centros_validos=centros_validos,
    )

    cups_sin_contrato = detect_cups_sin_contrato(data_sheet, indices)
    logger.info(
        "detect_all_problems_odontologia - Cups Sin Contrato encontrados: %d",
        len(cups_sin_contrato),
    )

    # Build responsable_cierra mapping
    responsable_cierra: dict[str, str] = {}
    responsable_cierra_idx = indices.get("responsable_cierra")
    num_fact_idx = indices.get("numero_factura")
    if responsable_cierra_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            raw = data_sheet.cell(row=row, column=responsable_cierra_idx + 1).value
            resp = str(raw).strip() if raw else ""
            if resp and factura not in responsable_cierra:
                responsable_cierra[factura] = resp

    # Build fec_factura_map
    fec_factura_map: dict[str, str] = {}
    fec_factura_idx = indices.get("fec_factura")
    if fec_factura_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            raw = data_sheet.cell(row=row, column=fec_factura_idx + 1).value
            val = str(raw).strip() if raw else ""
            if val and factura not in fec_factura_map:
                fec_factura_map[factura] = val

    # Build normalized rows for unified 6-column display
    from app.services.odontologia.normalized_rows import build_odontologia_normalized_rows
    normalized_rows_od = build_odontologia_normalized_rows(
        decimales=decimales,
        doble_tipo=doble_tipo,
        ruta_dup=ruta_dup,
        profesionales=profesionales,
        cantidades=cantidades,
        tipo_id_edad=tipo_id_edad,
        tipo_id_entidad=tipo_id_entidad,
        centro_costo=centro_costo,
        ide_contrato=ide_contrato,
        responsable_cierra=responsable_cierra,
        entidad_afiliacion_comparison=entidad_afiliacion_comparison,
        tipo_usuario=tipo_usuario_od,
        fec_factura_map=fec_factura_map,
        cups_sin_contrato=cups_sin_contrato,
    )

    resultado: dict[str, Any] = {
        "area": AREA_ODONTOLOGIA,
        "problemas": {
            "normalizados": normalized_rows_od,
            "decimales": decimales,
            "doble_tipo_procedimiento": doble_tipo,
            "ruta_duplicada": ruta_dup,
            "profesionales": profesionales,
            "cantidades_anomalas": cantidades,
            "tipo_identificacion_edad": tipo_id_edad,
            "tipo_identificacion_entidad": tipo_id_entidad,
            "codigo_entidad_vs_afiliacion": entidad_afiliacion_comparison,
            "tipo_usuario": tipo_usuario_od,
            "centro_costo": centro_costo,
            "ide_contrato": ide_contrato,
            "cups_sin_contrato": cups_sin_contrato,
        },
        "totales": {
            "decimales": len(decimales),
            "doble_tipo_procedimiento": len(doble_tipo),
            "ruta_duplicada": len(ruta_dup),
            "profesionales": len(profesionales),
            "cantidades_anomalas": len(cantidades),
            "tipo_identificacion_edad": len(tipo_id_edad),
            "tipo_identificacion_entidad": len(tipo_id_entidad),
            "centro_costo": len(centro_costo),
            "ide_contrato": len(ide_contrato),
            "codigo_entidad_vs_afiliacion": len(entidad_afiliacion_comparison),
            "tipo_usuario": len(tipo_usuario_od),
            "cups_sin_contrato": len(cups_sin_contrato),
        },
        "missing_columns": [],
    }

    # Enrich errors with responsable from mapping
    if responsable_cierra:
        for problem_type, problems in resultado["problemas"].items():
            if not isinstance(problems, list):
                continue
            for p in problems:
                factura = p.get("factura")
                if factura and factura in responsable_cierra:
                    p["responsable"] = responsable_cierra[factura]
                elif "responsable" not in p:
                    p["responsable"] = ""
    else:
        for problem_type, problems in resultado["problemas"].items():
            if not isinstance(problems, list):
                continue
            for p in problems:
                if "responsable" not in p:
                    p["responsable"] = ""

    return resultado, responsable_cierra
