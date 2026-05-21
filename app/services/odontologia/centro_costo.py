"""Detector de centro de costo específico de odontología.

Dos modos de operación:
1. permitir_todos_centros = True (validación desactivada):
   - Solo se permiten: "ODONTOLOGIA" y "SERVICIOS ODONTOLOGIA -EXTRAMURALES"
   - Cualquier otro centro es error

2. permitir_todos_centros = False (validación activada con días):
   - Por defecto: Centro debe ser "ODONTOLOGIA"
   - Si el profesional tiene días seleccionados en el calendario
     Y la fecha de factura coincide con uno de esos días -> Centro debe ser EXTRAMURAL
"""

from __future__ import annotations

import logging
from datetime import datetime
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import CENTRO_COSTO_ODONTOLOGIA, CENTRO_COSTO_EXTRAMURAL
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_centro_costo_odontologia(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
    profesional_dias: dict[str, list[int]] | None = None,
    permitir_todos_centros: bool = False,
    centros_validos: list[str] | None = None,
) -> list[dict[str, str]]:
    """
    Detecta facturas con problemas de centro de costo en Odontología.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas
        profesional_dias: Dict {identificacion: [dias]} con días seleccionados
        permitir_todos_centros: Si True, solo permite ODONTOLOGIA y EXTRAMURAL
        centros_validos: Lista personalizada de centros válidos

    Returns:
        Lista de dicts con keys: "factura", "centro_actual", "centro_deberia",
        "profesional", "fec_factura"
    """
    problemas: list[dict[str, str]] = []

    # Valores por defecto
    if centros_validos is None:
        centros_validos = [CENTRO_COSTO_ODONTOLOGIA, CENTRO_COSTO_EXTRAMURAL]

    num_fact_idx = indices.get("numero_factura")
    centro_costo_idx = indices.get("centro_costo")
    fec_factura_idx = indices.get("fec_factura")
    profesional_id_idx = indices.get("profesional_identificacion")

    if num_fact_idx is None or centro_costo_idx is None:
        logger.warning("Columnas necesarias no encontradas para validar centro de costo")
        return []

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        centro_costo = data_sheet.cell(row=row, column=centro_costo_idx + 1).value
        centro_costo_str = str(centro_costo).strip().upper() if centro_costo else ""

        # Obtener fecha de factura
        fec_factura = data_sheet.cell(row=row, column=fec_factura_idx + 1).value if fec_factura_idx is not None else None
        dia_factura: int | None = None
        fec_factura_dt: datetime | None = None

        if fec_factura:
            try:
                if isinstance(fec_factura, datetime):
                    dia_factura = fec_factura.day
                    fec_factura_dt = fec_factura
                elif isinstance(fec_factura, (int, float)):
                    from datetime import datetime as dt, timedelta
                    excel_date = int(fec_factura)
                    fec_factura_dt = dt(1900, 1, 1) + timedelta(days=excel_date - 1)
                    dia_factura = fec_factura_dt.day
                elif isinstance(fec_factura, str):
                    formatos = [
                        "%Y-%m-%d %H:%M:%S",
                        "%Y-%m-%d",
                        "%d/%m/%Y", "%d-%m-%Y",
                        "%d-%b-%Y", "%b %d, %Y", "%d %b %Y",
                        "%m/%d/%Y", "%Y/%m/%d",
                        "%d.%m.%Y", "%Y.%m.%d",
                    ]
                    for fmt in formatos:
                        try:
                            fec_factura_dt = datetime.strptime(fec_factura.strip(), fmt)
                            dia_factura = fec_factura_dt.day
                            break
                        except ValueError:
                            continue
            except Exception:
                pass

        # Obtener identificación del profesional
        profesional_id: str | None = None
        if profesional_id_idx is not None:
            pid = data_sheet.cell(row=row, column=profesional_id_idx + 1).value
            if pid:
                profesional_id = str(pid).strip()

        # Determinar centro correcto según el modo
        centro_correcto: str | None = None
        if not permitir_todos_centros:
            dias_profesional: list[int] = []
            if profesional_dias and profesional_id and profesional_id in profesional_dias:
                dias_profesional = profesional_dias[profesional_id]

            if dia_factura and dias_profesional and dia_factura in dias_profesional:
                centro_correcto = CENTRO_COSTO_EXTRAMURAL
            else:
                centro_correcto = CENTRO_COSTO_ODONTOLOGIA

        # Validar
        if centros_validos is None:
            centros_validos = [CENTRO_COSTO_ODONTOLOGIA, CENTRO_COSTO_EXTRAMURAL]

        # Caso 1: Centro no está en la lista de válidos
        if centro_costo_str not in centros_validos:
            problema: dict[str, str] = {
                "factura": factura_str,
                "centro_actual": centro_costo_str,
                "centro_deberia": centro_correcto if centro_correcto else " o ".join(centros_validos),
                "profesional": profesional_id or "",
                "fec_factura": fec_factura_dt.strftime("%Y-%m-%d") if fec_factura_dt else "",
            }
            problemas.append(problema)
        # Caso 2: Validación activada Y centro no coincide con el esperado
        elif not permitir_todos_centros and centro_correcto and centro_costo_str != centro_correcto:
            problema = {
                "factura": factura_str,
                "centro_actual": centro_costo_str,
                "centro_deberia": centro_correcto,
                "profesional": profesional_id or "",
                "fec_factura": fec_factura_dt.strftime("%Y-%m-%d") if fec_factura_dt else "",
            }
            problemas.append(problema)

    return problemas
