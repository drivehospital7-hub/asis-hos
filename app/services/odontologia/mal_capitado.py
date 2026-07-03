"""Detector de facturas mal capitadas en odontología.

Reglas:
- Códigos G03XB01 y A02BB01 DEBEN tener Número Factura con prefijo "FEV"
- Si Número Factura tiene prefijo CAP -> Cód Entidad Cobrar debe ser ESS118
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    CODIGOS_MAL_CAPITADO,
    PREFIJO_FACTURA_MAL_CAPITADO,
    PREFIJO_FACTURA_CAP,
    ENTIDAD_REQUERIDA_CAP,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_mal_capitado(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con códigos G03XB01 o A02BB01 que NO tienen prefijo FEV.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo", "procedimiento",
        "observacion", "ide_contrato_actual"
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")
    ide_contrato_idx = indices.get("ide_contrato")

    if num_fact_idx is None or codigo_idx is None:
        logger.warning("MAL CAPITADO - Columnas necesarias no encontradas")
        return []

    problemas: list[dict[str, str]] = []
    facturas_procesadas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_procesadas:
            continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip().upper() if codigo else ""

        # Solo verificar si el código es uno de los MAL CAPITADO
        if codigo_str not in CODIGOS_MAL_CAPITADO:
            continue

        # Verificar prefijo FEV
        tiene_prefijo_fev = factura_str.upper().startswith(PREFIJO_FACTURA_MAL_CAPITADO)

        if not tiene_prefijo_fev:
            procedimiento = ""
            if procedimiento_idx is not None:
                proc_value = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc_value).strip() if proc_value else ""

            ide_contrato_actual = ""
            if ide_contrato_idx is not None:
                ide_val = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value
                ide_contrato_actual = str(ide_val).strip() if ide_val else ""

            problemas.append({
                "factura": factura_str,
                "codigo": codigo_str,
                "procedimiento": procedimiento,
                "ide_contrato_actual": ide_contrato_actual,
                "observacion": "En caso de ser IVE pasar a Evento",
            })
            facturas_procesadas.add(factura_str)

    # Regla: Si Número Factura tiene prefijo CAP -> Cód Entidad Cobrar debe ser ESS118
    codigo_entidad_cobrar_idx = indices.get("codigo_entidad_cobrar")
    if num_fact_idx is not None and codigo_entidad_cobrar_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura_str = normalize_invoice(numero_factura)
            if not factura_str:
                continue

            if factura_str.upper().startswith(PREFIJO_FACTURA_CAP):
                entidad_val = data_sheet.cell(row=row, column=codigo_entidad_cobrar_idx + 1).value
                entidad_str = str(entidad_val).strip() if entidad_val else ""

                if entidad_str != ENTIDAD_REQUERIDA_CAP:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": "",
                        "procedimiento": "",
                        "ide_contrato_actual": "",
                        "observacion": (
                            f"Número Factura con prefijo CAP requiere "
                            f"Cód Entidad Cobrar={ENTIDAD_REQUERIDA_CAP} "
                            f"(actual: {entidad_str})"
                        ),
                    })

    return problemas
