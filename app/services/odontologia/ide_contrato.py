"""Detector de IDE Contrato específico de odontología (PyP).

Reglas:
- ESS118 + Código PyP -> IDE debe ser 970 o 974
- ESS118 + Código NO PyP -> IDE debe ser 969 o 973
- ESSC18 + Código PyP -> IDE debe ser 975
- ESSC18 + Código NO PyP -> IDE debe ser 968
- (y otras combinaciones de entidad + código)
"""

from __future__ import annotations

import logging
from datetime import datetime
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    PYP_CUPS_CODES,
    IDE_CONTRATO_MULTIPLE_ESS118_PYP,
    IDE_CONTRATO_MULTIPLE_ESS118_NO_PYP,
    IDE_CONTRATO_MULTIPLE_ESSC18_PYP,
    IDE_CONTRATO_MULTIPLE_ESSC18_NO_PYP,
    IDE_CONTRATO_MULTIPLE_EPSS41_PYP,
    IDE_CONTRATO_MULTIPLE_EPSS41_NO_PYP,
    IDE_CONTRATO_MULTIPLE_EPS037_PYP,
    IDE_CONTRATO_MULTIPLE_EPS037_NO_PYP,
    IDE_CONTRATO_MULTIPLE_EPSI05_PYP,
    IDE_CONTRATO_MULTIPLE_EPSI05_NO_PYP,
    IDE_CONTRATO_MULTIPLE_EPSIC5_PYP,
    IDE_CONTRATO_MULTIPLE_EPSIC5_NO_PYP,
    IDE_CONTRATO_MULTIPLE_RES001_PYP,
    IDE_CONTRATO_MULTIPLE_RES001_NO_PYP,
    IDE_CONTRATO_MULTIPLE_ESS062_PYP,
    IDE_CONTRATO_MULTIPLE_ESS062_NO_PYP,
    IDE_CONTRATO_MULTIPLE_ESSC62_PYP,
    IDE_CONTRATO_MULTIPLE_ESSC62_NO_PYP,
    IDE_CONTRATO_MULTIPLE_0001_PYP,
    IDE_CONTRATO_MULTIPLE_0001_NO_PYP,
    IDE_CONTRATO_MULTIPLE_EPSS005_PYP,
    IDE_CONTRATO_MULTIPLE_EPSS005_NO_PYP,
    IDE_CONTRATO_MULTIPLE_EPSC005_PYP,
    IDE_CONTRATO_MULTIPLE_EPSC005_NO_PYP,
    IDE_CONTRATO_MULTIPLE_86_NO_PYP,
    IDE_CONTRATO_MULTIPLE_86000_PYP,
    IDE_CONTRATO_MULTIPLE_86000_NO_PYP,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


RES001_MESES_HISTORICO: dict[int, tuple[frozenset[str], frozenset[str]]] = {
    5: (frozenset({"993", "954"}), frozenset({"992", "953"})),  # Mayo: nuevos + históricos
}


def _determinar_ide_esperado(
    codigo_entidad_str: str,
    codigo_str: str,
    mes: int | None = None,
) -> tuple[frozenset[str] | None, str]:
    """
    Determina el IDE Contrato esperado según entidad y código.

    Args:
        codigo_entidad_str: Código de entidad (ESS118, ESSC18, etc.)
        codigo_str: Código CUPS del procedimiento
        mes: Mes de factura (1-12, None si no disponible)

    Returns:
        Tuple de (set de IDE esperados, nota descriptiva)
    """
    es_pyp = codigo_str in PYP_CUPS_CODES

    if codigo_entidad_str == "ESS118":
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_ESS118_PYP, "ESS118 + PyP"
        return IDE_CONTRATO_MULTIPLE_ESS118_NO_PYP, "ESS118 + NO PyP"

    if codigo_entidad_str == "ESSC18":
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_ESSC18_PYP, "ESSC18 + PyP"
        return IDE_CONTRATO_MULTIPLE_ESSC18_NO_PYP, "ESSC18 + NO PyP"

    if codigo_entidad_str == "EPSS41":
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_EPSS41_PYP, "EPSS41 + PyP"
        return IDE_CONTRATO_MULTIPLE_EPSS41_NO_PYP, "EPSS41 + NO PyP"

    if codigo_entidad_str == "EPS037":
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_EPS037_PYP, "EPS037 + PyP"
        return IDE_CONTRATO_MULTIPLE_EPS037_NO_PYP, "EPS037 + NO PyP"

    if codigo_entidad_str == "EPSI05":
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_EPSI05_PYP, "EPSI05 + PyP"
        return IDE_CONTRATO_MULTIPLE_EPSI05_NO_PYP, "EPSI05 + NO PyP"

    if codigo_entidad_str == "EPSIC5":
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_EPSIC5_PYP, "EPSIC5 + PyP"
        return IDE_CONTRATO_MULTIPLE_EPSIC5_NO_PYP, "EPSIC5 + NO PyP"

    if codigo_entidad_str == "RES001":
        if mes is not None and mes in RES001_MESES_HISTORICO:
            pyp_set, no_pyp_set = RES001_MESES_HISTORICO[mes]
            if es_pyp:
                return pyp_set, f"RES001 + PyP (mes {mes})"
            return no_pyp_set, f"RES001 + NO PyP (mes {mes})"
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_RES001_PYP, "RES001 + PyP"
        return IDE_CONTRATO_MULTIPLE_RES001_NO_PYP, "RES001 + NO PyP"

    if codigo_entidad_str == "ESS062":
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_ESS062_PYP, "ESS062 + PyP"
        return IDE_CONTRATO_MULTIPLE_ESS062_NO_PYP, "ESS062 + NO PyP"

    if codigo_entidad_str == "ESSC62":
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_ESSC62_PYP, "ESSC62 + PyP"
        return IDE_CONTRATO_MULTIPLE_ESSC62_NO_PYP, "ESSC62 + NO PyP"

    if codigo_entidad_str == "0001":
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_0001_PYP, "0001 + PyP"
        return IDE_CONTRATO_MULTIPLE_0001_NO_PYP, "0001 + NO PyP"

    if codigo_entidad_str == "EPSS005":
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_EPSS005_PYP, "EPSS005 + PyP"
        return IDE_CONTRATO_MULTIPLE_EPSS005_NO_PYP, "EPSS005 + NO PyP"

    if codigo_entidad_str == "EPSC005":
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_EPSC005_PYP, "EPSC005 + PyP"
        return IDE_CONTRATO_MULTIPLE_EPSC005_NO_PYP, "EPSC005 + NO PyP"

    if codigo_entidad_str == "86" and not es_pyp:
        return IDE_CONTRATO_MULTIPLE_86_NO_PYP, "86 + NO PyP"

    if codigo_entidad_str == "86000":
        if es_pyp:
            return IDE_CONTRATO_MULTIPLE_86000_PYP, "86000 + PyP"
        return IDE_CONTRATO_MULTIPLE_86000_NO_PYP, "86000 + NO PyP"

    # Entidad sin regla específica
    return None, ""


def detect_ide_contrato_odontologia(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con problemas de IDE Contrato en Odontología.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo", "cod_entidad",
        "ide_actual", "ide_deberia", "nota"
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_entidad_idx = indices.get("codigo_entidad_cobrar")
    codigo_idx = indices.get("codigo")
    ide_contrato_idx = indices.get("ide_contrato")
    fec_factura_idx = indices.get("fec_factura")

    if None in (num_fact_idx, codigo_entidad_idx, codigo_idx, ide_contrato_idx):
        logger.warning(
            "IDE Contrato - Columnas necesarias no encontradas"
        )
        return []

    problemas: list[dict[str, str]] = []
    facturas_procesadas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue
        if factura_str in facturas_procesadas:
            continue

        codigo_entidad = data_sheet.cell(row=row, column=codigo_entidad_idx + 1).value
        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        ide_contrato = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value

        if codigo_entidad is None or codigo is None or ide_contrato is None:
            continue

        codigo_entidad_str = str(codigo_entidad).strip().upper()
        codigo_str = str(codigo).strip().upper()
        ide_str = str(ide_contrato).strip()

        # Extraer mes de fec_factura para reglas históricas
        mes: int | None = None
        if fec_factura_idx is not None and codigo_entidad_str == "RES001":
            fecha_raw = data_sheet.cell(row=row, column=fec_factura_idx + 1).value
            if fecha_raw:
                try:
                    fecha_dt = datetime.strptime(str(fecha_raw).strip(), "%Y-%m-%d %H:%M:%S")
                    mes = fecha_dt.month
                except (ValueError, TypeError):
                    pass

        # Determinar IDE esperado según entidad y código
        ide_esperado_set, nota = _determinar_ide_esperado(codigo_entidad_str, codigo_str, mes)

        if ide_esperado_set is None:
            # Entidad no tiene regla específica
            continue

        if ide_str not in ide_esperado_set:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "cod_entidad": codigo_entidad_str,
                    "ide_actual": ide_str,
                    "ide_deberia": " o ".join(sorted(ide_esperado_set)),
                    "nota": nota,
                })
                facturas_procesadas.add(factura_str)

    return problemas
