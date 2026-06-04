"""Detector de profesionales de odontología no válidos.

Reglas (Odontología):
- "Código Profesional" DEBE estar en PROFESIONALES_ODONTOLOGIA_VALIDACION
- HIGIENISTA: Solo puede usar códigos en PYP_CODES_HIGIENISTA
- ODONTOLOGO: Puede usar cualquier código EXCEPTO los que están en PYP_CODES_HIGIENISTA
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    PROFESIONALES_ODONTOLOGIA_VALIDACION,
    PYP_CODES_HIGIENISTA,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_profesionales_odontologia(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con profesionales no válidos en Odontología.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo_profesional", "nombre",
        "tipo", "profesional_area", "procedimiento", "regla", "problema"
    """
    num_fact_idx = indices.get("numero_factura")
    cod_prof_idx = indices.get("codigo_profesional")
    codigo_idx = indices.get("codigo")

    if num_fact_idx is None or cod_prof_idx is None:
        return []

    problemas: list[dict[str, str]] = []
    facturas_procesadas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_procesadas:
            continue

        cod_profesional = data_sheet.cell(row=row, column=cod_prof_idx + 1).value
        cod_profesional_str = str(cod_profesional).strip() if cod_profesional else ""

        if not cod_profesional_str:
            continue

        # Buscar profesional en el diccionario
        profesional_info = PROFESIONALES_ODONTOLOGIA_VALIDACION.get(cod_profesional_str)

        if profesional_info is None:
            problemas.append({
                "factura": factura_str,
                "codigo_profesional": cod_profesional_str,
                "nombre": "",
                "tipo": "",
                "profesional_area": "",
                "procedimiento": "",
                "regla": "Profesional debe estar en listado",
                "problema": "Profesional no existe en el listado de Odontología",
            })
            facturas_procesadas.add(factura_str)
            continue

        # Obtener código del procedimiento
        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value if codigo_idx else None
        codigo_str = str(codigo).strip() if codigo else ""

        tipo_profesional = profesional_info.get("tipo", "")

        # Validar según tipo de profesional
        if tipo_profesional == "HIGIENISTA" and codigo_str not in PYP_CODES_HIGIENISTA:
            problemas.append({
                "factura": factura_str,
                "codigo_profesional": cod_profesional_str,
                "nombre": profesional_info.get("nombre", ""),
                "tipo": "HIGIENISTA",
                "profesional_area": "HIGIENISTA",
                "procedimiento": codigo_str,
                "regla": "Solo códigos PYP",
                "problema": "HIGIENISTA no puede usar código no PYP",
            })
            facturas_procesadas.add(factura_str)

        elif tipo_profesional == "ODONTOLOGO" and codigo_str in PYP_CODES_HIGIENISTA:
            problemas.append({
                "factura": factura_str,
                "codigo_profesional": cod_profesional_str,
                "nombre": profesional_info.get("nombre", ""),
                "tipo": "ODONTOLOGO",
                "profesional_area": "ODONTOLOGO",
                "procedimiento": codigo_str,
                "regla": "No códigos PYP (excepto 890203)",
                "problema": "ODONTOLOGO no puede usar código PYP",
            })
            facturas_procesadas.add(factura_str)

    return problemas
