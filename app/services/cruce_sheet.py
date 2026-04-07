"""Servicio para manejo de hoja CruceFacturas.

Este módulo se encarga de crear y configurar la hoja CruceFacturas
donde se listan las facturas para cruce.
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.constants import CRUCE_FACTURAS_SHEET, CRUCE_HEADERS
from app.utils.formatting import (
    find_column_letter_by_header,
    create_header_style,
    create_data_row_style,
    auto_adjust_column_width,
)

logger = logging.getLogger(__name__)


def get_or_create_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
    """
    Obtiene una hoja existente o la crea si no existe.
    
    Args:
        workbook: Libro de Excel
        sheet_name: Nombre de la hoja
    
    Returns:
        Worksheet existente o recién creada
    """
    if sheet_name in workbook.sheetnames:
        logger.debug("Hoja '%s' ya existe, retornando existente", sheet_name)
        return workbook[sheet_name]
    
    logger.info("Creando hoja '%s'", sheet_name)
    return workbook.create_sheet(title=sheet_name)


def apply_cruce_headers(
    sheet: Worksheet,
    headers: dict[str, str] | None = None,
) -> dict[str, Any]:
    """
    Aplica los headers a la hoja CruceFacturas.
    
    Args:
        sheet: Hoja CruceFacturas
        headers: Dict de celda -> valor. Si es None usa CRUCE_HEADERS
    
    Returns:
        Dict con información de los headers aplicados
    """
    if headers is None:
        headers = CRUCE_HEADERS
    
    header_style = create_header_style()
    
    for cell, value in headers.items():
        cell_obj = sheet[cell]
        cell_obj.value = value
        cell_obj.font = header_style["font"]
        cell_obj.fill = header_style["fill"]
        cell_obj.border = header_style["border"]
        cell_obj.alignment = header_style["alignment"]
        logger.debug("Header aplicado: %s = '%s'", cell, value)
    
    logger.info(
        "Headers aplicados a hoja '%s': %s",
        sheet.title,
        list(headers.keys()),
    )
    
    return {
        "sheet": sheet.title,
        "headers": headers,
    }


def create_cruce_facturas_sheet(workbook: Workbook) -> tuple[Worksheet, dict[str, Any]]:
    """
    Crea y configura la hoja CruceFacturas completa.
    
    Esta función:
    1. Crea la hoja si no existe
    2. Aplica los headers predefinidos con estilo
    3. Aplica estilo a filas de datos
    4. Ajusta el ancho de columnas automáticamente
    
    Args:
        workbook: Libro de Excel
    
    Returns:
        Tupla (worksheet, info_dict)
    """
    sheet = get_or_create_sheet(workbook, CRUCE_FACTURAS_SHEET)
    
    # Insertar fila vacía arriba
    sheet.insert_rows(1)
    
    # Aplicar headers en la fila 2
    header_style = create_header_style()
    
    for cell, value in CRUCE_HEADERS.items():
        cell_obj = sheet[cell]
        cell_obj.value = value
        cell_obj.font = header_style["font"]
        cell_obj.fill = header_style["fill"]
        cell_obj.border = header_style["border"]
        cell_obj.alignment = header_style["alignment"]
        logger.debug("Header aplicado: %s = '%s'", cell, value)
    
    # Aplicar estilo a filas de datos (comienzan en fila 3)
    data_style = create_data_row_style()
    for row in range(3, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = data_style["fill"]
            cell.border = data_style["border"]
            cell.alignment = data_style["alignment"]
    
    # Ajustar ancho de columnas
    column_widths = auto_adjust_column_width(sheet)
    
    return sheet, {
        "rule": "cruce_facturas_headers",
        "sheet": CRUCE_FACTURAS_SHEET,
        "cells": CRUCE_HEADERS,
        "column_widths": column_widths,
    }
