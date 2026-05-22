"""Servicio para manejo de hoja Revisión.

Este módulo detecta problemas en las facturas y los lista
en la hoja "Revision" para revisión manual.

Problemas detectados:
- Decimales: Facturas con valores decimales en Vlr. Subsidiado o Vlr. Procedimiento
- Doble tipo procedimiento: Facturas con más de un tipo de procedimiento
- Ruta duplicada: Pacientes con >= 3 facturas en Promoción y Prevención
- Convenio de procedimiento: Procedimientos que no corresponden al convenio
- Cantidades: Facturas con cantidades anómalas
"""

from __future__ import annotations

import logging
from collections import defaultdict
from datetime import datetime
from typing import Any

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    CONVENIO_ASISTENCIAL,
    CONVENIO_PYP,
    REVISION_SHEET,
    REVISION_HEADERS,
    URGENCIA_REVISION_HEADERS,
    TARGET_PROCEDURES,
    PYP_CUPS_CODES,
    RUTA_DUPLICADA_THRESHOLD,
    CANTIDAD_CONSULTAS_MIN,
    CANTIDAD_MAX,
    CANTIDAD_PYP_MIN,
    AREA_ODONTOLOGIA,
    AREA_URGENCIAS,
    AREA_EQUIPOS_BASICOS,
    CODIGO_TIPO_PROCEDIMIENTO_DIAGNOSTICO,
    LABORATORIO_NO,
    CENTRO_COSTO_APOYO_DIAGNOSTICO,
    CODIGOS_EXCEPTUADOS,
    URGENCIA_DATA_ROW_BACKGROUND_COLOR,
    CODIGOS_PYP_URGENCIAS,
    CENTRO_COSTO_PYP_URGENCIAS,
    CODIGOS_QUIROFANO_URGENCIAS,
    CENTRO_COSTO_QUIROFANO_URGENCIAS,
    CODIGOS_LABORATORIO_URGENCIAS,
    CODIGOS_LABORATORIO_URGENCIAS_REVERSE,
    CENTRO_COSTO_LABORATORIO_URGENCIAS,
    CENTRO_COSTO_ODONTOLOGIA,
    CENTRO_COSTO_EXTRAMURAL,
    CENTRO_COSTO_EQUIPOS_BASICOS,
    PROFESIONALES_ODONTOLOGIA,
    PROFESIONALES_ODONTOLOGIA_VALIDACION,
    PROFESIONALES_EQUIPOS_BASICOS,
    PROFESIONALES_URGENCIAS,
    CODIGOS_TRABAJADORA_SOCIAL,
    CODIGOS_PSICOLOGA,
    CODIGOS_NUTRICIONISTA,
    CODIGOS_FISIOTERAPEUTA,
    CODIGOS_JEFE_ENFERMERIA,
    CODIGOS_EXCLUIDOS_MEDICO,
    EXCEPCIONES_BACTERIOLOGA,
    LABORATORIO_NO,
    # IDE Contrato Urgencias
    CODIGO_IDE_CONTRATO_URGENCIAS,
    ENTIDAD_IDE_CONTRATO_URGENCIAS,
    IDE_CONTRATO_REQUERIDO_URGENCIAS,
    CODIGO_IDE_CONTRATO_861801_EPSI05,
    ENTIDAD_IDE_CONTRATO_861801_EPSI05,
    IDE_CONTRATO_REQUERIDO_861801_EPSI05,
    CODIGO_IDE_CONTRATO_890405_EPSI05,
    ENTIDAD_IDE_CONTRATO_890405_EPSI05,
    IDE_CONTRATO_CON_INSERCION_890405_EPSI05,
    IDE_CONTRATO_SIN_INSERCION_890405_EPSI05,
    CODIGO_INSERCION_BUSCAR,
    # Nueva regla EPSIC5
    CODIGO_IDE_CONTRATO_EPSIC5,
    ENTIDAD_IDE_CONTRATO_EPSIC5,
    IDE_CONTRATO_REQUERIDO_EPSIC5,
    CODIGO_IDE_CONTRATO_890405_EPSIC5,
    ENTIDAD_IDE_CONTRATO_890405_EPSIC5,
    IDE_CONTRATO_CON_INSERCION_890405_EPSIC5,
    IDE_CONTRATO_SIN_INSERCION_890405_EPSIC5,
    # Nueva regla ESS118 + Código 735301
    CODIGO_IDE_CONTRATO_735301_ESS118,
    ENTIDAD_IDE_CONTRATO_ESS118_NUEVOS,
    IDE_CONTRATO_MULTIPLE_ESS118_NUEVOS,
    # Nueva regla ESS118 + Código 906340 -> IDE Contrato debe ser 839
    CODIGO_IDE_CONTRATO_906340_ESS118,
    ENTIDAD_IDE_CONTRATO_906340_ESS118,
    IDE_CONTRATO_REQUERIDO_906340_ESS118,
    # ESS118 + 735301/861801 -> 970 o 974
    CODIGO_IDE_CONTRATO_735301_ESS118,
    CODIGO_IDE_CONTRATO_861801_ESS118,
    ENTIDAD_IDE_CONTRATO_ESS118_NUEVOS,
    IDE_CONTRATO_MULTIPLE_ESS118_NUEVOS,
    # ESS118 + 890405 -> 974
    CODIGO_IDE_CONTRATO_890405_ESS118,
    ENTIDAD_IDE_CONTRATO_890405_ESS118,
    IDE_CONTRATO_REQUERIDO_890405_ESS118,
    # ESS118 + 890205 -> 970
    CODIGO_IDE_CONTRATO_890205_ESS118,
    ENTIDAD_IDE_CONTRATO_890205_ESS118,
    IDE_CONTRATO_REQUERIDO_890205_ESS118,
    # Nueva regla ESSC18 + Código 906340 -> IDE Contrato debe ser 842
    CODIGO_IDE_CONTRATO_906340_ESSC18,
    ENTIDAD_IDE_CONTRATO_ESSC18,
    IDE_CONTRATO_REQUERIDO_906340_ESSC18,
    # Nueva regla ESSC18 + Código 861801 -> IDE Contrato debe ser 975
    CODIGO_IDE_CONTRATO_861801_ESSC18,
    IDE_CONTRATO_REQUERIDO_861801_ESSC18,
    # Nueva regla ESSC18 + Código 890405 -> IDE Contrato según inserción
    CODIGO_IDE_CONTRATO_890405_ESSC18,
    IDE_CONTRATO_CON_INSERCION_890405_ESSC18,
    IDE_CONTRATO_SIN_INSERCION_890405_ESSC18,
    # Nueva regla EPS037 + Código 906340 -> IDE Contrato debe ser 962
    CODIGO_IDE_CONTRATO_906340_EPS037,
    ENTIDAD_IDE_CONTRATO_EPS037,
    IDE_CONTRATO_REQUERIDO_906340_EPS037,
    # Nueva regla EPS037 + Código 861801 -> IDE Contrato debe ser 961
    CODIGO_IDE_CONTRATO_861801_EPS037,
    IDE_CONTRATO_REQUERIDO_861801_EPS037,
    # Nueva regla EPS037 + Código 890405 -> IDE Contrato según inserción
    CODIGO_IDE_CONTRATO_890405_EPS037,
    IDE_CONTRATO_CON_INSERCION_890405_EPS037,
    IDE_CONTRATO_SIN_INSERCION_890405_EPS037,
    # Nueva regla EPSS41 + Código 906340 -> IDE 959
    CODIGO_IDE_CONTRATO_906340_EPSS41,
    IDE_CONTRATO_REQUERIDO_906340_EPSS41,
    # Nueva regla EPSS41 + Código 861801 -> IDE 958
    CODIGO_IDE_CONTRATO_861801_EPSS41,
    IDE_CONTRATO_REQUERIDO_861801_EPSS41,
    # Nueva regla EPSS41 + Código 890405 -> IDE según inserción
    CODIGO_IDE_CONTRATO_890405_EPSS41,
    IDE_CONTRATO_CON_INSERCION_890405_EPSS41,
    IDE_CONTRATO_SIN_INSERCION_890405_EPSS41,
    # Nueva regla ESS062 + Código 861801 -> IDE Contrato debe ser 922
    CODIGO_IDE_CONTRATO_861801_ESS062,
    ENTIDAD_IDE_CONTRATO_ESS062,
    IDE_CONTRATO_REQUERIDO_861801_ESS062,
    # Nueva regla ESS062 + Código 890405 -> IDE Contrato según inserción
    CODIGO_IDE_CONTRATO_890405_ESS062,
    IDE_CONTRATO_CON_INSERCION_890405_ESS062,
    IDE_CONTRATO_SIN_INSERCION_890405_ESS062,
    # Nueva regla ESSC62 + Código 861801 -> IDE Contrato debe ser 863
    CODIGO_IDE_CONTRATO_861801_ESSC62,
    ENTIDAD_IDE_CONTRATO_ESSC62,
    IDE_CONTRATO_REQUERIDO_861801_ESSC62,
    # Nueva regla ESSC62 + Código 890405 -> IDE Contrato según si tiene 890405
    CODIGO_IDE_CONTRATO_890405_ESSC62,
    CODIGO_A_BUSCAR_890405_ESSC62,
    IDE_CONTRATO_CON_INSERCION_890405_ESSC62,
    IDE_CONTRATO_SIN_INSERCION_890405_ESSC62,
    # Urgencias - Entidad -> IDE Contrato
    URGENCIA_ENTIDAD_CONTRATO,
    URGENCIA_ENTIDAD_MULTIPLE_CONTRATO,
    # Urgencias 86000 + Código 861801 -> IDE Contrato debe ser 920
    CODIGO_IDE_CONTRATO_861801_86000,
    ENTIDAD_IDE_CONTRATO_861801_86000,
    IDE_CONTRATO_REQUERIDO_861801_86000,
    # Urgencias 86000 + Código 890405 -> IDE Contrato 919 (con 861801) o 920 (sin 861801)
    CODIGO_IDE_CONTRATO_890405_86000,
    ENTIDAD_IDE_CONTRATO_890405_86000,
    IDE_CONTRATO_CON_INSERCION_890405_86000,
    IDE_CONTRATO_SIN_INSERCION_890405_86000,
    # Urgencias RES004 + Código 861801 -> IDE Contrato debe ser 908
    # Urgencias RES004 + Código 861801 -> IDE Contrato debe ser 908
    CODIGO_IDE_CONTRATO_861801_RES004,
    ENTIDAD_IDE_CONTRATO_861801_RES004,
    IDE_CONTRATO_REQUERIDO_861801_RES004,
    # Urgencias RES004 + Código 890405 -> IDE Contrato 908 (con 861801) o 909 (sin 861801)
    CODIGO_IDE_CONTRATO_890405_RES004,
    ENTIDAD_IDE_CONTRATO_890405_RES004,
    IDE_CONTRATO_CON_INSERCION_890405_RES004,
    IDE_CONTRATO_SIN_INSERCION_890405_RES004,
    CODIGO_INSERCION_BUSCAR_RES004,
    # ESS118 + Procedimientos PyP -> IDE Contrato 970 o 974
    ENTIDAD_IDE_CONTRATO_ESS118_PYP,
    IDE_CONTRATO_MULTIPLE_ESS118_PYP,
    IDE_CONTRATO_MULTIPLE_ESS118_NO_PYP,
    # ESSC18 + Procedimientos PyP -> IDE Contrato 975
    ENTIDAD_IDE_CONTRATO_ESSC18_PYP,
    IDE_CONTRATO_MULTIPLE_ESSC18_PYP,
    IDE_CONTRATO_MULTIPLE_ESSC18_NO_PYP,
    # EPSS41 + Procedimientos PyP -> IDE Contrato 955 o 958
    ENTIDAD_IDE_CONTRATO_EPSS41_PYP,
    IDE_CONTRATO_MULTIPLE_EPSS41_PYP,
    IDE_CONTRATO_MULTIPLE_EPSS41_NO_PYP,
    # EPS037 + Procedimientos PyP -> IDE Contrato 961
    ENTIDAD_IDE_CONTRATO_EPS037_PYP,
    IDE_CONTRATO_MULTIPLE_EPS037_PYP,
    IDE_CONTRATO_MULTIPLE_EPS037_NO_PYP,
    # EPSI05 + Procedimientos PyP -> IDE Contrato 977
    ENTIDAD_IDE_CONTRATO_EPSI05_PYP,
    IDE_CONTRATO_MULTIPLE_EPSI05_PYP,
    IDE_CONTRATO_MULTIPLE_EPSI05_NO_PYP,
    # EPSIC5 + Procedimientos PyP -> IDE Contrato 979
    ENTIDAD_IDE_CONTRATO_EPSIC5_PYP,
    IDE_CONTRATO_MULTIPLE_EPSIC5_PYP,
    IDE_CONTRATO_MULTIPLE_EPSIC5_NO_PYP,
    # RES001 + Procedimientos PyP -> IDE Contrato 954
    ENTIDAD_IDE_CONTRATO_RES001_PYP,
    IDE_CONTRATO_MULTIPLE_RES001_PYP,
    IDE_CONTRATO_MULTIPLE_RES001_NO_PYP,
    # ESS062 + Procedimientos PyP -> IDE Contrato 922
    ENTIDAD_IDE_CONTRATO_ESS062_PYP,
    IDE_CONTRATO_MULTIPLE_ESS062_PYP,
    IDE_CONTRATO_MULTIPLE_ESS062_NO_PYP,
    # ESSC62 + Procedimientos PyP -> IDE Contrato 863
    ENTIDAD_IDE_CONTRATO_ESSC62_PYP,
    IDE_CONTRATO_MULTIPLE_ESSC62_PYP,
    IDE_CONTRATO_MULTIPLE_ESSC62_NO_PYP,
    # 0001 + Procedimientos PyP -> IDE Contrato 17
    ENTIDAD_IDE_CONTRATO_0001_PYP,
    IDE_CONTRATO_MULTIPLE_0001_PYP,
    IDE_CONTRATO_MULTIPLE_0001_NO_PYP,
    # EPSS005 + Procedimientos PyP -> IDE Contrato 933
    ENTIDAD_IDE_CONTRATO_EPSS005_PYP,
    IDE_CONTRATO_MULTIPLE_EPSS005_PYP,
    IDE_CONTRATO_MULTIPLE_EPSS005_NO_PYP,
    # EPSC005 + Procedimientos PyP -> IDE Contrato 932
    ENTIDAD_IDE_CONTRATO_EPSC005_PYP,
    IDE_CONTRATO_MULTIPLE_EPSC005_PYP,
    IDE_CONTRATO_MULTIPLE_EPSC005_NO_PYP,
    # 86 + Procedimientos NO PyP -> IDE Contrato 911
    ENTIDAD_IDE_CONTRATO_86_NO_PYP,
    IDE_CONTRATO_MULTIPLE_86_NO_PYP,
    # 86000 + Procedimientos PyP -> IDE Contrato 920
    ENTIDAD_IDE_CONTRATO_86000_PYP,
    IDE_CONTRATO_MULTIPLE_86000_PYP,
    IDE_CONTRATO_MULTIPLE_86000_NO_PYP,
    # Urgencias Capita - Listado de códigos CUPS válidos
    URGENCIAS_CAPITA_CUPS_CODES,
    # Equipos Básicos - Reglas independientes (comparte PYP_CUPS_CODES con Odontología)
    EQUIPOS_BASICOS_RUTA_DUPLICADA_THRESHOLD,
    EQUIPOS_BASICOS_CANTIDAD_CONSULTAS_MIN,
    EQUIPOS_BASICOS_CANTIDAD_MAX,
    EQUIPOS_BASICOS_CANTIDAD_PYP_MIN,
    PYP_CODES_ONLY_ODONTOLOGO,
    PYP_CODES_HIGIENISTA,
    # MAL CAPITADO - Códigos que requieren prefijo FEV en Número Factura
    CODIGOS_MAL_CAPITADO,
    PREFIJO_FACTURA_MAL_CAPITADO,
    PREFIJO_FACTURA_CAP,
    ENTIDAD_REQUERIDA_CAP,
    # Urgencias - Cups Equivalentes (890205 -> 890405)
    CODIGO_CUPS_EQUIVALENTE_890205,
    CODIGO_CUPS_EQUIVALENTE_SUSTITUTO_890405,
    ENTIDADES_PERMITIDAS_890205,
    # Urgencias - Cantidades max 1 para ciertos códigos
    URGENCIAS_CODIGOS_CANTIDAD_MAX_1,
    URGENCIAS_SOAT_CODIGOS_CANTIDAD_MAX_1,
    URGENCIAS_NO_SOAT_CODIGOS_CANTIDAD_MAX_1,
    # Hospitalización - Reglas de cantidades
    CODIGO_HOSPITALIZACION_ESTANCIA,
    CODIGO_HOSPITALIZACION_CAMAS,
    HORAS_POR_DIA,
    # SOAT - Reglas de urgencias y hospitalización
    VALOR_TARIFARIO_SOAT,
    CODIGO_SOAT_SALA_OBSERVACION_LARGA,
    CODIGO_SOAT_SALA_OBSERVACION_CORTA,
    CODIGOS_SOAT_OBLIGATORIOS_SALA,
    CODIGO_SOAT_URGENCIAS_PROHIBIDO,
    CODIGOS_SOAT_HOSPITALIZACION_PROHIBIDOS,
    CODIGOS_SOAT_CANTIDAD_OBLIGATORIA,
    CODIGOS_SOAT_HOSPITALIZACION_CANTIDAD,
    CODIGOS_SOAT_HOSPITALIZACION_OBLIGATORIOS,
    # ⚠️ Revisión Necesaria - Códigos exentos de cantidad
    CODIGOS_REVISION_CANTIDAD_EXENTOS,
    LABORATORIO_REVISION_EXENTO,
    CODIGO_TIPO_PROCEDIMIENTO_REVISION_LAB,
    CODIGOS_TIPO_PROC_09_12,
    CODIGO_EXENTO_V03AN0101,
    CANTIDAD_MAX_09_12,
    CODIGO_ESPECIAL_02_LAB,
    CANTIDAD_MAX_02_LAB,
    CANTIDAD_MAX_02_LAB_903883,
    CODIGOS_LIMITE_ESPECIFICO,
)

from app.utils.formatting import (
    create_header_style,
    create_data_row_style,
    create_urgencia_header_style,
    create_urgencia_data_row_style,
    auto_adjust_column_width,
)

# Importar reglas transversales
from app.services.transversales import (
    detect_decimales,
    detect_tipo_documento_edad,
    detect_codigo_entidad_vs_entidad_afiliacion,
    detect_tipo_usuario,
    # Nuevos módulos extraídos (Fase 2)
    get_column_indices,
    detect_doble_tipo_procedimiento,
    detect_ruta_duplicada,
    detect_cantidades_anomalas,
)

# Importar módulos de odontología extraídos (Fase 4)
from app.services.odontologia.profesionales import (
    detect_profesionales_odontologia,
)
from app.services.odontologia.centro_costo import (
    detect_centro_costo_odontologia as detect_centro_costo_odontologia_ext,
)
from app.services.odontologia.ide_contrato import (
    detect_ide_contrato_odontologia,
)
from app.services.odontologia.mal_capitado import (
    detect_mal_capitado,
)
from app.services.odontologia.detect_all import (
    detect_all_problems_odontologia,
)

# Importar módulos de urgencias extraídos (Fase 5a)
from app.services.urgencias.cantidades_urgencias import (
    detect_cantidades_urgencias,
)
from app.services.urgencias.cantidades_soat_urgencias import (
    detect_cantidades_soat_urgencias,
)
from app.services.urgencias.cantidades_soat_hospitalizacion import (
    detect_cantidades_soat_hospitalizacion,
)
from app.services.urgencias.hospitalizacion import (
    detect_cantidades_hospitalizacion as detect_cantidades_hospitalizacion_ext,
)

logger = logging.getLogger(__name__)


def _normalize_header(value: Any) -> str:
    """Normaliza un header a minúsculas sin espacios extra."""
    return str(value).strip().lower() if value is not None else ""


def _normalize_invoice(value: Any) -> str | None:
    """Normaliza un número de factura a string."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and value == int(value):
        return str(int(value))
    return str(value).strip() or None


def _get_column_indices(headers: list[Any]) -> tuple[dict[str, int | None], list[str]]:
    """
    Mapea nombres de columna a sus índices.

    REQUIERE COINCIDENCIA EXACTA - NO infiere nombres similares.
    Si una columna no coincide exactamente, retorna None y la reporta en la lista de errores.

    Args:
        headers: Lista de nombres de columna del Excel

    Returns:
        Tuple de (dict con nombre de columna -> índice 0-based o None,
                  lista de columnas NO encontradas)

    Note:
        Esta función ahora delega a transversales/column_indices.py.
        Se eliminará en Fase 7 cuando todos los consumidores usen get_column_indices directamente.
    """
    required_headers: dict[str, str] = {
        "numero_factura": "Número Factura",
        "vlr_subsidiado": "Vlr. Subsidiado",
        "vlr_procedimiento": "Vlr. Procedimiento",
        "codigo_tipo_procedimiento": "Código Tipo Procedimiento",
        "tipo_procedimiento": "Tipo Procedimiento",
        "codigo": "Código",
        "codigo_equiv": "Cód. Equivalente CUPS",
        "procedimiento": "Procedimiento",
        "identificacion": "Nº Identificación",
        "convenio_facturado": "Convenio Facturado",
        "cantidad": "Cantidad",
        "laboratorio": "Laboratorio",
        "centro_costo": "Centro Costo",
        "codigo_entidad_cobrar": "Cód Entidad Cobrar",
        "entidad_cobrar": "Entidad Cobrar",
        "entidad_afiliacion": "Entidad Afiliación",
        "tipo_factura_descripcion": "Tipo Factura Descripción",
        "ide_contrato": "IDE Contrato",
        "tipo_identificacion": "Tipo Identificación",
        "fec_nacimiento": "Fec. Nacimiento",
        "fec_factura": "Fec. Factura",
        "fecha_cierre": "Fecha Cierre",
        "profesional_identificacion": "Identificación Profesional",
        "profesional_atiende": "Profesional Atiende",
        "codigo_profesional": "Código Profesional",
        "responsable_cierra": "Responsable Cierra Facturar",
        "tarifario": "Tarifario",
        "tipo_usuario": "Tipo Usuario",
    }
    return get_column_indices(headers, required_headers)


def _detect_decimals(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict]:
    """Detecta facturas con valores decimales.

    Note: Delega a transversales/decimales.py.
    Se eliminará en Fase 7.
    """
    return [
        {"factura": factura, "valores": ""}
        for factura in detect_decimales(data_sheet, indices)
    ]


def _detect_doble_tipo_procedimiento(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict]:
    """Detecta facturas con más de un tipo de procedimiento.

    Note: Delega a transversales/doble_tipo_procedimiento.py.
    Se eliminará en Fase 7.
    """
    return detect_doble_tipo_procedimiento(data_sheet, indices)


def _detect_ruta_duplicada(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict]:
    """Detecta pacientes con múltiples facturas en PyP.

    Note: Delega a transversales/ruta_duplicada.py.
    Se eliminará en Fase 7.
    """
    from app.constants import RUTA_DUPLICADA_THRESHOLD
    return detect_ruta_duplicada(data_sheet, indices, threshold=RUTA_DUPLICADA_THRESHOLD)


def _detect_ruta_duplicada_equipos_basicos(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict]:
    """Detecta pacientes con múltiples facturas en PyP (Equipos Básicos).

    Note: Delega a transversales/ruta_duplicada.py.
    Se eliminará en Fase 7.
    """
    from app.constants import EQUIPOS_BASICOS_RUTA_DUPLICADA_THRESHOLD
    return detect_ruta_duplicada(data_sheet, indices, threshold=EQUIPOS_BASICOS_RUTA_DUPLICADA_THRESHOLD)


def _detect_tipo_identificacion_edad(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """Detecta facturas donde el tipo de identificación no coincide con la edad.

    Note: Delega a transversales/tipo_documento_edad.py.
    Se eliminará en Fase 7.
    """
    problemas = detect_tipo_documento_edad(data_sheet, indices)
    return [
        {
            "factura": p["factura"],
            "tipo_actual": p["tipo_actual"],
            "tipo_deberia": p["tipo_deberia"],
            "edad": str(p["edad_anios"]),
        }
        for p in problemas
    ]


def _detect_cantidades_anomalas(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict]:
    """Detecta facturas con cantidades anómalas.

    Note: Delega a transversales/cantidades_anomalas.py.
    Se eliminará en Fase 7.
    """
    from app.constants import (
        CANTIDAD_CONSULTAS_MIN,
        CANTIDAD_MAX,
        CANTIDAD_PYP_MIN,
    )
    return detect_cantidades_anomalas(
        data_sheet,
        indices,
        cantidad_consultas_min=CANTIDAD_CONSULTAS_MIN,
        cantidad_max_general=CANTIDAD_MAX,
        cantidad_pyp_min=CANTIDAD_PYP_MIN,
    )


def _detect_cantidades_anomalas_equipos_basicos(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict]:
    """Detecta facturas con cantidades anómalas (Equipos Básicos).

    Note: Delega a transversales/cantidades_anomalas.py.
    Se eliminará en Fase 7.
    """
    from app.constants import (
        EQUIPOS_BASICOS_CANTIDAD_CONSULTAS_MIN,
        EQUIPOS_BASICOS_CANTIDAD_MAX,
        EQUIPOS_BASICOS_CANTIDAD_PYP_MIN,
    )
    # Equipos Básicos requiere columna procedimiento (más estricto)
    if indices.get("procedimiento") is None:
        return []
    return detect_cantidades_anomalas(
        data_sheet,
        indices,
        cantidad_consultas_min=EQUIPOS_BASICOS_CANTIDAD_CONSULTAS_MIN,
        cantidad_max_general=EQUIPOS_BASICOS_CANTIDAD_MAX,
        cantidad_pyp_min=EQUIPOS_BASICOS_CANTIDAD_PYP_MIN,
    )


def _detect_profesionales_odontologia(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con profesionales no válidos en Odontología.

    NOTA: Ahora delega en app/services/odontologia/profesionales.py.
    Se eliminará en Fase 7.
    """
    return detect_profesionales_odontologia(data_sheet, indices)


def _detect_mal_capitado(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con códigos G03XB01 o A02BB01 que NO tienen prefijo FEV en Número Factura.

    NOTA: Ahora delega en app/services/odontologia/mal_capitado.py.
    Se eliminará en Fase 7.
    """
    return detect_mal_capitado(data_sheet, indices)


def _detect_cantidades_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta facturas con cantidades anómalas en Urgencias.

    Note: Delega a urgencias/cantidades_urgencias.py.
    Se eliminará en Fase 7.
    """
    return detect_cantidades_urgencias(data_sheet, indices)


def _detect_cantidades_soat_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta facturas SOAT Urgencias con cantidades incorrectas.

    Note: Delega a urgencias/cantidades_soat_urgencias.py.
    Se eliminará en Fase 7.
    """
    return detect_cantidades_soat_urgencias(data_sheet, indices)


def _detect_cantidades_soat_hospitalizacion(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta facturas SOAT Hospitalización con cantidades incorrectas.

    Note: Delega a urgencias/cantidades_soat_hospitalizacion.py.
    Se eliminará en Fase 7.
    """
    return detect_cantidades_soat_hospitalizacion(data_sheet, indices)


def _detect_cantidades_hospitalizacion(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta facturas con cantidades incorrectas en Hospitalización.

    Note: Delega a urgencias/hospitalizacion.py.
    Se eliminará en Fase 7.
    """
    return detect_cantidades_hospitalizacion_ext(data_sheet, indices)


def _detect_profesionales_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con profesionales no válidos en Urgencias.

    Reglas (Urgencias):
    - "Código Profesional" DEBE estar en PROFESIONALES_URGENCIAS
    - TRABAJADORA SOCIAL: solo puede usar código 890409

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo_profesional", "nombre", "tipo", "profesional_area", "procedimiento", "regla", "problema"
    """
    logger.warning("=== _detect_profesionales_urgencias ===")
    logger.warning("Indices encontrados: %s", indices)
    
    num_fact_idx = indices.get("numero_factura")
    cod_prof_idx = indices.get("codigo_profesional")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")
    
    logger.warning("numero_factura idx: %s, codigo_profesional idx: %s, codigo idx: %s, procedimiento idx: %s",
              num_fact_idx, cod_prof_idx, codigo_idx, procedimiento_idx)
    
    if num_fact_idx is None or cod_prof_idx is None:
        logger.warning("NO se encontró numero_factura o codigo_profesional en los índices")
        return []

    problemas = []
    facturas_procesadas: set[str] = set()
    
    # Log de las primeras 5 filas para debug
    logger.warning("=== MUESTREO 5 PRIMERAS FILAS PROFESIONALES ===")
    for row in range(2, min(7, data_sheet.max_row + 1)):
        num_fact = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        cod_prof = data_sheet.cell(row=row, column=cod_prof_idx + 1).value
        codigo_val = ""
        proc_val = ""
        if codigo_idx is not None:
            codigo_val = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_val = str(codigo_val).strip() if codigo_val else ""
        if procedimiento_idx is not None:
            proc_val = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
            proc_val = str(proc_val).strip()[:30] if proc_val else ""
        
        logger.warning("Fila %d: factura=%s, cod_prof=%s, codigo=%s, proc=%s",
                    row, num_fact, cod_prof, codigo_val, proc_val)

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = _normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_procesadas:
            continue

        cod_profesional = data_sheet.cell(row=row, column=cod_prof_idx + 1).value
        cod_profesional_str = str(cod_profesional).strip() if cod_profesional else ""

        if not cod_profesional_str:
            continue

        # Buscar profesional en el diccionario de Urgencias
        profesional_info = PROFESIONALES_URGENCIAS.get(cod_profesional_str)

        if profesional_info is None:
            logger.warning("Profesional no encontrado en lista: %s", cod_profesional_str)
            # Obtener procedimiento para mostrar en el error
            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""
            
            problemas.append({
                "factura": factura_str,
                "codigo_profesional": cod_profesional_str,
                "nombre": "",
                "tipo": "",
                "profesional_area": "",
                "procedimiento": procedimiento,
                "regla": "Profesional debe estar en listado",
                "problema": "Profesional no existe en el listado de Urgencias",
            })
            facturas_procesadas.add(factura_str)
            continue

        # Validación por tipo de profesional
        tipo_profesional = profesional_info.get("tipo", "")
        
        # Si es TRABAJADORA SOCIAL, validar código 890409
        if tipo_profesional == "TRABAJADORA SOCIAL" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""
            
            # Obtener procedimiento
            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""
            
            if codigo_str and codigo_str not in CODIGOS_TRABAJADORA_SOCIAL:
                codigos_validos = ", ".join(sorted(CODIGOS_TRABAJADORA_SOCIAL))
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "TRABAJADORA SOCIAL",
                    "profesional_area": "TRABAJADORA SOCIAL",
                    "procedimiento": procedimiento,
                    "regla": f"Código debe ser uno de: {codigos_validos}",
                    "problema": f"TRABAJADORA SOCIAL con código no permitido ({codigo_str}). Debería usar {codigos_validos}",
                })
                facturas_procesadas.add(factura_str)
        
        # Si es PSICOLOGA, validar códigos 890408 o 35102
        if tipo_profesional == "PSICOLOGA" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""

            # Obtener procedimiento
            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""

            if codigo_str and codigo_str not in CODIGOS_PSICOLOGA:
                codigos_validos = ", ".join(sorted(CODIGOS_PSICOLOGA))
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "PSICOLOGA",
                    "profesional_area": "PSICOLOGA",
                    "procedimiento": procedimiento,
                    "regla": f"Código debe ser uno de: {codigos_validos}",
                    "problema": f"PSICOLOGA con código no permitido ({codigo_str}). Debería usar {codigos_validos}",
                })
                facturas_procesadas.add(factura_str)
        
        # Si es NUTRICIONISTA, validar códigos 890406 o 37602
        if tipo_profesional == "NUTRICIONISTA" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""

            # Obtener procedimiento
            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""

            if codigo_str and codigo_str not in CODIGOS_NUTRICIONISTA:
                codigos_validos = ", ".join(sorted(CODIGOS_NUTRICIONISTA))
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "NUTRICIONISTA",
                    "profesional_area": "NUTRICIONISTA",
                    "procedimiento": procedimiento,
                    "regla": f"Código debe ser uno de: {codigos_validos}",
                    "problema": f"NUTRICIONISTA con código no permitido ({codigo_str}). Debería usar {codigos_validos}",
                })
                facturas_procesadas.add(factura_str)
        
        # Si es FISIOTERAPEUTA, validar código 890412, 890411 o 29117
        if tipo_profesional == "FISIOTERAPEUTA" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""
            
            # Obtener procedimiento
            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""
            
            codigos_validos = ", ".join(sorted(CODIGOS_FISIOTERAPEUTA))
            if codigo_str and codigo_str not in CODIGOS_FISIOTERAPEUTA:
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "FISIOTERAPEUTA",
                    "profesional_area": "FISIOTERAPEUTA",
                    "procedimiento": procedimiento,
                    "regla": f"Código debe ser {codigos_validos}",
                    "problema": f"FISIOTERAPEUTA con código no permitido ({codigo_str}). Debería usar {codigos_validos}",
                })
                facturas_procesadas.add(factura_str)
        
        # Si es JEFE ENFERMERIA, validar códigos 861801, 890205, 890405, 990211, 29116, 39360
        if tipo_profesional == "JEFE ENFERMERIA" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""
            
            # Obtener procedimiento
            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""
            
            codigos_validos = ", ".join(sorted(CODIGOS_JEFE_ENFERMERIA))
            if codigo_str and codigo_str not in CODIGOS_JEFE_ENFERMERIA:
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "JEFE ENFERMERIA",
                    "profesional_area": "JEFE ENFERMERIA",
                    "procedimiento": procedimiento,
                    "regla": f"Código debe ser {codigos_validos}",
                    "problema": f"JEFE ENFERMERIA con código no permitido ({codigo_str}). Debería usar {codigos_validos}",
                })
                facturas_procesadas.add(factura_str)
        
        # Si es BACTERIOLOGA, validar Código Tipo Procedimiento = 02 o 05 y Laboratorio = "Si"
        if tipo_profesional == "BACTERIOLOGA":
            # Obtener código de procedimiento para verificar excepciones
            codigo_proc = ""
            if codigo_idx is not None:
                codigo_proc = data_sheet.cell(row=row, column=codigo_idx + 1).value
                codigo_proc = str(codigo_proc).strip() if codigo_proc else ""
            
            # Si es excepción -> skip validación, no dar error
            if codigo_proc in EXCEPCIONES_BACTERIOLOGA:
                facturas_procesadas.add(factura_str)
                continue
            
            codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")
            laboratorio_idx = indices.get("laboratorio")
            
            # Obtener valores
            codigo_tipo = ""
            laboratorio = ""
            
            if codigo_tipo_proc_idx is not None:
                codigo_tipo = data_sheet.cell(row=row, column=codigo_tipo_proc_idx + 1).value
                codigo_tipo = str(codigo_tipo).strip() if codigo_tipo else ""
            
            if laboratorio_idx is not None:
                laboratorio = data_sheet.cell(row=row, column=laboratorio_idx + 1).value
                laboratorio = str(laboratorio).strip().upper() if laboratorio else ""
            
            # Validar: Código Tipo Procedimiento debe ser 02 o 05 Y Laboratorio debe ser "Si"
            es_tipo_valido = codigo_tipo in ("02", "05", CODIGO_TIPO_PROCEDIMIENTO_DIAGNOSTICO)
            es_laboratorio_si = laboratorio == "SI"
            
            # Obtener procedimiento
            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""
            
            if not (es_tipo_valido and es_laboratorio_si):
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "BACTERIOLOGA",
                    "profesional_area": "BACTERIOLOGA",
                    "procedimiento": procedimiento,
                    "regla": "Código Tipo=02/05 + Laboratorio=Si",
                    "problema": "LABORATORIO NO IDENTIFICADO: BACTERIOLOGA requiere Código Tipo Procedimiento=02/05 y Laboratorio=Si",
                })
                facturas_procesadas.add(factura_str)
        
        # Si es MEDICO, NO puede usar códigos de otros profesionales ni regla de laboratorio
        if tipo_profesional == "MEDICO" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""
            
            # Verificar si usa código excluido
            if codigo_str and codigo_str in CODIGOS_EXCLUIDOS_MEDICO:
                # Obtener procedimiento
                procedimiento = ""
                if procedimiento_idx is not None:
                    proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                    procedimiento = str(proc).strip() if proc else ""
                
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "MEDICO",
                    "profesional_area": "MEDICO",
                    "procedimiento": procedimiento,
                    "regla": f"No usar: {', '.join(sorted(CODIGOS_EXCLUIDOS_MEDICO))}",
                    "problema": f"MEDICO con código no permitido ({codigo_str}). Código reservado para otro tipo de profesional",
                })
                facturas_procesadas.add(factura_str)
                continue
            
            # Verificar si cumple regla de laboratorio (que es de BACTERIOLOGA)
            codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")
            laboratorio_idx = indices.get("laboratorio")
            
            codigo_tipo = ""
            laboratorio = ""
            
            if codigo_tipo_proc_idx is not None:
                codigo_tipo = data_sheet.cell(row=row, column=codigo_tipo_proc_idx + 1).value
                codigo_tipo = str(codigo_tipo).strip() if codigo_tipo else ""
            
            if laboratorio_idx is not None:
                laboratorio = data_sheet.cell(row=row, column=laboratorio_idx + 1).value
                laboratorio = str(laboratorio).strip().upper() if laboratorio else ""
            
            # Obtener procedimiento
            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""
            
            # Si tiene código normal pero cumple regla de laboratorio = error
            es_tipo_lab = codigo_tipo in ("02", "05", CODIGO_TIPO_PROCEDIMIENTO_DIAGNOSTICO)
            es_lab_si = laboratorio == "SI"
            
            if codigo_str and es_tipo_lab and es_lab_si:
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "MEDICO",
                    "profesional_area": "MEDICO",
                    "procedimiento": procedimiento,
                    "regla": "No usar Tipo=02/05 + Lab=Si (reservado BACTERIOLOGA)",
                    "problema": "MEDICO no puede usar código de Laboratorio (Tipo 02/05 + Lab=Si). Reserved for BACTERIOLOGA",
                })
                facturas_procesadas.add(factura_str)

    if problemas:
        logger.warning("=== ERRORES PROFESIONALES URGENCIAS: %d ===", len(problemas))
        for p in problemas:
            logger.warning("- Factura: %s, Profesional: %s (%s), Área: %s, Código: %s, Problema: %s",
                p.get("factura"), p.get("codigo_profesional"), p.get("nombre"),
                p.get("profesional_area"), p.get("procedimiento"), p.get("problema"))

    return problemas


def _detect_profesionales_equipos_basicos(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con profesionales no válidos o procedimientos no permitidos.

    Reglas (Equipos Básicos):
    - "Código Profesional" DEBE estar en PROFESIONALES_EQUIPOS_BASICOS
    - HIGIENISTA: Solo puede usar códigos PYP (excepto 890203)
    - ODONTOLOGO: Puede usar cualquier código EXCEPTO los PYP (excepto 890203)

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo_profesional", "nombre", "tipo", "problema"
    """
    num_fact_idx = indices["numero_factura"]
    cod_prof_idx = indices["codigo_profesional"]
    codigo_idx = indices["codigo"]

    if None in (num_fact_idx, cod_prof_idx) or codigo_idx is None:
        return []

    problemas = []
    facturas_procesadas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = _normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_procesadas:
            continue

        cod_profesional = data_sheet.cell(row=row, column=cod_prof_idx + 1).value
        cod_profesional_str = str(cod_profesional).strip() if cod_profesional else ""

        if not cod_profesional_str:
            continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip() if codigo else ""

        # Buscar profesional en el diccionario
        profesional_info = PROFESIONALES_EQUIPOS_BASICOS.get(cod_profesional_str)

        if profesional_info is None:
            problemas.append({
                "factura": factura_str,
                "codigo_profesional": cod_profesional_str,
                "nombre": "",
                "tipo": "",
                "problema": "Profesional no existe en el listado de Equipos Básicos",
            })
            facturas_procesadas.add(factura_str)
        elif profesional_info.get("tipo") == "HIGIENISTA":
            # Higienista: solo puede usar códigos de PYP_CODES_HIGIENISTA
            if codigo_str and codigo_str not in PYP_CODES_HIGIENISTA:
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "HIGIENISTA",
                    "problema": f"Higienista con código no permitido ({codigo_str})",
                })
                facturas_procesadas.add(factura_str)
        elif profesional_info.get("tipo") == "ODONTOLOGO":
            # Odontólogo: no puede usar códigos de PYP_CODES_HIGIENISTA
            # (pero SÍ puede usar 890203)
            if codigo_str and codigo_str in PYP_CODES_HIGIENISTA:
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "ODONTOLOGO",
                    "problema": f"Odontólogo con código PYP no permitido ({codigo_str})",
                })
                facturas_procesadas.add(factura_str)

    return problemas


def _detect_ide_contrato_odontologia(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con problemas de IDE Contrato en Odontología.

    NOTA: Ahora delega en app/services/odontologia/ide_contrato.py.
    Se eliminará en Fase 7.
    """
    return detect_ide_contrato_odontologia(data_sheet, indices)


def _detect_centro_costo_odontologia(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
    profesional_dias: dict[str, list[int]] | None = None,
    permitir_todos_centros: bool = False,
    centros_validos: list[str] | None = None,
) -> list[dict[str, str]]:
    """
    Detecta facturas con problemas de centro de costo en Odontología.

    NOTA: Ahora delega en app/services/odontologia/centro_costo.py.
    Se eliminará en Fase 7.
    """
    return detect_centro_costo_odontologia_ext(
        data_sheet,
        indices,
        profesional_dias=profesional_dias,
        permitir_todos_centros=permitir_todos_centros,
        centros_validos=centros_validos,
    )


def _get_codigos_no_en_db_ess118(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Retorna lista de problemas de códigos CUPS que NO están en la DB.
    
    Regla: IDE Contrato = 969 Y Código Tipo Procedimiento no es 9,12,13 
           Y código NO está en tabla procedimiento → ERROR
    
    Nota: Se consulta la tabla procedimiento de PostgreSQL (no la DB externa).
    
    Returns:
        Lista de dicts con keys: "factura", "codigo", "procedimiento", "entidad"
    """
    # Cargar códigos válidos de la tabla procedimiento relacionados con nota_hoja = 3
    from app.database import SessionLocal
    from app.models import Procedimiento, NotasTecnicas

    try:
        db = SessionLocal()
        cups_validos = set(
            row.cups
            for row in db.query(Procedimiento.cups)
            .join(NotasTecnicas, NotasTecnicas.id_procedimiento == Procedimiento.id)
            .filter(NotasTecnicas.id_nota_hoja == 3)
            .distinct()
            .all()
        )
        db.close()
    except Exception as e:
        logger.warning("No se pudo conectar a DB para validar códigos: %s", e)
        return []
    
    if not cups_validos:
        logger.warning("No hay códigos en tabla procedimiento para nota_hoja=3")
        return []
    
    logger.info("Códigos válidos (nota_hoja=3): %d", len(cups_validos))
    
    codigo_idx = indices.get("codigo")
    ide_contrato_idx = indices.get("ide_contrato")
    codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")
    num_fact_idx = indices.get("numero_factura")
    proc_idx = indices.get("procedimiento")
    codigo_entidad_idx = indices.get("codigo_entidad_cobrar")
    
    if codigo_idx is None:
        return []
    
    problemas = []
    
    for row in range(2, data_sheet.max_row + 1):
        # Verificar IDE Contrato = 969
        ide_contrato = None
        if ide_contrato_idx is not None:
            ide_contrato = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value
        
        ide_str = str(ide_contrato).strip() if ide_contrato else ""
        
        # Solo procesar si IDE = 969
        if ide_str != "969":
            continue
        
        # Excluir Código Tipo Procedimiento = 09, 12, 13
        if codigo_tipo_proc_idx is not None:
            codigo_tipo = data_sheet.cell(row=row, column=codigo_tipo_proc_idx + 1).value
            if codigo_tipo and str(codigo_tipo).strip() in ["09", "12", "13"]:
                continue
        
        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        if not codigo:
            continue
        
        codigo_str = str(codigo).strip()
        
        # Verificar si existe en la tabla procedimiento
        if codigo_str not in cups_validos:
            # Agregar problema individual por cada fila
            factura = ""
            if num_fact_idx is not None:
                factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value or ""
            
            procedimiento = ""
            if proc_idx is not None:
                procedimiento = data_sheet.cell(row=row, column=proc_idx + 1).value or ""
            
            entidad = ""
            if codigo_entidad_idx is not None:
                entidad = data_sheet.cell(row=row, column=codigo_entidad_idx + 1).value or ""
            
            problemas.append({
                "factura": str(factura),
                "codigo": codigo_str,
                "procedimiento": str(procedimiento),
                "entidad": str(entidad),
            })
    
    return problemas


def _detect_centro_costo_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
    problemas_codigos_no_en_db: list[dict[str, str]] | None = None,
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    """
    Detecta facturas con problemas de centro de costo, IDE contrato y cups equivalentes.
    AHORA DELEGA a modulos extraidos en app/services/urgencias/.

    Reglas originales delegadas:
    - Centro de costo: detect_centro_costo_urgencias()
    - IDE Contrato: detect_ide_contrato_urgencias()
    - CUPS equivalentes: detect_cups_equivalentes()
    - Sala de observacion: detect_sala_observacion()
    - Hospitalizacion codigos: detect_hospitalizacion_codes()

    Args:
        data_sheet: Hoja de datos
        indices: Indices de columnas
        problemas_codigos_no_en_db: No usado directamente (los modulos lo manejan)

    Returns:
        Tuple de tres listas: (problemas_centros, problemas_ide_contrato, problemas_cups_equivalentes)
    """
    logger.info("Indices detectados para urgencias: %s", indices)

    num_fact_idx = indices.get("numero_factura")

    if num_fact_idx is None:
        return []

    # Delegar a modulos extraidos
    from app.services.urgencias import (
        detect_centro_costo_urgencias,
        detect_cups_equivalentes,
        detect_hospitalizacion_codes,
        detect_ide_contrato_urgencias,
        detect_sala_observacion,
    )

    problemas_centros = detect_centro_costo_urgencias(data_sheet, indices)
    problemas_ide_contrato = detect_ide_contrato_urgencias(data_sheet, indices)

    problemas_cups_equivalentes: list[dict[str, str]] = []
    problemas_cups_equivalentes.extend(detect_cups_equivalentes(data_sheet, indices))
    problemas_cups_equivalentes.extend(detect_sala_observacion(data_sheet, indices))
    problemas_cups_equivalentes.extend(detect_hospitalizacion_codes(data_sheet, indices))

    logger.info(
        "Centro Costo Urgencias - Total: centros=%d, ide_contrato=%d, cups_equiv=%d",
        len(problemas_centros), len(problemas_ide_contrato), len(problemas_cups_equivalentes),
    )

    return problemas_centros, problemas_ide_contrato, problemas_cups_equivalentes



def _detect_ide_contrato_reverse_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con problemas de IDE Contrato REVERSE (sin entidad).

    Dado un IDE Contrato, verifica que el Código CUPS corresponda al esperado.

    Reglas REVERSE (sin entidad):
    - IDE 986 → Código debe ser 906340
    - IDE 977 → Código puede ser 861801 (siempre) o 890405 (solo si la
      identificación NO tiene 861801 en otra factura; si la tiene,
      890405 debería ser IDE 976)

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo", "ide_contrato", "codigo_deberia"
    """
    from app.constants import IDE_CONTRATO_REVERSE

    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    ide_contrato_idx = indices.get("ide_contrato")
    ident_idx = indices.get("identificacion")

    if num_fact_idx is None or codigo_idx is None or ide_contrato_idx is None:
        logger.warning(
            "IDE Contrato REVERSE - Columnas necesarias no encontradas: "
            "numero_factura=%s, codigo=%s, ide_contrato=%s",
            num_fact_idx, codigo_idx, ide_contrato_idx,
        )
        return []

    # PASO 1: Recolectar identificaciones que tienen código 861801
    # (estas identificaciones NO deberían usar 890405 con IDE 977)
    identificaciones_con_861801: set[str] = set()
    for row in range(2, data_sheet.max_row + 1):
        ident = data_sheet.cell(row=row, column=ident_idx + 1).value if ident_idx is not None else None
        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip().upper() if codigo else ""

        if codigo_str == "861801" and ident:
            identificaciones_con_861801.add(str(ident).strip())

    logger.info("IDE REVERSE: Identificaciones con 861801: %d", len(identificaciones_con_861801))

    problemas = []
    facturas_procesadas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = _normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_procesadas:
            continue

        ide_contrato = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value
        ide_str = str(ide_contrato).strip() if ide_contrato else ""

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip() if codigo else ""

        ident = data_sheet.cell(row=row, column=ident_idx + 1).value if ident_idx is not None else None
        ident_str = str(ident).strip() if ident else ""

        # Verificar regla 986
        if ide_str == "986":
            codigo_esperado = IDE_CONTRATO_REVERSE.get("986")
            if codigo_str != codigo_esperado:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": codigo_esperado,
                    "observacion": f"IDE 986 → Código debe ser {codigo_esperado}",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 977
        if ide_str == "977":
            if codigo_str == "861801":
                # Válido
                continue
            elif codigo_str == "890405":
                # Depende: si la identificación tiene 861801 en otra factura,
                # entonces 890405 debería ser IDE 976 (ERROR)
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "976",
                        "observacion": "890405 con IDE 977 inválido - Identificación tiene 861801, debería ser IDE 976",
                    })
                    facturas_procesadas.add(factura_str)
                # Si NO tiene 861801, es válido
                continue
            else:
                # Cualquier otro código con IDE 977 es error
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": f"IDE 977 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 979
        if ide_str == "979":
            if codigo_str == "861801":
                # Válido
                continue
            elif codigo_str == "890405":
                # Depende: si la identificación tiene 861801 en otra factura,
                # entonces 890405 debería ser 967 (ERROR)
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "967",
                        "observacion": "890405 con IDE 979 inválido - Identificación tiene 861801, debería ser 967",
                    })
                    facturas_procesadas.add(factura_str)
                # Si NO tiene 861801, es válido
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 979 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 839
        if ide_str == "839":
            codigo_esperado = "906340"
            if codigo_str != codigo_esperado:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": codigo_esperado,
                    "observacion": f"IDE 839 → Código debe ser {codigo_esperado}",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 842
        if ide_str == "842":
            codigo_esperado = "906340"
            if codigo_str != codigo_esperado:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": codigo_esperado,
                    "observacion": f"IDE 842 → Código debe ser {codigo_esperado}",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 958
        if ide_str == "958":
            if codigo_str == "861801":
                # Válido
                continue
            elif codigo_str == "890405":
                # Depende: si la identificación tiene 861801 en otra factura,
                # entonces 890405 debería ser 959 (ERROR)
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "959",
                        "observacion": "890405 con IDE 958 inválido - Identificación tiene 861801, debería ser 959",
                    })
                    facturas_procesadas.add(factura_str)
                # Si NO tiene 861801, es válido
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 958 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 961
        if ide_str == "961":
            if codigo_str == "861801":
                # Válido
                continue
            elif codigo_str == "890405":
                # Depende: si la identificación tiene 861801 en otra factura,
                # entonces 890405 debería ser 962 (ERROR)
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "962",
                        "observacion": "890405 con IDE 961 inválido - Identificación tiene 861801, debería ser 962",
                    })
                    facturas_procesadas.add(factura_str)
                # Si NO tiene 861801, es válido
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 961 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 922
        if ide_str == "922":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "921",
                        "observacion": "890405 con IDE 922 inválido - Identificación tiene 861801, debería ser 921",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 922 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 863
        if ide_str == "863":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "862",
                        "observacion": "890405 con IDE 863 inválido - Identificación tiene 861801, debería ser 862",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 863 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 975
        if ide_str == "975":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "968",
                        "observacion": "890405 con IDE 975 inválido - Identificación tiene 861801, debería ser 968",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 975 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 920
        if ide_str == "920":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "919",
                        "observacion": "890405 con IDE 920 inválido - Identificación tiene 861801, debería ser 919",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 920 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 908
        if ide_str == "908":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "909",
                        "observacion": "890405 con IDE 908 inválido - Identificación tiene 861801, debería ser 909",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 908 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 970 (ESS118): 735301, 861801 o 890205
        if ide_str == "970":
            codigos_permitidos = {"735301", "861801", "890205"}
            if codigo_str in codigos_permitidos:
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "735301, 861801 o 890205",
                    "observacion": "IDE 970 (ESS118) → Código debe ser 735301, 861801 o 890205",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 974 (ESS118): 735301, 861801 o 890405
        if ide_str == "974":
            codigos_permitidos = {"735301", "861801", "890405"}
            if codigo_str in codigos_permitidos:
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "735301, 861801 o 890405",
                    "observacion": "IDE 974 (ESS118) → Código debe ser 735301, 861801 o 890405",
                })
                facturas_procesadas.add(factura_str)
            continue

    return problemas


def _log_verificacion_codigos_ess118(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[str]:
    """
    Verifica códigos CUPS con IDE Contrato = 969 contra URGENCIAS_CAPITA_CUPS_CODES.
    
    Compara cada código que tiene IDE=969 contra el listado de códigos válidos
    de Urgencias Capita (definido en constants.py).
    
    Returns:
        Lista de códigos no encontrados en URGENCIAS_CAPITA_CUPS_CODES
    """
    # Usar el listado hardcodedo de urgencias capita
    cups_validos = URGENCIAS_CAPITA_CUPS_CODES
    
    if not cups_validos:
        logger.warning("No hay códigos en URGENCIAS_CAPITA_CUPS_CODES")
        return set()
    
    logger.info("Verificando códigos IDE=969 contra URGENCIAS_CAPITA_CUPS_CODES (%d códigos válidos)", len(cups_validos))
    
    # Usar claves del diccionario indices
    codigo_idx = indices.get("codigo")
    ide_contrato_idx = indices.get("ide_contrato")
    codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")
    
    if codigo_idx is None:
        logger.warning("No hay índice de Código")
        return set()
    
    # Collect códigos únicos con IDE = 969
    codigos_ide_969 = set()
    
    for row in range(2, data_sheet.max_row + 1):
        # Verificar IDE = 969
        ide_contrato = None
        if ide_contrato_idx is not None:
            ide_contrato = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value
        
        ide_str = str(ide_contrato).strip() if ide_contrato else ""
        
        if ide_str != "969":
            continue
        
        # Verificar excepción: Código Tipo Procedimiento = 09, 12, 13 → no incluir
        codigo_tipo = None
        if codigo_tipo_proc_idx:
            codigo_tipo = data_sheet.cell(row=row, column=codigo_tipo_proc_idx + 1).value
        
        if codigo_tipo and str(codigo_tipo).strip() in ["09", "12", "13"]:
            continue
        
        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        if codigo:
            codigos_ide_969.add(str(codigo).strip())
    
    if not codigos_ide_969:
        logger.info("No hay códigos con IDE=969")
        return set()
    
    logger.info("Códigos únicos con IDE=969: %d", len(codigos_ide_969))
    
    # Verificar cada código contra URGENCIAS_CAPITA_CUPS_CODES
    codigos_no_encontrados = set()
    
    for codigo in codigos_ide_969:
        if codigo not in cups_validos:
            codigos_no_encontrados.add(codigo)
    
    if codigos_no_encontrados:
        logger.warning("Códigos NO en URGENCIAS_CAPITA_CUPS_CODES (%d): %s",
                     len(codigos_no_encontrados), sorted(codigos_no_encontrados))
    else:
        logger.info("Todos los códigos con IDE=969 están en URGENCIAS_CAPITA_CUPS_CODES")
    
    return codigos_no_encontrados


def _log_resumen_ide_contrato(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> None:
    """
    Log de resumen de valores únicos de código y entidad para debug de reglas IDE Contrato.
    """
    codigo_idx = indices.get("codigo")
    codigo_entidad_idx = indices.get("codigo_entidad_cobrar")
    ide_contrato_idx = indices.get("ide_contrato")
    
    if codigo_idx is None or codigo_entidad_idx is None:
        logger.warning("No hay índices de código o entidad para resumir")
        return
    
    codigos_set = set()
    entidades_set = set()
    ide_contratos_set = set()
    
    for row in range(2, min(data_sheet.max_row + 1, 500)):  # Limitado a primeras 500 filas
        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        entidad = data_sheet.cell(row=row, column=codigo_entidad_idx + 1).value
        
        if codigo:
            codigos_set.add(str(codigo).strip())
        if entidad:
            entidades_set.add(str(entidad).strip())
        
        if ide_contrato_idx is not None:
            ide = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value
            if ide:
                ide_contratos_set.add(str(ide).strip())
    
    # Mostrar las primeras 10 filas de datos crudos
    logger.warning("=== PRIMERAS FILAS DATOS IDE CONTRATO ===")
    for row in range(2, min(data_sheet.max_row + 1, 12)):
        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_entidad = data_sheet.cell(row=row, column=codigo_entidad_idx + 1).value if codigo_entidad_idx is not None else None
        ide = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value if ide_contrato_idx is not None else None
        factura = data_sheet.cell(row=row, column=indices.get("numero_factura", 0) + 1).value if indices.get("numero_factura") is not None else None
        
        logger.warning("Fila %d: Factura=%s, Código=%s, CódEntidad=%s, IDE=%s",
                       row, factura, codigo, codigo_entidad, ide)
    logger.warning("==========================================")
    
    logger.warning("=== RESUMEN DATOS EXCEL PARA REGLAS IDE CONTRATO ===")
    logger.warning("Códigos únicos encontrados (%d): %s", len(codigos_set), sorted(codigos_set))
    logger.warning("Códigos Entidad únicos encontrados (%d): %s", len(entidades_set), sorted(entidades_set))
    logger.warning("IDE Contrato únicos encontrados (%d): %s", len(ide_contratos_set), sorted(ide_contratos_set))
    logger.warning("=========================================================")


def _write_column(sheet: Worksheet, column: int, values: list[str], start_row: int = 2) -> None:
    """Escribe una lista de valores en una columna."""
    for i, value in enumerate(values, start=start_row):
        sheet.cell(row=i, column=column, value=value)


def _detect_revision_entidad_86_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con Cód Entidad Cobrar = 86 que requieren revisión manual.

    En Urgencias, cuando el código entidad cobrar es '86', se marca la factura
    como revisión necesaria (NO es un error de validación, es una advertencia).

    Returns:
        Lista de dicts con keys: 'factura', 'codigo', 'procedimiento', 'entidad', 'ide_contrato'
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_entidad_idx = indices.get("codigo_entidad_cobrar")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")
    ide_contrato_idx = indices.get("ide_contrato")

    if None in (num_fact_idx, codigo_entidad_idx):
        logger.warning(
            "Revision Entidad 86 - Columnas necesarias no encontradas: "
            "numero_factura=%s, codigo_entidad_cobrar=%s",
            num_fact_idx,
            codigo_entidad_idx,
        )
        return []

    revision_items: list[dict[str, str]] = []
    facturas_vistas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = _normalize_invoice(numero_factura)
        if not factura_str:
            continue

        # Saltar si ya registramos esta factura
        if factura_str in facturas_vistas:
            continue

        codigo_entidad = data_sheet.cell(row=row, column=codigo_entidad_idx + 1).value
        if codigo_entidad is None:
            continue

        codigo_entidad_str = str(codigo_entidad).strip().upper()
        if codigo_entidad_str != "86":
            continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value if codigo_idx is not None else ""
        procedimiento = data_sheet.cell(row=row, column=procedimiento_idx + 1).value if procedimiento_idx is not None else ""
        ide_contrato = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value if ide_contrato_idx is not None else ""

        facturas_vistas.add(factura_str)
        revision_items.append({
            "factura": factura_str,
            "codigo": str(codigo).strip() if codigo else "",
            "procedimiento": str(procedimiento).strip() if procedimiento else "",
            "entidad": codigo_entidad_str,
            "ide_contrato": str(ide_contrato).strip() if ide_contrato else "",
        })

    logger.info(
        "Revision Entidad 86 - Filas procesadas: %d, Facturas únicas encontradas: %d",
        data_sheet.max_row - 1,
        len(revision_items),
    )
    return revision_items


def _detect_revision_cantidad_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta filas con cantidad anómala que requieren revisión manual.

    Reglas:
    - General: Cantidad > 1 (excepto códigos exentos)
    - 02+Lab=No: Cantidad > 2 (código 903883: límite 5)
    - 09/12: Cantidad > 20 (código V03AN0101: siempre permitido)

    Returns:
        Lista de dicts con keys: 'factura', 'codigo', 'procedimiento',
        'cantidad', 'tipo_procedimiento', 'laboratorio'
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")
    cantidad_idx = indices.get("cantidad")
    tipo_proc_idx = indices.get("tipo_procedimiento")
    laboratorio_idx = indices.get("laboratorio")
    codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")

    if None in (num_fact_idx, codigo_idx, cantidad_idx, tipo_proc_idx, laboratorio_idx):
        logger.warning(
            "Revision Cantidad - Columnas necesarias no encontradas: "
            "numero_factura=%s, codigo=%s, cantidad=%s, "
            "tipo_procedimiento=%s, laboratorio=%s",
            num_fact_idx, codigo_idx, cantidad_idx,
            tipo_proc_idx, laboratorio_idx,
        )
        return []

    revision_items: list[dict[str, str]] = []

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = _normalize_invoice(numero_factura)
        if not factura_str:
            continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip().upper() if codigo else ""

        # Excepción 1: código está en la lista de exentos
        if codigo_str in CODIGOS_REVISION_CANTIDAD_EXENTOS:
            continue

        cantidad = data_sheet.cell(row=row, column=cantidad_idx + 1).value
        if not isinstance(cantidad, (int, float)):
            continue

        # Excepción 2: código con límite específico
        if codigo_str in CODIGOS_LIMITE_ESPECIFICO:
            max_cant = CODIGOS_LIMITE_ESPECIFICO[codigo_str]
            if cantidad <= max_cant:
                continue
            # superó el límite → cae al append (lo marco como revisión)

        tipo_proc = data_sheet.cell(row=row, column=tipo_proc_idx + 1).value
        tipo_proc_str = str(tipo_proc).strip() if tipo_proc else ""

        laboratorio = data_sheet.cell(row=row, column=laboratorio_idx + 1).value
        laboratorio_str = str(laboratorio).strip() if laboratorio else ""

        # Leer Código Tipo Procedimiento una sola vez
        codigo_tipo_proc_str = ""
        if codigo_tipo_proc_idx is not None:
            codigo_tipo_proc = data_sheet.cell(row=row, column=codigo_tipo_proc_idx + 1).value
            codigo_tipo_proc_str = str(codigo_tipo_proc).strip() if codigo_tipo_proc else ""

        # --- Regla para 02 + Lab=No: cantidad máxima 2 (903883: máximo 5) ---
        if (codigo_tipo_proc_str == CODIGO_TIPO_PROCEDIMIENTO_REVISION_LAB
                and laboratorio_str == LABORATORIO_REVISION_EXENTO):
            # 903883 tiene límite propio (5)
            if codigo_str == CODIGO_ESPECIAL_02_LAB:
                if cantidad <= CANTIDAD_MAX_02_LAB_903883:
                    continue
            # General: límite 2
            elif cantidad <= CANTIDAD_MAX_02_LAB:
                continue
            # > límite → cae al append

        # --- Regla para 09/12: cantidad máxima 20 (excepto V03AN0101) ---
        if codigo_tipo_proc_str in CODIGOS_TIPO_PROC_09_12:
            # V03AN0101 siempre permitido
            if codigo_str == CODIGO_EXENTO_V03AN0101:
                continue
            # Si cantidad <= 20, está dentro del límite
            if cantidad <= CANTIDAD_MAX_09_12:
                continue
            # cantidad > 20 → cae al append

        # --- Regla general: cantidad > 1 ---
        else:
            if cantidad <= 1:
                continue

        procedimiento = ""
        if procedimiento_idx is not None:
            proc_value = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
            procedimiento = str(proc_value).strip() if proc_value else ""

        revision_items.append({
            "factura": factura_str,
            "codigo": codigo_str,
            "procedimiento": procedimiento,
            "cantidad": cantidad,
            "tipo_procedimiento": tipo_proc_str,
            "laboratorio": laboratorio_str,
        })

    logger.info(
        "Revision Cantidad - Filas procesadas: %d, Items encontrados: %d",
        data_sheet.max_row - 1,
        len(revision_items),
    )
    return revision_items


def _build_urgencias_normalized_rows(
    problemas_centros: list[dict],
    problemas_ide_contrato: list[dict],
    problemas_cups_equivalentes: list[dict],
    mal_capitado: list[dict],
    cantidades_urgencias: list[dict],
    cantidades_soat_urgencias: list[dict],
    cantidades_hospitalizacion: list[dict],
    cantidades_soat_hospitalizacion: list[dict],
    responsables_map: dict[str, str],
    decimales: list[str] | None = None,
    tipo_identificacion_edad: list[dict] | None = None,
    profesionales: list[dict] | None = None,
    entidad_afiliacion_comparison: list[dict] | None = None,
    fecha_cierre_vacia_map: dict[str, bool] | None = None,
    tipo_usuario: list[dict] | None = None,
    revision_entidad_86: list[dict] | None = None,
    revision_cantidad: list[dict] | None = None,
) -> list[dict[str, str]]:
    """
    Normaliza todos los tipos de error de Urgencias en filas de 6 columnas.
    
    Formato de salida: cada dict tiene:
        - tipo_error: str
        - factura: str
        - responsable_cierra: str
        - descripcion: str
        - procedimiento: str (Var1)
        - detalle: str (Var2)
        - fecha_cierre_vacia: bool
    
    Args:
        problemas_centros: Lista de errores de centros de costo
        problemas_ide_contrato: Lista de errores de IDE Contrato
        problemas_cups_equivalentes: Lista de errores de CUPS equivalentes
        mal_capitado: Lista de errores MAL CAPITADO
        cantidades_urgencias: Lista de errores de cantidades en Urgencias
        cantidades_hospitalizacion: Lista de errores de cantidades en Hospitalización
        responsables_map: Dict {factura: responsable}
        decimales: Lista de errores de decimales (opcional)
        tipo_identificacion_edad: Lista de errores de tipo identificación/edad (opcional)
        profesionales: Lista de errores de profesionales (opcional)
        entidad_afiliacion_comparison: Lista de errores de entidad vs afiliación (opcional)
        fecha_cierre_vacia_map: Dict {factura: True si Fecha Cierre está vacía} (opcional)
        revision_entidad_86: Lista de revisiones necesarias para entidad 86 (opcional)
        revision_cantidad: Lista de revisiones necesarias por cantidad > 1 (opcional)
    
    Returns:
        Lista de dicts normalizados listos para escribir en Excel o renderizar en HTML
    """
    rows: list[dict[str, str]] = []

    def _get_fecha_cierre_vacia(factura: str) -> bool:
        if fecha_cierre_vacia_map is None:
            return False
        return fecha_cierre_vacia_map.get(factura, False)

    def _get_responsable(factura: str) -> str:
        return responsables_map.get(factura, "")

    def _build_procedimiento(codigo: str, procedimiento: str) -> str:
        """Construye 'Código - Nombre' si ambos existen, sino el que esté presente."""
        codigo = str(codigo).strip() if codigo else ""
        procedimiento = str(procedimiento).strip() if procedimiento else ""
        if codigo and procedimiento:
            return f"{codigo} - {procedimiento}"
        return codigo or procedimiento or ""

    def _build_ide_contrato_descripcion(ide_deberia: str) -> str:
        """Construye descripción de IDE Contrato, omitiendo el prefix para errores de DB."""
        if ide_deberia in ("Código no en DB", "CÓDIGO NO EN DB"):
            return ide_deberia
        return f"IDE Contrato debería ser {ide_deberia}"

    # --- Centros de Costo ---
    for item in problemas_centros:
        factura = item.get("factura", str(item.get("invoice", "")))
        codigo = item.get("codigo", "")
        proc = item.get("procedimiento", "")
        rows.append({
            "tipo_error": "Centros de Costo",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": f"Centro de costo debería ser {item.get('centro_deberia', 'N/A')}",
            "procedimiento": _build_procedimiento(codigo, proc),
            "detalle": item.get("centro_actual", ""),
            "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
        })

    # --- IDE Contrato ---
    for item in problemas_ide_contrato:
        factura = item.get("factura", "")
        codigo = item.get("codigo", "")
        proc = item.get("procedimiento", "")
        rows.append({
            "tipo_error": "IDE Contrato",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": _build_ide_contrato_descripcion(item.get('ide_contrato_deberia', 'N/A')),
            "procedimiento": _build_procedimiento(codigo, proc),
            "detalle": item.get("ide_contrato_actual", ""),
            "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
        })

    # --- Cups Equivalentes ---
    for item in problemas_cups_equivalentes:
        factura = item.get("factura", "")
        codigo_raw = item.get("codigo", "")
        proc_raw = item.get("procedimiento", "")
        estancia_str = item.get("estancia_str", "")
        # codigo puede ser str o list, normalizar
        if isinstance(codigo_raw, list):
            codigo_str = ", ".join(str(c) for c in codigo_raw)
        else:
            codigo_str = str(codigo_raw)
        proc_str = str(proc_raw).strip() if proc_raw else ""
        # Procedimiento: _build_sala_proc ya devuelve "código - nombre" tal cual del Excel
        # Si no hay nombre, cae a solo código
        proc_final = proc_str if proc_str else codigo_str
        # Detalle: estancia en día+hora
        if estancia_str:
            detalle = f"Estancia: {estancia_str}"
        else:
            detalle = codigo_str
        rows.append({
            "tipo_error": "Cups Equivalentes",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": item.get("accion", ""),
            "procedimiento": proc_final,
            "detalle": detalle,
            "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
        })

    # --- MAL CAPITADO ---
    for item in mal_capitado:
        factura = item.get("factura", "")
        codigo = item.get("codigo", "")
        proc = item.get("procedimiento", "")
        rows.append({
            "tipo_error": "MAL CAPITADO",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": item.get("observacion", ""),
            "procedimiento": _build_procedimiento(codigo, proc),
            "detalle": item.get("ide_contrato_actual", ""),
            "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
        })

    # --- Cantidades (Urgencias) ---
    for item in cantidades_urgencias:
        factura = item.get("factura", "")
        codigo = item.get("codigo", "")
        proc = item.get("procedimiento", "")
        cantidad = item.get("cantidad", "")
        rows.append({
            "tipo_error": "Cantidades",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": f"Cantidad {cantidad} debe ser ≤ 1 en Urgencias",
            "procedimiento": _build_procedimiento(codigo, proc),
            "detalle": str(cantidad),
            "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
        })

    # --- Cantidades SOAT Urgencias ---
    for item in cantidades_soat_urgencias:
        factura = item.get("factura", "")
        codigo = item.get("codigo", "")
        proc = item.get("procedimiento", "")
        cantidad = item.get("cantidad", "")
        rows.append({
            "tipo_error": "Cantidades SOAT",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": f"Cantidad {cantidad} debe ser = 1 (SOAT Urgencias)",
            "procedimiento": _build_procedimiento(codigo, proc),
            "detalle": str(cantidad),
            "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
        })

    # --- Cantidades Hospitalización ---
    for item in cantidades_hospitalizacion:
        factura = item.get("factura", "")
        codigo = item.get("codigo", "")
        proc = item.get("procedimiento", "")
        cantidad = item.get("cantidad", "")
        cantidad_esperada = item.get("cantidad_esperada", "")
        rows.append({
            "tipo_error": "Cantidades Hospitalización",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": f"Cantidad {cantidad} debería ser {cantidad_esperada}",
            "procedimiento": _build_procedimiento(codigo, proc),
            "detalle": str(cantidad),
            "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
        })

    # --- Cantidades SOAT Hospitalización ---
    for item in cantidades_soat_hospitalizacion:
        factura = item.get("factura", "")
        codigo = item.get("codigo", "")
        proc = item.get("procedimiento", "")
        cantidad = item.get("cantidad", "")
        cantidad_esperada = item.get("cantidad_esperada", "")
        rows.append({
            "tipo_error": "Cantidades SOAT Hospitalización",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": f"Cantidad {cantidad} debería ser {cantidad_esperada} (SOAT Hospitalización)",
            "procedimiento": _build_procedimiento(codigo, proc),
            "detalle": str(cantidad),
            "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
        })

    # --- Decimales ---
    if decimales:
        for factura in decimales:
            rows.append({
                "tipo_error": "Decimales",
                "factura": factura,
                "responsable_cierra": _get_responsable(factura),
                "descripcion": "Valores con decimales",
                "procedimiento": "Vlr. Procedimiento",
                "detalle": "Vlr. Subsidiado",
                "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
            })

    # --- Tipo Identificación / Edad ---
    if tipo_identificacion_edad:
        for item in tipo_identificacion_edad:
            factura = item.get("factura", "")
            num_id = item.get("numero_identificacion", "")
            anios = item.get("edad_anios", "")
            meses = item.get("edad_meses", "")
            tipo_actual = item.get("tipo_actual", "")
            tipo_deberia = item.get("tipo_deberia", "")
            rows.append({
                "tipo_error": "Tipo Identificación / Edad",
                "factura": factura,
                "responsable_cierra": _get_responsable(factura),
                "descripcion": f"Tipo actual {tipo_actual} debería ser {tipo_deberia}",
                "procedimiento": num_id,
                "detalle": f"{anios} años {meses} meses",
                "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
            })

    # --- Profesionales ---
    if profesionales:
        for item in profesionales:
            factura = item.get("factura", "")
            cod_prof = item.get("codigo_profesional", "")
            proc_nombre = item.get("procedimiento", "")
            nombre = item.get("nombre", "")
            rows.append({
                "tipo_error": "Profesionales",
                "factura": factura,
                "responsable_cierra": _get_responsable(factura),
                "descripcion": item.get("problema", item.get("regla", "")),
                "procedimiento": _build_procedimiento(cod_prof, proc_nombre),
                "detalle": f"Cód: {cod_prof}" if cod_prof else "",
                "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
            })

    # --- Código Entidad vs Afiliación ---
    if entidad_afiliacion_comparison:
        for item in entidad_afiliacion_comparison:
            factura = item.get("factura", "")
            # Procedimiento: Cód Entidad Cobrar + nombre de Entidad Cobrar
            cod = item.get("codigo_entidad_cobrar", "")
            nombre = item.get("entidad_cobrar_nombre", "")
            proc_entidad = f"{cod} - {nombre}" if cod and nombre else cod
            rows.append({
                "tipo_error": "Código Entidad vs Afiliación",
                "factura": factura,
                "responsable_cierra": _get_responsable(factura),
                "descripcion": item.get("problema", ""),
                "procedimiento": proc_entidad,
                "detalle": f"Afiliación: {item.get('entidad_afiliacion', '')}",
                "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
                "_header_override": "Entidad de factura",
            })

# --- Tipo Usuario ---
    if tipo_usuario:
        for item in tipo_usuario:
            factura = item.get("factura", "")
            tipo_actual = item.get("tipo_actual", "")
            rows.append({
                "tipo_error": "Tipo Usuario",
                "factura": factura,
                "responsable_cierra": _get_responsable(factura),
                "descripcion": "Revisar tipo usuario en Targetero",
                "procedimiento": "",
                "detalle": tipo_actual,
                "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
            })

    # --- ⚠️ Revisión Necesaria: Entidad 86 ---
    if revision_entidad_86:
        for item in revision_entidad_86:
            factura = item.get("factura", "")
            codigo = item.get("codigo", "")
            proc = item.get("procedimiento", "")
            rows.append({
                "tipo_error": "⚠️ Revisión Necesaria",
                "factura": factura,
                "responsable_cierra": _get_responsable(factura),
                "descripcion": "Cód Entidad Cobrar = 86 requiere revisión manual",
                "procedimiento": _build_procedimiento(codigo, proc),
                "detalle": "86",
                "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
            })

    if revision_cantidad:
        for item in revision_cantidad:
            factura = item.get("factura", "")
            codigo = item.get("codigo", "")
            proc = item.get("procedimiento", "")
            cantidad = item.get("cantidad", "")
            rows.append({
                "tipo_error": "⚠️ Revisión Necesaria",
                "factura": factura,
                "responsable_cierra": _get_responsable(factura),
                "descripcion": "Cantidad > 1 con código no exento requiere revisión manual",
                "procedimiento": _build_procedimiento(codigo, proc),
                "detalle": f"Cant: {cantidad}",
                "fecha_cierre_vacia": _get_fecha_cierre_vacia(factura),
            })

    return rows


def _build_odontologia_normalized_rows(
    decimales: list[dict] | list[str],
    doble_tipo: list[dict],
    ruta_dup: list[dict],
    profesionales: list[dict],
    cantidades: list[dict],
    tipo_id_edad: list[dict],
    centro_costo: list[dict],
    ide_contrato: list[dict],
    responsable_cierra: dict[str, str],
    entidad_afiliacion_comparison: list[dict] | None = None,
    tipo_usuario: list[dict] | None = None,
) -> list[dict[str, str]]:
    """
    Normaliza todos los tipos de error de Odontología/Equipos Básicos en filas de 6 columnas.

    Args:
        decimales: Lista de dicts con "factura", "valores" o lista de strings (solo facturas)
        doble_tipo: Lista de dicts con "factura", "tipos"
        ruta_dup: Lista de dicts con "identificacion", "facturas", "cantidad"
        profesionales: Lista de dicts con "factura", "codigo_profesional", "procedimiento", ...
        cantidades: Lista de dicts con "factura", "tipo_procedimiento", "cantidad", ...
        tipo_id_edad: Lista de dicts con "factura", "tipo_actual", "tipo_deberia", "edad"
        centro_costo: Lista de dicts con "factura", "centro_actual", "centro_deberia", ...
        ide_contrato: Lista de dicts con "factura", "codigo", "cod_entidad", "ide_actual", ...
        responsable_cierra: Dict {factura: responsable}

    Returns:
        Lista de dicts normalizados con tipo_error, factura, responsable_cierra,
        descripcion, procedimiento (Var1), detalle (Var2)
    """
    rows: list[dict[str, str]] = []

    def _get_responsable(factura: str) -> str:
        return responsable_cierra.get(factura, "")

    def _build_procedimiento(codigo: str, procedimiento: str) -> str:
        codigo = str(codigo).strip() if codigo else ""
        procedimiento = str(procedimiento).strip() if procedimiento else ""
        if codigo and procedimiento:
            return f"{codigo} - {procedimiento}"
        return codigo or procedimiento or ""

    # --- Decimales ---
    for item in decimales:
        if isinstance(item, dict):
            factura = item.get("factura", "")
            valores = item.get("valores", "")
        else:
            factura = str(item) if item else ""
            valores = ""
        rows.append({
            "tipo_error": "Decimales",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": f"Valores con decimales: {valores}" if valores else "Valores con decimales",
            "procedimiento": valores,
            "detalle": "",
        })

    # --- Doble tipo procedimiento ---
    for item in doble_tipo:
        factura = item.get("factura", "")
        tipos = item.get("tipos", "")
        rows.append({
            "tipo_error": "Doble tipo procedimiento",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": f"Múltiples tipos de procedimiento",
            "procedimiento": "",
            "detalle": tipos,
        })

    # --- Ruta Duplicada ---
    for item in ruta_dup:
        identificacion = item.get("identificacion", "")
        facturas_list = item.get("facturas", "")
        cantidad = item.get("cantidad", 0)
        rows.append({
            "tipo_error": "Ruta Duplicada",
            "factura": identificacion,
            "responsable_cierra": _get_responsable(identificacion),
            "descripcion": f"Paciente con {cantidad} facturas en PyP",
            "procedimiento": facturas_list,
            "detalle": identificacion,
        })

    # --- Profesionales (reemplaza Convenio de procedimiento) ---
    for item in profesionales:
        factura = item.get("factura", "")
        cod_prof = item.get("codigo_profesional", "")
        proc_nombre = item.get("procedimiento", "")
        problema = item.get("problema", "")
        regla = item.get("regla", "")
        rows.append({
            "tipo_error": "Convenio de procedimiento",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": problema or regla,
            "procedimiento": _build_procedimiento(cod_prof, proc_nombre),
            "detalle": problema or "",
        })

    # --- Cantidades ---
    for item in cantidades:
        factura = item.get("factura", "")
        tipo_proc = item.get("tipo_procedimiento", "")
        cantidad_val = item.get("cantidad", "")
        problema = item.get("problema", "")
        rows.append({
            "tipo_error": "Cantidades",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": problema or f"Cantidad anómala: {cantidad_val}",
            "procedimiento": tipo_proc,
            "detalle": str(cantidad_val),
        })

    # --- Tipo Identificación vs Edad ---
    for item in tipo_id_edad:
        factura = item.get("factura", "")
        tipo_actual = item.get("tipo_actual", "")
        tipo_deberia = item.get("tipo_deberia", "")
        edad = item.get("edad", "")
        rows.append({
            "tipo_error": "Tipo Identificación",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": f"{tipo_actual} debería ser {tipo_deberia}",
            "procedimiento": f"Edad: {edad} años",
            "detalle": "",
        })

    # --- Centro Costo ---
    for item in centro_costo:
        factura = item.get("factura", "")
        centro_actual = item.get("centro_actual", "")
        centro_deberia = item.get("centro_deberia", "")
        rows.append({
            "tipo_error": "Centro Costo",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": f"Centro de costo debería ser {centro_deberia}",
            "procedimiento": "",
            "detalle": centro_actual,
        })

    # --- IDE Contrato ---
    for item in ide_contrato:
        factura = item.get("factura", "")
        codigo = item.get("codigo", "")
        ide_actual = item.get("ide_actual", "")
        ide_deberia = item.get("ide_deberia", "")
        nota = item.get("nota", "")
        rows.append({
            "tipo_error": "IDE Contrato",
            "factura": factura,
            "responsable_cierra": _get_responsable(factura),
            "descripcion": f"IDE Contrato debería ser {ide_deberia} ({nota})" if nota else f"IDE Contrato debería ser {ide_deberia}",
            "procedimiento": _build_procedimiento(codigo, ""),
            "detalle": ide_actual,
        })

    # --- Código Entidad vs Afiliación ---
    if entidad_afiliacion_comparison:
        for item in entidad_afiliacion_comparison:
            factura = item.get("factura", "")
            cod = item.get("codigo_entidad_cobrar", "")
            nombre = item.get("entidad_cobrar_nombre", "")
            proc_entidad = f"{cod} - {nombre}" if cod and nombre else cod
            rows.append({
                "tipo_error": "Código Entidad vs Afiliación",
                "factura": factura,
                "responsable_cierra": _get_responsable(factura),
                "descripcion": item.get("problema", ""),
                "procedimiento": proc_entidad,
                "detalle": f"Afiliación: {item.get('entidad_afiliacion', '')}",
            })

    # --- Tipo Usuario ---
    if tipo_usuario:
        for item in tipo_usuario:
            factura = item.get("factura", "")
            tipo_actual = item.get("tipo_actual", "")
            rows.append({
                "tipo_error": "Tipo Usuario",
                "factura": factura,
                "responsable_cierra": _get_responsable(factura),
                "descripcion": "Revisar tipo usuario en Targetero",
                "procedimiento": "",
                "detalle": tipo_actual,
            })

    return rows


def create_revision_sheet(
    workbook: Workbook,
    data_sheet: Worksheet,
    area: str = AREA_ODONTOLOGIA,
    profesional_dias: dict[str, list[int]] | None = None,
    permitir_todos_centros: bool = False,
) -> dict[str, Any]:
    """
    Crea la hoja Revision con los problemas detectados.
    
    Args:
        workbook: Libro de Excel (debe tener una hoja activa con datos)
        data_sheet: Hoja de datos a analizar
        area: Área del sistema ("odontologia" o "urgencias")
        profesional_dias: Dict {identificacion: [dias]} con días seleccionados por profesional
        permitir_todos_centros: Si True, solo permite ODONTOLOGIA y EXTRAMURAL
    
    Returns:
        Dict con información de los problemas encontrados
    """
    sheet = workbook.create_sheet(title=REVISION_SHEET)
    
    # Insertar fila vacía arriba
    sheet.insert_rows(1)
    
    # Obtener índices de columnas (coincidencia exacta - reporta faltantes)
    headers = [
        data_sheet.cell(row=1, column=col).value
        for col in range(1, data_sheet.max_column + 1)
    ]
    indices, missing_columns = _get_column_indices(headers)
    
    # Si hay columnas faltantes, incluir en el resultado para mostrar al usuario
    if missing_columns:
        logger.error("Columnas faltantes en el Excel: %s", missing_columns)
    
    # Seleccionar headers según el área
    if area == AREA_URGENCIAS:
        revision_headers = URGENCIA_REVISION_HEADERS
        header_style = create_urgencia_header_style()
    else:
        revision_headers = REVISION_HEADERS
        header_style = create_header_style()
    
    # Aplicar headers con estilo en fila 2
    for col, header in revision_headers.items():
        cell = sheet.cell(row=2, column=col, value=header)
        cell.font = header_style["font"]
        cell.fill = header_style["fill"]
        cell.border = header_style["border"]
        cell.alignment = header_style["alignment"]
    
# Detectar problemas según el área
    if area == AREA_URGENCIAS:
        # --- Construir mapa responsable_cierra ---
        responsable_cierra: dict[str, str] = {}
        responsable_cierra_idx = indices.get("responsable_cierra")
        num_fact_idx = indices.get("numero_factura")
        if responsable_cierra_idx is not None and num_fact_idx is not None:
            for row in range(2, data_sheet.max_row + 1):
                numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
                factura = _normalize_invoice(numero)
                if not factura:
                    continue
                raw = data_sheet.cell(row=row, column=responsable_cierra_idx + 1).value
                resp = str(raw).strip() if raw else ""
                if resp and factura not in responsable_cierra:
                    responsable_cierra[factura] = resp

        # --- Detectar todos los problemas ---
        logger.warning("=== VERIFICANDO CÓDIGOS ESS118 CONTRA DB ===")
        problemas_codigos_no_en_db = _get_codigos_no_en_db_ess118(data_sheet, indices)
        codigos_no_en_db_set = {item["codigo"] for item in problemas_codigos_no_en_db}

        if problemas_codigos_no_en_db:
            logger.warning("Procedimientos NO encontrados en DB para ESS118 (%d errores): %s",
                        len(problemas_codigos_no_en_db), sorted(codigos_no_en_db_set))
        else:
            logger.warning("Todos los códigos de ESS118 están en DB")

        try:
            result = _detect_centro_costo_urgencias(data_sheet, indices, problemas_codigos_no_en_db)
            if isinstance(result, tuple) and len(result) >= 3:
                problemas_centros, problemas_ide_contrato, problemas_cups_equivalentes = result[0], result[1], result[2]
            else:
                logger.error("Función retornó formato inesperado: %s", type(result))
                problemas_centros, problemas_ide_contrato, problemas_cups_equivalentes = [], [], []
        except Exception as exc:
            logger.exception("Error en _detect_centro_costo_urgencias: %s", exc)
            problemas_centros, problemas_ide_contrato, problemas_cups_equivalentes = [], [], []

        mal_capitado = _detect_mal_capitado(data_sheet, indices)
        cantidades_urgencias = _detect_cantidades_urgencias(data_sheet, indices)
        cantidades_hospitalizacion = _detect_cantidades_hospitalizacion(data_sheet, indices)

        # --- ⚠️ Revisión Necesaria: Entidad 86 ---
        revision_entidad_86 = _detect_revision_entidad_86_urgencias(data_sheet, indices)
        logger.info(
            "create_revision_sheet - Revision Entidad 86 encontradas: %d",
            len(revision_entidad_86),
        )

        # --- ⚠️ Revisión Necesaria: Cantidad > 1 ---
        revision_cantidad = _detect_revision_cantidad_urgencias(data_sheet, indices)
        logger.info(
            "create_revision_sheet - Revision Cantidad encontradas: %d",
            len(revision_cantidad),
        )

        # --- Normalizar todos los errores a filas de 6 columnas ---
        normalized_rows = _build_urgencias_normalized_rows(
            problemas_centros=problemas_centros,
            problemas_ide_contrato=problemas_ide_contrato,
            problemas_cups_equivalentes=problemas_cups_equivalentes,
            mal_capitado=mal_capitado,
            cantidades_urgencias=cantidades_urgencias,
            cantidades_soat_urgencias=[],  # No aplica para Odontología
            cantidades_hospitalizacion=cantidades_hospitalizacion,
            cantidades_soat_hospitalizacion=[],  # No aplica para Odontología
            responsables_map=responsable_cierra,
            revision_entidad_86=revision_entidad_86,
            revision_cantidad=revision_cantidad,
        )

        # --- Escribir filas normalizadas en Excel ---
        # Buscar si hay rows con _header_override para la columna 5 (Procedimiento)
        has_header_override = any("_header_override" in row for row in normalized_rows)
        for i, row_data in enumerate(normalized_rows, start=3):
            sheet.cell(row=i, column=1, value=row_data["tipo_error"])
            sheet.cell(row=i, column=2, value=row_data["factura"])
            sheet.cell(row=i, column=3, value=row_data["responsable_cierra"])
            sheet.cell(row=i, column=4, value=row_data["descripcion"])
            # Columna 5: usar _header_override si existe, si no usar procedimiento
            col5_value = row_data.get("_header_override", row_data.get("procedimiento", ""))
            sheet.cell(row=i, column=5, value=col5_value)
            sheet.cell(row=i, column=6, value=row_data["detalle"])

        problemas_encontrados = {
            "normalizados": normalized_rows,
            "totales_por_tipo": {
                "Centros de Costo": len(problemas_centros),
                "IDE Contrato": len(problemas_ide_contrato),
                "Cups Equivalentes": len(problemas_cups_equivalentes),
                "MAL CAPITADO": len(mal_capitado),
                "Cantidades": len(cantidades_urgencias),
                "Cantidades Hospitalización": len(cantidades_hospitalizacion),
                "⚠️ Revisión Necesaria": len(revision_entidad_86) + len(revision_cantidad),
            },
        }
    else:
# Odontología / Equipos Básicos: detectar todos los problemas
        decimales = _detect_decimals(data_sheet, indices)
        doble_tipo = _detect_doble_tipo_procedimiento(data_sheet, indices)
        ruta_dup = _detect_ruta_duplicada(data_sheet, indices)
        cantidades = _detect_cantidades_anomalas(data_sheet, indices)
        tipo_id_edad = _detect_tipo_identificacion_edad(data_sheet, indices)
        profesionales = _detect_profesionales_odontologia(data_sheet, indices)
        centro_costo = _detect_centro_costo_odontologia(data_sheet, indices)
        ide_contrato = _detect_ide_contrato_odontologia(data_sheet, indices)

        # Construir mapa responsable_cierra
        responsable_cierra_map: dict[str, str] = {}
        rci = indices.get("responsable_cierra")
        nfi = indices.get("numero_factura")
        if rci is not None and nfi is not None:
            for row in range(2, data_sheet.max_row + 1):
                num = data_sheet.cell(row=row, column=nfi + 1).value
                fac = _normalize_invoice(num)
                if not fac:
                    continue
                raw = data_sheet.cell(row=row, column=rci + 1).value
                resp = str(raw).strip() if raw else ""
                if resp and fac not in responsable_cierra_map:
                    responsable_cierra_map[fac] = resp

        # Normalizar a filas de 6 columnas
        normalized_rows = _build_odontologia_normalized_rows(
            decimales=decimales,
            doble_tipo=doble_tipo,
            ruta_dup=ruta_dup,
            profesionales=profesionales,
            cantidades=cantidades,
            tipo_id_edad=tipo_id_edad,
            centro_costo=centro_costo,
            ide_contrato=ide_contrato,
            responsable_cierra=responsable_cierra_map,
        )

        # Escribir filas normalizadas en Excel
        for i, row_data in enumerate(normalized_rows, start=3):
            sheet.cell(row=i, column=1, value=row_data["tipo_error"])
            sheet.cell(row=i, column=2, value=row_data["factura"])
            sheet.cell(row=i, column=3, value=row_data["responsable_cierra"])
            sheet.cell(row=i, column=4, value=row_data["descripcion"])
            sheet.cell(row=i, column=5, value=row_data["procedimiento"])
            sheet.cell(row=i, column=6, value=row_data["detalle"])

        problemas_encontrados = {
            "normalizados": normalized_rows,
            "totales_por_tipo": {
                "Decimales": len(decimales),
                "Doble tipo procedimiento": len(doble_tipo),
                "Ruta Duplicada": len(ruta_dup),
                "Convenio de procedimiento": len(profesionales),
                "Cantidades": len(cantidades),
                "Tipo Identificación": len(tipo_id_edad),
                "Centro Costo": len(centro_costo),
                "IDE Contrato": len(ide_contrato),
            },
        }
    
    # Aplicar estilo a filas de datos (fila 3+) según el área
    if area == AREA_URGENCIAS:
        data_style = create_urgencia_data_row_style()
    else:
        data_style = create_data_row_style()
    
    for row in range(3, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = data_style["fill"]
            cell.border = data_style["border"]
            cell.alignment = data_style["alignment"]
    
    # Ajustar ancho de columnas automáticamente
    column_widths = auto_adjust_column_width(sheet)
    
    # Logging según el área
    if area == AREA_URGENCIAS:
        logger.info(
            "Hoja Revision Urgencias creada - Total filas normalizadas: %d",
            len(normalized_rows),
        )
    else:
        logger.info(
            "Hoja Revision Odontología/EB creada - Total filas normalizadas: %d",
            len(normalized_rows),
        )

    # Build resultado según el área
    if area == AREA_URGENCIAS:
        return {
            "rule": "create_revision_sheet",
            "sheet": REVISION_SHEET,
            "area": area,
            "headers": list(URGENCIA_REVISION_HEADERS.values()),
            "normalized_rows_count": len(normalized_rows),
            "problemas": problemas_encontrados,
            "column_widths": column_widths,
            "missing_columns": missing_columns,
        }
    else:
        return {
            "rule": "create_revision_sheet",
            "sheet": REVISION_SHEET,
            "area": area,
            "headers": list(REVISION_HEADERS.values()),
            "normalized_rows_count": len(normalized_rows),
            "problemas": problemas_encontrados,
            "column_widths": column_widths,
            "missing_columns": missing_columns,
        }


def detect_all_problems(
    data_sheet: Worksheet,
    area: str = AREA_ODONTOLOGIA,
    profesional_dias: dict[str, list[int]] | None = None,
    permitir_todos_centros: bool = False,
) -> tuple[dict[str, Any], dict[str, str]]:
    """
    Detecta todos los problemas en las facturas SIN crear hoja Excel.
    
    Retorna: (resultado_dict, responsables_map)
    
    Args:
        data_sheet: Hoja de Excel con los datos
        area: Área del sistema ("odontologia", "urgencias" o "equipos_basicos")
        profesional_dias: Dict {identificacion: [dias]} con días seleccionados por profesional
        permitir_todos_centros: Si True, solo permite ODONTOLOGIA y EXTRAMURAL sin validación por fecha
    
    Returns:
        (resultado_dict, responsables_map)
    """
    # Obtener índices de columnas (coincidencia exacta - reporta faltantes)
    headers = [
        data_sheet.cell(row=1, column=col).value
        for col in range(1, data_sheet.max_column + 1)
    ]
    indices, missing_columns = _get_column_indices(headers)
    
    # Si hay columnas faltantes, incluir en el resultado para mostrar al usuario
    if missing_columns:
        logger.error("Columnas faltantes en el Excel: %s", missing_columns)
    
    if area == AREA_URGENCIAS:
        # Urgencias: detectar códigos NO en DB con IDE=969
        # Excluir Código Tipo Procedimiento = 09, 12, 13
        
        # Debug: mostrar índices encontrados en el dict 'indices'
        logger.warning("=== DEBUG: Indices del dict para ESS118 ===")
        logger.warning(f"  Código: {indices.get('codigo')}")
        logger.warning(f"  Cód. Equivalente CUPS: {indices.get('codigo_equiv')}")
        logger.warning(f"  Código Tipo Procedimiento: {indices.get('codigo_tipo_procedimiento')}")
        logger.warning(f"  Codigo_Entidad: {indices.get('codigo_entidad_cobrar')}")
        logger.warning(f"  IDE Contrato: {indices.get('ide_contrato')}")
        
        logger.warning("=== VERIFICANDO CÓDIGOS ESS118 CONTRA DB ===")
        problemas_codigos_no_en_db = _get_codigos_no_en_db_ess118(data_sheet, indices)
        
        # Extraer códigos únicos para logging
        codigos_no_en_db_set = {item["codigo"] for item in problemas_codigos_no_en_db}
        
        if problemas_codigos_no_en_db:
            logger.warning("Procedimientos NO en DB (ESS118 + IDE=969): %d errores, códigos: %s",
                        len(problemas_codigos_no_en_db), sorted(codigos_no_en_db_set))
        else:
            logger.warning("No hay códigos sin DB con IDE=969 para ESS118")
        
        # Debug: mostrar valores de las primeras filas ESS118
        logger.warning("=== DEBUG: 5 primeras filas ESS118 ===")
        codigo_equiv_idx = indices.get("codigo_equiv")
        codigo_tipo_idx = indices.get("codigo_tipo_procedimiento")
        codigo_entidad_idx = indices.get("codigo_entidad_cobrar")
        ide_idx = indices.get("ide_contrato")
        
        count = 0
        for row in range(2, data_sheet.max_row + 1):
            entidad = data_sheet.cell(row=row, column=codigo_entidad_idx + 1).value if codigo_entidad_idx is not None else None
            if entidad and "ESS118" in str(entidad).upper():
                cod_equiv = data_sheet.cell(row=row, column=codigo_equiv_idx + 1).value if codigo_equiv_idx is not None else None
                cod_tipo = data_sheet.cell(row=row, column=codigo_tipo_idx + 1).value if codigo_tipo_idx is not None else None
                ide = data_sheet.cell(row=row, column=ide_idx + 1).value if ide_idx is not None else None
                logger.warning(f"  Fila {row}: equiv={cod_equiv}, tipo={cod_tipo}, IDE={ide}")
                count += 1
                if count >= 5:
                    break
        
        # Llamar función de detección con manejo de errores defensivo
        try:
            result = _detect_centro_costo_urgencias(data_sheet, indices, problemas_codigos_no_en_db)
            if isinstance(result, tuple) and len(result) == 3:
                problemas_centros, problemas_ide_contrato, problemas_cups_equivalentes = result
            else:
                logger.error("Función _detect_centro_costo_urgencias retornó formato inesperado: %s", type(result))
                problemas_centros, problemas_ide_contrato, problemas_cups_equivalentes = [], [], []
        except Exception as exc:
            logger.exception("Error en _detect_centro_costo_urgencias: %s", exc)
            problemas_centros, problemas_ide_contrato, problemas_cups_equivalentes = [], [], []
        
        # Agregar TODOS los procedimientos no encontrados en DB (no solo IDE=969)
        # como errores separados en ide_contrato
        for problema in problemas_codigos_no_en_db:
            problemas_ide_contrato.append({
                "factura": problema.get("factura", ""),
                "ide_contrato_actual": "N/A",
                "ide_contrato_deberia": "Código no en DB",
                "procedimiento": problema.get("procedimiento", ""),
                "codigo": problema.get("codigo", ""),
                "entidad": problema.get("entidad", ""),
            })
        
        logger.info("Agregados %d procedimientos sin DB a problemas_ide_contrato", len(problemas_codigos_no_en_db))
        
        # reglas transversales
        decimales = detect_decimales(data_sheet, indices)
        tipo_identificacion_edad = detect_tipo_documento_edad(data_sheet, indices)
        # Nueva regla: Cód Entidad Cobrar vs Entidad Afiliación (solo loggear las 5 primeras filas)
        entidad_afiliacion_comparison = detect_codigo_entidad_vs_entidad_afiliacion(
            data_sheet, indices, limit_log=5
        )

        # Nueva regla transversal: Tipo Usuario
        tipo_usuario = detect_tipo_usuario(data_sheet, indices)

        # Validación profesionales (solo Urgencias)
        profesionales = _detect_profesionales_urgencias(data_sheet, indices)
        logger.info("detect_all_problems - Urgencias, Profesionales encontrados: %d", len(profesionales))

        # Validación MAL CAPITADO
        mal_capitado = _detect_mal_capitado(data_sheet, indices)
        logger.info("detect_all_problems - Urgencias, MAL CAPITADO encontrados: %d", len(mal_capitado))

        # Validación Cantidades Urgencias (Tipo Factura = "Urgencias" + códigos específicos con cantidad > 1)
        cantidades_urgencias = _detect_cantidades_urgencias(data_sheet, indices)
        logger.info("detect_all_problems - Urgencias, Cantidades Urgencias encontradas: %d", len(cantidades_urgencias))

        # Validación Cantidades SOAT Urgencias (Tarifario = SOAT + Tipo Factura = Urgencias + códigos 39145, 38114, 38915, 39131 con cantidad != 1)
        cantidades_soat_urgencias = _detect_cantidades_soat_urgencias(data_sheet, indices)
        logger.info("detect_all_problems - Urgencias, Cantidades SOAT Urgencias encontradas: %d", len(cantidades_soat_urgencias))

        # Validación Cantidades Hospitalización (Tipo Factura = "Hospitalización" + códigos 129B02/890601)
        cantidades_hospitalizacion = _detect_cantidades_hospitalizacion(data_sheet, indices)
        logger.info("detect_all_problems - Urgencias, Cantidades Hospitalización encontradas: %d", len(cantidades_hospitalizacion))

        # Validación Cantidades SOAT Hospitalización (Tarifario = SOAT + Tipo Factura = Hospitalización + códigos 38114, 39131)
        cantidades_soat_hospitalizacion = _detect_cantidades_soat_hospitalizacion(data_sheet, indices)
        logger.info("detect_all_problems - Urgencias, Cantidades SOAT Hospitalización encontradas: %d", len(cantidades_soat_hospitalizacion))

        # Validación IDE Contrato REVERSE (sin entidad)
        ide_contrato_reverse = _detect_ide_contrato_reverse_urgencias(data_sheet, indices)
        logger.info("detect_all_problems - Urgencias, IDE Contrato REVERSE encontrados: %d", len(ide_contrato_reverse))

        # --- ⚠️ Revisión Necesaria: Entidad 86 ---
        revision_entidad_86 = _detect_revision_entidad_86_urgencias(data_sheet, indices)
        logger.info(
            "detect_all_problems - Revision Entidad 86 encontradas: %d",
            len(revision_entidad_86),
        )

        # --- ⚠️ Revisión Necesaria: Cantidad > 1 ---
        revision_cantidad = _detect_revision_cantidad_urgencias(data_sheet, indices)
        logger.info(
            "detect_all_problems - Revision Cantidad encontradas: %d",
            len(revision_cantidad),
        )

        logger.info("detect_all_problems (Urgencias): problemas_centros=%d, problemas_ide_contrato=%d, decimales=%d, tipo_id_edad=%d, entidad_afiliacion=%d, profesionales=%d, mal_capitado=%d, cantidades_urgencias=%d, cantidades_hospitalizacion=%d, ide_contrato_reverse=%d, revision_entidad_86=%d, revision_cantidad=%d",
len(problemas_centros), len(problemas_ide_contrato), len(decimales), len(tipo_identificacion_edad), len(entidad_afiliacion_comparison), len(profesionales), len(mal_capitado), len(cantidades_urgencias), len(cantidades_hospitalizacion), len(ide_contrato_reverse), len(revision_entidad_86), len(revision_cantidad))
        
        # Filtrar errores: si la misma factura+código tiene prioridad 1 y prioridad 2, mostrar solo prioridad 1
        # Agrupar por factura+código
        errores_por_factura_codigo = {}
        for item in problemas_centros:
            key = (item.get("factura", ""), item.get("codigo", ""))
            prioridad = item.get("prioridad", 1)
            if key not in errores_por_factura_codigo:
                errores_por_factura_codigo[key] = []
            errores_por_factura_codigo[key].append((item, prioridad))
        
        # Filtrar: quedarse con prioridad 1 si coexiste con prioridad 2
        problemas_centros_filtrados = []
        for key, items in errores_por_factura_codigo.items():
            prioridades = [p for _, p in items]
            if 1 in prioridades:
                # Hay prioridad 1, filtrar solo esos
                for item, p in items:
                    if p == 1:
                        problemas_centros_filtrados.append(item)
            else:
                # Solo hay prioridad 2, incluir todos
                for item, _ in items:
                    problemas_centros_filtrados.append(item)
        
        logger.info("FILTRO centros_de_costos: %d -> %d (eliminados %d errores de prioridad 2 porque hay prioridad 1 para misma factura+código)",
                 len(problemas_centros), len(problemas_centros_filtrados), len(problemas_centros) - len(problemas_centros_filtrados))
        
        # Incluir TODOS los campos en el resultado
        
        # Build responsable_cierra mapping
        responsable_cierra = {}
        responsable_cierra_idx = indices.get("responsable_cierra")
        num_fact_idx = indices.get("numero_factura")
        if responsable_cierra_idx is not None and num_fact_idx is not None:
            for row in range(2, data_sheet.max_row + 1):
                numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
                factura = _normalize_invoice(numero)
                if not factura:
                    continue
                raw = data_sheet.cell(row=row, column=responsable_cierra_idx + 1).value
                resp = str(raw).strip() if raw else ""
                if resp and factura not in responsable_cierra:
                    responsable_cierra[factura] = resp
        
        # Build fecha_cierre_vacia mapping: factura -> True si Fecha Cierre está vacía
        fecha_cierre_vacia: dict[str, bool] = {}
        fecha_cierre_idx = indices.get("fecha_cierre")
        num_fact_idx = indices.get("numero_factura")
        if fecha_cierre_idx is not None and num_fact_idx is not None:
            for row in range(2, data_sheet.max_row + 1):
                numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
                factura = _normalize_invoice(numero)
                if not factura:
                    continue
                fecha_cierre_val = data_sheet.cell(row=row, column=fecha_cierre_idx + 1).value
                # Si ALGUNA fila tiene Fecha Cierre vacía, marcar toda la factura
                if not fecha_cierre_val or str(fecha_cierre_val).strip() == "":
                    fecha_cierre_vacia[factura] = True
                elif factura not in fecha_cierre_vacia:
                    fecha_cierre_vacia[factura] = False
        
        logger.info("Fecha Cierre vacía detectada para %d facturas", 
                     sum(1 for v in fecha_cierre_vacia.values() if v))

        # Build normalized rows for unified 6-column display
        normalized_rows = _build_urgencias_normalized_rows(
            problemas_centros=problemas_centros_filtrados,
            problemas_ide_contrato=problemas_ide_contrato,
            problemas_cups_equivalentes=problemas_cups_equivalentes,
            mal_capitado=mal_capitado,
            cantidades_urgencias=cantidades_urgencias,
            cantidades_soat_urgencias=cantidades_soat_urgencias,
            cantidades_hospitalizacion=cantidades_hospitalizacion,
            cantidades_soat_hospitalizacion=cantidades_soat_hospitalizacion,
            responsables_map=responsable_cierra,
            decimales=decimales,
            tipo_identificacion_edad=tipo_identificacion_edad,
            profesionales=profesionales,
            entidad_afiliacion_comparison=entidad_afiliacion_comparison,
            fecha_cierre_vacia_map=fecha_cierre_vacia,
            tipo_usuario=tipo_usuario,
            revision_entidad_86=revision_entidad_86,
            revision_cantidad=revision_cantidad,
        )

        resultado = {
            "area": area,
            "problemas": {
                "normalizados": normalized_rows,
                "centros_de_costos": [
                    {
                        "tipo_factura": item.get("tipo_factura") or "-",
                        "factura": item["factura"],
                        "codigo": item.get("codigo", ""),
                        "procedimiento": item.get("procedimiento", ""),
                        "centro_actual": item["centro_actual"],
                        "centro_deberia": item["centro_deberia"],
                        "prioridad": item.get("prioridad", 1),  # Default prioridad 1 si no existe
                    }
                    for item in problemas_centros_filtrados
                ],
                "ide_contrato": [
                    {
                        "factura": item["factura"],
                        "ide_contrato_actual": item["ide_contrato_actual"],
                        "ide_contrato_deberia": item["ide_contrato_deberia"],
                        # Incluir campos adicionales si existen
                        "procedimiento": item.get("procedimiento", ""),
                        "codigo": item.get("codigo", ""),
                        "entidad": item.get("entidad", ""),
                        "nota": item.get("nota", ""),
                    }
                    for item in problemas_ide_contrato
                ],
                "cups_equivalentes": [
                    {
                        "factura": item["factura"],
                        "codigo": item["codigo"],
                        "codigo_equiv": item["codigo_equiv"],
                        "accion": item["accion"],
                    }
                    for item in problemas_cups_equivalentes
                ],
                # reglas transversales
                "decimales": decimales,
                "tipo_identificacion_edad": tipo_identificacion_edad,
                "codigo_entidad_vs_afiliacion": entidad_afiliacion_comparison,
                "tipo_usuario": tipo_usuario,
                "profesionales": profesionales,
                "mal_capitado": mal_capitado,
                "cantidades_urgencias": cantidades_urgencias,
                "cantidades_soat_urgencias": cantidades_soat_urgencias,
                "cantidades_hospitalizacion": cantidades_hospitalizacion,
                "cantidades_soat_hospitalizacion": cantidades_soat_hospitalizacion,
                "revision_entidad_86": revision_entidad_86,
                "revision_cantidad": revision_cantidad,
            },
            "totales": {
                "centros_de_costos": len(problemas_centros),
                "ide_contrato": len(problemas_ide_contrato),
                "cups_equivalentes": len(problemas_cups_equivalentes),
                "decimales": len(decimales),
                "tipo_identificacion_edad": len(tipo_identificacion_edad),
                "codigo_entidad_vs_afiliacion": len(entidad_afiliacion_comparison),
                "tipo_usuario": len(tipo_usuario),
                "profesionales": len(profesionales),
                "mal_capitado": len(mal_capitado),
                "cantidades_urgencias": len(cantidades_urgencias),
                "cantidades_soat_urgencias": len(cantidades_soat_urgencias),
                "cantidades_hospitalizacion": len(cantidades_hospitalizacion),
                "cantidades_soat_hospitalizacion": len(cantidades_soat_hospitalizacion),
                "revision_entidad_86": len(revision_entidad_86),
                "revision_cantidad": len(revision_cantidad),
            },
            "missing_columns": missing_columns,  # Columnas no encontradas (coincidencia exacta)
            "codigos_sin_db_ide_969": sorted(codigos_no_en_db_set) if problemas_codigos_no_en_db else [],
        }

        # Enrich errors with responsable from mapping
        if responsable_cierra:
            for problem_type, problems in resultado["problemas"].items():
                for p in problems:
                    # Skip non-dict items (e.g., decimales is list[str] in Urgencias)
                    if not isinstance(p, dict):
                        continue
                    factura = p.get("factura")
                    if factura and factura in responsable_cierra:
                        p["responsable"] = responsable_cierra[factura]
                    elif "responsable" not in p:
                        p["responsable"] = ""
        else:
            # Ensure all problems have responsable key
            for problem_type, problems in resultado["problemas"].items():
                for p in problems:
                    if not isinstance(p, dict):
                        continue
                    if "responsable" not in p:
                        p["responsable"] = ""

        return resultado, responsable_cierra
    elif area == AREA_EQUIPOS_BASICOS:
        # Equipos Básicos: usar reglas independientes configurables
        decimales = _detect_decimals(data_sheet, indices)
        doble_tipo = _detect_doble_tipo_procedimiento(data_sheet, indices)
        ruta_dup = _detect_ruta_duplicada_equipos_basicos(data_sheet, indices)
        # Nota: La validación de "convenio incorrecto" ya está incluida en _detect_profesionales_equipos_basicos
        cantidades = _detect_cantidades_anomalas_equipos_basicos(data_sheet, indices)
        tipo_id_edad = _detect_tipo_identificacion_edad(data_sheet, indices)
        
        logger.info("create_revision_sheet - Equipos Básicos, Llamando _detect_ide_contrato_odontologia")
        ide_contrato = _detect_ide_contrato_odontologia(data_sheet, indices)
        logger.info("create_revision_sheet - Equipos Básicos, IDE Contrato encontrados: %d", len(ide_contrato))
        
        # Validación profesionales (solo Equipos Básicos)
        profesionales = _detect_profesionales_equipos_basicos(data_sheet, indices)
        logger.info("create_revision_sheet - Equipos Básicos, Profesionales encontrados: %d", len(profesionales))
        
        # Regla transversal: Cód Entidad Cobrar vs Entidad Afiliación
        entidad_afiliacion_comparison = detect_codigo_entidad_vs_entidad_afiliacion(
            data_sheet, indices, limit_log=5
        )
        
        # Validación centro de costo (solo EQUIPOS BASICOS ODONTOLOGIA)
        centro_costo = _detect_centro_costo_odontologia(
            data_sheet, 
            indices, 
            profesional_dias=profesional_dias,
            permitir_todos_centros=permitir_todos_centros,
            centros_validos=[CENTRO_COSTO_EQUIPOS_BASICOS],
        )
        
        # Build responsable_cierra mapping
        responsable_cierra = {}
        responsable_cierra_idx = indices.get("responsable_cierra")
        num_fact_idx = indices.get("numero_factura")
        if responsable_cierra_idx is not None and num_fact_idx is not None:
            for row in range(2, data_sheet.max_row + 1):
                numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
                factura = _normalize_invoice(numero)
                if not factura:
                    continue
                raw = data_sheet.cell(row=row, column=responsable_cierra_idx + 1).value
                resp = str(raw).strip() if raw else ""
                if resp and factura not in responsable_cierra:
                    responsable_cierra[factura] = resp
        
        # Build normalized rows for unified 6-column display
        normalized_rows_eb = _build_odontologia_normalized_rows(
            decimales=decimales,
            doble_tipo=doble_tipo,
            ruta_dup=ruta_dup,
            profesionales=profesionales,
            cantidades=cantidades,
            tipo_id_edad=tipo_id_edad,
            centro_costo=centro_costo,
            ide_contrato=ide_contrato,
            responsable_cierra=responsable_cierra,
        )

        resultado = {
            "area": area,
            "problemas": {
                "normalizados": normalized_rows_eb,
                "decimales": decimales,
                "doble_tipo_procedimiento": doble_tipo,
                "ruta_duplicada": ruta_dup,
                "cantidades_anomalas": cantidades,
                "tipo_identificacion_edad": tipo_id_edad,
                "centro_costo": centro_costo,
                "ide_contrato": ide_contrato,
                "profesionales": profesionales,
            },
            "totales": {
                "decimales": len(decimales),
                "doble_tipo_procedimiento": len(doble_tipo),
                "ruta_duplicada": len(ruta_dup),
                "cantidades_anomalas": len(cantidades),
                "tipo_identificacion_edad": len(tipo_id_edad),
                "centro_costo": len(centro_costo),
                "ide_contrato": len(ide_contrato),
                "profesionales": len(profesionales),
                "codigo_entidad_vs_afiliacion": len(entidad_afiliacion_comparison),
            },
            "es_equipos_basicos": True,
            "missing_columns": missing_columns,
        }
        
        # Enrich errors with responsable from mapping
        if responsable_cierra:
            for problem_type, problems in resultado["problemas"].items():
                for p in problems:
                    factura = p.get("factura")
                    if factura and factura in responsable_cierra:
                        p["responsable"] = responsable_cierra[factura]
                    elif "responsable" not in p:
                        p["responsable"] = ""
        else:
            # Ensure all problems have responsable key
            for problem_type, problems in resultado["problemas"].items():
                for p in problems:
                    if "responsable" not in p:
                        p["responsable"] = ""
        
        return resultado, responsable_cierra
    else:
# Odontología estándar: todas las validaciones
        # NOTA: Ahora delega en app/services/odontologia/detect_all.py.
        # Se eliminará en Fase 7 cuando se unifique el flujo.
        resultado, responsable_cierra = detect_all_problems_odontologia(
            data_sheet,
            indices,
            profesional_dias=profesional_dias,
            permitir_todos_centros=permitir_todos_centros,
        )
        resultado["missing_columns"] = missing_columns
        return resultado, responsable_cierra
