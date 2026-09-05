"""Orquestador de detección de problemas para Hospitalización.

Agrupa detectores transversales + específicos de Hospitalización.
"""

from __future__ import annotations

import logging
from typing import Any, Callable

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import AREA_HOSPITALIZACION
from app.constants.base import is_evidence_audit_enabled, is_rule_engine_enabled
from app.services.transversales import (
    normalize_invoice,
)
from app.services.normalized_rows import build_normalized_rows

# Module-level flag: skip evidence/audit DB writes when testing
_PERSIST = is_evidence_audit_enabled()
logger = logging.getLogger(__name__)


def _get_hospitalizacion_detectors() -> list[Callable]:
    """Returns list of Hospitalización-specific detector callables.
    
    Used by tipo_factura_registry for lazy import.
    """
    from app.services.hospitalizacion.cantidades_hospitalizacion import (
        detect_cantidades_hospitalizacion,
    )
    from app.services.hospitalizacion.hospitalizacion_codes import (
        detect_hospitalizacion_codes,
    )
    from app.services.hospitalizacion.centro_costo_hospitalizacion import (
        detect_centro_costo_hospitalizacion,
    )
    from app.services.hospitalizacion.cantidades_soat_hospitalizacion import (
        detect_cantidades_soat_hospitalizacion,
    )
    from app.services.urgencias.ide_contrato_urgencias import (
        detect_ide_contrato_urgencias,
    )
    from app.services.transversales.detect_copago_entidad import (
        detect_copago_entidad_urgencias,
    )

    return [
        detect_centro_costo_hospitalizacion,
        detect_ide_contrato_urgencias,
        detect_cantidades_hospitalizacion,
        detect_cantidades_soat_hospitalizacion,
        detect_hospitalizacion_codes,
        detect_copago_entidad_urgencias,
    ]


def detect_all_problems_hospitalizacion(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> tuple[dict[str, Any], dict[str, str]]:
    """Detecta TODOS los problemas en facturas de Hospitalización.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        (resultado_dict, responsables_map)
    """
    from app.services.transversales import (
        detect_decimales,
        detect_tipo_documento_edad,
        detect_tipo_identificacion_entidad,
        detect_codigo_entidad_vs_entidad_afiliacion,
        detect_tipo_usuario,
    )
    from app.services.hospitalizacion.cantidades_hospitalizacion import (
        detect_cantidades_hospitalizacion,
    )
    from app.services.hospitalizacion.hospitalizacion_codes import (
        detect_hospitalizacion_codes,
    )
    from app.services.hospitalizacion.centro_costo_hospitalizacion import (
        detect_centro_costo_hospitalizacion,
    )
    from app.services.hospitalizacion.cantidades_soat_hospitalizacion import (
        detect_cantidades_soat_hospitalizacion,
    )
    from app.services.urgencias.ide_contrato_urgencias import (
        detect_ide_contrato_urgencias,
    )
    from app.services.urgencias.profesionales_urgencias import detect_profesionales_urgencias
    from app.services.transversales.detect_copago_entidad import (
        detect_copago_entidad_urgencias,
    )
    from app.services.transversales.procedimiento_contratado import detect_cups_sin_contrato

    # 1. Centro Costo + IDE Contrato
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            problemas_centros = RuleBasedDetector("centro_costo_hospitalizacion_valido", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    problemas_ide_contrato = detect_ide_contrato_urgencias(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            problemas_ide_contrato = RuleBasedDetector("ide_contrato_hospitalizacion_valido", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()

    # 2. Cups Equivalentes (Hospitalización codes — group rules via engine)
    problemas_cups_equivalentes: list[dict[str, str]] = []
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            problemas_cups_equivalentes.extend(
                RuleBasedDetector("hosp_codigos_oblig_mayor24h", session).detect(data_sheet, indices, persist=_PERSIST)
            )
            problemas_cups_equivalentes.extend(
                RuleBasedDetector("hosp_codigos_oblig_menor24h", session).detect(data_sheet, indices, persist=_PERSIST)
            )
            problemas_cups_equivalentes.extend(
                RuleBasedDetector("hosp_codigos_prohibidos", session).detect(data_sheet, indices, persist=_PERSIST)
            )
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    else:
        problemas_cups_equivalentes.extend(detect_hospitalizacion_codes(data_sheet, indices))

    # 3. Detectores transversales (con toggle engine)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            decimales = RuleBasedDetector("valores_decimales", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    else:
        decimales = []
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            r1 = RuleBasedDetector("tipo_documento_edad_menor_7", session).detect(data_sheet, indices, persist=_PERSIST)
            r2 = RuleBasedDetector("tipo_documento_edad_mayor_18", session).detect(data_sheet, indices, persist=_PERSIST)
            r3 = RuleBasedDetector("tipo_documento_edad_7_17", session).detect(data_sheet, indices, persist=_PERSIST)
            r4 = RuleBasedDetector("tipo_documento_edad_as_menor", session).detect(data_sheet, indices, persist=_PERSIST)
            r5 = RuleBasedDetector("tipo_documento_edad_ms_mayor", session).detect(data_sheet, indices, persist=_PERSIST)
            r6 = RuleBasedDetector("tipo_documento_edad_cn_invalido", session).detect(data_sheet, indices, persist=_PERSIST)
            r7 = RuleBasedDetector("tipo_documento_edad_ce_invalido", session).detect(data_sheet, indices, persist=_PERSIST)
            tipo_identificacion_edad = r1 + r2 + r3 + r4 + r5 + r6 + r7
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    else:
        tipo_identificacion_edad = []
    tipo_identificacion_entidad = detect_tipo_identificacion_entidad(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            r1 = RuleBasedDetector("tipo_id_requiere_entidad_86000", session).detect(data_sheet, indices, persist=_PERSIST)
            r2 = RuleBasedDetector("entidad_86000_requiere_as_ms", session).detect(data_sheet, indices, persist=_PERSIST)
            tipo_identificacion_entidad = r1 + r2
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    entidad_afiliacion_comparison = detect_codigo_entidad_vs_entidad_afiliacion(
        data_sheet, indices, limit_log=5
    )
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            entidad_afiliacion_comparison = RuleBasedDetector("codigo_entidad", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    tipo_usuario = detect_tipo_usuario(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            tipo_usuario = RuleBasedDetector("tipo_usuario_valido", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()

    # 4. Detectores específicos de Hospitalización
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            cantidades_hospitalizacion = RuleBasedDetector("cantidades_hospitalizacion", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    else:
        cantidades_hospitalizacion = []
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            cantidades_soat_hospitalizacion = RuleBasedDetector("cantidades_soat_hospitalizacion", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    else:
        cantidades_soat_hospitalizacion = []
    copago_entidad = detect_copago_entidad_urgencias(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            copago_entidad = RuleBasedDetector("copago_entidad_valido", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()

    profesionales = detect_profesionales_urgencias(data_sheet, indices, tipos_validos={"Hospitalización"})
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            profesionales = RuleBasedDetector("profesional_hospitalizacion_valido", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    logger.info(
        "detect_all_problems_hospitalizacion - Profesionales encontrados: %d",
        len(profesionales),
    )

    cups_sin_contrato = detect_cups_sin_contrato(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            cups_sin_contrato = RuleBasedDetector("cups_sin_contrato", session).detect(data_sheet, indices, persist=_PERSIST)
            if _PERSIST:
                session.commit()
            else:
                session.rollback()
        finally:
            session.close()
    logger.info(
        "detect_all_problems_hospitalizacion - Cups Sin Contrato encontrados: %d",
        len(cups_sin_contrato),
    )

    # 5. Filtrar centros de costo por prioridad
    errores_por_factura_codigo: dict[tuple[str, str], list[tuple[dict, int]]] = {}
    for item in problemas_centros:
        key = (item.get("factura", ""), item.get("codigo", ""))
        prioridad = item.get("prioridad", 1)
        if key not in errores_por_factura_codigo:
            errores_por_factura_codigo[key] = []
        errores_por_factura_codigo[key].append((item, prioridad))

    problemas_centros_filtrados = []
    for key, items in errores_por_factura_codigo.items():
        prioridades = [p for _, p in items]
        if 1 in prioridades:
            for item, p in items:
                if p == 1:
                    problemas_centros_filtrados.append(item)
        else:
            for item, _ in items:
                problemas_centros_filtrados.append(item)

    # 6. Build responsable_cierra mapping
    responsable_cierra: dict[str, str] = {}
    responsable_cierra_idx = indices.get("responsable_cierra")
    num_fact_idx = indices.get("numero_factura")
    if responsable_cierra_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            raw = data_sheet.cell(row=row, column=responsable_cierra_idx + 1).value
            resp = str(raw).strip() if raw else ""
            if resp and factura not in responsable_cierra:
                responsable_cierra[factura] = resp

    # 7. Build fecha_cierre_vacia mapping
    fecha_cierre_vacia: dict[str, bool] = {}
    fecha_cierre_idx = indices.get("fecha_cierre")
    if fecha_cierre_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            fecha_cierre_val = data_sheet.cell(row=row, column=fecha_cierre_idx + 1).value
            if not fecha_cierre_val or str(fecha_cierre_val).strip() == "":
                fecha_cierre_vacia[factura] = True
            elif factura not in fecha_cierre_vacia:
                fecha_cierre_vacia[factura] = False

    # 8. Build fec_factura_map
    fec_factura_map: dict[str, str] = {}
    fec_factura_idx = indices.get("fec_factura")
    if fec_factura_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            raw = data_sheet.cell(row=row, column=fec_factura_idx + 1).value
            val = str(raw).strip() if raw else ""
            if val and factura not in fec_factura_map:
                fec_factura_map[factura] = val

    # 9. Build normalized rows
    error_groups = {
        "Centros de Costo": problemas_centros_filtrados,
        "IDE Contrato": problemas_ide_contrato,
        "Cups Equivalentes": problemas_cups_equivalentes,
        "Cantidades Hospitalización": cantidades_hospitalizacion,
        "Cantidades SOAT Hospitalización": cantidades_soat_hospitalizacion,
        "Decimales": decimales,
        "Tipo Identificación / Edad": tipo_identificacion_edad,
        "Código Entidad vs Afiliación": entidad_afiliacion_comparison + tipo_identificacion_entidad,
        "Tipo Usuario": tipo_usuario,
        "Copago vs Entidad": copago_entidad,
        "Profesionales": profesionales,
        "Cups Sin Contrato": cups_sin_contrato,
    }
    normalized_rows = build_normalized_rows(
        error_groups=error_groups,
        responsables_map=responsable_cierra,
        fec_factura_map=fec_factura_map,
        fecha_cierre_vacia_map=fecha_cierre_vacia,
    )

    # 10. Build resultado
    resultado: dict[str, Any] = {
        "area": AREA_HOSPITALIZACION,
        "problemas": {
            "normalizados": normalized_rows,
            "centros_de_costos": [
                {
                    "tipo_factura": item.get("tipo_factura") or "-",
                    "factura": item["factura"],
                    "codigo": item.get("codigo", ""),
                    "procedimiento": item.get("procedimiento", ""),
                    "centro_actual": item.get("centro_actual", ""),
                    "centro_deberia": item.get("centro_deberia", ""),
                    "prioridad": item.get("prioridad", 1),
                }
                for item in problemas_centros_filtrados
            ],
            "ide_contrato": [
                {
                    "factura": item["factura"],
                    "ide_contrato_actual": item.get("ide_contrato_actual", item.get("ide_contrato", "")),
                    "ide_contrato_deberia": item.get("ide_contrato_deberia", ""),
                    "procedimiento": item.get("procedimiento", ""),
                    "codigo": item.get("codigo", ""),
                    "entidad": item.get("entidad", ""),
                    "nota": item.get("nota", ""),
                }
                for item in problemas_ide_contrato
            ],
            "cups_equivalentes": [
                {
                    "factura": item["factura"],
                    "codigo": item["codigo"],
                    "codigo_equiv": item["codigo_equiv"],
                    "accion": item["accion"],
                }
                for item in problemas_cups_equivalentes
            ],
            "decimales": decimales,
            "tipo_identificacion_edad": tipo_identificacion_edad,
            "tipo_identificacion_entidad": tipo_identificacion_entidad,
            "codigo_entidad_vs_afiliacion": entidad_afiliacion_comparison,
            "tipo_usuario": tipo_usuario,
            "cantidades_hospitalizacion": cantidades_hospitalizacion,
            "cantidades_soat_hospitalizacion": cantidades_soat_hospitalizacion,
            "copago_entidad": copago_entidad,
            "profesionales": profesionales,
            "cups_sin_contrato": cups_sin_contrato,
        },
        "totales": {
            "centros_de_costos": len(problemas_centros),
            "ide_contrato": len(problemas_ide_contrato),
            "cups_equivalentes": len(problemas_cups_equivalentes),
            "decimales": len(decimales),
            "tipo_identificacion_edad": len(tipo_identificacion_edad),
            "tipo_identificacion_entidad": len(tipo_identificacion_entidad),
            "codigo_entidad_vs_afiliacion": len(entidad_afiliacion_comparison),
            "tipo_usuario": len(tipo_usuario),
            "cantidades_hospitalizacion": len(cantidades_hospitalizacion),
            "cantidades_soat_hospitalizacion": len(cantidades_soat_hospitalizacion),
            "copago_entidad": len(copago_entidad),
            "profesionales": len(profesionales),
            "cups_sin_contrato": len(cups_sin_contrato),
        },
        "missing_columns": [],
    }

    # 11. Enrich errors with responsable
    for problem_type, problems in resultado["problemas"].items():
        for p in problems:
            if not isinstance(p, dict):
                continue
            factura = p.get("factura")
            if factura and factura in responsable_cierra:
                p["responsable"] = responsable_cierra[factura]
            elif "responsable" not in p:
                p["responsable"] = ""

    return resultado, responsable_cierra
