"""Detector de códigos obligatorios/prohibidos en Hospitalización.

Extraído de urgencias/hospitalizacion.py como parte de la reorganización por tipo_factura.
"""

from __future__ import annotations

import logging
from datetime import datetime
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    CODIGOS_HOSPITALIZACION_OBLIGATORIOS,
    CODIGOS_HOSPITALIZACION_PROHIBIDOS,
    CODIGOS_SOAT_HOSPITALIZACION_OBLIGATORIOS,
    CODIGOS_SOAT_HOSPITALIZACION_PROHIBIDOS,
    VALOR_TARIFARIO_SOAT,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def _format_estancia_hosp(horas: float | None) -> str:
    """Formatea estancia en días + horas."""
    if horas is None:
        return "N/A"
    dias = int(horas // 24)
    hrs = int(horas % 24)
    if dias > 0:
        return f"{dias}d {hrs}h"
    return f"{hrs}h"


def detect_hospitalizacion_codes(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """Detecta problemas de códigos obligatorios/prohibidos en Hospitalización."""
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    codigo_entidad_cobrar_idx = indices.get("codigo_entidad_cobrar")
    tipo_factura_descripcion_idx = indices.get("tipo_factura_descripcion")
    proc_idx = indices.get("procedimiento")
    tarifario_idx = indices.get("tarifario")
    fec_factura_idx = indices.get("fec_factura")
    fecha_cierre_idx = indices.get("fecha_cierre")

    if num_fact_idx is None or tipo_factura_descripcion_idx is None:
        logger.warning("Hospitalización Códigos - Columnas necesarias no encontradas")
        return []

    factura_hospitalizacion_data: dict[str, dict[str, Any]] = {}

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        tipo_factura_cell = data_sheet.cell(row=row, column=tipo_factura_descripcion_idx + 1).value
        tipo_factura_str = str(tipo_factura_cell).strip() if tipo_factura_cell else ""

        if tipo_factura_str != "Hospitalización":
            continue

        codigo_cell = data_sheet.cell(row=row, column=codigo_idx + 1).value if codigo_idx else None
        codigo_normalized = str(codigo_cell).strip() if codigo_cell else ""

        estancia_horas_hosp = None
        if fec_factura_idx is not None and fecha_cierre_idx is not None:
            fec_factura_cell_hosp = data_sheet.cell(row=row, column=fec_factura_idx + 1).value
            fecha_cierre_cell_hosp = data_sheet.cell(row=row, column=fecha_cierre_idx + 1).value
            if fec_factura_cell_hosp and fecha_cierre_cell_hosp:
                try:
                    fec_factura_dt_hosp = datetime.strptime(str(fec_factura_cell_hosp).strip(), "%Y-%m-%d %H:%M:%S")
                    fecha_cierre_dt_hosp = datetime.strptime(str(fecha_cierre_cell_hosp).strip(), "%Y-%m-%d %H:%M:%S")
                    diferencia_hosp = fecha_cierre_dt_hosp - fec_factura_dt_hosp
                    estancia_horas_hosp = diferencia_hosp.total_seconds() / 3600
                except (ValueError, TypeError):
                    estancia_horas_hosp = None

        entidad_hosp = ""
        if codigo_entidad_cobrar_idx is not None:
            entidad_cell_hosp = data_sheet.cell(row=row, column=codigo_entidad_cobrar_idx + 1).value
            entidad_hosp = str(entidad_cell_hosp).strip() if entidad_cell_hosp else ""

        tarifario_hosp = ""
        if tarifario_idx is not None:
            tarifario_cell_hosp = data_sheet.cell(row=row, column=tarifario_idx + 1).value
            tarifario_hosp = str(tarifario_cell_hosp).strip() if tarifario_cell_hosp else ""

        if factura_str not in factura_hospitalizacion_data:
            factura_hospitalizacion_data[factura_str] = {
                "entidad": entidad_hosp,
                "tarifario": tarifario_hosp,
                "codigos_hospitalizacion": set(),
                "codigos_prohibidos": set(),
                "estancia_horas": estancia_horas_hosp,
                "tiene_soat_hosp_prohibido": False,
                "codigos_soat_hosp_obligatorios": set(),
            }
        else:
            if estancia_horas_hosp is not None:
                previa = factura_hospitalizacion_data[factura_str].get("estancia_horas")
                if previa is None or estancia_horas_hosp > previa:
                    factura_hospitalizacion_data[factura_str]["estancia_horas"] = estancia_horas_hosp
            if not factura_hospitalizacion_data[factura_str].get("entidad") and entidad_hosp:
                factura_hospitalizacion_data[factura_str]["entidad"] = entidad_hosp
            if tarifario_hosp and tarifario_hosp.upper() == VALOR_TARIFARIO_SOAT:
                factura_hospitalizacion_data[factura_str]["tarifario"] = tarifario_hosp
            elif not factura_hospitalizacion_data[factura_str].get("tarifario") and tarifario_hosp:
                factura_hospitalizacion_data[factura_str]["tarifario"] = tarifario_hosp

        if codigo_normalized in CODIGOS_HOSPITALIZACION_OBLIGATORIOS:
            factura_hospitalizacion_data[factura_str]["codigos_hospitalizacion"].add(codigo_normalized)
        if codigo_normalized in CODIGOS_HOSPITALIZACION_PROHIBIDOS:
            factura_hospitalizacion_data[factura_str]["codigos_prohibidos"].add(codigo_normalized)
        if tarifario_hosp.upper() == VALOR_TARIFARIO_SOAT and codigo_normalized in CODIGOS_SOAT_HOSPITALIZACION_PROHIBIDOS:
            factura_hospitalizacion_data[factura_str]["tiene_soat_hosp_prohibido"] = True
        if tarifario_hosp.upper() == VALOR_TARIFARIO_SOAT and codigo_normalized in CODIGOS_SOAT_HOSPITALIZACION_OBLIGATORIOS:
            factura_hospitalizacion_data[factura_str]["codigos_soat_hosp_obligatorios"].add(codigo_normalized)

    problemas_cups_equivalentes: list[dict[str, Any]] = []
    facturas_hosp_reportadas: set[str] = set()

    for factura_str, data in factura_hospitalizacion_data.items():
        codigos_hosp = data.get("codigos_hospitalizacion", set())
        codigos_prohibidos = data.get("codigos_prohibidos", set())
        estancia_horas = data.get("estancia_horas")
        tarifario = data.get("tarifario", "")

        if tarifario.upper() == VALOR_TARIFARIO_SOAT:
            codigos_soat_hosp_obligatorios = data.get("codigos_soat_hosp_obligatorios", set())
            faltan_soat_oblig = CODIGOS_SOAT_HOSPITALIZACION_OBLIGATORIOS - codigos_soat_hosp_obligatorios
            if faltan_soat_oblig:
                estancia_str = _format_estancia_hosp(estancia_horas) if estancia_horas is not None else "N/A"
                problemas_cups_equivalentes.append({
                    "factura": factura_str,
                    "codigo": list(codigos_soat_hosp_obligatorios),
                    "codigo_equiv": "",
                    "accion": f"SOAT Hospitalización debe tener: {', '.join(sorted(faltan_soat_oblig))}",
                    "procedimiento": "",
                    "estancia_str": estancia_str,
                })
                facturas_hosp_reportadas.add(factura_str)

            tiene_soat_hosp_prohibido = data.get("tiene_soat_hosp_prohibido", False)
            if tiene_soat_hosp_prohibido:
                estancia_str = _format_estancia_hosp(estancia_horas) if estancia_horas is not None else "N/A"
                problemas_cups_equivalentes.append({
                    "factura": factura_str,
                    "codigo": list(codigos_prohibidos),
                    "codigo_equiv": "",
                    "accion": f"SOAT Hospitalización no puede tener: {', '.join(CODIGOS_SOAT_HOSPITALIZACION_PROHIBIDOS)}",
                    "procedimiento": "",
                    "estancia_str": estancia_str,
                })
                facturas_hosp_reportadas.add(factura_str)
            continue

        if estancia_horas is None:
            continue

        if estancia_horas > 24:
            codigos_obligatorios_hosp = CODIGOS_HOSPITALIZACION_OBLIGATORIOS
        else:
            codigos_obligatorios_hosp = {"890601H", "129B02"}

        estancia_str = _format_estancia_hosp(estancia_horas)

        faltan = codigos_obligatorios_hosp - codigos_hosp
        if faltan:
            problemas_cups_equivalentes.append({
                "factura": factura_str,
                "codigo": list(codigos_hosp),
                "codigo_equiv": "",
                "accion": (
                    f"Hospitalización ({'>24h' if estancia_horas > 24 else '<=24h'}) "
                    f"debe tener: {', '.join(sorted(codigos_obligatorios_hosp))} "
                    f"(faltan: {', '.join(faltan)})"
                ),
                "procedimiento": "",
                "estancia_str": estancia_str,
            })
            facturas_hosp_reportadas.add(factura_str)

        if codigos_prohibidos:
            problemas_cups_equivalentes.append({
                "factura": factura_str,
                "codigo": list(codigos_prohibidos),
                "codigo_equiv": "",
                "accion": f"Hospitalización no puede tener: {', '.join(codigos_prohibidos)}",
                "procedimiento": "",
                "estancia_str": estancia_str,
            })
            facturas_hosp_reportadas.add(factura_str)

    if problemas_cups_equivalentes:
        logger.info("Hospitalización Códigos - Problemas encontrados: %d", len(problemas_cups_equivalentes))

    return problemas_cups_equivalentes
