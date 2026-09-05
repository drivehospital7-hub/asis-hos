"""Export del Listado de Exámenes a Excel (.xlsx) en memoria.

Espeja ``control_errores_export.py`` (estilo control-novedades: header verde
oscuro 1B5E20 con fuente blanca bold, columnas de ancho 20, bordes finos,
rellenos alternados E8F5E9/blanco, freeze A2) y porta las semánticas de filtro
de ``composeListadoView`` (lib/examenes.ts): rango de fechas excluye sin-fecha;
``q`` no vacío busca sobre el listado COMPLETO ignorando el rango; orden
date-desc con hora desc dentro del día y sin-fecha al final.

Una fila por ÍTEM con N° secuencial; columnas = ``CSV_HEADERS`` (11). El tc usa
la variante CSV (X→SI/AUTH→AUTH/vacío), NO la de impresión (``tcDisplay``).
"""

from __future__ import annotations

import logging
import re
import unicodedata
from datetime import date
from functools import cmp_to_key
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.constants.examenes import CSV_HEADERS

logger = logging.getLogger(__name__)

HEADERS: list[str] = CSV_HEADERS

COLUMN_WIDTH = 20

HEADER_FILL = PatternFill("solid", fgColor="1B5E20")
HEADER_FONT = Font(bold=True, color="FFFFFF")

ROW_FILL_LIGHT = PatternFill("solid", fgColor="E8F5E9")
ROW_FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
DATA_FONT = Font(color="000000")

THIN_BORDER = Border(
    left=Side(style="thin", color="A5D6A7"),
    right=Side(style="thin", color="A5D6A7"),
    top=Side(style="thin", color="A5D6A7"),
    bottom=Side(style="thin", color="A5D6A7"),
)

# dd/mm/yyyy con hora opcional (paridad `listadoFechaInfo`, lib/examenes.ts).
_FECHA_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})(?:\s|$)")
_HORA_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{1,2})")


# ─── Filtros (paridad composeListadoView, lib/examenes.ts) ──────────────────


def _fold(value: str | None) -> str:
    """Normaliza búsqueda: NFD + quita marcas combinantes (preservando Ñ/ñ) + trim + UPPER.

    Espeja ``normalizeListadoQuery``: "Álvaro" ≈ "alvaro", pero "MUÑOZ" sigue
    distinguible de "MUNOZ" (Ñ preservada).
    """
    raw = str(value or "")
    _PH_U = "\uE000"
    _PH_L = "\uE001"
    raw = raw.replace("Ñ", _PH_U).replace("ñ", _PH_L)
    folded = "".join(
        ch for ch in unicodedata.normalize("NFD", raw)
        if not unicodedata.combining(ch)
    )
    folded = folded.replace(_PH_U, "Ñ").replace(_PH_L, "ñ")
    return folded.strip().upper()


def _day_key(hora: str | None) -> str:
    """Clave ISO ``yyyy-mm-dd`` de una hora es-CO; inválida → ``sin-fecha``.

    Validación round-trip igual que ``listadoFechaInfo``: fechas imposibles
    (31/02) y formatos no parseables caen en ``sin-fecha`` (nunca se dropean).
    """
    match = _FECHA_RE.match(str(hora or ""))
    if match:
        day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        try:
            date(year, month, day)
        except ValueError:
            return "sin-fecha"
        return f"{year:04d}-{month:02d}-{day:02d}"
    return "sin-fecha"


def _hora_minutes(hora: str | None) -> int:
    """Minutos desde medianoche de ``dd/mm/yyyy hh:mm``; sin hora → 0."""
    match = _HORA_RE.match(str(hora or ""))
    if not match:
        return 0
    return int(match.group(4)) * 60 + int(match.group(5))


def _cmp(a: dict, b: dict) -> int:
    """Comparador desc: día desc, hora desc dentro del día, sin-fecha al final.

    Port exacto del comparador de ``sortByDateDesc`` (estable en empates).
    """
    ka = _day_key(a.get("hora", ""))
    kb = _day_key(b.get("hora", ""))
    if ka == "sin-fecha" or kb == "sin-fecha":
        if ka == kb:
            return 0
        return 1 if ka == "sin-fecha" else -1
    if ka != kb:
        return 1 if ka < kb else -1
    return _hora_minutes(b.get("hora", "")) - _hora_minutes(a.get("hora", ""))


def _in_range(pf: dict, from_: str | None, to: str | None) -> bool:
    """Rango inclusivo [from_, to] en ISO; sin-fecha NUNCA está en rango (A5)."""
    day_key = _day_key(pf.get("hora", ""))
    if day_key == "sin-fecha":
        return False
    if from_ is not None and day_key < from_:
        return False
    if to is not None and day_key > to:
        return False
    return True


def _search_listado(listado: list[dict], query: str) -> list[dict]:
    """Substring plegado sobre paciente | cedula | facturador | items (EX-29).

    ``q`` vacío → el listado de entrada sin cambios (misma referencia).
    """
    q = _fold(query)
    if not q:
        return listado
    matches = []
    for pf in listado:
        haystack = [
            _fold(pf.get("paciente", "")),
            _fold(pf.get("cedula", "")),
            _fold(pf.get("facturador", "")),
        ]
        for item in pf.get("items", []):
            haystack.append(_fold(item.get("cod", "")))
            haystack.append(_fold(item.get("nom", "")))
        if any(q in h for h in haystack):
            matches.append(pf)
    return matches


def compose_listado_view(
    listado: list[dict], from_: str | None, to: str | None, query: str
) -> list[dict]:
    """Conjunto exhibido con las mismas reglas que ``composeListadoView`` (D4).

    Rango activo (cualquier extremo no-None) → excluye sin-fecha; ``q`` no
    vacío → búsqueda global sobre el listado COMPLETO ignorando el rango;
    resultado ordenado date-desc (hora desc dentro del día, sin-fecha al
    final). Sin rango y sin ``q`` → todo el listado ordenado.
    """
    ranged = (
        [pf for pf in listado if _in_range(pf, from_, to)]
        if from_ is not None or to is not None
        else listado
    )
    displayed = _search_listado(listado, query) if _fold(query) else ranged
    return sorted(displayed, key=cmp_to_key(_cmp))


# ─── Builder del workbook (paridad control_errores_export.py) ───────────────


def _cantidad(item: dict) -> int:
    """Cantidad normalizada (EX-27): ausente/NaN/<1 → 1; fracciones truncadas."""
    try:
        q = int(float(item.get("cantidad")))
    except (TypeError, ValueError, OverflowError):
        q = 0
    return q if q >= 1 else 1


def _tc(value: str | None) -> str:
    """Variante CSV del mapeo (X→SI, AUTH→AUTH, resto vacío) — EX-14, D5."""
    if value == "X":
        return "SI"
    if value == "AUTH":
        return "AUTH"
    return ""


def _sanitize_for_excel(value):
    """OWASP formula-injection mitigation: prefix =, +, -, @ with single quote.

    openpyxl stores strings starting with '=' as live formulas (data_type 'f');
    prefixing with "'" forces text (data_type 's') and Excel shows the literal.
    """
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def build_listado_export_workbook(prefacturas: list[dict]) -> BytesIO:
    """Construye el workbook .xlsx en memoria con las prefacturas dadas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTH

    n = 0
    for row_index, pf in enumerate(prefacturas):
        for item in pf.get("items", []):
            n += 1
            ws.append([
                n,
                _sanitize_for_excel(pf.get("paciente", "")),
                _sanitize_for_excel(pf.get("cedula", "")),
                _sanitize_for_excel(item.get("cod", "")),
                _sanitize_for_excel(item.get("nom", "")),
                _cantidad(item),
                _tc(item.get("neps", "")),
                _tc(item.get("mall", "")),
                _tc(item.get("emss", "")),
                _sanitize_for_excel(pf.get("facturador", "")),
                _sanitize_for_excel(pf.get("hora", "")),
            ])
            row_idx = ws.max_row
            row_fill = ROW_FILL_LIGHT if row_index % 2 == 0 else ROW_FILL_WHITE
            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = row_fill
                cell.border = THIN_BORDER
                cell.font = DATA_FONT

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def filename_export(from_: str | None, to: str | None) -> str:
    """Nombre del archivo: ``Listado_Lab_HospitalOrito_{label}.xlsx`` (EX-34).

    Label = ``{from}_{to}`` (solo ``from`` → ``{from}_hasta``; solo ``to`` →
    etiqueta que empieza con "_", igual que el frontend); sin rango →
    ``Todos_los_meses``.
    """
    label = f"{from_ or ''}_{to or 'hasta'}" if (from_ or to) else "Todos_los_meses"
    return f"Listado_Lab_HospitalOrito_{label}.xlsx"