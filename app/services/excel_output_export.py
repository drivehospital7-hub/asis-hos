from __future__ import annotations

import logging
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

from app.services.excel_column_headers import ALLOWED_EXCEL_SUFFIXES
from app.utils.input_data import (
    resolve_safe_excel_in_input,
    resolve_safe_excel_in_output,
)

logger = logging.getLogger(__name__)

CRUCE_FACTURAS_SHEET = "CruceFacturas"

# Columns to keep in the export (hide all others)
COLUMNS_TO_KEEP = {
    "Entidad Cobrar",
    "Profesional Atiende",
    "Fec. Factura",
    "Fecha Cierre",
    "Número Factura",
    "Tipo Entidad Cobrar",
    "Convenio Facturado",
    "Procedimiento",
    "Tipo Identificación",
    "Edad Completa",
    "Nº Identificación",
    "Primer Apellido",
    "Responsable Cierra Facturar",
    "Vlr. Procedimiento",
    "Vlr. Subsidiado",
    "Cantidad",
    "Segundo Apellido",
    "Primer Nombre",
    "Segundo Nombre",
    "Sexo",
    "Fec. Nacimiento",
    "Cita",
    "Tipo Cita",
    "Centro Costo"
}


def _validate_input_path(path: Path) -> str | None:
    if not path.is_file():
        return f"Archivo no encontrado: {path.name}"
    if path.suffix.lower() not in ALLOWED_EXCEL_SUFFIXES:
        return f"Formato no soportado: {path.suffix.lower()}"
    return None


def _get_or_create_sheet(workbook: Workbook, title: str) -> Worksheet:
    if title in workbook.sheetnames:
        return workbook[title]
    return workbook.create_sheet(title=title)


def _find_numero_factura_column(sheet: Worksheet) -> str | None:
    """Find the column letter for 'Número Factura' in the first row."""
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value == "Número Factura":
            return sheet.cell(row=1, column=col).column_letter
    return None


def _apply_conditional_formulas_to_cruce_facturas(
    workbook: Workbook, cruce_sheet: Worksheet, data_sheet: Worksheet
) -> None:
    """
    Apply conditional formatting to CruceFacturas sheet and data sheet.
    The formatting will highlight cells when invoice numbers match the headers.
    """
    numero_factura_col = _find_numero_factura_column(data_sheet)
    if not numero_factura_col:
        logger.warning("No se encontró columna 'Número Factura' en la hoja de datos")
        return
    
    # Get max row in data sheet
    max_data_row = data_sheet.max_row
    
    # Define column configurations for CruceFacturas sheet
    # Each tuple: (column_letter, header_cell, color_rgb)
    columns_config = [
        ("B", "B1", "92D050"),  # Green
        ("D", "D1", "FFC000"),  # Yellow
        ("F", "F1", "FF0000"),  # Red
    ]
    
    for col_letter, header_cell, color_rgb in columns_config:
        # Apply conditional formatting with fill color
        # Formula checks if the invoice number in the header matches any value in the data column
        fill = PatternFill(start_color=color_rgb, end_color=color_rgb, fill_type="solid")
        rule = FormulaRule(
            formula=[f"COUNTIF({data_sheet.title}!${numero_factura_col}$2:${numero_factura_col}${max_data_row},{cruce_sheet.title}!{header_cell})>0"],
            fill=fill
        )
        
        # Apply to the range in the cruce sheet (just the header cell for now, but can be extended)
        cruce_range = f"{col_letter}1:{col_letter}{max_data_row + 50}"
        cruce_sheet.conditional_formatting.add(cruce_range, rule)
        
        logger.info(f"Conditional formatting applied to {cruce_range} with color {color_rgb}")
    
    # Also apply conditional formatting to the data sheet's numero_factura column
    # for each of the three header values
    fill_configs = [
        ("B", "B1", "92D050"),  # Green for B column
        ("D", "D1", "FFC000"),  # Yellow for D column
        ("F", "F1", "FF0000"),  # Red for F column
    ]
    
    for col_letter, header_cell, color_rgb in fill_configs:
        fill = PatternFill(start_color=color_rgb, end_color=color_rgb, fill_type="solid")
        # Formula that checks if the current cell appears anywhere in the CruceFacturas column
        # Uses COUNTIF with entire column range (e.g., B:B) to avoid scope errors
        rule = FormulaRule(
            formula=[f"=COUNTIF({cruce_sheet.title}!{col_letter}:{col_letter}, {numero_factura_col}2)>0"],
            fill=fill
        )
        
        data_range = f"{numero_factura_col}2:{numero_factura_col}{max_data_row}"
        data_sheet.conditional_formatting.add(data_range, rule)
        
        logger.info(f"Conditional formatting applied to {data_range} checking {cruce_sheet.title}!{col_letter}:{col_letter}")


def _apply_cruce_facturas_headers(workbook: Workbook) -> dict[str, Any]:
    cruce_sheet = _get_or_create_sheet(workbook, CRUCE_FACTURAS_SHEET)
    headers = {
        "B1": "Facturas Ok",
        "D1": "Facturas Pendientes",
        "F1": "PDFs de Facturas",
    }
    for cell, value in headers.items():
        cruce_sheet[cell] = value
    
    # Apply conditional formulas and formatting
    data_sheet = workbook.active
    _apply_conditional_formulas_to_cruce_facturas(workbook, cruce_sheet, data_sheet)
    
    return {
        "rule": "cruce_facturas_headers",
        "sheet": CRUCE_FACTURAS_SHEET,
        "cells": headers,
        "conditional_formulas": "applied",
    }


def _create_revision_sheet(workbook: Workbook) -> dict[str, Any]:
    """Create a revision sheet with headers for formulas and check for decimal values."""
    sheet = workbook.create_sheet(title="revision")
    
    # Add header for decimales formula
    sheet["A1"] = "Decimales"
    
    # Get the data sheet (after filtering)
    data_sheet = workbook.active
    
    # Find column indices for the required columns
    headers = []
    for col in range(1, data_sheet.max_column + 1):
        cell_value = data_sheet.cell(row=1, column=col).value
        headers.append(cell_value)
    
    # Find indices of required columns
    numero_factura_idx = None
    vlr_subsidiado_idx = None
    vlr_procedimiento_idx = None
    
    for i, header in enumerate(headers):
        if header == "Número Factura":
            numero_factura_idx = i
        elif header == "Vlr. Subsidiado":
            vlr_subsidiado_idx = i
        elif header == "Vlr. Procedimiento":
            vlr_procedimiento_idx = i
    
    # Check for decimal values and collect invoice numbers
    decimal_invoices = []
    if numero_factura_idx is not None:
        for row in range(2, data_sheet.max_row + 1):  # Start from row 2 (data rows)
            numero_factura = data_sheet.cell(row=row, column=numero_factura_idx + 1).value
            
            # Check Vlr. Subsidiado for decimals
            has_decimals = False
            if vlr_subsidiado_idx is not None:
                vlr_subsidiado = data_sheet.cell(row=row, column=vlr_subsidiado_idx + 1).value
                if vlr_subsidiado is not None and isinstance(vlr_subsidiado, (int, float)):
                    if isinstance(vlr_subsidiado, float) and vlr_subsidiado % 1 != 0:
                        has_decimals = True
            
            # Check Vlr. Procedimiento for decimals
            if not has_decimals and vlr_procedimiento_idx is not None:
                vlr_procedimiento = data_sheet.cell(row=row, column=vlr_procedimiento_idx + 1).value
                if vlr_procedimiento is not None and isinstance(vlr_procedimiento, (int, float)):
                    if isinstance(vlr_procedimiento, float) and vlr_procedimiento % 1 != 0:
                        has_decimals = True
            
            # If decimals found, add invoice number to list
            if has_decimals and numero_factura is not None:
                decimal_invoices.append(str(numero_factura))
    
    # Write decimal invoice numbers to revision sheet
    for i, invoice in enumerate(decimal_invoices, 2):  # Start from row 2
        sheet.cell(row=i, column=1, value=invoice)
    
    return {
        "rule": "create_revision_sheet", 
        "sheet": "revision", 
        "headers": ["Decimales"],
        "decimal_invoices_found": len(decimal_invoices)
    }


def _filter_columns(workbook: Workbook, sheet_name: str | None, header_row: int) -> None:
    """Hide columns not in the keep list after deleting first two rows."""
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
    
    logger.info("Before processing: max_row=%s, max_column=%s", sheet.max_row, sheet.max_column)
    
    # Unmerge cells in the first two rows to allow deletion
    for merged_range in list(sheet.merged_cells):
        if merged_range.min_row <= 2:
            sheet.unmerge_cells(str(merged_range))
            logger.info("Unmerged cells: %s", merged_range)
    
    # Delete the first two rows to move headers to the top
    sheet.delete_rows(1, 2)
    
    logger.info("After deleting rows: max_row=%s, max_column=%s", sheet.max_row, sheet.max_column)
    
    # Read headers from row 1 (now the original row 3)
    headers = []
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        headers.append(cell_value)
    
    logger.info("Headers after delete: %s", headers[:10])  # Log first 10 headers
    
    # Find indices of columns to keep (0-based)
    indices_to_keep = [i for i, h in enumerate(headers) if h in COLUMNS_TO_KEEP]
    
    logger.info("Indices to keep: %s", indices_to_keep)
    
    # Hide columns not in the keep list
    for col_idx in range(len(headers)):
        if col_idx not in indices_to_keep:
            col_letter = sheet.cell(row=1, column=col_idx + 1).column_letter
            sheet.column_dimensions[col_letter].hidden = True
            logger.info("Hiding column %s (%s)", col_letter, headers[col_idx])
    
    if not indices_to_keep:
        logger.warning("No matching columns found to keep in sheet %s", sheet.title)


def _apply_rules(workbook: Workbook) -> list[dict[str, Any]]:
    rules = [_apply_cruce_facturas_headers, _create_revision_sheet]
    applied: list[dict[str, Any]] = []
    for rule in rules:
        applied.append(rule(workbook))
    return applied


def export_excel_with_cruce_facturas(*, filename: str, sheet_name: str | None = None, header_row: int = 0) -> dict[str, Any]:
    logger.info("Iniciando exportación a output: %s", filename)
    source_path, source_error = resolve_safe_excel_in_input(filename)
    if source_error:
        return {"status": "error", "data": {}, "errors": [source_error]}
    assert source_path is not None

    validation_error = _validate_input_path(source_path)
    if validation_error:
        return {"status": "error", "data": {}, "errors": [validation_error]}

    output_path, output_error = resolve_safe_excel_in_output(source_path.name)
    if output_error:
        return {"status": "error", "data": {}, "errors": [output_error]}
    assert output_path is not None

    try:
        shutil.copy2(source_path, output_path)
        workbook = load_workbook(output_path)
        
        # Filter columns in the data sheet
        _filter_columns(workbook, sheet_name, header_row)
        
        applied_rules = _apply_rules(workbook)
        workbook.save(output_path)
    except Exception as exc:
        logger.exception("Error exportando Excel a output")
        return {"status": "error", "data": {}, "errors": [str(exc)]}

    logger.info("Exportación completada: %s", output_path.name)
    return {
        "status": "success",
        "data": {
            "input_file": source_path.name,
            "output_file": output_path.name,
            "output_path": str(output_path),
            "sheet": CRUCE_FACTURAS_SHEET,
            "headers_written": ["B1", "D1", "F1"],
            "applied_rules": applied_rules,
        },
        "errors": [],
    }
