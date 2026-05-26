"""Detección de revisión necesaria para entidad 86 en Urgencias.

Extraído de app/services/revision_sheet.py._detect_revision_entidad_86_urgencias
como parte de la Fase 7 (cleanup).
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_revision_entidad_86_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con Cód Entidad Cobrar = 86 que requieren revisión manual.

    En Urgencias, cuando el código entidad cobrar es '86', se marca la factura
    como revisión necesaria (NO es un error de validación, es una advertencia).

    Returns:
        Lista de dicts con keys: 'factura', 'codigo', 'procedimiento', 'entidad', 'ide_contrato'
    """
    tipo_factura_idx = indices.get("tipo_factura_descripcion")
    num_fact_idx = indices.get("numero_factura")
    codigo_entidad_idx = indices.get("codigo_entidad_cobrar")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")
    ide_contrato_idx = indices.get("ide_contrato")

    if None in (tipo_factura_idx, num_fact_idx, codigo_entidad_idx):
        logger.warning(
            "Revision Entidad 86 - Columnas necesarias no encontradas: "
            "tipo_factura_descripcion=%s, numero_factura=%s, codigo_entidad_cobrar=%s",
            tipo_factura_idx,
            num_fact_idx,
            codigo_entidad_idx,
        )
        return []

    revision_items: list[dict[str, str]] = []
    facturas_vistas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        # Filtrar por tipo_factura_descripcion = "Urgencias"
        tipo_factura = data_sheet.cell(row=row, column=tipo_factura_idx + 1).value
        tipo_factura_str = str(tipo_factura).strip() if tipo_factura else ""
        if tipo_factura_str != "Urgencias":
            continue

        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        if factura_str in facturas_vistas:
            continue

        codigo_entidad = data_sheet.cell(row=row, column=codigo_entidad_idx + 1).value
        if codigo_entidad is None:
            continue

        codigo_entidad_str = str(codigo_entidad).strip().upper()
        if codigo_entidad_str != "86":
            continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value if codigo_idx is not None else ""
        procedimiento = data_sheet.cell(row=row, column=procedimiento_idx + 1).value if procedimiento_idx is not None else ""
        ide_contrato = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value if ide_contrato_idx is not None else ""

        facturas_vistas.add(factura_str)
        revision_items.append({
            "factura": factura_str,
            "codigo": str(codigo).strip() if codigo else "",
            "procedimiento": str(procedimiento).strip() if procedimiento else "",
            "entidad": codigo_entidad_str,
            "ide_contrato": str(ide_contrato).strip() if ide_contrato else "",
        })

    logger.info(
        "Revision Entidad 86 - Filas procesadas: %d, Facturas únicas encontradas: %d",
        data_sheet.max_row - 1,
        len(revision_items),
    )
    return revision_items
