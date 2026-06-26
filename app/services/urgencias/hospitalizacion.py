"""Detector de reglas de Hospitalización.

Reglas de cantidades:
- Código 129B02 (Estancia): cantidad esperada = días_estancia + 1
- Código 890601 (Camas): cantidad esperada = días_redondeados_arriba
- Código 890601H: cantidad debe ser ≤ 1 (solo cuando Tarifario NO es SOAT)

Reglas de códigos:
- SOAT: debe tener 39133, 38114, 39131; NO puede tener 39145, 38915
- No-SOAT: >24h requiere 129B02, 890601H, 890601; <=24h requiere 129B02, 890601H
- Prohibidos: 05DSB01, 5DSB01, 890701
"""

from __future__ import annotations

import logging
from datetime import datetime
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    CODIGO_HOSPITALIZACION_CAMAS,
    CODIGO_HOSPITALIZACION_ESTANCIA,
    CODIGOS_HOSPITALIZACION_OBLIGATORIOS,
    CODIGOS_HOSPITALIZACION_PROHIBIDOS,
    CODIGOS_SOAT_HOSPITALIZACION_OBLIGATORIOS,
    CODIGOS_SOAT_HOSPITALIZACION_PROHIBIDOS,
    HORAS_POR_DIA,
    URGENCIAS_NO_SOAT_CODIGOS_CANTIDAD_MAX_1,
    VALOR_TARIFARIO_SOAT,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_cantidades_hospitalizacion(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta facturas con cantidades incorrectas en Hospitalización.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo", "procedimiento", "cantidad",
        "cantidad_esperada", "estancia_dias", "tipo_factura"
    """
    tipo_factura_idx = indices.get("tipo_factura_descripcion")
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")
    cantidad_idx = indices.get("cantidad")
    fec_factura_idx = indices.get("fec_factura")
    fecha_cierre_idx = indices.get("fecha_cierre")
    tarifario_idx = indices.get("tarifario")

    if None in (tipo_factura_idx, num_fact_idx, codigo_idx, cantidad_idx):
        logger.warning("Cantidades Hospitalización - Columnas necesarias no encontradas")
        return []

    problemas = []

    for row in range(2, data_sheet.max_row + 1):
        tipo_factura = data_sheet.cell(row=row, column=tipo_factura_idx + 1).value
        tipo_factura_str = str(tipo_factura).strip() if tipo_factura else ""

        # Solo procesar si Tipo Factura = "Hospitalización"
        if tipo_factura_str != "Hospitalización":
            continue

        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip().upper() if codigo else ""

        # Solo procesar códigos 129B02, 890601 y 890601H
        codigos_hosp_calculados = {CODIGO_HOSPITALIZACION_ESTANCIA, CODIGO_HOSPITALIZACION_CAMAS}
        if codigo_str not in codigos_hosp_calculados and codigo_str not in URGENCIAS_NO_SOAT_CODIGOS_CANTIDAD_MAX_1:
            continue

        cantidad = data_sheet.cell(row=row, column=cantidad_idx + 1).value
        if not isinstance(cantidad, (int, float)):
            continue

        # Obtener tarifario (necesario para reglas condicionales como 890601H)
        tarifario = data_sheet.cell(row=row, column=tarifario_idx + 1).value if tarifario_idx is not None else None
        tarifario_str = str(tarifario).strip().upper() if tarifario else ""

        # Calcular estancia en horas y días
        estancia_horas = 0
        fec_factura_cell = data_sheet.cell(row=row, column=fec_factura_idx + 1).value if fec_factura_idx else None
        fecha_cierre_cell = data_sheet.cell(row=row, column=fecha_cierre_idx + 1).value if fecha_cierre_idx else None

        if fec_factura_cell and fecha_cierre_cell:
            try:
                fec_factura_dt = datetime.strptime(str(fec_factura_cell).strip(), "%Y-%m-%d %H:%M:%S")
                fecha_cierre_dt = datetime.strptime(str(fecha_cierre_cell).strip(), "%Y-%m-%d %H:%M:%S")
                diferencia = fecha_cierre_dt - fec_factura_dt
                estancia_horas = diferencia.total_seconds() / 3600
            except (ValueError, TypeError):
                estancia_horas = 0

        estancia_dias_ceiling = -(-int(estancia_horas) // HORAS_POR_DIA)  # Ceiling division
        estancia_dias_floor = int(estancia_horas) // HORAS_POR_DIA  # Floor division (días completos)

        procedimiento = ""
        if procedimiento_idx is not None:
            proc_value = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
            procedimiento = str(proc_value).strip() if proc_value else ""

        es_error = False
        cantidad_esperada = None

        if codigo_str == CODIGO_HOSPITALIZACION_ESTANCIA:
            # 129B02: cantidad = días_completos + 1 (solo cuenta días completos, no el parcial)
            cantidad_esperada = estancia_dias_floor + 1
            if cantidad != cantidad_esperada:
                es_error = True
                logger.warning(
                    "CANTIDAD HOSPITALIZACIÓN 129B02 - Factura='%s', Fila=%d, Estancia=%.1fh (%d días completos), Cantidad=%s (esperado=%d)",
                    factura_str, row, estancia_horas, estancia_dias_floor, cantidad, cantidad_esperada
                )

        elif codigo_str == CODIGO_HOSPITALIZACION_CAMAS:
            # 890601: cantidad = días_completos (floor), NO puede existir si < 24h
            if estancia_horas < HORAS_POR_DIA:
                # < 24h -> ERROR: no puede haber 890601
                es_error = True
                cantidad_esperada = 0  # Indica que no debería existir
                logger.warning(
                    "CANTIDAD HOSPITALIZACIÓN 890601 - Factura='%s', Fila=%d, Estancia=%.1fh (<24h) -> NO DEBE EXISTIR",
                    factura_str, row, estancia_horas
                )
            else:
                cantidad_esperada = estancia_dias_floor
                if cantidad != cantidad_esperada:
                    es_error = True
                    logger.warning(
                        "CANTIDAD HOSPITALIZACIÓN 890601 - Factura='%s', Fila=%d, Estancia=%.1fh (%d días completos), Cantidad=%s (esperado=%d)",
                        factura_str, row, estancia_horas, estancia_dias_floor, cantidad, cantidad_esperada
                    )

        elif codigo_str in URGENCIAS_NO_SOAT_CODIGOS_CANTIDAD_MAX_1:
            # 890601H: cantidad debe ser ≤ 1 (solo cuando NO es SOAT)
            if tarifario_str != VALOR_TARIFARIO_SOAT and cantidad > 1:
                es_error = True
                cantidad_esperada = 1
                logger.warning(
                    "CANTIDAD HOSPITALIZACIÓN 890601H - Factura='%s', Fila=%d, Cantidad=%s (debe ser <=1, tarifario=%s)",
                    factura_str, row, cantidad, tarifario_str
                )

        if es_error:
            problemas.append({
                "factura": factura_str,
                "codigo": codigo_str,
                "procedimiento": procedimiento,
                "cantidad": cantidad,
                "cantidad_esperada": cantidad_esperada,
                "estancia_horas": round(estancia_horas, 1),
                "estancia_dias": estancia_dias_floor,
                "tipo_factura": tipo_factura_str,
                "fila": row,
            })

    if problemas:
        logger.info("Cantidades Hospitalización - Problemas encontrados: %d", len(problemas))

    return problemas


def _format_estancia_hosp(horas: float | None) -> str:
    """Formatea estancia en días + horas."""
    if horas is None:
        return "N/A"
    dias = int(horas // 24)
    hrs = int(horas % 24)
    if dias > 0:
        return f"{dias}d {hrs}h"
    return f"{hrs}h"


def detect_hospitalizacion_codes(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta problemas de códigos obligatorios/prohibidos en Hospitalización.

    Reglas extraídas de _detect_centro_costo_urgencias (revision_sheet.py):
    - SOAT: debe tener 39133, 38114, 39131; NO puede tener 39145, 38915
    - No-SOAT: >24h requiere 129B02, 890601H, 890601; <=24h requiere 129B02, 890601H
    - Prohibidos: 05DSB01, 5DSB01, 890701

    Returns:
        Lista de dicts con keys compatibles con problemas_cups_equivalentes:
        "factura", "codigo", "codigo_equiv", "accion", "procedimiento", "estancia_str"
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    codigo_entidad_cobrar_idx = indices.get("codigo_entidad_cobrar")
    tipo_factura_descripcion_idx = indices.get("tipo_factura_descripcion")
    proc_idx = indices.get("procedimiento")
    tarifario_idx = indices.get("tarifario")
    fec_factura_idx = indices.get("fec_factura")
    fecha_cierre_idx = indices.get("fecha_cierre")

    if num_fact_idx is None or tipo_factura_descripcion_idx is None:
        logger.warning("Hospitalización Códigos - Columnas necesarias no encontradas")
        return []

    # ----- Paso 1: Recolectar datos de hospitalización por factura
    factura_hospitalizacion_data: dict[str, dict[str, Any]] = {}

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        tipo_factura_cell = data_sheet.cell(row=row, column=tipo_factura_descripcion_idx + 1).value
        tipo_factura_str = str(tipo_factura_cell).strip() if tipo_factura_cell else ""

        # Solo procesar Hospitalización
        if tipo_factura_str != "Hospitalización":
            continue

        codigo_cell = data_sheet.cell(row=row, column=codigo_idx + 1).value if codigo_idx else None
        codigo_normalized = str(codigo_cell).strip() if codigo_cell else ""

        # Calcular estancia para esta fila
        estancia_horas_hosp = None
        if fec_factura_idx is not None and fecha_cierre_idx is not None:
            fec_factura_cell_hosp = data_sheet.cell(row=row, column=fec_factura_idx + 1).value
            fecha_cierre_cell_hosp = data_sheet.cell(row=row, column=fecha_cierre_idx + 1).value
            if fec_factura_cell_hosp and fecha_cierre_cell_hosp:
                try:
                    fec_factura_dt_hosp = datetime.strptime(str(fec_factura_cell_hosp).strip(), "%Y-%m-%d %H:%M:%S")
                    fecha_cierre_dt_hosp = datetime.strptime(str(fecha_cierre_cell_hosp).strip(), "%Y-%m-%d %H:%M:%S")
                    diferencia_hosp = fecha_cierre_dt_hosp - fec_factura_dt_hosp
                    estancia_horas_hosp = diferencia_hosp.total_seconds() / 3600
                except (ValueError, TypeError):
                    estancia_horas_hosp = None

        # Capturar entidad
        entidad_hosp = ""
        if codigo_entidad_cobrar_idx is not None:
            entidad_cell_hosp = data_sheet.cell(row=row, column=codigo_entidad_cobrar_idx + 1).value
            entidad_hosp = str(entidad_cell_hosp).strip() if entidad_cell_hosp else ""

        # Capturar tarifario
        tarifario_hosp = ""
        if tarifario_idx is not None:
            tarifario_cell_hosp = data_sheet.cell(row=row, column=tarifario_idx + 1).value
            tarifario_hosp = str(tarifario_cell_hosp).strip() if tarifario_cell_hosp else ""

        if factura_str not in factura_hospitalizacion_data:
            factura_hospitalizacion_data[factura_str] = {
                "entidad": entidad_hosp,
                "tarifario": tarifario_hosp,
                "codigos_hospitalizacion": set(),
                "codigos_prohibidos": set(),
                "estancia_horas": estancia_horas_hosp,
                "tiene_soat_hosp_prohibido": False,
                "codigos_soat_hosp_obligatorios": set(),
            }
        else:
            if estancia_horas_hosp is not None:
                previa = factura_hospitalizacion_data[factura_str].get("estancia_horas")
                if previa is None or estancia_horas_hosp > previa:
                    factura_hospitalizacion_data[factura_str]["estancia_horas"] = estancia_horas_hosp
            if not factura_hospitalizacion_data[factura_str].get("entidad") and entidad_hosp:
                factura_hospitalizacion_data[factura_str]["entidad"] = entidad_hosp
            if tarifario_hosp and tarifario_hosp.upper() == VALOR_TARIFARIO_SOAT:
                factura_hospitalizacion_data[factura_str]["tarifario"] = tarifario_hosp
            elif not factura_hospitalizacion_data[factura_str].get("tarifario") and tarifario_hosp:
                factura_hospitalizacion_data[factura_str]["tarifario"] = tarifario_hosp

        # Colectar códigos de hospitalización
        if codigo_normalized in CODIGOS_HOSPITALIZACION_OBLIGATORIOS:
            factura_hospitalizacion_data[factura_str]["codigos_hospitalizacion"].add(codigo_normalized)
        if codigo_normalized in CODIGOS_HOSPITALIZACION_PROHIBIDOS:
            factura_hospitalizacion_data[factura_str]["codigos_prohibidos"].add(codigo_normalized)
        if tarifario_hosp.upper() == VALOR_TARIFARIO_SOAT and codigo_normalized in CODIGOS_SOAT_HOSPITALIZACION_PROHIBIDOS:
            factura_hospitalizacion_data[factura_str]["tiene_soat_hosp_prohibido"] = True
        if tarifario_hosp.upper() == VALOR_TARIFARIO_SOAT and codigo_normalized in CODIGOS_SOAT_HOSPITALIZACION_OBLIGATORIOS:
            factura_hospitalizacion_data[factura_str]["codigos_soat_hosp_obligatorios"].add(codigo_normalized)

    # ----- Paso 2: Validar problemas de hospitalización
    problemas_cups_equivalentes: list[dict[str, Any]] = []
    facturas_hosp_reportadas: set[str] = set()

    for factura_str, data in factura_hospitalizacion_data.items():
        codigos_hosp = data.get("codigos_hospitalizacion", set())
        codigos_prohibidos = data.get("codigos_prohibidos", set())
        estancia_horas = data.get("estancia_horas")
        entidad = data.get("entidad", "")
        tarifario = data.get("tarifario", "")

        # ----- Regla SOAT: Hospitalización debe tener 39133, 38114, 39131
        if tarifario.upper() == VALOR_TARIFARIO_SOAT:
            codigos_soat_hosp_obligatorios = data.get("codigos_soat_hosp_obligatorios", set())
            codigos_requeridos = set(CODIGOS_SOAT_HOSPITALIZACION_OBLIGATORIOS)
            # Excepción: 39131 no es obligatorio si estancia < 24h
            if estancia_horas is not None and estancia_horas < 24:
                codigos_requeridos.discard("39131")
            faltan_soat_oblig = codigos_requeridos - codigos_soat_hosp_obligatorios
            if faltan_soat_oblig:
                estancia_str = _format_estancia_hosp(estancia_horas) if estancia_horas is not None else "N/A"
                problemas_cups_equivalentes.append({
                    "factura": factura_str,
                    "codigo": list(codigos_soat_hosp_obligatorios),
                    "codigo_equiv": "",
                    "accion": f"SOAT Hospitalización debe tener: {', '.join(sorted(faltan_soat_oblig))}",
                    "procedimiento": "",
                    "estancia_str": estancia_str,
                })
                facturas_hosp_reportadas.add(factura_str)

            # ----- Regla SOAT: NO puede tener 39145, 38915
            tiene_soat_hosp_prohibido = data.get("tiene_soat_hosp_prohibido", False)
            if tiene_soat_hosp_prohibido:
                estancia_str = _format_estancia_hosp(estancia_horas) if estancia_horas is not None else "N/A"
                problemas_cups_equivalentes.append({
                    "factura": factura_str,
                    "codigo": list(codigos_prohibidos),
                    "codigo_equiv": "",
                    "accion": f"SOAT Hospitalización no puede tener: {', '.join(CODIGOS_SOAT_HOSPITALIZACION_PROHIBIDOS)}",
                    "procedimiento": "",
                    "estancia_str": estancia_str,
                })
                facturas_hosp_reportadas.add(factura_str)
            continue  # Saltar reglas no-SOAT

        # ----- Reglas no-SOAT
        if estancia_horas is None:
            continue

        # Determinar códigos obligatorios según estancia
        if estancia_horas > 24:
            codigos_obligatorios_hosp = CODIGOS_HOSPITALIZACION_OBLIGATORIOS  # 3 códigos
        else:
            codigos_obligatorios_hosp = {"890601H", "129B02"}  # solo 2 códigos

        estancia_str = _format_estancia_hosp(estancia_horas)

        # Validar códigos obligatorios
        faltan = codigos_obligatorios_hosp - codigos_hosp
        if faltan:
            problemas_cups_equivalentes.append({
                "factura": factura_str,
                "codigo": list(codigos_hosp),
                "codigo_equiv": "",
                "accion": (
                    f"Hospitalización ({'>24h' if estancia_horas > 24 else '<=24h'}) "
                    f"debe tener: {', '.join(sorted(codigos_obligatorios_hosp))} "
                    f"(faltan: {', '.join(faltan)})"
                ),
                "procedimiento": "",
                "estancia_str": estancia_str,
            })
            facturas_hosp_reportadas.add(factura_str)
            logger.warning(
                "DETECTADO cups equiv HOSPITALIZACION: factura=%s, estancia=%.1fh, tiene=%s, faltan=%s",
                factura_str, estancia_horas or 0, list(codigos_hosp), list(faltan),
            )

        # Validar códigos prohibidos (05DSB01, 5DSB01)
        if codigos_prohibidos:
            problemas_cups_equivalentes.append({
                "factura": factura_str,
                "codigo": list(codigos_prohibidos),
                "codigo_equiv": "",
                "accion": f"Hospitalización no puede tener: {', '.join(codigos_prohibidos)}",
                "procedimiento": "",
                "estancia_str": estancia_str,
            })
            facturas_hosp_reportadas.add(factura_str)
            logger.warning(
                "DETECTADO cups equiv HOSPITALIZACION PROHIBIDO: factura=%s, prohibidos=%s",
                factura_str, list(codigos_prohibidos),
            )

    if problemas_cups_equivalentes:
        logger.info("Hospitalización Códigos - Problemas encontrados: %d", len(problemas_cups_equivalentes))

    return problemas_cups_equivalentes
