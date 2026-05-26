"""Detección de códigos CUPS que no están en la base de datos.

Extraído de app/services/revision_sheet.py._get_codigos_no_en_db_ess118
como parte de la Fase 7 (cleanup).
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def get_codigos_no_en_db_ess118(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Retorna lista de problemas de códigos CUPS que NO están en la DB.

    Regla: IDE Contrato = 969 Y Código Tipo Procedimiento no es 9,12,13
           Y código NO está en tabla procedimiento → ERROR

    Nota: Se consulta la tabla procedimiento de PostgreSQL (no la DB externa).

    Returns:
        Lista de dicts con keys: "factura", "codigo", "procedimiento", "entidad"
    """
    # Cargar códigos válidos de la tabla procedimiento relacionados con nota_hoja = 3
    from app.database import SessionLocal
    from app.models import Procedimiento, NotasTecnicas

    try:
        db = SessionLocal()
        cups_validos = set(
            row.cups
            for row in db.query(Procedimiento.cups)
            .join(NotasTecnicas, NotasTecnicas.id_procedimiento == Procedimiento.id)
            .filter(NotasTecnicas.id_nota_hoja == 3)
            .distinct()
            .all()
        )
        db.close()
    except Exception as e:
        logger.warning("No se pudo conectar a DB para validar códigos: %s", e)
        return []

    if not cups_validos:
        logger.warning("No hay códigos en tabla procedimiento para nota_hoja=3")
        return []

    logger.info("Códigos válidos (nota_hoja=3): %d", len(cups_validos))

    tipo_factura_idx = indices.get("tipo_factura_descripcion")
    codigo_idx = indices.get("codigo")
    ide_contrato_idx = indices.get("ide_contrato")
    codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")
    num_fact_idx = indices.get("numero_factura")
    proc_idx = indices.get("procedimiento")
    codigo_entidad_idx = indices.get("codigo_entidad_cobrar")

    if tipo_factura_idx is None or codigo_idx is None:
        return []

    problemas = []

    for row in range(2, data_sheet.max_row + 1):
        tipo_factura = data_sheet.cell(row=row, column=tipo_factura_idx + 1).value
        tipo_factura_str = str(tipo_factura).strip() if tipo_factura else ""

        # Solo procesar si Tipo Factura = "Urgencias"
        if tipo_factura_str != "Urgencias":
            continue

        # Verificar IDE Contrato = 969
        ide_contrato = None
        if ide_contrato_idx is not None:
            ide_contrato = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value

        ide_str = str(ide_contrato).strip() if ide_contrato else ""

        # Solo procesar si IDE = 969
        if ide_str != "969":
            continue

        # Excluir Código Tipo Procedimiento = 09, 12, 13
        if codigo_tipo_proc_idx is not None:
            codigo_tipo = data_sheet.cell(row=row, column=codigo_tipo_proc_idx + 1).value
            if codigo_tipo and str(codigo_tipo).strip() in ["09", "12", "13"]:
                continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        if not codigo:
            continue

        codigo_str = str(codigo).strip()

        # Verificar si existe en la tabla procedimiento
        if codigo_str not in cups_validos:
            factura = ""
            if num_fact_idx is not None:
                factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value or ""

            procedimiento = ""
            if proc_idx is not None:
                procedimiento = data_sheet.cell(row=row, column=proc_idx + 1).value or ""

            entidad = ""
            if codigo_entidad_idx is not None:
                entidad = data_sheet.cell(row=row, column=codigo_entidad_idx + 1).value or ""

            problemas.append({
                "factura": str(factura),
                "codigo": codigo_str,
                "procedimiento": str(procedimiento),
                "entidad": str(entidad),
            })

    return problemas
