"""Detección de IDE Contrato REVERSE (sin entidad) para Urgencias.

Extraído de app/services/revision_sheet.py._detect_ide_contrato_reverse_urgencias
como parte de la Fase 7 (cleanup).
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import IDE_CONTRATO_REVERSE
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_ide_contrato_reverse_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con problemas de IDE Contrato REVERSE (sin entidad).

    Dado un IDE Contrato, verifica que el Código CUPS corresponda al esperado.

    Reglas REVERSE (sin entidad):
    - IDE 986 → Código debe ser 906340
    - IDE 977 → Código puede ser 861801 (siempre) o 890405 (solo si la
      identificación NO tiene 861801 en otra factura; si la tiene,
      890405 debería ser IDE 976)

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo", "ide_contrato", "codigo_deberia"
    """
    tipo_factura_idx = indices.get("tipo_factura_descripcion")
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    ide_contrato_idx = indices.get("ide_contrato")
    ident_idx = indices.get("identificacion")

    if None in (tipo_factura_idx, num_fact_idx, codigo_idx, ide_contrato_idx):
        logger.warning(
            "IDE Contrato REVERSE - Columnas necesarias no encontradas: "
            "numero_factura=%s, codigo=%s, ide_contrato=%s",
            num_fact_idx, codigo_idx, ide_contrato_idx,
        )
        return []

    # PASO 1: Recolectar identificaciones que tienen código 861801
    # (estas identificaciones NO deberían usar 890405 con IDE 977)
    identificaciones_con_861801: set[str] = set()
    for row in range(2, data_sheet.max_row + 1):
        ident = data_sheet.cell(row=row, column=ident_idx + 1).value if ident_idx is not None else None
        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip().upper() if codigo else ""

        if codigo_str == "861801" and ident:
            identificaciones_con_861801.add(str(ident).strip())

    logger.info("IDE REVERSE: Identificaciones con 861801: %d", len(identificaciones_con_861801))

    problemas = []
    facturas_procesadas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        tipo_factura = data_sheet.cell(row=row, column=tipo_factura_idx + 1).value
        tipo_factura_str = str(tipo_factura).strip() if tipo_factura else ""

        # Solo procesar si Tipo Factura = "Urgencias"
        if tipo_factura_str != "Urgencias":
            continue

        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_procesadas:
            continue

        ide_contrato = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value
        ide_str = str(ide_contrato).strip() if ide_contrato else ""

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip() if codigo else ""

        ident = data_sheet.cell(row=row, column=ident_idx + 1).value if ident_idx is not None else None
        ident_str = str(ident).strip() if ident else ""

        # Verificar regla 986
        if ide_str == "986":
            codigo_esperado = IDE_CONTRATO_REVERSE.get("986")
            if codigo_str != codigo_esperado:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": codigo_esperado,
                    "observacion": f"IDE 986 → Código debe ser {codigo_esperado}",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 977
        if ide_str == "977":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "976",
                        "observacion": "890405 con IDE 977 inválido - Identificación tiene 861801, debería ser 976",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 977 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 979
        if ide_str == "979":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "967",
                        "observacion": "890405 con IDE 979 inválido - Identificación tiene 861801, debería ser 967",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 979 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 839
        if ide_str == "839":
            codigo_esperado = "906340"
            if codigo_str != codigo_esperado:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": codigo_esperado,
                    "observacion": f"IDE 839 → Código debe ser {codigo_esperado}",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 842
        if ide_str == "842":
            codigo_esperado = "906340"
            if codigo_str != codigo_esperado:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": codigo_esperado,
                    "observacion": f"IDE 842 → Código debe ser {codigo_esperado}",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 958
        if ide_str == "958":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "959",
                        "observacion": "890405 con IDE 958 inválido - Identificación tiene 861801, debería ser 959",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 958 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 961
        if ide_str == "961":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "962",
                        "observacion": "890405 con IDE 961 inválido - Identificación tiene 861801, debería ser 962",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 961 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 922
        if ide_str == "922":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "921",
                        "observacion": "890405 con IDE 922 inválido - Identificación tiene 861801, debería ser 921",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 922 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 863
        if ide_str == "863":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "862",
                        "observacion": "890405 con IDE 863 inválido - Identificación tiene 861801, debería ser 862",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 863 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 975
        if ide_str == "975":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "968",
                        "observacion": "890405 con IDE 975 inválido - Identificación tiene 861801, debería ser 968",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 975 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 920
        if ide_str == "920":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "919",
                        "observacion": "890405 con IDE 920 inválido - Identificación tiene 861801, debería ser 919",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 920 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 908
        if ide_str == "908":
            if codigo_str == "861801":
                continue
            elif codigo_str == "890405":
                if ident_str in identificaciones_con_861801:
                    problemas.append({
                        "factura": factura_str,
                        "codigo": codigo_str,
                        "ide_contrato": ide_str,
                        "codigo_deberia": "909",
                        "observacion": "890405 con IDE 908 inválido - Identificación tiene 861801, debería ser 909",
                    })
                    facturas_procesadas.add(factura_str)
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "861801 o 890405",
                    "observacion": "IDE 908 → Código debe ser 861801 o 890405 (sin 861801 en identificación)",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 970 (ESS118): 735301, 861801 o 890205
        if ide_str == "970":
            codigos_permitidos = {"735301", "861801", "890205"}
            if codigo_str in codigos_permitidos:
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "735301, 861801 o 890205",
                    "observacion": "IDE 970 (ESS118) → Código debe ser 735301, 861801 o 890205",
                })
                facturas_procesadas.add(factura_str)
            continue

        # Verificar regla 974 (ESS118): 735301, 861801 o 890405
        if ide_str == "974":
            codigos_permitidos = {"735301", "861801", "890405"}
            if codigo_str in codigos_permitidos:
                continue
            else:
                problemas.append({
                    "factura": factura_str,
                    "codigo": codigo_str,
                    "ide_contrato": ide_str,
                    "codigo_deberia": "735301, 861801 o 890405",
                    "observacion": "IDE 974 (ESS118) → Código debe ser 735301, 861801 o 890405",
                })
                facturas_procesadas.add(factura_str)
            continue

    return problemas
