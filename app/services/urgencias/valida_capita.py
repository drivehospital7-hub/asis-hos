"""Detector de códigos CUPS no permitidos en facturas CAP (cápita) de Urgencias.

Regla:
- Si Número Factura empieza con prefijo "CAP", el código CUPS debe pertenecer
  al listado URGENCIAS_CAPITA_CUPS_CODES.
- Si el código NO está en el listado, se marca como error.
- EXCLUYE filas con Código Tipo Procedimiento = "09" o "12" (no se procesan).
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants.urgencias import (
    CODIGOS_TIPO_PROC_09_12,
    URGENCIAS_CAPITA_CUPS_CODES,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)

PREFIJO_CAP = "CAP"


def detect_capita_cups_invalidos(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta códigos CUPS en facturas con prefijo CAP que NO están
    en el listado URGENCIAS_CAPITA_CUPS_CODES.

    Excluye filas con Código Tipo Procedimiento = 09 o 12.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo", "procedimiento", "observacion"
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")
    tipo_proc_idx = indices.get("codigo_tipo_procedimiento")

    if num_fact_idx is None or codigo_idx is None:
        logger.warning("VALIDA CAPITA - Columnas necesarias no encontradas")
        return []

    problemas: list[dict[str, str]] = []

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        # Solo facturas con prefijo CAP
        if not factura_str.upper().startswith(PREFIJO_CAP):
            continue

        # Excluir Tipo Procedimiento 09 y 12 (farmacia, materiales)
        if tipo_proc_idx is not None:
            tipo_proc_val = data_sheet.cell(row=row, column=tipo_proc_idx + 1).value
            tipo_proc_str = str(tipo_proc_val).strip() if tipo_proc_val else ""
            if tipo_proc_str in CODIGOS_TIPO_PROC_09_12:
                continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip().upper() if codigo else ""

        if not codigo_str:
            continue

        # Si el código está en el listado CAPITA, es válido
        if codigo_str in URGENCIAS_CAPITA_CUPS_CODES:
            continue

        procedimiento = ""
        if procedimiento_idx is not None:
            proc_value = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
            procedimiento = str(proc_value).strip() if proc_value else ""

        problemas.append({
            "factura": factura_str,
            "codigo": codigo_str,
            "procedimiento": procedimiento,
            "observacion": (
                f"Código {codigo_str} no está en el listado CAPITA. "
                "Factura con prefijo CAP solo permite códigos del listado URGENCIAS CAPITA CUPS."
            ),
        })

    if problemas:
        logger.info(
            "VALIDA CAPITA - Códigos no CAPITA en facturas CAP encontrados: %d",
            len(problemas),
        )

    return problemas
