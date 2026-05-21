"""Detector de cantidades SOAT anómalas en facturas de Hospitalización.

Regla: Si Tarifario = "SOAT" y Tipo Factura Descripción = "Hospitalización":
- Código 38114: cantidad = días_estancia + 1
- Código 39131: cantidad = días_estancia
- Código 39133: cantidad debe ser ≤ 1
"""

from __future__ import annotations

import logging
from datetime import datetime
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    CODIGOS_SOAT_HOSPITALIZACION_CANTIDAD,
    HORAS_POR_DIA,
    URGENCIAS_SOAT_CODIGOS_CANTIDAD_MAX_1,
    VALOR_TARIFARIO_SOAT,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_cantidades_soat_hospitalizacion(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta facturas SOAT Hospitalización con cantidades incorrectas.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo", "procedimiento", "cantidad",
        "cantidad_esperada", "estancia_dias", "tipo_factura"
    """
    tipo_factura_idx = indices.get("tipo_factura_descripcion")
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")
    cantidad_idx = indices.get("cantidad")
    fec_factura_idx = indices.get("fec_factura")
    fecha_cierre_idx = indices.get("fecha_cierre")
    tarifario_idx = indices.get("tarifario")

    if None in (tipo_factura_idx, num_fact_idx, codigo_idx, cantidad_idx, tarifario_idx):
        logger.warning("Cantidades SOAT Hospitalización - Columnas necesarias no encontradas")
        return []

    problemas = []

    for row in range(2, data_sheet.max_row + 1):
        tipo_factura = data_sheet.cell(row=row, column=tipo_factura_idx + 1).value
        tipo_factura_str = str(tipo_factura).strip() if tipo_factura else ""

        # Solo procesar si Tipo Factura = "Hospitalización"
        if tipo_factura_str != "Hospitalización":
            continue

        # Verificar si es tarifario SOAT
        tarifario = data_sheet.cell(row=row, column=tarifario_idx + 1).value
        tarifario_str = str(tarifario).strip().upper() if tarifario else ""
        if tarifario_str != VALOR_TARIFARIO_SOAT:
            continue

        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip().upper() if codigo else ""

        # Solo procesar códigos 38114, 39131 y los de SOAT + Hospitalización (39133)
        if codigo_str not in CODIGOS_SOAT_HOSPITALIZACION_CANTIDAD and codigo_str not in URGENCIAS_SOAT_CODIGOS_CANTIDAD_MAX_1:
            continue

        cantidad = data_sheet.cell(row=row, column=cantidad_idx + 1).value
        if not isinstance(cantidad, (int, float)):
            continue

        # Calcular estancia en horas y días
        estancia_horas = 0
        fec_factura_cell = data_sheet.cell(row=row, column=fec_factura_idx + 1).value if fec_factura_idx else None
        fecha_cierre_cell = data_sheet.cell(row=row, column=fecha_cierre_idx + 1).value if fecha_cierre_idx else None

        if fec_factura_cell and fecha_cierre_cell:
            try:
                fec_factura_dt = datetime.strptime(str(fec_factura_cell).strip(), "%Y-%m-%d %H:%M:%S")
                fecha_cierre_dt = datetime.strptime(str(fecha_cierre_cell).strip(), "%Y-%m-%d %H:%M:%S")
                diferencia = fecha_cierre_dt - fec_factura_dt
                estancia_horas = diferencia.total_seconds() / 3600
            except (ValueError, TypeError):
                estancia_horas = 0

        estancia_dias_floor = int(estancia_horas) // HORAS_POR_DIA  # Días completos

        procedimiento = ""
        if procedimiento_idx is not None:
            proc_value = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
            procedimiento = str(proc_value).strip() if proc_value else ""

        es_error = False
        cantidad_esperada = None

        if codigo_str == "38114":
            # 38114: cantidad = días_completos + 1
            cantidad_esperada = estancia_dias_floor + 1
            if cantidad != cantidad_esperada:
                es_error = True
                logger.warning(
                    "CANTIDAD SOAT HOSPITALIZACIÓN 38114 - Factura='%s', Fila=%d, Estancia=%.1fh (%d días completos), Cantidad=%s (esperado=%d)",
                    factura_str, row, estancia_horas, estancia_dias_floor, cantidad, cantidad_esperada
                )

        elif codigo_str == "39131":
            # 39131: cantidad = días_completos (<24h = 0 días -> cantidad 0)
            cantidad_esperada = estancia_dias_floor
            if cantidad != cantidad_esperada:
                es_error = True
                logger.warning(
                    "CANTIDAD SOAT HOSPITALIZACIÓN 39131 - Factura='%s', Fila=%d, Estancia=%.1fh (%d días completos), Cantidad=%s (esperado=%d)",
                    factura_str, row, estancia_horas, estancia_dias_floor, cantidad, cantidad_esperada
                )

        elif codigo_str in URGENCIAS_SOAT_CODIGOS_CANTIDAD_MAX_1:
            # 39133: cantidad debe ser ≤ 1
            if cantidad > 1:
                es_error = True
                cantidad_esperada = 1
                logger.warning(
                    "CANTIDAD SOAT HOSPITALIZACIÓN 39133 - Factura='%s', Fila=%d, Cantidad=%s (debe ser <=1)",
                    factura_str, row, cantidad
                )

        if es_error:
            problemas.append({
                "factura": factura_str,
                "codigo": codigo_str,
                "procedimiento": procedimiento,
                "cantidad": cantidad,
                "cantidad_esperada": cantidad_esperada,
                "estancia_dias": estancia_dias_floor,
                "tipo_factura": tipo_factura_str,
            })

    if problemas:
        logger.info("Cantidades SOAT Hospitalización - Problemas encontrados: %d", len(problemas))

    return problemas
