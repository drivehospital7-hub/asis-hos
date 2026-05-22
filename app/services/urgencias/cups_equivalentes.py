"""Detector de problemas de CUPS equivalentes en Urgencias.

Reglas extraídas de _detect_centro_costo_urgencias (revision_sheet.py):
- Código 890201 -> debe usarse 890701
- Código 129B01 -> debe usarse 129B02
- Código 890205 -> debe usarse 890405 (excepto ESS118/ESSC18)
- Código 939402 + Hospitalización -> Error
- Código 12333 + Hospitalización -> Error
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    CODIGO_CUPS_EQUIVALENTE_890205,
    CODIGO_CUPS_EQUIVALENTE_SUSTITUTO_890405,
    CODIGO_CUPS_HOSPITALIZACION_PROHIBIDO,
    CODIGO_12333_HOSPITALIZACION_PROHIBIDO,
    ENTIDADES_PERMITIDAS_890205,
    ERROR_12333_HOSPITALIZACION_NO_PERMITIDO,
    ERROR_HOSPITALIZACION_NO_PERMITIDO,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_cups_equivalentes(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta facturas con problemas de CUPS equivalentes en Urgencias.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo", "codigo_equiv",
        "accion", "procedimiento"
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    codigo_entidad_cobrar_idx = indices.get("codigo_entidad_cobrar")
    tipo_factura_descripcion_idx = indices.get("tipo_factura_descripcion")
    proc_idx = indices.get("procedimiento")

    if num_fact_idx is None or codigo_idx is None:
        logger.warning("CUPS Equivalentes - Columnas necesarias no encontradas")
        return []

    problemas_cups_equivalentes: list[dict[str, Any]] = []

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        codigo = (
            data_sheet.cell(row=row, column=codigo_idx + 1).value
            if codigo_idx is not None else None
        )
        codigo_excluir = str(codigo).strip() if codigo else ""

        if not codigo_excluir:
            continue

        codigo_entidad_str = ""
        if codigo_entidad_cobrar_idx is not None:
            codigo_entidad_val = data_sheet.cell(row=row, column=codigo_entidad_cobrar_idx + 1).value
            codigo_entidad_str = str(codigo_entidad_val).strip() if codigo_entidad_val else ""

        tipo_factura_str = ""
        if tipo_factura_descripcion_idx is not None:
            tipo_factura_val = data_sheet.cell(row=row, column=tipo_factura_descripcion_idx + 1).value
            tipo_factura_str = str(tipo_factura_val).strip() if tipo_factura_val else ""

        proc_str = ""
        if proc_idx is not None:
            proc_val = data_sheet.cell(row=row, column=proc_idx + 1).value
            proc_str = str(proc_val).strip() if proc_val else ""

        # ----- Regla: Código CUPS 939402 + Tipo Factura=Hospitalización -> Error
        if codigo_excluir == CODIGO_CUPS_HOSPITALIZACION_PROHIBIDO and tipo_factura_str == "Hospitalización":
            problemas_cups_equivalentes.append({
                "factura": factura_str,
                "codigo": codigo_excluir,
                "codigo_equiv": "",
                "accion": ERROR_HOSPITALIZACION_NO_PERMITIDO,
                "procedimiento": proc_str,
            })
            logger.info(
                "REGLA (939402-Hospitalización): Fila %s: Código=%s, Tipo Factura=Hospitalización -> Error: %s",
                row, codigo_excluir, ERROR_HOSPITALIZACION_NO_PERMITIDO,
            )

        # ----- Regla: Código CUPS 12333 + Tipo Factura=Hospitalización -> Error
        # (Originalmente en pre-recorrido de hospitalización)
        if codigo_excluir == CODIGO_12333_HOSPITALIZACION_PROHIBIDO and tipo_factura_str == "Hospitalización":
            problemas_cups_equivalentes.append({
                "factura": factura_str,
                "codigo": codigo_excluir,
                "codigo_equiv": "",
                "accion": ERROR_12333_HOSPITALIZACION_NO_PERMITIDO,
                "procedimiento": proc_str,
            })
            logger.warning(
                "REGLA (12333-Hospitalización) AGREGADA: Fila %s: Código=%s, Tipo Factura=Hospitalización -> Error: %s",
                row, codigo_excluir, ERROR_12333_HOSPITALIZACION_NO_PERMITIDO,
            )

        # ----- Regla: Código = 890201 -> ERROR (debe usarse 890701)
        if codigo_excluir == "890201":
            logger.warning(
                "DETECTADO cups equiv error: factura=%s, codigo=%s", factura_str, codigo_excluir
            )
            problemas_cups_equivalentes.append({
                "factura": factura_str,
                "codigo": codigo_excluir,
                "codigo_equiv": "",
                "accion": "Usar 890701",
                "procedimiento": proc_str,
            })
            logger.info(
                "REGLA (Cups equivalentes): Fila %s: Código=%s -> debe usarse 890701",
                row, codigo_excluir,
            )

        # ----- Regla: Código = 129B01 -> ERROR (debe usarse 129B02)
        if codigo_excluir == "129B01":
            logger.warning(
                "DETECTADO cups equiv error: factura=%s, codigo=%s", factura_str, codigo_excluir
            )
            problemas_cups_equivalentes.append({
                "factura": factura_str,
                "codigo": codigo_excluir,
                "codigo_equiv": "",
                "accion": "Usar 129B02",
                "procedimiento": proc_str,
            })
            logger.info(
                "REGLA (Cups equivalentes): Fila %s: Código=%s -> debe usarse 129B02",
                row, codigo_excluir,
            )

        # ----- Regla: Código=890205 + Entidad distinta de ESS118/ESSC18 -> ERROR
        if codigo_excluir == CODIGO_CUPS_EQUIVALENTE_890205:
            if codigo_entidad_str not in ENTIDADES_PERMITIDAS_890205:
                logger.warning(
                    "DETECTADO cups equiv error: factura=%s, codigo=%s, entidad=%s",
                    factura_str, codigo_excluir, codigo_entidad_str,
                )
                problemas_cups_equivalentes.append({
                    "factura": factura_str,
                    "codigo": codigo_excluir,
                    "codigo_equiv": "",
                    "accion": f"Usar {CODIGO_CUPS_EQUIVALENTE_SUSTITUTO_890405}",
                    "procedimiento": proc_str,
                })
                logger.info(
                    "REGLA (Cups equivalentes): Fila %s: Código=%s, Entidad=%s -> debe usarse %s",
                    row, codigo_excluir, codigo_entidad_str,
                    CODIGO_CUPS_EQUIVALENTE_SUSTITUTO_890405,
                )

    if problemas_cups_equivalentes:
        logger.info("CUPS Equivalentes - Problemas encontrados: %d", len(problemas_cups_equivalentes))

    return problemas_cups_equivalentes
