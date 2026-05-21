"""Detector de cantidades SOAT anómalas en facturas de Urgencias.

Regla: Si Tarifario = "SOAT" y Tipo Factura Descripción = "Urgencias",
entonces los códigos 39145, 38114, 38915, 39131 deben tener cantidad = 1.
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import CODIGOS_SOAT_CANTIDAD_OBLIGATORIA, VALOR_TARIFARIO_SOAT
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_cantidades_soat_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta facturas SOAT Urgencias con códigos 39145, 38114, 38915, 39131
    que tienen cantidad != 1.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo", "procedimiento",
        "cantidad", "tipo_factura"
    """
    tipo_factura_idx = indices.get("tipo_factura_descripcion")
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")
    cantidad_idx = indices.get("cantidad")
    tarifario_idx = indices.get("tarifario")

    if None in (tipo_factura_idx, num_fact_idx, codigo_idx, cantidad_idx, tarifario_idx):
        logger.warning("Cantidades SOAT Urgencias - Columnas necesarias no encontradas")
        return []

    problemas = []
    facturas_procesadas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        tipo_factura = data_sheet.cell(row=row, column=tipo_factura_idx + 1).value
        tipo_factura_str = str(tipo_factura).strip() if tipo_factura else ""

        # Solo procesar si Tipo Factura = "Urgencias"
        if tipo_factura_str != "Urgencias":
            continue

        # Verificar si es tarifario SOAT
        tarifario = data_sheet.cell(row=row, column=tarifario_idx + 1).value
        tarifario_str = str(tarifario).strip().upper() if tarifario else ""
        if tarifario_str != VALOR_TARIFARIO_SOAT:
            continue

        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_procesadas:
            continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip().upper() if codigo else ""

        # Verificar si el código está en la lista de códigos SOAT con cantidad obligatoria = 1
        if codigo_str not in CODIGOS_SOAT_CANTIDAD_OBLIGATORIA:
            continue

        cantidad = data_sheet.cell(row=row, column=cantidad_idx + 1).value
        if not isinstance(cantidad, (int, float)):
            continue

        # Validar: cantidad debe ser = 1
        if cantidad != 1:
            procedimiento = ""
            if procedimiento_idx is not None:
                proc_value = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc_value).strip() if proc_value else ""

            problemas.append({
                "factura": factura_str,
                "codigo": codigo_str,
                "procedimiento": procedimiento,
                "cantidad": cantidad,
                "tipo_factura": tipo_factura_str,
            })
            facturas_procesadas.add(factura_str)
            logger.warning(
                "CANTIDAD SOAT URGENCIAS - Factura='%s', Código='%s', Cantidad=%s (debe ser =1)",
                factura_str, codigo_str, cantidad
            )

    if problemas:
        logger.info("Cantidades SOAT Urgencias - Problemas encontrados: %d", len(problemas))

    return problemas
