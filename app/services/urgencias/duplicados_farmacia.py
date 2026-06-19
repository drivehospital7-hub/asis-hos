"""Detector de revisión necesaria para código J07BG01 en Urgencias.

Marca filas donde el código sea "J07BG01" y la cantidad sea mayor a 1
en la misma factura. Requiere revisión manual.
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)

CODIGO_REVISION_FARMACIA = "J07BG01"


def detect_duplicados_farmacia(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """Detecta filas con código J07BG01 y cantidad > 1.

    Itera por todas las filas de la hoja. Si encuentra una fila donde
    el código es "J07BG01" y la cantidad es mayor a 1, la marca como
    revisión necesaria.

    Args:
        data_sheet: Hoja de Excel con los datos.
        indices: Índices de columnas.

    Returns:
        Lista de dicts con keys: factura, codigo, cantidad.
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    cantidad_idx = indices.get("cantidad")

    # Guard: columnas requeridas faltantes
    if None in (num_fact_idx, codigo_idx, cantidad_idx):
        logger.warning(
            "Duplicados Farmacia - Columnas necesarias no encontradas: "
            "numero_factura=%s, codigo=%s, cantidad=%s",
            num_fact_idx,
            codigo_idx,
            cantidad_idx,
        )
        return []

    resultados: list[dict[str, Any]] = []
    facturas_vistas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        # Leer código
        codigo_val = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo_val).strip().upper() if codigo_val else ""
        if codigo_str != CODIGO_REVISION_FARMACIA:
            continue

        # Evitar duplicados por factura
        if factura_str in facturas_vistas:
            continue

        # Leer cantidad
        cantidad_val = data_sheet.cell(row=row, column=cantidad_idx + 1).value
        if not isinstance(cantidad_val, (int, float)):
            continue
        if cantidad_val <= 1:
            continue

        facturas_vistas.add(factura_str)
        resultados.append({
            "factura": factura_str,
            "codigo": codigo_str,
            "cantidad": cantidad_val,
        })

    if resultados:
        logger.info(
            "Duplicados Farmacia (J07BG01>1) - %d facturas encontradas",
            len(resultados),
        )
    else:
        logger.info("Duplicados Farmacia - No se encontraron facturas con J07BG01>1")

    return resultados
