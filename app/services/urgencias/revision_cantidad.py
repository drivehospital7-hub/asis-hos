"""Detección de revisión necesaria por cantidad anómala en Urgencias.

Extraído de app/services/revision_sheet.py._detect_revision_cantidad_urgencias
como parte de la Fase 7 (cleanup).
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    CANTIDAD_MAX_02_LAB,
    CANTIDAD_MAX_02_LAB_903883,
    CANTIDAD_MAX_09_12,
    CODIGO_ESPECIAL_02_LAB,
    CODIGO_EXENTO_V03AN0101,
    CODIGOS_LIMITE_ESPECIFICO,
    CODIGOS_REVISION_CANTIDAD_EXENTOS,
    CODIGOS_TIPO_PROC_09_12,
    CODIGO_TIPO_PROCEDIMIENTO_REVISION_LAB,
    LABORATORIO_REVISION_EXENTO,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_revision_cantidad_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta filas con cantidad anómala que requieren revisión manual.

    Reglas:
    - General: Cantidad > 1 (excepto códigos exentos)
    - 02+Lab=No: Cantidad > 2 (código 903883: límite 5)
    - 09/12: Cantidad > 20 (código V03AN0101: siempre permitido)

    Returns:
        Lista de dicts con keys: 'factura', 'codigo', 'procedimiento',
        'cantidad', 'tipo_procedimiento', 'laboratorio'
    """
    tipo_factura_idx = indices.get("tipo_factura_descripcion")
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")
    cantidad_idx = indices.get("cantidad")
    tipo_proc_idx = indices.get("tipo_procedimiento")
    laboratorio_idx = indices.get("laboratorio")
    codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")

    if None in (tipo_factura_idx, num_fact_idx, codigo_idx, cantidad_idx, tipo_proc_idx, laboratorio_idx):
        logger.warning(
            "Revision Cantidad - Columnas necesarias no encontradas: "
            "tipo_factura=%s, numero_factura=%s, codigo=%s, cantidad=%s, "
            "tipo_procedimiento=%s, laboratorio=%s",
            tipo_factura_idx, num_fact_idx, codigo_idx, cantidad_idx,
            tipo_proc_idx, laboratorio_idx,
        )
        return []

    revision_items: list[dict[str, str]] = []

    for row in range(2, data_sheet.max_row + 1):
        tipo_factura = data_sheet.cell(row=row, column=tipo_factura_idx + 1).value
        tipo_factura_str = str(tipo_factura).strip() if tipo_factura else ""

        # Solo procesar si Tipo Factura = "Urgencias"
        if tipo_factura_str != "Urgencias":
            continue

        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip().upper() if codigo else ""

        # Excepción 1: código está en la lista de exentos
        if codigo_str in CODIGOS_REVISION_CANTIDAD_EXENTOS:
            continue

        cantidad = data_sheet.cell(row=row, column=cantidad_idx + 1).value
        if not isinstance(cantidad, (int, float)):
            continue

        # Excepción 2: código con límite específico
        if codigo_str in CODIGOS_LIMITE_ESPECIFICO:
            max_cant = CODIGOS_LIMITE_ESPECIFICO[codigo_str]
            if cantidad <= max_cant:
                continue

        tipo_proc = data_sheet.cell(row=row, column=tipo_proc_idx + 1).value
        tipo_proc_str = str(tipo_proc).strip() if tipo_proc else ""

        laboratorio = data_sheet.cell(row=row, column=laboratorio_idx + 1).value
        laboratorio_str = str(laboratorio).strip() if laboratorio else ""

        # Leer Código Tipo Procedimiento una sola vez
        codigo_tipo_proc_str = ""
        if codigo_tipo_proc_idx is not None:
            codigo_tipo_proc = data_sheet.cell(row=row, column=codigo_tipo_proc_idx + 1).value
            codigo_tipo_proc_str = str(codigo_tipo_proc).strip() if codigo_tipo_proc else ""

        # --- Regla para 02 + Lab=No: cantidad máxima 2 (903883: máximo 5) ---
        if (codigo_tipo_proc_str == CODIGO_TIPO_PROCEDIMIENTO_REVISION_LAB
                and laboratorio_str == LABORATORIO_REVISION_EXENTO):
            if codigo_str == CODIGO_ESPECIAL_02_LAB:
                if cantidad <= CANTIDAD_MAX_02_LAB_903883:
                    continue
            elif cantidad <= CANTIDAD_MAX_02_LAB:
                continue

        # --- Regla para 09/12: cantidad máxima 20 (excepto V03AN0101) ---
        if codigo_tipo_proc_str in CODIGOS_TIPO_PROC_09_12:
            if codigo_str == CODIGO_EXENTO_V03AN0101:
                continue
            if cantidad <= CANTIDAD_MAX_09_12:
                continue

        # --- Regla general: cantidad > 1 ---
        else:
            if cantidad <= 1:
                continue

        procedimiento = ""
        if procedimiento_idx is not None:
            proc_value = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
            procedimiento = str(proc_value).strip() if proc_value else ""

        revision_items.append({
            "factura": factura_str,
            "codigo": codigo_str,
            "procedimiento": procedimiento,
            "cantidad": cantidad,
            "tipo_procedimiento": tipo_proc_str,
            "laboratorio": laboratorio_str,
        })

    logger.info(
        "Revision Cantidad - Filas procesadas: %d, Items encontrados: %d",
        data_sheet.max_row - 1,
        len(revision_items),
    )
    return revision_items
