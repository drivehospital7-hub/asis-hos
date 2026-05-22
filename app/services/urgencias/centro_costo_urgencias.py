"""Detector de problemas de centro de costo en facturas de Urgencias.

Reglas extraídas de _detect_centro_costo_urgencias (revision_sheet.py):
- Regla 1 a 9 de centro de costo
- Reglas REVERSE de centro de costo
- Reglas de tipo factura (Intramural, Ambulatoria, Hospitalización)
- Validación de centro de costo válido para Urgencias
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    CENTRO_COSTO_APOYO_DIAGNOSTICO,
    CENTRO_COSTO_FARMACIA,
    CENTRO_COSTO_HOSPITALIZACION_ESTANCIA,
    CENTRO_COSTO_LABORATORIO_URGENCIAS,
    CENTRO_COSTO_PYP_URGENCIAS,
    CENTRO_COSTO_QUIROFANO_URGENCIAS,
    CENTRO_COSTO_TRASLADOS,
    CENTRO_COSTO_URGENCIAS,
    CENTROS_COSTO_VALIDOS_URGENCIAS,
    CODIGOS_EXCEPTUADOS,
    CODIGOS_HOSPITALIZACION_ESTANCIA,
    CODIGOS_LABORATORIO_URGENCIAS,
    CODIGOS_LABORATORIO_URGENCIAS_REVERSE,
    CODIGOS_PYP_URGENCIAS,
    CODIGOS_QUIROFANO_URGENCIAS,
    CODIGO_TIPO_PROCEDIMIENTO_DIAGNOSTICO,
    CODIGO_TIPO_PROCEDIMIENTO_TRASLADOS,
    LABORATORIO_NO,
    VALOR_TARIFARIO_FARMACIA,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_centro_costo_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta facturas con problemas de centro de costo en Urgencias.

    Reglas de centro de costo (1-9), REVERSE, y reglas de tipo de factura.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "tipo_factura", "centro_actual",
        "centro_deberia", "codigo", "procedimiento", "prioridad", "regla"
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")
    codigo_idx = indices.get("codigo")
    laboratorio_idx = indices.get("laboratorio")
    centro_costo_idx = indices.get("centro_costo")
    codigo_entidad_cobrar_idx = indices.get("codigo_entidad_cobrar")
    tipo_factura_descripcion_idx = indices.get("tipo_factura_descripcion")
    proc_idx = indices.get("procedimiento")
    tarifario_idx = indices.get("tarifario")

    if num_fact_idx is None or centro_costo_idx is None:
        logger.warning("Centro Costo Urgencias - Columnas necesarias no encontradas")
        return []

    problemas_centros: list[dict[str, Any]] = []

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        # Leer valores de columnas
        codigo_tipo_proc = (
            data_sheet.cell(row=row, column=codigo_tipo_proc_idx + 1).value
            if codigo_tipo_proc_idx is not None else None
        )
        codigo = (
            data_sheet.cell(row=row, column=codigo_idx + 1).value
            if codigo_idx is not None else None
        )
        laboratorio = (
            data_sheet.cell(row=row, column=laboratorio_idx + 1).value
            if laboratorio_idx is not None else None
        )
        centro_costo = (
            data_sheet.cell(row=row, column=centro_costo_idx + 1).value
        )
        codigo_entidad_cobrar = (
            data_sheet.cell(row=row, column=codigo_entidad_cobrar_idx + 1).value
            if codigo_entidad_cobrar_idx is not None else None
        )
        tipo_factura_descripcion = (
            data_sheet.cell(row=row, column=tipo_factura_descripcion_idx + 1).value
            if tipo_factura_descripcion_idx is not None else None
        )
        procedimiento = (
            data_sheet.cell(row=row, column=proc_idx + 1).value
            if proc_idx is not None else None
        )
        tarifario = (
            data_sheet.cell(row=row, column=tarifario_idx + 1).value
            if tarifario_idx is not None else None
        )

        # Normalizar strings
        codigo_str = str(codigo_tipo_proc).strip() if codigo_tipo_proc else ""
        codigo_excluir = str(codigo).strip() if codigo else ""
        laboratorio_str = str(laboratorio).strip() if laboratorio else ""
        centro_costo_str = str(centro_costo).strip() if centro_costo else ""
        codigo_entidad_str = str(codigo_entidad_cobrar).strip() if codigo_entidad_cobrar else ""
        tipo_factura_str = str(tipo_factura_descripcion).strip() if tipo_factura_descripcion else ""
        proc_str = str(procedimiento).strip() if procedimiento else ""
        tarifario_str = str(tarifario).strip() if tarifario else ""

        # ----- Regla: Centro de costo debe ser uno de los valores válidos
        if centro_costo_str and centro_costo_str not in CENTROS_COSTO_VALIDOS_URGENCIAS:
            problemas_centros.append({
                "factura": factura_str,
                "tipo_factura": tipo_factura_str,
                "centro_actual": centro_costo_str,
                "centro_deberia": "Centro de costo no válido para Urgencias",
                "codigo": codigo_excluir,
                "procedimiento": proc_str,
                "prioridad": 1,
                "regla": "CENTRO_INVALIDO",
            })
            logger.info(
                "REGLA (CENTRO INVALIDO): Fila %s: Centro=%s no está en lista válida",
                row, centro_costo_str,
            )

        # ----- Regla 9: Tarifario="Suministros, Medicamentos" -> Centro debe ser FARMACIA
        if tarifario_str == VALOR_TARIFARIO_FARMACIA:
            if centro_costo_str != CENTRO_COSTO_FARMACIA:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": CENTRO_COSTO_FARMACIA,
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "REGLA9",
                })
                logger.info(
                    "REGLA9: Fila %s: Tarifario=%s, Centro=%s (debe ser %s)",
                    row, tarifario_str, centro_costo_str, CENTRO_COSTO_FARMACIA,
                )

        # ----- Regla 1: Código=02 + Laboratorio=No + Centro != IMAGENOLOGIA
        regla_1_activa = (
            codigo_str == CODIGO_TIPO_PROCEDIMIENTO_DIAGNOSTICO
            and laboratorio_str == LABORATORIO_NO
        )
        es_exceptuado = codigo_excluir in CODIGOS_EXCEPTUADOS
        if regla_1_activa and not es_exceptuado and centro_costo_str != CENTRO_COSTO_APOYO_DIAGNOSTICO:
            problemas_centros.append({
                "factura": factura_str,
                "tipo_factura": tipo_factura_str,
                "centro_actual": centro_costo_str,
                "centro_deberia": CENTRO_COSTO_APOYO_DIAGNOSTICO,
                "codigo": codigo_excluir,
                "procedimiento": proc_str,
                "prioridad": 1,
            })
            logger.info(
                "REGLA1: Fila %s: Código=02, Lab=No, Centro incorrecto (Centro: '%s', CódigoProc: '%s')",
                row, centro_costo, codigo_excluir,
            )

        # ----- Regla 1 REVERSE: Centro=APOYO DIAGNOSTICO -> Código=02 Y Laboratorio=No
        if centro_costo_str == CENTRO_COSTO_APOYO_DIAGNOSTICO:
            if codigo_str != CODIGO_TIPO_PROCEDIMIENTO_DIAGNOSTICO or laboratorio_str != LABORATORIO_NO:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": f"Código={CODIGO_TIPO_PROCEDIMIENTO_DIAGNOSTICO} + Laboratorio={LABORATORIO_NO}",
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "REVERSE1",
                })
                logger.info(
                    "REGLA1-REVERSE: Fila %s: Centro=APOYO DIAGNOSTICO pero Código=%s (debe ser 02) o Lab=%s (debe ser No)",
                    row, codigo_str, laboratorio_str,
                )

        # ----- Regla 2: Código=14 + Centro != TRASLADOS
        if codigo_str == CODIGO_TIPO_PROCEDIMIENTO_TRASLADOS:
            if centro_costo_str != CENTRO_COSTO_TRASLADOS:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": CENTRO_COSTO_TRASLADOS,
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                })
                logger.info(
                    "REGLA2: Fila %s: Código=14, Centro distinto a TRASLADOS", row,
                )

        # ----- Regla 2 REVERSE: Centro=TRASLADOS -> Código debe ser 14
        if centro_costo_str == CENTRO_COSTO_TRASLADOS:
            if codigo_str != CODIGO_TIPO_PROCEDIMIENTO_TRASLADOS:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": f"Código={CODIGO_TIPO_PROCEDIMIENTO_TRASLADOS}",
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "REVERSE2",
                })
                logger.info(
                    "REGLA2-REVERSE: Fila %s: Centro=TRASLADOS pero Código=%s (debe ser 14)",
                    row, codigo_str,
                )

        # ----- Regla 3: Código en PYP + Centro != PROCEDIMIENTO PYP
        if codigo_excluir in CODIGOS_PYP_URGENCIAS:
            if centro_costo_str != CENTRO_COSTO_PYP_URGENCIAS:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": CENTRO_COSTO_PYP_URGENCIAS,
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                })

        # ----- Regla 3 REVERSE: Centro=PYP -> Código debe ser PYP
        if centro_costo_str == CENTRO_COSTO_PYP_URGENCIAS:
            if codigo_excluir not in CODIGOS_PYP_URGENCIAS:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": "Procedimiento con mal uso de centro de costo PYP",
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "REVERSE3",
                })
                logger.info(
                    "REGLA3-REVERSE: Fila %s: Centro=PYP pero Código=%s (debe ser alguno de %s)",
                    row, codigo_excluir, CODIGOS_PYP_URGENCIAS,
                )

        # ----- Regla 4: Código en QUIRÓFANOS + Centro != QUIRÓFANOS
        if codigo_excluir in CODIGOS_QUIROFANO_URGENCIAS:
            if centro_costo_str != CENTRO_COSTO_QUIROFANO_URGENCIAS:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": CENTRO_COSTO_QUIROFANO_URGENCIAS,
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                })
                logger.info(
                    "REGLA4: Fila %s: Código=%s, Centro incorrecto (Centro: '%s')",
                    row, codigo_excluir, centro_costo_str,
                )

        # ----- Regla 4 REVERSE: Centro=QUIRÓFANOS -> Código debe ser QUIRÓFANOS
        if centro_costo_str == CENTRO_COSTO_QUIROFANO_URGENCIAS:
            if codigo_excluir not in CODIGOS_QUIROFANO_URGENCIAS:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": "Procedimiento con mal uso de centro de costo QUIRÓFANOS",
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "REVERSE4",
                })
                logger.info(
                    "REGLA4-REVERSE: Fila %s: Centro=QUIRÓFANOS pero Código=%s (debe ser alguno de %s)",
                    row, codigo_excluir, CODIGOS_QUIROFANO_URGENCIAS,
                )

        # ----- Regla 5: Código en laboratorio + ESS118 + Intramural -> Centro LABORATORIO
        if codigo_excluir in CODIGOS_LABORATORIO_URGENCIAS:
            if codigo_entidad_str == "ESS118" and tipo_factura_str == "Intramural":
                centro_valido = centro_costo_str in (
                    CENTRO_COSTO_LABORATORIO_URGENCIAS,
                    f"{CENTRO_COSTO_LABORATORIO_URGENCIAS}.",
                )
                if not centro_valido:
                    problemas_centros.append({
                        "factura": factura_str,
                        "tipo_factura": tipo_factura_str,
                        "centro_actual": centro_costo_str,
                        "centro_deberia": CENTRO_COSTO_LABORATORIO_URGENCIAS,
                        "codigo": codigo_excluir,
                        "procedimiento": proc_str,
                        "prioridad": 1,
                    })
                    logger.info(
                        "REGLA5: Fila %s: Código=%s, ESS118+Intramural, Centro incorrecto (Centro: '%s')",
                        row, codigo_excluir, centro_costo_str,
                    )

        # ----- Regla 5 REVERSE: Centro=LABORATORIO -> Código en lista + Tipo=Intramural
        if centro_costo_str == CENTRO_COSTO_LABORATORIO_URGENCIAS:
            tipo_id_idx = indices.get("tipo_identificacion")
            tipo_identificacion_val = None
            if tipo_id_idx is not None:
                tipo_identificacion_val = data_sheet.cell(row=row, column=tipo_id_idx + 1).value
            tipo_identificacion_str = str(tipo_identificacion_val).strip() if tipo_identificacion_val else ""

            if tipo_identificacion_str == "CN":
                codigos_validos = CODIGOS_LABORATORIO_URGENCIAS | CODIGOS_LABORATORIO_URGENCIAS_REVERSE
            else:
                codigos_validos = CODIGOS_LABORATORIO_URGENCIAS

            if tipo_factura_str != "Intramural" or codigo_excluir not in codigos_validos:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": "Procedimiento con mal uso de centro de costo LABORATORIO + Tipo=Intramural",
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "REVERSE5",
                })
                logger.info(
                    "REGLA5-REVERSE: Fila %s: Centro=LABORATORIO pero Tipo=%s (debe ser Intramural) o Código=%s (no está en lista)",
                    row, tipo_factura_str, codigo_excluir,
                )

        # ----- Regla 9 REVERSE: Centro=FARMACIA -> Tarifario debe ser farmacia
        if centro_costo_str == CENTRO_COSTO_FARMACIA:
            if tarifario_str != VALOR_TARIFARIO_FARMACIA:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": f"Tarifario debe ser {VALOR_TARIFARIO_FARMACIA}",
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "REVERSE9",
                })
                logger.info(
                    "REGLA9-REVERSE: Fila %s: Centro=FARMACIA pero Tarifario=%s (debe ser %s)",
                    row, tarifario_str, VALOR_TARIFARIO_FARMACIA,
                )

        # ----- Regla: Tipo=Intramural + Cód Entidad != ESS118 -> Centro LABORATORIO
        if tipo_factura_str == "Intramural" and codigo_entidad_str and codigo_entidad_str != "ESS118":
            if centro_costo_str != CENTRO_COSTO_LABORATORIO_URGENCIAS:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": CENTRO_COSTO_LABORATORIO_URGENCIAS,
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "INTRAMURAL_OTRAS_ENTIDADES",
                })
                logger.info(
                    "REGLA (INTRAMURAL-NO-ESS118): Fila %s: Tipo=Intramural, Entidad=%s (no ESS118), Centro=%s (debe ser LABORATORIO)",
                    row, codigo_entidad_str, centro_costo_str,
                )

        # ----- Regla: Tipo Factura=Ambulatoria -> Centro debe ser PYP
        if tipo_factura_str == "Ambulatoria":
            if centro_costo_str != CENTRO_COSTO_PYP_URGENCIAS:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": CENTRO_COSTO_PYP_URGENCIAS,
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                    "regla": "AMBULATORIA_PYP",
                })
                logger.info(
                    "REGLA (AMBULATORIA-PYP): Fila %s: Tipo=Ambulatoria pero Centro=%s (debe ser PYP)",
                    row, centro_costo_str,
                )

        # ----- Regla 8: Código en {890601H, 39133} -> Centro = HOSPITALIZACIÓN - ESTANCIA GENERAL
        if codigo_excluir in CODIGOS_HOSPITALIZACION_ESTANCIA:
            if centro_costo_str != CENTRO_COSTO_HOSPITALIZACION_ESTANCIA:
                problemas_centros.append({
                    "factura": factura_str,
                    "tipo_factura": tipo_factura_str,
                    "centro_actual": centro_costo_str,
                    "centro_deberia": CENTRO_COSTO_HOSPITALIZACION_ESTANCIA,
                    "codigo": codigo_excluir,
                    "procedimiento": proc_str,
                    "prioridad": 1,
                })
                logger.info(
                    "REGLA8: Fila %s: Código=%s, Centro=%s (debe ser %s)",
                    row, codigo_excluir, centro_costo_str, CENTRO_COSTO_HOSPITALIZACION_ESTANCIA,
                )

        # ----- Regla: Tipo Factura=Hospitalización + Centro=URGENCIAS -> Error
        if tipo_factura_str == "Hospitalización" and centro_costo_str == CENTRO_COSTO_URGENCIAS:
            problemas_centros.append({
                "factura": factura_str,
                "tipo_factura": tipo_factura_str,
                "centro_actual": centro_costo_str,
                "centro_deberia": CENTRO_COSTO_HOSPITALIZACION_ESTANCIA,
                "codigo": codigo_excluir,
                "procedimiento": proc_str,
                "prioridad": 2,
            })

        # ----- Regla: Tipo Factura=Urgencias + Centro=HOSPITALIZACIÓN -> Error
        if tipo_factura_str == "Urgencias" and centro_costo_str == CENTRO_COSTO_HOSPITALIZACION_ESTANCIA:
            problemas_centros.append({
                "factura": factura_str,
                "tipo_factura": tipo_factura_str,
                "centro_actual": centro_costo_str,
                "centro_deberia": CENTRO_COSTO_URGENCIAS,
                "codigo": codigo_excluir,
                "procedimiento": proc_str,
                "prioridad": 2,
            })

    if problemas_centros:
        logger.info("Centro Costo Urgencias - Problemas encontrados: %d", len(problemas_centros))

    return problemas_centros
