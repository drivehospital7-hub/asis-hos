"""Detector de sala de observación en facturas de Urgencias.

Extraído de _detect_centro_costo_urgencias (revision_sheet.py).

Reglas:
- Estancia en sala de observación según entidad y horas
- Códigos obligatorios (890701, 890601) cuando hay sala
- Reglas SOAT de sala de observación (38114, 38915, 39145, 39131)
- Códigos prohibidos por entidad (129B02 para ESS118/ESSC18)
- 890601H prohibido en Urgencias
- 05DSB01 prohibido en entidades no ESS
"""

from __future__ import annotations

import logging
from datetime import datetime
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    CODIGOS_SALA_OBSERVACION_OBLIGATORIOS,
    CODIGO_SALA_OBSERVACION_CORTA,
    CODIGO_SALA_OBSERVACION_LARGA_ESS,
    CODIGO_SALA_OBSERVACION_LARGA_OTRAS,
    CODIGO_SALA_PROHIBIDO_ESS,
    CODIGOS_SOAT_OBLIGATORIOS_SALA,
    CODIGO_SOAT_SALA_OBSERVACION_CORTA,
    CODIGO_SOAT_SALA_OBSERVACION_LARGA,
    CODIGO_SOAT_URGENCIAS_PROHIBIDO,
    CODIGO_05DSB01_PROHIBIDO_OTRAS,
    CODIGO_URGENCIAS_PROHIBIDO,
    ENTIDADES_ESS_PERMITIDO_05DSB01,
    ENTIDADES_ESS_PROHIBIDO_129B02,
    ENTIDADES_SALA_OBSERVACION_05DSB01,
    VALOR_TARIFARIO_SOAT,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def _format_estancia(horas: float | None) -> str:
    """Formatea estancia en días + horas."""
    if horas is None:
        return "N/A"
    dias = int(horas // 24)
    hrs = int(horas % 24)
    if dias > 0:
        return f"{dias}d {hrs}h"
    return f"{hrs}h"


def _build_sala_proc(factura: str, codigos_sala: set[str], factura_sala_procedimiento: dict[str, str]) -> str:
    """Construye 'código - nombre' real del Excel para sala de observación."""
    if not codigos_sala:
        return ""
    proc_nombre = factura_sala_procedimiento.get(factura, "")
    primer_codigo = next(iter(codigos_sala), "")
    if primer_codigo and proc_nombre:
        return f"{primer_codigo} - {proc_nombre}"
    return primer_codigo


def detect_sala_observacion(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta problemas de sala de observación en facturas de Urgencias.

    Dos pasos:
    1. Recolectar datos por factura (entidad, estancia, códigos de sala)
    2. Validar reglas de sala de observación

    Returns:
        Lista de dicts con keys compatibles con problemas_cups_equivalentes:
        "factura", "codigo", "codigo_equiv", "accion", "procedimiento", "estancia_str"
    """
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    codigo_entidad_cobrar_idx = indices.get("codigo_entidad_cobrar")
    tipo_factura_descripcion_idx = indices.get("tipo_factura_descripcion")
    proc_idx = indices.get("procedimiento")
    tarifario_idx = indices.get("tarifario")
    fec_factura_idx = indices.get("fec_factura")
    fecha_cierre_idx = indices.get("fecha_cierre")

    if num_fact_idx is None or tipo_factura_descripcion_idx is None:
        logger.warning("Sala Observación - Columnas necesarias no encontradas")
        return []

    # ----- Paso 1: Recolectar datos de sala de observación por factura
    factura_sala_data: dict[str, dict[str, Any]] = {}
    factura_sala_procedimiento: dict[str, str] = {}

    for row in range(2, data_sheet.max_row + 1):
        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        tipo_factura_cell = data_sheet.cell(row=row, column=tipo_factura_descripcion_idx + 1).value
        tipo_factura_str = str(tipo_factura_cell).strip() if tipo_factura_cell else ""

        # Solo procesar facturas de Urgencias
        if tipo_factura_str != "Urgencias":
            continue

        # Leer tarifario de esta fila
        tarifario_cell = data_sheet.cell(row=row, column=tarifario_idx + 1).value if tarifario_idx else None
        tarifario_str = str(tarifario_cell).strip() if tarifario_cell else ""

        if factura_str not in factura_sala_data:
            # Entidad
            entidad_cell = data_sheet.cell(row=row, column=codigo_entidad_cobrar_idx + 1).value if codigo_entidad_cobrar_idx else None
            entidad_str = str(entidad_cell).strip() if entidad_cell else ""

            # Fechas para estancia
            fec_factura_cell = data_sheet.cell(row=row, column=fec_factura_idx + 1).value if fec_factura_idx else None
            fecha_cierre_cell = data_sheet.cell(row=row, column=fecha_cierre_idx + 1).value if fecha_cierre_idx else None

            estancia_horas = None
            if fec_factura_cell and fecha_cierre_cell:
                try:
                    fec_factura_dt = datetime.strptime(str(fec_factura_cell).strip(), "%Y-%m-%d %H:%M:%S")
                    fecha_cierre_dt = datetime.strptime(str(fecha_cierre_cell).strip(), "%Y-%m-%d %H:%M:%S")
                    diferencia = fecha_cierre_dt - fec_factura_dt
                    estancia_horas = diferencia.total_seconds() / 3600
                except (ValueError, TypeError):
                    estancia_horas = None

            factura_sala_data[factura_str] = {
                "entidad": entidad_str,
                "tipo_factura": tipo_factura_str,
                "estancia_horas": estancia_horas,
                "tarifario": tarifario_str,
                "codigos_sala": set(),
                "codigos_urgencias_obligatorios": set(),
                "codigos_soat_obligatorios": set(),
                "tiene_890601h": False,
                "tiene_soat_prohibido": False,
            }
        else:
            # Actualizar tarifario: SOAT siempre tiene prioridad
            if tarifario_str and tarifario_str.upper() == VALOR_TARIFARIO_SOAT:
                factura_sala_data[factura_str]["tarifario"] = tarifario_str

        # Recolectar códigos de sala de observación
        codigo_cell = data_sheet.cell(row=row, column=codigo_idx + 1).value if codigo_idx else None
        proc_cell = data_sheet.cell(row=row, column=proc_idx + 1).value if proc_idx else None
        if codigo_cell:
            codigo_normalized = str(codigo_cell).strip()
            if codigo_normalized in (
                CODIGO_SALA_OBSERVACION_CORTA,
                CODIGO_SALA_OBSERVACION_LARGA_ESS,
                CODIGO_SALA_OBSERVACION_LARGA_OTRAS,
                CODIGO_SOAT_SALA_OBSERVACION_LARGA,
                CODIGO_SOAT_SALA_OBSERVACION_CORTA,
            ):
                factura_sala_data[factura_str]["codigos_sala"].add(codigo_normalized)
                if proc_cell:
                    factura_sala_procedimiento[factura_str] = str(proc_cell).strip()

            # Recolectar códigos obligatorios de urgencias (890701 y 890601)
            if codigo_normalized in CODIGOS_SALA_OBSERVACION_OBLIGATORIOS:
                factura_sala_data[factura_str]["codigos_urgencias_obligatorios"].add(codigo_normalized)

            # Recolectar códigos obligatorios SOAT si tiene sala SOAT
            if tarifario_str.upper() == VALOR_TARIFARIO_SOAT and codigo_normalized in CODIGOS_SOAT_OBLIGATORIOS_SALA:
                factura_sala_data[factura_str]["codigos_soat_obligatorios"].add(codigo_normalized)

            # Colectar si tiene código prohibido SOAT Urgencias (39133)
            if tarifario_str.upper() == VALOR_TARIFARIO_SOAT and codigo_normalized == CODIGO_SOAT_URGENCIAS_PROHIBIDO:
                factura_sala_data[factura_str]["tiene_soat_prohibido"] = True

            # Colectar si tiene 890601H (prohibido en Urgencias)
            if codigo_normalized == CODIGO_URGENCIAS_PROHIBIDO:
                factura_sala_data[factura_str]["tiene_890601h"] = True

    # ----- Paso 2: Validar problemas de sala de observación
    problemas_cups_equivalentes: list[dict[str, Any]] = []
    facturas_estancia_reportadas: set[str] = set()
    facturas_urgencia_reportadas: set[str] = set()
    facturas_ess_129b02_reportadas: set[str] = set()

    for factura_str, data in factura_sala_data.items():
        entidad = data["entidad"]
        estancia_horas = data["estancia_horas"]
        codigos_sala = data["codigos_sala"]
        tarifario = data.get("tarifario", "")
        tipo_factura = data.get("tipo_factura", "")

        if estancia_horas is None:
            continue

        es_problema = False
        accion = ""
        codigo_requerido = None

        # ----- Regla SOAT: Sala de Observación
        if tarifario.upper() == VALOR_TARIFARIO_SOAT:
            if estancia_horas <= 2:
                codigo_requerido = None
            elif estancia_horas > 6:
                codigo_requerido = CODIGO_SOAT_SALA_OBSERVACION_LARGA
            else:
                codigo_requerido = CODIGO_SOAT_SALA_OBSERVACION_CORTA

            if codigo_requerido is not None and codigo_requerido not in codigos_sala:
                es_problema = True
                accion = f"Usar {codigo_requerido}"
        else:
            if estancia_horas <= 2:
                codigo_requerido = None
            elif entidad in ENTIDADES_SALA_OBSERVACION_05DSB01:
                codigo_requerido = CODIGO_SALA_OBSERVACION_LARGA_ESS if estancia_horas > 6 else CODIGO_SALA_OBSERVACION_CORTA
            else:
                codigo_requerido = CODIGO_SALA_OBSERVACION_LARGA_OTRAS if estancia_horas > 6 else CODIGO_SALA_OBSERVACION_CORTA

            if codigo_requerido is None:
                codigos_incorrectos = codigos_sala - {CODIGO_SALA_OBSERVACION_CORTA}
                if codigos_incorrectos:
                    es_problema = True
                    accion = f"Estancia <2h: no requiere sala observación (si tiene, usar 5DSB01)"
            elif codigo_requerido not in codigos_sala:
                es_problema = True
                accion = f"Usar {codigo_requerido}"

        # ----- Regla: Si tiene sala, ESS118/ESSC18 no puede tener 129B02
        codigos_urgencia_obligatorios = data.get("codigos_urgencias_obligatorios", set())
        if codigos_sala:
            if entidad in ENTIDADES_ESS_PROHIBIDO_129B02 and CODIGO_SALA_PROHIBIDO_ESS in codigos_sala:
                if factura_str not in facturas_ess_129b02_reportadas:
                    estancia_str = _format_estancia(estancia_horas)
                    problemas_cups_equivalentes.append({
                        "factura": factura_str,
                        "codigo": list(codigos_sala),
                        "codigo_equiv": "",
                        "accion": f"ESS118/ESSC18 no puede tener {CODIGO_SALA_PROHIBIDO_ESS} - usar 05DSB01 para >6h",
                        "procedimiento": _build_sala_proc(factura_str, codigos_sala, factura_sala_procedimiento),
                        "estancia_str": estancia_str,
                    })
                    facturas_ess_129b02_reportadas.add(factura_str)

            # ----- Regla: Urgencias no puede tener 890601H
            tiene_890601h = data.get("tiene_890601h", False)
            if tipo_factura == "Urgencias" and tiene_890601h:
                if factura_str not in facturas_urgencia_reportadas:
                    estancia_str = _format_estancia(estancia_horas)
                    problemas_cups_equivalentes.append({
                        "factura": factura_str,
                        "codigo": list(codigos_sala),
                        "codigo_equiv": "",
                        "accion": f"Urgencias no puede tener {CODIGO_URGENCIAS_PROHIBIDO}",
                        "procedimiento": _build_sala_proc(factura_str, codigos_sala, factura_sala_procedimiento),
                        "estancia_str": estancia_str,
                    })
                    facturas_urgencia_reportadas.add(factura_str)

            # ----- Regla: 05DSB01 prohibido en entidades no ESS
            if tipo_factura == "Urgencias" and CODIGO_05DSB01_PROHIBIDO_OTRAS in codigos_sala:
                if entidad not in ENTIDADES_ESS_PERMITIDO_05DSB01:
                    if factura_str not in facturas_urgencia_reportadas:
                        estancia_str = _format_estancia(estancia_horas)
                        problemas_cups_equivalentes.append({
                            "factura": factura_str,
                            "codigo": list(codigos_sala),
                            "codigo_equiv": "",
                            "accion": f"Entidad {entidad} no puede tener {CODIGO_05DSB01_PROHIBIDO_OTRAS} - usar 5DSB01 o 129B02",
                            "procedimiento": _build_sala_proc(factura_str, codigos_sala, factura_sala_procedimiento),
                            "estancia_str": estancia_str,
                        })
                        facturas_urgencia_reportadas.add(factura_str)

            # ----- Validar códigos obligatorios no-SOAT (890701 y 890601)
            if tarifario.upper() != VALOR_TARIFARIO_SOAT and tipo_factura == "Urgencias":
                activadores_no_soat = {
                    CODIGO_SALA_OBSERVACION_CORTA,
                    CODIGO_SALA_OBSERVACION_LARGA_ESS,
                    CODIGO_SALA_OBSERVACION_LARGA_OTRAS,
                }
                if codigos_sala & activadores_no_soat:
                    faltan_obligatorios = CODIGOS_SALA_OBSERVACION_OBLIGATORIOS - codigos_urgencia_obligatorios
                    if faltan_obligatorios:
                        if factura_str not in facturas_urgencia_reportadas:
                            estancia_str = _format_estancia(estancia_horas)
                            problemas_cups_equivalentes.append({
                                "factura": factura_str,
                                "codigo": list(codigos_sala),
                                "codigo_equiv": "",
                                "accion": f"Agregar códigos obligatorios: {', '.join(faltan_obligatorios)}",
                                "procedimiento": _build_sala_proc(factura_str, codigos_sala, factura_sala_procedimiento),
                                "estancia_str": estancia_str,
                            })
                            facturas_urgencia_reportadas.add(factura_str)

            # ----- Regla SOAT: Si tiene 38114 o 38915 -> debe tener 39145 y 39131
            codigos_soat_obligatorios = data.get("codigos_soat_obligatorios", set())
            if codigos_sala & {CODIGO_SOAT_SALA_OBSERVACION_LARGA, CODIGO_SOAT_SALA_OBSERVACION_CORTA}:
                if tipo_factura == "Urgencias" and tarifario.upper() == VALOR_TARIFARIO_SOAT:
                    faltan_soat = CODIGOS_SOAT_OBLIGATORIOS_SALA - codigos_soat_obligatorios
                    if faltan_soat:
                        if factura_str not in facturas_urgencia_reportadas:
                            estancia_str = _format_estancia(estancia_horas)
                            problemas_cups_equivalentes.append({
                                "factura": factura_str,
                                "codigo": list(codigos_sala),
                                "codigo_equiv": "",
                                "accion": f"SOAT Urgencias debe tener: {', '.join(faltan_soat)}",
                                "procedimiento": _build_sala_proc(factura_str, codigos_sala, factura_sala_procedimiento),
                                "estancia_str": estancia_str,
                            })
                            facturas_urgencia_reportadas.add(factura_str)

            # ----- Regla SOAT: Urgencias NO puede tener código 39133
            tiene_soat_prohibido = data.get("tiene_soat_prohibido", False)
            if (tipo_factura == "Urgencias"
                    and tarifario.upper() == VALOR_TARIFARIO_SOAT
                    and tiene_soat_prohibido):
                if factura_str not in facturas_urgencia_reportadas:
                    estancia_str = _format_estancia(estancia_horas)
                    problemas_cups_equivalentes.append({
                        "factura": factura_str,
                        "codigo": list(codigos_sala),
                        "codigo_equiv": "",
                        "accion": f"SOAT Urgencias no puede tener {CODIGO_SOAT_URGENCIAS_PROHIBIDO}",
                        "procedimiento": _build_sala_proc(factura_str, codigos_sala, factura_sala_procedimiento),
                        "estancia_str": estancia_str,
                    })
                    facturas_urgencia_reportadas.add(factura_str)

        # ----- Reportar error de estancia (no SOAT)
        if es_problema:
            if factura_str not in facturas_estancia_reportadas:
                estancia_str = _format_estancia(estancia_horas)
                problemas_cups_equivalentes.append({
                    "factura": factura_str,
                    "codigo": ", ".join(codigos_sala) if codigos_sala else "ninguno",
                    "codigo_equiv": "",
                    "accion": accion,
                    "procedimiento": _build_sala_proc(factura_str, codigos_sala, factura_sala_procedimiento),
                    "estancia_str": estancia_str,
                })
                facturas_estancia_reportadas.add(factura_str)

    if problemas_cups_equivalentes:
        logger.info("Sala Observación - Problemas encontrados: %d", len(problemas_cups_equivalentes))

    return problemas_cups_equivalentes
