"""Detector de problemas de IDE Contrato en facturas de Urgencias.

Utiliza el mapping de reglas en ide_contrato_rules.py y las reglas genéricas
de app.constants.urgencias para validar IDE Contrato.

Reglas extraídas de _detect_centro_costo_urgencias (revision_sheet.py):
- ~30 reglas de IDE Contrato por (código, entidad)
- Reglas condicionales por inserción (código 861801)
- Regla especial ESSC62 por presencia de 890405
- Reglas genéricas entidad->contrato (Regla 29)
- Reglas de entidades con múltiples contratos (Regla 30)
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    URGENCIA_ENTIDAD_CONTRATO,
    URGENCIA_ENTIDAD_MULTIPLE_CONTRATO,
)
from app.services.transversales.normalize import normalize_invoice
from app.services.urgencias.ide_contrato_rules import (
    IDE_INSERTION_RULES,
    IDE_MULTIPLE_RULES,
    IDE_SIMPLE_RULES,
    IDE_ESSC62_890405_RULES,
)

logger = logging.getLogger(__name__)


def detect_ide_contrato_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta facturas con problemas de IDE Contrato en Urgencias.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo", "entidad",
        "ide_contrato_actual", "ide_contrato_deberia" y opcionales
    """
    tipo_factura_idx = indices.get("tipo_factura_descripcion")
    num_fact_idx = indices.get("numero_factura")
    ident_idx = indices.get("identificacion")
    codigo_idx = indices.get("codigo")
    codigo_entidad_cobrar_idx = indices.get("codigo_entidad_cobrar")
    proc_idx = indices.get("procedimiento")
    ide_contrato_idx = indices.get("ide_contrato")

    if None in (tipo_factura_idx, num_fact_idx, codigo_idx, ide_contrato_idx, codigo_entidad_cobrar_idx):
        logger.warning("IDE Contrato Urgencias - Columnas necesarias no encontradas")
        return []

    # ----- Pre-recorrido 1: Identificaciones que tienen código 861801
    # (para reglas condicionales de inserción)
    identificaciones_con_insercion: set[str] = set()
    if ident_idx is not None and codigo_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero_ident = data_sheet.cell(row=row, column=ident_idx + 1).value
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            if numero_ident and codigo:
                ident_normalized = str(numero_ident).strip()
                codigo_normalized = str(codigo).strip()
                if codigo_normalized == "861801":
                    identificaciones_con_insercion.add(ident_normalized)

    # ----- Pre-recorrido 2: Identificaciones que tienen código 890405
    # (para regla ESSC62 que usa 890405 en lugar de 861801)
    identificaciones_con_890405: set[str] = set()
    if ident_idx is not None and codigo_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero_ident = data_sheet.cell(row=row, column=ident_idx + 1).value
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            if numero_ident and codigo:
                ident_normalized = str(numero_ident).strip()
                codigo_normalized = str(codigo).strip()
                if codigo_normalized == "890405":
                    identificaciones_con_890405.add(codigo_normalized)

    problemas_ide_contrato: list[dict[str, Any]] = []

    # ----- Loop principal: validar IDE Contrato por fila
    for row in range(2, data_sheet.max_row + 1):
        tipo_factura = data_sheet.cell(row=row, column=tipo_factura_idx + 1).value
        tipo_factura_str = str(tipo_factura).strip() if tipo_factura else ""

        # Solo procesar si Tipo Factura = "Urgencias"
        if tipo_factura_str != "Urgencias":
            continue

        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str:
            continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_excluir = str(codigo).strip() if codigo else ""

        if not codigo_excluir:
            continue

        codigo_entidad_val = data_sheet.cell(row=row, column=codigo_entidad_cobrar_idx + 1).value
        codigo_entidad_str = str(codigo_entidad_val).strip() if codigo_entidad_val else ""

        if not codigo_entidad_str:
            continue

        ide_contrato_val = data_sheet.cell(row=row, column=ide_contrato_idx + 1).value
        ide_contrato_str = str(ide_contrato_val).strip() if ide_contrato_val else ""

        proc_str = ""
        if proc_idx is not None:
            proc_val = data_sheet.cell(row=row, column=proc_idx + 1).value
            proc_str = str(proc_val).strip() if proc_val else ""

        ident_str = ""
        if ident_idx is not None:
            ident_val = data_sheet.cell(row=row, column=ident_idx + 1).value
            ident_str = str(ident_val).strip() if ident_val else ""

        # ----- Aplicar reglas SIMPLES (exact match)
        for rule in IDE_SIMPLE_RULES:
            if codigo_excluir == rule["codigo"] and codigo_entidad_str == rule["entidad"]:
                if ide_contrato_str != rule["expected"]:
                    problemas_ide_contrato.append({
                        "factura": factura_str,
                        "procedimiento": proc_str,
                        "codigo": codigo_excluir,
                        "entidad": codigo_entidad_str,
                        "ide_contrato_actual": ide_contrato_str,
                        "ide_contrato_deberia": rule["expected"],
                    })
                    logger.debug(
                        "Fila %s: Entidad=%s, Código=%s, IDE incorrecto (Actual: '%s', Esperado: %s)",
                        row, codigo_entidad_str, codigo_excluir,
                        ide_contrato_str, rule["expected"],
                    )

        # ----- Aplicar reglas CONDICIONALES por inserción 861801
        for rule in IDE_INSERTION_RULES:
            if codigo_excluir == rule["codigo"] and codigo_entidad_str == rule["entidad"]:
                tiene_insercion = ident_str in identificaciones_con_insercion
                ide_esperado = rule["expected_with"] if tiene_insercion else rule["expected_without"]
                if ide_contrato_str != ide_esperado:
                    problemas_ide_contrato.append({
                        "factura": factura_str,
                        "procedimiento": proc_str,
                        "codigo": codigo_excluir,
                        "entidad": codigo_entidad_str,
                        "ide_contrato_actual": ide_contrato_str,
                        "ide_contrato_deberia": ide_esperado,
                        "tiene_insercion": tiene_insercion,
                    })
                    logger.debug(
                        "Fila %s: Entidad=%s, Código=%s, IDE incorrecto (Actual: '%s', Esperado: '%s', Inserción: %s)",
                        row, codigo_entidad_str, codigo_excluir,
                        ide_contrato_str, ide_esperado, tiene_insercion,
                    )

        # ----- Aplicar reglas ESSC62 (condicional por 890405)
        for rule in IDE_ESSC62_890405_RULES:
            if codigo_excluir == rule["codigo"] and codigo_entidad_str == rule["entidad"]:
                tiene_890405_flag = ident_str in identificaciones_con_890405
                ide_esperado = rule["expected_with"] if tiene_890405_flag else rule["expected_without"]
                if ide_contrato_str != ide_esperado:
                    problemas_ide_contrato.append({
                        "factura": factura_str,
                        "procedimiento": proc_str,
                        "codigo": codigo_excluir,
                        "entidad": codigo_entidad_str,
                        "ide_contrato_actual": ide_contrato_str,
                        "ide_contrato_deberia": ide_esperado,
                        "tiene_890405": tiene_890405_flag,
                    })
                    logger.debug(
                        "Fila %s: Entidad=%s, Código=%s, IDE incorrecto (Actual: '%s', Esperado: '%s', Tiene 890405: %s)",
                        row, codigo_entidad_str, codigo_excluir,
                        ide_contrato_str, ide_esperado, tiene_890405_flag,
                    )

        # ----- Aplicar reglas MÚLTIPLES (cualquier IDE del set)
        for rule in IDE_MULTIPLE_RULES:
            if codigo_excluir == rule["codigo"] and codigo_entidad_str == rule["entidad"]:
                if ide_contrato_str not in rule["expected_set"]:
                    problemas_ide_contrato.append({
                        "factura": factura_str,
                        "procedimiento": proc_str,
                        "codigo": codigo_excluir,
                        "entidad": codigo_entidad_str,
                        "ide_contrato_actual": ide_contrato_str,
                        "ide_contrato_deberia": f"uno de: {sorted(rule['expected_set'])}",
                        "nota": rule["note"],
                    })
                    logger.debug(
                        "Fila %s: Entidad=%s, Código=%s, IDE incorrecto (Actual: '%s', Esperado uno de: %s)",
                        row, codigo_entidad_str, codigo_excluir,
                        ide_contrato_str, sorted(rule["expected_set"]),
                    )

        # ----- Regla 29: Entidad -> IDE Contrato (mapeo directo, sin importar código)
        if codigo_entidad_str in URGENCIA_ENTIDAD_CONTRATO:
            # Excluir casos especiales que se manejan en reglas anteriores
            es_excluido = False
            if codigo_entidad_str == "86000" and codigo_excluir == "861801":
                es_excluido = True  # Se maneja en simple rule
            elif codigo_entidad_str == "86000" and codigo_excluir == "890405":
                es_excluido = True  # Se maneja en insertion rule
            elif codigo_entidad_str == "RES004" and codigo_excluir == "861801":
                es_excluido = True  # Se maneja en simple rule
            elif codigo_entidad_str == "RES004" and codigo_excluir == "890405":
                es_excluido = True  # Se maneja en insertion rule
            elif codigo_entidad_str in URGENCIA_ENTIDAD_MULTIPLE_CONTRATO:
                es_excluido = True  # Se maneja en multiple rule

            if not es_excluido:
                ide_contrato_requerido = URGENCIA_ENTIDAD_CONTRATO[codigo_entidad_str]
                if ide_contrato_str != ide_contrato_requerido:
                    problemas_ide_contrato.append({
                        "factura": factura_str,
                        "procedimiento": proc_str,
                        "codigo": codigo_excluir,
                        "entidad": codigo_entidad_str,
                        "ide_contrato_actual": ide_contrato_str,
                        "ide_contrato_deberia": ide_contrato_requerido,
                        "nota": "Regla Entidad->Contrato",
                    })
                    logger.debug(
                        "Fila %s: Entidad=%s, IDE incorrecto (Actual: '%s', Esperado: %s)",
                        row, codigo_entidad_str, ide_contrato_str, ide_contrato_requerido,
                    )

        # ----- Regla 30: Entidad con múltiples contratos válidos
        if codigo_entidad_str in URGENCIA_ENTIDAD_MULTIPLE_CONTRATO:
            contratos_validos = URGENCIA_ENTIDAD_MULTIPLE_CONTRATO[codigo_entidad_str]
            if ide_contrato_str not in contratos_validos:
                problemas_ide_contrato.append({
                    "factura": factura_str,
                    "procedimiento": proc_str,
                    "codigo": codigo_excluir,
                    "entidad": codigo_entidad_str,
                    "ide_contrato_actual": ide_contrato_str,
                    "ide_contrato_deberia": f"uno de: {sorted(contratos_validos)}",
                    "nota": "Entidad con múltiples contratos válidos",
                })
                logger.debug(
                    "Fila %s: Entidad=%s, IDE incorrecto (Actual: '%s', Esperado uno de: %s)",
                    row, codigo_entidad_str, ide_contrato_str, sorted(contratos_validos),
                )

    if problemas_ide_contrato:
        logger.info("IDE Contrato Urgencias - Problemas encontrados: %d", len(problemas_ide_contrato))

    return problemas_ide_contrato
