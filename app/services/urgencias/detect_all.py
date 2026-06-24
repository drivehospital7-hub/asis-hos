"""Orquestador de detección de problemas para Urgencias.

Agrupa detectores transversales + específicos de Urgencias.
Usa el builder compartido de normalized_rows.
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import AREA_URGENCIAS
from app.constants.base import is_rule_engine_enabled
from app.services.transversales import (
    detect_decimales,
    detect_tipo_documento_edad,
    detect_tipo_identificacion_entidad,
    detect_tipo_usuario,
    normalize_invoice,
)
from app.services.urgencias.centro_costo_urgencias import (
    detect_centro_costo_urgencias,
)
from app.services.urgencias.ide_contrato_urgencias import (
    detect_ide_contrato_urgencias,
)
from app.services.urgencias.cups_equivalentes import detect_cups_equivalentes
from app.services.urgencias.sala_observacion import detect_sala_observacion
from app.services.urgencias.cantidades_urgencias import (
    detect_cantidades_urgencias,
)
from app.services.urgencias.cantidades_soat_urgencias import (
    detect_cantidades_soat_urgencias,
)
from app.services.urgencias.mal_capitado import detect_mal_capitado
from app.services.urgencias.ide_contrato_reverse import detect_ide_contrato_reverse_urgencias
from app.services.urgencias.profesionales_urgencias import detect_profesionales_urgencias
from app.services.transversales.detect_copago_entidad import (
    detect_copago_entidad_urgencias,
)
from app.services.transversales.procedimiento_contratado import detect_cups_sin_contrato
from app.services.urgencias.revision_cantidad import detect_revision_cantidad_urgencias
from app.services.urgencias.revision_entidad_86 import detect_revision_entidad_86_urgencias
from app.services.urgencias.duplicados_farmacia import detect_duplicados_farmacia
from app.services.normalized_rows import build_normalized_rows

logger = logging.getLogger(__name__)


def detect_all_problems_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> tuple[dict[str, Any], dict[str, str]]:
    """Detecta TODOS los problemas en facturas de urgencias.

    Incluye detectores transversales y específicos de urgencias.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        (resultado_dict, responsables_map)
    """
    # 1. Centro Costo + IDE Contrato + CUPS equivalentes
    problemas_centros = detect_centro_costo_urgencias(data_sheet, indices)
    problemas_ide_contrato = detect_ide_contrato_urgencias(data_sheet, indices)

    problemas_cups_equivalentes: list[dict[str, str]] = []
    problemas_cups_equivalentes.extend(detect_cups_equivalentes(data_sheet, indices))
    problemas_cups_equivalentes.extend(detect_sala_observacion(data_sheet, indices))

    # 3. Detectores transversales
    decimales = detect_decimales(data_sheet, indices)
    tipo_identificacion_edad = detect_tipo_documento_edad(data_sheet, indices)
    tipo_identificacion_entidad = detect_tipo_identificacion_entidad(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            r1 = RuleBasedDetector("tipo_id_requiere_entidad_86000", session).detect(data_sheet, indices)
            r2 = RuleBasedDetector("entidad_86000_requiere_as_ms", session).detect(data_sheet, indices)
            tipo_identificacion_entidad = r1 + r2
            session.commit()
        finally:
            session.close()
    tipo_usuario = detect_tipo_usuario(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            tipo_usuario = RuleBasedDetector("tipo_usuario_valido", session).detect(data_sheet, indices)
            session.commit()
        finally:
            session.close()

    # 5. Detectores específicos de urgencias
    profesionales = detect_profesionales_urgencias(data_sheet, indices)
    logger.info(
        "detect_all_problems_urgencias - Profesionales encontrados: %d",
        len(profesionales),
    )

    mal_capitado = detect_mal_capitado(data_sheet, indices)
    logger.info(
        "detect_all_problems_urgencias - MAL CAPITADO encontrados: %d",
        len(mal_capitado),
    )

    cantidades_urgencias = detect_cantidades_urgencias(data_sheet, indices)
    logger.info(
        "detect_all_problems_urgencias - Cantidades Urgencias encontradas: %d",
        len(cantidades_urgencias),
    )

    cantidades_soat_urgencias = detect_cantidades_soat_urgencias(data_sheet, indices)
    logger.info(
        "detect_all_problems_urgencias - Cantidades SOAT Urgencias encontradas: %d",
        len(cantidades_soat_urgencias),
    )

    ide_contrato_reverse = detect_ide_contrato_reverse_urgencias(
        data_sheet, indices
    )
    logger.info(
        "detect_all_problems_urgencias - IDE Contrato REVERSE encontrados: %d",
        len(ide_contrato_reverse),
    )

    revision_entidad_86 = detect_revision_entidad_86_urgencias(
        data_sheet, indices
    )
    logger.info(
        "detect_all_problems_urgencias - Revision Entidad 86 encontradas: %d",
        len(revision_entidad_86),
    )

    revision_cantidad = detect_revision_cantidad_urgencias(data_sheet, indices)
    logger.info(
        "detect_all_problems_urgencias - Revision Cantidad encontradas: %d",
        len(revision_cantidad),
    )

    copago_entidad = detect_copago_entidad_urgencias(data_sheet, indices)
    if is_rule_engine_enabled():
        from app.services.engine.rule_based_detector import RuleBasedDetector
        from app.database import get_session
        session = get_session()
        try:
            copago_entidad = RuleBasedDetector("copago_entidad_valido", session).detect(data_sheet, indices)
            session.commit()
        finally:
            session.close()
    logger.info(
        "detect_all_problems_urgencias - Copago vs Entidad encontrados: %d",
        len(copago_entidad),
    )

    duplicados_farmacia = detect_duplicados_farmacia(data_sheet, indices)
    logger.info(
        "detect_all_problems_urgencias - Duplicados Farmacia encontrados: %d",
        len(duplicados_farmacia),
    )

    # 6a. Cups Sin Contrato
    cups_sin_contrato = detect_cups_sin_contrato(data_sheet, indices)
    logger.info(
        "detect_all_problems_urgencias - Cups Sin Contrato encontrados: %d",
        len(cups_sin_contrato),
    )

    # 6. Filtrar centros de costo por prioridad
    errores_por_factura_codigo: dict[tuple[str, str], list[tuple[dict, int]]] = {}
    for item in problemas_centros:
        key = (item.get("factura", ""), item.get("codigo", ""))
        prioridad = item.get("prioridad", 1)
        if key not in errores_por_factura_codigo:
            errores_por_factura_codigo[key] = []
        errores_por_factura_codigo[key].append((item, prioridad))

    problemas_centros_filtrados = []
    for key, items in errores_por_factura_codigo.items():
        prioridades = [p for _, p in items]
        if 1 in prioridades:
            for item, p in items:
                if p == 1:
                    problemas_centros_filtrados.append(item)
        else:
            for item, _ in items:
                problemas_centros_filtrados.append(item)

    logger.info(
        "FILTRO centros_de_costos: %d -> %d",
        len(problemas_centros),
        len(problemas_centros_filtrados),
    )

    # 7. Build responsable_cierra mapping
    responsable_cierra: dict[str, str] = {}
    responsable_cierra_idx = indices.get("responsable_cierra")
    num_fact_idx = indices.get("numero_factura")
    if responsable_cierra_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            raw = data_sheet.cell(row=row, column=responsable_cierra_idx + 1).value
            resp = str(raw).strip() if raw else ""
            if resp and factura not in responsable_cierra:
                responsable_cierra[factura] = resp

    # 8. Build fecha_cierre_vacia mapping
    fecha_cierre_vacia: dict[str, bool] = {}
    fecha_cierre_idx = indices.get("fecha_cierre")
    if fecha_cierre_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            fecha_cierre_val = data_sheet.cell(
                row=row, column=fecha_cierre_idx + 1
            ).value
            if not fecha_cierre_val or str(fecha_cierre_val).strip() == "":
                fecha_cierre_vacia[factura] = True
            elif factura not in fecha_cierre_vacia:
                fecha_cierre_vacia[factura] = False

    # 9. Build fec_factura_map
    fec_factura_map: dict[str, str] = {}
    fec_factura_idx = indices.get("fec_factura")
    if fec_factura_idx is not None and num_fact_idx is not None:
        for row in range(2, data_sheet.max_row + 1):
            numero = data_sheet.cell(row=row, column=num_fact_idx + 1).value
            factura = normalize_invoice(numero)
            if not factura:
                continue
            raw = data_sheet.cell(row=row, column=fec_factura_idx + 1).value
            val = str(raw).strip() if raw else ""
            if val and factura not in fec_factura_map:
                fec_factura_map[factura] = val

    # 10. Build normalized rows (shared builder)
    error_groups = {
        "Centros de Costo": problemas_centros_filtrados,
        "IDE Contrato": problemas_ide_contrato,
        "Cups Equivalentes": problemas_cups_equivalentes,
        "MAL CAPITADO": mal_capitado,
        "Cantidades": cantidades_urgencias,
        "Cantidades SOAT": cantidades_soat_urgencias,
        "Decimales": decimales,
        "Tipo Identificación / Edad": tipo_identificacion_edad,
        "Profesionales": profesionales,
        "Código Entidad vs Afiliación": tipo_identificacion_entidad,
        "Tipo Usuario": tipo_usuario,
        "⚠️ Revisión Necesaria": revision_entidad_86 + revision_cantidad,
        "Copago vs Entidad": copago_entidad,
        "Duplicados Farmacia": duplicados_farmacia,
        "Cups Sin Contrato": cups_sin_contrato,
    }
    normalized_rows = build_normalized_rows(
        error_groups=error_groups,
        responsables_map=responsable_cierra,
        fec_factura_map=fec_factura_map,
        fecha_cierre_vacia_map=fecha_cierre_vacia,
    )

    # 11. Build resultado dict
    resultado: dict[str, Any] = {
        "area": AREA_URGENCIAS,
        "problemas": {
            "normalizados": normalized_rows,
            "centros_de_costos": [
                {
                    "tipo_factura": item.get("tipo_factura") or "-",
                    "factura": item["factura"],
                    "codigo": item.get("codigo", ""),
                    "procedimiento": item.get("procedimiento", ""),
                    "centro_actual": item["centro_actual"],
                    "centro_deberia": item["centro_deberia"],
                    "prioridad": item.get("prioridad", 1),
                }
                for item in problemas_centros_filtrados
            ],
            "ide_contrato": [
                {
                    "factura": item["factura"],
                    "ide_contrato_actual": item["ide_contrato_actual"],
                    "ide_contrato_deberia": item["ide_contrato_deberia"],
                    "procedimiento": item.get("procedimiento", ""),
                    "codigo": item.get("codigo", ""),
                    "entidad": item.get("entidad", ""),
                    "nota": item.get("nota", ""),
                }
                for item in problemas_ide_contrato
            ],
            "cups_equivalentes": [
                {
                    "factura": item["factura"],
                    "codigo": item["codigo"],
                    "codigo_equiv": item["codigo_equiv"],
                    "accion": item["accion"],
                }
                for item in problemas_cups_equivalentes
            ],
            # reglas transversales
            "decimales": decimales,
            "tipo_identificacion_edad": tipo_identificacion_edad,
            "tipo_identificacion_entidad": tipo_identificacion_entidad,
            "codigo_entidad_vs_afiliacion": [],
            "tipo_usuario": tipo_usuario,
            # reglas urgencias
            "profesionales": profesionales,
            "mal_capitado": mal_capitado,
            "cantidades_urgencias": cantidades_urgencias,
            "cantidades_soat_urgencias": cantidades_soat_urgencias,
            "revision_entidad_86": revision_entidad_86,
            "revision_cantidad": revision_cantidad,
            "copago_entidad": copago_entidad,
            "duplicados_farmacia": duplicados_farmacia,
            "cups_sin_contrato": cups_sin_contrato,
        },
        "totales": {
            "centros_de_costos": len(problemas_centros),
            "ide_contrato": len(problemas_ide_contrato),
            "cups_equivalentes": len(problemas_cups_equivalentes),
            "decimales": len(decimales),
            "tipo_identificacion_edad": len(tipo_identificacion_edad),
            "tipo_identificacion_entidad": len(tipo_identificacion_entidad),
            "codigo_entidad_vs_afiliacion": 0,
            "tipo_usuario": len(tipo_usuario),
            "profesionales": len(profesionales),
            "mal_capitado": len(mal_capitado),
            "cantidades_urgencias": len(cantidades_urgencias),
            "cantidades_soat_urgencias": len(cantidades_soat_urgencias),
            "revision_entidad_86": len(revision_entidad_86),
            "revision_cantidad": len(revision_cantidad),
            "copago_entidad": len(copago_entidad),
            "duplicados_farmacia": len(duplicados_farmacia),
            "cups_sin_contrato": len(cups_sin_contrato),
        },
        "missing_columns": [],
    }

    # 12. Enrich errors with responsable from mapping
    if responsable_cierra:
        for problem_type, problems in resultado["problemas"].items():
            for p in problems:
                if not isinstance(p, dict):
                    continue
                factura = p.get("factura")
                if factura and factura in responsable_cierra:
                    p["responsable"] = responsable_cierra[factura]
                elif "responsable" not in p:
                    p["responsable"] = ""
    else:
        for problem_type, problems in resultado["problemas"].items():
            for p in problems:
                if not isinstance(p, dict):
                    continue
                if "responsable" not in p:
                    p["responsable"] = ""

    return resultado, responsable_cierra
