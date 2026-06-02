"""Detección de profesionales no válidos en Urgencias.

Extraído de app/services/revision_sheet.py._detect_profesionales_urgencias
como parte de la Fase 7 (cleanup).
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    CODIGOS_EXCLUIDOS_MEDICO,
    CODIGOS_FISIOTERAPEUTA,
    CODIGOS_JEFE_ENFERMERIA,
    CODIGOS_NUTRICIONISTA,
    CODIGOS_ODONTOLOGO,
    CODIGOS_PSICOLOGA,
    CODIGOS_TRABAJADORA_SOCIAL,
    CODIGO_TIPO_PROCEDIMIENTO_DIAGNOSTICO,
    EXCEPCIONES_BACTERIOLOGA,
    PROFESIONALES_URGENCIAS,
)
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_profesionales_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, str]]:
    """
    Detecta facturas con profesionales no válidos en Urgencias.

    Reglas (Urgencias):
    - "Código Profesional" DEBE estar en PROFESIONALES_URGENCIAS
    - TRABAJADORA SOCIAL: solo puede usar código 890409

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo_profesional", "nombre", "tipo",
        "profesional_area", "procedimiento", "regla", "problema"
    """
    logger.warning("=== detect_profesionales_urgencias ===")
    logger.warning("Indices encontrados: %s", indices)

    tipo_factura_idx = indices.get("tipo_factura_descripcion")
    num_fact_idx = indices.get("numero_factura")
    cod_prof_idx = indices.get("codigo_profesional")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")

    logger.warning(
        "numero_factura idx: %s, codigo_profesional idx: %s, codigo idx: %s, procedimiento idx: %s",
        num_fact_idx, cod_prof_idx, codigo_idx, procedimiento_idx,
    )

    if None in (tipo_factura_idx, num_fact_idx, cod_prof_idx):
        logger.warning("NO se encontró numero_factura o codigo_profesional en los índices")
        return []

    problemas = []
    facturas_procesadas: set[str] = set()

    # Log de las primeras 5 filas para debug
    logger.warning("=== MUESTREO 5 PRIMERAS FILAS PROFESIONALES ===")
    for row in range(2, min(7, data_sheet.max_row + 1)):
        num_fact = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        cod_prof = data_sheet.cell(row=row, column=cod_prof_idx + 1).value
        codigo_val = ""
        proc_val = ""
        if codigo_idx is not None:
            codigo_val = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_val = str(codigo_val).strip() if codigo_val else ""
        if procedimiento_idx is not None:
            proc_val = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
            proc_val = str(proc_val).strip()[:30] if proc_val else ""

        logger.warning(
            "Fila %d: factura=%s, cod_prof=%s, codigo=%s, proc=%s",
            row, num_fact, cod_prof, codigo_val, proc_val,
        )

    for row in range(2, data_sheet.max_row + 1):
        tipo_factura = data_sheet.cell(row=row, column=tipo_factura_idx + 1).value
        tipo_factura_str = str(tipo_factura).strip() if tipo_factura else ""

        # Solo procesar si Tipo Factura = "Urgencias"
        if tipo_factura_str != "Urgencias":
            continue

        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_procesadas:
            continue

        cod_profesional = data_sheet.cell(row=row, column=cod_prof_idx + 1).value
        cod_profesional_str = str(cod_profesional).strip() if cod_profesional else ""

        if not cod_profesional_str:
            continue

        # Buscar profesional en el diccionario de Urgencias
        profesional_info = PROFESIONALES_URGENCIAS.get(cod_profesional_str)

        if profesional_info is None:
            logger.warning("Profesional no encontrado en lista: %s", cod_profesional_str)
            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""

            problemas.append({
                "factura": factura_str,
                "codigo_profesional": cod_profesional_str,
                "nombre": "",
                "tipo": "",
                "profesional_area": "",
                "procedimiento": procedimiento,
                "regla": "Profesional debe estar en listado",
                "problema": "Profesional no existe en el listado de Urgencias",
            })
            facturas_procesadas.add(factura_str)
            continue

        # Validación por tipo de profesional
        tipo_profesional = profesional_info.get("tipo", "")

        # Si es TRABAJADORA SOCIAL, validar código 890409
        if tipo_profesional == "TRABAJADORA SOCIAL" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""

            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""

            if codigo_str and codigo_str not in CODIGOS_TRABAJADORA_SOCIAL:
                codigos_validos = ", ".join(sorted(CODIGOS_TRABAJADORA_SOCIAL))
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "TRABAJADORA SOCIAL",
                    "profesional_area": "TRABAJADORA SOCIAL",
                    "procedimiento": procedimiento,
                    "regla": f"Código debe ser uno de: {codigos_validos}",
                    "problema": f"TRABAJADORA SOCIAL con código no permitido ({codigo_str}). Debería usar {codigos_validos}",
                })
                facturas_procesadas.add(factura_str)

        # Si es PSICOLOGA, validar códigos 890408 o 35102
        if tipo_profesional == "PSICOLOGA" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""

            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""

            if codigo_str and codigo_str not in CODIGOS_PSICOLOGA:
                codigos_validos = ", ".join(sorted(CODIGOS_PSICOLOGA))
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "PSICOLOGA",
                    "profesional_area": "PSICOLOGA",
                    "procedimiento": procedimiento,
                    "regla": f"Código debe ser uno de: {codigos_validos}",
                    "problema": f"PSICOLOGA con código no permitido ({codigo_str}). Debería usar {codigos_validos}",
                })
                facturas_procesadas.add(factura_str)

        # Si es NUTRICIONISTA, validar códigos 890406 o 37602
        if tipo_profesional == "NUTRICIONISTA" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""

            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""

            if codigo_str and codigo_str not in CODIGOS_NUTRICIONISTA:
                codigos_validos = ", ".join(sorted(CODIGOS_NUTRICIONISTA))
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "NUTRICIONISTA",
                    "profesional_area": "NUTRICIONISTA",
                    "procedimiento": procedimiento,
                    "regla": f"Código debe ser uno de: {codigos_validos}",
                    "problema": f"NUTRICIONISTA con código no permitido ({codigo_str}). Debería usar {codigos_validos}",
                })
                facturas_procesadas.add(factura_str)

        # Si es FISIOTERAPEUTA, validar código 890412, 890411 o 29117
        if tipo_profesional == "FISIOTERAPEUTA" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""

            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""

            codigos_validos = ", ".join(sorted(CODIGOS_FISIOTERAPEUTA))
            if codigo_str and codigo_str not in CODIGOS_FISIOTERAPEUTA:
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "FISIOTERAPEUTA",
                    "profesional_area": "FISIOTERAPEUTA",
                    "procedimiento": procedimiento,
                    "regla": f"Código debe ser {codigos_validos}",
                    "problema": f"FISIOTERAPEUTA con código no permitido ({codigo_str}). Debería usar {codigos_validos}",
                })
                facturas_procesadas.add(factura_str)

        # Si es JEFE ENFERMERIA, validar códigos 861801, 890205, 890405, 990211, 29116, 39360
        if tipo_profesional == "JEFE ENFERMERIA" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""

            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""

            codigos_validos = ", ".join(sorted(CODIGOS_JEFE_ENFERMERIA))
            if codigo_str and codigo_str not in CODIGOS_JEFE_ENFERMERIA:
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "JEFE ENFERMERIA",
                    "profesional_area": "JEFE ENFERMERIA",
                    "procedimiento": procedimiento,
                    "regla": f"Código debe ser {codigos_validos}",
                    "problema": f"JEFE ENFERMERIA con código no permitido ({codigo_str}). Debería usar {codigos_validos}",
                })
                facturas_procesadas.add(factura_str)

        # Si es BACTERIOLOGA, validar Código Tipo Procedimiento = 02 o 05 y Laboratorio = "Si"
        if tipo_profesional == "BACTERIOLOGA":
            codigo_proc = ""
            if codigo_idx is not None:
                codigo_proc = data_sheet.cell(row=row, column=codigo_idx + 1).value
                codigo_proc = str(codigo_proc).strip() if codigo_proc else ""

            # Si es excepción -> skip validación, no dar error
            if codigo_proc in EXCEPCIONES_BACTERIOLOGA:
                facturas_procesadas.add(factura_str)
                continue

            codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")
            laboratorio_idx = indices.get("laboratorio")

            codigo_tipo = ""
            laboratorio = ""

            if codigo_tipo_proc_idx is not None:
                codigo_tipo = data_sheet.cell(row=row, column=codigo_tipo_proc_idx + 1).value
                codigo_tipo = str(codigo_tipo).strip() if codigo_tipo else ""

            if laboratorio_idx is not None:
                laboratorio = data_sheet.cell(row=row, column=laboratorio_idx + 1).value
                laboratorio = str(laboratorio).strip().upper() if laboratorio else ""

            # Validar: Código Tipo Procedimiento debe ser 02 o 05 Y Laboratorio debe ser "Si"
            es_tipo_valido = codigo_tipo in ("02", "05", CODIGO_TIPO_PROCEDIMIENTO_DIAGNOSTICO)
            es_laboratorio_si = laboratorio == "SI"

            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""

            if not (es_tipo_valido and es_laboratorio_si):
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "BACTERIOLOGA",
                    "profesional_area": "BACTERIOLOGA",
                    "procedimiento": procedimiento,
                    "regla": "Código Tipo=02/05 + Laboratorio=Si",
                    "problema": "LABORATORIO NO IDENTIFICADO: BACTERIOLOGA requiere Código Tipo Procedimiento=02/05 y Laboratorio=Si",
                })
                facturas_procesadas.add(factura_str)

        # Si es ODONTOLOGO, validar código 890403
        if tipo_profesional == "ODONTOLOGO" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""

            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""

            if codigo_str and codigo_str not in CODIGOS_ODONTOLOGO:
                codigos_validos = ", ".join(sorted(CODIGOS_ODONTOLOGO))
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "ODONTOLOGO",
                    "profesional_area": "ODONTOLOGO",
                    "procedimiento": procedimiento,
                    "regla": f"Código debe ser {codigos_validos}",
                    "problema": f"ODONTOLOGO con código no permitido ({codigo_str}). Debería usar {codigos_validos}",
                })
                facturas_procesadas.add(factura_str)

        # Si es MEDICO, NO puede usar códigos de otros profesionales ni regla de laboratorio
        if tipo_profesional == "MEDICO" and codigo_idx is not None:
            codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
            codigo_str = str(codigo).strip() if codigo else ""

            # Verificar si usa código excluido
            if codigo_str and codigo_str in CODIGOS_EXCLUIDOS_MEDICO:
                procedimiento = ""
                if procedimiento_idx is not None:
                    proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                    procedimiento = str(proc).strip() if proc else ""

                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "MEDICO",
                    "profesional_area": "MEDICO",
                    "procedimiento": procedimiento,
                    "regla": f"No usar: {', '.join(sorted(CODIGOS_EXCLUIDOS_MEDICO))}",
                    "problema": f"MEDICO con código no permitido ({codigo_str}). Código reservado para otro tipo de profesional",
                })
                facturas_procesadas.add(factura_str)
                continue

            # Verificar si cumple regla de laboratorio (que es de BACTERIOLOGA)
            codigo_tipo_proc_idx = indices.get("codigo_tipo_procedimiento")
            laboratorio_idx = indices.get("laboratorio")

            codigo_tipo = ""
            laboratorio = ""

            if codigo_tipo_proc_idx is not None:
                codigo_tipo = data_sheet.cell(row=row, column=codigo_tipo_proc_idx + 1).value
                codigo_tipo = str(codigo_tipo).strip() if codigo_tipo else ""

            if laboratorio_idx is not None:
                laboratorio = data_sheet.cell(row=row, column=laboratorio_idx + 1).value
                laboratorio = str(laboratorio).strip().upper() if laboratorio else ""

            procedimiento = ""
            if procedimiento_idx is not None:
                proc = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc).strip() if proc else ""

            # Si tiene código normal pero cumple regla de laboratorio = error
            es_tipo_lab = codigo_tipo in ("02", "05", CODIGO_TIPO_PROCEDIMIENTO_DIAGNOSTICO)
            es_lab_si = laboratorio == "SI"

            if codigo_str and es_tipo_lab and es_lab_si:
                problemas.append({
                    "factura": factura_str,
                    "codigo_profesional": cod_profesional_str,
                    "nombre": profesional_info.get("nombre", ""),
                    "tipo": "MEDICO",
                    "profesional_area": "MEDICO",
                    "procedimiento": procedimiento,
                    "regla": "No usar Tipo=02/05 + Lab=Si (reservado BACTERIOLOGA)",
                    "problema": "MEDICO no puede usar código de Laboratorio (Tipo 02/05 + Lab=Si). Reserved for BACTERIOLOGA",
                })
                facturas_procesadas.add(factura_str)

    if problemas:
        logger.warning("=== ERRORES PROFESIONALES URGENCIAS: %d ===", len(problemas))
        for p in problemas:
            logger.warning(
                "- Factura: %s, Profesional: %s (%s), Área: %s, Código: %s, Problema: %s",
                p.get("factura"), p.get("codigo_profesional"), p.get("nombre"),
                p.get("profesional_area"), p.get("procedimiento"), p.get("problema"),
            )

    return problemas
