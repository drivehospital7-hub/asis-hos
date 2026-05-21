"""Detector de cantidades anómalas en facturas de Urgencias.

Regla: Cuando Tipo Factura Descripción = "Urgencias", los siguientes códigos
CUPS deben tener cantidad <= 1:
- 05DSB01, 5DSB01, 890601, 890701, 129B02, 12333
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.constants import URGENCIAS_CODIGOS_CANTIDAD_MAX_1
from app.services.transversales.normalize import normalize_invoice

logger = logging.getLogger(__name__)


def detect_cantidades_urgencias(
    data_sheet: Worksheet,
    indices: dict[str, int | None],
) -> list[dict[str, Any]]:
    """
    Detecta facturas con cantidades anómalas en Urgencias.

    Args:
        data_sheet: Hoja de Excel con los datos
        indices: Índices de columnas

    Returns:
        Lista de dicts con keys: "factura", "codigo", "procedimiento",
        "cantidad", "tipo_factura"
    """
    tipo_factura_idx = indices.get("tipo_factura_descripcion")
    num_fact_idx = indices.get("numero_factura")
    codigo_idx = indices.get("codigo")
    procedimiento_idx = indices.get("procedimiento")
    cantidad_idx = indices.get("cantidad")

    if None in (tipo_factura_idx, num_fact_idx, codigo_idx, cantidad_idx):
        logger.warning("Cantidades Urgencias - Columnas necesarias no encontradas")
        return []

    problemas = []
    facturas_procesadas: set[str] = set()

    for row in range(2, data_sheet.max_row + 1):
        tipo_factura = data_sheet.cell(row=row, column=tipo_factura_idx + 1).value
        tipo_factura_str = str(tipo_factura).strip() if tipo_factura else ""

        # Solo procesar si Tipo Factura = "Urgencias"
        if tipo_factura_str != "Urgencias":
            continue

        numero_factura = data_sheet.cell(row=row, column=num_fact_idx + 1).value
        factura_str = normalize_invoice(numero_factura)
        if not factura_str or factura_str in facturas_procesadas:
            continue

        codigo = data_sheet.cell(row=row, column=codigo_idx + 1).value
        codigo_str = str(codigo).strip().upper() if codigo else ""

        # Verificar si el código está en la lista restringida
        if codigo_str not in URGENCIAS_CODIGOS_CANTIDAD_MAX_1:
            continue

        cantidad = data_sheet.cell(row=row, column=cantidad_idx + 1).value
        if not isinstance(cantidad, (int, float)):
            continue

        # Validar: cantidad debe ser <= 1
        if cantidad > 1:
            procedimiento = ""
            if procedimiento_idx is not None:
                proc_value = data_sheet.cell(row=row, column=procedimiento_idx + 1).value
                procedimiento = str(proc_value).strip() if proc_value else ""

            problemas.append({
                "factura": factura_str,
                "codigo": codigo_str,
                "procedimiento": procedimiento,
                "cantidad": cantidad,
                "tipo_factura": tipo_factura_str,
            })
            facturas_procesadas.add(factura_str)
            logger.warning(
                "CANTIDAD URGENCIAS - Factura='%s', Código='%s', Cantidad=%s (debe ser <=1)",
                factura_str, codigo_str, cantidad
            )

    if problemas:
        logger.info("Cantidades Urgencias - Problemas encontrados: %d", len(problemas))

    return problemas
