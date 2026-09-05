"""RuleEvaluationEngine — orchestrates the full rule evaluation flow.

Loads rules, evaluates condition trees against sheet rows, collects evidence,
and returns structured detection results.
"""

from __future__ import annotations

import logging
from typing import Any, TYPE_CHECKING

from app.models import Regla, Condicion, ResultadoAuditoria
from app.services.engine.context import EvaluationContext
from app.services.engine.condition_evaluator import ConditionEvaluator
from app.services.engine.evidence_collector import EvidenceCollector
from app.services.engine.exception_handler import ExceptionHandler
from app.services.engine.group_evaluator import GroupEvaluator
from app.services.engine.rule_resolver import RuleResolver

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet
    from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


class RuleEvaluationEngine:
    """Orchestrates rule loading, condition evaluation, exception handling,
    and evidence collection for a single rule against an Excel sheet.

    Usage:
        engine = RuleEvaluationEngine(session)
        results = engine.evaluate_sheet("valores_decimales", data_sheet, indices)

    The engine supports two data-paths:
    - **Worksheet path** (legacy): Provide ``data_sheet`` and ``indices``.
    - **RowStore path** (fast): Provide ``rows`` (list[dict]) plus ``indices``.

    When ``rows`` is given, all ``data_sheet.cell()`` lookups are replaced with
    O(1) dict accesses from ``rows[row - 2]``.
    """

    def __init__(self, session: "Session") -> None:
        self._session = session
        self._resolver = RuleResolver()
        self._evaluator = ConditionEvaluator()
        self._exception_handler = ExceptionHandler()
        self._evidence_collector = EvidenceCollector()

    def evaluate_sheet(
        self,
        rule_name: str,
        data_sheet: "Worksheet | None" = None,
        indices: dict[str, int | None] | None = None,
        persist: bool = True,
        rows: list[dict[str, Any]] | None = None,
        evidence_collector: "EvidenceCollector | None" = None,
    ) -> list[dict[str, Any]]:
        """Evaluate a single rule against all rows.

        Supports two data-source paths:
        - **Worksheet path** (legacy): Pass ``data_sheet`` and ``indices``.
        - **RowStore path** (fast): Pass ``rows`` (list[dict]) plus ``indices``.

        When both are provided, ``rows`` takes precedence for data access.

        Args:
            rule_name: Rule name to evaluate (e.g., 'valores_decimales').
            data_sheet: openpyxl Worksheet with invoice data.
            indices: Column name → 0-based column index mapping.
            persist: If True (default), record evidence and audit. If False,
                     skip all DB writes — only return detection results.
            rows: Optional list of dicts (RowStore) for O(1) dict access path.
            evidence_collector: Optional external EvidenceCollector. When
                provided, the engine uses it instead of its own internal
                collector and does NOT call ``flush_batch()`` — the caller
                owns the flush lifecycle.

        Returns:
            List of detection dicts with keys: factura, problema, regla, severidad,
            and optional rule-specific keys.
        """
        # Load the rule
        rule = self._load_rule_by_name(rule_name)
        if rule is None:
            logger.warning("Rule not found: %s", rule_name)
            return []

        # Load conditions and build tree
        conditions = self._load_conditions(rule.id)
        tree = self._evaluator.build_tree(conditions)
        if tree is None:
            logger.warning("No condition tree for rule: %s", rule_name)
            return []

        # Load param configs
        param_configs = rule.parametros or []
        if not param_configs:
            param_configs = [{}]  # Default: single evaluation with no overrides

        # Determine which collector to use
        collector = evidence_collector if evidence_collector is not None else self._evidence_collector

        # ── Group-by routing ──────────────────────────────────────────────
        first_param = param_configs[0] if param_configs else {}
        if isinstance(first_param, dict) and first_param.get("group_by"):
            return self._evaluate_sheet_group_by(
                rule, tree, data_sheet, indices, first_param,
                persist=persist, rows=rows, evidence_collector=evidence_collector,
            )

        results: list[dict[str, Any]] = []

        # ── Optimization: cache exceptions once per rule (not per row) ─────
        exception_result = self._exception_handler.query_exceptions(rule, self._session)

        # ── Optimization: lazy date.edad — only resolve when tree references it
        needs_edad = self._evaluator.tree_uses_field(tree, "date.edad")
        needs_edad_meses = self._evaluator.tree_uses_field(tree, "date.edad_meses")

        # ── Determine row iteration bounds ────────────────────────────────
        # rows[0] → Excel row 2, rows[n-1] → Excel row len(rows)+1
        if rows is not None:
            max_row = len(rows) + 1
        else:
            max_row = data_sheet.max_row  # type: ignore[union-attr]

        # Iterate over rows (row 1 = header, data starts at row 2)
        for row in range(2, max_row + 1):
            if rows is not None:
                row_data = rows[row - 2]
                factura = _extract_factura(row_data)
                if not factura:
                    continue
            else:
                row_data, factura = self._build_row_context(data_sheet, row, indices)  # type: ignore[arg-type]
                if not factura:
                    continue

            ctx = EvaluationContext(invoice_data=row_data, indices=indices, session=self._session)

            # Check exceptions (uses cached list to avoid per-row DB query)
            # Pass row_data directly — skip unnecessary EvaluationContext
            effect, overrides = self._exception_handler.apply_exceptions(
                rule, row_data, self._session, cached_exc=exception_result,
            )
            if effect == "skip":
                logger.debug("Rule %s skipped for factura %s", rule_name, factura)
                continue

            # Evaluate with each param config
            for config_idx, params in enumerate(param_configs):
                # When params is empty (default), reuse row_data directly
                # to avoid an unnecessary dict copy on every row.
                if isinstance(params, dict) and not params and not overrides:
                    merged_data = row_data
                elif overrides:
                    merged_data = {**row_data, **params, **overrides}
                else:
                    merged_data = {**row_data, **(params if isinstance(params, dict) else {})}

                eval_ctx = EvaluationContext(
                    invoice_data=merged_data,
                    indices=indices,
                    session=self._session,
                )

                # Pre-resolve common computed fields (only when tree references them)
                if needs_edad:
                    date_edad = self._resolve_computed("date.edad", eval_ctx)
                    if date_edad is not None:
                        eval_ctx.invoice_data["date.edad"] = date_edad
                if needs_edad_meses:
                    date_edad_meses = self._resolve_computed("date.edad_meses", eval_ctx)
                    if date_edad_meses is not None:
                        eval_ctx.invoice_data["date.edad_meses"] = date_edad_meses

                eval_result = self._evaluator.evaluate(tree, eval_ctx, collect_trace=False)
                outcome = eval_result.get("outcome", False)
                error_msg = eval_result.get("error")

                # Determine result status
                if error_msg:
                    final_outcome = "ERROR"
                elif outcome:
                    final_outcome = "MATCH"
                else:
                    final_outcome = "NO_MATCH"

                # ── Two-phase evaluation: fast path first, trace only when needed
                # When persist=True AND outcome is MATCH/ERROR, re-evaluate
                # with full trace for evidence collection. NO_MATCH rows skip
                # the second evaluation entirely (saves ~95% of trace dicts).
                if persist and final_outcome != "NO_MATCH":
                    eval_result = self._evaluator.evaluate(tree, eval_ctx, collect_trace=True)
                    # Re-extract outcome/error from full trace result (should be identical)
                    outcome = eval_result.get("outcome", False)
                    error_msg = eval_result.get("error")
                    # Record evidence (immutable snapshot)
                    collector.record(
                        regla_id=rule.id,
                        regla_version=rule.version,
                        dominio=rule.dominio,
                        factura=factura,
                        param_config_id=config_idx if param_configs else None,
                        outcome=final_outcome,
                        arbol_evaluado=eval_result.get("trace", {}),
                        snapshot_fila=row_data,
                        error_mensaje=error_msg,
                    )

                # If MATCH, add to detection results
                if outcome and not error_msg:
                    problem = {
                        "factura": factura,
                        "problema": rule.descripcion or rule.nombre,
                        "regla": f"#{rule.id}",
                        "severidad": rule.severidad,
                        "param_config_id": config_idx,
                    }
                    # Include relevant Excel row data for display in /procesar
                    for field in ("codigo", "codigo_equiv", "procedimiento", "tipo_identificacion",
                                  "codigo_entidad_cobrar", "tipo_procedimiento", "vlr_subsidiado",
                                  "vlr_procedimiento", "cantidad", "convenio_facturado",
                                  "centro_costo", "ide_contrato", "entidad_cobrar",
                                  "entidad_afiliacion", "tipo_usuario", "vlr_copago",
                                  "codigo_tipo_procedimiento", "laboratorio", "tarifario",
                                  "tipo_factura_descripcion", "responsable_cierra",
                                  "profesional_atiende", "identificacion",
                                  "fec_nacimiento", "fec_factura", "edad",
                                  "date.edad", "date.edad_meses", "numero_identificacion"):
                        if field in row_data:
                            problem[field] = row_data[field]
                        elif field in eval_ctx.invoice_data:
                            problem[field] = eval_ctx.invoice_data[field]
                    results.append(problem)

        if persist:
            if evidence_collector is None:
                # Only flush if WE own the collector — external collector is
                # the caller's responsibility
                evidencias = self._evidence_collector.flush_batch(self._session)

                # Create ResultadoAuditoria for each evidence record
                for ev in evidencias:
                    if ev.outcome == "MATCH":
                        resultado_str = "FAIL"
                    elif ev.outcome == "ERROR":
                        resultado_str = "ERROR"
                    else:
                        resultado_str = "PASS"

                    ra = ResultadoAuditoria(
                        evidencia_id=ev.id,
                        regla_id=ev.regla_id,
                        regla_version=ev.regla_version,
                        factura=ev.factura,
                        param_config_id=ev.param_config_id,
                        resultado=resultado_str,
                        severidad=rule.severidad,
                        mensaje=ev.error_mensaje or rule.descripcion,
                        detalles={"outcome": ev.outcome},
                    )
                    self._session.add(ra)

                self._session.flush()

        return results

    def evaluate_sheet_domain(
        self,
        domain: str,
        data_sheet: "Worksheet | None" = None,
        indices: dict[str, int | None] | None = None,
        rows: list[dict[str, Any]] | None = None,
        evidence_collector: "EvidenceCollector | None" = None,
    ) -> list[dict[str, Any]]:
        """Evaluate all active rules for a domain (including transversal) against a sheet.

        Loads rules via RuleResolver, evaluates each rule in priority order,
        and returns the combined detection results.

        Args:
            domain: Domain filter (e.g., 'odontologia', 'urgencias').
            data_sheet: openpyxl Worksheet with invoice data.
            indices: Column name → 0-based column index mapping.
            rows: Optional list of dicts (RowStore) for O(1) dict access path.
            evidence_collector: Optional external EvidenceCollector.

        Returns:
            Combined list of detection dicts from all active rules.
            Empty list if no rules found or no problems detected.
        """
        rules = self._resolver.resolve(domain, self._session)
        all_results: list[dict[str, Any]] = []
        for rule in rules:
            results = self.evaluate_sheet(
                rule.nombre, data_sheet, indices,
                rows=rows, evidence_collector=evidence_collector,
            )
            all_results.extend(results)
        return all_results

    def _evaluate_sheet_group_by(
        self,
        rule: Regla,
        tree: dict,
        data_sheet: "Worksheet | None",
        indices: dict[str, int | None] | None,
        param_config: dict[str, Any],
        persist: bool = True,
        rows: list[dict[str, Any]] | None = None,
        evidence_collector: "EvidenceCollector | None" = None,
    ) -> list[dict[str, Any]]:
        """Evaluate a group-by rule: pre-scan → partition → aggregate → evaluate.

        Args:
            rule: The loaded Regla ORM object.
            tree: Condition tree root node.
            data_sheet: openpyxl Worksheet.
            indices: Column name → index mapping.
            param_config: The parametros config dict with group_by + aggregations.
            persist: If True (default), record evidence and audit. If False,
                     skip all DB writes.
            rows: Optional list of dicts (RowStore) for O(1) dict access.
            evidence_collector: Optional external EvidenceCollector.

        Returns:
            List of detection dicts with factura, problema, regla, severidad.
        """
        group_by_field = param_config.get("group_by", "numero_factura")
        filter_field = param_config.get("filter_field")
        filter_value = param_config.get("filter_value")
        agg_configs = param_config.get("aggregations", [])

        # Determine which collector to use
        collector = evidence_collector if evidence_collector is not None else self._evidence_collector

        # 1. Pre-scan: build groups (optionally filtered)
        groups = GroupEvaluator.build_groups(
            data_sheet, indices, group_by_field, filter_field, filter_value, rows=rows,
        )
        if not groups:
            logger.debug("Group-by rule %s: no groups found", rule.nombre)
            return []

        # 2. Prepare rule info for GroupEvaluator
        rule_info = {
            "id": rule.id,
            "version": rule.version,
            "dominio": rule.dominio,
            "nombre": rule.nombre,
            "descripcion": rule.descripcion,
            "severidad": rule.severidad,
        }

        # 3. Delegate to GroupEvaluator
        results = GroupEvaluator.evaluate(
            groups=groups,
            data_sheet=data_sheet,
            indices=indices,
            agg_configs=agg_configs,
            condition_tree=tree,
            condition_evaluator=self._evaluator,
            rule_info=rule_info,
            evidence_collector=collector,
            record_evidence=persist,
            rows=rows,
        )

        if persist:
            if evidence_collector is None:
                # 4. Flush evidence and create ResultadoAuditoria (same as row-by-row path)
                evidencias = self._evidence_collector.flush_batch(self._session)
                for ev in evidencias:
                    if ev.outcome == "MATCH":
                        resultado_str = "FAIL"
                    elif ev.outcome == "ERROR":
                        resultado_str = "ERROR"
                    else:
                        resultado_str = "PASS"

                    ra = ResultadoAuditoria(
                        evidencia_id=ev.id,
                        regla_id=ev.regla_id,
                        regla_version=ev.regla_version,
                        factura=ev.factura,
                        param_config_id=ev.param_config_id,
                        resultado=resultado_str,
                        severidad=rule.severidad,
                        mensaje=ev.error_mensaje or rule.descripcion,
                        detalles={"outcome": ev.outcome},
                    )
                    self._session.add(ra)

                self._session.flush()

        return results

    # ── Internal helpers ──────────────────────────────────────────────────

    def _load_rule_by_name(self, rule_name: str) -> Regla | None:
        """Load a single rule by name."""
        return (
            self._session.query(Regla)
            .filter(Regla.nombre == rule_name)
            .filter(Regla.activo == True)  # noqa: E712
            .first()
        )

    def _load_conditions(self, regla_id: int) -> list[dict]:
        """Load all conditions for a rule and convert to dicts."""
        rows = (
            self._session.query(Condicion)
            .filter(Condicion.regla_id == regla_id)
            .order_by(Condicion.padre_id.asc().nullsfirst(), Condicion.orden.asc())
            .all()
        )
        return [
            {
                "id": c.id,
                "regla_id": c.regla_id,
                "padre_id": c.padre_id,
                "tipo": c.tipo,
                "operador": c.operador,
                "fuente_datos": c.fuente_datos,
                "valor_esperado": c.valor_esperado,
                "orden": c.orden,
            }
            for c in rows
        ]

    def _build_row_context(
        self,
        data_sheet: "Worksheet",
        row: int,
        indices: dict[str, int | None],
    ) -> tuple[dict[str, Any], str | None]:
        """Extract row data from Excel worksheet using column indices.

        Returns:
            (row_data_dict, factura_string) where factura_string may be None.
        """
        row_data: dict[str, Any] = {}
        factura = None

        for col_name, col_idx in indices.items():
            if col_idx is None:
                continue
            value = data_sheet.cell(row=row, column=col_idx + 1).value
            row_data[col_name] = value
            if col_name in ("numero_factura", "factura") and value is not None:
                factura = str(value).strip()

        return row_data, factura

    def _resolve_computed(self, path: str, ctx: "EvaluationContext") -> Any:
        """Resolve a computed field like 'date.edad' for inclusion in problem output."""
        from app.services.engine.providers import PROVIDER_REGISTRY
        prefix = path.split(".")[0] if "." in path else path
        provider = PROVIDER_REGISTRY.get(prefix)
        if provider:
            try:
                return provider.resolve(path, ctx)
            except Exception:
                pass
        return None


def _extract_factura(row_data: dict[str, Any]) -> str | None:
    """Extract the invoice identifier from a row dict.

    Tries ``numero_factura`` first, then ``factura``. Returns ``None`` if
    neither key holds a non-empty string.
    """
    raw = row_data.get("numero_factura") or row_data.get("factura")
    if raw is not None:
        stripped = str(raw).strip()
        if stripped:
            return stripped
    return None
