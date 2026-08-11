"""GroupEvaluator — evaluates condition trees against GROUPS of rows.

Pre-scan → partition → aggregate → evaluate → merge lifecycle.
Keeps the row-by-row ConditionEvaluator path untouched.
"""

from __future__ import annotations

import logging
from typing import Any, TYPE_CHECKING

from app.services.engine.context import EvaluationContext

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet
    from app.services.engine.condition_evaluator import ConditionEvaluator
    from app.services.engine.evidence_collector import EvidenceCollector

logger = logging.getLogger(__name__)


class GroupEvaluator:
    """Evaluates conditions against GROUPS of rows instead of individual rows.

    Lifecycle:
        1. build_groups() — pre-scan rows, key by factura.
        2. _build_group_data() — for each group, compute aggregate values.
        3. evaluate() — evaluate condition tree against group-level data.

    Supported aggregation functions:
        - distinct_count(field) → number of distinct values in group
        - group_size → number of rows in group
        - sum(field) → sum of numeric values in group

    The evaluator supports two data-paths:
    - **Worksheet path**: Provide ``data_sheet`` and ``indices``.
    - **RowStore path**: Provide ``rows`` (list[dict]) — dict accesses replace
      ``data_sheet.cell()`` calls.
    """

    @staticmethod
    def _cell_value(
        row_idx: int,
        field: str,
        data_sheet: "Worksheet | None",
        indices: dict[str, int | None],
        rows: list[dict] | None = None,
    ) -> Any:
        """Get a cell value from either a Worksheet or a RowStore dict.

        When ``rows`` is provided, uses O(1) dict lookup: ``rows[row_idx-2].get(field)``.
        Otherwise falls back to ``data_sheet.cell(row_idx, col_idx+1).value``.

        Args:
            row_idx: 1-based row number (matches Excel row numbering).
            field: Column name key (snake_case).
            data_sheet: openpyxl Worksheet (ignored when ``rows`` is set).
            indices: Column name → 0-based column index mapping.
            rows: Optional list of dicts (RowStore) for fast path.

        Returns:
            Cell value or ``None`` if the field does not exist.
        """
        if rows is not None:
            return rows[row_idx - 2].get(field)
        field_idx = indices.get(field)
        if field_idx is None:
            return None
        return data_sheet.cell(row=row_idx, column=field_idx + 1).value  # type: ignore[union-attr]

    @staticmethod
    def build_groups(
        data_sheet: "Worksheet | None" = None,
        indices: dict[str, int | None] | None = None,
        group_by_field: str | list[str] = "numero_factura",
        filter_field: str | None = None,
        filter_value: str | None = None,
        rows: list[dict[str, Any]] | None = None,
    ) -> dict[str, list[int]]:
        """Pre-scan: build groups from sheet data keyed by factura.

        Supports both Worksheet and RowStore paths. When ``rows`` is provided,
        data is read from the dict list instead of ``data_sheet.cell()``.

        Args:
            data_sheet: openpyxl Worksheet with invoice data.
            indices: Column name → 0-based column index mapping.
            group_by_field: Column(s) to group by. Can be a single field name
                (str, default: numero_factura) or a list of field names for
                composite grouping (e.g., ["identificacion", "codigo", "dx"]).
            filter_field: Optional column to filter rows before grouping.
            filter_value: Required value for filter_field when filtering.
            rows: Optional list of dicts (RowStore) for O(1) dict access path.

        Returns:
            Dict mapping composite key string → list of 1-based row numbers.
            Empty dict if any group-by column is missing.
        """
        groups: dict[str, list[int]] = {}

        if rows is not None:
            # ── RowStore path ──────────────────────────────────────────────
            for idx, row_dict in enumerate(rows):
                row = idx + 2  # 1-based row number (row 2 = rows[0])

                # Apply row filter if configured
                if filter_field is not None:
                    val = str(row_dict.get(filter_field) or "").strip().upper()
                    if val != (filter_value or "").upper():
                        continue

                # Build composite key
                if isinstance(group_by_field, list):
                    key_parts: list[str] = [
                        str(row_dict.get(f, "") or "").strip()
                        for f in group_by_field
                    ]
                    key = "\t".join(key_parts)
                    if not any(key_parts):
                        continue
                else:
                    factura = str(row_dict.get(group_by_field) or "").strip()
                    if not factura:
                        continue
                    key = factura

                if key not in groups:
                    groups[key] = []
                groups[key].append(row)

            return groups

        # ── Worksheet path (legacy) ────────────────────────────────────────
        # Support composite group_by (list of fields)
        if isinstance(group_by_field, list):
            group_indices: list[int | None] = [indices.get(f) for f in group_by_field]  # type: ignore[union-attr]
            if any(idx is None for idx in group_indices):
                return groups
        else:
            num_fact_idx = indices.get(group_by_field)  # type: ignore[union-attr]
            if num_fact_idx is None:
                return groups

        filter_idx = indices.get(filter_field) if filter_field else None  # type: ignore[union-attr]

        for row in range(2, data_sheet.max_row + 1):  # type: ignore[union-attr]
            # Apply row filter if configured
            if filter_idx is not None:
                val = str(data_sheet.cell(row=row, column=filter_idx + 1).value or "").strip().upper()  # type: ignore[union-attr]
                if val != filter_value.upper():
                    continue

            # Build composite key
            if isinstance(group_by_field, list):
                key_parts = []
                for idx in group_indices:
                    val = data_sheet.cell(row=row, column=idx + 1).value  # type: ignore[union-attr]
                    key_parts.append(str(val).strip() if val is not None else "")
                key = "\t".join(key_parts)  # Tab-separated for uniqueness
                if not any(k for k in key_parts):  # Skip if all empty
                    continue
            else:
                factura = str(
                    data_sheet.cell(row=row, column=num_fact_idx + 1).value or ""  # type: ignore[union-attr]
                ).strip()
                if not factura:
                    continue
                key = factura

            if key not in groups:
                groups[key] = []
            groups[key].append(row)

        return groups

    @staticmethod
    def _build_group_data(
        factura: str,
        group_rows: list[int],
        data_sheet: "Worksheet | None",
        indices: dict[str, int | None],
        agg_configs: list[dict[str, Any]],
        rows: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        """Compute aggregate data for a group of rows.

        For each config in agg_configs, computes the specified aggregation
        and stores the result under the target field name.

        Args:
            factura: Group key (invoice number).
            group_rows: List of 1-based row numbers in this group.
            data_sheet: openpyxl Worksheet.
            indices: Column name → index mapping.
            agg_configs: List of dicts with keys:
                - function: "distinct_count" | "group_size" | "sum" | etc.
                - field: Source column name (not needed for group_size)
                - target: Output field name (default: {function}_{field})
            rows: Optional list of dicts (RowStore) for O(1) dict access.

        Returns:
            Dict with aggregated values plus "numero_factura" key.
        """
        agg_data: dict[str, Any] = {"numero_factura": factura}

        for config in agg_configs:
            func = config.get("function", "")
            field = config.get("field", "")
            target = config.get("target") or (
                f"{func}_{field}" if field else func
            )

            if func == "distinct_count":
                agg_data[target] = GroupEvaluator._agg_distinct_count(
                    group_rows, data_sheet, indices, field, rows=rows,
                )
            elif func == "group_size":
                agg_data[target] = len(group_rows)
            elif func == "sum":
                agg_data[target] = GroupEvaluator._agg_sum(
                    group_rows, data_sheet, indices, field, rows=rows,
                )
            elif func == "collect_set":
                agg_data[target] = GroupEvaluator._agg_collect_set(
                    group_rows, data_sheet, indices, field, rows=rows,
                )
            elif func == "collect_value_counts":
                fields = config.get("fields", [field])
                agg_data[target] = GroupEvaluator._agg_collect_value_counts(
                    group_rows, data_sheet, indices, fields, rows=rows,
                )
            elif func == "compute_horas":
                agg_data[target] = GroupEvaluator._agg_compute_horas(
                    group_rows, data_sheet, indices,
                    config.get("field1", ""),
                    config.get("field2", ""),
                    rows=rows,
                )
            elif func == "collect_group_keys":
                agg_data[target] = GroupEvaluator._agg_collect_group_keys(
                    group_rows, data_sheet, indices, field, rows=rows,
                )
            else:
                logger.warning("Unknown aggregation function: %s", func)

        return agg_data

    @staticmethod
    def _agg_distinct_count(
        group_rows: list[int],
        data_sheet: "Worksheet | None",
        indices: dict[str, int | None],
        field: str,
        rows: list[dict[str, Any]] | None = None,
    ) -> int:
        """Count distinct non-None values of a field across rows."""
        values: set[str] = set()
        if rows is not None:
            for row_idx in group_rows:
                val = rows[row_idx - 2].get(field)
                if val is not None:
                    values.add(str(val).strip())
        else:
            field_idx = indices.get(field)
            if field_idx is None:
                return 0
            for row_idx in group_rows:
                val = data_sheet.cell(row=row_idx, column=field_idx + 1).value  # type: ignore[union-attr]
                if val is not None:
                    values.add(str(val).strip())
        return len(values)

    @staticmethod
    def _agg_sum(
        group_rows: list[int],
        data_sheet: "Worksheet | None",
        indices: dict[str, int | None],
        field: str,
        rows: list[dict[str, Any]] | None = None,
    ) -> float:
        """Sum numeric values of a field across rows."""
        total = 0.0
        if rows is not None:
            for row_idx in group_rows:
                val = rows[row_idx - 2].get(field)
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        else:
            field_idx = indices.get(field)
            if field_idx is None:
                return total
            for row_idx in group_rows:
                val = data_sheet.cell(row=row_idx, column=field_idx + 1).value  # type: ignore[union-attr]
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        return total

    @staticmethod
    def _agg_collect_set(
        group_rows: list[int],
        data_sheet: "Worksheet | None",
        indices: dict[str, int | None],
        field: str,
        rows: list[dict[str, Any]] | None = None,
    ) -> list[str]:
        """Collect unique non-None values of a field across rows.

        Returns a list (not set) for JSONB serialization compatibility.
        """
        values: set[str] = set()
        if rows is not None:
            for row_idx in group_rows:
                val = rows[row_idx - 2].get(field)
                if val is not None:
                    values.add(str(val).strip())
        else:
            field_idx = indices.get(field)
            if field_idx is None:
                return []
            for row_idx in group_rows:
                val = data_sheet.cell(row=row_idx, column=field_idx + 1).value  # type: ignore[union-attr]
                if val is not None:
                    values.add(str(val).strip())
        return list(values)

    @staticmethod
    def _agg_collect_group_keys(
        group_rows: list[int],
        data_sheet: "Worksheet | None",
        indices: dict[str, int | None],
        field: str,
        rows: list[dict[str, Any]] | None = None,
    ) -> list[str]:
        """Collect unique values of a field across group rows.

        Returns a sorted list of unique string values for the given field
        across all rows in the group. Used by duplicado_id_codigo engine
        rules to list facturas in a duplicate group.

        Args:
            group_rows: List of 1-based row numbers in this group.
            data_sheet: openpyxl Worksheet.
            indices: Column name → index mapping.
            field: Source column name to collect values from.
            rows: Optional list of dicts (RowStore) for fast path.

        Returns:
            Sorted list of unique string values. Empty if field is missing.
        """
        values: set[str] = set()
        if rows is not None:
            for row_idx in group_rows:
                val = rows[row_idx - 2].get(field)
                if val is not None:
                    values.add(str(val).strip())
        else:
            field_idx = indices.get(field)
            if field_idx is None:
                return []
            for row_idx in group_rows:
                val = data_sheet.cell(row=row_idx, column=field_idx + 1).value  # type: ignore[union-attr]
                if val is not None:
                    values.add(str(val).strip())
        return sorted(values)

    @staticmethod
    def _agg_collect_value_counts(
        group_rows: list[int],
        data_sheet: "Worksheet | None",
        indices: dict[str, int | None],
        fields: list[str],
        rows: list[dict[str, Any]] | None = None,
    ) -> list[dict[str, object]]:
        """Count occurrences of (field1, field2) pairs across rows.

        Returns a list of dicts with keys: codigo, cantidad, count.
        JSONB-compatible output format.
        """
        from collections import Counter
        pairs: Counter = Counter()
        if rows is not None:
            for row_idx in group_rows:
                key = tuple(
                    str(rows[row_idx - 2].get(f, "") or "").strip()
                    for f in fields
                )
                pairs[key] += 1
        else:
            field_indices = [indices.get(f) for f in fields]
            for row_idx in group_rows:
                key = tuple(
                    str(data_sheet.cell(row=row_idx, column=col_idx + 1).value or "").strip()  # type: ignore[union-attr]
                    for col_idx in field_indices
                    if col_idx is not None
                )
                pairs[key] += 1
        result = []
        for k, v in pairs.items():
            entry: dict[str, object] = {"count": v}
            for i, field_name in enumerate(fields):
                if i < len(k):
                    entry[field_name] = k[i]
            result.append(entry)
        return result

    @staticmethod
    def _agg_compute_horas(
        group_rows: list[int],
        data_sheet: "Worksheet | None",
        indices: dict[str, int | None],
        field1: str,
        field2: str,
        rows: list[dict[str, Any]] | None = None,
    ) -> float:
        """Compute absolute hours between two datetime fields.

        Takes the first row in the group where both fields are non-None.
        Supports datetime objects, ISO strings (%Y-%m-%d %H:%M:%S),
        and Excel serial dates (numeric days since 1899-12-30).

        Returns 0.0 if either column is missing or no valid date pair found.
        """
        from datetime import datetime, timedelta

        def _parse_date(value):
            """Parse a cell value into a datetime."""
            if isinstance(value, datetime):
                return value
            if isinstance(value, str):
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S"):
                    try:
                        return datetime.strptime(value, fmt)
                    except ValueError:
                        continue
            if isinstance(value, (int, float)):
                serial = value
                if serial > 59:
                    serial -= 1  # Adjust for Lotus 123 leap year bug
                try:
                    return datetime(1899, 12, 30) + timedelta(days=serial)
                except (ValueError, OverflowError):
                    pass
            return None

        if rows is not None:
            for row_idx in group_rows:
                val1 = rows[row_idx - 2].get(field1)
                val2 = rows[row_idx - 2].get(field2)
                if val1 is None or val2 is None:
                    continue
                dt1 = _parse_date(val1)
                dt2 = _parse_date(val2)
                if dt1 is not None and dt2 is not None:
                    diff_seconds = (dt2 - dt1).total_seconds()
                    return abs(diff_seconds) / 3600.0
        else:
            idx1 = indices.get(field1)
            idx2 = indices.get(field2)
            if idx1 is None or idx2 is None:
                return 0.0
            for row_idx in group_rows:
                val1 = data_sheet.cell(row=row_idx, column=idx1 + 1).value  # type: ignore[union-attr]
                val2 = data_sheet.cell(row=row_idx, column=idx2 + 1).value  # type: ignore[union-attr]
                if val1 is None or val2 is None:
                    continue
                dt1 = _parse_date(val1)
                dt2 = _parse_date(val2)
                if dt1 is not None and dt2 is not None:
                    diff_seconds = (dt2 - dt1).total_seconds()
                    return abs(diff_seconds) / 3600.0

        return 0.0

    @staticmethod
    def evaluate(
        groups: dict[str, list[int]],
        data_sheet: "Worksheet | None",
        indices: dict[str, int | None],
        agg_configs: list[dict[str, Any]],
        condition_tree: dict | None,
        condition_evaluator: "ConditionEvaluator",
        rule_info: dict[str, Any],
        evidence_collector: "EvidenceCollector",
        record_evidence: bool = True,
        rows: list[dict[str, Any]] | None = None,
    ) -> list[dict[str, Any]]:
        """Evaluate a group-by rule against all groups.

        For each group:
        1. Compute aggregated data via _build_group_data.
        2. Build an EvaluationContext with the aggregated data.
        3. Evaluate the condition tree via ConditionEvaluator.
        4. Record evidence (if record_evidence=True) and collect MATCH results.

        Args:
            groups: Dict mapping factura → list of row numbers.
            data_sheet: openpyxl Worksheet.
            indices: Column name → index mapping.
            agg_configs: Aggregation configurations from rule parametros.
            condition_tree: Root node of the condition tree.
            condition_evaluator: ConditionEvaluator instance.
            rule_info: Dict with id, version, dominio, nombre, descripcion, severidad.
            evidence_collector: EvidenceCollector for audit trail.
            record_evidence: If False, skip evidence recording (speed during testing).
            rows: Optional list of dicts (RowStore) for O(1) dict access path.

        Returns:
            List of detection dicts with factura, problema, regla, severidad.
        """
        results: list[dict[str, Any]] = []

        for factura, group_rows in groups.items():
            # 1. Compute aggregated data
            group_data = GroupEvaluator._build_group_data(
                factura, group_rows, data_sheet, indices, agg_configs, rows=rows,
            )

            # 2. Build evaluation context with aggregated data
            ctx = EvaluationContext(invoice_data=group_data, indices=indices)

            # 3. Evaluate condition tree
            eval_result = condition_evaluator.evaluate(condition_tree, ctx)
            outcome = eval_result.get("outcome", False)
            error_msg = eval_result.get("error")

            # 4. Determine final outcome
            if error_msg:
                final_outcome = "ERROR"
            elif outcome:
                final_outcome = "MATCH"
            else:
                final_outcome = "NO_MATCH"

            # 5. Record evidence (immutable snapshot) — skip when testing
            if record_evidence:
                evidence_collector.record(
                    regla_id=rule_info["id"],
                    regla_version=rule_info["version"],
                    dominio=rule_info["dominio"],
                    factura=factura,
                    outcome=final_outcome,
                    arbol_evaluado=eval_result.get("trace", {}),
                    snapshot_fila=group_data,
                    error_mensaje=error_msg,
                )

            # 6. If MATCH, add to detection results
            if outcome and not error_msg:
                problem = {
                    "factura": factura,
                    "problema": rule_info.get("descripcion") or rule_info.get("nombre", ""),
                    "regla": f"#{rule_info.get('id', '')}",
                    "severidad": rule_info.get("severidad", "error"),
                }
                # Include aggregate data in problem dict for toggle post-processing
                for key, val in group_data.items():
                    if key not in ("numero_factura",) and val is not None:
                        problem[key] = val
                results.append(problem)

        return results
