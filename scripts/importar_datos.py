"""Importar datos iniciales desde Excel a DB PostgreSQL (batch optimizado).

Uso:
    python scripts/importar_datos.py

Orden: eps_contratado → procedimiento → nota_hoja
Usa add_all() + un solo commit por tabla para máxima velocidad.
"""

import sys
import logging
from pathlib import Path

sys.path.insert(0, ".")

import openpyxl

from app.database import get_session
from app.models import EpsContratado, Procedimiento, NotaHoja

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stdout,
)
logger = logging.getLogger("import")

DATA_DIR = Path("data/import")


def import_eps_contratado():
    """Importa eps_contratado desde entidades.xlsx (batch)."""
    path = DATA_DIR / "entidades.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    total = ws.max_row - 1
    objects = []
    errors = 0
    seen_cods = set()

    for row in range(2, ws.max_row + 1):
        cod = str(ws.cell(row=row, column=1).value or "").strip()
        eps = str(ws.cell(row=row, column=2).value or "").strip()
        reg = str(ws.cell(row=row, column=3).value or "SUBSIDIADO").strip()
        if not cod or not eps:
            errors += 1
            continue
        if cod in seen_cods:
            errors += 1
            continue
        seen_cods.add(cod)
        objects.append(EpsContratado(cod_contrato=cod, eps=eps, regimen=reg))

    db = get_session()
    try:
        db.add_all(objects)
        db.commit()
        logger.info("[EPS_CONTRATADO] %d importados, %d errores (de %d)", len(objects), errors, total)
    except Exception as e:
        db.rollback()
        logger.exception("Error en eps_contratado: %s", e)
    finally:
        db.close()
        wb.close()

    return len(objects)


def import_procedimiento():
    """Importa procedimiento desde procedimiento.xlsx (batch)."""
    path = DATA_DIR / "procedimiento.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    total = ws.max_row - 1
    errors = 0
    proc_map = {}  # cups → nombre (última aparición gana)

    for row in range(2, ws.max_row + 1):
        cups = str(ws.cell(row=row, column=1).value or "").strip()
        nombre = str(ws.cell(row=row, column=2).value or "").strip()
        if not cups:
            errors += 1
            continue
        proc_map[cups] = nombre  # sobrescribe si ya existe → última gana

    objects = [Procedimiento(cups=cups, procedimiento=nombre) for cups, nombre in proc_map.items()]

    db = get_session()
    try:
        # Batch de 5000 para no saturar memoria
        BATCH = 5000
        for i in range(0, len(objects), BATCH):
            batch = objects[i : i + BATCH]
            db.add_all(batch)
            db.commit()
            logger.info("  procedimiento: %d/%d", min(i + BATCH, len(objects)), total)

        logger.info("[PROCEDIMIENTO] %d importados (última aparición gana), %d filas vacías (de %d)", len(objects), errors, total)
    except Exception as e:
        db.rollback()
        logger.exception("Error en procedimiento: %s", e)
    finally:
        db.close()
        wb.close()

    return len(objects)


def import_nota_hoja():
    """Importa nota_hoja desde notas.xlsx (batch)."""
    path = DATA_DIR / "notas.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    total = ws.max_row - 1
    objects = []
    errors = 0

    for row in range(2, ws.max_row + 1):
        nota = str(ws.cell(row=row, column=1).value or "").strip()
        if nota:
            objects.append(NotaHoja(nota=nota))
        else:
            errors += 1

    db = get_session()
    try:
        db.add_all(objects)
        db.commit()
        logger.info("[NOTA_HOJA] %d importados, %d errores (de %d)", len(objects), errors, total)
    except Exception as e:
        db.rollback()
        logger.exception("Error en nota_hoja: %s", e)
    finally:
        db.close()
        wb.close()

    return len(objects)


def main():
    logger.info("=" * 50)
    logger.info("INICIANDO IMPORTACION DE DATOS")
    logger.info("=" * 50)

    logger.info("\n1) eps_contratado (entidades.xlsx)...")
    eps = import_eps_contratado()

    logger.info("\n2) procedimiento (procedimiento.xlsx)...")
    proc = import_procedimiento()

    logger.info("\n3) nota_hoja (notas.xlsx)...")
    notas = import_nota_hoja()

    logger.info("\n" + "=" * 50)
    logger.info("RESUMEN FINAL")
    logger.info("=" * 50)
    logger.info("  eps_contratado: %d", eps)
    logger.info("  procedimiento:  %d", proc)
    logger.info("  nota_hoja:      %d", notas)
    logger.info("=" * 50)


if __name__ == "__main__":
    main()
