"""Quick test: odontologia through unified processor."""
from openpyxl import Workbook
from app.services.exporter import _do_detect_problems
from app.constants import AREA_UNIFICADA
import os

wb = Workbook()
ws = wb.active
headers = [
    'Número Factura','Vlr. Subsidiado','Vlr. Procedimiento','Código Tipo Procedimiento',
    'Tipo Procedimiento','Código','Cód. Equivalente CUPS','Procedimiento',
    'Nº Identificación','Convenio Facturado','Cantidad','Laboratorio',
    'Centro Costo','Cód Entidad Cobrar','Entidad Cobrar','Entidad Afiliación',
    'Tipo Factura Descripción','IDE Contrato','Tipo Identificación',
    'Fec. Nacimiento','Fec. Factura','Fecha Cierre',
    'Identificación Profesional','Profesional Atiende','Código Profesional',
    'Responsable Cierra Facturar','Tarifario','Tipo Usuario','Vlr. Copago',
    'Nº Reingreso','Cód. Dx Principal'
]
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)

ws.cell(row=2, column=1, value='ODT001')
ws.cell(row=2, column=9, value='123456')
ws.cell(row=2, column=7, value='997002')
ws.cell(row=2, column=8, value='Control Placa')
ws.cell(row=2, column=17, value='Odontología')
ws.cell(row=2, column=6, value='997002')

filename = 'app/data/input/_test_odontologia.xlsx'
wb.save(filename)
try:
    result = _do_detect_problems(filename='_test_odontologia.xlsx', area=AREA_UNIFICADA)
    print('Status:', result['status'])
    if result['status'] == 'error':
        print('Errors:', result.get('errors'))
    else:
        problemas = result.get('data', {}).get('problemas', {})
        tipos = problemas.get('tipos_procesados', [])
        print('Tipos procesados:', tipos)
        norm = problemas.get('problemas', {}).get('normalizados', [])
        print('Normalized rows:', len(norm))
        if norm:
            print('First:', norm[0].get('tipo_error'))
finally:
    os.unlink(filename)
